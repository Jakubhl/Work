import customtkinter
import tkinter as tk
from openpyxl import load_workbook
import subprocess
import os
import ipaddress

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")
root=customtkinter.CTk()
root.geometry("1200x900")
root.title("IP nastavovač")

def path_check(path_raw,only_repair = None):
    path=path_raw
    backslash = "\ "
    if backslash[0] in path:
        newPath = path.replace(backslash[0], '/')
        path = newPath
    if path.endswith('/') == False:
        newPath = path + "/"
        path = newPath
    #oprava mezery v nazvu
    path = r"{}".format(path)
    if not os.path.exists(path) and only_repair == None:
        return False
    else:
        return path

def add_colored_line(text_widget, text, color,font=None,delete_line = None):
    """
    Vloží řádek do console
    """
    text_widget.configure(state=tk.NORMAL)
    if font == None:
        font = ("Arial",16)
    if delete_line != None:
        text_widget.delete("current linestart","current lineend")
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert("current lineend",text, color)
    else:
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert(tk.END,"    > "+ text+"\n", color)

    text_widget.configure(state=tk.DISABLED)

class IP_assignment: # Umožňuje procházet obrázky a přitom například vybrané přesouvat do jiné složky
    """
    Umožňuje procházet obrázky a přitom například vybrané přesouvat do jiné složky

    - umožňuje: měnit rychlost přehrávání, přiblížení, otočení obrázku
    - reaguje na klávesové zkratky
    """
    def __init__(self,root):
        self.root = root
        self.rows_taken = 0
        self.all_rows = []
        self.project_list = []
        self.number_of_parameters = 4
        app_path = os.getcwd()
        app_path = path_check(app_path,True)
        self.excel_file_path = app_path + "saved_adresses_2.xlsx"
        self.options_list = ["Ethernet",
                             "Ethernet 1",
                             "Ethernet 2",
                             "Ethernet 3",
                             "Ethernet 4",
                             "Ethernet 5",
                             "Wi-Fi"
                             ]

        self.create_widgets()

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def close_window(self,window):
        window.update_idletasks()
        window.destroy()

    def read_excel_data(self):
        self.all_rows = []
        self.project_list = []
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row:
                row_array.append(items)
            self.project_list.append(row_array[0])
            self.all_rows.append(row_array)
        self.rows_taken = int(len(self.all_rows))
        #print(self.all_rows)
        workbook.close()
            
    def save_excel_data(self,project_name,IP_adress,mask,notes):
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook.active
        row_to_print = self.rows_taken +1
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = ip adresa
        worksheet['B' + str(row_to_print)] = IP_adress
        #C = maska
        worksheet['C' + str(row_to_print)] = mask
        #D = poznamky
        worksheet['D' + str(row_to_print)] = notes

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def get_notes(self):
        notes_legit_rows = []
        notes = str(self.notes_input.get("1.0", tk.END))
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    def check_ip_and_mask(self,input_value):
        input_splitted = input_value.split(".")
        if len(input_splitted) == 4:
            return input_value
        else:
            return False

    def save_new_project_data(self,child_root):
        project_name = str(self.name_input.get())
        IP_adress = str(self.IP_adress_input.get())
        IP_adress = self.check_ip_and_mask(IP_adress)
        mask = str(self.mask_input.get())
        mask = self.check_ip_and_mask(mask)
        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
            errors += 1
        if IP_adress == False and errors == 0:
            add_colored_line(self.console,f"Neplatná IP adresa","red",None,True)
            errors += 1
        if mask == False and errors == 0:
            add_colored_line(self.console,f"Neplatná maska","red",None,True)
            errors += 1
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()
            self.save_excel_data(project_name,IP_adress,mask,notes)
            self.close_window(child_root)
            self.make_project_cells(only_one_new=True)
            add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)

    def delete_project(self):
        self.read_excel_data()
        wanted_project = str(self.search_input.get().replace(" ",""))
        project_found = False
        print(self.project_list)
        for i in range(0,len(self.project_list)):
            if self.project_list[i] == wanted_project and len(str(self.project_list[i])) == len(str(wanted_project)):
                row_index = self.project_list.index(wanted_project)
                workbook = load_workbook(self.excel_file_path)
                worksheet = workbook.active
                worksheet.delete_rows(row_index+1)
                workbook.save(self.excel_file_path)
                workbook.close()
                self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                project_found = True
                add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                break
        if project_found == False:
            add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)

    def add_new_project(self):
        child_root=customtkinter.CTk()
        child_root.geometry("520x500")
        child_root.title("Nový projekt")    

        project_name =      customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
        self.name_input =        customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        IP_adress =         customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
        self.IP_adress_input =   customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        mask =              customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
        self.mask_input =        customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =             customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        self.notes_input =       customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=120)
        self.console = tk.Text(child_root, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)

        save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root),font=("Arial",20,"bold"),corner_radius=0)

        project_name.   grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.name_input.     grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        IP_adress.      grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        self.IP_adress_input.grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        mask.           grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        self.mask_input.     grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        notes.          grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.    grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.console.   grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        save_button.        grid(column = 0,row=10,pady = 5,padx =165,sticky = tk.W)

        self.IP_adress_input.delete("0","300")
        self.IP_adress_input.insert("0","192.168.100.241")
        self.mask_input.delete("0","300")
        self.mask_input.insert("0","255.255.255.0")
        if str(self.search_input.get()).replace(" ","") != "":
            self.name_input.delete("0","300")
            self.name_input.insert("0",str(self.search_input.get()))

        child_root.mainloop()

    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused:
            return True
        else:
            return False

    def change_computer_ip(self,button_row):
        #button_row je id stisknuteho tlacitka... =0 od spodu
        ip = str(self.all_rows[button_row][1])
        mask = str(self.all_rows[button_row][2])
        # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
        interface_name = str(self.drop_down_options.get())
        powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
        # subprocess.run(["powershell.exe", "-Command", "Start-Process", "powershell.exe", "-Verb", "RunAs", "-ArgumentList", f"'-Command {powershell_command}'"])
        try:
            subprocess.run(["powershell.exe", "-Command",powershell_command],check=True)
            add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla přenastavena na: {ip}","green",None,True)
        except subprocess.CalledProcessError as e:
            add_colored_line(self.main_console,f"Chyba, aplikace musí být spuštěna, jako administrátor. (případně, nemáte tuto adresu již uloženou u jiného zařízení?)","red",None,True)

    def clicked_on_project(self,e,widget_id):
        self.search_input.delete("0","300")
        self.search_input.insert("0",str(self.all_rows[widget_id][0]))
        
    def make_project_cells(self,only_one_new=None):
        self.read_excel_data()
        padx_list = [10,220,400,600,800]
        if only_one_new == None:
            self.clear_frame(self.project_tree)
            # y = widgets ve smeru y, x = widgets ve smeru x
            for y in range(0,len(self.all_rows)):
                project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2)
                project_frame.pack(pady=0,padx=5,fill="x",expand=False,side = "bottom",anchor="w")
                # binding the click on widget
                #project_frame.bind("<Button-1>",self.clicked_on_project)
                project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                for x in range(0,len(self.all_rows[y])):
                    if x == 0:
                        button =  customtkinter.CTkButton(master = project_frame,width = 160,height=30,text = self.all_rows[y][x], command = lambda widget_id = y: self.change_computer_ip(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                        button.grid(column = 0,row=y,pady = 0,padx =padx_list[x],sticky = tk.W)
                    else:
                        parameter =  customtkinter.CTkLabel(master = project_frame,height=30,text = self.all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        parameter.grid(column = 0,row=y,pady = 5,padx =padx_list[x],sticky = tk.W)
        else:
            project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2)
            project_frame.pack(pady=0,padx=5,fill="x",expand=False,side = "bottom",anchor="w")
            y = len(self.all_rows)-1
            for x in range(0,len(self.all_rows[y])):
                if x == 0:
                    button =  customtkinter.CTkButton(master = project_frame,width = 160,height=30,text = self.all_rows[y][x], command = lambda widget_id=y: self.change_computer_ip(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                    button.grid(column = 0,row=y,pady = 0,padx =padx_list[x],sticky = tk.W)
                    # binding the click on widget
                    project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                else:
                    parameter =  customtkinter.CTkLabel(master = project_frame,height=30,text = self.all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                    parameter.grid(column = 0,row=y,pady = 5,padx =padx_list[x],sticky = tk.W)

    def create_widgets(self):
        main_widgets = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =  customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        main_widgets.pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        # project_tree.grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)

        project_label =  customtkinter.CTkLabel(master = main_widgets, width = 20,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input = customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=150,height=30,placeholder_text="Název projektu",corner_radius=0)
        button_search =  customtkinter.CTkButton(master = main_widgets, width = 20,height=30,text = "Vyhledat", command = lambda: self.test(),font=("Arial",16,"bold"),corner_radius=0)
        button_add =  customtkinter.CTkButton(master = main_widgets, width = 20,height=30,text = "Nový projekt", command = lambda: self.add_new_project(),font=("Arial",16,"bold"),corner_radius=0)
        button_remove = customtkinter.CTkButton(master = main_widgets, width = 80,height=30,text = "Smazat projekt", command =  lambda: self.delete_project(),font=("Arial",16,"bold"),corner_radius=0)
        self.drop_down_options = customtkinter.CTkOptionMenu(master = main_widgets,width=100,height=30,values=self.options_list,font=("Arial",16,"bold"),corner_radius=0)

        self.main_console = tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)

        project_label.grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.search_input.grid(column = 0,row=0,pady = 5,padx =90,sticky = tk.W)
        button_search.grid(column = 0,row=0,pady = 5,padx =245,sticky = tk.W)
        button_add.grid(column = 0,row=0,pady = 5,padx =320,sticky = tk.W)
        button_remove.grid(column = 0,row=0,pady = 5,padx =425,sticky = tk.W)
        self.drop_down_options.grid(column = 0,row=0,pady = 5,padx =550,sticky = tk.W)
        self.main_console.grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        
        self.make_project_cells()

        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)

IP_assignment(root)
root.mainloop()
