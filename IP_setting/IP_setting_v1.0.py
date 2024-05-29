import customtkinter
import tkinter as tk
from openpyxl import load_workbook
import subprocess
import os
import re
import time
import threading
import psutil
import socket

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")
root=customtkinter.CTk()
root.geometry("1200x900")
root.title("IP manager v3.4")

def path_check(path_raw,only_repair = None):
    path=path_raw
    backslash = "\ "
    if backslash[0] in path:
        newPath = path.replace(backslash[0], '/')
        path = newPath
    if path.endswith('/') == False:
        newPath = path + "/"
        path = newPath
    #oprava mezery v nazvu
    path = r"{}".format(path)
    if not os.path.exists(path) and only_repair == None:
        return False
    else:
        return path

def add_colored_line(text_widget, text, color,font=None,delete_line = None):
    """
    Vloží řádek do console
    """
    text_widget.configure(state=tk.NORMAL)
    if font == None:
        font = ("Arial",16)
    if delete_line != None:
        text_widget.delete("current linestart","current lineend")
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert("current lineend",text, color)
    else:
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert(tk.END,"    > "+ text+"\n", color)

    text_widget.configure(state=tk.DISABLED)

class IP_assignment: # Umožňuje měnit statickou IP a mountit disky
    """
    Umožňuje měnit nastavení statických IP adres
    """

    def __init__(self,root):
        self.root = root
        self.rows_taken = 0
        self.all_rows = []
        self.project_list = []
        self.app_path = os.getcwd()
        self.app_path = path_check(self.app_path,True)
        self.excel_file_path = self.app_path + "saved_addresses_2.xlsx"
        #default:
        self.connection_option_list = ["Ethernet",
                             "Ethernet 1",
                             "Ethernet 2",
                             "Ethernet 3",
                             "Ethernet 4",
                             "Ethernet 5",
                             "Wi-Fi"
                             ]
        self.default_connection_option = 0
        self.last_project_name = ""
        self.last_project_ip = ""
        self.last_project_mask = ""
        self.last_project_notes = ""
        self.last_project_id = ""

        self.last_project_disc_letter = ""
        self.last_project_ftp = ""
        self.last_project_username = ""
        self.last_project_password = ""

        self.managing_disc = False
        self.connection_status = None

        # okno oblibene/ normal
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["Settings"]
        show_favourite = worksheet['B' + str(3)].value
        if int(show_favourite) == 1:
            self.show_favourite = True
        else:
            self.show_favourite = False
        workbook.close()

        self.make_project_favourite = False
        self.favourite_list = []

        self.read_excel_data()
        self.create_widgets()

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def close_window(self,window):
        window.update_idletasks()
        window.destroy()

    def read_excel_data(self):
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"

        workbook = load_workbook(self.excel_file_path)
        # seznam vsech statickych ip adres
        self.all_rows = []
        self.project_list = []
        self.favourite_list = []
        worksheet = workbook[excel_worksheet]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:4]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            if len(row_array) < 4:
                row_array.append("")
            self.project_list.insert(0,row_array[0])
            self.all_rows.insert(0,row_array)
            for items in row[4:5]:
                self.favourite_list.insert(0,items)
            

        # seznam vsech ftp pripojeni k diskum
        self.disc_all_rows = []
        self.disc_project_list = []  
        worksheet = workbook["disc_list"]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:6]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            """if len(row_array) < 4:
                row_array.append("")"""
            self.disc_project_list.insert(0,row_array[0])
            self.disc_all_rows.insert(0,row_array)

        # ukladani nastavenych hodnot
        worksheet = workbook["Settings"]
        saved_def_con_option = worksheet['B' + str(1)].value
        self.default_connection_option = int(saved_def_con_option)
        # seznam interfaců
        self.connection_option_list = []
        all_options = worksheet['B' + str(2)].value
        all_options = str(all_options).split(",")
        for i in range (0,len(all_options)):
            if all_options[i] != "":
                self.connection_option_list.append(all_options[i])
                     
    def save_excel_data(self,project_name,IP_adress,mask,notes,only_edit = None,force_row_to_print=None,fav_status = None):
        workbook = load_workbook(self.excel_file_path)
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        worksheet = workbook[excel_worksheet]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.all_rows)) +1
            if only_edit != None:
                #pouze změna na temtýž řádku
                row_to_print = (len(self.all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        if notes.replace(" ","") == "" or notes == None:
            notes = ""
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = ip adresa
        worksheet['B' + str(row_to_print)] = IP_adress
        #C = maska
        worksheet['C' + str(row_to_print)] = mask
        #D = poznamky
        worksheet['D' + str(row_to_print)] = notes
        #E = oblibenost
        if fav_status != None:
            worksheet['E' + str(row_to_print)] = fav_status

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def save_excel_data_disc(self,project_name,disc_letter,ftp_address,username,password,notes,only_edit = None,force_row_to_print=None):
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["disc_list"]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.disc_all_rows)) +1
            if only_edit != None:
                #pouze změna na temtýž řádku
                row_to_print = (len(self.disc_all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = písmeno disku, označení...
        worksheet['B' + str(row_to_print)] = disc_letter
        #C = ftp adresa
        worksheet['C' + str(row_to_print)] = ftp_address
        #D = uživatelské jméno
        worksheet['D' + str(row_to_print)] = username
        #E = heslo
        worksheet['E' + str(row_to_print)] = password
        #F = poznamky
        worksheet['F' + str(row_to_print)] = notes

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def get_notes(self):
        notes_legit_rows = []
        notes = str(self.notes_input.get("1.0", tk.END))
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    def check_ip_and_mask(self,input_value):
        input_splitted = input_value.split(".")
        if len(input_splitted) == 4:
            return input_value
        else:
            return False

    def switch_fav_status(self,operation:str,project_given=None,new_project = None,deleting_whole_project = False):
        # selected_project = self.all_rows[self.last_project_id]
        if project_given == None:
            selected_project = str(self.search_input.get())
            if selected_project not in self.project_list:
                add_colored_line(self.main_console,"Nebyl vložen projekt",color="red",font=None,delete_line=True)
                return
            else:
                selected_project = self.all_rows[self.project_list.index(selected_project)]
        else:
            selected_project = project_given
        if self.show_favourite == False:
            if operation == "add_favourite":
                # swich statusu:
                if new_project:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],None,None,fav_status=1)                    
                else:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=1)
                # přepnutí
                self.show_favourite = True
                self.read_excel_data()
                # do tohoto prostředí uložím na začátek
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()
            
            elif operation == "del_favourite":
                # swich statusu:
                if not deleting_whole_project:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=0)
                # přepnutí
                self.show_favourite = True
                self.read_excel_data()
                # z tohoto prostředí smažu
                self.delete_project(wanted_project=selected_project[0],silence=True)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()

            elif operation == "rewrite_favourite":
                # přepnutí
                self.show_favourite = True
                self.read_excel_data()
                # nejprve popnu stary projekt, s povodnim jmenem
                # poté insertnu pozměněný
                print(self.all_rows,"the favourite projects")
                the_id_to_pop = self.project_list.index(self.last_project_name)
                self.all_rows.pop(the_id_to_pop)
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()

        elif self.show_favourite:
            # z aktuálního prostředí smažu
            self.delete_project(wanted_project=selected_project[0],silence=True)
            # musim prepnout prostředí jen kvůli změně statusu
            self.show_favourite = False
            self.read_excel_data()
            match_found = False
            for i in range(0,len(self.project_list)):
                if self.project_list[i] == selected_project[0] and len(str(self.project_list[i])) == len(str(selected_project[0])):
                    row_index = self.project_list.index(selected_project[0])
                    match_found = True
            if match_found:
                row = len(self.all_rows) - row_index
                self.save_excel_data(self.all_rows[row_index][0],self.all_rows[row_index][1],self.all_rows[row_index][2],self.all_rows[row_index][3],None,row,fav_status=0)

            # přepnutí zpět
            self.show_favourite = True
            self.read_excel_data()
        
        if operation == "with_refresh":
            add_colored_line(self.main_console,f"Projekt: {selected_project[0]} byl odebrán z oblíbených","green",None,True)
            self.make_project_cells(no_read=True)

    def save_new_project_data(self,child_root,only_edit = None,make_fav=False):
        project_name = str(self.name_input.get())
        IP_adress = str(self.IP_adress_input.get())
        IP_adress = self.check_ip_and_mask(IP_adress)
        mask = str(self.mask_input.get())
        mask = self.check_ip_and_mask(mask)
        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
            errors += 1
        if IP_adress == False and errors == 0:
            add_colored_line(self.console,f"Neplatná IP adresa","red",None,True)
            errors += 1
        if mask == False and errors == 0:
            add_colored_line(self.console,f"Neplatná maska","red",None,True)
            errors += 1
        # poznamky nejsou povinne
        if errors ==0:
            if make_fav:
                fav_status = 1
            else:
                fav_status = 0
            self.read_excel_data()

            if only_edit == None: # pridavam novy projekt
                if make_fav:
                    new_project = [project_name,IP_adress,mask,notes,1]
                    self.switch_fav_status("add_favourite",new_project,new_project=True)
                    add_colored_line(self.main_console,f"Přidán nový oblíbený projekt: {project_name}","green",None,True)
                else:
                    self.save_excel_data(project_name,IP_adress,mask,notes,None,None,fav_status)
                    add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)
            else:
                # kdyz edituji muze mit projekt jiz prideleny status
                current_fav_status = self.is_project_favourite(self.last_project_id)
                self.save_excel_data(project_name,IP_adress,mask,notes,True,None,fav_status=current_fav_status)
                if make_fav:
                    project_with_changes = [project_name,IP_adress,mask,notes,current_fav_status]
                    self.switch_fav_status("add_favourite",project_with_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} úspěšně pozměněn a přidán do oblíbených","green",None,True)
                elif make_fav == False:
                    project_with_changes = [project_name,IP_adress,mask,notes,current_fav_status]
                    self.switch_fav_status("del_favourite",project_with_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} úspěšně pozměněn a odebrán z oblíbených","green",None,True)
                elif current_fav_status == 1: 
                    #nedoslo ke zmene statusu, ale mohlo dojit ke zmene - proto prepsat v oblibenych - vzdy se jedna o oblibene...
                    project_with_changes = [project_name,IP_adress,mask,notes,current_fav_status]
                    self.switch_fav_status("rewrite_favourite",project_with_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} úspěšně pozměněn","green",None,True)
                else:
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} úspěšně pozměněn","green",None,True)

                
            
            self.close_window(child_root)
            self.make_project_cells()
    
    def save_new_project_data_disc(self,child_root,only_edit = None):
        project_name =  str(self.name_input.get())
        disc_letter =   str(self.disc_letter_input.get())
        ftp_address =   str(self.FTP_adress_input.get())
        username =      str(self.username_input.get())
        password =      str(self.password_input.get())

        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
            errors += 1
        
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()
            if only_edit == None:
                self.save_excel_data_disc(project_name,disc_letter,ftp_address,username,password,notes)
            else:
                self.save_excel_data_disc(project_name,disc_letter,ftp_address,username,password,notes,True)
            self.close_window(child_root)
            if only_edit == None:
                self.make_project_cells_disc()
                add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)
            else: #musi byt proveden reset
                self.make_project_cells_disc()
                add_colored_line(self.main_console,f"Projekt: {project_name} úspěšně pozměněn","green",None,True)

    def delete_project(self,wanted_project=None,silence=None):
        remove_favourite_as_well = False
        if wanted_project == None:
            self.read_excel_data()
            wanted_project = str(self.search_input.get())
        project_found = False
        workbook = load_workbook(self.excel_file_path)
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        worksheet = workbook[excel_worksheet]

        for i in range(0,len(self.project_list)):
            if self.project_list[i] == wanted_project and len(str(self.project_list[i])) == len(str(wanted_project)) and project_found == False:
                row_index = self.project_list.index(wanted_project)
                print(self.favourite_list[row_index],"  ",self.all_rows[row_index],self.show_favourite)
                worksheet.delete_rows(len(self.all_rows)-row_index)
                project_found = True
                #pokud ma status oblibenosti, tak vymazat i z oblibenych:
                if self.favourite_list[row_index] == 1 and self.show_favourite == False:
                    remove_favourite_as_well = True
                    deleted_project = self.all_rows[row_index]
            workbook.save(self.excel_file_path)
            workbook.close()

        if silence == None:
            if project_found:
                add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
            else:
                add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)
        
        if remove_favourite_as_well:
            self.switch_fav_status("del_favourite",deleted_project,deleting_whole_project = True)


    def delete_project_disc(self):
        self.read_excel_data()
        wanted_project = str(self.search_input.get())
        project_found = False
        for i in range(0,len(self.disc_project_list)):
            if self.disc_project_list[i] == wanted_project and len(str(self.disc_project_list[i])) == len(str(wanted_project)):
                row_index = self.disc_project_list.index(wanted_project)
                workbook = load_workbook(self.excel_file_path)
                worksheet = workbook["disc_list"]
                worksheet.delete_rows(len(self.disc_all_rows)-row_index)
                workbook.save(self.excel_file_path)
                workbook.close()
                self.make_project_cells_disc() #refresh = cele zresetovat, jine: id, poradi...
                project_found = True
                add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                break
        if project_found == False:
            add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)

    def copy_previous_project(self,disc=None):
        if self.last_project_name == "":
            add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
        else:
            self.name_input.delete("0","300")
            self.name_input.insert("0",str(self.last_project_name))
            if disc == None:
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0",str(self.last_project_ip))
                self.mask_input.delete("0","300")
                self.mask_input.insert("0",str(self.last_project_mask))
                self.notes_input.insert(tk.END,str(self.last_project_notes))
            else:
                self.disc_letter_input.delete("0","300")
                self.disc_letter_input.insert("0",str(self.last_project_disc_letter))
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0",str(self.last_project_ftp))
                self.username_input.delete("0","300")
                self.username_input.insert("0",str(self.last_project_username))
                self.password_input.delete("0","300")
                self.password_input.insert("0",str(self.last_project_password))
                self.notes_input.insert(tk.END,str(self.last_project_notes))

    def make_favourite_toggle_via_edit(self,e,new_project = False):
        def do_favourite():
            self.make_fav_btn.configure(text = "🐘",font=("Arial",130),text_color = "pink")
            self.make_fav_label.configure(text = "Oblíbený ❤️")
        
        def unfavourite():
            self.make_fav_btn.configure(text = "❌",font=("Arial",100),text_color = "red")
            self.make_fav_label.configure(text = "Neoblíbený")

        if new_project:
            if self.make_project_favourite == None or self.make_project_favourite == True:
                self.make_project_favourite = False
                unfavourite()
            else:
                self.make_project_favourite = True
                do_favourite()
        else:
            if self.is_project_favourite(self.last_project_id):
                self.make_project_favourite = False
                unfavourite()
            else:
                self.make_project_favourite = True
                do_favourite()
        
    def add_new_project(self,edit = None):
        if self.show_favourite:
            #přepnutí do hlavního prostředí
            self.show_favourite_toggle(True)

        child_root=customtkinter.CTk()
        child_root.geometry("520x750")
        if edit:
            child_root.title("Editovat projekt: "+self.last_project_name)
        else:
            child_root.title("Nový projekt")  

        project_name =    customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
        copy_check =      customtkinter.CTkButton(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: self.copy_previous_project())
        self.name_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        
        if edit:
            if self.is_project_favourite(self.last_project_id):
                self.make_project_favourite = None #init hodnota
                self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Oblíbený ❤️",font=("Arial",20,"bold"))
                fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
                self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "🐘",font=("Arial",130),text_color = "pink")
            else:
                self.make_project_favourite = None #init hodnota
                self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Neoblíbený",font=("Arial",20,"bold"))
                fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
                self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "❌",font=("Arial",100),text_color = "red")
        else: # defaultne neoblibeny
            self.make_project_favourite = None #init hodnota
            self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Neoblíbený",font=("Arial",20,"bold"))
            fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
            self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "❌",font=("Arial",100),text_color = "red")

        IP_adress =            customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
        self.IP_adress_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        mask =                 customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
        self.mask_input =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =                customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        self.notes_input =     customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=370)
        self.console =         tk.Text(child_root, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,True,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,None,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)

        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        copy_check.             grid(column = 0,row=0,pady = 5,padx =240,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        # if edit:
        self.make_fav_label.grid(column = 0,row=1,pady = 5,padx =240,sticky = tk.W)
        IP_adress.              grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        # if edit:
        fav_frame.          grid(row=3,column=0,padx=240,sticky=tk.W,rowspan=4)
        fav_frame.          grid_propagate(0)
        self.make_fav_btn.  grid(column=0,row=0)
        if edit:
            self.make_fav_btn.  bind("<Button-1>",lambda e: self.make_favourite_toggle_via_edit(e))
        else:
            self.make_fav_btn.  bind("<Button-1>",lambda e: self.make_favourite_toggle_via_edit(e,new_project=True))

        self.IP_adress_input.   grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        mask.                   grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        self.mask_input.        grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        notes.                  grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=10,pady = 5,padx =165,sticky = tk.W)

        if edit:
            self.copy_previous_project()
        else:
            self.IP_adress_input.delete("0","300")
            self.IP_adress_input.insert("0","192.168.000.000")
            self.mask_input.delete("0","300")
            self.mask_input.insert("0","255.255.255.0")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))


        child_root.mainloop()

    def add_new_project_disc(self,edit = None):
        child_root=customtkinter.CTk()
        child_root.geometry("520x800")
        if edit == None:
            child_root.title("Nový projekt")
        else:
            child_root.title("Editovat projekt: "+self.last_project_name)

        project_name =              customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
        copy_check =                customtkinter.CTkButton(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: self.copy_previous_project(True))
        self.name_input =           customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        disc_letter =               customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Písmeno disku: ",font=("Arial",20,"bold"))
        self.disc_letter_input =    customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        FTP_adress =                customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
        self.FTP_adress_input =     customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=500,height=30,corner_radius=0)
        user =                      customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Uživatelské jméno: ",font=("Arial",20,"bold"))
        self.username_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        password =                  customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Heslo: ",font=("Arial",20,"bold"))
        self.password_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =                     customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        self.notes_input =          customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=260)
        self.console =              tk.Text(child_root, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit == None:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disc(child_root),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disc(child_root,True),font=("Arial",20,"bold"),corner_radius=0)
        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        copy_check.             grid(column = 0,row=0,pady = 5,padx =230,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        disc_letter.            grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.disc_letter_input. grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        FTP_adress.             grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        self.FTP_adress_input.  grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        user.                   grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        self.username_input.    grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        password.               grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.password_input.    grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        notes.                  grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=11,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=12,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=13,pady = 5,padx =165,sticky = tk.W)

        if edit == None:
            self.disc_letter_input.delete("0","300")
            self.disc_letter_input.insert("0","P")
            self.FTP_adress_input.delete("0","300")
            self.FTP_adress_input.insert("0","\\\\192.168.000.000\\")
            self.username_input.delete("0","300")
            #self.username_input.insert("0","Vision")
            self.password_input.delete("0","300")
            #self.password_input.insert("0","")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))
        else:
            self.copy_previous_project(True)

        child_root.mainloop()

    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused:
            return True
        else:
            return False

    def change_computer_ip(self,button_row):
        #button_row je id stisknuteho tlacitka... =0 od vrchu
        ip = str(self.all_rows[button_row][1])
        mask = str(self.all_rows[button_row][2])
        # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
        interface_name = str(self.drop_down_options.get())
        powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
        # subprocess.run(["powershell.exe", "-Command", "Start-Process", "powershell.exe", "-Verb", "RunAs", "-ArgumentList", f"'-Command {powershell_command}'"])
        try:
            subprocess.run(["powershell.exe", "-Command",powershell_command],check=True)
            add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla přenastavena na: {ip}","green",None,True)
        except subprocess.CalledProcessError as e:
            add_colored_line(self.main_console,f"Chyba, aplikace musí být spuštěna, jako administrátor. (případně, nemáte tuto adresu již uloženou u jiného připojení?)","red",None,True)

    def check_given_input(self):
        given_data = self.search_input.get()
        if given_data == "":
            found = None
            return found
        found = False

        if self.managing_disc == False:
            for i in range(0,len(self.all_rows)):
                if given_data == self.all_rows[i][0]:
                    self.last_project_name =    str(self.all_rows[i][0])
                    self.last_project_ip =      str(self.all_rows[i][1])
                    self.last_project_mask =    str(self.all_rows[i][2])
                    self.last_project_notes =   str(self.all_rows[i][3])
                    self.last_project_id = i
                    found = True
        else:
            for i in range(0,len(self.disc_all_rows)):
                if given_data == self.disc_all_rows[i][0]:
                    self.last_project_name =        str(self.disc_all_rows[i][0])
                    self.last_project_disc_letter = str(self.disc_all_rows[i][1])
                    self.last_project_ftp =         str(self.disc_all_rows[i][2])
                    self.last_project_username =    str(self.disc_all_rows[i][3])
                    self.last_project_password =    str(self.disc_all_rows[i][4])
                    self.last_project_notes =       str(self.disc_all_rows[i][5])
                    self.last_project_id = i
                    found = True
            
        return found    

    def clicked_on_project(self,e,widget_id,hearth=None):
        self.search_input.delete("0","300")
        if self.managing_disc == False:
            self.search_input.insert("0",str(self.all_rows[widget_id][0]))
        else:
            self.search_input.insert("0",str(self.disc_all_rows[widget_id][0]))

        self.check_given_input()
        if hearth == "favourite":
            add_colored_line(self.main_console,f"Projekt: {self.all_rows[widget_id][0]} byl odebrán z oblíbených","green",None,True)
            self.switch_fav_status("del_favourite")
            #refresh obrazku oblibenosti:
            self.make_project_cells()
        elif hearth == "no_favourite":
            add_colored_line(self.main_console,f"Projekt: {self.all_rows[widget_id][0]} byl přidán do oblíbených","green",None,True)
            self.switch_fav_status("add_favourite")
            self.make_project_cells()

            print("❤️",hearth)
        print(widget_id)

    def is_project_favourite(self,array_index):
        try:
            fav_status = int(self.favourite_list[array_index])
            if fav_status == 1:
                return True
            else:
                return False
        except IndexError:
            print(array_index," index error fav_status")

    def make_project_cells(self,no_read = None):
        if no_read == None:
            self.read_excel_data()
        # padx_list = [10,190,390,390,650]
        padx_list = [60,240,440,440,700]
        self.clear_frame(self.project_tree)
        column1 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        column2 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "IPv4 adresa: ",font=("Arial",20,"bold"))
        column3 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        column1.grid(column = 0,row=0,pady = 5,padx =padx_list[0],sticky = tk.W)
        column2.grid(column = 0,row=0,pady = 5,padx =padx_list[1],sticky = tk.W)
        column3.grid(column = 0,row=0,pady = 5,padx =padx_list[3],sticky = tk.W)
        # y = widgets ve smeru y, x = widgets ve smeru x
        # nejprve vypis oblibene, potom zbytek
        for y in range(0,len(self.all_rows)):
            # ♡,♥,❤️
            project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=50)
            project_frame.grid(row=y+1,column=0,padx=10,sticky=tk.W)
            project_frame.grid_propagate(0)
            
            is_favourite = self.is_project_favourite(y)
            if is_favourite:
                filled_hearth =  customtkinter.CTkLabel(master = project_frame, width = 45,height=45,text = "🐘",font=("Arial",35),text_color="pink")
                filled_hearth.grid(column = 0,row=0,pady = 2,padx =2)
                filled_hearth.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id,"favourite"))
            else:
                unfilled_hearth =  customtkinter.CTkLabel(master = project_frame, width =45,height=45,text = "♡",font=("Arial",40),text_color="red")
                unfilled_hearth.grid(column = 0,row=0,pady = 2,padx =2)
                unfilled_hearth.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id,"no_favourite"))
            
            for x in range(0,len(self.all_rows[y])):
                if x != 2: #nevypisujeme masku
                    if x == 0:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=180)
                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        button =  customtkinter.CTkButton(master = project_frame,width = 160,text = self.all_rows[y][x], command = lambda widget_id = y: self.change_computer_ip(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                        button.grid(column = 0,row=0,pady = 10,padx =10)
                    else:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=200)
                        if x==3: #frame s poznamkami...
                            project_frame.configure(width=750)
                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        parameter =  customtkinter.CTkLabel(master = project_frame,text = self.all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        parameter.grid(column = 0,row=0,pady = 10,padx =10,sticky=tk.W)
    
    def make_project_cells_disc(self,no_read = None):
        if no_read == None:
            self.read_excel_data()
        padx_list = [10,190,240,0,0,640]
        self.clear_frame(self.project_tree)

        column1 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        column2 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
        column3 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        column1.grid(column = 0,row=0,pady = 5,padx =padx_list[0],sticky = tk.W)
        column2.grid(column = 0,row=0,pady = 5,padx =padx_list[2],sticky = tk.W)
        column3.grid(column = 0,row=0,pady = 5,padx =padx_list[5],sticky = tk.W)
        # y = widgets ve smeru y, x = widgets ve smeru x
        for y in range(0,len(self.disc_all_rows)):
            for x in range(0,len(self.disc_all_rows[y])):
                if x == 0:
                    project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=180)
                    project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                    project_frame.grid_propagate(0)
                    # binding the click on widget
                    project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                    button =  customtkinter.CTkButton(master = project_frame,width = 160,text = self.disc_all_rows[y][x], command = lambda widget_id = y: self.map_disc(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                    button.grid(column = 0,row=0,pady = 10,padx =10)
                else:
                    if x != 3 and x != 4:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=400)
                        if x == 1: #frame s písmenem disku
                            project_frame.configure(width=50)
                        if x == 5: #frame s poznamkami...
                            project_frame.configure(width=750)
                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        parameter =  customtkinter.CTkLabel(master = project_frame,text = self.disc_all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        parameter.grid(column = 0,row=0,pady = 10,padx =10,sticky=tk.W)

    def edit_project(self):
        result = self.check_given_input()
        if result == True:
            if self.managing_disc == False:
                self.add_new_project(True)
            else:
                self.add_new_project_disc(True)
        elif result == None:
            add_colored_line(self.main_console,f"Vyberte projekt pro editaci","orange",None,True)
        else:
            add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
    
    def refresh_explorer(self):
        refresh_explorer="taskkill /f /im explorer.exe"
        subprocess.run(refresh_explorer, shell=True)
        refresh_explorer="start explorer.exe"
        subprocess.run(refresh_explorer, shell=True)

    def delete_disc(self,child_root):
        drive_letter = str(self.drive_letter_input.get())
        if len(str(self.DL_manual_entry.get())) > 0:
            drive_letter = str(self.DL_manual_entry.get())
        
        delete_command = "net use " + drive_letter +": /del"
        subprocess.run(delete_command, shell=True)

        self.refresh_explorer()

        add_colored_line(self.main_console,f"Disky s označením {drive_letter} byly odpojeny","orange",None,True)
        self.close_window(child_root)

    def delete_disc_option_menu(self):
        child_root=customtkinter.CTk()
        child_root.geometry("520x200")
        child_root.title("Odpojování síťového disku")
        found_drive_letters=[]

        for i in range(0,len(self.disc_all_rows)):
            if not self.disc_all_rows[i][1] in found_drive_letters:
                found_drive_letters.append(self.disc_all_rows[i][1])
        
        label =                     customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte disk nebo vyhledejte manuálně: ",font=("Arial",20,"bold"))
        self.drive_letter_input =   customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=found_drive_letters,corner_radius=0)
        self.DL_manual_entry =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manuálně")
        del_button =                customtkinter.CTkButton(master = child_root, width = 200,height=30,text = "Odpojit", command = lambda: self.delete_disc(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        
        label.                      grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.drive_letter_input.    grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.DL_manual_entry.       grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        del_button.                 grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)

        child_root.mainloop()

    def map_disc(self,button_row):
        Drive_letter = str(self.disc_all_rows[button_row][1])
        ftp_adress = str(self.disc_all_rows[button_row][2])
        # raw_ftp_address = r"{}".format(ftp_adress)
        # ftp_adress = raw_ftp_address
        
        user = str(self.disc_all_rows[button_row][3])
        password = str(self.disc_all_rows[button_row][4])

        delete_command = "net use " + Drive_letter + ": /del"
        subprocess.run(delete_command, shell=True)
        # second_command = "net use " + Drive_letter + ": " + ftp_adress + " /user:" + user + " " + password + " /persistent:No"
        second_command = "net use " + Drive_letter + ": " + ftp_adress + " " + password + " /user:" + user# + " " + " /persistent:No"
        print(second_command)

        def call_subprocess():
            """process = subprocess.Popen(second_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
            self.connection_status = process.returncode
            print("STDOUT:", stdout)
            print("STDERR:", stderr)
            print("Return Code:", self.connection_status)"""

            self.connection_status = subprocess.call(second_command,shell=True,text=True)

        
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

        time_start = time.time()
        while self.connection_status==None:
            time.sleep(0.05)
            if time.time() - time_start > 3:
                print("terminated due to runtime error")
                break\

        if self.connection_status == 0:
             add_colored_line(self.main_console,f"Disk úspěšně připojen","green",None,True)
             self.refresh_explorer()
        else:
             add_colored_line(self.main_console,f"Připojení selhalo (nesedí vlastní IP adresa? ixon? musí být zvolena alespoň 1 složka...)","red",None,True)

    def get_ipv4_addresses(self):
        # Run the ipconfig command
        result = subprocess.run(['ipconfig'], capture_output=True, text=True)
        # Regular expression to match the IPv4 address
        ipv4_pattern = re.compile(r'IPv4 Address[.\s]*: ([\d.]+)')
        # Dictionary to store interface names and their IPv4 addresses
        ipv4_addresses = []
        # Split the output by lines
        lines = result.stdout.splitlines()
        current_interface = None
        # Iterate over each line to find interface names and IPv4 addresses
        for line in lines:
            if line.strip():
                # Detect interface name
                if line[0].isalpha():
                    current_interface = line.strip()
                else:
                    # Detect IPv4 address for the current interface
                    match = ipv4_pattern.search(line)
                    if match and current_interface:
                        ipv4_addresses.append(current_interface)
                        ipv4_addresses.append(match.group(1))
                        #ipv4_addresses[current_interface] = match.group(1)
        
        return ipv4_addresses

    def option_change(self,args):
        self.default_connection_option = self.connection_option_list.index(self.drop_down_options.get())
        #pamatovat si naposledy zvoleny zpusob pripojeni:
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["Settings"]
        worksheet['B' + str(1)] = int(self.default_connection_option)
        workbook.save(filename=self.excel_file_path)
        workbook.close()
        # ziskat data o aktualnim pripojeni
        current_connection = self.get_ipv4_addresses()
        message = ""
        for items in current_connection:
            message = message + items + " "
        add_colored_line(self.main_console,f"Současné připojení: {message}","white",None,True)

        #ziskat soucasna nastaveni na ruznych pripojeni
        self.get_current_ip_list()
        if  self.static_label2.winfo_exists():
            self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
    
    def make_project_first(self,purpouse=None):
        result = self.check_given_input()
        if result == True:
            #zmena poradi
            project = self.all_rows[self.last_project_id]
            favourite_status = self.favourite_list[self.last_project_id]
            self.all_rows.pop(self.last_project_id)
            self.all_rows.insert(0,project)
            self.favourite_list.pop(self.last_project_id)
            self.favourite_list.insert(0,favourite_status)
            #self.all_rows.append(project)
            #if save == True:
            for i in range(0,len(self.all_rows)):
                row = (len(self.all_rows)-1)-i
                self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=self.favourite_list[i])

            self.make_project_cells()
            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} nalezen","green",None,True)
            elif purpouse != "silent":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} přesunut na začátek","green",None,True)
        elif result == None:
            add_colored_line(self.main_console,f"Nejprve vyberte projekt","orange",None,True)
        else:
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

    def make_project_first_disc(self,purpouse = None):
        result = self.check_given_input()
        if result == True:
            #zmena poradi
            project = self.disc_all_rows[self.last_project_id]
            self.disc_all_rows.pop(self.last_project_id)
            self.disc_all_rows.insert(0,project)

            for i in range(0,len(self.disc_all_rows)):
                row = (len(self.disc_all_rows)-1)-i
                
                self.save_excel_data_disc(self.disc_all_rows[i][0],self.disc_all_rows[i][1],self.disc_all_rows[i][2],self.disc_all_rows[i][3],self.disc_all_rows[i][4],self.disc_all_rows[i][5],None,row+1)

            self.make_project_cells_disc()

            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.disc_all_rows[0][0]} nalezen","green",None,True)
            else:
                add_colored_line(self.main_console,f"Projekt {self.disc_all_rows[0][0]} přesunut na začátek","green",None,True)
        elif result == None:
            add_colored_line(self.main_console,f"Nejprve vyberte projekt","orange",None,True)
        else:
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

    def get_current_ip_address(self,interface_name):
        # Get network interfaces and their addresses
        addresses = psutil.net_if_addrs()
        # Check if the specified interface exists
        if interface_name in addresses:
            for addr in addresses[interface_name]:
                if addr.family == socket.AF_INET:  # IPv4 address
                    return addr.address
        else:
            return "Nenalezeno"

    def get_current_ip_list(self):
        self.current_address_list = []
        for items in self.connection_option_list:
            found_address = self.get_current_ip_address(items)
            self.current_address_list.append(found_address)
    
    def manage_interfaces(self,child_root,given_input,operation = None):
        index =0
        changes_were_made = False
        if operation == "add_new":
            if given_input.replace(" ","") != "":
                self.connection_option_list.insert(0,given_input)
                add_colored_line(self.console,f"Přidán nový interface: {given_input}","green",None,True)
                changes_were_made = True
            else:
                add_colored_line(self.console,"Vložte název","red",None,True)
        
        elif operation == "remove":
            if given_input.replace(" ","") != "":
                try:
                    index = self.connection_option_list.index(given_input)
                    self.connection_option_list.pop(index)
                    add_colored_line(self.console,f"Interface: {given_input} smazán","orange",None,True)
                    changes_were_made = True
                except ValueError:
                    add_colored_line(self.console,"Interface nenalezen","red",None,True)
                
            elif self.interface_input.get() != "":
                given_input = self.interface_input.get()
                index = self.connection_option_list.index(given_input)
                self.connection_option_list.pop(index)
                add_colored_line(self.console,f"Interface: {given_input} smazán","orange",None,True)
                changes_were_made = True
            else:
                add_colored_line(self.console,"Vyberte interface","red",None,True)

        # ulozeni zmen
        if changes_were_made == True:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["Settings"]
            self.default_connection_option = 0
            excel_string_of_options = ""
            worksheet['B' + str(1)] = int(self.default_connection_option)
            for items in self.connection_option_list:
                if items != "":
                    excel_string_of_options = excel_string_of_options + str(items) + ","
            worksheet['B' + str(2)] = excel_string_of_options
            workbook.save(filename=self.excel_file_path)
            workbook.close()
        
        # zvoleni noveho interfacu
        self.drop_down_options.configure(values = self.connection_option_list)
        self.drop_down_options.set(self.connection_option_list[self.default_connection_option])
        self.interface_input.configure(values = self.connection_option_list)
        self.interface_input.set(self.connection_option_list[self.default_connection_option])
        # self.close_window(child_root)

    def connection_option_setting_menu(self):
        self.read_excel_data()
        child_root=customtkinter.CTk()
        child_root.geometry("520x300")
        child_root.title("Nastavení možností připojení (interface list)")

        label =             customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte nebo vyhledejte manuálně: ",font=("Arial",20,"bold"))
        self.interface_input = customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=self.connection_option_list,corner_radius=0)
        manual_entry_interface = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manuálně")
        add_button =        customtkinter.CTkButton(master = child_root, width = 150,height=30,text = "Přidat", command = lambda: self.manage_interfaces(child_root,manual_entry_interface.get(),"add_new"),font=("Arial",20,"bold"),corner_radius=0,fg_color="green")
        del_button =        customtkinter.CTkButton(master = child_root, width = 150,height=30,text = "Smazat", command = lambda: self.manage_interfaces(child_root,manual_entry_interface.get(),"remove"),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        self.console =      tk.Text(child_root, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)
        
        label.              grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.interface_input.grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        manual_entry_interface.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        add_button.         grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        del_button.         grid(column = 0,row=3,pady = 5,padx =170,sticky = tk.W)
        self.console.       grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        
        child_root.mainloop()

    def show_favourite_toggle(self,keep_search_input = False): # hlavni prepinaci tlacitko oblibene/ neoblibene
        if self.show_favourite == True:
            self.show_favourite = False
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            self.show_only_fav.configure(text = "Oblíbené")
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            # self.option_change("")
            # self.make_project_cells()
            self.button_remove_main.configure(command = lambda: self.delete_project())
        else: 
            # favourite window
            self.show_favourite = True
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            self.show_only_fav.configure(text = "Všechny projekty")
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            # self.option_change("")
            
            self.button_remove_main.configure(command = lambda: self.switch_fav_status("with_refresh"))

    def create_widgets(self):
        self.clear_frame(self.root)
        self.managing_disc = False
        main_widgets = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =  customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        main_widgets.pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        # project_tree.grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)

        project_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =     customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=150,height=30,placeholder_text="Název projektu",corner_radius=0)
        button_search =         customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Vyhledat",command =  lambda: self.make_project_first("search"),font=("Arial",16,"bold"),corner_radius=0)
        self.button_add_main =  customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Nový projekt", command = lambda: self.add_new_project(),font=("Arial",16,"bold"),corner_radius=0)
        self.button_remove_main = customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Smazat projekt", command =  lambda: self.delete_project(),font=("Arial",16,"bold"),corner_radius=0)
        self.button_edit_main = customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",16,"bold"),corner_radius=0)
        button_make_first =     customtkinter.CTkButton(master = main_widgets, width = 190,height=30,text = "Přesunout na začátek",command =  lambda: self.make_project_first(),font=("Arial",16,"bold"),corner_radius=0)
        if self.show_favourite:
            self.show_only_fav =    customtkinter.CTkButton(master = main_widgets, width = 190,height=30,text = "Všechny projekty",command =  lambda: self.show_favourite_toggle(),font=("Arial",16,"bold"),corner_radius=0)
        else:
            self.show_only_fav =    customtkinter.CTkButton(master = main_widgets, width = 190,height=30,text = "Oblíbené",command =  lambda: self.show_favourite_toggle(),font=("Arial",16,"bold"),corner_radius=0)

        connect_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=30,text = "Připojení: ",font=("Arial",20,"bold"))
        self.drop_down_options = customtkinter.CTkOptionMenu(master = main_widgets,width=200,height=30,values=self.connection_option_list,font=("Arial",16,"bold"),corner_radius=0,command=  self.option_change)
        # "⚙️", "⚒", "🔧", "🔩"
        button_settings =       customtkinter.CTkButton(master = main_widgets, width = 30,height=30,text="⚒",command =  lambda: self.connection_option_setting_menu(),font=("Arial",22,"bold"),corner_radius=0)
        static_label =          customtkinter.CTkLabel(master = main_widgets, height=30,text = "Static:",font=("Arial",20,"bold"))
        self.static_label2 =    customtkinter.CTkLabel(master = main_widgets, height=30,text = "",font=("Arial",20,"bold"))
        button_change_window =  customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Připojování k síťovým diskům",command =  lambda: self.create_widgets_disc(),font=("Arial",16,"bold"),corner_radius=0,fg_color="green")

        self.main_console = tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",20),state=tk.DISABLED)

        project_label.      grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.  grid(column = 0,row=0,pady = 5,padx =100,sticky = tk.W)
        button_search.      grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        self.button_add_main.grid(column = 0,row=0,pady = 5,padx =360,sticky = tk.W)
        self.button_remove_main.grid(column = 0,row=0,pady = 5,padx =465,sticky = tk.W)
        self.button_edit_main.grid(column = 0,row=0,pady = 5,padx =590,sticky = tk.W)
        button_make_first.  grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)
        self.show_only_fav. grid(column = 0,row=0,pady = 5,padx =915,sticky = tk.W)

        connect_label.          grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.drop_down_options. grid(column = 0,row=1,pady = 0,padx =110,sticky = tk.W)
        button_settings.        grid(column = 0,row=1,pady = 0,padx =315,sticky = tk.W)
        static_label.           grid(column = 0,row=1,pady = 0,padx =355,sticky = tk.W)
        self.static_label2.     grid(column = 0,row=1,pady = 0,padx =420,sticky = tk.W)
        button_change_window.   grid(column = 0,row=1,pady = 0,padx =590,sticky = tk.W)
        
        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        
        self.drop_down_options.set(self.connection_option_list[self.default_connection_option])
        self.option_change("")

        self.make_project_cells(True)
        self.get_current_ip_list()
        self.static_label2.configure(text=self.current_address_list[self.default_connection_option])

        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                #self.root.after(0, lambda:self.root.state('normal'))
                self.root.state('normal')
                self.root.geometry("210x500")
            elif int(current_width) ==210:
                self.root.geometry("1200x900")
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')
        self.root.bind("<f>",maximalize_window)

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first("search")
        self.search_input.bind("<Return>",call_search)

    def create_widgets_disc(self):
        self.clear_frame(self.root)
        self.managing_disc = True
        main_widgets = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =  customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        main_widgets.pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        # project_tree.grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)

        project_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =     customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=150,height=30,placeholder_text="Název projektu",corner_radius=0)
        button_search =         customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Vyhledat",command =  lambda: self.make_project_first_disc("search"),font=("Arial",16,"bold"),corner_radius=0)
        button_add =            customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Nový projekt", command = lambda: self.add_new_project_disc(),font=("Arial",16,"bold"),corner_radius=0)
        button_remove =         customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Smazat projekt", command =  lambda: self.delete_project_disc(),font=("Arial",16,"bold"),corner_radius=0)
        button_edit =           customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",16,"bold"),corner_radius=0)
        button_make_first =     customtkinter.CTkButton(master = main_widgets, width = 200,height=30,text = "Přesunout na začátek",command =  lambda: self.make_project_first_disc(),font=("Arial",16,"bold"),corner_radius=0)
        
        button_change_window = customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Měnit IP adresu",command =  lambda: self.create_widgets(),font=("Arial",16,"bold"),corner_radius=0,fg_color="green")
        delete_disc          = customtkinter.CTkButton(master = main_widgets, width = 100,height=30,text = "Odpojit síťový disk",command =  lambda: self.delete_disc_option_menu(),font=("Arial",16,"bold"),corner_radius=0,fg_color="red")

        self.main_console = tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",20),state=tk.DISABLED)

        project_label.      grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.  grid(column = 0,row=0,pady = 5,padx =100,sticky = tk.W)
        button_search.      grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        button_add.         grid(column = 0,row=0,pady = 5,padx =360,sticky = tk.W)
        button_remove.      grid(column = 0,row=0,pady = 5,padx =465,sticky = tk.W)
        button_edit.        grid(column = 0,row=0,pady = 5,padx =590,sticky = tk.W)
        button_make_first.  grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)

        button_change_window.grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        delete_disc         .grid(column = 0,row=1,pady = 5,padx =140,sticky = tk.W)
        
        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.option_change("")
        self.make_project_cells_disc(True)

        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                #self.root.after(0, lambda:self.root.state('normal'))
                self.root.state('normal')
                self.root.geometry("210x500")
            elif int(current_width) ==210:
                self.root.geometry("1200x900")
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')
        self.root.bind("<f>",maximalize_window)

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first_disc("search")
        self.search_input.bind("<Return>",call_search)

IP_assignment(root)
root.mainloop()
