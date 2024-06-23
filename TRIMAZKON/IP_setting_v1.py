import customtkinter
import tkinter as tk
from openpyxl import load_workbook
import subprocess
import os
import re
import time
import threading
import psutil
import socket
import win32api
import win32file
from PIL import Image
import sys
import ctypes

def path_check(path_raw,only_repair = None):
    path=path_raw
    backslash = "\ "
    if backslash[0] in path:
        newPath = path.replace(backslash[0], '/')
        path = newPath
    if path.endswith('/') == False:
        newPath = path + "/"
        path = newPath
    #oprava mezery v nazvu
    path = r"{}".format(path)
    if not os.path.exists(path) and only_repair == None:
        return False
    else:
        return path

def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def add_colored_line(text_widget, text, color,font=None,delete_line = None):
    """
    Vlo≈æ√≠ ≈ô√°dek do console
    """
    text_widget.configure(state=tk.NORMAL)
    if font == None:
        font = ("Arial",16)
    if delete_line != None:
        text_widget.delete("current linestart","current lineend")
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert("current lineend",text, color)
    else:
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert(tk.END,"    > "+ text+"\n", color)

    text_widget.configure(state=tk.DISABLED)
    
def check_network_drive_status(drive_path):
    global checking_done
    checking_done = False
    try:
        # Attempt to access a file or directory on the network drive
        # drive_path = drive_letter + "\\"
        # print("drive_path",drive_path)
        drive_path = drive_path[0:3]
        def call_subprocess():
            global checking_done
            if os.path.exists(drive_path):
                os.listdir(drive_path)
                checking_done = True
                return True
            else:
                checking_done = True
                return False
            
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

        time_start = time.time()
        while checking_done==False:
            time.sleep(0.05)
            if time.time() - time_start > 1:
                print("terminated due to runtime error")
                return False
            
        run_background.join()
        return True

    except FileNotFoundError:
        return False
    except OSError:
        return False

def list_mapped_disks(whole_format=None):
    drives = win32api.GetLogicalDriveStrings()
    print("drives",drives)
    drives = drives.split('\000')[:-1]
    remote_drives = []
    for drive in drives:
        if win32file.GetDriveType(drive) == win32file.DRIVE_REMOTE:
            if whole_format:
                remote_drives.append(drive)
            else:
                remote_drives.append(drive[0:1])
    return remote_drives

class IP_assignment: # Umo≈æ≈àuje mƒõnit statickou IP a mountit disky
    """
    Umo≈æ≈àuje mƒõnit nastaven√≠ statick√Ωch IP adres
    """

    def __init__(self,root,callback_function,window_mode,initial_path):
        self.initial_path = initial_path
        self.window_mode = window_mode
        self.callback = callback_function
        self.root = root
        self.app_icon = 'images/logo_TRIMAZKON.ico'
        self.rows_taken = 0
        self.all_rows = []
        self.project_list = []
        # self.app_path = os.getcwd()
        # self.app_path = path_check(self.app_path,True)
        # self.excel_file_path = self.app_path + "saved_addresses_2.xlsx"
        self.excel_file_path = initial_path + "saved_addresses_2.xlsx"
        # print(self.excel_file_path)
        self.last_project_name = ""
        self.last_project_ip = ""
        self.last_project_mask = ""
        self.last_project_notes = ""
        self.last_project_id = ""
        self.last_project_disk_letter = ""
        self.last_project_ftp = ""
        self.last_project_username = ""
        self.last_project_password = ""
        self.managing_disk = False
        self.connection_status = None
        self.make_project_favourite = False
        self.favourite_list = []
        self.default_connection_option = 0
        self.connection_option_list = []
        self.default_disk_status_behav = 0
        try:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["Settings"]
            # z worksheetu nastaven√≠ ƒçtu z√°kladn√≠ zvolen√Ω interface p≈ôipojen√≠
            # - v≈°echny mo≈æn√© intefaces
            # - defaultn√≠ okno zobrazen√≠ (obl√≠ben√©/ v≈°echny/ disky)
            # - defaultn√≠ velikost okna - pamatuje si nejmen≈°√≠ zvolenou
            saved_def_con_option = worksheet['B' + str(1)].value
            self.default_connection_option = int(saved_def_con_option)

            # self.connection_option_list = []
            # all_options = worksheet['B' + str(2)].value
            # all_options = str(all_options).split(",")
            # for i in range (0,len(all_options)):
            #     if all_options[i] != "":
            #         self.connection_option_list.append(all_options[i])

            def_show_favourite = worksheet['B' + str(3)].value
            if int(def_show_favourite) == 1:
                self.show_favourite = True
            else:
                self.show_favourite = False
            
            def_show_disk = worksheet['B' + str(4)].value
            if int(def_show_disk) == 1:
                self.create_widgets_disk(init=True)
            else:
                self.create_widgets(init=True)

            def_window_size = worksheet['B' + str(5)].value
            if def_window_size == 2:
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")

            self.default_disk_status_behav = int(worksheet['B' + str(6)].value)
            workbook.close()
        except Exception as e:
            self.connection_option_list = ["data nenalezena"]
            self.show_favourite = False
            self.create_widgets(init=True,excel_load_error=True)
            print(f"Nejd≈ô√≠v zav≈ôete soubor {self.excel_file_path} Chyba: {e}")

    def call_menu(self): # Tlaƒç√≠tko menu (konec, n√°vrat do menu)
        """
        Funkce ƒçist√≠ v≈°echny zaplnƒõn√© r√°meƒçky a funguje, jako tlaƒç√≠tko zpƒõt do hlavn√≠ho menu trimazkonu
        """
        self.clear_frame(self.root)
        self.callback()

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def close_window(self,window):
        window.update_idletasks()
        window.destroy()

    def fill_interfaces(self):
        """
        Vr√°t√≠:
        - seznam interfac≈Ø [0]
        - seznam p≈ôipojen√Ωch interfac≈Ø [1]
        """
        process = subprocess.Popen("netsh interface show interface",
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        try:
            stdout_str = stdout.decode('utf-8')
            data = str(stdout_str)
        except UnicodeDecodeError:
            try:
                stdout_str = stdout.decode('cp1250')
                data = str(stdout_str)
            except UnicodeDecodeError:
                data = str(stdout)
            
        print("netsh data: ", data)
        lines = data.strip().splitlines()
        interfaces = []
        interface_statuses =[]
        # Process each line starting from the second line
        for line in lines[2:]:  # Skip the first two lines (headers and separator)
            # Split the line based on spaces
            values = line.split()
            if len(values) >= 4:  # Ensure there are enough columns
                # Combine the last columns into Interface Name, handle spaces in the name
                interface_name = ' '.join(values[3:])
                interface_statuses.append(values[1])
                interfaces.append(interface_name)

        print(interfaces)
        print("status: ", interface_statuses)
        connected_interfaces =[]
        for items in interface_statuses:
            if items != "Odpojen" and items != "Odpojeno" and items != "Disconnected":
                if interfaces[interface_statuses.index(items)] not in connected_interfaces:
                    connected_interfaces.append(interfaces[interface_statuses.index(items)])
        print("online: ", connected_interfaces)
        
        return [interfaces,connected_interfaces]

    def read_excel_data(self):
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"

        workbook = load_workbook(self.excel_file_path)
        # seznam vsech statickych ip adres
        self.all_rows = []
        self.project_list = []
        self.favourite_list = []
        worksheet = workbook[excel_worksheet]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:4]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            if len(row_array) < 4:
                row_array.append("")
            self.project_list.insert(0,row_array[0])
            self.all_rows.insert(0,row_array)
            for items in row[4:5]:
                self.favourite_list.insert(0,items)
            

        # seznam vsech ftp pripojeni k diskum
        self.disk_all_rows = []
        self.disk_project_list = []  
        worksheet = workbook["disk_list"]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:6]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            """if len(row_array) < 4:
                row_array.append("")"""
            self.disk_project_list.insert(0,row_array[0])
            self.disk_all_rows.insert(0,row_array)

        # ukladani nastavenych hodnot
        worksheet = workbook["Settings"]
        saved_def_con_option = worksheet['B' + str(1)].value
        self.default_connection_option = int(saved_def_con_option)

        self.default_disk_status_behav = int(worksheet['B' + str(6)].value)
        # seznam interfac≈Ø
        # self.fill_interfaces()
        # self.connection_option_list = []
        # all_options = worksheet['B' + str(2)].value
        # all_options = str(all_options).split(",")
        # for i in range (0,len(all_options)):
        #     if all_options[i] != "":
        #         self.connection_option_list.append(all_options[i])
        workbook.close()
                     
    def save_excel_data(self,project_name,IP_adress,mask,notes,only_edit = None,force_row_to_print=None,fav_status = None):
        workbook = load_workbook(self.excel_file_path)
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        worksheet = workbook[excel_worksheet]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.all_rows)) +1
            if only_edit != None:
                #pouze zmƒõna na temt√Ω≈æ ≈ô√°dku
                row_to_print = (len(self.all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        if notes.replace(" ","") == "" or notes == None:
            notes = ""
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = ip adresa
        worksheet['B' + str(row_to_print)] = IP_adress
        #C = maska
        worksheet['C' + str(row_to_print)] = mask
        #D = poznamky
        worksheet['D' + str(row_to_print)] = notes
        #E = oblibenost
        if fav_status != None:
            worksheet['E' + str(row_to_print)] = fav_status

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def save_excel_data_disk(self,project_name,disk_letter,ftp_address,username,password,notes,only_edit = None,force_row_to_print=None):
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["disk_list"]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.disk_all_rows)) +1
            if only_edit != None:
                #pouze zmƒõna na temt√Ω≈æ ≈ô√°dku
                row_to_print = (len(self.disk_all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = p√≠smeno disku, oznaƒçen√≠...
        worksheet['B' + str(row_to_print)] = disk_letter
        #C = ftp adresa
        worksheet['C' + str(row_to_print)] = ftp_address
        #D = u≈æivatelsk√© jm√©no
        worksheet['D' + str(row_to_print)] = username
        #E = heslo
        worksheet['E' + str(row_to_print)] = password
        #F = poznamky
        worksheet['F' + str(row_to_print)] = notes

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def get_notes(self):
        notes_legit_rows = []
        notes = str(self.notes_input.get("1.0", tk.END))
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    def check_ip_and_mask(self,input_value):
        input_splitted = input_value.split(".")
        if len(input_splitted) == 4:
            return input_value
        else:
            return False

    def switch_fav_status(self,operation:str,project_given=None,change_status = False):
        # selected_project = self.all_rows[self.last_project_id]
        if project_given == None:
            selected_project = str(self.search_input.get())
            if selected_project not in self.project_list:
                add_colored_line(self.main_console,"Nebyl vlo≈æen projekt",color="red",font=None,delete_line=True)
                return
            else:
                selected_project = self.all_rows[self.project_list.index(selected_project)]
        else:
            selected_project = project_given
        if self.show_favourite == False:
            if operation == "add_favourite":
                if change_status:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=1)
                self.show_favourite = True
                self.read_excel_data()
                # do tohoto prost≈ôed√≠ ulo≈æ√≠m na zaƒç√°tek
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # p≈ôepnut√≠ zpƒõt
                self.show_favourite = False
                self.read_excel_data()
            
            elif operation == "del_favourite":
                if change_status:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=0)
                # p≈ôepnut√≠
                self.show_favourite = True
                self.read_excel_data()
                # z tohoto prost≈ôed√≠ sma≈æu
                self.delete_project(wanted_project=selected_project[0],silence=True)
                # p≈ôepnut√≠ zpƒõt
                self.show_favourite = False
                self.read_excel_data()

            elif operation == "rewrite_favourite":
                # p≈ôepnut√≠
                self.show_favourite = True
                self.read_excel_data()
                # nejprve popnu stary projekt, s povodnim jmenem
                # pot√© insertnu pozmƒõnƒõn√Ω
                the_id_to_pop = self.project_list.index(self.last_project_name)
                self.all_rows.pop(the_id_to_pop)
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # p≈ôepnut√≠ zpƒõt
                self.show_favourite = False
                self.read_excel_data()

        elif self.show_favourite:
            # z aktu√°ln√≠ho prost≈ôed√≠ sma≈æu
            self.delete_project(wanted_project=selected_project[0],silence=True)
            # musim prepnout prost≈ôed√≠ jen kv≈Øli zmƒõnƒõ statusu
            self.show_favourite = False
            self.read_excel_data()
            match_found = False
            for i in range(0,len(self.project_list)):
                if self.project_list[i] == selected_project[0] and len(str(self.project_list[i])) == len(str(selected_project[0])):
                    row_index = self.project_list.index(selected_project[0])
                    match_found = True
            if match_found:
                row = len(self.all_rows) - row_index
                self.save_excel_data(self.all_rows[row_index][0],self.all_rows[row_index][1],self.all_rows[row_index][2],self.all_rows[row_index][3],None,row,fav_status=0)

            # p≈ôepnut√≠ zpƒõt
            self.show_favourite = True
            self.read_excel_data()
        
        if operation == "with_refresh":
            add_colored_line(self.main_console,f"Projekt: {selected_project[0]} byl odebr√°n z obl√≠ben√Ωch","green",None,True)
            self.make_project_cells(no_read=True)

    def save_new_project_data(self,child_root,only_edit = None,make_fav=False):
        project_name = str(self.name_input.get())
        IP_adress = str(self.IP_adress_input.get())
        IP_adress = self.check_ip_and_mask(IP_adress)
        mask = str(self.mask_input.get())
        mask = self.check_ip_and_mask(mask)
        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jm√©no projektu","red",None,True)
            errors += 1
        if IP_adress == False and errors == 0:
            add_colored_line(self.console,f"Neplatn√° IP adresa","red",None,True)
            errors += 1
        if mask == False and errors == 0:
            add_colored_line(self.console,f"Neplatn√° maska","red",None,True)
            errors += 1
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()

            if only_edit == None: # pridavam novy projekt
                if make_fav:
                    new_project = [project_name,IP_adress,mask,notes,1]
                    self.switch_fav_status("add_favourite",new_project)
                    add_colored_line(self.main_console,f"P≈ôid√°n nov√Ω obl√≠ben√Ω projekt: {project_name}","green",None,True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=None,force_row_to_print=None,fav_status=1)                
                else:
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=None,force_row_to_print=None,fav_status=0)
                    add_colored_line(self.main_console,f"P≈ôid√°n nov√Ω projekt: {project_name}","green",None,True)
            else:
                # kdyz edituji muze mit projekt jiz prideleny status
                current_fav_status = self.is_project_favourite(self.last_project_id)
                if make_fav and current_fav_status == 0:
                    project_with_changes = [project_name,IP_adress,mask,notes,current_fav_status]
                    self.switch_fav_status("add_favourite",project_with_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} √∫spƒõ≈°nƒõ pozmƒõnƒõn a p≈ôid√°n do obl√≠ben√Ωch","green",None,True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=1)
                elif make_fav == False and current_fav_status == 1:
                    project_without_changes = self.all_rows[self.last_project_id]
                    self.switch_fav_status("del_favourite",project_without_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} √∫spƒõ≈°nƒõ pozmƒõnƒõn a odebr√°n z obl√≠ben√Ωch","green",None,True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=0)
                elif make_fav and current_fav_status == 1:
                    #nedoslo ke zmene statusu, ale mohlo dojit ke zmene - proto prepsat v oblibenych - vzdy se jedna o oblibene...
                    project_with_changes = [project_name,IP_adress,mask,notes,current_fav_status]
                    self.switch_fav_status("rewrite_favourite",project_with_changes)
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} √∫spƒõ≈°nƒõ pozmƒõnƒõn","green",None,True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=current_fav_status)
                else:
                    add_colored_line(self.main_console,f"Projekt: {self.last_project_name} √∫spƒõ≈°nƒõ pozmƒõnƒõn","green",None,True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=current_fav_status)

            self.close_window(child_root)
            self.make_project_cells()
    
    def save_new_project_data_disk(self,child_root,only_edit = None):
        project_name =  str(self.name_input.get())
        disk_letter =   str(self.disk_letter_input.get())
        ftp_address =   str(self.FTP_adress_input.get())
        username =      str(self.username_input.get())
        password =      str(self.password_input.get())

        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jm√©no projektu","red",None,True)
            errors += 1
        elif disk_letter.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste p√≠smeno disku","red",None,True)
            errors += 1
        elif ftp_address.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste adresu","red",None,True)
            errors += 1
        # elif username.replace(" ","") == "":
        #     add_colored_line(self.console,f"Nezadali jste p≈ôihla≈°ovac√≠ jm√©no","red",None,True)
        #     errors += 1
        # elif password.replace(" ","") == "":
        #     add_colored_line(self.console,f"Nezadali jste p≈ôihla≈°ovac√≠ heslo","red",None,True)
        #     errors += 1
        
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()
            if only_edit == None:
                self.save_excel_data_disk(project_name,disk_letter,ftp_address,username,password,notes)
            else:
                self.save_excel_data_disk(project_name,disk_letter,ftp_address,username,password,notes,True)
            self.close_window(child_root)
            if only_edit == None:
                self.make_project_cells_disk()
                add_colored_line(self.main_console,f"P≈ôid√°n nov√Ω projekt: {project_name}","green",None,True)
            else: #musi byt proveden reset
                self.make_project_cells_disk()
                add_colored_line(self.main_console,f"Projekt: {project_name} √∫spƒõ≈°nƒõ pozmƒõnƒõn","green",None,True)

    def delete_project(self,wanted_project=None,silence=None):
        remove_favourite_as_well = False
        if wanted_project == None:
            self.read_excel_data()
            wanted_project = str(self.search_input.get())
        project_found = False
        workbook = load_workbook(self.excel_file_path)
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        worksheet = workbook[excel_worksheet]

        for i in range(0,len(self.project_list)):
            if self.project_list[i] == wanted_project and len(str(self.project_list[i])) == len(str(wanted_project)) and project_found == False:
                row_index = self.project_list.index(wanted_project)
                worksheet.delete_rows(len(self.all_rows)-row_index)
                project_found = True
                #pokud ma status oblibenosti, tak vymazat i z oblibenych:
                if self.favourite_list[row_index] == 1 and self.show_favourite == False:
                    remove_favourite_as_well = True
                    deleted_project = self.all_rows[row_index]
            workbook.save(self.excel_file_path)
            workbook.close()

        if silence == None:
            if project_found:
                add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstranƒõn","orange",None,True)
                self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
            elif wanted_project.replace(" ","") == "":
                add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout lev√Ωm na parametry dan√©ho projektu nebo prav√Ωm na tlaƒç√≠ko projektu)","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Zadan√Ω projekt: {wanted_project} nebyl nalezen","red",None,True)
        
        if remove_favourite_as_well:
            self.switch_fav_status("del_favourite",deleted_project)

    def delete_project_disk(self):
        self.read_excel_data()
        wanted_project = str(self.search_input.get())
        project_found = False
        if wanted_project.replace(" ","") != "":
            for i in range(0,len(self.disk_project_list)):
                if self.disk_project_list[i] == wanted_project and len(str(self.disk_project_list[i])) == len(str(wanted_project)):
                    row_index = self.disk_project_list.index(wanted_project)
                    workbook = load_workbook(self.excel_file_path)
                    worksheet = workbook["disk_list"]
                    worksheet.delete_rows(len(self.disk_all_rows)-row_index)
                    workbook.save(self.excel_file_path)
                    workbook.close()
                    self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                    project_found = True
                    add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstranƒõn","orange",None,True)
                    break
            if project_found == False:
                add_colored_line(self.main_console,f"Zadan√Ω projekt: {wanted_project} nebyl nalezen","red",None,True)
        else:
            add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout lev√Ωm na parametry dan√©ho projektu nebo prav√Ωm na tlaƒç√≠ko projektu)","orange",None,True)

    def copy_previous_project(self,disk=None):
        if self.last_project_name == "":
            add_colored_line(self.console,"Nen√≠ vybr√°n ≈æ√°dn√Ω projekt","red",None,True)
        else:
            self.name_input.delete("0","300")
            self.name_input.insert("0",str(self.last_project_name))
            if disk == None:
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0",str(self.last_project_ip))
                self.mask_input.delete("0","300")
                self.mask_input.insert("0",str(self.last_project_mask))
                self.notes_input.insert(tk.END,str(self.last_project_notes))
            else:
                self.disk_letter_input.delete("0","300")
                self.disk_letter_input.insert("0",str(self.last_project_disk_letter))
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0",str(self.last_project_ftp))
                self.username_input.delete("0","300")
                self.username_input.insert("0",str(self.last_project_username))
                self.password_input.delete("0","300")
                self.password_input.insert("0",str(self.last_project_password))
                self.notes_input.insert(tk.END,str(self.last_project_notes))

    def make_favourite_toggle_via_edit(self,e):
        def do_favourite():
            self.make_fav_btn.configure(text = "üêò",font=("Arial",130),text_color = "pink")
            self.make_fav_label.configure(text = "Obl√≠ben√Ω ‚ù§Ô∏è")
        
        def unfavourite():
            self.make_fav_btn.configure(text = "‚ùå",font=("Arial",100),text_color = "red")
            self.make_fav_label.configure(text = "Neobl√≠ben√Ω")

        if self.make_project_favourite:
            self.make_project_favourite = False
            unfavourite()
        else:
            self.make_project_favourite = True
            do_favourite()

    def add_new_project(self,edit = None):
        if self.show_favourite:
            #p≈ôepnut√≠ do hlavn√≠ho prost≈ôed√≠
            self.show_favourite_toggle(True)
        child_root=customtkinter.CTk()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"520x750+{x+50}+{y+80}")
        # child_root.wm_iconbitmap(self.initial_path+'images/logo_TRIMAZKON.ico')
        child_root.wm_iconbitmap(resource_path(self.app_icon))
        # child_root.geometry("520x750")
        if edit:
            child_root.title("Editovat projekt: "+self.last_project_name)
        else:
            child_root.title("Nov√Ω projekt")
        
        project_name =    customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "N√°zev projektu: ",font=("Arial",20,"bold"))
        copy_check =      customtkinter.CTkButton(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,text="Kop√≠rovat p≈ôedchoz√≠ projekt",command= lambda: self.copy_previous_project())
        self.name_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        
        if edit:
            if self.is_project_favourite(self.last_project_id):
                self.make_project_favourite = True #init hodnota
                self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Obl√≠ben√Ω ‚ù§Ô∏è",font=("Arial",20,"bold"))
                fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
                self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "üêò",font=("Arial",130),text_color = "pink")
            else:
                self.make_project_favourite = False #init hodnota
                self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Neobl√≠ben√Ω",font=("Arial",20,"bold"))
                fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
                self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "‚ùå",font=("Arial",100),text_color = "red")
        else: # defaultne neoblibeny
            self.make_project_favourite = False #init hodnota
            self.make_fav_label =   customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Neobl√≠ben√Ω",font=("Arial",20,"bold"))
            fav_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=150,width=150)
            self.make_fav_btn =     customtkinter.CTkLabel(master = fav_frame, width = 150,height=150,text = "‚ùå",font=("Arial",100),text_color = "red")

        IP_adress =            customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
        self.IP_adress_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        mask =                 customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
        self.mask_input =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =                customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Pozn√°mky: ",font=("Arial",20,"bold"))
        self.notes_input =     customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=370)
        self.console =         tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Ulo≈æit", command = lambda: self.save_new_project_data(child_root,True,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Ulo≈æit", command = lambda: self.save_new_project_data(child_root,None,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)

        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        copy_check.             grid(column = 0,row=0,pady = 5,padx =240,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.make_fav_label.    grid(column = 0,row=1,pady = 5,padx =240,sticky = tk.W)
        IP_adress.              grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        fav_frame.              grid(row=3,column=0,padx=240,sticky=tk.W,rowspan=4)
        fav_frame.              grid_propagate(0)
        self.make_fav_btn.      grid(column=0,row=0)
        self.make_fav_btn.      bind("<Button-1>",lambda e: self.make_favourite_toggle_via_edit(e))
        self.IP_adress_input.   grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        mask.                   grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        self.mask_input.        grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        notes.                  grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=10,pady = 5,padx =165,sticky = tk.W)

        if edit:
            self.copy_previous_project()
        else:
            self.IP_adress_input.delete("0","300")
            self.IP_adress_input.insert("0","192.168.000.000")
            self.mask_input.delete("0","300")
            self.mask_input.insert("0","255.255.255.0")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))


        child_root.mainloop()

    def add_new_project_disk(self,edit = None):
        child_root=customtkinter.CTk()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"520x800+{x+50}+{y+100}")
        # child_root.wm_iconbitmap(self.initial_path+'images/logo_TRIMAZKON.ico')
        child_root.wm_iconbitmap(resource_path(self.app_icon))
        # child_root.geometry("520x800")
        if edit == None:
            child_root.title("Nov√Ω projekt")
        else:
            child_root.title("Editovat projekt: "+self.last_project_name)

        project_name =              customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "N√°zev projektu: ",font=("Arial",20,"bold"))
        copy_check =                customtkinter.CTkButton(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,text="Kop√≠rovat p≈ôedchoz√≠ projekt",command= lambda: self.copy_previous_project(True))
        self.name_input =           customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        disk_letter =               customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "P√≠smeno disku: ",font=("Arial",20,"bold"))
        self.disk_letter_input =    customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        FTP_adress =                customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
        self.FTP_adress_input =     customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=500,height=30,corner_radius=0)
        user =                      customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "U≈æivatelsk√© jm√©no: ",font=("Arial",20,"bold"))
        self.username_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        password =                  customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Heslo: ",font=("Arial",20,"bold"))
        self.password_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =                     customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Pozn√°mky: ",font=("Arial",20,"bold"))
        self.notes_input =          customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=260)
        self.console =              tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit == None:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Ulo≈æit", command = lambda: self.save_new_project_data_disk(child_root),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Ulo≈æit", command = lambda: self.save_new_project_data_disk(child_root,True),font=("Arial",20,"bold"),corner_radius=0)
        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        copy_check.             grid(column = 0,row=0,pady = 5,padx =230,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        disk_letter.            grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.disk_letter_input. grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        FTP_adress.             grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        self.FTP_adress_input.  grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        user.                   grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        self.username_input.    grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        password.               grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.password_input.    grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        notes.                  grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=11,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=12,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=13,pady = 5,padx =165,sticky = tk.W)

        if edit == None:
            self.disk_letter_input.delete("0","300")
            self.disk_letter_input.insert("0","P")
            self.FTP_adress_input.delete("0","300")
            self.FTP_adress_input.insert("0","\\\\192.168.000.000\\")
            self.username_input.delete("0","300")
            #self.username_input.insert("0","Vision")
            self.password_input.delete("0","300")
            #self.password_input.insert("0","")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))
        else:
            self.copy_previous_project(True)

        child_root.mainloop()

    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused or ".!ctktextbox" in currently_focused:
            return True
        else:
            return False

    def make_sure_ip_changed(self,interface_name,ip,command):
        def close_prompt(child_root):
            self.close_window(child_root)
        def run_as_admin(child_root):
            # Vy≈æ√°d√°n√≠ admin pr√°v: nefunkƒçn√≠ ve vscode
            self.close_window(child_root)
            def is_admin():
                try:
                    return ctypes.windll.shell32.IsUserAnAdmin()
                except:
                    return False
            if not is_admin():
                ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
                sys.exit()
        def open_app_as_admin_prompt():
            child_root=customtkinter.CTk()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"620x150+{x+300}+{y+300}")  
            child_root.title("Upozornƒõn√≠")
            child_root.wm_iconbitmap(resource_path(self.app_icon))
            proceed_label = customtkinter.CTkLabel(master = child_root,text = "P≈ôejete si znovu spustit aplikaci, jako administr√°tor?",font=("Arial",25))
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin(child_root))
            button_no =     customtkinter.CTkButton(master = child_root,text = "Zru≈°it",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
            proceed_label   .pack(pady=(15,0),padx=10,anchor="w",expand=False,side = "top")
            button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            child_root.mainloop()

        interface_index = self.connection_option_list.index(interface_name)
        def call_subprocess():
            if ip == self.current_address_list[interface_index]:
                add_colored_line(self.main_console,f"Pro interface {interface_name} je ji≈æ tato adresa ({ip}) nastavena","orange",None,True)
                return
            elif ip in self.current_address_list:
                add_colored_line(self.main_console,f"Chyba, adresa je ji≈æ pou≈æ√≠v√°na pro jin√Ω interface","red",None,True)
                return
            win_change_ip_time = 7
            for i in range(0,win_change_ip_time):
                add_colored_line(self.main_console,f"ƒåek√°m, a≈æ windows provede zmƒõny: {7-i} s...","white",None,True)
                self.option_change("",silent=True)
                time.sleep(1)

            self.option_change("",silent=True)
            if ip == self.current_address_list[interface_index]:
                add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla p≈ôenastavena na: {ip}","green",None,True)
                self.make_project_cells(no_read=True)
            else:
                add_colored_line(self.main_console,f"Chyba, neplatn√° adresa nebo dan√Ω inteface odpojen od tohoto za≈ô√≠zen√≠ (pro nastavov√°n√≠ odpojen√Ωch interfac≈Ø spus≈•tƒõ aplikaci jako administr√°tor)","red",None,True)
                open_app_as_admin_prompt()

        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

    def change_computer_ip(self,button_row):
        def connected_interface(interface,ip,mask):
            try:
                # Construct the netsh command
                netsh_command = f"netsh interface ip set address \"{interface}\" static {ip} {mask}"
                powershell_command = [
                    'powershell.exe',
                    '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                ]
                process = subprocess.Popen(powershell_command,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                
                stdout, stderr = process.communicate()
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                if stderr_str:
                    print(f"Error occurred: {stderr_str}")
                else:
                    print(f"Command executed successfully:\n{stdout_str}")

            except Exception as e:
                print(f"Exception occurred: {str(e)}")


            self.make_sure_ip_changed(interface_name,ip,"")
        
        ip = str(self.all_rows[button_row][1])
        mask = str(self.all_rows[button_row][2])
        # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
        interface_name = str(self.drop_down_options.get())
        powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
        try:
            process = subprocess.Popen(['powershell.exe', '-Command', powershell_command],
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr =process.communicate()
            stdout_str = stdout.decode('utf-8')
            stderr_str = stderr.decode('utf-8')
            print(f"V√ùSTUP Z IP SETTING: {str(stdout_str)}")

            if len(str(stdout_str)) > 7:
                raise subprocess.CalledProcessError(1, powershell_command, stdout_str)
            if stderr_str:
                raise subprocess.CalledProcessError(1, powershell_command, stderr_str)

            # add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla p≈ôenastavena na: {ip}","green",None,True)
            self.make_sure_ip_changed(interface_name,ip,"")

        except subprocess.CalledProcessError as e:
            if "Run as administrator" in str(stdout_str):
                add_colored_line(self.main_console,f"Chyba, tato funkce mus√≠ b√Ωt spu≈°tƒõna s administr√°torsk√Ωmi pr√°vy","red",None,True)
                connected_interface(interface_name,ip,mask)
                
            elif "Invalid address" in str(stdout_str):
                add_colored_line(self.main_console,f"Chyba, neplatn√° IP adresa","red",None,True)
            else:
                add_colored_line(self.main_console,f"Chyba, Nem√°te tuto adresu ji≈æ nastavenou pro jin√Ω interface? (nebo dan√Ω interface na tomto za≈ô√≠zen√≠ neexistuje)","red",None,True)
        except Exception as e:
            # Handle any other exceptions that may occur
            add_colored_line(self.main_console, f"Nastala neoƒçek√°van√° chyba: {e}", "red", None, True)

    def check_given_input(self):
        given_data = self.search_input.get()
        if given_data == "":
            found = None
            return found
        found = False

        if self.managing_disk == False:
            for i in range(0,len(self.all_rows)):
                if given_data == self.all_rows[i][0]:
                    self.last_project_name =    str(self.all_rows[i][0])
                    self.last_project_ip =      str(self.all_rows[i][1])
                    self.last_project_mask =    str(self.all_rows[i][2])
                    self.last_project_notes =   str(self.all_rows[i][3])
                    self.last_project_id = i
                    found = True
        else:
            for i in range(0,len(self.disk_all_rows)):
                if given_data == self.disk_all_rows[i][0]:
                    self.last_project_name =        str(self.disk_all_rows[i][0])
                    self.last_project_disk_letter = str(self.disk_all_rows[i][1])
                    self.last_project_ftp =         str(self.disk_all_rows[i][2])
                    self.last_project_username =    str(self.disk_all_rows[i][3])
                    self.last_project_password =    str(self.disk_all_rows[i][4])
                    self.last_project_notes =       str(self.disk_all_rows[i][5])
                    self.last_project_id = i
                    found = True
            
        return found    

    def clicked_on_project(self,e,widget_id,hearth=None):
        self.search_input.delete("0","300")
        if self.managing_disk == False:
            self.search_input.insert("0",str(self.all_rows[widget_id][0]))
        else:
            self.search_input.insert("0",str(self.disk_all_rows[widget_id][0]))

        self.check_given_input()
        if hearth == "favourite":
            add_colored_line(self.main_console,f"Projekt: {self.all_rows[widget_id][0]} byl odebr√°n z obl√≠ben√Ωch","green",None,True)
            self.switch_fav_status("del_favourite",change_status = True)
            #refresh obrazku oblibenosti:
            self.make_project_cells()
        elif hearth == "no_favourite":
            add_colored_line(self.main_console,f"Projekt: {self.all_rows[widget_id][0]} byl p≈ôid√°n do obl√≠ben√Ωch","green",None,True)
            self.switch_fav_status("add_favourite",change_status = True)
            self.make_project_cells()

            print("‚ù§Ô∏è",hearth)
        print(widget_id)

    def is_project_favourite(self,array_index):
        try:
            fav_status = int(self.favourite_list[array_index])
            if fav_status == 1:
                return True
            else:
                return False
        except IndexError:
            print(array_index," index error fav_status")

    def show_only_notes(self,e,widget_id,disk = None):
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        if disk:
            project_note = self.disk_all_rows[widget_id][5]
            project_name = self.disk_all_rows[widget_id][0]
        else:
            project_note = self.all_rows[widget_id][3]
            project_name = self.all_rows[widget_id][0]

        if project_note != "":
            note_window=customtkinter.CTk()
            # note_window.geometry("520x500")
            note_window.geometry(f"+{x+500}+{y+200}")
            # note_window.wm_iconbitmap(self.initial_path+'images/logo_TRIMAZKON.ico')
            note_window.wm_iconbitmap(resource_path(self.app_icon))
            
            note_window.title(f"Pozn√°mky k projektu: {project_name}")
            notes = customtkinter.CTkTextbox(master = note_window,font=("Arial",20),width=520,height=500)
            notes.grid(column = 0,row=0)
            # notes.grid_propagate(0)
            notes.insert(tk.END,project_note)
            notes.configure(state=tk.DISABLED)

            note_window.mainloop()

    def make_project_cells(self,no_read = None):
        def on_enter(e,new_text,widget):
            widget.configure(text = new_text)   

        def on_leave(e,new_text,widget):
            widget.configure(text = new_text)

        def filter_text_input(text):
            legit_rows = []
            legit_notes = ""
            rows = text.split("\n")
            for i in range(0,len(rows)):
                if rows[i].replace(" ","") != "":
                    legit_rows.append(rows[i])

            for i in range(0,len(legit_rows)): 
                if i == len(legit_rows)-1:
                    legit_notes = legit_notes + legit_rows[i]
                else:
                    legit_notes = legit_notes + legit_rows[i]+ "\n"
            return legit_notes
        
        def save_changed_notes(notes,row):
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["ip_address_list"]
            worksheet['D' + str(len(self.all_rows)-row)] = notes
            workbook.save(filename=self.excel_file_path)
            workbook.close()

        def on_enter_entry(e,widget,row_of_widget):
            widget.delete("1.0",tk.END)
            widget.insert(tk.END,str(self.all_rows[row_of_widget][3]))

        def on_leave_entry(e,widget,row_of_widget):
            notes_before = filter_text_input(str(self.all_rows[row_of_widget][3]))
            notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
            if notes_before != notes_after:
                print("Poznamka byla zmenena")
                self.all_rows[row_of_widget][3] = notes_after
                save_changed_notes(notes_after,row_of_widget)

            if "\n" in self.all_rows[row_of_widget][3]:
                notes_rows = self.all_rows[row_of_widget][3].split("\n")
                first_row = notes_rows[0]
                widget.delete("1.0",tk.END)
                widget.insert(tk.END,str(first_row))

            self.root.focus_set() # unfocus widget

        def shrink_frame(e,widget):
            new_height = 50
            if isinstance(widget,list):
                widget[0].configure(height = new_height)
                widget[1].configure(height = new_height-10)
            else:
                widget.configure(height = new_height)

        def expand_frame(e,widget,row_of_widget):
            filtered_input = filter_text_input(self.all_rows[row_of_widget][3])
            self.all_rows[row_of_widget][3] = filtered_input
            if "\n" in self.all_rows[row_of_widget][3]:
                notes_rows = self.all_rows[row_of_widget][3].split("\n")
                expanded_dim = (len(notes_rows)) * 35
                if isinstance(widget,list):
                    widget[0].configure(height = expanded_dim)
                    widget[1].configure(height = expanded_dim-10)
                else:
                    widget.configure(height = expanded_dim)
            else:
                return

        if no_read == None:
            self.read_excel_data()
        # padx_list = [10,190,390,390,650]
        padx_list = [60,240,440,440,700]
        self.clear_frame(self.project_tree)
        column1 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        column2 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "IPv4 adresa: ",font=("Arial",20,"bold"))
        column3 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Pozn√°mky: ",font=("Arial",20,"bold"))
        column1.grid(column = 0,row=0,pady = 5,padx =padx_list[0],sticky = tk.W)
        column2.grid(column = 0,row=0,pady = 5,padx =padx_list[1],sticky = tk.W)
        column3.grid(column = 0,row=0,pady = 5,padx =padx_list[3],sticky = tk.W)
        # y = widgets ve smeru y, x = widgets ve smeru x
        # nejprve vypis oblibene, potom zbytek
        for y in range(0,len(self.all_rows)):
            # ‚ô°,‚ô•,‚ù§Ô∏è
            project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=50)
            project_frame.grid(row=y+1,column=0,padx=10,sticky=tk.W)
            project_frame.grid_propagate(0)
            
            is_favourite = self.is_project_favourite(y)
            if is_favourite:
                filled_hearth =  customtkinter.CTkLabel(master = project_frame, width = 45,height=45,text = "üêò",font=("Arial",35),text_color="pink")
                filled_hearth.grid(column = 0,row=0,pady = 2,padx =2)
                filled_hearth.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id,"favourite"))
            else:
                unfilled_hearth =  customtkinter.CTkLabel(master = project_frame, width =45,height=45,text = "‚ô°",font=("Arial",40),text_color="red")
                unfilled_hearth.grid(column = 0,row=0,pady = 2,padx =2)
                unfilled_hearth.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id,"no_favourite"))
            
            for x in range(0,len(self.all_rows[y])):
                if x != 2: #nevypisujeme masku
                    if x == 0:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=180)
                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        button =  customtkinter.CTkButton(master = project_frame,width = 160,text = self.all_rows[y][x], command = lambda widget_id = y: self.change_computer_ip(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                        button.grid(column = 0,row=0,pady = 10,padx =10)
                        # zkop√≠rovat prav√Ωm klikem na button
                        button.bind("<Button-3>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                    elif x == 1:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=200)
                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        parameter =  customtkinter.CTkLabel(master = project_frame,text = self.all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        parameter.grid(column = 0,row=0,pady = 10,padx =10,sticky=tk.W)
                        parameter.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        # if x == 1: #frame s ip adresou
                        ip_addr = self.all_rows[y][x]
                        if ip_addr in self.current_address_list:
                            project_frame.configure(fg_color = "green")
                            project_frame.bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                            project_frame.bind("<Leave>",lambda e, ip = ip_addr, widget = parameter: on_leave(e,ip,widget))
                            parameter.bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                            parameter.bind("<Leave>",lambda e, ip = ip_addr, widget = parameter: on_leave(e,ip,widget))
                    elif x == 3:
                        notes_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=200)
                        notes_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        notes_frame.grid_propagate(0)
                        # binding the click on widget
                        notes_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        # notes =  customtkinter.CTkLabel(master = project_frame,text = self.all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        notes =  customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),width = 2200,height=40,corner_radius=0)
                        notes.grid(column = 0,row=0,pady = 5,padx =5,sticky=tk.W)
                        notes.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        # notes.insert(tk.END,str(self.all_rows[y][x]))
                        # if x==3: #frame s poznamkami...
                        notes_frame.configure(width=2200)
                        if "\n" in self.all_rows[y][x]:
                            notes_rows = self.all_rows[y][x].split("\n")
                            first_row = notes_rows[0]
                            # notes.configure(text =first_row) #vepsat pouze prvni radek
                            notes.delete("1.0",tk.END)
                            notes.insert(tk.END,str(first_row))
                        else:
                            notes.insert(tk.END,str(self.all_rows[y][x]))
                        
                        notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(e,widget,row))
                        notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(e,widget))
                        notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(e,widget,row))
                        notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(e,widget,row))

    def make_project_cells_disk(self,no_read = None,disk_statuses = False):
        def filter_text_input(text):
            legit_rows = []
            legit_notes = ""
            rows = text.split("\n")
            for i in range(0,len(rows)):
                if rows[i].replace(" ","") != "":
                    legit_rows.append(rows[i])

            for i in range(0,len(legit_rows)): 
                if i == len(legit_rows)-1:
                    legit_notes = legit_notes + legit_rows[i]
                else:
                    legit_notes = legit_notes + legit_rows[i]+ "\n"
            return legit_notes
        
        def save_changed_notes(notes,row):
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["disk_list"]
            worksheet['F' + str(len(self.disk_all_rows)-row)] = notes
            workbook.save(filename=self.excel_file_path)
            workbook.close()

        def on_enter_entry(e,widget,row_of_widget):
            widget.delete("1.0",tk.END)
            widget.insert(tk.END,str(self.disk_all_rows[row_of_widget][5]))

        def on_leave_entry(e,widget,row_of_widget):
            notes_before = filter_text_input(str(self.disk_all_rows[row_of_widget][5]))
            notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
            if notes_before != notes_after:
                print("Poznamka byla zmenena")
                self.disk_all_rows[row_of_widget][5] = notes_after
                save_changed_notes(notes_after,row_of_widget)

            if "\n" in self.disk_all_rows[row_of_widget][5]:
                notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                first_row = notes_rows[0]
                widget.delete("1.0",tk.END)
                widget.insert(tk.END,str(first_row))
                
            self.root.focus_set() # unfocus widget

        def shrink_frame(e,widget):
            new_height = 50
            if isinstance(widget,list):
                widget[0].configure(height = new_height)
                widget[1].configure(height = new_height-10)
            else:
                widget.configure(height = new_height)

        def expand_frame(e,widget,row_of_widget):
            filtered_input = filter_text_input(self.disk_all_rows[row_of_widget][5])
            self.disk_all_rows[row_of_widget][5] = filtered_input
            if "\n" in self.disk_all_rows[row_of_widget][5]:
                notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                expanded_dim = (len(notes_rows)) * 35
                if isinstance(widget,list):
                    widget[0].configure(height = expanded_dim)
                    widget[1].configure(height = expanded_dim-10)
                else:
                    widget.configure(height = expanded_dim)
            else:
                return

        if no_read == None:
            self.read_excel_data()
        padx_list = [10,190,240,0,0,640]
        self.clear_frame(self.project_tree)
        if self.default_disk_status_behav == 1:
            disk_statuses = True
        if disk_statuses:
            mapped_disks = list_mapped_disks(whole_format = True)
        column1 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Projekt: ",font=("Arial",20,"bold"))
        column2 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
        column3 =  customtkinter.CTkLabel(master = self.project_tree, width = 20,height=30,text = "Pozn√°mky: ",font=("Arial",20,"bold"))
        column1.grid(column = 0,row=0,pady = 5,padx =padx_list[0],sticky = tk.W)
        column2.grid(column = 0,row=0,pady = 5,padx =padx_list[2],sticky = tk.W)
        column3.grid(column = 0,row=0,pady = 5,padx =padx_list[5],sticky = tk.W)
        # y = widgets ve smeru y, x = widgets ve smeru x
        for y in range(0,len(self.disk_all_rows)):
            for x in range(0,len(self.disk_all_rows[y])):
                if x == 0:
                    project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=180)
                    project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                    project_frame.grid_propagate(0)
                    # binding the click on widget
                    project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                    button =  customtkinter.CTkButton(master = project_frame,width = 160,text = self.disk_all_rows[y][x], command = lambda widget_id = y: self.map_disk(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                    button.grid(column = 0,row=0,pady = 10,padx =10)
                    # zkop√≠rovat prav√Ωm klikem na button
                    button.bind("<Button-3>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                else:
                    if x != 3 and x != 4 and x != 5:
                        project_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=400)
                        if x == 1: #frame s p√≠smenem disku
                            project_frame.configure(width=50)
                            if disk_statuses:
                                for i in range(0,len(mapped_disks)):
                                    if mapped_disks[i][0:1] == self.disk_all_rows[y][x]:
                                        drive_status = check_network_drive_status(mapped_disks[i])
                                        if drive_status == True:
                                            project_frame.configure(fg_color = "green")
                                        else:
                                            project_frame.configure(fg_color = "red")

                        project_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        project_frame.grid_propagate(0)
                        # binding the click on widget
                        project_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        parameter =  customtkinter.CTkLabel(master = project_frame,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),justify='left')
                        parameter.grid(column = 0,row=0,pady = 10,padx =10,sticky=tk.W)
                        parameter.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))

                    if x==5: #frame s poznamkami...
                        notes_frame =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,fg_color="black",border_width=2,height=50,width=200)
                        notes_frame.grid(row=y+1,column=0,padx=padx_list[x],sticky=tk.W)
                        notes_frame.grid_propagate(0)
                        notes_frame.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        notes =  customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),width = 2200,height=40,corner_radius=0)
                        notes.grid(column = 0,row=0,pady = 5,padx =5,sticky=tk.W)
                        notes.bind("<Button-1>",lambda e, widget_id = y: self.clicked_on_project(e, widget_id))
                        notes_frame.configure(width=2200)
                        if "\n" in self.disk_all_rows[y][x]:
                            notes_rows = self.disk_all_rows[y][x].split("\n")
                            first_row = notes_rows[0]
                            # notes.configure(text =first_row) #vepsat pouze prvni radek
                            notes.delete("1.0",tk.END)
                            notes.insert(tk.END,str(first_row))
                        else:
                            notes.insert(tk.END,str(self.disk_all_rows[y][x]))
                        
                        notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(e,widget,row))
                        notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(e,widget))
                        notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(e,widget,row))
                        notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(e,widget,row))

    def edit_project(self):
        result = self.check_given_input()
        if result == True:
            if self.managing_disk == False:
                self.add_new_project(True)
            else:
                self.add_new_project_disk(True)
        elif result == None:
            add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout lev√Ωm na parametry dan√©ho projektu nebo prav√Ωm na tlaƒç√≠ko projektu)","orange",None,True)
        else:
            add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
    
    def refresh_explorer(self,refresh_disk=None):
        """
        refresh_disk = udelat nove v≈°echni widgets (make_project_cells_disk())
        """
        refresh_explorer="taskkill /f /im explorer.exe"
        subprocess.run(refresh_explorer, shell=True)
        refresh_explorer="start explorer.exe"
        subprocess.run(refresh_explorer, shell=True)
        if refresh_disk:
            self.make_project_cells_disk(disk_statuses=True)

    def delete_disk(self,child_root):
        drive_letter = str(self.drive_letter_input.get())
        if len(str(self.DL_manual_entry.get())) > 0:
            drive_letter = str(self.DL_manual_entry.get())
        
        delete_command = "net use " + drive_letter +": /del"
        process = subprocess.Popen(delete_command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
        stdout, stderr= process.communicate()
        if "Is it OK to continue disconnecting and force them closed?" in stdout:
            add_colored_line(self.main_console,f"Disk je pr√°vƒõ pou≈æ√≠v√°n, nejprve jej zav≈ôete","red",None,True)
            self.close_window(child_root)
        else:
            self.refresh_explorer()
            add_colored_line(self.main_console,f"Disky s oznaƒçen√≠m {drive_letter} byly odpojeny","orange",None,True)
            self.make_project_cells_disk(no_read=True) #refresh
            self.close_window(child_root)

    def delete_disk_option_menu(self):
        child_root=customtkinter.CTk()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"+{x+250}+{y+200}")
        # child_root.wm_iconbitmap(self.initial_path+'images/logo_TRIMAZKON.ico')
        child_root.wm_iconbitmap(resource_path(self.app_icon))
        # child_root.geometry("520x200")
        child_root.title("Odpojov√°n√≠ s√≠≈•ov√©ho disku")
        
        found_drive_letters=[]
        for i in range(0,len(self.disk_all_rows)):
            if not self.disk_all_rows[i][1] in found_drive_letters:
                found_drive_letters.append(self.disk_all_rows[i][1])

        mapped_disks = list_mapped_disks()
        for i in range(0,len(mapped_disks)):
            if not mapped_disks[i] in found_drive_letters:
                found_drive_letters.append(mapped_disks[i])

        label =                     customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte disk nebo vyhledejte manu√°lnƒõ: ",font=("Arial",20,"bold"))
        self.drive_letter_input =   customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=found_drive_letters,corner_radius=0)
        self.DL_manual_entry =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manu√°lnƒõ")
        # del_button =                customtkinter.CTkButton(master = child_root, width = 200,height=30,text = "Odpojit", command = lambda: self.delete_disk(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        del_button =                customtkinter.CTkButton(master = child_root, width = 200,height=30,text = "Odpojit", command = lambda: self.delete_disk(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        label.                      grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.drive_letter_input.    grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.DL_manual_entry.       grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        del_button.                 grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        child_root.mainloop()

    def map_disk(self,button_row):
        Drive_letter = str(self.disk_all_rows[button_row][1])
        ftp_adress = str(self.disk_all_rows[button_row][2])
        # raw_ftp_address = r"{}".format(ftp_adress)
        # ftp_adress = raw_ftp_address
        
        user = str(self.disk_all_rows[button_row][3])
        password = str(self.disk_all_rows[button_row][4])

        delete_command = "net use " + Drive_letter + ": /del"
        subprocess.run(delete_command, shell=True)
        # second_command = "net use " + Drive_letter + ": " + ftp_adress + " /user:" + user + " " + password + " /persistent:No"
        if user != "" or password != "":
            second_command = "net use " + Drive_letter + ": " + ftp_adress + " " + password + " /user:" + user# + " /persistent:No"
        else:
            second_command = "net use " + Drive_letter + ": " + ftp_adress
        print("calling: ",second_command)

        def call_subprocess():
            """process = subprocess.Popen(second_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
            self.connection_status = process.returncode
            print("STDOUT:", stdout)
            print("STDERR:", stderr)
            print("Return Code:", self.connection_status)"""
            self.connection_status = subprocess.call(second_command,shell=True,text=True)
  
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

        time_start = time.time()
        while self.connection_status==None:
            time.sleep(0.05)
            if time.time() - time_start > 3:
                print("terminated due to runtime error")
                break

        if self.connection_status == 0:
             add_colored_line(self.main_console,f"Disk √∫spƒõ≈°nƒõ p≈ôipojen","green",None,True)
             self.refresh_explorer()
             self.make_project_cells_disk(no_read=True)
             def open_explorer(path):
                if os.path.exists(path):
                    os.startfile(path)
                else:
                    print(f"The path {path} does not exist.")

             open_explorer(Drive_letter + ":\\")
        else:
             add_colored_line(self.main_console,f"P≈ôipojen√≠ selhalo (ixon? mus√≠ b√Ωt zvolena alespo≈à 1 slo≈æka na disku...)","red",None,True)

    def get_ipv4_addresses(self):
        # Run the ipconfig command
        result = subprocess.run(['ipconfig'], capture_output=True, text=True)
        # Regular expression to match the IPv4 address
        ipv4_pattern = re.compile(r'IPv4 Address[.\s]*: ([\d.]+)')
        # Dictionary to store interface names and their IPv4 addresses
        ipv4_addresses = []
        # Split the output by lines
        lines = result.stdout.splitlines()
        current_interface = None
        # Iterate over each line to find interface names and IPv4 addresses
        for line in lines:
            if line.strip():
                # Detect interface name
                if line[0].isalpha():
                    current_interface = line.strip()
                else:
                    # Detect IPv4 address for the current interface
                    match = ipv4_pattern.search(line)
                    if match and current_interface:
                        ipv4_addresses.append(current_interface)
                        ipv4_addresses.append(match.group(1))
                        #ipv4_addresses[current_interface] = match.group(1)
        
        return ipv4_addresses

    def option_change(self,args,only_console = False,silent = False):
        """
        Vol√° get_current_ip_list(), aktualizuje souƒçasnƒõ nastaven√© adresy
        - only console: vyp√≠≈°e do konzole aktu√°ln√≠ p≈ôipojen√≠
        - silent: nevypisuje do konzole
        """
        if not only_console:
            try:
                self.default_connection_option = self.connection_option_list.index(self.drop_down_options.get())
            except ValueError as e:
                print(f"Error: {e}")
                self.default_connection_option = 0

            #pamatovat si naposledy zvoleny zpusob pripojeni:
            self.save_setting_parameter(parameter="change_def_conn_option",status=int(self.default_connection_option))
            self.get_current_ip_list()
            if self.static_label2.winfo_exists():
                self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
        if not silent:
            # ziskat data o aktualnim pripojeni
            current_connection = self.get_ipv4_addresses()
            message = ""
            for items in current_connection:
                message = message + items + " "
            if message == "":
                message = "nenalezeno"
            add_colored_line(self.main_console,f"Souƒçasn√© p≈ôipojen√≠: {message}","white",None,True)

    def make_project_first(self,purpouse=None):
        """
        purpouse:
        - search
        - silent
        """
        result = self.check_given_input()
        if result == True:
            #zmena poradi
            project = self.all_rows[self.last_project_id]
            favourite_status = self.favourite_list[self.last_project_id]
            self.all_rows.pop(self.last_project_id)
            self.all_rows.insert(0,project)
            self.favourite_list.pop(self.last_project_id)
            self.favourite_list.insert(0,favourite_status)
            #self.all_rows.append(project)
            #if save == True:
            for i in range(0,len(self.all_rows)):
                row = (len(self.all_rows)-1)-i
                self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=self.favourite_list[i])

            self.make_project_cells()
            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} nalezen","green",None,True)
            elif purpouse != "silent":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} p≈ôesunut na zaƒç√°tek","green",None,True)
        elif result == None and purpouse != "silent":
            if purpouse == "search":
                add_colored_line(self.main_console,f"Vlo≈æte hledan√Ω projekt do vyhled√°v√°n√≠","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout lev√Ωm na parametry dan√©ho projektu nebo prav√Ωm na tlaƒç√≠ko projektu)","orange",None,True)
        elif purpouse != "silent":
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

    def make_project_first_disk(self,purpouse = None):
        result = self.check_given_input()
        if result == True:
            #zmena poradi
            project = self.disk_all_rows[self.last_project_id]
            self.disk_all_rows.pop(self.last_project_id)
            self.disk_all_rows.insert(0,project)

            for i in range(0,len(self.disk_all_rows)):
                row = (len(self.disk_all_rows)-1)-i
                
                self.save_excel_data_disk(self.disk_all_rows[i][0],self.disk_all_rows[i][1],self.disk_all_rows[i][2],self.disk_all_rows[i][3],self.disk_all_rows[i][4],self.disk_all_rows[i][5],None,row+1)

            self.make_project_cells_disk()

            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} nalezen","green",None,True)
            else:
                add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} p≈ôesunut na zaƒç√°tek","green",None,True)
        elif result == None:
            if purpouse == "search":
                add_colored_line(self.main_console,f"Vlo≈æte hledan√Ω projekt do vyhled√°v√°n√≠","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout lev√Ωm na parametry dan√©ho projektu nebo prav√Ωm na tlaƒç√≠ko projektu)","orange",None,True)
        else:
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

    def get_current_ip_list(self):
        def get_current_ip_address(interface_name):
        # Get network interfaces and their addresses
            addresses = psutil.net_if_addrs()
            # Check if the specified interface exists
            if interface_name in addresses:
                addr_count = 0
                print(f"Adresy interfacu: {interface_name}")
                for addr in addresses[interface_name]:
                    print(f"{addr}")
                    # prvni AF_INET je pridelena automaticky, druha je privatni, nastavena DHCP
                    if addr.family == socket.AF_INET:  # IPv4 address
                        if addr_count == 1:
                            return addr.address
                        addr_count +=1
                        remembered_addr = addr
                if addr_count == 1:
                    # print(addr.family,addr.address)
                    # return "Nenalezeno"
                    if "AddressFamily.AF_INET:" in str(remembered_addr):
                        return remembered_addr.address
                    else:
                        return "Nenalezeno"
            else:
                return "Nenalezeno"
        self.current_address_list = []
        for items in self.connection_option_list:
            found_address = get_current_ip_address(items)
            self.current_address_list.append(found_address)

    def save_setting_parameter(self,parameter,status):
        """
        list of parameters:\n

        change_def_conn_option\n
        new_conn_options\n
        change_def_ip_window\n
        change_def_main_window\n
        change_def_window_size\n
        change_def_disk_behav\n
        """
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["Settings"]
        if parameter == "change_def_conn_option":
            row = 1
        elif parameter == "new_conn_options":
            row = 2
        elif parameter == "change_def_ip_window":
            row = 3
        elif parameter == "change_def_main_window":
            row = 4
        elif parameter == "change_def_window_size":
            row = 5
        elif parameter == "change_def_disk_behav":
            row = 6
        worksheet['B' + str(row)] = status
        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def show_favourite_toggle(self,keep_search_input = False,determine_status = None): # hlavni prepinaci tlacitko oblibene/ neoblibene
        if self.show_favourite and (determine_status == None or determine_status == "all"):
            self.show_favourite = False
            window_status = 0
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            # self.show_only_fav.configure(text = "Obl√≠ben√©")
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.search_input.configure(placeholder_text="N√°zev projektu")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            self.button_remove_main.configure(command = lambda: self.delete_project())
            self.save_setting_parameter(parameter="change_def_ip_window",status=window_status)
            self.button_switch_favourite_ip. configure(fg_color="black")
            self.button_switch_all_ip.       configure(fg_color="#212121")

        elif self.show_favourite == False and (determine_status == None or determine_status == "fav"):
            # favourite window
            self.show_favourite = True
            window_status = 1
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            # self.show_only_fav.configure(text = "V≈°echny projekty")
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.search_input.configure(placeholder_text="N√°zev projektu")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            self.button_remove_main.configure(command = lambda: self.switch_fav_status("with_refresh"))
            self.save_setting_parameter(parameter="change_def_ip_window",status=window_status)
            self.button_switch_favourite_ip. configure(fg_color="#212121")
            self.button_switch_all_ip.       configure(fg_color="black")

    def refresh_interfaces(self,all = False):
        interfaces_data = self.fill_interfaces()
        self.connection_option_list = interfaces_data[0]
        self.drop_down_options.configure(values = self.connection_option_list)
        online_list_text = ""
        if len(interfaces_data[1]) > 0:
            for data in interfaces_data[1]:
                online_list_text = online_list_text + str(data) +", "
            online_list_text = online_list_text[:-2] # odebrat ƒç√°rku s mezerou

        self.online_list.configure(text=online_list_text)
        if all:
            self.option_change("")

    def create_widgets(self,fav_status = None,init=None,excel_load_error = False):
        if not excel_load_error:
            if init:
                if self.window_mode == "max":
                    self.save_setting_parameter(parameter="change_def_window_size",status=1)
                else:
                    self.save_setting_parameter(parameter="change_def_window_size",status=0)
            if fav_status:
                self.show_favourite = True
                self.save_setting_parameter(parameter="change_def_ip_window",status=1)
            if fav_status == False:
                self.show_favourite = False
                self.save_setting_parameter(parameter="change_def_ip_window",status=0)
            
            self.save_setting_parameter(parameter="change_def_main_window",status=0)
        
        self.clear_frame(self.root)
        self.managing_disk = False
        menu_cards =            customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        main_widgets =          customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =     customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)

        menu_cards.             pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        # logo =                  customtkinter.CTkImage(Image.open(self.initial_path+"images/jhv_logo.png"),size=(300, 100))
        logo =                  customtkinter.CTkImage(Image.open(resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =            customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
        image_logo.             pack(pady=5,padx=15,expand=True,side = "right",anchor="e")
        main_widgets.           pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.      pack(pady=5,padx=5,fill="both",expand=True,side = "top")

        main_menu_button =                  customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        self.button_switch_all_ip =         customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - v≈°echny",command =  lambda: self.show_favourite_toggle(determine_status="all"),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        self.button_switch_favourite_ip =   customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - obl√≠ben√©",command =  lambda: self.show_favourite_toggle(determine_status="fav"),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_disk =                customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "S√≠≈•ov√© disky",command =  lambda: self.create_widgets_disk(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")

        project_label =             customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =         customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=160,height=40,placeholder_text="N√°zev projektu",corner_radius=0)
        button_search =             customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first("search"),font=("Arial",20,"bold"),corner_radius=0)
        self.button_add_main =      customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Nov√Ω projekt", command = lambda: self.add_new_project(),font=("Arial",20,"bold"),corner_radius=0)
        self.button_remove_main =   customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Smazat projekt", command =  lambda: self.delete_project(),font=("Arial",20,"bold"),corner_radius=0)
        self.button_edit_main =     customtkinter.CTkButton(master = main_widgets, width = 160,height=40,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
        button_make_first =         customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "P≈ôesunout na zaƒç√°tek",command =  lambda: self.make_project_first(),font=("Arial",20,"bold"),corner_radius=0)
        if self.show_favourite:
            self.button_switch_favourite_ip. configure(fg_color="#212121")
            self.button_switch_all_ip.       configure(fg_color="black")
        else:
            self.button_switch_favourite_ip. configure(fg_color="black")
            self.button_switch_all_ip.       configure(fg_color="#212121")

        connect_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "P≈ôipojen√≠: ",font=("Arial",20,"bold"))
        self.drop_down_options = customtkinter.CTkOptionMenu(master = main_widgets,width=200,height=40,font=("Arial",20,"bold"),corner_radius=0,command=  self.option_change)
        # "‚öôÔ∏è", "‚öí", "üîß", "üî©"
        button_settings =       customtkinter.CTkButton(master = main_widgets, width = 40,height=40,text="‚öí",command =  lambda: self.refresh_interfaces(all=True),font=("",22),corner_radius=0) #refresh interface status≈Ø
        static_label =          customtkinter.CTkLabel(master = main_widgets, height=40,text = "Static:",font=("Arial",20,"bold"))
        self.static_label2 =    customtkinter.CTkLabel(master = main_widgets, height=40,text = "",font=("Arial",22,"bold"),bg_color="black")
        online_label =          customtkinter.CTkLabel(master = main_widgets, height=40,text = "Online: ",font=("Arial",22,"bold"))
        self.online_list =           customtkinter.CTkLabel(master = main_widgets, height=40,text = "",font=("Arial",22,"bold"))

        self.main_console = tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
        main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        self.button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        self.button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        image_logo.                     pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
        project_label.          grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.      grid(column = 0,row=0,pady = 5,padx =90,sticky = tk.W)
        button_search.          grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        self.button_add_main.   grid(column = 0,row=0,pady = 5,padx =410,sticky = tk.W)
        self.button_remove_main.grid(column = 0,row=0,pady = 5,padx =565,sticky = tk.W)
        self.button_edit_main.  grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)
        button_make_first.      grid(column = 0,row=0,pady = 5,padx =885,sticky = tk.W)
        connect_label.          grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.drop_down_options. grid(column = 0,row=1,pady = 0,padx =110,sticky = tk.W)
        button_settings.        grid(column = 0,row=1,pady = 0,padx =315,sticky = tk.W)
        static_label.           grid(column = 0,row=1,pady = 0,padx =365,sticky = tk.W)
        self.static_label2.     grid(column = 0,row=1,pady = 0,padx =430,sticky = tk.W,ipadx = 10,ipady = 2)
        online_label.           grid(column = 0,row=1,pady = 0,padx =620,sticky = tk.W)
        self.online_list.           grid(column = 0,row=1,pady = 0,padx =700,sticky = tk.W)

        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.refresh_interfaces()
        # aktualizace hodnot nab√≠dky
        if self.default_connection_option < len(self.connection_option_list):
            # nastaven√≠ naposledy zvolen√©ho interfacu
            self.drop_down_options.set(self.connection_option_list[self.default_connection_option])
        else:
            self.default_connection_option = 0
            self.save_setting_parameter(parameter="change_def_conn_option",status=0)
            self.drop_down_options.set(self.connection_option_list[self.default_connection_option])

        if not excel_load_error:
            self.option_change("")
            self.make_project_cells()
            self.get_current_ip_list()
            self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
        else:
            add_colored_line(self.main_console,f"Nejprve pros√≠m zav≈ôete soubor {self.excel_file_path}","red",None,True)


        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                #self.root.after(0, lambda:self.root.state('normal'))
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")
                # self.root.geometry("210x500")
                self.save_setting_parameter(parameter="change_def_window_size",status=2)
            elif int(current_width) ==260:
                self.root.geometry("1200x900")
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')
                self.save_setting_parameter(parameter="change_def_window_size",status=1)

        self.root.bind("<f>",lambda e: maximalize_window(e))

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first("search")
        self.search_input.bind("<Return>",call_search)
        self.root.mainloop()

    def call_make_cells_disk(self):
        self.make_project_cells_disk()

    def setting_window(self):
        def save_new_behav():
            nonlocal checkbox
            if int(checkbox.get()) == 0:
                self.default_disk_status_behav = 0
                self.save_setting_parameter(parameter="change_def_disk_behav",status=0)
            elif int(checkbox.get()) == 1:
                self.default_disk_status_behav = 1
                self.make_project_cells_disk(no_read=True)
                self.save_setting_parameter(parameter="change_def_disk_behav",status=1)

        child_root=customtkinter.CTk()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"500x200+{x+350}+{y+180}")
        child_root.wm_iconbitmap(resource_path(self.app_icon))
        child_root.title("Nastaven√≠")
        main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0)
        main_frame.     pack(expand=True,fill="both",side="left")
        label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "Chov√°n√≠ p≈ôi vstupu do menu \"S√≠≈•ov√© disky\"",font=("Arial",20,"bold"))
        checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "P≈ôi spu≈°tƒõn√≠ aktualizovat statusy disk≈Ø",font=("Arial",16,"bold"),command=lambda: save_new_behav())
        button_close =  customtkinter.CTkButton(master = main_frame, width = 150,height=40,text = "Zav≈ô√≠t",command =  lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
        label.          pack(pady = 10,padx=10,side="top")
        checkbox.       pack(pady = 10,padx=10,side="top")
        button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

        if self.default_disk_status_behav == 1:
            checkbox.select()
        child_root.mainloop()

    def create_widgets_disk(self,init=None):
        if init:
            if self.window_mode == "max":
                self.save_setting_parameter(parameter="change_def_window_size",status=1)
            else:
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
        self.clear_frame(self.root)
        self.managing_disk = True
        self.save_setting_parameter(parameter="change_def_main_window",status=1)
        menu_cards =            customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        main_widgets =          customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =     customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        menu_cards.             pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        # logo =                  customtkinter.CTkImage(Image.open(self.initial_path+"images/jhv_logo.png"),size=(300, 100))
        logo =                  customtkinter.CTkImage(Image.open(resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =            customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
        image_logo.             pack(pady=5,padx=15,expand=True,side = "right",anchor="e")
        main_widgets.           pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.      pack(pady=5,padx=5,fill="both",expand=True,side = "top")

        main_menu_button =              customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_all_ip =          customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - v≈°echny",command =  lambda: self.create_widgets(fav_status=False),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_favourite_ip =    customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - obl√≠ben√©",command =  lambda: self.create_widgets(fav_status=True),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_disk =            customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "S√≠≈•ov√© disky",command =  lambda: self.create_widgets_disk(),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        project_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =     customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=160,height=40,placeholder_text="N√°zev projektu",corner_radius=0)
        button_search =         customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first_disk("search"),font=("Arial",20,"bold"),corner_radius=0)
        button_add =            customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Nov√Ω projekt", command = lambda: self.add_new_project_disk(),font=("Arial",20,"bold"),corner_radius=0)
        button_remove =         customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Smazat projekt", command =  lambda: self.delete_project_disk(),font=("Arial",20,"bold"),corner_radius=0)
        button_edit =           customtkinter.CTkButton(master = main_widgets, width = 160,height=40,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
        button_make_first =     customtkinter.CTkButton(master = main_widgets, width = 250,height=40,text = "P≈ôesunout na zaƒç√°tek",command =  lambda: self.make_project_first_disk(),font=("Arial",20,"bold"),corner_radius=0)
        button_settings =       customtkinter.CTkButton(master = main_widgets, width = 40,height=40,text="‚öôÔ∏è",command =  lambda: self.setting_window(),font=("",22),corner_radius=0)
        delete_disk =           customtkinter.CTkButton(master = main_widgets, width = 250,height=40,text = "Odpojit s√≠≈•ov√Ω disk",command =  lambda: self.delete_disk_option_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        reset =                 customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "Reset exploreru",command =  lambda: self.refresh_explorer(refresh_disk=True),font=("Arial",20,"bold"),corner_radius=0)
        refresh =               customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "Refresh status≈Ø",command =  lambda: self.make_project_cells_disk(disk_statuses=True),font=("Arial",20,"bold"),corner_radius=0)
        
        as_admin_label =        customtkinter.CTkLabel(master = main_widgets,text = "",font=("Arial",20,"bold"))
        self.main_console =     tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)

        main_menu_button.          pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_disk.        pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        image_logo.                pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
        project_label.      grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.  grid(column = 0,row=0,pady = 5,padx =90,sticky = tk.W)
        button_search.      grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        button_add.         grid(column = 0,row=0,pady = 5,padx =410,sticky = tk.W)
        button_remove.      grid(column = 0,row=0,pady = 5,padx =565,sticky = tk.W)
        button_edit.        grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)
        button_make_first.  grid(column = 0,row=0,pady = 5,padx =885,sticky = tk.W)
        button_settings.    grid(column = 0,row=0,pady = 5,padx =1140,sticky = tk.W)
        delete_disk.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        reset.              grid(column = 0,row=1,pady = 5,padx =265,sticky = tk.W)
        refresh.            grid(column = 0,row=1,pady = 5,padx =470,sticky = tk.W)

        as_admin_label.     grid(column = 0,row=1,pady = 5,padx =675,sticky = tk.W)
        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.option_change("",only_console=True)
        
        def is_admin():
            try:
                return ctypes.windll.shell32.IsUserAnAdmin()
            except:
                return False
            
        if is_admin():
            as_admin_label.configure(text = "Aplikace je spu≈°tƒõna, jako administr√°tor\n(mapovat disky lze pouze na u≈æivatelsk√©m √∫ƒçtu)",text_color = "orange")

        def maximalize_window(e):
            self.root.update_idletasks()
            self.root.update()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                #self.root.after(0, lambda:self.root.state('normal'))
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")
                # self.root.geometry("210x500")
                self.save_setting_parameter(parameter="change_def_window_size",status=2)
            elif int(current_width) ==260:
                self.root.geometry("1200x900")
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')
                self.save_setting_parameter(parameter="change_def_window_size",status=1)
            self.root.update_idletasks()
            self.root.update()
        self.root.bind("<f>",lambda e: maximalize_window(e))

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first_disk("search")
        self.search_input.bind("<Return>",call_search)

        def call_refresh(e):
            self.refresh_explorer(refresh_disk=True)
        self.root.bind("<F5>",lambda e: call_refresh(e))

        self.root.update()
        self.call_make_cells_disk()
        self.root.mainloop()

# IP_assignment(root)
# root.mainloop()