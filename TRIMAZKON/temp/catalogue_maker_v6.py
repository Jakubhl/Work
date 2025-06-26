import customtkinter
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl import load_workbook
import xlwings as xw
import string
from PIL import Image as PILImage
from PIL import ImageTk
from datetime import datetime
from tkinter import filedialog
import os
import xml.etree.ElementTree as ET
# import sharepoint_download as download_database
import sys
import pyperclip
import copy
import json
import tkinter.font as tkFont
import pyodbc
import time
import threading

initial_path = ""
testing = True

if testing:
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    root=customtkinter.CTk()
    # root.geometry("1200x900")
    root.title("Catalogue maker v3.0")
    database_filename  = "Sharepoint_databaze.xlsx"
    root.state('zoomed')

    def set_zoom(zoom_factor):
        root.tk.call('tk', 'scaling', zoom_factor / 100)
        customtkinter.set_widget_scaling(zoom_factor / 100) 
    set_zoom(80)

class Tools:
    config_json_filename = "TRIMAZKON.json"
    @classmethod
    def add_colored_line(cls,text_widget, text, color,font=None,delete_line = None):
        """
        Vloží řádek do console
        """
        text_widget.configure(state=tk.NORMAL)
        if font == None:
            font = ("Arial",22)
        if delete_line != None:
            text_widget.delete("current linestart","current lineend")
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert("current lineend",text, color)
        else:
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert(tk.END,"    > "+ text+"\n", color)

        text_widget.configure(state=tk.DISABLED)
        
    @classmethod
    def make_wrapping(cls,text):
        # text = re.sub(r'\n{3,}', '\n', str(text)) # odstraní více jak tři mezery za sebou
        lines = text.split("\n")
        whole_new_string = ""
        number_of_chars = 0
        max_num_of_chars_one_line = 35

        fitted_lines = []
        for line in lines:
            line = line.rstrip()
            if len(line) > max_num_of_chars_one_line:
                text_splitted = line.split(" ")
                # text_splitted = [x for x in text_splitted if x]
                new_string = ""
                for items in text_splitted:
                    number_of_chars += len(items)
                    if number_of_chars > max_num_of_chars_one_line:
                        if new_string == "": # osetreni odsazeni na prvnim radku
                            new_string += str(items) + " "
                            number_of_chars = len(items)
                        else:
                            new_string += "\n" + str(items) + " "
                            number_of_chars = len(items)
                    else: 
                        new_string += str(items) + " "

                fitted_lines.append(new_string + "\n")
            else:
                if line == "":
                    fitted_lines.append("\n")
                else:
                    fitted_lines.append(line+"\n")

        for items in fitted_lines:
            whole_new_string += items

        if whole_new_string.endswith("\n"):
            whole_new_string = whole_new_string.rstrip("\n")

        return whole_new_string
    
    @classmethod
    def browseDirectories(cls,visible_files,start_path=None,file_type = [("All files", "*.*")]): # Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat
        """
        Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat

        Vstupní data:

        0: visible_files = "all" / "only_dirs"\n
        1: start_path = None -optimalni, docasne se ulozi posledni nastavena cesta v exploreru

        Výstupní data:

        0: výstupní chybová hlášení
        1: opravená cesta
        2: nazev vybraneho souboru (option: all)
        """
        corrected_path = ""
        output= ""
        name_of_selected_file = ""
        if start_path == None:
            # start_path = Tools.resource_path(os.getcwd())
            start_path = initial_path
        start_path = Tools.path_check(start_path)
        # pripad vyberu files, aby byly viditelne
        if visible_files == "all":
            if(start_path != "" and start_path != False):
                foldername_path = filedialog.askopenfile(initialdir = start_path,
                                                        title = "Klikněte na soubor v požadované cestě",
                                                        filetypes=file_type)
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:           
                foldername_path = filedialog.askopenfile(initialdir = "/",
                                                        title = "Klikněte na soubor v požadované cestě",
                                                        filetypes=file_type)
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"

        # pripad vyberu slozek
        if visible_files == "only_dirs":
            if(start_path != ""):
                path_to_directory = filedialog.askdirectory(initialdir = start_path, title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:
                path_to_directory = filedialog.askdirectory(initialdir = "/", title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"

        check = Tools.path_check(path_to_directory)
        corrected_path = check
        return [output,corrected_path,name_of_selected_file]

    @classmethod
    def path_check(cls,path_raw,only_repair = None):
        if path_raw == None:
            return
        path=path_raw
        backslash = "\\"
        if backslash[0] in path:
            newPath = path.replace(backslash[0], '/')
            path = newPath
        if path.endswith('/') == False:
            newPath = path + "/"
            path = newPath
        #oprava mezery v nazvu
        path = r"{}".format(path)
        if not os.path.exists(path) and only_repair == None:
            return False
        else:
            return path

    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        # if hasattr(sys, '_MEIPASS'):
        #     return os.path.join(sys._MEIPASS, relative_path)
        # return os.path.join(os.path.abspath("."), relative_path)
        BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.abspath(".")
        return os.path.join(BASE_DIR, relative_path)
    
    @classmethod
    def read_json_config(cls): # Funkce vraci data z configu
        """
        Funkce vrací data z konfiguračního souboru

        \nCATALOGUE SETTINGS\n
        - database_filename
        - catalogue_filename
        - metadata_filename
        - subwindow_behav
        - default_export_suffix
        - default_path
        - render_mode
        - path_history_list
        - hover_info_trigger_mode

        \ndb_settings\n
        - server_history_list
        - username_history_list
        - db_name_list
        - user_id
        """
        # default_labels = string_database.default_setting_database

        if os.path.exists(initial_path+cls.config_json_filename):
            try:
                output_data = []
                with open(initial_path+cls.config_json_filename, "r") as file:
                    output_data = json.load(file)

                # print("config data: ", output_data, len(output_data))
                return output_data

            except Exception as e:
                print(f"Nejdřív zavřete soubor {cls.config_json_filename} Chyba: {e}")
        else:
            print(f"Chybí konfigurační soubor {cls.config_json_filename}")
            return False
        
    @classmethod
    def save_to_json_config(cls,input_data,which_settings,which_parameter,language_force = "cz"): # Funkce zapisuje data do souboru configu
        """
        Funkce zapisuje data do konfiguračního souboru

        vraci vystupni zpravu: report

        which_settings je bud: 
        - app_settings
        - sort_conv_settings
        - del_settings
        - image_browser_settings
        - catalogue_settings
        - ip_settings

        \nCATALOGUE_SETTINGS\n
        - database_filename
        - catalogue_filename
        - metadata_filename
        - subwindow_behav
        - default_export_suffix
        - default_path
        - render_mode
        - path_history_list
        - hover_info_trigger_mode

        \ndb_settings\n
        - server_history_list
        - username_history_list
        - db_name_list
        - user_id
        """
        def get_input_data_format():
            if isinstance(input_data,list):
                return input_data
            elif isinstance(input_data,str):
                return str(input_data)
            elif isinstance(input_data,int):
                return int(input_data)
        
        if os.path.exists(initial_path + cls.config_json_filename):
            with open(initial_path+cls.config_json_filename, "r") as file:
                config_data = json.load(file)

            if which_settings not in config_data:
                config_data[which_settings] = {}

            config_data[which_settings][which_parameter] = get_input_data_format()

            with open(initial_path+cls.config_json_filename, "w") as file:
                json.dump(config_data, file, indent=4)

        else:
            print("Chybí konfigurační soubor (nelze ukládat změny)")
            return "Chybí konfigurační soubor (nelze ukládat změny)"
   
    @classmethod
    def add_new_path_to_history(cls,new_path,which_settings):
        if new_path == "delete_history":
            Tools.save_to_json_config([],which_settings,"path_history_list")
            return

        current_paths = Tools.read_json_config()[which_settings]["path_history_list"]
        if new_path not in current_paths:
            if len(current_paths) > 9:
                current_paths.pop()
            # current_paths.append(str(new_path))
            current_paths.insert(0,str(new_path))
            Tools.save_to_json_config(current_paths,which_settings,"path_history_list")
    
    @classmethod
    def call_path_context_menu(cls,master,entry_widget,menu_btn,items_given = False,combine_path_items=False,given_path=None):
        if items_given==False:
            path_history = Tools.read_json_config()["catalogue_settings"]["path_history_list"]
        else:
            path_history = items_given
        print(path_history)
        def insert_path(path):
            entry_widget.delete("0","200")
            if combine_path_items:
                def check_dir(path):
                    if os.path.isfile(path):
                        path = os.path.dirname(path) + "/"
                    return path
                entry_widget.insert("0", check_dir(given_path) + path)
            else:
                entry_widget.insert("0", path)
        if len(path_history) > 0:
            # path_context_menu = tk.Menu(master, tearoff=0,fg="white",bg="black")
            path_context_menu = tk.Menu(master,tearoff=0,fg="white",bg="#202020",activebackground="#606060")

            for i in range(0,len(path_history)):
                path_context_menu.add_command(label=path_history[i], command=lambda row_path = path_history[i]: insert_path(row_path),font=("Arial",22,"bold"))
                if i < len(path_history)-1:
                    path_context_menu.add_separator()
                    
            path_context_menu.tk_popup(menu_btn.winfo_rootx(),menu_btn.winfo_rooty()+40)

    @classmethod
    def make_table_for_db_export(cls,station_list,controller_list):
        """
        data = [
            (1, 'Alice', 25),
            (2, 'Bob', 30),
            (3, 'Charlie', 22)
        ]

        """
        table_to_return = []

        current_date = datetime.now()
        date_string = current_date.strftime("%d.%m.%Y %H:%M:%S")
        user_id = "1111"

        def get_all_cables():
            for station in station_list:
                for camera in station["camera_list"]:
                    if camera["cable"] != "":
                        table_to_return.append((user_id,station["name"],"PŘÍSLUŠENSTVÍ","OMRON","kabel ke kameře","",camera["cable"],"",camera["controller"],date_string))
        
        def get_all_cameras():
            for station in station_list:
                for camera in station["camera_list"]:
                    table_to_return.append((user_id,station["name"],"KAMERY","OMRON","kamera","",camera["type"],"",camera["controller"],date_string))

        def get_all_optics_and_lights():
            for station in station_list:
                for camera in station["camera_list"]:
                    for optics in camera["optics_list"]:
                        if not "light_status" in optics:
                            table_to_return.append((user_id,station["name"],"OPTIKA","OMRON","objektiv","",optics["type"],optics["alternative"],camera["type"],date_string))
                        elif int(optics["light_status"]) != 1:
                            table_to_return.append((user_id,station["name"],"OPTIKA","OMRON","objektiv","",optics["type"],optics["alternative"],camera["type"],date_string))
                        else:
                            table_to_return.append((user_id,station["name"],"PŘÍSLUŠENSTVÍ","smart view","světlo","",optics["type"],optics["alternative"],camera["type"],date_string))

        def get_all_controllers():
            for controller in controller_list:
                table_to_return.append((user_id,"","KONTROLERY","OMRON","kontroler",controller["name"],controller["type"],"","",date_string))
                
        def get_all_accessories():
            for controller in controller_list:
                for acc in controller["accessory_list"]:
                    table_to_return.append((user_id,"","PŘÍSLUŠENSTVÍ","","příslušenství","",acc["type"],"",controller["type"],date_string))

        get_all_cables()
        get_all_cameras()
        get_all_optics_and_lights()
        get_all_controllers()
        get_all_accessories()
        print(table_to_return)
        return table_to_return

class Save_prog_metadata:
    def __init__(self,console,controller_database=[],station_list=[],project_name="",xml_file_path=""):
        self.controller_database = controller_database
        self.station_list = station_list
        self.project_name = project_name
        self.main_console = console
        self.xml_file_path = xml_file_path

    def store_xml_data(self):
        # KONTROLERY ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        root1 = ET.Element("metadata")
        controller_list = ET.SubElement(root1, "controllers")
        for controllers in self.controller_database:
            controller = ET.SubElement(controller_list, "controller")
            for contr_key, contr_value in controllers.items():
                if contr_key == "accessory_list":
                    accessories = ET.SubElement(controller, "accessory_list")
                    for accessory in contr_value:
                        accessory_element = ET.SubElement(accessories, "accessory")
                        for acc_key, acc_value in accessory.items():
                            acc_child = ET.SubElement(accessory_element, acc_key)
                            acc_child.text = str(acc_value)  # Ensure value is a string
                else:
                    contr_child = ET.SubElement(controller, contr_key)
                    contr_child.text = str(contr_value)

        # STANICE ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        station_list = ET.SubElement(root1, "station_list")
        for stations in self.station_list:
            station_element = ET.SubElement(station_list, "station")
            for key, value in stations.items():
                if key == "camera_list":
                    cameras = ET.SubElement(station_element, "camera_list")
                    for camera in value:
                        camera_element = ET.SubElement(cameras, "camera")
                        for cam_key, cam_value in camera.items():
                            if cam_key == "optics_list":
                                optics = ET.SubElement(camera_element, "optics_list")
                                for optic in cam_value:
                                    optic_element = ET.SubElement(optics, "optic")
                                    for opt_key, opt_value in optic.items():
                                        opt_child = ET.SubElement(optic_element, opt_key)
                                        opt_child.text = str(opt_value)  # Ensure value is a string
                            else:
                                cam_child = ET.SubElement(camera_element, cam_key)
                                cam_child.text = str(cam_value)  # Ensure value is a string
                
                elif key == "image_list":
                    images = ET.SubElement(station_element, "image_list")
                    for image in value:
                        image_element = ET.SubElement(images, "image")
                        image_element.text = image

                else:
                    child = ET.SubElement(station_element, key)
                    child.text = str(value)  # Ensure value is a string
            
        # NÁZEV PROJEKTU ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        project_name = ET.SubElement(root1,"project_name")
        if self.project_name == None:
            self.project_name = ""
        project_name.text = str(self.project_name)

        # ULOŽENÍ ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        tree1 = ET.ElementTree(root1)
        try:
            tree1.write(self.xml_file_path, encoding="utf-8", xml_declaration=True)
            Tools.add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně uložen","green",None,True)
        except Exception as e:
            Tools.add_colored_line(self.main_console,f"Neočekávaná chyba {e}","red",None,True)

    def read_xml_data(self,file_path):
        stations = self.read_stations_xml(file_path)
        controllers = self.read_controllers(file_path)
        project_name = self.read_project_name(file_path)
        print("stations, controllers, project_name: ",[stations,controllers,project_name])
        return [stations,controllers,project_name]

    def read_controllers(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        controllers = []
        controller_list = root.find("controllers")
        for controller in controller_list.findall("controller"):
            controller_data = {}
            for child in controller:
                if child.tag == "accessory_list":
                    accessory_list = []
                    for accessory in child.findall("accessory"):
                        accessory_data = {}
                        for acc_child in accessory:
                            if acc_child.text is not None:
                                accessory_data[acc_child.tag] = acc_child.text
                            else:
                                accessory_data[acc_child.tag] = ""
                        accessory_list.append(accessory_data)
                    controller_data[child.tag] = accessory_list
                else:                        
                    if child.text is not None:
                        controller_data[child.tag] = child.text
                    else:
                        controller_data[child.tag] = ""

            controllers.append(controller_data)
        
        return controllers
     
    def read_stations_xml(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        stations = []
        station_list = root.find("station_list")
        for station in station_list.findall("station"):
            station_data = {}
            for child in station:
                if child.tag == "camera_list":
                    camera_list = []
                    for camera in child.findall("camera"):
                        camera_data = {}
                        for cam_child in camera:
                            if cam_child.tag == "optics_list":
                                optics_list = []
                                for optic in cam_child.findall("optic"):
                                    optic_data = {}
                                    for opt_child in optic:
                                        if opt_child.text is not None:
                                            optic_data[opt_child.tag] = opt_child.text
                                        else:
                                            optic_data[opt_child.tag] = ""
                                    optics_list.append(optic_data)
                                camera_data[cam_child.tag] = optics_list
                            else:
                                if cam_child.text is not None:
                                    camera_data[cam_child.tag] = cam_child.text
                                else:
                                    camera_data[cam_child.tag] = ""
                        camera_list.append(camera_data)
                    station_data[child.tag] = camera_list
                
                elif child.tag == "image_list":
                    image_list = []
                    for image in child.findall("image"):
                        if image.text is not None:
                            image_list.append(image.text)
                    station_data[child.tag] = image_list
                    
                else:
                    if child.text is not None:
                        station_data[child.tag] = child.text
                    else:
                        station_data[child.tag] = ""

                         
            stations.append(station_data)
        return stations

    def read_project_name(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        project_name = ""
        
        project_name_element = root.find("project_name")
        if project_name_element is not None:
            project_name = project_name_element.text
        else:
            project_name = ""
        
        return project_name

class FakeContextMenu(customtkinter.CTkScrollableFrame):
    def __init__(self, parent, values, command=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.command = command
        self.buttons = []
        self.one_button_height = 50
        self._scrollbar.configure(width=30)
        self._scrollbar.configure(corner_radius=10)

        for val in values:
            btn = customtkinter.CTkButton(self, text=str(val), font=("Arial", 20), fg_color="transparent", hover_color="gray25",
                                command=lambda v=val: self.on_select(v))
            btn.pack(fill="x", pady=2,expand=True)
            self.one_button_height = btn._current_height
            self.buttons.append(btn)

    def on_select(self, value):
        if self.command:
            self.command(value)

class ToplevelWindow:
    @classmethod
    def export_to_db_window(cls,root,app_icon_path,server_connection,project_name,table_to_export,callback,main_console):
        try:
            cursor = server_connection.cursor()
            cursor.execute("SELECT @@SERVERNAME")
            current_server_name = str(cursor.fetchone()[0])
            cursor.close()
        except Exception as e:
            print(e)
            callback()
            return

        def close_window(window):
            window.destroy()

        def export_table():
            def check_table():
                cursor.execute(f"""
                IF NOT EXISTS (
                    SELECT * FROM INFORMATION_SCHEMA.TABLES
                    WHERE TABLE_NAME = '{table_name}' AND TABLE_SCHEMA = 'dbo'
                )
                BEGIN
                    CREATE TABLE [dbo].[{table_name}] (
                        id_user INT,
                        stanice NVARCHAR(100),
                        sw_kategorie NVARCHAR(100),
                        vyrobce NVARCHAR(100),
                        hw_kategorie NVARCHAR(100),
                        oznaceni NVARCHAR(100),
                        typ_zarizeni NVARCHAR(100),
                        alternativa NVARCHAR(100),
                        master_device NVARCHAR(100),
                        posledni_uprava NVARCHAR(100)
                    )
                END
                """)
            cursor = server_connection.cursor()
            table_name = str(table_entry.get())
            try:
                cursor.execute(f"USE {str(db_name_entry.get())}")
            except Exception as db_err:
                Tools.add_colored_line(console,f"Nepodařilo se přepnout do databáze ({db_err})","red",None,True)
                return
        
            try:
                check_table()
            except Exception as table_err:
                Tools.add_colored_line(console,f"Tabulku se nepodařilo vytvořit ({table_err})","red",None,True)
                return
                
            # insert_query = f"INSERT INTO {table_name} (id_user, stanice, sw_kategorie, vyrobce, hw_kategorie, oznaceni, typ_zarizeni, alternativa, master_device, posledni_uprava) VALUES (?,?,?,?,?,?,?,?,?,?)"
            insert_query = f"""
                INSERT INTO [dbo].[{table_name}] (
                    id_user, stanice, sw_kategorie, vyrobce, hw_kategorie, oznaceni, 
                    typ_zarizeni, alternativa, master_device, posledni_uprava
                ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """
            try:
                cursor.executemany(insert_query, table_to_export)
                server_connection.commit()
                Tools.add_colored_line(main_console,f"Tabulka byla úspěšně vyexportována","green",None,True)
                close_window(window)
            except Exception as export_err:
                Tools.add_colored_line(console,f"Tabulku se nepodařilo exportovat ({export_err})","red",None,True)
            finally:
                cursor.close()
            # server_connection.close()

        def init_fill_option_menu():
            cursor = server_connection.cursor()
            cursor.execute("SELECT name FROM sys.databases ORDER BY name")
            databases = [row[0] for row in cursor.fetchall()]
            cursor.close()
            db_name_entry.configure(values = databases)
            if len(databases) > 0:
                db_name_entry.set(databases[0])

        window = customtkinter.CTkToplevel(fg_color="#212121")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Možnosti exportu do databáze")
        label_column_width = 200
        top_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        server_frame =      customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        server_name =       customtkinter.CTkLabel(master = server_frame,text = "Aktuální připojení:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        server_entry =      customtkinter.CTkLabel(master = server_frame,text = current_server_name,font=("Arial",22,"bold"),anchor="w",text_color="green")
        server_name.        pack(pady=5,padx=5,anchor="w",side="left")
        server_entry.       pack(pady=5,padx=(5,0),anchor="w",side="left")
        server_frame.       pack(pady=5,padx=5,anchor="w",side="top",fill="x",expand=True)
        db_name_frame =     customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        db_name =           customtkinter.CTkLabel(master = db_name_frame,text = "Název databáze:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        db_name_entry =     customtkinter.CTkOptionMenu(master = db_name_frame,font=("Arial",22),dropdown_font=("Arial",22),values=[],height=50,corner_radius=0)
        db_name.            pack(pady=5,padx=5,anchor="w",side="left")
        db_name_entry.      pack(pady=5,padx=(5,0),anchor="w",side="left",fill="x",expand=True)
        db_name_frame.      pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        table_frame =       customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        table =             customtkinter.CTkLabel(master = table_frame,text = "Název tabulky:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        table_entry =       customtkinter.CTkEntry(master = table_frame,font=("Arial",22),height=50,corner_radius=0)
        table.              pack(pady=5,padx=5,anchor="w",side="left")
        table_entry.        pack(pady=5,padx=(5,0),anchor="w",side="left",fill="x",expand=True)
        table_frame.        pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        uid_frame =         customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        uid =               customtkinter.CTkLabel(master = uid_frame,text = "Uživatelské id:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        uid_entry =         customtkinter.CTkEntry(master = uid_frame,font=("Arial",22),height=50,corner_radius=0)
        uid.                pack(pady=5,padx=5,anchor="w",side="left")
        uid_entry.          pack(pady=5,padx=5,anchor="w",side="left",fill="x",expand=True)
        uid_frame.          pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        top_frame.          pack(pady=(10,5),padx=5,anchor="w",side="top",fill="both",expand=True)
        console =           tk.Text(window, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        console.            pack(pady = (0,10), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        button_frame =      customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        button_connect =    customtkinter.CTkButton(master = button_frame,text = "Exportovat",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: export_table())
        button_close =      customtkinter.CTkButton(master = button_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_close.       pack(pady = (0,10), padx = (5,10),side="right",anchor = "e")
        button_connect.     pack(pady = (0,10), padx = (5,10),side="right",anchor = "e")
        button_frame.       pack(pady=0,padx=0,anchor="w",side="top",fill="both")

        init_fill_option_menu()

        uid_entry.delete("0","200")
        uid_entry.insert("0", "1111")
        table_entry.delete("0","200")
        table_entry.insert("0", f"{project_name}_SW_CAMERA")

        root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        root.update_idletasks()
        window.geometry(f"{window.winfo_width()}x{window.winfo_height()}+{root._current_width/2}+{root._current_height/2}")
        window.after(100,window.focus_force())
        window.focus()
        return window

    @classmethod
    def db_login_window(cls,root,app_icon_path,db_label,callback,call_export,call_export_callback,main_console):
        window = customtkinter.CTkToplevel(fg_color="#212121")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Připojení k databázi")

        def close_window(window):
            window.destroy()

        def call_server_context_menu(parameter):
            """
            parameter:
            - server
            - login
            - db_name
            """
            config = Tools.read_json_config()
            if parameter == "server":
                history_list = config.get("db_settings", {}).get("server_history_list", [])
                entry_widget = server_entry
            elif parameter == "login":
                history_list = config.get("db_settings", {}).get("username_history_list", [])
                entry_widget = username_entry
            elif parameter == "db_name":
                history_list = config.get("db_settings", {}).get("db_name_list", [])
                entry_widget = db_name_entry

            print("current history: ",history_list)
            def insert_path(server_name):
                entry_widget.delete("0","200")
                entry_widget.insert("0", server_name)

            if len(history_list) > 0:
                path_context_menu = tk.Menu(window,tearoff=0,fg="white",bg="#202020",activebackground="#606060")
                for i in range(0,len(history_list)):
                    path_context_menu.add_command(label=history_list[i], command=lambda server_name = history_list[i]: insert_path(server_name),font=("Arial",22,"bold"))
                    if i < len(history_list)-1:
                        path_context_menu.add_separator()
                        
                path_context_menu.tk_popup(entry_widget.winfo_rootx(),entry_widget.winfo_rooty()+40)
            else:
                Tools.add_colored_line(console,"Prozatím žádná historie","orange",None,True)

        def connect_to_server():
            def add_new_param_to_history(uid_filled=False):
                def update_array(param,param_list):
                    if param not in param_list:
                        if str(param) == "":
                            return
                        if len(param_list) > 9:
                            param_list.pop()
                        param_list.insert(0,str(param))  

                config = Tools.read_json_config()
                server_history = config.get("db_settings", {}).get("server_history_list", [])
                db_name_history = config.get("db_settings", {}).get("db_name_list", [])
                login_history = config.get("db_settings", {}).get("username_history_list", [])
                update_array(server_name,server_history)
                update_array(database_name,db_name_history)
                Tools.save_to_json_config(server_history,"db_settings","server_history_list")
                Tools.save_to_json_config(db_name_history,"db_settings","db_name_list")
                if uid_filled:
                    update_array(uid,login_history)
                    Tools.save_to_json_config(login_history,"db_settings","username_history_list")
            

            if len(str(server_entry.get()).replace(" ","")) == 0:
                Tools.add_colored_line(console,"Zadejte název serveru","red",None,True)
                return
            # if len(str(db_name_entry.get()).replace(" ","")) == 0:
            #     Tools.add_colored_line(console,"Zadejte název databáze","red",None,True)
            #     return
            database_name = str(db_name_entry.get())
            server_name = str(server_entry.get())
            
            if str(username_entry.get()).replace(" ","") == "" or str(pwd_entry.get()).replace(" ","") == "":
                try:
                    conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};'
                        rf'SERVER={server_name};'
                        # rf'DATABASE={database_name};'
                        r'Trusted_Connection=yes;'
                    )
                    conn = pyodbc.connect(conn_str)
                    add_new_param_to_history()
                    Tools.add_colored_line(main_console,f"Úspěšně připojeno k serveru: {server_name}","green",None,True)
                    db_label.configure(text_color = "green",text=f"Přihlášen k: {server_name}")
                    callback(conn)
                    close_window(window)
                    if call_export:
                        call_export_callback()

                except Exception as e:
                    Tools.add_colored_line(console,f"K serveru: {server_name} se nepodařilo se připojit ({e})","red",None,True)
                    return
            else:
                uid = str(username_entry.get())
                pwd = str(pwd_entry.get())
                try:
                    conn_str = (
                        r'DRIVER={ODBC Driver 17 for SQL Server};'
                        rf'SERVER={server_name};'
                        # rf'DATABASE={database_name};'
                        rf"UID={uid};"
                        rf"PWD={pwd};"
                        r'Trusted_Connection=no;'
                    )
                    conn = pyodbc.connect(conn_str)
                    add_new_param_to_history(uid_filled=True)
                    Tools.add_colored_line(main_console,f"Úspěšně připojeno k serveru: {server_name}","green",None,True)
                    db_label.configure(text_color = "green",text=f"Přihlášen k: {server_name}")
                    callback(conn)
                    close_window(window)
                    if call_export:
                        call_export_callback()
                except Exception as e:
                    Tools.add_colored_line(console,f"K serveru: {server_name} se nepodařilo se připojit ({e})","red",None,True)
                    return
                
        def init_fill_entry():
            config = Tools.read_json_config()
            server_list = config.get("db_settings", {}).get("server_history_list", [])
            if len(server_list)>0:
                server_entry.delete("0","200")
                server_entry.insert("0", server_list[0])
            login_list = config.get("db_settings", {}).get("username_history_list", [])
            if len(login_list)>0:
                username_entry.delete("0","200")
                username_entry.insert("0", login_list[0])
            db_name_list = config.get("db_settings", {}).get("db_name_list", [])
            if len(db_name_list)>0:
                db_name_entry.delete("0","200")
                db_name_entry.insert("0", db_name_list[0])
                
        label_column_width = 200
        top_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        server_frame =      customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        server_name =       customtkinter.CTkLabel(master = server_frame,text = "Název serveru:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        context_menu_button = customtkinter.CTkButton(master = server_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        server_entry =      customtkinter.CTkEntry(master = server_frame,font=("Arial",22),height=50,corner_radius=0)
        server_name.        pack(pady=5,padx=5,anchor="w",side="left")
        server_entry.       pack(pady=5,padx=(5,0),anchor="w",side="left",fill="x",expand=True)
        context_menu_button.pack(pady=5,padx=(0,5),anchor="w",side="left")
        server_frame.       pack(pady=5,padx=5,anchor="w",side="top",fill="x",expand=True)
        db_name_frame =     customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        db_name =           customtkinter.CTkLabel(master = db_name_frame,text = "Název databáze:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        db_name_entry =     customtkinter.CTkEntry(master = db_name_frame,font=("Arial",22),height=50,corner_radius=0)
        db_context =        customtkinter.CTkButton(master = db_name_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        db_name.            pack(pady=5,padx=5,anchor="w",side="left")
        db_name_entry.      pack(pady=5,padx=(5,0),anchor="w",side="left",fill="x",expand=True)
        db_context.         pack(pady=5,padx=(0,5),anchor="w",side="left")
        db_name_frame.      pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        login_frame =       customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        username =          customtkinter.CTkLabel(master = login_frame,text = "Login:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        username_entry =    customtkinter.CTkEntry(master = login_frame,font=("Arial",22),height=50,corner_radius=0)
        username_context =  customtkinter.CTkButton(master = login_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        username.           pack(pady=5,padx=5,anchor="w",side="left")
        username_entry.     pack(pady=5,padx=(5,0),anchor="w",side="left",fill="x",expand=True)
        username_context.   pack(pady=5,padx=(0,5),anchor="w",side="left")
        login_frame.        pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        pwd_frame =         customtkinter.CTkFrame(master = top_frame,corner_radius=0,fg_color="#212121",border_width=1)
        pwd =               customtkinter.CTkLabel(master = pwd_frame,text = "Heslo:",font=("Arial",22,"bold"),width=label_column_width,anchor="w")
        pwd_entry =         customtkinter.CTkEntry(master = pwd_frame,font=("Arial",22),height=50,corner_radius=0)
        pwd.                pack(pady=5,padx=5,anchor="w",side="left")
        pwd_entry.          pack(pady=5,padx=5,anchor="w",side="left",fill="x",expand=True)
        pwd_frame.          pack(pady=5,padx=5,anchor="w",side="top",fill="x")
        top_frame.          pack(pady=(10,5),padx=5,anchor="w",side="top",fill="both",expand=True)
        console =           tk.Text(window, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        console.            pack(pady = (0,10), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        button_frame =      customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        button_connect =    customtkinter.CTkButton(master = button_frame,text = "Připojit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: connect_to_server())
        button_close =      customtkinter.CTkButton(master = button_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_close.       pack(pady = (0,10), padx = (5,10),side="right",anchor = "e")
        button_connect.     pack(pady = (0,10), padx = (5,10),side="right",anchor = "e")
        button_frame.       pack(pady=0,padx=0,anchor="w",side="top",fill="both")
        context_menu_button.bind("<Button-1>", lambda e: call_server_context_menu("server"))
        db_context.         bind("<Button-1>", lambda e: call_server_context_menu("db_name"))
        username_context.   bind("<Button-1>", lambda e: call_server_context_menu("login"))
        init_fill_entry()

        root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        root.update_idletasks()
        window.geometry(f"{window.winfo_width()}x{window.winfo_height()}+{root._current_width/2}+{root._current_height/2}")
        window.after(100,window.focus_force())
        window.focus()
        return window

    @classmethod
    def save_prog_options_window(cls,
                                 root,
                                 app_icon_path,
                                 custom_controller_database,
                                 main_console,
                                 station_list,
                                 project_name,
                                 callback,
                                 callback_save_last_file,
                                 last_file = None,
                                 last_path = "",
                                 default_xml_file_name="_metadata_catalogue",
                                 default_path = "",
                                 exit_status = False,
                                 only_save = True):
        """
        okno s možnostmi uložení rozdělaného projektu
        """
        window = customtkinter.CTkToplevel(fg_color="#212121")
        # window.geometry(f"1015x350+{self.x+200}+{self.y+50}")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        if only_save:
            window.title("Možnosti uložení projektu")
        else:
            window.title("Možnosti importování projektu")
        subwindow = ""

        def confirm_window(final_path):
            nonlocal subwindow
            def call_save(final_path):
                nonlocal subwindow
                save_file(final_path)
                subwindow.destroy()
            subwindow = customtkinter.CTkToplevel(fg_color="#212121")
            subwindow.after(200, lambda: subwindow.iconbitmap(app_icon_path))
            subwindow.geometry(f"700x150+{root.winfo_rootx()+250}+{root.winfo_rooty()+100}")
            subwindow.title("Potvrdit přepsání souboru")
            top_frame =         customtkinter.CTkFrame(master = subwindow,corner_radius=0,fg_color="#212121")
            warning_icon =      customtkinter.CTkLabel(master = top_frame,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/warning.png")),size=(50,50)),bg_color="#212121")
            export_label =      customtkinter.CTkLabel(master = top_frame,text = "V zadané cestě se soubor s tímto názvem již vyskytuje, přepsat?",font=("Arial",22,"bold"),justify = "left",text_color="orange")
            button_yes =        customtkinter.CTkButton(master = subwindow,text = "Ano",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command = lambda: call_save(final_path))
            button_no =         customtkinter.CTkButton(master = subwindow,text = "Ne",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command = lambda: subwindow.destroy())
            warning_icon.       pack(pady=10,padx=30,side = "left",anchor="w")
            export_label.       pack(pady=10,padx=10,side = "left",anchor="w")      
            top_frame.          pack(pady=0,padx=0,expand=False,side = "top",anchor="w")
            button_no.          pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            button_yes.         pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            subwindow.update()
            subwindow.update_idletasks()
            subwindow.focus_force()
            subwindow.focus()
            subwindow.grab_set()

        def close_window(window,button = None):
            nonlocal subwindow
            try:
                if subwindow.winfo_exists():
                    subwindow.destroy()
            except Exception:
                pass
            window.destroy()
            if button:
                callback_save_last_file(None,None,None)

        def create_path(path_inserted):
            nonlocal export_name
            nonlocal default_xml_file_name
            file_name = export_name.get()
            if file_name =="":
                file_name = default_xml_file_name
            path = path_inserted + file_name
            if not path.endswith(".xml"):
                path = path + ".xml"
            print(path)
            return path

        def save_file(final_path):
            nonlocal window
            # ukladani posledne zadaneho nazvu souboru:
            nonlocal export_name
            nonlocal export_path
            path_inserted = export_path.get()
            file_name = export_name.get()
            # samotne ukladani vsech dat:
            save_prog = Save_prog_metadata(station_list=station_list,project_name=project_name,controller_database=custom_controller_database,console=console,xml_file_path=final_path)
            save_prog.store_xml_data()
            Tools.add_colored_line(main_console,f"Data úspěšně uložena do: {final_path}","green",None,True)
            close_window(window)
            callback_save_last_file(file_name,path_inserted,None,True)

        def call_save_file(window):
            nonlocal console
            nonlocal export_path
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                final_path = create_path(path_inserted)
                if os.path.exists(final_path):
                    confirm_window(final_path)
                else:
                    save_file(final_path)
            else:
                Tools.add_colored_line(console,"Zadaná cesta pro uložení je neplatná","red",None,True)
        
        def call_load_file(window):
            nonlocal console
            nonlocal export_path
            nonlocal export_name
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                final_path = create_path(path_inserted)
                save_prog = Save_prog_metadata(station_list=station_list,project_name=project_name,controller_database=custom_controller_database,console=console)
                try:
                    received_data = save_prog.read_xml_data(final_path)
                    Tools.add_colored_line(main_console,f"Data úspěšně nahrána z: {final_path}","green",None,True)
                    callback(received_data)
                    # ulozit posledně načtený soubor
                    file_name = export_name.get()
                    callback_save_last_file(file_name,path_inserted,None,True)
                    window.destroy()
                except Exception:
                    Tools.add_colored_line(console,f"Soubor .xml je neplatný: {final_path}","red",None,True)
                    # window.destroy()
            else:
                Tools.add_colored_line(console,f"V zadané cestě nebyl nalezen soubor .xml s názvem {export_name.get()}","red",None,True)

        def call_browse_directories(what_search,file_extension = [("All files", "*.*")]):
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if what_search == "only_dirs":
                output = Tools.browseDirectories(what_search)
                if str(output[1]) != "/":
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    Tools.add_new_path_to_history(str(output[1]),"catalogue_settings")
                    Tools.add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            else:
                output = Tools.browseDirectories(what_search,file_type=file_extension)
                if str(output[1]) != "/":
                    export_name.delete(0,300)
                    name_without_extension = str(output[2])[:-4]
                    export_name.insert(0, name_without_extension)
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    Tools.add_new_path_to_history(str(output[1]),"catalogue_settings")
                    Tools.add_colored_line(console,"Byla vložena cesta a název souboru","green",None,True)
            print(output[0])

            window.focus_force()
            window.focus()

        def search_for_xmls(path):
            found_files = []
            for files in os.listdir(path):
                if ".xml" in files:
                    if not files in found_files:
                        found_files.append(files)
            return found_files

        def save_current_path():
            path_inserted = str(export_path.get())
            if path_inserted.replace(" ","") != "":
                checked_path = Tools.path_check(path_inserted)
                if checked_path != False:
                    Tools.add_new_path_to_history(checked_path,"catalogue_settings")
                    Tools.save_to_json_config(checked_path,"catalogue_settings","default_path")
                    callback_save_last_file(None,None,checked_path)
                    Tools.add_colored_line(console,f"Zvolená cesta uložena: {checked_path}","green",None,True)

        export_frame =          customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        export_label =          customtkinter.CTkLabel(master = export_frame,text = "Zadejte název souboru:",font=("Arial",22,"bold"))
        export_name_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button  =  customtkinter.CTkButton(master = export_name_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_name =           customtkinter.CTkEntry(master = export_name_frame,font=("Arial",20),width=730,height=50,corner_radius=0)
        explorer_btn_name =     customtkinter.CTkButton(master = export_name_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("all",[("XML files", "*.xml"),("All files", "*.*")]))
        format_entry =          customtkinter.CTkOptionMenu(master = export_name_frame,font=("Arial",22),dropdown_font=("Arial",22),values=[".xml"],width=200,height=50,corner_radius=0)
        context_menu_button     .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_name             .pack(pady = 5, padx = 0,anchor="w",fill="x",expand=True,side="left")
        format_entry            .pack(pady = 5, padx = (5,10),anchor="e",expand=False,side="right")
        explorer_btn_name       .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        export_label2 =         customtkinter.CTkLabel(master = export_frame,text = "Zadejte cestu, kam soubor uložit:",font=("Arial",22,"bold"))
        export_path_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button2  = customtkinter.CTkButton(master = export_path_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_path =           customtkinter.CTkEntry(master = export_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        explorer_btn =          customtkinter.CTkButton(master = export_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("only_dirs"))
        save_path_btn =         customtkinter.CTkButton(master = export_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: save_current_path())
        context_menu_button2    .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_path             .pack(pady = 5, padx = (0,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn           .pack(pady = 5, padx = (5,10),anchor="e",expand=False,side="right")
        explorer_btn            .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        console =               tk.Text(export_frame, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        button_frame =          customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        button_load =           customtkinter.CTkButton(master = button_frame,text = "Nahrát",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_load_file(window))
        button_save =           customtkinter.CTkButton(master = button_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_save_file(window))
        button_exit =           customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window,True))
        button_exit             .pack(pady = 10, padx = (5,10),expand=False,side="right",anchor = "e")
        if not only_save:
            button_load         .pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")
        else:
            button_save         .pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")
        export_frame            .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        export_label            .pack(pady=(15,5),padx=10,anchor="w",expand=False,side="top")
        export_name_frame       .pack(expand=False,side="top",anchor="n",fill="x")
        export_label2           .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        export_path_frame       .pack(expand=False,side="top",anchor="n",fill="x")
        console                 .pack(pady = (10,0), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        button_frame            .pack(pady = 0, padx = (0),side="top",fill="x",anchor = "w")

        context_menu_button.bind("<Button-1>", lambda e: Tools.call_path_context_menu(window,export_name,context_menu_button,search_for_xmls(export_path.get())))
        context_menu_button2.bind("<Button-1>", lambda e: Tools.call_path_context_menu(window,export_path,context_menu_button2))

        if not only_save:
            export_label.configure(text = "Vyberte soubor:")
            export_label2.configure(text = "Zadejte cestu k souboru:")

        if exit_status:
            button_load.configure(state = "disabled")

        checked_last_path = Tools.path_check(last_path)
        default_path = Tools.path_check(default_path)
        if checked_last_path != False and checked_last_path != None and checked_last_path.replace(" ","") != "" and checked_last_path.replace(" ","") != "/":
            initial_path = str(checked_last_path)
            export_path.insert("0",Tools.resource_path(str(checked_last_path)))
            Tools.add_colored_line(console,"Byla vložena posledně zvolená cesta","green",None,True)

        elif default_path != False and default_path != None and default_path.replace(" ","") != "" and default_path.replace(" ","") != "/":
            initial_path = str(default_path)
            export_path.insert("0",Tools.resource_path(str(default_path)))
            Tools.add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)
        else:
            # initial_path = Tools.path_check(os.getcwd())
            export_path.insert("0",Tools.resource_path(str(initial_path)))

        found_xmls = search_for_xmls(initial_path)
        
        # posledni ulozeny/ nacteny soubor:
        if last_file != None and last_file.replace(" ","") != "":
            export_name.insert("0",str(last_file))
        # první soubor nalezeny ve slozce:
        elif len(found_xmls) > 0:
            export_name.insert("0",str(found_xmls[0].replace(".xml","")))
            print("nalezené soubory xml: ",found_xmls)
        # default název + název projektu:
        else:
            export_name.insert("0",str(project_name) + default_xml_file_name)

        root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        # x = root.winfo_rootx()
        # y = root.winfo_rooty()
        # window.geometry(f"{window.winfo_width()}x{window.winfo_height()}+{x+250}+{y+150}")
        window.after(100,window.focus_force())
        window.focus()
        return window

    @classmethod
    def load_prog_window(cls,
                        root,
                        app_icon_path,
                        custom_controller_database,
                        main_console,
                        station_list,
                        project_name,
                        callback,
                        callback_save_last_file,
                        last_file = None,
                        last_path = "",
                        default_xml_file_name="_metadata_catalogue",
                        default_path = "",
                        exit_status = False
                        ):
        """
        okno s možnostmi uložení rozdělaného projektu
        """
        window = customtkinter.CTkToplevel(fg_color="#212121")
        # window.geometry(f"1015x350+{self.x+200}+{self.y+50}")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Možnosti importování projektu")
        subwindow = ""

        def close_window(window,button = None):
            nonlocal subwindow
            try:
                if subwindow.winfo_exists():
                    subwindow.destroy()
            except Exception:
                pass
            window.destroy()
            if button:
                callback_save_last_file(None,None,None)

        def create_path(path_inserted):
            nonlocal export_name
            nonlocal default_xml_file_name
            file_name = export_name.get()
            if file_name =="":
                file_name = default_xml_file_name
            path = path_inserted + file_name
            if not path.endswith(".xml"):
                path = path + ".xml"
            print(path)
            return path
        
        def call_load_file(window):
            nonlocal console
            nonlocal export_path
            nonlocal export_name
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                final_path = create_path(path_inserted)
                save_prog = Save_prog_metadata(station_list=station_list,project_name=project_name,controller_database=custom_controller_database,console=console)
                try:
                    received_data = save_prog.read_xml_data(final_path)
                    Tools.add_colored_line(main_console,f"Data úspěšně nahrána z: {final_path}","green",None,True)
                    callback(received_data)
                    # ulozit posledně načtený soubor
                    file_name = export_name.get()
                    callback_save_last_file(file_name,path_inserted,None,True)
                    window.destroy()
                except Exception:
                    Tools.add_colored_line(console,f"Soubor .xml je neplatný: {final_path}","red",None,True)
                    # window.destroy()
            else:
                Tools.add_colored_line(console,f"V zadané cestě nebyl nalezen soubor .xml s názvem {export_name.get()}","red",None,True)

        def check_dir(path):
            if os.path.isfile(path):
                path = os.path.dirname(path) + "/"
            return path

        def call_browse_directories(what_search,file_extension = [("All files", "*.*")]):
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if what_search == "only_dirs":
                output = Tools.browseDirectories(what_search)
                if str(output[1]) != "/":
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    Tools.add_new_path_to_history(str(output[1]),"catalogue_settings")
                    Tools.add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            else:
                start_path_to_give = check_dir(export_path.get())
                if not os.path.exists(start_path_to_give):
                    start_path_to_give = None
                output = Tools.browseDirectories(what_search,start_path=start_path_to_give,file_type=file_extension)
                if str(output[1]) != "/":
                    export_name.delete(0,300)
                    name_without_extension = str(output[2])[:-4]
                    export_name.insert(0, name_without_extension)
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    Tools.add_new_path_to_history(str(output[1]),"catalogue_settings")
                    Tools.add_colored_line(console,"Byla vložena cesta a název souboru","green",None,True)
            print(output[0])

            window.focus_force()
            window.focus()

        def search_for_xmls(path):
            found_files = []
            for files in os.listdir(path):
                if ".xml" in files:
                    if not files in found_files:
                        found_files.append(files)
            return found_files

        def save_current_path():
            path_inserted = str(export_path.get())
            if path_inserted.replace(" ","") != "":
                checked_path = Tools.path_check(path_inserted)
                if checked_path != False:
                    Tools.add_new_path_to_history(checked_path,"catalogue_settings")
                    Tools.save_to_json_config(checked_path,"catalogue_settings","default_path")
                    callback_save_last_file(None,None,checked_path)
                    Tools.add_colored_line(console,f"Zvolená cesta uložena: {checked_path}","green",None,True)

        export_frame =          customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        export_label =          customtkinter.CTkLabel(master = export_frame,text = "Vyberte soubor:",font=("Arial",22,"bold"))
        export_name_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button  =  customtkinter.CTkButton(master = export_name_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_name =           customtkinter.CTkEntry(master = export_name_frame,font=("Arial",20),width=730,height=50,corner_radius=0)
        explorer_btn_name =     customtkinter.CTkButton(master = export_name_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("all",[("XML files", "*.xml"),("All files", "*.*")]))
        format_entry =          customtkinter.CTkOptionMenu(master = export_name_frame,font=("Arial",22),dropdown_font=("Arial",22),values=[".xml"],width=200,height=50,corner_radius=0)
        context_menu_button     .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_name             .pack(pady = 5, padx = 0,anchor="w",fill="x",expand=True,side="left")
        format_entry            .pack(pady = 5, padx = (5,10),anchor="e",expand=False,side="right")
        explorer_btn_name       .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        export_label2 =         customtkinter.CTkLabel(master = export_frame,text = "Zadejte cestu k souboru/ souborům xml:",font=("Arial",22,"bold"))
        export_path_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button2  = customtkinter.CTkButton(master = export_path_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_path =           customtkinter.CTkEntry(master = export_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        explorer_btn =          customtkinter.CTkButton(master = export_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("only_dirs"))
        save_path_btn =         customtkinter.CTkButton(master = export_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: save_current_path())
        context_menu_button2    .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_path             .pack(pady = 5, padx = (0,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn           .pack(pady = 5, padx = 5,anchor="e",expand=False,side="right")
        explorer_btn            .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        console =               tk.Text(export_frame, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        button_frame =          customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        button_load =           customtkinter.CTkButton(master = button_frame,text = "Nahrát",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_load_file(window))
        button_exit =           customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window,True))
        button_exit             .pack(pady = 10, padx = (5,10),expand=False,side="right",anchor = "e")
        button_load             .pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")
        export_frame            .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        export_label            .pack(pady=(15,5),padx=10,anchor="w",expand=False,side="top")
        export_name_frame       .pack(expand=False,side="top",anchor="n",fill="x")
        export_label2           .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        export_path_frame       .pack(expand=False,side="top",anchor="n",fill="x")
        console                 .pack(pady = (10,0), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        button_frame            .pack(pady = 0, padx = (0),side="top",fill="x",anchor = "w")
        context_menu_button.    bind("<Button-1>", lambda e: Tools.call_path_context_menu(window,export_name,context_menu_button,search_for_xmls(export_path.get())))
        context_menu_button2.   bind("<Button-1>", lambda e: Tools.call_path_context_menu(window,export_path,context_menu_button2))

        if exit_status:
            button_load.configure(state = "disabled")

        checked_last_path = Tools.path_check(last_path)
        default_path = Tools.path_check(default_path)
        if checked_last_path != False and checked_last_path != None and checked_last_path.replace(" ","") != "" and checked_last_path.replace(" ","") != "/":
            initial_path = str(checked_last_path)
            export_path.insert("0",Tools.resource_path(str(checked_last_path)))
            Tools.add_colored_line(console,"Byla vložena posledně zvolená cesta","green",None,True)

        elif default_path != False and default_path != None and default_path.replace(" ","") != "" and default_path.replace(" ","") != "/":
            initial_path = str(default_path)
            export_path.insert("0",Tools.resource_path(str(default_path)))
            Tools.add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)
        else:
            # initial_path = Tools.path_check(os.getcwd())
            export_path.insert("0",Tools.resource_path(str(initial_path)))

        found_xmls = search_for_xmls(initial_path)
        
        # posledni ulozeny/ nacteny soubor:
        if last_file != None and last_file.replace(" ","") != "":
            export_name.insert("0",str(last_file))
        # první soubor nalezeny ve slozce:
        elif len(found_xmls) > 0:
            export_name.insert("0",str(found_xmls[0].replace(".xml","")))
            print("nalezené soubory xml: ",found_xmls)
        # default název + název projektu:
        else:
            export_name.insert("0",str(project_name) + default_xml_file_name)

        root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        # x = root.winfo_rootx()
        # y = root.winfo_rooty()
        # window.geometry(f"{window.winfo_width()}x{window.winfo_height()}+{x+250}+{y+150}")
        window.after(100,window.focus_force())
        window.focus()
        return window

    @classmethod
    def setting_window(cls,
                       root,
                       app_icon_path,
                       default_excel_name,
                       default_xml_name,
                       window_status,
                       callback,
                       default_database_filename,
                       detailed_view_status,
                       render_mode = "fast",
                       hover_trigger_mode="1"):
        def close_window(window):
            window.destroy()

        def save_changes():
            def filter_input(data):
                forbidden_formats = [".","xml","xlsm","xlsx"]
                for formats in forbidden_formats:
                    data = data.replace(formats,"")
                return data

            if checkbox.get() == 1:
                window_status = 1
            else:
                window_status = 0
            default_excel_name = filter_input(str(excel_name_label_entry.get()))
            default_xml_name = filter_input(str(xml_name_label_entry.get()))
            default_database_filename = filter_input(str(default_database_name_entry.get()))

            input_data = [default_excel_name,default_xml_name,window_status,default_database_filename]
            Tools.save_to_json_config(input_data[0],"catalogue_settings","catalogue_filename")
            Tools.save_to_json_config(input_data[1],"catalogue_settings","metadata_filename")
            Tools.save_to_json_config(input_data[2],"catalogue_settings","subwindow_behav")
            Tools.save_to_json_config(input_data[3] + ".xlsx","catalogue_settings","database_filename")

            callback(input_data)
            close_window(window)

        def open_all_data():
            callback(["open_all_cmd",show_all_data_chckbx.get()])

        def switch_hover_trigger_mode():
            Tools.save_to_json_config(str(checkbox2.get()),"catalogue_settings","hover_info_trigger_mode")
            callback(["hover_info_trigger_mode",str(checkbox2.get())])

        def switch_render_mode(mode):
            if mode == "fast":
                fast_render_mode.select()
                precise_render_mode.deselect()
                Tools.save_to_json_config("fast","catalogue_settings","render_mode")
                callback(["set_render_mode","fast"])
            else:
                fast_render_mode.deselect()
                precise_render_mode.select()
                Tools.save_to_json_config("precise","catalogue_settings","render_mode")
                callback(["set_render_mode","precise"])

        window = customtkinter.CTkToplevel(fg_color="#212121")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Nastavení")
        main_frame =                    customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        option1_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        checkbox =                      customtkinter.CTkCheckBox(master = option1_frame, text = "Okna editování otevírat maximalizované",font=("Arial",22,"bold"))#,command=lambda: save_new_behav_notes()
        checkbox2 =                     customtkinter.CTkCheckBox(master = option1_frame, text = "Aut. zobrazovat detailní info pod kurzorem myši",font=("Arial",22,"bold"),command=lambda: switch_hover_trigger_mode())
        checkbox.                       pack(pady = (20,0), padx = 10,anchor="w")
        checkbox2.                      pack(pady = 10, padx = 10,anchor="w")
        option2_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        xml_name_label =                customtkinter.CTkLabel(master = option2_frame,text = "Nastavte základní název pro ukládání rozpracovaného projektu:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        xml_name_frame =                customtkinter.CTkFrame(master = option2_frame,corner_radius=0,fg_color="#212121")
        xml_name_label_entry =          customtkinter.CTkEntry(master = xml_name_frame,font=("Arial",20),corner_radius=0)
        xml_extension_label =           customtkinter.CTkLabel(master = xml_name_frame,text = ".xml",font=("Arial",22,"bold"),justify = "left",anchor="w")
        xml_extension_label.            pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        xml_name_label_entry.           pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        xml_name_label.                 pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        xml_name_frame.                 pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        option3_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        excel_name_label =              customtkinter.CTkLabel(master = option3_frame,text = "Nastavte základní název pro ukládání excelu:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        excel_name_frame =              customtkinter.CTkFrame(master = option3_frame,corner_radius=0,fg_color="#212121")
        excel_name_label_entry =        customtkinter.CTkEntry(master = excel_name_frame,font=("Arial",20),corner_radius=0)
        excel_extension_label =         customtkinter.CTkLabel(master = excel_name_frame,text = ".xlsm/ .xlsx",font=("Arial",22,"bold"),justify = "left",anchor="w")
        excel_extension_label.          pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        excel_name_label_entry.         pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        excel_name_label.               pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        excel_name_frame.               pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        option4_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        default_database_name =         customtkinter.CTkLabel(master = option4_frame,text = "Nastavte základní název souboru databáze produktů:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        default_database_name_warning = customtkinter.CTkLabel(master = option4_frame,text = "(název se musí shodovat s názvem souboru na sharepointu)",font=("Arial",22),justify = "left",anchor="w",text_color="orange")
        default_database_name_frame =   customtkinter.CTkFrame(master = option4_frame,corner_radius=0,fg_color="#212121")
        default_database_name_entry =   customtkinter.CTkEntry(master = default_database_name_frame,font=("Arial",20),corner_radius=0)
        database_extension_label =      customtkinter.CTkLabel(master = default_database_name_frame,text = ".xlsx",font=("Arial",22,"bold"),justify = "left",anchor="w")
        database_extension_label.       pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        default_database_name_entry.    pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        default_database_name.          pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        default_database_name_warning.  pack(pady = 0, padx = 10,fill="x",anchor="w",side="top")
        default_database_name_frame.    pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")

        option5_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        show_all_data_chckbx =          customtkinter.CTkCheckBox(master = option5_frame,text = "Zobrazit všechna data (rozbalit vše)",font=("Arial",22,"bold"),command=lambda: open_all_data())
        show_all_data_chckbx.           pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
    
        option6_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1,fg_color="#212121")
        render_mode_label =             customtkinter.CTkLabel(master = option6_frame,text = "Nastavte chování vykreslování:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        checkbox_frame =                customtkinter.CTkFrame(master = option6_frame,corner_radius=0,fg_color="#212121")
        fast_render_mode =              customtkinter.CTkCheckBox(master = checkbox_frame,text = "Rychlé",font=("Arial",22,"bold"),command=lambda: switch_render_mode("fast"))
        precise_render_mode =           customtkinter.CTkCheckBox(master = checkbox_frame,text = "Precizní",font=("Arial",22,"bold"),command=lambda: switch_render_mode("precise"))
        fast_render_mode.               pack(pady = 10, padx = 10,anchor="w",side="left")
        precise_render_mode.            pack(pady = 10, padx = 10,anchor="w",side="left")
        render_mode_label.              pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        checkbox_frame.                 pack(pady = 0, padx = 10,anchor="w",side="top")
    
        button_save =                   customtkinter.CTkButton(master = main_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_changes())
        button_exit =                   customtkinter.CTkButton(master = main_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        main_frame.                     pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 10,ipadx=10)
        option1_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option2_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option3_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option4_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option5_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option6_frame.                  pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top",ipadx=5,ipady=5)
        button_exit.                    pack(pady = 10, padx = (5,10),expand=False,side="right",anchor = "e")
        button_save.                    pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")

        excel_name_label_entry.insert(0,str(default_excel_name))
        xml_name_label_entry.insert(0,str(default_xml_name))
        default_database_name_entry.insert(0,str(default_database_filename.replace(".xlsx","")))
        if detailed_view_status == True:
            show_all_data_chckbx.select()

        if window_status == 1:
            checkbox.select()

        if hover_trigger_mode == "1":
            checkbox2.select()

        if render_mode == "fast":
            fast_render_mode.select()
        else:
            precise_render_mode.select()

        root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        window_height = window.winfo_height()
        window_width = window.winfo_width()
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        window.geometry(f"{window_width}x{window_height}+{x+200}+{y+100}")
        window.focus_force()
        window.focus()
        return window

    @classmethod
    def export_option_window(cls,root,app_icon_path,export_to_excel_callback,excel_format_list,favourite_excel_format,last_path,default_path,default_excel_filename,project_name = ""):
        child_root = customtkinter.CTkToplevel(fg_color="#212121")
        child_root.after(200, lambda: child_root.iconbitmap(app_icon_path))
        child_root.title("Možnosti exportování souboru")

        def get_excel_path():
            nonlocal export_path
            nonlocal export_name
            nonlocal format_entry
            name_inserted = export_name.get()
            path_inserted = export_path.get()
            path_inserted = Tools.resource_path(path_inserted)
            if path_inserted.replace(" ","") == "":
                return None
            else:
                print("Cesta pro export: ",path_inserted + name_inserted + "." + format_entry.get())
                if name_inserted.endswith(".xlsx") or name_inserted.endswith(".xlsm"):
                    return path_inserted + name_inserted
                else:
                    return path_inserted + name_inserted + "." + format_entry.get()

        def call_save_file(child_root):
            nonlocal console
            nonlocal export_path
            path_inserted = export_path.get()

            if os.path.exists(path_inserted):
                excel_path_with_name = get_excel_path()
                if os.path.exists(excel_path_with_name): # kontrola souboru se stejným názvem
                    nonlocal click_count
                    nonlocal previous_path
                    click_count += 1
                    Tools.add_colored_line(console,f"Cesta již obsahuje soubor se stejným názvem, při druhém kliknutí na \"Uložit\" bude přepsán","orange",None,True)
                    if click_count > 1 and previous_path == excel_path_with_name: # když podruhé a nebyla změněna cesta
                        favourite_format = str(format_entry.get())
                        export_success = export_to_excel_callback(excel_path_with_name,favourite_format,path_inserted)
                        if export_success == True:
                            close_window(child_root)
                        else:
                            Tools.add_colored_line(console,export_success,"red",None,True)

                    elif click_count > 1 and previous_path != excel_path_with_name:
                        click_count =1
                    previous_path = excel_path_with_name
                else:
                    favourite_format = str(format_entry.get())
                    export_success = export_to_excel_callback(excel_path_with_name,favourite_format,path_inserted)
                    if export_success == True:
                        close_window(child_root)
                    else:
                        Tools.add_colored_line(console,export_success,"red",None,True)
            else:
                Tools.add_colored_line(console,"Zadaná cesta pro uložení je neplatná","red",None,True)

        def close_window(child_root):
            try:
                root.unbind("<Button-1>")
            except Exception:
                pass
            # child_root.quit()
            child_root.destroy()

        def call_browse_directories():
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            output = Tools.browseDirectories("only_dirs")
            if str(output[1]) != "/":
                export_path.delete(0,300)
                export_path.insert(0, str(output[1]))
                Tools.add_new_path_to_history(str(output[1]),"catalogue_settings")
                Tools.add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            print(output[0])
            child_root.focus()
            child_root.focus_force()

        def search_for_xlsxs(path):
            found_files = []
            for files in os.listdir(path):
                if ".xlsx" in files or ".xlsm" in files:
                    if not files in found_files:
                        found_files.append(files)
            return found_files

        def save_current_path():
            path_inserted = str(export_path.get())
            if path_inserted.replace(" ","") != "":
                checked_path = Tools.path_check(path_inserted)
                if checked_path != False:
                    export_to_excel_callback(None,None,path_inserted)
                    Tools.add_colored_line(console,f"Zvolená cesta uložena: {path_inserted}","green",None,True)
                    Tools.save_to_json_config(checked_path,"catalogue_settings","default_path")
                    Tools.add_new_path_to_history(checked_path,"catalogue_settings")
                    
        click_count = 0
        previous_path = ""
        export_frame =          customtkinter.CTkFrame(master = child_root,corner_radius=0,fg_color="#212121")
        export_label =          customtkinter.CTkLabel(master = export_frame,text = "Zadejte název souboru:",font=("Arial",22,"bold"))
        export_name_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button  =  customtkinter.CTkButton(master = export_name_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_name =           customtkinter.CTkEntry(master = export_name_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        format_entry =          customtkinter.CTkOptionMenu(master = export_name_frame,font=("Arial",22),dropdown_font=("Arial",22),width=200,height=50,values=excel_format_list,corner_radius=0)
        context_menu_button     .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_name             .pack(pady = 5, padx = (0,0),anchor="w",fill="x",expand=True,side="left")
        format_entry            .pack(pady = 5, padx = (5,10),anchor="e",expand=False,side="right")
        export_label2 =         customtkinter.CTkLabel(master = export_frame,text = "Zadejte cestu, kam soubor uložit:",font=("Arial",22,"bold"))
        export_path_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0,fg_color="#212121")
        context_menu_button2 =  customtkinter.CTkButton(master = export_path_frame, width = 50,height=50, text = "V",font=("Arial",20,"bold"),corner_radius=0,fg_color="#505050")
        export_path =           customtkinter.CTkEntry(master = export_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        save_path_btn =         customtkinter.CTkButton(master = export_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: save_current_path())
        explorer_btn =          customtkinter.CTkButton(master = export_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories())
        context_menu_button2    .pack(pady = 5, padx = (10,0),anchor="w",side="left")
        export_path             .pack(pady = 5, padx = (0,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn           .pack(pady = 5, padx = (5,10),anchor="e",expand=False,side="right")
        explorer_btn            .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        console =               tk.Text(export_frame, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        button_save =           customtkinter.CTkButton(master = export_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_save_file(child_root))
        button_exit =           customtkinter.CTkButton(master = export_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(child_root))
        export_frame            .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        export_label            .pack(pady=(15,5),padx=10,anchor="w",expand=False,side="top")
        export_name_frame       .pack(expand=True,side="top",anchor="n",fill="x")
        export_label2           .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        export_path_frame       .pack(expand=True,side="top",anchor="n",fill="x")
        console                 .pack(pady = (10,0), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        button_exit             .pack(pady = 10, padx = (5,10),expand=False,side="right",anchor = "e")
        button_save             .pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")
        context_menu_button.    bind("<Button-1>", lambda e: Tools.call_path_context_menu(child_root,export_name,context_menu_button,search_for_xlsxs(export_path.get())))
        context_menu_button2.   bind("<Button-1>", lambda e: Tools.call_path_context_menu(child_root,export_path,context_menu_button2))

        excel_filename = default_excel_filename
        if str(project_name.replace(" ","")) != "":
            excel_filename = default_excel_filename + "_projekt_" + str(project_name)
        export_name.insert("0",excel_filename)

        default_path = Tools.path_check(default_path)
        checked_last_path = Tools.path_check(last_path)
        
        if checked_last_path != False and checked_last_path != None and checked_last_path.replace(" ","") != "" and checked_last_path.replace(" ","") != "/":
            initial_path = Tools.resource_path(checked_last_path)
            Tools.add_colored_line(console,"Byla vložena poslední zvolená cesta","green",None,True)
        elif default_path != False and default_path != None and default_path.replace(" ","") != "" and default_path.replace(" ","") != "/":
            initial_path = Tools.resource_path(default_path)
            Tools.add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)
        # else:
        #     # initial_path = Tools.resource_path(Tools.path_check(os.getcwd()))
        #     initial_path = initial_path

        export_path.insert("0",str(initial_path))
        format_entry.set(favourite_excel_format)

        root.bind("<Button-1>",lambda e: close_window(child_root))
        child_root.update()
        child_root.update_idletasks()
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{x+200}+{y+100}")
        child_root.focus()
        child_root.focus_force()
        return child_root

    @classmethod
    def excel_manual_window(cls,root,app_icon_path):
        window = customtkinter.CTkToplevel()
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        window.geometry(f"1200x580+{x+100}+{y+200}")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Manual")

        manual_frame =  customtkinter.CTkFrame(master=window,corner_radius=0,height=100,fg_color="#212121")
        manual_frame    .pack(pady=0,padx=0,expand=False,side = "right",anchor="e",ipady = 10,ipadx = 10)
        manual =        customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/excel_manual.png")),size=(1200,520))
        manual_label =  customtkinter.CTkLabel(master = manual_frame,text = "",image =manual,bg_color="#212121")
        manual_label    .pack(pady=0,padx=0,expand=True)
        button_exit =   customtkinter.CTkButton(master = manual_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: window.destroy())
        button_exit     .pack(pady=10,padx=10,expand=True,side = "bottom",anchor = "e")

        root.bind("<Button-1>",lambda e: window.destroy())
        window.grab_set()
        window.focus_force()
        return window

    def __init__(self,root,controller_databases = [[],[]],callback = None,custom_controller_database = [],accessory_databases=[[],[],[]],changes_check = False):
        self.controller_database = controller_databases[0]
        self.controller_notes_database = controller_databases[1]
        self.custom_controller_database = custom_controller_database
        self.accessory_database = accessory_databases[0]
        self.whole_accessory_database = accessory_databases[1]
        self.accessory_notes_database = accessory_databases[2]
        self.changes_check = changes_check
        self.root = root
        self.callback_function = callback
        self.x = self.root.winfo_rootx()
        self.y = self.root.winfo_rooty()
        self.controller_color_list = [
            "",
            "#1E90FF",  # Dodger Blue
            "#32CD32",  # Lime Green
            "#FF4500",  # Orange Red
            "#8A2BE2",  # Blue Violet
            "#00CED1",  # Dark Turquoise
            "#DC143C",  # Crimson
            "#FF6347",  # Tomato
            "#FF69B4",  # Hot Pink
            "#7FFF00",  # Chartreuse
            "#FFD700"  # Gold
        ]
        self.controller_color_pointer = 0
        self.accessory_database_pointer = 0
        self.one_segment_width = 450
        self.app_icon_path = Tools.resource_path('images\\logo_TRIMAZKON.ico')
        self.autosearch_menu = None

    def save_check(self,menu_callback,save_metadata_callback):
        if self.changes_check == False:
            print("no changes made after last save")
            menu_callback()
            return
        
        window = customtkinter.CTkToplevel(fg_color="#212121")
        window.geometry(f"650x130+{self.x+80}+{self.y+150}")
        window.after(200, lambda: window.iconbitmap(self.app_icon_path))
        window.title("Upozornění")

        def clicked_save():
            print("saving")
            window.destroy()
            save_metadata_callback(True)
        
        def clicked_cancel():
            print("cancelling")
            window.destroy()
            menu_callback()

        top_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="#212121")
        warning_icon =      customtkinter.CTkLabel(master = top_frame,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/warning.png")),size=(50,50)),bg_color="#212121")
        proceed_label =     customtkinter.CTkLabel(master = top_frame,text = "Nemáte uložený rozpracovaný projekt!",font=("Arial",20,"bold"),anchor="w",justify="left")
        warning_icon        .pack(pady=10,padx=30,side = "left",anchor="w")
        proceed_label       .pack(pady=10,padx=10,side = "left",anchor="w")
        button_yes =        customtkinter.CTkButton(master = window,text = "Uložit",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: clicked_save())
        button_no =         customtkinter.CTkButton(master = window,text = "Neukládat",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  clicked_cancel())
        top_frame           .pack(pady=0,padx=0,expand=False,side = "top",anchor="w")
        button_no           .pack(pady = 5, padx = (0,10),anchor="w",expand=False,side="right")
        button_yes          .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        window.grab_set()
        window.focus_force()
        window.wait_window()
        return window

    def new_controller_window(self,childroot,controller = None,edit = False,accessory_index =0,only_accessory=False):
        """
        vrací:
        - zvolený kontroler z databáze
        - název (pojmenování) kontoleru
        - barva kontroleru
        - ftp adresa
        - jméno uživatele
        - heslo uživatele
        - accessory list, příslušenství ke kontroleru
        - poznámky ke kontroleru
        """
        def save_contoller():
            db_error_found = False
            save_status = save_changes()
            if save_status == "db_error":
                db_error_found = True
            if controller_entry.get() not in self.controller_database:
                controller_entry.configure(fg_color = "#bd1931",border_color = "red")
                return
            elif db_error_found:
                return
            else:
                controller_entry.configure(fg_color = "#343638",border_color = "#565B5E")
            notes = notes_input.get("1.0", tk.END)
            notes = Tools.make_wrapping(notes)
            try:
                color_chosen = controller_color.cget("fg_color")
            except Exception:
                color_chosen = ""
            print("chosen color: ",color_chosen)
            output = [controller_entry.get(),controller_name_entry.get(),color_chosen,IP_adress_entry.get(),username_entry.get(),password_entry.get(),controller["accessory_list"],notes]
            close_window(window)
            self.callback_function(output)

        def check_used_colors():
            used_colors=[]
            # print("controller database input",self.custom_controller_database)
            for items in self.custom_controller_database:
                if items["color"] != "": # možnost nezvolit žádnou barvu
                    if items["color"] in self.controller_color_list:
                        if controller != None:
                            if not items["color"] == controller["color"]:
                                used_colors.append(items["color"])
                                self.controller_color_list.pop(self.controller_color_list.index(items["color"]))
                        elif not edit:
                            used_colors.append(items["color"])
                            self.controller_color_list.pop(self.controller_color_list.index(items["color"]))
            print("used colors: ",used_colors)
        check_used_colors()

        def switch_color():
            self.controller_color_pointer += 1
            if self.controller_color_pointer > len(self.controller_color_list)-1:
                self.controller_color_pointer = 0

            new_color = self.controller_color_list[self.controller_color_pointer]
            if new_color != "":
                controller_color.configure(fg_color = new_color,hover_color = new_color)
            else:
                controller_color.configure(fg_color= "#212121",hover_color="#212121")

        def close_window(window):
            if not edit:
                try:
                    childroot.unbind("<Button-1>")
                except Exception:
                    pass
            window.destroy()

        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(self.app_icon_path))
        window_height = 870
        window.geometry(f"{self.one_segment_width}x{window_height}+{self.x+150}+{self.y+5}")
        if edit:
            current_name = controller["name"]
            current_type = controller["type"]
            current_ip = controller["ip"]
            current_username = controller["username"]
            current_password = controller["password"]
            current_color = controller["color"]
            current_notes = controller["notes"]

            window.title(f"Editování kontroleru: {current_name} ({current_type})")
        else:
            window.title("Nový kontroler")
            controller = {
                "accessory_list": []
            }
        
        def save_changes():
            """
            only accessory...
            """
            notes = str(notes_input3.get("0.0", tk.END))
            notes = Tools.make_wrapping(notes)
            accessory_item = str(hw_type_entry.get())

            if accessory_item not in self.whole_accessory_database:
                hw_type_entry.configure(fg_color = "#bd1931",border_color = "red")
                return "db_error"
            else:
                hw_type_entry.configure(fg_color = "#343638",border_color = "#565B5E")
            try:
                # if accessory_item != "":
                controller["accessory_list"][accessory_index]["type"] = accessory_item
                controller["accessory_list"][accessory_index]["description"] = notes
                # elif accessory_item == "" and len(controller["accessory_list"])==1:
                #     controller["accessory_list"] = []

            except IndexError:
                if accessory_item != "" and notes != "\n":
                    new_accessory = {
                    "type": accessory_item,
                    "description":notes,
                    }
                    controller["accessory_list"].append(new_accessory)
            except TypeError: # pokud je jako index vložen None
                if accessory_item != "" and notes != "\n":
                    new_accessory = {
                    "type": accessory_item,
                    "description":notes,
                    }
                    controller["accessory_list"].append(new_accessory)
            
            # print("acc_list --------- ",controller["accessory_list"])

        def next_accessory():
            nonlocal accessory_index
            accessory_index += 1
            if accessory_index < len(controller["accessory_list"]):
                accessory_index -= 1
                save_status =save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                accessory_index += 1

            else: # TLACITKO +:
                # program nedopusti pridani noveho accessory pokud neni alespon vyplnen typ nebo poznamka
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    accessory_index -= 1
                    save_status =save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                    if save_status == "db_error":
                        return
                    accessory_index += 1
                else:
                    accessory_index -= 1
                    hw_type_entry.configure(fg_color = "#343638",border_color = "#565B5E")

            initial_prefill() # prefill s novým indexem

        def previous_accessory():
            nonlocal accessory_index
            accessory_index -= 1
            if accessory_index > -1:
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    accessory_index += 1
                    save_status =save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                    if save_status == "db_error":
                        return
                    accessory_index -= 1
                else:
                    hw_type_entry.configure(fg_color = "#343638",border_color = "#565B5E")

                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                accessory_index += 1

        def import_notes(operation = ""):
            notes_string = ""
            if operation == "controller":
                current_controller = controller_entry.get()
                if current_controller != "":
                    controller_notes = str(self.controller_notes_database[self.controller_database.index(current_controller)])
                    if controller_notes != "":
                        notes_string = notes_string + controller_notes + "\n\n"
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",notes_string)
            else:
                current_accessory = hw_type_entry.get()
                if current_accessory != "":
                    notes_string = notes_string + str(self.accessory_notes_database[self.whole_accessory_database.index(current_accessory)])
                notes_input3.delete("1.0",tk.END)
                notes_input3.insert("1.0",notes_string)

        def remaping_characters(event):
            if event.char == 'ì':
                event.widget.insert(tk.INSERT, 'ě')
                return "break"  # Stop the event from inserting the original character
            elif event.char == 'è':
                event.widget.insert(tk.INSERT, 'č')
                return "break"  # Stop the event from inserting the original character
            elif event.char == 'ø':
                event.widget.insert(tk.INSERT, 'ř')
                return "break"  # Stop the event from inserting the original character

        def manage_option_menu(e,values,entry_widget,mirror=None,auto_search_call=False):
            def on_item_selected(value):
                # if auto_search_call:
                entry_widget.delete(0,200)
                entry_widget.insert(0,str(value))
                # else:
                #     entry_widget.set(str(value))
                context_window.destroy()

            if len(values) == 0:
                return

            screen_x = window.winfo_pointerx()
            screen_y = window.winfo_pointery()
            parent_x = window.winfo_rootx()+e.x
            parent_y = window.winfo_rooty()+e.y
            x = screen_x - parent_x +entry_widget.winfo_width()
            y = screen_y - parent_y +entry_widget.winfo_height()

            if auto_search_call:
                screen_x = entry_widget.winfo_rootx()
                screen_y = entry_widget.winfo_rooty() + entry_widget.winfo_height()+5

            font = tkFont.Font(family="Arial", size=20)
            max_width_px = 40
            try:
                max_width_px = max(font.measure(str(val)) for val in values) + 40  # Add some padding
            except Exception as e:
                pass
            context_window = customtkinter.CTkToplevel(window)
            context_window.overrideredirect(True)
            context_window.configure(bg="black")
            listbox = FakeContextMenu(context_window, values, command=on_item_selected, width=max_width_px)
            listbox.pack(fill="both",expand=True)
            window.bind("<Button-1>", lambda e: context_window.destroy(), "+")

            max_visible_items = 50
            visible_items = min(len(values), max_visible_items)
            total_height = visible_items * int(listbox.one_button_height)+20
            window.update_idletasks()
            if total_height > window._current_height-20-y:
                total_height = window._current_height-20-y

            if mirror == True: #priznak aby pri maximalizovani nelezlo mimo obrazovku (doprava)
                screen_x=screen_x-max_width_px
            context_window.geometry(f"{max_width_px}x{total_height}+{screen_x}+{screen_y}")
            if auto_search_call:
                self.autosearch_menu = context_window

        def autosearch_engine(e,which_item):
            """
            which_item:
            - controller
            - accessory
            """
            if self.autosearch_menu != None:
                self.autosearch_menu.destroy()
                self.autosearch_menu = None

            if which_item == "controller":
                entry_widget = controller_entry
                database = self.controller_database
            elif which_item == "accessory":
                entry_widget = hw_type_entry
                database = self.whole_accessory_database

            entry_widget.update_idletasks()
            currently_inserted = str(entry_widget.get()).strip().lower()
            if len(str(currently_inserted))==0:
                return
            found_itemss = []

            for items in database:
                item_str = str(items).lower()
                if currently_inserted in str(item_str):
                # if item_str.startswith(currently_inserted):
                    found_itemss.append(str(items))

            found_itemss = sorted(found_itemss)
            # print(found_itemss)
            manage_option_menu(e,found_itemss,entry_widget,auto_search_call=True)

        icon_small = 45
        icon_large = 49
        # KONTROLER ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        main_frame =                customtkinter.CTkFrame(master = window,corner_radius=0,border_width=3,fg_color="#212121")
        controller_frame =          customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_width=3,fg_color="#212121")
        controller_type =           customtkinter.CTkLabel(master = controller_frame,text = "Typ kontroleru: ",font=("Arial",22,"bold"))
        controller_select_frame =   customtkinter.CTkFrame(master = controller_frame,corner_radius=0,fg_color="#212121")
        controller_entry =          customtkinter.CTkEntry(master = controller_select_frame,font=("Arial",22),corner_radius=0,height=50)
        controller_search =         customtkinter.CTkLabel(master = controller_select_frame,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        controller_search.          bind("<Enter>",lambda e: controller_search._image.configure(size=(icon_large,icon_large)))
        controller_search.          bind("<Leave>",lambda e: controller_search._image.configure(size=(icon_small,icon_small)))
        controller_search.          bind("<Button-1>",lambda e: manage_option_menu(e,self.controller_database,controller_entry))
        controller_name =           customtkinter.CTkLabel(master = controller_frame,text = "Název (interní označení): ",font=("Arial",22,"bold"))
        controller_name_entry =     customtkinter.CTkEntry(master = controller_frame,font=("Arial",22),corner_radius=0,height=50)
        controller_color =          customtkinter.CTkButton(master = controller_frame,corner_radius=0,text="Podbarvení kontroleru",font=("Arial",22,"bold"),height=50,command=lambda:switch_color(),border_width=3)
        IP_adress =                 customtkinter.CTkLabel(master = controller_frame,text = "IP adresa: ",font=("Arial",22,"bold"))
        IP_adress_entry =           customtkinter.CTkEntry(master = controller_frame,font=("Arial",22),corner_radius=0,height=50)
        username =                  customtkinter.CTkLabel(master = controller_frame,text = "Jméno: ",font=("Arial",22,"bold"))
        username_entry =            customtkinter.CTkEntry(master = controller_frame,font=("Arial",22),corner_radius=0,height=50)
        password =                  customtkinter.CTkLabel(master = controller_frame,text = "Heslo: ",font=("Arial",22,"bold"))
        password_entry =            customtkinter.CTkEntry(master = controller_frame,font=("Arial",22),corner_radius=0,height=50,placeholder_text="*******")
        note_label_frame =          customtkinter.CTkFrame(master = controller_frame,corner_radius=0,fg_color="#212121")
        note_label =                customtkinter.CTkLabel(master = note_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes_btn =          customtkinter.CTkButton(master = note_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("controller"))
        note_label.                 pack(pady = 5, padx = (10,0),anchor="w",side="left")
        import_notes_btn.           pack(pady = 5, padx = (10,0),anchor="w",side="left")
        notes_input =               customtkinter.CTkTextbox(master = controller_frame,font=("Arial",22),corner_radius=0)
        controller_type.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        controller_entry.           pack(pady=(10,0),padx=(5,5),side = "left",anchor = "w",fill ="x",expand = True)
        controller_search.          pack(pady=(10,0),padx=(0,5),side = "left",anchor = "w")
        controller_select_frame.    pack(pady = 5, padx = 10,anchor="w",side="top",fill="x")
        controller_name.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        controller_name_entry.      pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        controller_color.           pack(pady=0,padx=10,side =      "top",anchor = "w",fill="x")
        IP_adress.                  pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        IP_adress_entry.            pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        username.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        username_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        password.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        password_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        note_label_frame.           pack(pady = 5, padx = 10,side="top",fill="both")
        notes_input.                pack(pady = 5, padx = 10,side="top",fill="both",expand = True)
        notes_input.                bind("<Key>",remaping_characters)
        controller_name_entry.      bind("<Key>",remaping_characters)
        username_entry.             bind("<Key>",remaping_characters)
        password_entry.             bind("<Key>",remaping_characters)
        controller_entry.           bind("<KeyRelease>",lambda e: autosearch_engine(e,"controller"))
        selected_color = self.controller_color_list[self.controller_color_pointer]
        if selected_color != "":
            controller_color.configure(fg_color=selected_color,hover_color=selected_color)
        else:
            controller_color.configure(fg_color= "#212121",hover_color="#212121")

        # PŘÍSLUŠENSTVÍ ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        accessory_frame =           customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_width=3,fg_color="#212121")
        counter_frame_acc =         customtkinter.CTkFrame(master = accessory_frame,corner_radius=0,fg_color="#212121")
        button_prev_acc =           customtkinter.CTkButton(master = counter_frame_acc,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_accessory())
        counter_acc =               customtkinter.CTkLabel(master = counter_frame_acc,text = "0/0",font=("Arial",22,"bold"))
        button_next_acc =           customtkinter.CTkButton(master = counter_frame_acc,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_accessory())
        button_prev_acc.            pack(pady = 0, padx = (5,0),anchor="w",side="left")
        counter_acc.                pack(pady = 0, padx = (5,0),anchor="w",side="left")
        button_next_acc.            pack(pady = 0, padx = (5,0),anchor="w",side="left")
        accessory_label =           customtkinter.CTkLabel(master = accessory_frame,text = "Příslušenství:",font=("Arial",22,"bold"))
        option_menu_frame_acc =     customtkinter.CTkFrame(master = accessory_frame,corner_radius=0,fg_color="#212121")
        hw_type_entry =             customtkinter.CTkEntry(master = option_menu_frame_acc,font=("Arial",22),corner_radius=0,height=50)
        acc_search =                customtkinter.CTkLabel(master = option_menu_frame_acc,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        acc_search.                 bind("<Enter>",lambda e: acc_search._image.configure(size=(icon_large,icon_large)))
        acc_search.                 bind("<Leave>",lambda e: acc_search._image.configure(size=(icon_small,icon_small)))
        acc_search.                 bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_accessory_database,hw_type_entry))
        hw_type_entry.              pack(pady = 5, padx = (5,0),anchor="w",side="left",fill="x",expand = True)
        acc_search.                 pack(pady = 5, padx = (5,0),anchor="w",side="left")
        note3_label_frame =         customtkinter.CTkFrame(master = accessory_frame,corner_radius=0,fg_color="#212121")
        note3_label =               customtkinter.CTkLabel(master = note3_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes3_btn =         customtkinter.CTkButton(master = note3_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes())
        note3_label.                pack(pady = 5, padx = (10,0),anchor="w",side="left")
        import_notes3_btn.          pack(pady = 5, padx = (10,0),anchor="w",side="left")
        notes_input3 =              customtkinter.CTkTextbox(master = accessory_frame,font=("Arial",22),corner_radius=0)
        counter_frame_acc.          pack(pady=(10,0),padx=3,anchor="n",side = "top")
        accessory_label.            pack(pady=(15,5),padx=10,anchor="w",side = "top")
        option_menu_frame_acc.      pack(pady = 5, padx = 10,anchor="w",side="top",fill="x")
        note3_label_frame.          pack(pady = 0, padx = 3,anchor="w",side="top",fill="x")
        notes_input3.               pack(pady = 5, padx = 10,side="top",fill="both",expand = True)
        if not only_accessory:
            window.geometry(f"{2*self.one_segment_width}x{window_height}+{self.x+150}+{self.y+5}")
            controller_frame.       pack(pady = 0, padx = 0,fill="both",expand = True,anchor="n",side="left",ipady = 3,ipadx = 3)
        accessory_frame.            pack(pady = 0, padx = 0,fill="both",expand = True,anchor="n",side="left",ipady = 3,ipadx = 3)
        main_frame.                 pack(pady = 0, padx = 0,fill="both",expand = True,anchor="n",side="top")
        bottom_frame =              customtkinter.CTkFrame(master=window,corner_radius=0,fg_color="#212121")
        button_save =               customtkinter.CTkButton(master = bottom_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_contoller())
        button_exit =               customtkinter.CTkButton(master = bottom_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_exit.                pack(pady=10,padx=(5,10),side = "right",anchor="e")
        button_save.                pack(pady=10,padx=5,side = "right",anchor="e")
        bottom_frame.               pack(pady = 0, padx = 0,fill="x",anchor="s",side="bottom")
        notes_input3.               bind("<Key>",remaping_characters)
        hw_type_entry.              bind("<KeyRelease>",lambda e: autosearch_engine(e,"accessory"))

        if edit:
            IP_adress_entry.insert(0,str(current_ip))
            controller_entry.delete(0,300)
            controller_entry.insert(0,str(current_type))
            controller_name_entry.insert(0,str(current_name))
            username_entry.insert(0,str(current_username))
            password_entry.insert(0,str(current_password))
            if str(current_color) != "":
                controller_color.configure(fg_color = str(current_color))
            else:
                controller_color.configure(fg_color= "#212121",hover_color="#212121")

            notes_input.delete("1.0",tk.END)
            notes_input.insert("1.0",str(current_notes))
        else:
            IP_adress_entry.insert(0,"192.168.000.000")
            controller_name_entry.insert(0,"Kontroler " + str(len(self.custom_controller_database)+1) + " ")
            childroot.bind("<Button-1>",lambda e: close_window(window))

        def refresh_counters():
            nonlocal accessory_index
            nonlocal counter_acc
            try:
                counter_acc_state = str(accessory_index+1) + "/" + str(len(controller["accessory_list"]))
                counter_acc.configure(text = counter_acc_state)
            except Exception:
                pass

        def refresh_button_appearance():
            nonlocal accessory_index
            nonlocal button_prev_acc
            nonlocal button_next_acc

            def config_buttons(button_left,button_right,index,max_array_value):
                if index ==0:
                    button_left.configure(text = "",fg_color = "#636363")
                else:
                    button_left.configure(text = "<",fg_color = "#636363")

                if index == max_array_value:
                    button_right.configure(text = "+",fg_color = "green")
                else:
                    button_right.configure(text = ">",fg_color = "#636363")
            # pokud není accessory:
            try:
                config_buttons(button_prev_acc,button_next_acc,accessory_index,len(controller["accessory_list"])-1)
            except Exception as e:
                print(f"chyba při nastavování vzhledu tlačítek - accessory: {e}")

        def initial_prefill():
            nonlocal hw_type_entry
            nonlocal notes_input3
            nonlocal accessory_index

            try:
                if str(controller["accessory_list"][accessory_index]["type"]) in self.whole_accessory_database:
                    hw_type_entry.delete(0,300)
                    hw_type_entry.insert(0,str(controller["accessory_list"][accessory_index]["type"]))
                else:
                    hw_type_entry.delete(0,300)
                notes_input3.delete("1.0",tk.END)
                notes_input3.insert("1.0",str(controller["accessory_list"][accessory_index]["description"]))
            except TypeError: # pokud je v indexu None - defaultně nastavit index na nulu:
                try:
                    accessory_index = 0
                    if str(controller["accessory_list"][accessory_index]["type"]) in self.whole_accessory_database:
                        hw_type_entry.delete(0,300)
                        hw_type_entry.insert(0,str(controller["accessory_list"][accessory_index]["type"]))
                    else:
                        hw_type_entry.delete(0,300)
                    notes_input3.delete("1.0",tk.END)
                    notes_input3.insert("1.0",str(controller["accessory_list"][accessory_index]["description"]))
                except IndexError: #případ, že není accessory
                    hw_type_entry.delete(0,300)
                    notes_input3.delete("1.0",tk.END)
            except IndexError: #případ, že není accessory
                hw_type_entry.delete(0,300)
                notes_input3.delete("1.0",tk.END)

            refresh_counters()
            refresh_button_appearance()

        initial_prefill()
        self.root.bind("<Button-1>",lambda e: close_window(window))
        global opened_subwindow
        opened_subwindow = window
        window.update()
        window.update_idletasks()
        window.focus_force()
        window.focus()
        return window

class Insert_image:
    def __init__(self,root,childroot,image_paths,callback,window_scale,remembered_path=None):
        self.root = root
        self.childroot = childroot
        self.image_paths = image_paths
        self.callback_function = callback
        self.image_name = ""
        self.image_path_inserted = ""
        self.current_image_index = 0
        self.remembered_path = remembered_path
        self.window_scale = window_scale
        self.app_icon_path = Tools.resource_path('images\\logo_TRIMAZKON.ico')

    def calc_current_format(self,image_width,image_height,frame_width,frame_height): # Přepočítávání rozměrů obrázku do rozměru rámce podle jeho formátu + zooming
        """
        Přepočítávání rozměrů obrázku do rozměru rámce podle jeho formátu

        -vstupními daty jsou šířka a výška obrázku
        -přepočítávání pozicování obrázku a scrollbarů v závislosti na zoomu
        """
        image_width = image_width
        image_height = image_height
        image_ratio = image_width / image_height

        def rescale_image(): # Vmestnani obrazku do velikosti aktualni velikosti ramce podle jeho formatu
            if image_height > image_width:
                new_height = frame_height
                if image_width > frame_width:
                    new_width = int(new_height * image_ratio)
                else:
                    new_width = frame_width
                    if image_width > frame_width:
                        new_width = image_width
                    new_height = int(new_width / image_ratio)

            elif image_height < image_width:
                new_width = frame_width

                if image_height < frame_height:
                    new_height = int(new_width / image_ratio)
                else:
                    new_height = frame_height
                    if image_height < frame_height:
                        new_height = image_height
                    new_width = int(new_height * image_ratio)

            elif image_height == image_width:
                new_height = frame_height
                new_width = new_height

            #doublecheck
            if new_height > frame_height:
                new_height = frame_height
                new_width = int(new_height * image_ratio)
            if new_width > frame_width:
                new_width = frame_width
                new_height = int(new_width / image_ratio)

            return (new_height,new_width)
        
        new_height, new_width = rescale_image()
        return (int(new_width), int(new_height))

    def load_image_paths(self,refresh = False):
        if len(self.image_paths) == 0:
            Tools.add_colored_line(self.console,"Není přiřazena fotografie","orange",None,True)
            return
        current_path = str(self.image_paths[self.current_image_index])
        if current_path.endswith("/"):
            current_path = current_path[:-1]
        name_split = current_path.split("/")
        self.image_name = str(name_split[-1])
        self.image_path_inserted = current_path.replace(self.image_name,"")

        if refresh:
            if self.name_or_path.get() == 1:
                Tools.add_colored_line(self.console,self.image_name,"white",None,True)
            else:
                Tools.add_colored_line(self.console,self.image_path_inserted + self.image_name,"white",None,True)
            return
        
        self.image_frame.bind("<Button-3>", self.show_context_menu)

        if not os.path.isfile(str(self.image_path_inserted)+str(self.image_name)):
            self.image_frame.delete("lower")
            Tools.add_colored_line(self.console,f"Cesta k souboru neexistuje: {current_path}","red",None,True)
            return
        try:
            with PILImage.open(current_path) as opened_image:
                width,height = opened_image.size
                self.image_frame.update()
                self.image_frame.update_idletasks()
                dimensions = self.calc_current_format(width,
                                                    height,
                                                    self.image_frame.winfo_width(),
                                                    self.image_frame.winfo_height())

                resized = opened_image.resize(size=dimensions)
                self.tk_image = ImageTk.PhotoImage(resized)
                self.image_frame.delete("lower")
                main_image = self.image_frame.create_image(0, 0,anchor=tk.NW, image=self.tk_image,tag = "lower")
                self.image_frame.tag_lower(main_image)
                if self.name_or_path.get() == 1:
                    Tools.add_colored_line(self.console,self.image_name,"white",None,True)
                else:
                    Tools.add_colored_line(self.console,self.image_path_inserted + self.image_name,"white",None,True)

        except Exception as e:
            error_message = f"Obrázek: {self.image_name} je poškozen. {e}"
            Tools.add_colored_line(self.console,error_message,"red",None,True)
            # print(error_message)
            self.image_name = ""
            self.image_path_inserted = ""
            self.image_frame.delete("lower")
            return error_message
        
    def show_context_menu(self,event):
        self.context_menu.tk_popup(event.x_root, event.y_root)

    def image_menu_gui(self):
        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(self.app_icon_path))
        window.title("Možnosti vložení fotografie ke stanici")
        subwindow = ""

        def close_window(window):
            nonlocal subwindow
            try:
                if subwindow.winfo_exists():
                    subwindow.destroy()
            except Exception:
                pass
            # window.grab_release()
            window.destroy()

        def call_browse_directories(context_menu = False):
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if context_menu:
                name_split = self.image_paths[self.current_image_index].split("/")
                self.remembered_path = str(self.image_paths[self.current_image_index].replace(str(name_split[-1]),""))
                if self.image_paths[self.current_image_index].endswith("/"):
                    self.remembered_path = str(self.image_paths[self.current_image_index].replace(str(name_split[-2]),""))

            filetypes = [
                ("Image files", "*.png;*.jpg;*.bmp"),
                ("All files", "*.*")
            ]
            output = Tools.browseDirectories("all",start_path=self.remembered_path,file_type=filetypes)
            if str(output[1]) != "/":
                image_path.delete(0,300)
                image_path.insert(0, str(output[1])+str(output[2]))
                self.remembered_path = str(output[1])
                Tools.add_colored_line(self.console,"Byla vložena cesta a název souboru","green",None,True)
                add_image_path()
                
            if self.childroot != None:
                self.childroot.focus_force()
                self.childroot.focus()
            window.focus_force()
            window.focus()

        def add_image_path():
            checked_path = Tools.path_check(image_path.get(),only_repair=True)
            if checked_path == False or checked_path.replace(" ","") == "" or checked_path.replace(" ","") == "/":
                Tools.add_colored_line(self.console,"Cesta k souboru je neplatná","red",None,True)
                return
            
            if checked_path not in self.image_paths:
                self.image_paths.append(checked_path)
                self.callback_function(self.image_paths,self.remembered_path)
                next_image(force_index=len(self.image_paths)-1)
            else:
                Tools.add_colored_line(self.console,"Soubor už je přidán","orange",None,True)

        def next_image(force_index = False):
            if not force_index:
                self.current_image_index +=1
                if self.current_image_index == len(self.image_paths):
                    self.current_image_index = 0
            else:
                self.current_image_index = force_index
            self.image_number.configure(text = str(self.current_image_index+1)+"/"+str(len(self.image_paths)))
            self.load_image_paths()
        
        def previous_image():
            self.current_image_index -=1
            if self.current_image_index < 0:
                self.current_image_index = len(self.image_paths)-1
            self.image_number.configure(text = str(self.current_image_index+1)+"/"+str(len(self.image_paths)))
            self.load_image_paths()

        def remove_file(file):
            self.image_paths.pop(self.image_paths.index(file))
            self.callback_function(self.image_paths,self.remembered_path)
            previous_image()

        load_photo_frame =  customtkinter.CTkFrame(master = window,corner_radius=0)
        image_path_label =  customtkinter.CTkLabel(master = load_photo_frame,text = "Zadejte cestu k fotografii:",font=("Arial",22,"bold"))
        image_path_frame =  customtkinter.CTkFrame(master = load_photo_frame,corner_radius=0)
        image_path =        customtkinter.CTkEntry(master = image_path_frame,font=("Arial",20),width=580,height=50,corner_radius=0)
        explorer_btn =      customtkinter.CTkButton(master = image_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories())
        save_path_btn =     customtkinter.CTkButton(master = image_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: add_image_path())
        image_path          .pack(pady = 5, padx = (10,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn       .pack(pady = 5, padx = 5,anchor="e",expand=False,side="right")
        explorer_btn        .pack(pady = 5, padx = (5,0),anchor="e",expand=False,side="right")
        self.console =      tk.Text(load_photo_frame, wrap="none", height=0,background="#212121",font=("Arial",22),state=tk.DISABLED,foreground="#565B5E",borderwidth=3)
        controls_frame =    customtkinter.CTkFrame(master = load_photo_frame,corner_radius=0,height=50)
        self.name_or_path = customtkinter.CTkCheckBox(master = controls_frame, text = "Název/ cesta",font=("Arial",22,"bold"),command=lambda: self.load_image_paths(refresh=True))
        button_left =       customtkinter.CTkButton(master = controls_frame,text = "<",font=("Arial",30,"bold"),width = 150,height=50,corner_radius=0,command=lambda: previous_image())
        self.image_number = customtkinter.CTkLabel(master = controls_frame,text = f"1/{str(len(self.image_paths))}",font=("Arial",22,"bold"))
        button_right =      customtkinter.CTkButton(master = controls_frame,text = ">",font=("Arial",30,"bold"),width = 150,height=50,corner_radius=0,command=lambda: next_image())
        self.name_or_path   .pack(pady = 10, padx = 10,anchor="w",side="left")
        button_left         .pack(pady = 0, padx = 10,anchor="w",side="left")
        self.image_number   .pack(pady = 0, padx = 10,anchor="w",side="left")
        button_right        .pack(pady = 0, padx = 10,anchor="w",side="left")
        self.image_frame =  tk.Canvas(master=load_photo_frame,bg="#212121",highlightthickness=0)
        buttons_frame =     customtkinter.CTkFrame(master = load_photo_frame,corner_radius=0)
        button_exit =       customtkinter.CTkButton(master = buttons_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_exit         .pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        load_photo_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        image_path_label    .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        image_path_frame    .pack(expand=False,side="top",anchor="n",fill="x")
        controls_frame      .pack(expand=False,side="top",anchor="n",fill="x")
        self.console        .pack(pady = (5,5), padx =10,anchor="w",expand=False,fill="x",side="top",ipady=3,ipadx=5)
        self.image_frame    .pack(pady = 5, padx = 5,expand=True,side="top",fill="both",anchor="n")
        buttons_frame       .pack(pady = 0, padx = 0,expand=False,side="top",fill="x")

        self.context_menu = tk.Menu(window,tearoff=0,fg="white",bg="#202020",activebackground="#606060")
        self.context_menu.add_command(label="Otevřít cestu", command=lambda: call_browse_directories(context_menu = True),font=("Arial",22,"bold"))
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Odstranit cestu", command=lambda: remove_file(self.image_paths[self.current_image_index]),font=("Arial",22,"bold"))
        self.root.bind("<Button-1>",lambda e: close_window(window),"+")
        if self.childroot != None:
            self.childroot.bind("<Button-1>",lambda e: close_window(window),"+")
        window.update()
        window.update_idletasks()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        window.geometry(f"1200x900+{x+200}+{y+50}")
        if self.window_scale == 1:
            window.state('zoomed')
        self.load_image_paths()
        window.focus_force()
        window.focus()
        window.grab_set()
        window.grab_release()

        window.bind("<Left>",lambda e: previous_image())
        window.bind("<Right>",lambda e: next_image())
        def mousewheel_handle(e):
            if e.delta < 0:
                previous_image()
            else:
                next_image()
        window.bind("<MouseWheel>",mousewheel_handle)

class Fill_details:
    @classmethod
    def station(cls,station):
        detail_info = ""
        detail_info = str(station["inspection_description"])

        return detail_info
    
    @classmethod
    def controller(cls,controller):
        detail_info = str(controller["detailed_name"])
        if not str(controller["notes"]) == "":
            detail_info = detail_info +"\n\n"+ str(controller["notes"])
        if not str(controller["ip"]) == "" and not controller["ip"] == "192.168.000.000":
            detail_info = detail_info + "\n\nIP: " + str(controller["ip"])
        if not str(controller["username"]) == "":
            detail_info = detail_info + "\nJméno: " + str(controller["username"])
        if not str(controller["password"]) == "":
            detail_info = detail_info + "\nHeslo: " + str(controller["password"])
        return detail_info
    
    @classmethod
    def camera(cls,camera):
        """
        Returns:
        - detail info string [0]
        - controller background color [1]
        """
        detail_info_cam = ""
        controller_fill = None

        if str(camera["controller_color"]) != "":
            try:
                color_modified = str(camera["controller_color"])[1:]
                controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                # ws[excel_cell].fill = controller_fill
            except Exception as e:
                print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                pass

        cable = str(camera["cable"])
        if cable != "" and not cable in str(camera["description"]):
            detail_info_cam = detail_info_cam + "Kabel: " + str(camera["cable"])+ "\n\n"
        detail_info_cam += str(camera["description"])

        return [detail_info_cam,controller_fill]
    
    @classmethod
    def optics(cls,optics):
        detail_info = ""
        if str(optics["alternative"]) != "":
            detail_info = "Alternativa: " + str(optics["alternative"]) + "\n\n"
            
        detail_info += str(optics["description"])
        return detail_info
    
    @classmethod
    def accessory(cls,accessory):
        detail_info = ""
        detail_info = str(accessory["description"])

        return detail_info

class Catalogue_gui:
    class ToolTip:
        def __init__(self, widget, text, root,unbind=False,subwindow_status=False,reverse=False):
            self.widget = widget
            self.text = text
            self.root = root
            self.tip_window = None
            self.subwindow_status = subwindow_status
            self.reverse = reverse
            if unbind:
                self.unbind_all("",self.widget)
            else:
                self.bind_it()

        def bind_it(self):
            self.widget.bind("<Enter>",lambda e,widget = self.widget: self.really_entering(e,widget))
            self.widget.bind("<Leave>",lambda e,widget = self.widget: self.really_leaving(e,widget))
            self.widget.bind("<Button-1>",lambda e,widget = self.widget:self.just_destroy(e,widget))

        def unbind_all(self,e,widget):
            try:
                self.tip_window.update_idletasks()
                # print("destroying")
                self.tip_window.destroy()
                self.root.after(0,self.tip_window.destroy)
            except Exception as ee:
                pass
            widget.unbind("<Enter>")
            widget.unbind("<Leave>")
            widget.unbind("<Button-1>")

        def just_destroy(self,e,widget,unbind=True):
            # if self.tip_window:
            try:
                self.tip_window.update_idletasks()
            except Exception:
                pass
            try:
                self.tip_window.destroy()
                # self.root.after(0,self.tip_window.destroy)
            except Exception as ee:
                # print(ee)
                pass
            self.tip_window = None
            
        def really_entering(self,e,widget):
            if self.tip_window != None:
                return

            def show_tooltip_v2(e):
                screen_x = self.root.winfo_pointerx()
                screen_y = self.root.winfo_pointery()
                parent_x = self.root.winfo_rootx()+e.x
                parent_y = self.root.winfo_rooty()+e.y
                local_x = screen_x - parent_x +self.widget.winfo_width()
                local_y = screen_y - parent_y +self.widget.winfo_height()
                self.tip_window = customtkinter.CTkLabel(
                    self.root,
                    text=self.text,
                    font=("Arial", 20),
                    text_color="black",
                    bg_color= "white"
                )
                self.tip_window.place(x=-200,y=-200)
                self.tip_window.update_idletasks()
                if self.subwindow_status:
                    if self.reverse:
                        tip_window_width = self.tip_window._current_width
                        self.tip_window.place_configure(x=local_x-tip_window_width,y = local_y)
                    else:
                        self.tip_window.place_configure(x=local_x,y = local_y)
                else:
                    if self.reverse:
                        tip_window_width = self.tip_window._current_width
                        self.tip_window.place_configure(x=local_x-tip_window_width,y = local_y+10)
                    else:
                        self.tip_window.place_configure(x=local_x,y = local_y+10)
                # self.tip_window.place(x=local_x+tip_window_width/2,y = local_y)

            show_tooltip_v2(e)
            self.tip_window.bind("<Leave>",lambda e,widget = self.widget:self.really_leaving(e,widget))
        
        def really_leaving(self,e,widget):
            if self.tip_window == None:
                return

            x = widget.winfo_width()-1
            y = widget.winfo_height()-1
            if (e.x < 1 or e.x > x) or (e.y < 1 or e.y > y):
                try:
                    self.root.after(0,self.tip_window.destroy)
                    # self.tip_window.destroy()
                except Exception as e2:
                    print("error2")
                self.tip_window = None
    
    @classmethod
    def get_device_strings(cls,widget_tier):
        device_string_mapping = {
        2: ["Nová kamera","Editovat stanici","Odebrat stanici","Kopírovat stanici"],
        4: ["Nová optika/ světlo","Editovat kameru","Odebrat kameru","Kopírovat kameru"],
        6: ["","Editovat optiku/světlo","Odebrat optiku/světlo","Kopírovat optiku/světlo"],
        7: ["Nové příslušenství","Editovat kontroler","Odebrat kontroler",""],
        9: ["","Editovat příslušenství","Odebrat příslušenství",""],
        }
        return device_string_mapping.get(len(widget_tier))

    def __init__(self,root,
                 download_status,
                 callback_function,
                 window_size,
                #  database_filename,
                #  default_excel_name,
                #  default_xml_name,
                #  default_subwindow_status,
                #  default_file_extension,
                #  default_path,
                #  default_render_mode,
                initial_path_given):
        
        self.root = root
        global initial_path
        initial_path = initial_path_given
        self.download_status = download_status
        self.callback = callback_function
        self.path_for_callback = None
        self.app_icon_path = Tools.resource_path('images\\logo_TRIMAZKON.ico')
        if window_size == "max":
            self.root.state('zoomed')
            self.root.update()
        else:
            self.root.state('normal')
            self.root.geometry("1600x900")
            self.root.update()
        self.station_list = []
        self.temp_station_list = []
        self.default_block_width = 500
        self.format_list = ["xlsm","xlsx"]
        self.favourite_format = "xlsm"

        config_data = Tools.read_json_config()
        self.default_database_filename = config_data["catalogue_settings"]["database_filename"]
        self.default_excel_filename = config_data["catalogue_settings"]["catalogue_filename"]
        self.default_xml_file_name = config_data["catalogue_settings"]["metadata_filename"]
        self.default_subwindow_status = config_data["catalogue_settings"]["subwindow_behav"]
        self.default_export_extension = config_data["catalogue_settings"]["default_export_suffix"]
        self.default_path = config_data["catalogue_settings"]["default_path"]
        self.render_mode = config_data["catalogue_settings"]["render_mode"]
        self.show_tooltip = config_data["app_settings"]["tooltip_status"]
        try:
            self.hover_trigger_mode = config_data["catalogue_settings"]["hover_info_trigger_mode"]
        except KeyError:
            self.hover_trigger_mode = "1"
            Tools.save_to_json_config("1","catalogue_settings","hover_info_trigger_mode")

        try:
            if self.default_export_extension in self.format_list:
                self.favourite_format = self.default_export_extension
            elif self.default_export_extension.replace(".","") in self.format_list:
                self.favourite_format = self.default_export_extension.replace(".","")
        except Exception:
            pass
        self.current_block_id = ""
        self.controller_object_list = []
        self.custom_controller_drop_list = [""]
        self.chosen_manufacturer = "Omron"
        self.last_selected_widget = ["",""]
        self.controller_database = []
        self.controller_notes_database = []
        self.camera_type_database = [["no data"]]
        self.camera_notes_database = []
        self.whole_camera_type_database = []
        self.camera_cable_database = [["no data"]]
        self.cable_notes_database = []
        self.whole_camera_cable_database = []
        self.optics_database = [["no data"]]
        self.whole_optics_database =[]
        self.optics_notes_database = []
        self.light_database = [["no data"]]
        self.whole_light_database =[]
        self.light_notes_database = []
        self.accessory_database = [["no data"]]
        self.accessory_notes_database = []
        self.whole_accessory_database = []
        self.last_xml_filename = ""
        self.last_path_input = ""
        self.last_controller_index = 0
        self.opened_window = ""
        self.copy_memory = ""
        self.copy_widget_tier = None
        self.leave_expanded_widget = None
        self.leave_expanded_widget_tier = None
        self.current_db_connection = None
        self.autosearch_menu = None

        self.changes_made = False
        self.optic_light_option = "optic"
        self.detailed_view = False
        self.last_scroll_position = 0.0
        self.widget_list = [] #lists of every widget by station
        self.last_path_to_images = None
        self.read_database()
        self.create_main_widgets(initial=True)

    def close_window(self,window):
        window.update_idletasks()
        window.destroy()
    
    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused:
            return True
        else:
            return False

    def read_database(self):
        """
        Stahuje aktuální databázi do adresáře
        - 1. controller_database, controller_notes_database
        - 2. camera_database
        - 3. optics_database
        - 4. accessory_database
        """
        self.download_database_console_input = []
        if self.chosen_manufacturer == "Omron":
            column_index = 1
        else:
            column_index = 3

        if "Chyba" in self.download_status or "chyba" in self.download_status or "epodařilo" in self.download_status:
            text_color = "red"
        else:
            text_color = "green"
        self.download_database_console_input.append(self.download_status)
        self.download_database_console_input.append(text_color)

        # sharepoint_database_path = Tools.resource_path(Tools.path_check(os.getcwd()) + self.default_database_filename)
        sharepoint_database_path = initial_path + self.default_database_filename
        print(sharepoint_database_path)

        self.camera_database_pointer = 0
        self.optics_database_pointer = 0
        self.camera_cable_database_pointer = 0
        self.accessory_database_pointer = 0
        # if len(self.download_database_console_input) > 0:
        # if text_color == "red":
        #     return

        load_failed = False
        try:
            wb = load_workbook(filename=sharepoint_database_path)
        except Exception as err:
            load_failed = True
            self.download_database_console_input = []
            self.download_database_console_input.append(f"Chyba - selhalo načtení databáze produktů ({sharepoint_database_path})")
            self.download_database_console_input.append("red")

        if not load_failed:
            def fill_lists(wb,name_of_excel_sheet:str,empty_option = True):
                """
                - Vrací seznam produktů přečtecńých z databáze
                    - 0 = celá, kompletní databáze produktů
                    - 1 = databáze rozdělená podle *** ([[]])
                    - 2 = druhý parametr - nyní poznámky - kompletní databáze
                """
                nonlocal column_index
                database_section = [""]
                section_database = []
                whole_database = [""]
                notes_database = [""]
                if not empty_option:
                    whole_database = []
                    notes_database = []
                ws = wb[name_of_excel_sheet]
                # první parametr - typ produktu
                row_count = 0
                for row in ws.iter_rows(min_row=2,min_col=column_index, max_col=column_index,values_only=True):
                    if (row[0] is not None or str(row[0]) != "None") and "***" not in str(row[0]): 
                        database_section.append(str(row[0]))
                        whole_database.append(str(row[0]))
                    elif "***" in str(row[0]):
                        section_database.append(database_section)
                        database_section = []
                    row_count +=1
                if database_section != []:
                    section_database.append(database_section)

                # druhý parametr, dolplňující poznámky
                for row in ws.iter_rows(min_row=2,max_row=row_count+1,min_col=column_index+1, max_col=column_index+1,values_only=True):
                    if row[0] is not None and str(row[0]) != "None": 
                        notes_database.append(str(row[0]))
                    else:
                        notes_database.append("")
                    
                
                return [whole_database,section_database,notes_database]
            
            # KONTROLERY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kontrolery",empty_option = False)
            self.controller_database = read_database[0]
            self.controller_notes_database = read_database[2]
            # KAMERY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kamery",empty_option = True)
            self.whole_camera_type_database = read_database[0]
            self.camera_type_database = read_database[1]
            self.camera_notes_database = read_database[2]
            # KABELY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kabely",empty_option = True)
            self.whole_camera_cable_database = read_database[0]
            self.camera_cable_database = read_database[1]
            self.cable_notes_database = read_database[2]
            # OPTIKA ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Optika",empty_option = True)
            self.whole_optics_database = read_database[0]
            self.optics_database = read_database[1]
            self.optics_notes_database = read_database[2]
            # SVĚTLA ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            try:
                read_database = fill_lists(wb,"Světla",empty_option = True)
                self.whole_light_database = read_database[0]
                self.light_database = read_database[1]
                self.light_notes_database = read_database[2]
            except Exception:
                print("chybí list se světly")

            # PŘÍSLUŠENSTVÍ ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Přislušenství",empty_option = True)
            self.whole_accessory_database = read_database[0]
            self.accessory_database = read_database[1]
            self.accessory_notes_database = read_database[2]
            wb.close()
        
    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
        """
        self.clear_frame(self.root)
        if not  self.default_database_filename.endswith(".xlsx"):
            default_database_name_w_extension = self.default_database_filename + ".xlsx"
        else:
            default_database_name_w_extension = self.default_database_filename
        self.callback()
        # self.callback([default_database_name_w_extension,self.default_excel_filename,self.default_xml_file_name,self.default_subwindow_status,self.favourite_format,self.path_for_callback,self.render_mode])

    def switch_widget_info(self,args,widget_tier,widget):
        if len(widget_tier) == 2: #01-99 stanice
            station_index = int(widget_tier[:2])
            station_name = str(self.station_list[station_index]["name"])
            if widget._text != station_name:
                widget.configure(text=station_name,font = ("Arial",25,"bold"))
            else:
                description = Fill_details.station(self.station_list[station_index])
                widget.configure(text=description,font = ("Arial",25))
        
        elif len(widget_tier) == 7: # xxxxc01-xxxxc99 kontolery
            controller_index = int(widget_tier[5:7])
            if widget._text == str(self.controller_object_list[controller_index]["type"]):
                details = Fill_details.controller(self.controller_object_list[controller_index])
                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=str(self.controller_object_list[controller_index]["type"]),font = ("Arial",25,"bold"))

        elif len(widget_tier) == 4: # 0101-9999 kamery
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["type"]):
                details = Fill_details.camera(self.station_list[station_index]["camera_list"][camera_index])[0]
                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["type"]),font = ("Arial",25,"bold"))

        elif len(widget_tier) == 9: # xxxxc0101-xxxxc9999 prislusenstvi kontroleru
            controller_index = int(widget_tier[5:7])
            accessory_index = int(widget_tier[7:9])
            if widget._text == str(self.controller_object_list[controller_index]["accessory_list"][accessory_index]["type"]):
                description = Fill_details.accessory(self.controller_object_list[controller_index]["accessory_list"][accessory_index])
                widget.configure(text=description,font = ("Arial",25))
            else:
                widget.configure(text=str(self.controller_object_list[controller_index]["accessory_list"][accessory_index]["type"]),font = ("Arial",25,"bold"))
            
        elif len(widget_tier) == 6: # 010101-999999 optika
            details = ""
            addition = ""
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            optic_type = str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"])
            if optic_type in self.whole_light_database and optic_type != "":
                addition = "💡 "
            if widget._text == addition + optic_type:
                details = Fill_details.optics(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=addition+optic_type,font = ("Arial",25,"bold"))

    def select_block(self,args,widget_tier,widget):
        """
        - vyvoláno levým klikem
        - Vkládá widget tier do vyhledávače
        - mění názvy tlačítek v závislosti na nakliknutém zařízení
        """
        self.current_block_id = str(widget_tier)
        button_strings = Catalogue_gui.get_device_strings(self.current_block_id)
        
        if button_strings [0] == "":
            # self.new_device.configure(text = button_strings[0],state = tk.DISABLED)
            self.new_device.configure(text = button_strings[0])
            self.new_device.pack_forget()
            self.new_device.master.configure(width=1,height=1)
        else:
            # self.new_device.configure(text = button_strings[0],state = tk.NORMAL)
            self.new_device.configure(text = button_strings[0])
            self.new_device.pack(pady = 0, padx = (10,0),anchor="w",side="left")
        self.edit_device.configure(text = button_strings[1])
        self.del_device.configure(text = button_strings[2])

        self.button_copy.configure(text = button_strings[3])
        if self.copy_memory != "":
            if len(self.copy_widget_tier) == 2:
                button_label = "Vložit stanici"
            if len(self.copy_widget_tier) == 4:
                button_label = "Vložit kameru"
            if len(self.copy_widget_tier) == 6:
                button_label = "Vložit optiku/ světlo"
            self.button_copy.configure(text = button_label)

        if button_strings [3] == "":
            # self.button_copy.configure(state = tk.DISABLED)
            self.button_copy.pack_forget()
        else:
            # self.button_copy.configure(state = tk.NORMAL)
            self.button_copy.pack(pady = 0, padx = (10,0),anchor="w",side="left")

        if self.last_selected_widget[0] != "" and self.last_selected_widget[0].winfo_exists():
            # if self.last_selected_widget[0]._border_color.lower() != "#ffff00":
            if self.last_selected_widget[1] == "light":
                self.last_selected_widget[0].configure(border_color="#ffff00")
            else:
                self.last_selected_widget[0].configure(border_color="#636363")

        self.last_selected_widget[0] = widget
        self.last_selected_widget[1] = ""
        if widget._border_color.lower() == "#ffff00":
            self.last_selected_widget[1] = "light"

        # if widget._border_color.lower() != "#ffff00":
        widget.configure(border_color="white")

    def show_station_images(self,event,widget_id):
        def manage_photo_callback(updated_list,last_path):
            self.station_list[station_index]["image_list"] = updated_list
            self.last_path_to_images = last_path

        station_index = int(widget_id)
        image_list_given = []
        if "image_list" in self.station_list[station_index]:
            print("station indes:",station_index)
            image_list_given = self.station_list[station_index]["image_list"]
        
        show_im = Insert_image(self.root,None,image_list_given,manage_photo_callback,self.default_subwindow_status,self.last_path_to_images)
        show_im.image_menu_gui()

    def show_context_menu(self,event,widget_id):
        button_strings = Catalogue_gui.get_device_strings(str(widget_id))
        context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="#202020",activebackground="#606060")
        if len(str(widget_id)) == 2: # Station extra options
            context_menu.add_command(label="Nová stanice",font=("Arial",22,"bold"), command=lambda: self.manage_widgets("",str(widget_id),btn="add_line"))
            context_menu.add_separator()
            station_index = int(widget_id)
            if "image_list" in self.station_list[station_index]:
                context_menu.add_command(label="Načíst obrázky",font=("Arial",22,"bold"), command=lambda: self.show_station_images("",str(widget_id)))
                context_menu.add_separator()
            else:
                context_menu.add_command(label="Přidat obrázky",font=("Arial",22,"bold"), command=lambda: self.show_station_images("",str(widget_id)))
                context_menu.add_separator()
        if button_strings[0] != "":
            context_menu.add_command(label=button_strings[0],font=("Arial",22,"bold"),command=lambda: self.manage_widgets("",str(widget_id),btn="add_object"))
            context_menu.add_separator()
        context_menu.add_command(label=button_strings[1],font=("Arial",22,"bold"),command=lambda: self.edit_object("",str(widget_id),rewrite_temp = True))
        context_menu.add_separator()
        context_menu.add_command(label=button_strings[2],font=("Arial",22,"bold"),command=lambda: self.delete_block("",str(widget_id)))
        if button_strings[3] != "":
            context_menu.add_separator()
            context_menu.add_command(label=button_strings[3],font=("Arial",22,"bold"), command=lambda: self.copy_objects(str(widget_id)))
            if self.copy_memory != "":
                # context_menu.add_command(label="Vložit kopírované",font=("Arial",22,"bold"), command=lambda: self.copy_objects(str(widget_id),paste=True),state="disabled")
                context_menu.add_separator()
                context_menu.add_command(label="Vložit kopírované",font=("Arial",22,"bold"), command=lambda: self.copy_objects(str(widget_id),paste=True))

        if len(str(widget_id)) == 7: # Controller extra options
            controller_index = int(widget_id[5:7])
            controller_ip = self.controller_object_list[controller_index]["ip"]
            controller_username = self.controller_object_list[controller_index]["username"]
            controller_password = self.controller_object_list[controller_index]["password"]
            if controller_ip != "" and controller_ip != "192.168.000.000":
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat IP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(controller_ip))
                
            if controller_username != "":
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat uživ. jméno",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(controller_username))
            if controller_password != "":
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat heslo",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(controller_password))
        
        context_menu.tk_popup(event.x_root, event.y_root)
        # self.opened_window = context_menu
        
    def make_block(self,master_widget,height,width,fg_color,text,side,dummy_block = False,tier = "",border_color="#636363",anchor="w",fill=None):
        def opened_window_check():
            if self.opened_window == "":
                return False
            try:
                if self.opened_window.winfo_exists():
                    return True
                else:
                    return False
            except Exception as err:
                print(err)
                return False
            
        def leave_expanded(e,widget_tier,widget):
            # pokud někde rozliknutá, schovej:
            if self.leave_expanded_widget == widget:
                return
            
            if self.leave_expanded_widget != None: # to close the expanded widget
                try:
                    self.switch_widget_info(e,self.leave_expanded_widget_tier,self.leave_expanded_widget)
                    self.leave_expanded_widget.expanded_status = False
                except Exception as e:
                    print("exception error: ",e)
                    
            self.leave_expanded_widget = widget
            self.leave_expanded_widget.expanded_status = True
            self.leave_expanded_widget_tier = widget_tier

        def on_enter(e, widget_tier,widget):
            if not opened_window_check():
                if self.leave_expanded_widget != widget and widget.expanded_status == False:
                        self.switch_widget_info(e, widget_tier,widget)
                        widget.expanded_status = True
                        widget.update()
                        widget.update_idletasks()

        def on_leave(e, widget_tier,widget,flag=""):
            if not opened_window_check():
                if self.leave_expanded_widget != widget and widget.expanded_status == True:
                    self.switch_widget_info(e, widget_tier,widget)
                    widget.expanded_status = False
                    widget.update()
                    widget.update_idletasks()

        if dummy_block:
            dummy_block_widget =    customtkinter.CTkFrame(master=master_widget,corner_radius=0,height=height,width =width-10,fg_color="#212121")
            dummy_block_widget.     pack(pady = 0,padx =0,expand = False,side = side,anchor=anchor)
            return dummy_block_widget
        else:
            block_widget =      customtkinter.CTkFrame(master=master_widget,corner_radius=0,fg_color=fg_color,height=height,width =width,border_width= 2,border_color=border_color)
            block_widget.       pack(pady = (0,0),padx =0,expand = False,side = side,anchor=anchor)
            block_name =        customtkinter.CTkLabel(master = block_widget,text = text,font=("Arial",25,"bold"),width=block_widget.cget("width")-15,height=block_widget.cget("height")-15,justify = "left",anchor="w")
            if fill == None:
                block_name.pack(pady = 5,padx =5,expand = False)
            else:
                block_name.pack(pady = 5,padx =5,expand = False,fill=fill)
            
            block_name.bind("<Button-3>",lambda e, widget_tier=tier: self.show_context_menu(e, widget_tier),"+")
            block_name.bind("<Button-1>",lambda e, widget_tier=tier,widget = block_widget: self.select_block(e, widget_tier,widget),"+")

            if self.hover_trigger_mode == "1":
                block_name.         bind("<Enter>",lambda e, widget_tier=tier,widget = block_name: on_enter(e, widget_tier,widget))
                block_name.         bind("<Leave>",lambda e, widget_tier=tier,widget = block_name: on_leave(e, widget_tier,widget))
                # block_name.         bind("<Button-3>",lambda e, widget_tier=tier,widget = block_name: on_leave(e, widget_tier,widget,flag="context_menu"),"+")
                # block_name.         bind("<Button-3>",lambda e, widget_tier=tier,widget = block_name: leave_expanded(e, widget_tier,widget),"+")
                block_name.         bind("<Button-1>",lambda e, widget_tier=tier,widget = block_name: leave_expanded(e, widget_tier,widget),"+")
            else:
                block_widget.       bind("<Button-1>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
                block_name.         bind("<Button-1>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
            
            block_name.expanded_status = False


            return block_name
        
    def make_new_object(self,which_one,object_to_edit = None,cam_index = None,optic_index = None):
        """
        which_one:
        - station
        - camera
        - optic
        - accessory
        """
        if which_one == "station":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            
            camera = {
                "type": "",
                "controller": "",
                "controller_color": "",
                "cable": "",
                "optics_list": [optic],
                "description": "",
            }
            station = {
                "name": "",
                "inspection_description": "",
                "camera_list": [camera],
            }

            return station
        
        elif which_one == "camera":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            camera = {
                "type": "",
                "controller": "",
                "controller_color": "",
                "cable": "",
                "optics_list": [optic],
                "description": "",
            }
            object_to_edit["camera_list"].append(camera)
            return object_to_edit
        
        elif which_one == "optic":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }

            object_to_edit["camera_list"][cam_index]["optics_list"].append(optic)
            return object_to_edit
        
        elif which_one == "accessory":
            accessory = {
                "type": "",
                "description":"",
            }

            object_to_edit["accessory_list"].append(accessory)
            return object_to_edit

    def manage_widgets(self,args,widget_tier,btn,open_edit = True,rewrite_temp = True):
        if rewrite_temp:
            self.temp_station_list = copy.deepcopy(self.station_list)

        if btn == "add_line": # nova stanice
            new_station = self.make_new_object("station")
            if len(self.temp_station_list) > 0 and self.temp_station_list[0] != "00":
                self.temp_station_list.insert(int(widget_tier)+1,new_station)
                widget_tier = str(int(widget_tier)+1)
                if len(widget_tier) < 2:
                    widget_tier = "0" + widget_tier
                self.edit_object("",widget_tier) # logika pro vložení pod posledně zvolený widget
            else:
                self.temp_station_list.append(new_station)
                self.edit_object("",widget_tier,new_station=True)
            return
        
        if len(widget_tier) == 2: #01-99 stanice
            if btn == "add_object": # nova kamera ke stanici 0101-9999 kamery
                station_index = int(widget_tier[:2])
                station_with_new_camera = self.make_new_object("camera",object_to_edit = self.temp_station_list[station_index])
                self.temp_station_list[station_index] = station_with_new_camera
                if open_edit:
                    new_camera_index = len(self.temp_station_list[station_index]["camera_list"])-1
                    if new_camera_index > -1:
                        if len(str(new_camera_index)) == 1:
                            new_camera_index = "0" + str(new_camera_index)
                    else:
                        new_camera_index = "00"
                    self.edit_object("",widget_tier+str(new_camera_index),new_station=False)

        elif len(widget_tier) == 7: #xxxxc01-xxxxc99 kontrolery - tzn. nove prislusenstvi ke kontroleru
            if btn == "add_object":
                controller_index = int(widget_tier[5:7])
                controller_with_new_accessories = self.make_new_object("accessory",object_to_edit = self.controller_object_list[controller_index])
                self.controller_object_list[controller_index] = controller_with_new_accessories
                if open_edit:
                    self.edit_object("",widget_tier,new_station=False)
        
        elif len(widget_tier) == 4: # 0101-9999 kamery, nove bude pridano: 010101-999999 optika
            if btn == "add_object": # nova optika kamery
                station_index = int(widget_tier[:2])
                camera_index = int(widget_tier[2:])
                camera_with_new_optics = self.make_new_object("optic",object_to_edit = self.temp_station_list[station_index],cam_index = camera_index)
                self.temp_station_list[station_index] = camera_with_new_optics
                if open_edit:
                    new_optics_index = len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"])-1
                    if new_optics_index > -1:
                        if len(str(new_optics_index)) == 1:
                            new_optics_index = "0" + str(new_optics_index)
                    else:
                        new_optics_index = "00"
                    self.edit_object("",widget_tier+str(new_optics_index),new_station=False)
        
        print("widget_tier: ",widget_tier)

    def confirm_delete(self,to_del_object):
        def make_decision(decision):
            if decision == True:
                self.station_list.pop(to_del_object)
                self.make_project_widgets()
                Tools.add_colored_line(self.main_console,f"Stanice {station_name} byla úspěšně odstraněna","orange",None,True)
            close_window(child_root)

        def close_window(window):
            child_root.grab_release()
            self.root.unbind("<Button-1>")
            window.destroy()

        child_root = customtkinter.CTkToplevel(fg_color="#212121")
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"650x130+{x+80}+{y+80}")
        child_root.after(200, lambda: child_root.iconbitmap(self.app_icon_path))
        child_root.title("Upozornění")
        station_name = str(self.station_list[to_del_object]["name"])
        top_frame =         customtkinter.CTkFrame(master = child_root,corner_radius=0,fg_color="#212121")
        warning_icon =      customtkinter.CTkLabel(master = top_frame,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/warning.png")),size=(50,50)),bg_color="#212121")
        proceed_label =     customtkinter.CTkLabel(master = top_frame,text = f"Opravdu si přejete odstranit celou stanici ({station_name}),\nvčetně všech zařízení ke stanici připojených?",font=("Arial",20,"bold"),anchor="w",justify="left")
        warning_icon       .pack(pady=10,padx=30,side = "left",anchor="w")
        proceed_label      .pack(pady=10,padx=10,side = "left",anchor="w")
        button_yes =        customtkinter.CTkButton(master = child_root,text = "Pokračovat",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: make_decision(True))
        button_no =         customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  make_decision(False))
        top_frame           .pack(pady=0,padx=0,expand=False,side = "top",anchor="w")
        button_no           .pack(pady = 5, padx = (0,10),anchor="w",expand=False,side="right")
        button_yes          .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        self.root.          bind("<Button-1>",lambda e: close_window(child_root))
        child_root.update()
        child_root.update_idletasks()
        child_root.focus_force()
        child_root.focus() 
        child_root.grab_set()
        self.opened_window = child_root
        
    def delete_block(self,args,widget_tier):
        if "c" in widget_tier:
            #kontrolery:
            if len(widget_tier) == 7:
                station_index = int(widget_tier[:2])
                camera_index = int(widget_tier[2:4])
                # controller_index = int(widget_tier[5:7])
                self.station_list[station_index]["camera_list"][camera_index]["controller"] = None
                self.station_list[station_index]["camera_list"][camera_index]["controller_index"] = None

            # příslušenství ke kontroleru:
            if len(widget_tier) == 9:
                controller_index = int(widget_tier[5:7])
                accessory_index = int(widget_tier[7:9])
                self.controller_object_list[controller_index]["accessory_list"].pop(accessory_index)

            self.make_project_widgets()
            return
            
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            print("deleting",self.station_list[station_index])
            self.confirm_delete(station_index)
            return
        
        elif len(widget_tier) == 4: # 0101-9999 kamery
            camera_index = int(widget_tier[2:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index])
            self.station_list[station_index]["camera_list"].pop(camera_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"].pop(optic_index)
            
        elif len(widget_tier) == 8: # 01010101-99999999 prislusenstvi
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:6])
            accessory_index = int(widget_tier[6:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"].pop(accessory_index)
        
        #refresh
        self.make_project_widgets()

    def edit_object_gui_new(self,object:str,station_index,camera_index = None,optics_index = None,accessory_index = None,controller_index = None,all_parameters = False,new_station = False):
        """
        Object:
        - station
        - camera
        - optics
        """
        
        def save_changes(no_window_shut = False):
            db_error_found = False
            if object == "station" or all_parameters:
                self.temp_station_list[station_index]["name"] = new_name.get()
                filtered_description = Tools.make_wrapping(str(new_description.get("1.0", tk.END)))
                self.temp_station_list[station_index]["inspection_description"] = filtered_description

            if object == "camera" or all_parameters:
                camera_item = str(camera_type_entry.get())
                self.temp_station_list[station_index]["camera_list"][camera_index]["type"] = camera_item
                if not camera_item in self.whole_camera_type_database:
                    camera_type_entry.configure(fg_color = "#bd1931",border_color = "red")
                    # return "db_error"
                    db_error_found = True
                else:
                    camera_type_entry.configure(fg_color = "#343638",border_color = "#565B5E")

                self.temp_station_list[station_index]["camera_list"][camera_index]["controller"] = controller_entry.get()
                current_controller = controller_entry.get()
                controller_index = None
                if str(current_controller).replace(" ","") != "":
                    for controllers in self.controller_object_list:
                        if str(str(controllers["name"])+"("+controllers["type"]+")").replace(" ","") == str(current_controller).replace(" ",""):
                            controller_index = self.controller_object_list.index(controllers)
                            self.last_controller_index = controller_index+1 #musíme počítat s možností nemít žádný kontroler
                            break
                self.temp_station_list[station_index]["camera_list"][camera_index]["controller_index"] = controller_index
                cable_item = str(cam_cable_menu.get())
                self.temp_station_list[station_index]["camera_list"][camera_index]["cable"] = cable_item
                if not cable_item in self.whole_camera_cable_database:
                    cam_cable_menu.configure(fg_color = "#bd1931",border_color = "red")
                    # return "db_error"
                    db_error_found = True
                else:
                    cam_cable_menu.configure(fg_color = "#343638",border_color = "#565B5E")

                filtered_description = Tools.make_wrapping(str(notes_input.get("1.0", tk.END)))
                self.temp_station_list[station_index]["camera_list"][camera_index]["description"] = filtered_description
                
            if object == "optics" or "camera" or all_parameters:
                optic_type = str(optic_type_entry.get())
                if len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"]) > 0:
                    self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"] = optic_type
                    if not optic_type in self.whole_optics_database and not optic_type in self.whole_light_database:
                        optic_type_entry.configure(fg_color = "#bd1931",border_color = "red")
                        db_error_found = True
                        # return "db_error"
                    else:
                        optic_type_entry.configure(fg_color = "#343638",border_color = "#565B5E")

                    alternative_item = str(alternative_entry.get())
                    self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"] = alternative_item
                    if not alternative_item in self.whole_optics_database and not alternative_item in self.whole_light_database:
                        alternative_entry.configure(fg_color = "#bd1931",border_color = "red")
                        db_error_found = True
                        # return "db_error"
                    else:
                        alternative_entry.configure(fg_color = "#343638",border_color = "#565B5E")
                    filtered_description = Tools.make_wrapping(str(notes_input2.get("1.0", tk.END)))
                    self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"] = filtered_description
            if db_error_found:
                return "db_error"
            
            if not no_window_shut:
                self.station_list = copy.deepcopy(self.temp_station_list)
                self.make_project_widgets() #refresh
                self.close_window(child_root)

        def next_station():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index

            station_index += 1
            if station_index < len(self.temp_station_list):
                station_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                station_index += 1
                camera_index = 0
                optics_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # TLACITKO +:
                station_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                camera_index = 0
                optics_index = 0
                self.station_list = copy.deepcopy(self.temp_station_list)
                close_window(child_root)
                if station_index < 10:
                    widget_tier = "0" + str(station_index)
                else:
                    widget_tier = str(station_index)
                self.manage_widgets("",widget_tier,"add_line")

        def previous_station():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            station_index -= 1
            if station_index > -1:
                station_index += 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                station_index -= 1
                camera_index = 0
                optics_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                station_index += 1
                camera_index = 0
                optics_index = 0
            
        def next_camera():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            camera_index += 1
            if camera_index < len(self.temp_station_list[station_index]["camera_list"]):
                camera_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                camera_index += 1
                optics_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání

            else: # TLACITKO +:
                camera_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice¨
                if save_status == "db_error":
                    return
                camera_index += 1
                optics_index = 0
                accessory_index = 0
                if station_index < 10:
                    widget_tier_st = "0" + str(station_index)
                else:
                    widget_tier_st = str(station_index)

                print("camera st widget tier",widget_tier_st)
                self.manage_widgets("",widget_tier_st,"add_object",open_edit=False,rewrite_temp=False)
                initial_prefill() # prefill s novým indexem 

        def previous_camera():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            camera_index -= 1
            if camera_index > -1:
                camera_index += 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                camera_index -= 1
                optics_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                camera_index += 1

        def next_optic():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            optics_index += 1
            if optics_index < len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"]):
                optics_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                optics_index += 1
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání

            else: # TLACITKO +:
                optics_index -= 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                optics_index += 1
                if station_index < 10:
                    widget_tier_st = "0" + str(station_index)
                else:
                    widget_tier_st = str(station_index)
                if camera_index < 10:
                    widget_tier_cam = "0" + str(camera_index)
                else:
                    widget_tier_cam = str(camera_index)
                widget_tier = widget_tier_st + widget_tier_cam
                self.manage_widgets("",widget_tier,"add_object",open_edit=False,rewrite_temp=False)
                initial_prefill() # prefill s novým indexem

        def previous_optic():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            optics_index -= 1
            if optics_index > -1:
                optics_index += 1
                save_status = save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                if save_status == "db_error":
                    return
                optics_index -= 1
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                optics_index += 1

        def close_window(child_root):
            try:
                if opened_subwindow.winfo_exists():
                    self.close_window(opened_subwindow)
            except Exception:
                pass
            
            self.root.unbind("<Button-1>")
            child_root.destroy()

        def callback_new_controller(new_controller_data):
            print("saving new controller: ",new_controller_data)
            new_controller = {
                "type": new_controller_data[0],
                "name": new_controller_data[1],
                "color": new_controller_data[2],
                "ip": new_controller_data[3],
                "username": new_controller_data[4],
                "password": new_controller_data[5],
                "accessory_list": new_controller_data[6],
                "notes": new_controller_data[7],
            }
            print("Nový kontroler------ ",new_controller)
            self.controller_object_list.append(new_controller)
            new_drop_option = f"{new_controller['name']} ({new_controller['type']})"
            self.custom_controller_drop_list.append(new_drop_option)
            controller_entry.configure(values = self.custom_controller_drop_list)
            controller_entry.set(new_drop_option)
            controller_opt_menu_color("",only_color=new_controller["color"])
            child_root.focus_force()

        def import_notes(which):
            """
            - camera (Kontroler, Kamera, Kabel)
            - optics (Objektiv, Alternativní)
            - accessory
            """
            if which == "camera":
                current_camera = camera_type_entry.get()
                current_cable = cam_cable_menu.get()
                notes_string = ""
                if current_camera != "":
                    camera_notes = str(self.camera_notes_database[self.whole_camera_type_database.index(current_camera)])
                    if camera_notes != "":
                        notes_string = notes_string + "Kamera - popis: " + camera_notes + "\n\n"
                if current_cable != "":
                    cable_notes = str(self.cable_notes_database[self.whole_camera_cable_database.index(current_cable)]) 
                    if cable_notes != "":
                        notes_string = notes_string + "Kabel (" + str(current_cable) + "): " + cable_notes + "\n\n"
                
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",notes_string)
            
            elif which == "optics":
                current_optics = optic_type_entry.get()
                current_alternative = alternative_entry.get()
                notes_string = ""
                if current_optics != "":
                    optic_notes = str(self.optics_notes_database[self.whole_optics_database.index(current_optics)])
                    if optic_notes !="":
                        notes_string = notes_string + "Objektiv - popis: " + optic_notes + "\n\n"
                if current_alternative != "":
                    alternative_notes = str(self.optics_notes_database[self.whole_optics_database.index(current_alternative)])
                    if alternative_notes != "":
                        notes_string = notes_string + "Alternativní - popis: " + alternative_notes + "\n\n"
                
                notes_input2.delete("1.0",tk.END)
                notes_input2.insert("1.0",notes_string)

        def call_new_controller_gui():
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_new_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            self.opened_window = window.new_controller_window(child_root)

        def controller_opt_menu_color(*args,only_color = False):
            nonlocal controller_entry
            if not only_color:
                current_controller = str(*args)
                if str(current_controller).replace(" ","") != "":
                    for controllers in self.controller_object_list:
                        if controllers["color"] != "":
                            if (controllers["name"]+"("+controllers["type"]+")").replace(" ","") == current_controller.replace(" ",""):
                                controller_entry.configure(fg_color = str(controllers["color"]))
                                break
                else:
                    controller_entry.configure(fg_color = "#636363")
            else:
                controller_entry.configure(fg_color = str(only_color))

        def optics_lights_switch(reverse = False):
            if reverse:
                if self.optic_light_option == "optic":
                    self.optic_light_option = "light"
                else:
                    self.optic_light_option = "optic"

            if self.optic_light_option == "optic":
                self.optic_light_option = "light"
                optic_type.configure(text = "Typ světla:")
                optic_type_entry.   unbind("<KeyRelease>")
                alternative_entry.  unbind("<KeyRelease>")
                optic_type_entry.   bind("<KeyRelease>",lambda e: autosearch_engine(e,"lights"))
                alternative_entry.  bind("<KeyRelease>",lambda e: autosearch_engine(e,"lights_alternative"))

                light_checkbox.select()
                optics_checkbox.deselect()
                optic_search.unbind("<Button-1>")
                alternative_search.unbind("<Button-1>")
                optic_search.bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_light_database,optic_type_entry,mirror=True))
                alternative_search.bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_light_database,alternative_entry,mirror=True))

            else:
                self.optic_light_option = "optic"
                optic_type.configure(text = "Typ objektivu:")
                optic_type_entry.   unbind("<KeyRelease>")
                alternative_entry.  unbind("<KeyRelease>")
                optic_type_entry.   bind("<KeyRelease>",lambda e: autosearch_engine(e,"optics"))
                alternative_entry.  bind("<KeyRelease>",lambda e: autosearch_engine(e,"optics_alternative"))

                light_checkbox.deselect()
                optics_checkbox.select()
                optic_search.unbind("<Button-1>")
                alternative_search.unbind("<Button-1>")
                optic_search.bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_optics_database,optic_type_entry,mirror=True))
                alternative_search.bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_optics_database,alternative_entry,mirror=True))

        def remaping_characters(event):
            remap = {
                'ì': 'ě',
                'è': 'č',
                'ø': 'ř'
            }

            if event.char in remap:
                widget = event.widget
                replacement = remap[event.char]

                # Zjistit, zda je něco vybráno
                try:
                    selection_start = widget.index("sel.first")
                    selection_end = widget.index("sel.last")
                    widget.delete(selection_start, selection_end)
                    widget.insert(selection_start, replacement)
                except tk.TclError:
                    # Pokud nic není vybrané, vlož na pozici kurzoru
                    widget.insert(tk.INSERT, replacement)

                return "break"
        
        def add_photo():
            """
            Pozor pracuje se tu s temp station listem 
            - a bere se tu v potaz childroot
            """
            def add_photo_callback(updated_list,last_path):
                self.temp_station_list[station_index]["image_list"] = updated_list
                self.last_path_to_images = last_path
            image_list_given = []
            if "image_list" in self.temp_station_list[station_index]:
                image_list_given = self.temp_station_list[station_index]["image_list"]
            insert_image_class = Insert_image(self.root,
                                              child_root,
                                              image_list_given,
                                              add_photo_callback,
                                              self.default_subwindow_status,
                                              self.last_path_to_images)
            insert_image_class.image_menu_gui()
            
        def call_text_wrap(textbox_widget):
            wrapped_text = Tools.make_wrapping(str(textbox_widget.get("1.0", tk.END)))
            textbox_widget.delete("0.0","end")
            textbox_widget.insert("0.0",wrapped_text)

        def manage_option_menu(e,values,entry_widget,mirror=None,auto_search_call=False):
            def on_item_selected(value):
                # if auto_search_call:
                entry_widget.delete(0,200)
                entry_widget.insert(0,str(value))
                # else:
                #     entry_widget.set(str(value))
                window.destroy()

            if len(values) == 0:
                return

            screen_x = child_root.winfo_pointerx()
            screen_y = child_root.winfo_pointery()
            parent_x = child_root.winfo_rootx()+e.x
            parent_y = child_root.winfo_rooty()+e.y
            x = screen_x - parent_x +entry_widget.winfo_width()
            y = screen_y - parent_y +entry_widget.winfo_height()

            if auto_search_call:
                screen_x = entry_widget.winfo_rootx()
                screen_y = entry_widget.winfo_rooty() + entry_widget.winfo_height()+5

            font = tkFont.Font(family="Arial", size=20)
            max_width_px = 40
            try:
                max_width_px = max(font.measure(str(val)) for val in values) + 40  # Add some padding
            except Exception as e:
                pass
            window = customtkinter.CTkToplevel(child_root)
            window.overrideredirect(True)
            window.configure(bg="black")
            listbox = FakeContextMenu(window, values, command=on_item_selected, width=max_width_px)
            listbox.pack(fill="both",expand=True)
            child_root.bind("<Button-1>", lambda e: window.destroy(), "+")

            max_visible_items = 50
            visible_items = min(len(values), max_visible_items)
            total_height = visible_items * int(listbox.one_button_height)+20
            child_root.update_idletasks()
            if total_height > child_root._current_height-20-y:
                total_height = child_root._current_height-20-y

            if mirror == True: #priznak aby pri maximalizovani nelezlo mimo obrazovku (doprava)
                screen_x=screen_x-max_width_px
            window.geometry(f"{max_width_px}x{total_height}+{screen_x}+{screen_y}")
            if auto_search_call:
                self.autosearch_menu = window

        def autosearch_engine(e,which_item):
            """
            which_item:
            - camera
            - optics
            - optics_alternative
            - lights
            - lights_alternative
            - cables
            """
            if self.autosearch_menu != None:
                self.autosearch_menu.destroy()
                self.autosearch_menu = None

            if which_item == "camera":
                entry_widget = camera_type_entry
                database = self.whole_camera_type_database
            elif which_item == "optics":
                entry_widget = optic_type_entry
                database = self.whole_optics_database
            elif which_item == "optics_alternative":
                entry_widget = alternative_entry
                database = self.whole_optics_database
            elif which_item == "lights":
                entry_widget = optic_type_entry
                database = self.whole_light_database
            elif which_item == "lights_alternative":
                entry_widget = alternative_entry
                database = self.whole_light_database
            elif which_item == "cables":
                entry_widget = cam_cable_menu
                database = self.whole_camera_cable_database

            entry_widget.update_idletasks()
            currently_inserted = str(entry_widget.get()).strip().lower()
            if len(str(currently_inserted))==0:
                return
            found_itemss = []

            for items in database:
                item_str = str(items).lower()
                if currently_inserted in str(item_str):
                # if item_str.startswith(currently_inserted):
                    found_itemss.append(str(items))

            found_itemss = sorted(found_itemss)
            # print(found_itemss)
            manage_option_menu(e,found_itemss,entry_widget,auto_search_call=True)

        child_root = customtkinter.CTkToplevel()
        icon_small = 45
        icon_large = 49
        # STANICE ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        station_frame =             customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)
        station_name_label =        customtkinter.CTkLabel(master = station_frame,text = "Název stanice:",font=("Arial",22,"bold"))
        name_frame =                customtkinter.CTkFrame(master = station_frame,corner_radius=0)
        button_prev_st =            customtkinter.CTkButton(master = name_frame,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_station())
        new_name =                  customtkinter.CTkEntry(master = name_frame,font=("Arial",22),height=50,corner_radius=0)
        button_next_st =            customtkinter.CTkButton(master = name_frame,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_station())
        button_prev_st.             pack(pady = 5, padx = 0,anchor="w",expand=False,side="left")
        new_name.                   pack(pady = 5, padx = 0,anchor="w",expand=True,side="left",fill="x")
        button_next_st.             pack(pady = 5, padx = 0,anchor="w",expand=False,side="left")
        button_add_photo =          customtkinter.CTkButton(master = station_frame,text = "Přiřadit/ zobrazit fotografii",font=("Arial",22,"bold"),height=50,corner_radius=0,command=lambda: add_photo())
        description_label_frame =   customtkinter.CTkFrame(master = station_frame,corner_radius=0,fg_color="#212121")
        inspection_description =    customtkinter.CTkLabel(master = description_label_frame,text = "Popis inspekce:",font=("Arial",22,"bold"))
        wrap_text_btn =             customtkinter.CTkButton(master = description_label_frame,text = "Zarovnat text",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: call_text_wrap(new_description))
        inspection_description.     pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        wrap_text_btn.              pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        new_description =           customtkinter.CTkTextbox(master = station_frame,font=("Arial",22),width=450,height=600,corner_radius=0)
        station_name_label.         pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
        name_frame.                 pack(pady = 5, padx = 5,anchor="w",expand=False,side="top",fill="x")
        button_add_photo.           pack(pady=(5,5),padx=10,anchor="w",expand=False,side = "top",fill="x")
        description_label_frame.    pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        new_description.            pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        new_name.                   bind("<Key>",remaping_characters)
        new_description.            bind("<Key>",remaping_characters)

        # KAMERY ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        camera_frame =              customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)
        counter_frame_cam =         customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="#212121")
        button_prev_cam =           customtkinter.CTkButton(master = counter_frame_cam,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_camera())
        counter_cam =               customtkinter.CTkLabel(master = counter_frame_cam,text = "0/0",font=("Arial",22,"bold"))
        button_next_cam =           customtkinter.CTkButton(master = counter_frame_cam,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_camera())
        button_prev_cam.            pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        counter_cam.                pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_cam.            pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        camera_type =               customtkinter.CTkLabel(master = camera_frame,text = "Typ kamery:",font=("Arial",22,"bold"))
        option_menu_frame_cam =     customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="#212121")
        camera_type_entry =         customtkinter.CTkEntry(master = option_menu_frame_cam,font=("Arial",22),height=50,corner_radius=0)
        camera_search =             customtkinter.CTkLabel(master = option_menu_frame_cam,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        camera_search.              bind("<Enter>",lambda e: camera_search._image.configure(size=(icon_large,icon_large)))
        camera_search.              bind("<Leave>",lambda e: camera_search._image.configure(size=(icon_small,icon_small)))
        camera_search.              bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_camera_type_database,camera_type_entry))
        camera_type_entry.          pack(pady = 5, padx = (5,5),anchor="w",expand=True,side="left",fill="x")
        camera_search.              pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        camera_type_entry.          bind("<KeyRelease>",lambda e: autosearch_engine(e,"camera"))

        cam_cable =                 customtkinter.CTkLabel(master = camera_frame,text = "Kabel ke kameře:",font=("Arial",22,"bold"))
        option_menu_frame_cable =   customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="#212121")
        cam_cable_menu =            customtkinter.CTkEntry(master = option_menu_frame_cable,font=("Arial",22),height=50,corner_radius=0)
        cable_search =              customtkinter.CTkLabel(master = option_menu_frame_cable,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        cable_search.               bind("<Enter>",lambda e: cable_search._image.configure(size=(icon_large,icon_large)))
        cable_search.               bind("<Leave>",lambda e: cable_search._image.configure(size=(icon_small,icon_small)))
        cable_search.               bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_camera_cable_database,cam_cable_menu))
        cam_cable_menu.             pack(pady = 5, padx = (5,5),anchor="w",expand=True,side="left",fill="x")
        cable_search.               pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        cam_cable_menu.             bind("<KeyRelease>",lambda e: autosearch_engine(e,"cables"))

        controller =                customtkinter.CTkLabel(master = camera_frame,text = "Kontroler:",font=("Arial",22,"bold"))
        controller_frame =          customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="#212121")
        controller_entry =          customtkinter.CTkOptionMenu(master = controller_frame,font=("Arial",22),dropdown_font=("Arial",22),width=280,height=50,values=self.custom_controller_drop_list,corner_radius=0,fg_color="#212121",command=controller_opt_menu_color)
        new_controller =            customtkinter.CTkButton(master = controller_frame,text = "Přidat",font=("Arial",22,"bold"),width = 80,height=50,corner_radius=0,command=lambda: call_new_controller_gui())
        controller_entry.           pack(pady = 5, padx = (10,0),anchor="w",expand=True,side="left",fill="x")
        new_controller.             pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        note_label_frame =          customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="#212121")
        note_label =                customtkinter.CTkLabel(master = note_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes_btn =          customtkinter.CTkButton(master = note_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("camera"))
        wrap_text_btn2 =            customtkinter.CTkButton(master = note_label_frame,text = "Zarovnat text",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: call_text_wrap(notes_input))
        note_label.                 pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes_btn.           pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        wrap_text_btn2.             pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        notes_input =               customtkinter.CTkTextbox(master = camera_frame,font=("Arial",22),corner_radius=0,width=450,height=450)
        counter_frame_cam.          pack(pady=(10,0),padx= 3,anchor="n",expand=False,side="top")
        camera_type.                pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_cam.      pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        cam_cable.                  pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_cable.    pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        controller.                 pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        controller_frame.           pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        new_controller.             pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        note_label_frame.           pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        notes_input.                pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        notes_input.                bind("<Key>",remaping_characters)

        # OPTIKA --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        if "" in self.optics_database:
            self.optics_database.pop(self.optics_database.index(""))
        optics_frame =              customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)
        counter_frame_optics =      customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="#212121")
        button_prev_opt =           customtkinter.CTkButton(master = counter_frame_optics,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_optic())
        counter_opt =               customtkinter.CTkLabel(master = counter_frame_optics,text = "0/0",font=("Arial",22,"bold"))
        self.button_next_opt =      customtkinter.CTkButton(master = counter_frame_optics,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_optic())
        button_prev_opt.            pack(pady = 0, padx = (5,0),anchor="w",side="left")
        counter_opt.                pack(pady = 0, padx = (5,0),anchor="w",side="left")
        self.button_next_opt.       pack(pady = 0, padx = (5,0),anchor="w",side="left")
        checkbox_frame =            customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="#212121")
        light_checkbox =            customtkinter.CTkCheckBox(master = checkbox_frame, text = "Světla",font=("Arial",22,"bold"),command=lambda:optics_lights_switch())
        optics_checkbox =           customtkinter.CTkCheckBox(master = checkbox_frame, text = "Objektivy",font=("Arial",22,"bold"),command=lambda:optics_lights_switch())
        light_checkbox.             pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        optics_checkbox.            pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        optic_type =                customtkinter.CTkLabel(master = optics_frame,text = "Typ objektivu:",font=("Arial",22,"bold"))
        option_menu_frame_optic =   customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="#212121")
        optic_type_entry =          customtkinter.CTkEntry(master = option_menu_frame_optic,font=("Arial",22),height=50,corner_radius=0)
        optic_search =              customtkinter.CTkLabel(master = option_menu_frame_optic,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        optic_search.               bind("<Enter>",lambda e: optic_search._image.configure(size=(icon_large,icon_large)))
        optic_search.               bind("<Leave>",lambda e: optic_search._image.configure(size=(icon_small,icon_small)))
        optic_search.               bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_optics_database,optic_type_entry,mirror=True))
        optic_type_entry.           pack(pady = 5, padx = (5,5),anchor="w",expand=True,side="left",fill="x")
        optic_search.               pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        optic_type_entry.           bind("<KeyRelease>",lambda e: autosearch_engine(e,"optics"))
        alternative_type =          customtkinter.CTkLabel(master = optics_frame,text = "Alternativa:",font=("Arial",22,"bold"))
        option_menu_frame_alternative = customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="#212121")
        alternative_entry =         customtkinter.CTkEntry(master = option_menu_frame_alternative,font=("Arial",22),height=50,corner_radius=0)
        
        alternative_search =        customtkinter.CTkLabel(master = option_menu_frame_alternative,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/SearchWhite.png")),size=(icon_small,icon_small)),bg_color="#212121")
        alternative_search.         bind("<Enter>",lambda e: alternative_search._image.configure(size=(icon_large,icon_large)))
        alternative_search.         bind("<Leave>",lambda e: alternative_search._image.configure(size=(icon_small,icon_small)))
        alternative_search.         bind("<Button-1>",lambda e: manage_option_menu(e,self.whole_optics_database,alternative_entry,mirror=True))
        alternative_entry.          pack(pady = 5, padx = (5,5),anchor="w",expand=True,side="left",fill="x")
        alternative_search.         pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        alternative_entry.          bind("<KeyRelease>",lambda e: autosearch_engine(e,"optics_alternative"))
        
        note2_label_frame =         customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="#212121")
        note2_label =               customtkinter.CTkLabel(master = note2_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes2_btn =         customtkinter.CTkButton(master = note2_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("optics"))
        wrap_text_btn3 =            customtkinter.CTkButton(master = note2_label_frame,text = "Zarovnat text",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: call_text_wrap(notes_input2))
        note2_label.                pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes2_btn.          pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        wrap_text_btn3.             pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        notes_input2 =              customtkinter.CTkTextbox(master = optics_frame,font=("Arial",22),width=450,height=450,corner_radius=0,wrap= "word")
        counter_frame_optics.       pack(pady=(10,0),padx=3,anchor="n",side = "top")
        checkbox_frame.             pack(pady = 5, padx = 10,anchor="n",expand=False,side="top")
        optic_type.                 pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_optic.    pack(pady = (5,0), padx = 10,anchor="w",expand=False,side="top",fill="x")
        alternative_type.           pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_alternative.pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        note2_label_frame.          pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        notes_input2.               pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        notes_input2.               bind("<Key>",remaping_characters)
        optics_lights_switch(reverse = True)

        def refresh_counters():
            nonlocal station_index
            nonlocal optics_index
            nonlocal camera_index
            nonlocal counter_cam
            nonlocal counter_opt

            try:
                counter_cam_state = str(camera_index+1) + "/" + str(len(self.temp_station_list[station_index]["camera_list"]))
                counter_cam.configure(text = counter_cam_state)
            except Exception:
                pass
            try:
                counter_opt_state = str(optics_index+1) + "/" + str(len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"]))
                counter_opt.configure(text = counter_opt_state)
            except Exception:
                pass

        def refresh_button_appearance():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal button_prev_st
            nonlocal button_next_st
            nonlocal button_prev_cam
            nonlocal button_next_cam
            nonlocal button_prev_opt

            def unbind_tooltip(widget):
                widget.unbind("<Enter>")
                widget.unbind("<Leave>")
                widget.unbind("<Button-1>")

            def config_buttons(button_left,button_right,index,max_array_value,product = "stanice"):
                if index ==0:
                    button_left.event_generate("<Button-1>")
                    button_left.unbind("<Enter>")
                    button_left.configure(text = "",fg_color = "#636363")
                else:
                    button_left.configure(text = "<",fg_color = "#636363")
                    if self.show_tooltip == "ano":
                        unbind_tooltip(button_left)
                        child_root.after(100, lambda: Catalogue_gui.ToolTip(button_left,f" Předcházející {product} ",child_root,subwindow_status=True))

                if index == max_array_value:
                    button_right.configure(text = "+",fg_color = "green")
                    if self.show_tooltip == "ano":
                        unbind_tooltip(button_right)
                        child_root.after(100, lambda: Catalogue_gui.ToolTip(button_right,f" Nová {product} ",child_root,subwindow_status=True))
                else:
                    button_right.configure(text = ">",fg_color = "#636363")
                    if self.show_tooltip == "ano":
                        unbind_tooltip(button_right)
                        child_root.after(100, lambda: Catalogue_gui.ToolTip(button_right,f" Další {product} ",child_root,subwindow_status=True))

            config_buttons(button_prev_st,button_next_st,station_index,len(self.temp_station_list)-1)
            config_buttons(button_prev_cam,button_next_cam,camera_index,len(self.temp_station_list[station_index]["camera_list"])-1,product="kamera")
            config_buttons(button_prev_opt,self.button_next_opt,optics_index,len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"])-1,product="optika")

        def initial_prefill():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            new_name.delete(0,300)
            new_name.insert(0,str(self.temp_station_list[station_index]["name"]))
            new_description.delete("0.0","end")
            new_description.insert("0.0",str(self.temp_station_list[station_index]["inspection_description"]))
            # initial prefill - camera:
            try:
                if len(self.temp_station_list[station_index]["camera_list"]) == 0:
                    camera_type_entry.delete(0,300)
                    controller_entry.set("")
                    cam_cable_menu.delete(0,300)
                    notes_input.delete("1.0",tk.END)

                if str(self.temp_station_list[station_index]["camera_list"][camera_index]["type"]) in self.whole_camera_type_database:
                    camera_type_entry.delete(0,300)
                    camera_type_entry.insert(0,str(self.temp_station_list[station_index]["camera_list"][camera_index]["type"]))
                if str(self.temp_station_list[station_index]["camera_list"][camera_index]["controller"]) in self.custom_controller_drop_list:
                    controller_entry.set(str(self.temp_station_list[station_index]["camera_list"][camera_index]["controller"]))
                if str(self.temp_station_list[station_index]["camera_list"][camera_index]["cable"]) in self.whole_camera_cable_database:
                    cam_cable_menu.delete(0,300)
                    cam_cable_menu.insert(0,str(self.temp_station_list[station_index]["camera_list"][camera_index]["cable"]))

                
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",str(self.temp_station_list[station_index]["camera_list"][camera_index]["description"]))
            except TypeError as typeerr_msg:
                print("ERROR: ",typeerr_msg)
                camera_index = 0
                if len(self.temp_station_list[station_index]["camera_list"]) > 0:
                    if str(self.temp_station_list[station_index]["camera_list"][camera_index]["type"]) in self.whole_camera_type_database:
                        camera_type_entry.delete(0,300)
                        camera_type_entry.insert(0,str(self.temp_station_list[station_index]["camera_list"][camera_index]["type"]))
                    if self.last_controller_index < len(self.custom_controller_drop_list)-1:
                        controller_entry.set(self.custom_controller_drop_list[self.last_controller_index])
                    
                    try:
                        assigned_controller_index = int(self.temp_station_list[station_index]["camera_list"][camera_index]["controller_index"])
                        controller_entry.set(self.custom_controller_drop_list[assigned_controller_index])
                    except Exception:
                        pass
                    
                    if str(self.temp_station_list[station_index]["camera_list"][camera_index]["cable"]) in self.whole_camera_cable_database:
                        cam_cable_menu.delete(0,300)
                        cam_cable_menu.insert(0,str(self.temp_station_list[station_index]["camera_list"][camera_index]["cable"]))
                    notes_input.delete("1.0",tk.END)
                    notes_input.insert("1.0",str(self.temp_station_list[station_index]["camera_list"][camera_index]["description"]))
            except IndexError:
                camera_index = 0
                # bypass aby vychazeli indexy... neni osetřeno proti nule (kamer nebo objektivů) skoro nikde
                station_with_new_camera = self.make_new_object("camera",object_to_edit=self.temp_station_list[station_index])
                self.temp_station_list[station_index] = station_with_new_camera

            # initial prefill - optics:
            try:
                optic_type = str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"])
                optic_alternative = str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"])
                if optic_type in self.whole_optics_database:
                    if light_checkbox.get() != 1:
                        optics_lights_switch(reverse=True)
                    else:
                        optics_lights_switch()
                    optic_type_entry.delete(0,300)
                    optic_type_entry.insert(0,optic_type)

                elif optic_type in self.whole_light_database:
                    if light_checkbox.get() == 1:
                        optics_lights_switch(reverse=True)
                    else:
                        optics_lights_switch()
                    optic_type_entry.delete(0,300)
                    optic_type_entry.insert(0,optic_type)
                else:
                    optic_type_entry.delete(0,300)

                if optic_alternative in self.whole_optics_database or optic_alternative in self.whole_light_database:
                    alternative_entry.delete(0,300)
                    alternative_entry.insert(0,optic_alternative)
                else:
                    alternative_entry.delete(0,300)
                notes_input2.delete("1.0",tk.END)
                notes_input2.insert("1.0",str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"]))
            except TypeError:
                optics_index = 0
                optic_type = str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"])
                optic_alternative = str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"])
                if len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"]) > 0:
                    if optic_type in self.whole_optics_database:
                        if light_checkbox.get() != 1:
                            optics_lights_switch(reverse=True)
                        else:
                            optics_lights_switch()
                        optic_type_entry.delete(0,300)
                        optic_type_entry.insert(0,optic_type)

                    
                    elif optic_type in self.whole_light_database:
                        if light_checkbox.get() == 1:
                            optics_lights_switch(reverse=True)
                        else:
                            optics_lights_switch()
                        optic_type_entry.delete(0,300)
                        optic_type_entry.insert(0,optic_type)
                    else:
                        optic_type_entry.delete(0,300)

                    if optic_alternative in self.whole_optics_database or optic_alternative in self.whole_light_database:
                        alternative_entry.delete(0,300)
                        alternative_entry.insert(0,optic_alternative)
                    else:
                        alternative_entry.delete(0,300)
                    notes_input2.delete("1.0",tk.END)
                    notes_input2.insert("1.0",str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"]))
            except Exception:
                optics_index = 0

            refresh_counters()
            refresh_button_appearance()
            controller_opt_menu_color(controller_entry.get())

        initial_prefill()
        button_frame =  customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_frame    .pack(pady = 0, padx = 0,fill="x",anchor="s",expand=False,side="bottom")
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        one_segment_width = 450
        height = 850
        child_root.after(200, lambda: child_root.iconbitmap(self.app_icon_path))

        if object == "station":
            # width = 3*one_segment_width
            # child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování stanice: " + str(self.temp_station_list[station_index]["name"]))
            station_frame   .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            camera_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        elif object == "camera":
            # width = 2*one_segment_width
            # child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování kamery: " + str(self.temp_station_list[station_index]["camera_list"][camera_index]["type"]))
            camera_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        elif object == "optics":
            # width = one_segment_width
            # child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování optiky: " + str(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        button_save =   customtkinter.CTkButton(master = button_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_changes())
        button_exit =   customtkinter.CTkButton(master = button_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(child_root))
        button_exit     .pack(pady = 10, padx = (5,10),anchor="e",expand=False,side="right")
        button_save     .pack(pady = 10, padx = 5,anchor="e",expand=False,side="right")

        if self.default_subwindow_status == 1:
            child_root.state('zoomed')

        if self.show_tooltip == "ano":
            Catalogue_gui.ToolTip(wrap_text_btn," Zarovnat text na rozměr buňky ",child_root,subwindow_status=True)
            Catalogue_gui.ToolTip(wrap_text_btn2," Zarovnat text na rozměr buňky ",child_root,subwindow_status=True)
            Catalogue_gui.ToolTip(wrap_text_btn3," Zarovnat text na rozměr buňky ",child_root,subwindow_status=True,reverse=True)

        child_root.update()
        child_root.update_idletasks()
        child_root.focus_force()
        child_root.focus()
        self.opened_window = child_root
        
    def edit_object(self,args,widget_tier,new_station = False,rewrite_temp = False):
        if rewrite_temp:
            self.temp_station_list = copy.deepcopy(self.station_list)

        def callback_edited_controller(new_controller_data):
            nonlocal controller_index
            old_controller = f"{self.controller_object_list[controller_index]['name']} ({self.controller_object_list[controller_index]['type']})"
            for stations in self.station_list:
                for cameras in stations["camera_list"]:
                    if cameras["controller"] == old_controller:
                        cameras["controller"] = f"{new_controller_data[1]} ({new_controller_data[0]})"

            self.controller_object_list[controller_index]["type"] = new_controller_data[0]
            self.controller_object_list[controller_index]["name"] = new_controller_data[1]
            self.controller_object_list[controller_index]["color"] = new_controller_data[2]
            self.controller_object_list[controller_index]["ip"] = new_controller_data[3]
            self.controller_object_list[controller_index]["username"] = new_controller_data[4]
            self.controller_object_list[controller_index]["password"] = new_controller_data[5]
            self.controller_object_list[controller_index]["accessory_list"] = new_controller_data[6]
            self.controller_object_list[controller_index]["notes"] = new_controller_data[7]
            # refresh dropdownlist:
            self.custom_controller_drop_list = []
            for controllers in self.controller_object_list:
                new_drop_option = f"{controllers['name']} ({controllers['type']})"
                self.custom_controller_drop_list.append(new_drop_option)
            self.make_project_widgets()
            self.root.focus_force()
        if len(widget_tier) == 2: #01-99 stanice
            station_index = int(widget_tier[:2])
            if new_station:
                self.edit_object_gui_new("station",(len(self.temp_station_list)-1),all_parameters=True,new_station=new_station)
            else:
                print("editing",self.temp_station_list[station_index])
                # kdyz nova kamera, chci rovnou editovat tu nově přidanou
                current_cam_count = len(self.temp_station_list[station_index]["camera_list"])
                camera_index = 0
                if current_cam_count > 0:
                    camera_index = current_cam_count-1
                self.edit_object_gui_new("station",station_index,camera_index,all_parameters=True)

        elif len(widget_tier) == 4: # 0101-9999 kamery
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:])
            print("editing",self.temp_station_list[station_index]["camera_list"][camera_index])
            # kdyz edituju, chci prejit na posledni optiku...
            current_optics_count = len(self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"])
            optic_index = 0
            if current_optics_count > 0:
                optic_index = current_optics_count-1
            self.edit_object_gui_new("camera",station_index,camera_index,optic_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("editing",self.temp_station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.edit_object_gui_new("optics",station_index,camera_index,optic_index)
        
        elif len(widget_tier) == 7: # xxxxc01-xxxxc99 kontolery
            controller_index = int(widget_tier[5:7])
            current_acc_count = len(self.controller_object_list[controller_index]["accessory_list"])
            accessory_index = 0
            if current_acc_count > 0:
                accessory_index = current_acc_count-1
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_edited_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            self.opened_window = window.new_controller_window(childroot=None,controller=self.controller_object_list[controller_index],edit=True,accessory_index=accessory_index)

        elif len(widget_tier) == 9: # xxxxc0101-xxxxc9999 prislusenstvi kontoleru
            controller_index = int(widget_tier[5:7])
            accessory_index = int(widget_tier[7:9])
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_edited_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            self.opened_window = window.new_controller_window(childroot=None,controller=self.controller_object_list[controller_index],edit=True,accessory_index = accessory_index,only_accessory=True)

    def export_to_excel(self,path_with_name,favourite_format,path_inserted):
        if path_with_name == None and favourite_format == None: # only save path button
            self.last_path_input = path_inserted
            self.path_for_callback = path_inserted
            Tools.save_to_json_config(path_inserted,"catalogue_settings","default_path")
            return

        self.favourite_format = favourite_format
        Tools.save_to_json_config(self.favourite_format,"catalogue_settings","default_export_suffix")
        self.last_path_input = path_inserted
        save_excel_class = Save_excel(self.root,station_list = self.station_list,project_name = self.project_name_input.get(),excel_name=path_with_name,controller_list=self.controller_object_list,console = self.main_console)
        output = save_excel_class.main()
        return output
       
    def load_metadata_callback(self,input_data):
        self.station_list = input_data[0]
        print("loaded station list: ",self.station_list)

        self.controller_object_list = input_data[1]
        if str(input_data[2]) != "None":
            self.project_name_input.delete(0,300)
            self.project_name_input.insert(0,str(input_data[2]))

        self.custom_controller_drop_list = [""]
        for controllers in self.controller_object_list:
            new_drop_option = f"{controllers['name']} ({controllers['type']})"
            self.custom_controller_drop_list.append(new_drop_option)

        self.make_project_widgets(return_scroll=False)

    def call_save_metadata_gui(self,exiting_status = False,only_save_flag = False):
        def callback_save_last_input(filename,path_inserted,path_to_save,saving = False):
            nonlocal exiting_status

            if filename != None:
                self.last_xml_filename = filename
            if path_inserted != None:
                self.last_path_input = path_inserted
            if path_to_save != None:
                self.last_path_input = path_to_save
                self.path_for_callback = path_to_save

            if saving:
                print("save was made")
                self.changes_made = False
            if exiting_status:
                self.call_menu()

        if only_save_flag:
            self.opened_window = ToplevelWindow.save_prog_options_window(self.root,
                                                    self.app_icon_path,
                                                    self.controller_object_list,
                                                    self.main_console,
                                                    self.station_list,
                                                    self.project_name_input.get(),
                                                    self.load_metadata_callback,
                                                    callback_save_last_input,
                                                    self.last_xml_filename,
                                                    self.last_path_input,
                                                    self.default_xml_file_name,
                                                    self.default_path,
                                                    exit_status = exiting_status)
        else:
            self.opened_window = ToplevelWindow.load_prog_window(self.root,
                                            self.app_icon_path,
                                            self.controller_object_list,
                                            self.main_console,
                                            self.station_list,
                                            self.project_name_input.get(),
                                            self.load_metadata_callback,
                                            callback_save_last_input,
                                            self.last_xml_filename,
                                            self.last_path_input,
                                            self.default_xml_file_name,
                                            self.default_path,
                                            exit_status = exiting_status)

    def copy_objects(self,widget_tier,paste=False):
        """
        - self.copy_widget_tier = tier zkopirovaneho widgetu
        - widget_tier = aktuální, nakliknutý widget
        """
        if paste:
            try:
                button_strings = Catalogue_gui.get_device_strings(self.current_block_id)
                if len(self.copy_widget_tier) == 2:
                    station_index = int(widget_tier[:2])
                    if station_index +1 >len(self.station_list):
                        self.station_list.append(self.copy_memory)
                    else:
                        self.station_list.insert(station_index+1,self.copy_memory)

                    station_copyed = self.copy_memory["name"]
                    Tools.add_colored_line(self.main_console,f"Stanice {station_copyed} byla vložena pod {self.station_list[station_index]["name"]}","green",None,True)
                    self.copy_memory = ""
                    self.make_project_widgets()
                    self.button_copy.configure(text = button_strings[3])

                elif len(self.copy_widget_tier) == 4:
                    station_index = int(widget_tier[:2])
                    if len(widget_tier) < len(self.copy_widget_tier):
                        camera_index = len(self.station_list[station_index]["camera_list"])
                    else:
                        camera_index = int(widget_tier[2:4])

                    if camera_index +1 >len(self.station_list[station_index]["camera_list"]):
                        self.station_list[station_index]["camera_list"].append(self.copy_memory)
                    else:
                        self.station_list[station_index]["camera_list"].insert(camera_index+1,self.copy_memory)
                    Tools.add_colored_line(self.main_console,f"Kamera {self.copy_memory["type"]} byla přiřazena k: {self.station_list[station_index]["name"]}","green",None,True)
                    self.copy_memory = ""
                    self.make_project_widgets()
                    self.button_copy.configure(text = button_strings[3])

                elif len(self.copy_widget_tier) == 6:
                    station_index = int(widget_tier[:2])
                    if len(widget_tier) < 4:
                        Tools.add_colored_line(self.main_console,f"Pro vložení optiky definujte kameru nebo optiku pro vložení, nikoliv stanici...","red",None,True)
                        return
                    
                    camera_index = int(widget_tier[2:4])
                    if len(widget_tier) < 6:
                        optic_index = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"])
                    else:
                        optic_index = int(widget_tier[4:])

                    if optic_index +1 >len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"]):
                        self.station_list[station_index]["camera_list"][camera_index]["optics_list"].append(self.copy_memory)
                    else:
                        self.station_list[station_index]["camera_list"][camera_index]["optics_list"].insert(optic_index+1,self.copy_memory)
                    print("\n\n",self.station_list[station_index]["camera_list"][camera_index]["optics_list"])
                    Tools.add_colored_line(self.main_console,f"Optika {self.copy_memory["type"]} byla přiřazena k: {self.station_list[station_index]["camera_list"][camera_index]["type"]}","green",None,True)
                    self.copy_memory = ""
                    self.make_project_widgets()
                    self.button_copy.configure(text = button_strings[3])
                else:
                    Tools.add_colored_line(self.main_console,f"Pro vložení naklikněte jinou buňku než kontroler nebo příslušenství... (jsou navázány na kameru)","red",None,True)
            except Exception as e:
                Tools.add_colored_line(self.main_console,f"Neočekávaná chyba: {e}","red",None,True)

        else:
            try:
                if len(widget_tier) == 2:
                    station_index = int(widget_tier[:2])
                    self.copy_memory = copy.deepcopy(self.station_list[station_index])
                    self.copy_widget_tier = widget_tier
                    self.button_copy.configure(text = "Vložit stanici")
                    Tools.add_colored_line(self.main_console,f"Stanice {self.station_list[station_index]["name"]} byla zkopírována do schránky","green",None,True)

                elif len(widget_tier) == 4:
                    station_index = int(widget_tier[:2])
                    camera_index = int(widget_tier[2:4])
                    self.copy_memory = copy.deepcopy(self.station_list[station_index]["camera_list"][camera_index])
                    self.copy_widget_tier = widget_tier
                    self.button_copy.configure(text = "Vložit kameru")
                    Tools.add_colored_line(self.main_console,f"Kamera {self.station_list[station_index]["camera_list"][camera_index]["type"]} byla zkopírována do schránky","green",None,True)

                elif len(widget_tier) == 6:
                    station_index = int(widget_tier[:2])
                    camera_index = int(widget_tier[2:4])
                    optic_index = int(widget_tier[4:])
                    self.copy_memory = copy.deepcopy(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
                    self.copy_widget_tier = widget_tier
                    self.button_copy.configure(text = "Vložit optiku/ světlo")
                    Tools.add_colored_line(self.main_console,f"Optika {self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"]} byla zkopírována do schránky","green",None,True)

                else:
                    Tools.add_colored_line(self.main_console,f"Kontroler a příslušenství nelze kopírovat (jsou navázány na kameru)","red",None,True)
            except Exception as e:
                Tools.add_colored_line(self.main_console,f"Neočekávaná chyba: {e}","red",None,True)

    def create_main_widgets(self,initial=False):
        def call_manage_widgets(button):
            """
            add_line = přidat pouze stanici
            add_object = vše nové, ostatní
            """
            def get_last_st_position():
                next_st_widget_tier = len(self.station_list)-1
                if next_st_widget_tier < 10:
                    next_st_widget_tier = "0" + str(next_st_widget_tier)
                return str(next_st_widget_tier)

            widget_tier = ""
            widget_tier = self.current_block_id
            if len(self.station_list) == 0:
                widget_tier = "00"

            if button == "add_line":
                if widget_tier == "":
                    widget_tier = get_last_st_position()

                if widget_tier != "":
                    if len(widget_tier) > 2: # pokud je nakliknuteho neco jiného než stanice - přidej novou pod poslední
                        station_tier = widget_tier[:2]
                        self.current_block_id = station_tier
                        self.manage_widgets("",self.current_block_id,btn=button)
                    else:
                        self.manage_widgets("",widget_tier,btn=button)
                    return
            elif widget_tier != "" and self.current_block_id != "":
                self.manage_widgets("",widget_tier,btn=button)
                return
            
            Tools.add_colored_line(self.main_console,f"Nejprve zvolte pro co zařízení přidat","red",None,True)
        
        def call_edit_object():
            widget_tier = ""
            widget_tier = self.current_block_id
            if widget_tier != "":
                self.edit_object("",widget_tier,rewrite_temp = True)
            else:
                Tools.add_colored_line(self.main_console,f"Nejprve zvolte zařízení pro editaci","red",None,True)

        def call_delete_object():
            widget_tier = ""
            widget_tier = self.current_block_id
            if widget_tier != "":
                self.delete_block("",widget_tier)
            else:
                Tools.add_colored_line(self.main_console,f"Nejprve zvolte zařízení pro odebrání","red",None,True)

        def switch_manufacturer():
            if self.chosen_manufacturer == "Omron":
                # manufacturer_logo =             customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/keyence_logo.png")),size=(240, 50))
                self.chosen_manufacturer = "Keyence"
                self.camera_database_pointer = 0
                self.optics_database_pointer = 0
                self.camera_cable_database_pointer = 0
                self.accessory_database_pointer = 0
                # switch_manufacturer_image.configure(image = manufacturer_logo)
                switch_manufacturer_btn.configure(image = customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/keyence_logo.png")),size=(manuf_logo_width,manuf_logo_height)))

                self.read_database()
            elif self.chosen_manufacturer == "Keyence":
                # manufacturer_logo =             customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/omron_logo.png")),size=(240, 50))
                self.chosen_manufacturer = "Omron"
                self.camera_database_pointer = 0
                self.optics_database_pointer = 0
                self.camera_cable_database_pointer = 0
                self.accessory_database_pointer = 0
                # switch_manufacturer_image.configure(image = manufacturer_logo)
                switch_manufacturer_btn.configure(image = customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/omron_logo.png")),size=(manuf_logo_width,manuf_logo_height)))
                self.read_database()

        def call_setting_window():
            def apply_changes_callback(input_data):
                if input_data[0] == "open_all_cmd":
                    if input_data[1] == 1 and self.detailed_view != True:
                        self.detailed_view = True
                        self.make_project_widgets(return_scroll=False) # dá se očkávat nárůst - reset scrollbaru
                    elif input_data[1] == 0 and self.detailed_view != False:
                        self.detailed_view = False
                        self.make_project_widgets()
                    return
                
                elif input_data[0] == "set_render_mode":
                    self.render_mode = input_data[1]
                    return
                
                elif input_data[0] == "hover_info_trigger_mode":
                    self.hover_trigger_mode = input_data[1]
                    self.make_project_widgets()
                    return
                
                elif input_data[0] != "":
                    self.default_excel_filename = input_data[0]

                if input_data[1] != "":
                    self.default_xml_file_name = input_data[1]
                if input_data[2] != "":
                    self.default_subwindow_status = input_data[2]
                if input_data[3] != "":
                    self.default_database_filename = input_data[3]

            self.opened_window = ToplevelWindow.setting_window(self.root,
                                          self.app_icon_path,
                                          self.default_excel_filename,
                                          self.default_xml_file_name,
                                          self.default_subwindow_status,
                                          apply_changes_callback,
                                          self.default_database_filename,
                                          self.detailed_view,
                                          self.render_mode,
                                          self.hover_trigger_mode)
            
        def call_menu_routine():
            self.opened_window = ToplevelWindow(self.root,changes_check = self.changes_made).save_check(self.call_menu,self.call_save_metadata_gui)

        def remaping_characters(event):
            if event.char == 'ì':
                event.widget.insert(tk.INSERT, 'ě')
                return "break"  # Stop the event from inserting the original character
            elif event.char == 'è':
                event.widget.insert(tk.INSERT, 'č')
                return "break"  # Stop the event from inserting the original character
            elif event.char == 'ø':
                event.widget.insert(tk.INSERT, 'ř')
                return "break"  # Stop the event from inserting the original character

        def call_copy_object():
            widget_tier = ""
            widget_tier = self.current_block_id
            if widget_tier != "":
                if self.copy_memory != "": #vkládám přes tlačítko...
                    self.copy_objects(widget_tier,paste=True)
                    return

                self.copy_objects(widget_tier)

            else:
                Tools.add_colored_line(self.main_console,f"Nejprve zvolte, co chcete kopírovat","red",None,True)

        def call_export_window():
            self.opened_window = ToplevelWindow.export_option_window(self.root,
                                                                    self.app_icon_path,
                                                                    self.export_to_excel,
                                                                    self.format_list,
                                                                    self.favourite_format,
                                                                    self.last_path_input,
                                                                    self.default_path,
                                                                    self.default_excel_filename,
                                                                    str(self.project_name_input.get())
                                                                    )
        
        def call_db_login_window(call_export=False):
            def db_callback(connection):
                self.current_db_connection = connection
            def call_export_callback():
                call_db_export()
            self.opened_window = ToplevelWindow.db_login_window(self.root,
                                                                self.app_icon_path,
                                                                db_label,
                                                                db_callback,
                                                                call_export,
                                                                call_export_callback,
                                                                self.main_console
                                                                )
            
        def call_db_export():
            if self.current_db_connection == None:
                call_db_login_window(call_export = True)
                return

            table_to_export = Tools.make_table_for_db_export(self.station_list,self.controller_object_list)
            if len(table_to_export) == 0:
                Tools.add_colored_line(self.main_console,f"Není co exportovat","red",None,True)
                return
            
            def login_error():
                call_db_login_window(call_export = True)
            self.opened_window = ToplevelWindow.export_to_db_window(self.root,
                                                                    self.app_icon_path,
                                                                    self.current_db_connection,
                                                                    self.project_name_input.get(),
                                                                    table_to_export,
                                                                    login_error,
                                                                    self.main_console
                                                                    )
        icon_small = 45
        icon_large = 49

        def make_icon_and_text_button(master,text,icon):
            label_large = icon_large+10
            label_small = icon_small+10
            icon_text_button = customtkinter.CTkLabel(master = master,
                                                      width=label_large*2,height=label_large,
                                                      text = text,font=("Arial",22,"bold"),
                                                      image =customtkinter.CTkImage(PILImage.open(Tools.resource_path(f"images/{icon}.png")),size=(label_small/2,label_small/2)),
                                                      bg_color="#212121",justify="center",compound="top")
            icon_text_button.bind("<Enter>",lambda e: icon_text_button._image.configure(size=(label_large/2,label_large/2)))
            icon_text_button.bind("<Leave>",lambda e: icon_text_button._image.configure(size=(label_small/2,label_small/2)))
            return icon_text_button

        self.clear_frame(self.root)
        main_header =                   customtkinter.CTkFrame(master=self.root,corner_radius=0)
        main_header_left =              customtkinter.CTkFrame(master=main_header,corner_radius=0,fg_color="#212121")
        main_header_row0 =              customtkinter.CTkFrame(master=main_header_left,corner_radius=0,fg_color="#636363")
        buttons_frame =                 customtkinter.CTkFrame(master=main_header_left,corner_radius=0,fg_color="#212121")
        main_header_row1 =              customtkinter.CTkFrame(master=buttons_frame,corner_radius=0,fg_color="#212121")
        main_menu_button =              customtkinter.CTkButton(master = main_header_row0, width = 200,height=50,text = "MENU",command = lambda: call_menu_routine(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        main_menu_button.               pack(pady = (10,0),padx =(20,0),anchor = "s",side = "left")
        project_label =                 customtkinter.CTkLabel(master = main_header_row1,text = "Projekt:",font=("Arial",25,"bold"))
        self.project_name_input =       customtkinter.CTkEntry(master = main_header_row1,font=("Arial",22),width=250,height=50,placeholder_text="Název projektu",corner_radius=0)
        export_button =                 customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/export_excel.png")),size=(icon_small,icon_small)),bg_color="#212121")
        export_button.                  bind("<Enter>",lambda e: export_button._image.configure(size=(icon_large,icon_large)))
        export_button.                  bind("<Leave>",lambda e: export_button._image.configure(size=(icon_small,icon_small)))
        export_button.                  bind("<Button-1>",lambda e: call_export_window())
        save_button =                   customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/save_file.png")),size=(icon_small,icon_small)),bg_color="#212121")
        save_button.                    bind("<Enter>",lambda e: save_button._image.configure(size=(icon_large,icon_large)))
        save_button.                    bind("<Leave>",lambda e: save_button._image.configure(size=(icon_small,icon_small)))
        save_button.                    bind("<Button-1>",lambda e: self.call_save_metadata_gui(only_save_flag=True))
        load_button =                   customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/open_file.png")),size=(icon_small,icon_small)),bg_color="#212121")
        load_button.                    bind("<Enter>",lambda e: load_button._image.configure(size=(icon_large,icon_large)))
        load_button.                    bind("<Leave>",lambda e: load_button._image.configure(size=(icon_small,icon_small)))
        load_button.                    bind("<Button-1>",lambda e: self.call_save_metadata_gui())
        db_export =                     customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/db_upload.png")),size=(icon_small,icon_small)),bg_color="#212121")
        db_export.                      bind("<Enter>",lambda e: db_export._image.configure(size=(icon_large,icon_large)))
        db_export.                      bind("<Leave>",lambda e: db_export._image.configure(size=(icon_small,icon_small)))
        db_export.                      bind("<Button-1>",lambda e: call_db_export())
        button_settings =               customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/settings.png")),size=(icon_small,icon_small)),bg_color="#212121")
        button_settings.                bind("<Enter>",lambda e: button_settings._image.configure(size=(icon_large,icon_large)))
        button_settings.                bind("<Leave>",lambda e: button_settings._image.configure(size=(icon_small,icon_small)))
        button_settings.                bind("<Button-1>",lambda e: call_setting_window())
        new_station =                   make_icon_and_text_button(main_header_row1,"Nová stanice","green_plus")
        new_station.                    bind("<Button-1>",lambda e: call_manage_widgets("add_line"))
        new_device_frame =              customtkinter.CTkFrame(master=main_header_row1,corner_radius=0,fg_color="#212121")
        self.new_device =               make_icon_and_text_button(new_device_frame,"Nová kamera","green_plus")
        self.new_device.                bind("<Button-1>",lambda e: call_manage_widgets("add_object"))
        self.new_device.                pack(pady = 0, padx = (10,0),anchor="w",side="left")
        self.edit_device =              make_icon_and_text_button(main_header_row1,"Editovat stanici","edit")
        self.edit_device.               bind("<Button-1>",lambda e: call_edit_object())
        self.del_device =               make_icon_and_text_button(main_header_row1,"Odebrat stanici","delete_file")
        self.del_device.                bind("<Button-1>",lambda e: call_delete_object())
        self.button_copy =              make_icon_and_text_button(main_header_row1,"Kopírovat stanici","copy_file")
        self.button_copy.               bind("<Button-1>",lambda e: call_copy_object())
        db_login =                      customtkinter.CTkLabel(master = main_header_row1,width=icon_large,text = "",image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/login.png")),size=(icon_small,icon_small)),bg_color="#212121")
        db_login.                       bind("<Enter>",lambda e: db_login._image.configure(size=(icon_large,icon_large)))
        db_login.                       bind("<Leave>",lambda e: db_login._image.configure(size=(icon_small,icon_small)))
        db_label =                      customtkinter.CTkLabel(master = main_header_row1,text = "Nepřihlášen",font=("Arial",25,"bold"),text_color="red")
        db_login.                       bind("<Button-1>",lambda e: call_db_login_window())
        project_label.                  pack(pady = 0, padx = (10,0),anchor="w",side="left")
        self.project_name_input.        pack(pady = 0, padx = (10,0),anchor="w",side="left")
        save_button.                    pack(pady = 0, padx = (20,0),anchor="w",side="left")
        load_button.                    pack(pady = 0, padx = (20,0),anchor="w",side="left")
        db_export.                      pack(pady = 0, padx = (20,0),anchor="w",side="left")
        export_button.                  pack(pady = 0, padx = (15,0),anchor="w",side="left")
        button_settings.                pack(pady = 0, padx = (20,0),anchor="w",side="left")
        new_station.                    pack(pady = 0, padx = (20,0),anchor="w",side="left")
        new_device_frame.               pack(pady = 0, padx = (0,0),anchor="w",side="left",expand=False)
        self.edit_device.               pack(pady = 0, padx = (20,0),anchor="w",side="left")
        self.del_device.                pack(pady = 0, padx = (20,0),anchor="w",side="left")
        self.button_copy.               pack(pady = 0, padx = (20,0),anchor="w",side="left")
        db_label.                       pack(pady = 0, padx = (0,20),anchor="e",side="right")
        db_login.                       pack(pady = 0, padx = (0,10),anchor="e",side="right")
        self.project_name_input.        bind("<Key>",remaping_characters)
        image_frame =                   customtkinter.CTkFrame(master=main_header,corner_radius=0,fg_color="#212121")#,fg_color="#212121")
        logo =                          customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 102))
        image_logo =                    customtkinter.CTkLabel(master = image_frame,text = "",image =logo,bg_color="#212121")
        image_logo.                     pack(pady=(15,0),padx=0)
        console_frame=                  customtkinter.CTkFrame(master=self.root,corner_radius=0)
        manuf_logo_width = 240
        manuf_logo_height = 50
        manuf_logo_width_small = 216
        manuf_logo_height_small = 45
        switch_manufacturer_label =     customtkinter.CTkLabel(master = console_frame,text = "Změnit výrobce komponentů:",font=("Arial",25,"bold"))
        switch_manufacturer_btn =       customtkinter.CTkLabel(master = console_frame,
                                                    width=manuf_logo_width,height=manuf_logo_height,
                                                    text = "",font=("Arial",20,"bold"),
                                                    image =customtkinter.CTkImage(PILImage.open(Tools.resource_path("images/omron_logo.png")),size=(manuf_logo_width_small,manuf_logo_height_small)),
                                                    bg_color="#212121",justify="center",compound="bottom")
        switch_manufacturer_btn.        bind("<Enter>",lambda e: switch_manufacturer_btn._image.configure(size=(manuf_logo_width,manuf_logo_height)))
        switch_manufacturer_btn.        bind("<Leave>",lambda e: switch_manufacturer_btn._image.configure(size=(manuf_logo_width_small,manuf_logo_height_small)))
        switch_manufacturer_btn.        bind("<Button-1>",lambda e: switch_manufacturer())
        console_label =                 customtkinter.CTkLabel(master = console_frame,text = " > ",font=("Arial",40,"bold"))
        self.main_console =             tk.Text(console_frame, wrap="none", height=0,background="black",foreground="#565B5E",font=("Arial",22),state=tk.DISABLED,relief="flat")
        console_label.                  pack(pady = (0,10), padx =(10,5),anchor="w",side="left")
        self.main_console.              pack(pady = (0,10), padx =(0,10),anchor="w",expand=True,fill="x",side="left",ipady=3,ipadx=5)
        switch_manufacturer_btn.        pack(pady = (0,10), padx = (0,30),anchor="e",side="right")
        switch_manufacturer_label.      pack(pady = (0,10), padx = (10,5),anchor="e",side="right")

        column_labels =                 customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        self.project_tree =             customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        stations_column_header =        customtkinter.CTkLabel(master = column_labels,text = "Stanice",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        camera_column_header =          customtkinter.CTkLabel(master = column_labels,text = "Kamera",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        optics_column_header =          customtkinter.CTkLabel(master = column_labels,text = "Objektiv/ světla",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        controller_column_header =      customtkinter.CTkLabel(master = column_labels,text = "Kontrolery",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        accessory_column_header =       customtkinter.CTkLabel(master = column_labels,text = "Příslušenství",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        stations_column_header.         pack(pady=(15,0),padx=15,expand=False,side = "left")
        camera_column_header.           pack(pady=(15,0),padx=15,expand=False,side = "left")
        optics_column_header.           pack(pady=(15,0),padx=15,expand=False,side = "left")
        controller_column_header.       pack(pady=(15,0),padx=15,expand=False,side = "left")
        accessory_column_header.        pack(pady=(15,0),padx=15,expand=False,side = "left")
        main_header_row0.               pack(pady=0,padx=0,expand=False,fill="x",side = "top",anchor="w")
        main_header_row1.               pack(pady=(0,0),padx=0,expand=False,fill="x",side = "top",anchor="w")
        buttons_frame.                  pack(pady=0,padx=0,fill="x",side = "left",anchor="w",expand=True)
        image_frame.                    pack(pady=0,padx=0,side = "right",anchor="n",ipadx = 15)
        main_header_left.               pack(pady=0,padx=5,fill="both",side = "left",anchor="w",expand=True)
        main_header.                    pack(pady=0,padx=5,fill="x",side = "top",ipady = 10,ipadx = 10,anchor="w")
        console_frame.                  pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        column_labels.                  pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.              pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        self.make_project_widgets(initial = initial)
        Tools.add_colored_line(self.main_console,self.download_database_console_input[0],self.download_database_console_input[1],None,True)
        if self.show_tooltip == "ano":
            Catalogue_gui.ToolTip(export_button," Exporovat projekt ",self.root)
            Catalogue_gui.ToolTip(db_export," Exporovat projekt do databáze ",self.root)
            Catalogue_gui.ToolTip(button_settings," Nastavení ",self.root)
            Catalogue_gui.ToolTip(save_button," Uložit projekt ",self.root)
            Catalogue_gui.ToolTip(load_button," Nahrát projekt ",self.root)
            Catalogue_gui.ToolTip(db_login," Přihlásit do databáze ",self.root)

        def show_initial_context_menu(event):
            if len(self.station_list) == 0:   
                context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="#202020",activebackground="#606060",activeforeground="white")
                context_menu.add_command(label="Nová stanice",font=("Arial",22,"bold"),command=lambda: self.manage_widgets("","00",btn="add_line"))
                context_menu.tk_popup(event.x_root, event.y_root)

        self.root.bind("<Button-3>",lambda e: show_initial_context_menu(e))

        def unfocus_entry(e):
            self.root.focus_set()
        self.project_name_input.bind("<Leave>",lambda e:unfocus_entry(e))
        
        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1600:
                self.root.state('normal')
                self.root.geometry("1600x900")
            else:
                self.root.state('zoomed')

        self.root.bind("<f>",lambda e: maximalize_window(e))

        def unfocus_expanded(e):
            if self.leave_expanded_widget != None:
                x, y = self.root.winfo_pointerx(), self.root.winfo_pointery()
                hovered_widget = self.root.winfo_containing(x, y)
                if hovered_widget == self.leave_expanded_widget or str(hovered_widget) == str(self.leave_expanded_widget) + ".!label":
                    return
                try:
                    self.switch_widget_info(e,self.leave_expanded_widget_tier,self.leave_expanded_widget)
                    
                except Exception as e:
                    print("exception error: ",e)
            
            self.leave_expanded_widget = None
            self.leave_expanded_widget_tier = None
        self.root.bind("<Escape>",lambda e: unfocus_expanded(e))

    def make_project_widgets(self,initial = False,return_scroll = True):
        self.current_block_id = ""
        self.last_scroll_position = self.project_tree._parent_canvas.yview()[0]

        def upgrade_widget_heights(widget_list):
            """
            sets all widget heights accordingly to station frame
            """

            def get_max_height(camera_index):
                optic_list = widget_list[2][camera_index]
                optic_height = 0
                for optics in optic_list:
                    if self.render_mode == "precise":
                        optics.update_idletasks()
                    optic_height += optics._current_height

                acc_list = widget_list[4][camera_index]
                acc_height = 0
                for acc in acc_list:
                    if self.render_mode == "precise":
                        acc.update_idletasks()
                    acc_height += acc._current_height

                st_height = 0
                st = widget_list[0]
                if self.render_mode == "precise":
                    st.update_idletasks()
                st_height = st._current_height
                
                camera_height = 0
                camera = widget_list[1][camera_index]
                if self.render_mode == "precise":
                    camera.update_idletasks()
                camera_height = camera._current_height

                controller_height = 0
                if len(widget_list[3][x]) > 0: # controllers...
                    controller = widget_list[3][camera_index][0]
                    if self.render_mode == "precise":
                        controller.update_idletasks()
                    controller_height = controller._current_height
                
                max_height = max(optic_height,acc_height,camera_height,controller_height,st_height)
                if max_height <65:
                    max_height=65    

                return max_height

            all_heights = 0
            len_of_array = len(widget_list[1])
            if len_of_array > 0:
                for x in range(0,len(widget_list[1])):
                    camera_height = get_max_height(x)
                    all_heights += (camera_height+10)
                    if len(widget_list[3][x]) > 0: # controllers...
                        widget_list[3][x][0].configure(height=camera_height)
                    widget_list[1][x].configure(height=camera_height)

                    optic_list = widget_list[2][x]
                    len_of_array = len(optic_list)
                    if len_of_array>0:
                        second_height = camera_height/ len_of_array
                        for y in range(0,len(optic_list)):
                            if len(optic_list) == 1 or y == 0:
                                optic_list[y].configure(height=second_height)
                            else:
                                optic_list[y].configure(height=second_height-10)
                            
                    acc_list = widget_list[4][x]
                    len_of_array = len(acc_list)
                    if len_of_array>0:
                        second_height = camera_height/ len_of_array
                        for y in range(0,len(acc_list)):
                            if len(acc_list) == 1 or y == 0:
                                acc_list[y].configure(height=second_height)
                            else:
                                acc_list[y].configure(height=second_height-10)

            if all_heights < 65:
                widget_list[0].configure(height = 65)
            else:
                widget_list[0].configure(height = all_heights-10)

        def save_row_count(station_index):
            station_rows = 0
            for cameras in self.station_list[station_index]["camera_list"]:
                acc_count = 0
                if "controller_index" in cameras:
                    if cameras["controller_index"] != None and cameras["controller_index"] != "" and cameras["controller_index"] != "None":
                        acc_count = len(self.controller_object_list[int(cameras["controller_index"])]["accessory_list"]) 
                optics_count = len(cameras["optics_list"])
                cameras["row_count"] = max(acc_count,optics_count)

                if cameras["row_count"] == 0:
                    cameras["row_count"] = 1

                station_rows += cameras["row_count"]
            
            self.station_list[station_index]["row_count"] = station_rows

        if not initial:
            self.changes_made = True
        self.clear_frame(self.project_tree)
        default_height = 55

        # creating stations ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------        
        for i in range(0,len(self.station_list)):
            current_st_widget_list = []
            station_name = self.station_list[i]["name"]
            if i < 10:
                station_tier =  "0" + str(i) #01-99 
            else:
                station_tier =  str(i) #01-99

            station_camera_list = self.station_list[i]["camera_list"]
            camera_count = len(station_camera_list)
            station_frame = customtkinter.CTkFrame(master=self.project_tree,corner_radius=5,fg_color="#212121")
            station_frame.pack(pady=5,padx=0,side = "top",anchor = "w",expand = False)

            station_widget = self.make_block(master_widget=station_frame,height=default_height,width=self.default_block_width,fg_color="#181818",side = "left",text=station_name,tier=station_tier)
            if self.detailed_view:
                self.switch_widget_info("",station_tier,station_widget)

            # creating cameras ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            if camera_count == 0:
                dummy_cam =         self.make_block(master_widget=station_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
                dummy_opt =         self.make_block(master_widget=station_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
                dummy_controller =  self.make_block(master_widget=station_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
                dummy_acc =         self.make_block(master_widget=station_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)

            camera_widgets = []
            optics_widgets = []
            controllers_widgets = []
            accessory_widgets = []

            if camera_count > 1:
                main_camera_opt_frame = customtkinter.CTkFrame(master=station_frame,corner_radius=0,border_width=0,fg_color="#212121")
                main_camera_opt_frame.pack(pady=0,padx=0,side = "left",anchor="w",expand = False)

            for x in range(0,camera_count):
                if camera_count > 1:
                    camera_frame = customtkinter.CTkFrame(master=main_camera_opt_frame,corner_radius=0,border_width=0,fg_color="#212121") # frame with camera left and optics left and top
                    camera_frame.pack(pady=0,padx=0,side = "top",anchor = "w",expand = False)
                else:
                    camera_frame = customtkinter.CTkFrame(master=station_frame,corner_radius=0,border_width=0,fg_color="#212121") # frame with camera left and optics left and top
                    camera_frame.pack(pady=0,padx=0,side = "left",expand = False)

                camera_type = station_camera_list[x]["type"]
                try:
                    controller_index = int(station_camera_list[x]["controller_index"])
                    controller_color = self.controller_object_list[controller_index]["color"]
                    if controller_color == "": # nebyla zvolena žádná barva
                        controller_color = "#181818"
                except Exception: # pokud ke kamere neni prirazen zadny kontroler
                    controller_color = "#181818"
                    controller_index = None

                station_camera_optic_list = station_camera_list[x]["optics_list"]
                optic_count = len(station_camera_optic_list)
                if x < 10:
                    camera_tier =  station_tier + "0" + str(x) #0101-9999
                else:    
                    camera_tier =  station_tier + str(x) #0101-9999

                camera_widget = self.make_block(master_widget=camera_frame,height=default_height,width=self.default_block_width,fg_color=controller_color,side = "left",text=camera_type,tier = camera_tier,anchor="n")
                camera_widgets.append(camera_widget)
                if self.detailed_view:
                    self.switch_widget_info("",camera_tier,camera_widget)
                
                # creating optics for camera x------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
                if optic_count == 0:
                    dummy_opt = self.make_block(master_widget=camera_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
                camera_optics_widgets = []
                optic_frame_made =False
                for y in range(0,optic_count):
                    optic_type = station_camera_optic_list[y]["type"]
                    if y < 10:
                        optic_tier =  camera_tier + "0" + str(y) #010101-999999
                    else:
                        optic_tier =  camera_tier + str(y) #010101-999999
                    widget_border_color = "#636363"
                    if optic_type in self.whole_light_database and optic_type != "":
                        widget_border_color = "#FFFF00"
                        optic_type = "💡 "+ optic_type
                        self.station_list[i]["camera_list"][x]["optics_list"][y]["light_status"] = 1
                
                    if optic_count > 1 and optic_frame_made == False:
                        optics_frame = customtkinter.CTkFrame(master=camera_frame,corner_radius=0,border_width=0,fg_color="#212121")
                        optics_frame.pack(pady=0,padx=0,side = "left",anchor = "n",expand = False)
                        optic_frame_made=True

                    if optic_frame_made:
                        optic_widget = self.make_block(master_widget=optics_frame,height=default_height,width=self.default_block_width,fg_color="#181818",side = "top",text=optic_type,tier=optic_tier,border_color=widget_border_color,anchor="w")
                    else:
                        optic_widget = self.make_block(master_widget=camera_frame,height=default_height,width=self.default_block_width,fg_color="#181818",side = "left",text=optic_type,tier=optic_tier,border_color=widget_border_color,anchor="n")

                    if self.detailed_view:
                        self.switch_widget_info("",optic_tier,optic_widget)
                    camera_optics_widgets.append(optic_widget)

                controller_acc_widgets = []
                camera_controller_widgets = []
                if controller_index != None:
                    controller_frame = customtkinter.CTkFrame(master=camera_frame,corner_radius=0,border_width=0,fg_color="#212121") # frame with camera left and optics left and top
                    controller_frame.pack(pady=0,padx=0,side = "left",anchor = "n",expand = False)
                    if controller_index < 10:
                        controller_tier =  camera_tier + "c0" + str(controller_index) #xxxxc01-xxxxc99
                    else:
                        controller_tier =  camera_tier + "c" + str(controller_index) #xxxxc01-xxxxc99
                    self.controller_object_list[controller_index]["detailed_name"] = str(self.controller_object_list[controller_index]["name"]) + "(" + str(self.controller_object_list[controller_index]["type"]) + ")"
                    controller_color = self.controller_object_list[controller_index]["color"]
                    if controller_color == "": # nebyla zvolena žádná barva
                        controller_color = "#181818"

                    controller_widget = self.make_block(master_widget=controller_frame,height=default_height,width=self.default_block_width,fg_color=controller_color,
                                    side = "left",text=self.controller_object_list[controller_index]["type"],tier = controller_tier,anchor="n")

                    if self.detailed_view:
                        self.switch_widget_info("",controller_tier,controller_widget)
                    camera_controller_widgets.append(controller_widget)
                    accessory_list = self.controller_object_list[controller_index]["accessory_list"]
                    accessory_count = len(accessory_list)
                    for x in range(0,accessory_count):
                        accessory_type = accessory_list[x]["type"]
                        if x < 10:
                            accessory_tier =  controller_tier + "0" + str(x) #xxxxc0101-xxxxc9999
                        else:
                            accessory_tier =  controller_tier + str(x) #xxxxc0101-xxxxc9999

                        if accessory_count > 1:
                            accessory_widget = self.make_block(master_widget=controller_frame,height=default_height,width=self.default_block_width,fg_color="#181818",side = "top",text=accessory_type,tier = accessory_tier, anchor="w")
                        else:
                            accessory_widget = self.make_block(master_widget=controller_frame,height=default_height,width=self.default_block_width,fg_color="#181818",side = "left",text=accessory_type,tier = accessory_tier,anchor="w")
                        
                        if self.detailed_view:
                            self.switch_widget_info("",accessory_tier,accessory_widget)
                        controller_acc_widgets.append(accessory_widget)

                    if accessory_count == 0:
                        dummy_acc =     self.make_block(master_widget=controller_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)

                else:   
                    dummy_controller =  self.make_block(master_widget=camera_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
                    dummy_acc =         self.make_block(master_widget=camera_frame,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "left",text="",dummy_block=True)
            
                optics_widgets.append(camera_optics_widgets)
                accessory_widgets.append(controller_acc_widgets)
                controllers_widgets.append(camera_controller_widgets)
           
            current_st_widget_list.append(station_widget)
            current_st_widget_list.append(camera_widgets)
            current_st_widget_list.append(optics_widgets)
            current_st_widget_list.append(controllers_widgets)
            current_st_widget_list.append(accessory_widgets)
            upgrade_widget_heights(current_st_widget_list)
            save_row_count(station_index=i)
        if return_scroll:
            self.project_tree._parent_canvas.yview_moveto(self.last_scroll_position)

class Save_excel:
    def __init__(self,root,station_list,project_name,excel_name,controller_list,console):
        self.root = root
        self.main_console = console
        self.project_name = project_name
        self.station_list = station_list
        self.controller_list = controller_list
        self.values_start_row = 6
        self.excel_file_name = excel_name
        self.app_icon_path = Tools.resource_path('images\\logo_TRIMAZKON.ico')
        if self.excel_file_name == None:
            self.excel_file_name = "Katalog_kamerového_vybavení.xlsm"
        self.temp_excel_file_name = self.excel_file_name[:-5] + "_temp.xlsm"
        self.excel_rows_used = 0
        self.used_columns = ["A","B","C","D","E"]
        self.excel_column_width=50
        self.between_station_rows = []
        self.xlsx_format = False
        self.inventory_list = {"camera_list":[],
                               "optics_list":[],
                               "lights_list":[],
                               "controller_list":[],
                               "accessory_list":[],}

    def make_header(self,wb):
        ws = wb["Sheet"]
        if self.xlsx_format:
            ws["A"+str(self.values_start_row-1)] = "Stanice"
            ws["C"+str(self.values_start_row-1)] = "Kamera"
            ws["E"+str(self.values_start_row-1)] = "Optika/ světla"
            ws["G"+str(self.values_start_row-1)] = "Kontrolery"
            ws["I"+str(self.values_start_row-1)] = "Příslušenství"

        else:
            ws["A"+str(self.values_start_row-1)] = "Stanice"
            ws["B"+str(self.values_start_row-1)] = "Kamera"
            ws["C"+str(self.values_start_row-1)] = "Optika/ světla"
            ws["D"+str(self.values_start_row-1)] = "Kontrolery"
            ws["E"+str(self.values_start_row-1)] = "Příslušenství"

        ws["D"+str(1)] = "Číslo dokumentu:"
        ws["D"+str(2)] = "Verze dokumentu:"
        ws["E"+str(2)] = "AA"
        ws["D"+str(3)] = "Datum uvolnění:\n(dd.mm.rrrr)"

        image = Image(Tools.resource_path("images/jhv_logo2.png"))
        ws.add_image(image,"A1")
   
    def init_objects(self):
        """
        The excel_position and the hidden_values parameters of objects needs to be inited
        - only for objects containing more locations (mentioned above)
        - case of exporting again with some changes made
        """
        for controllers in self.controller_list:
            controllers["excel_position"] = []
            controllers["hidden_values"] = []
            for accessories in controllers["accessory_list"]:
                accessories["excel_position"] = []
                accessories["hidden_values"] = []

    def merge_cells(self,wb,merge_list:str):
        """
        cell range format: A1:A2
        """
        
        ws = wb.active
        for merge in merge_list:
            ws.merge_cells(merge)

    def update_sheet_vba_code(self,new_code):
        try:
            unsuccessfull = False
            app = xw.App(visible=False)
            wb = app.books.open(self.temp_excel_file_name)
            vb_project = wb.api.VBProject
            # vb_project.VBComponents.Add(1) # musi se pridat prazdny modul...
            code_module = vb_project.VBComponents("ThisWorkbook").CodeModule
            code_module.DeleteLines(1, code_module.CountOfLines)
            code_module.AddFromString(new_code)
            try:
                wb.save(self.excel_file_name)
            except Exception:
                unsuccessfull = True
            wb.close()
            app.quit()

            if os.path.exists(self.temp_excel_file_name): # nutná operace (vyuzivat temp soubor) kvůli zapisování vba
                os.remove(self.temp_excel_file_name)
            
            if unsuccessfull:
                return False
        except Exception as e:
            print("chyba: nejsou povolena práva na makra")
            wb.close()
            app.quit()
            if os.path.exists(self.temp_excel_file_name): # nutná operace (vyuzivat temp soubor) kvůli zapisování vba
                os.remove(self.temp_excel_file_name)
            return "rights_error"

    def get_cells_to_merge(self):
        """
        skládá se pole, které buňky je potřeba spojit\n
        + ukládají se pozice pro danou hodnotu v excelu k danému objektu pod key argument: excel position
        """
        last_row = self.values_start_row
        last_row_cam = self.values_start_row
        last_row_optics = self.values_start_row
        last_row_accessory = self.values_start_row
        rows_to_merge = []
        columns = ["A","B","C","D","E"]
        if self.xlsx_format:
            columns = ["A","C","E","G","I"]

        def check_for_dummy(last_cam=False):
            #fill the rest with dummies:
            addition = 0
            number_of_dummy = 0

            if last_cam:
                addition = 1
            if int(cameras["row_count"]) > len(cameras["optics_list"]):
                cam_row_count = (int(cameras["row_count"]) - len(cameras["optics_list"])) + addition
                for _ in range(0,cam_row_count):
                    dummy_opt = {
                        "type": "",
                        "alternative": "",
                        "excel_position": columns[2]+str(last_row_optics),
                        "description": ""
                    }
                    self.station_list[station_index]["camera_list"][camera_index]["optics_list"].append(dummy_opt)
                    number_of_dummy +=1
            return number_of_dummy

        for stations in self.station_list:
            station_index = self.station_list.index(stations)
            if stations["row_count"] > 1:
                self.station_list[station_index]["excel_position"] = columns[0]+str(last_row)
                rows_to_merge.append(columns[0] + str(last_row) + ":" + columns[0] + str(last_row + int(stations["row_count"]) - 1))
                last_row = last_row + (stations["row_count"])
            else:
                self.station_list[station_index]["excel_position"] = columns[0]+str(last_row)
                last_row = last_row + 1

            if len(stations["camera_list"]) == 0:
                last_row_cam = last_row_cam + 1
                last_row_optics = last_row_optics + 1
            cam_inc = 0
            for cameras in stations["camera_list"]:
                camera_index = self.station_list[station_index]["camera_list"].index(cameras)
                row_before_addition = last_row_cam
                if int(cameras["row_count"]) > 1:
                    self.station_list[station_index]["camera_list"][camera_index]["excel_position"] = columns[1]+str(last_row_cam)
                    rows_to_merge.append(columns[1] + str(last_row_cam) + ":"+columns[1] + str(last_row_cam + int(cameras["row_count"]) - 1))
                    # kontrolery maji stejný merge, pocet vsech radku ulozen v kamere (i kdyz je vetsi pocet prislusenstvi nez objektivu ke kamere)
                    rows_to_merge.append(columns[3] + str(last_row_cam) + ":"+columns[3] + str(last_row_cam + int(cameras["row_count"]) - 1))
                    last_row_cam = last_row_cam + (cameras["row_count"])
                else:
                    self.station_list[station_index]["camera_list"][camera_index]["excel_position"] = columns[1]+str(last_row_cam)
                    last_row_cam = last_row_cam + 1

                if cameras["controller"] != "":
                    ii = 0
                    for controllers in self.controller_list:
                        if str((controllers["name"]+"("+controllers["type"]+")")).replace(" ","") == str(cameras["controller"]).replace(" ",""):
                            controller_index = ii
                            # stejny kontroler muze byt soucasti vice kamer, proto pole s excel_position
                            try:
                                if not str(columns[3]+str(row_before_addition)) in self.controller_list[controller_index]["excel_position"]:
                                    self.controller_list[controller_index]["excel_position"].append(str(columns[3]+str(row_before_addition)))

                            except Exception as e:
                                self.controller_list[controller_index]["excel_position"] = [(columns[3]+str(row_before_addition))]

                            acc_count = len(controllers["accessory_list"])
                            if acc_count == 0:
                                last_row_accessory = last_row_accessory + 1
                            iii = 0
                            for accessories in controllers["accessory_list"]:
                                # nejprve hledam index abych mohl prepisovat přímo celý objekt
                                accessory_index = iii
                                try:
                                    if not (columns[4]+str(row_before_addition+iii)) in self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"]:
                                        self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"].append(columns[4]+str(row_before_addition+iii))
                                except Exception:
                                    self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"] = [columns[4]+str(row_before_addition+iii)]
                                last_row_accessory = last_row_accessory + 1

                                iii+=1
                            
                            dummy_start_row = row_before_addition+iii
                            acc_dummy_count = 0
                            if acc_count > 0:
                                acc_dummy_count = max(acc_count-int(cameras["row_count"]),int(cameras["row_count"])-acc_count)
                            
                            if acc_dummy_count>0:
                                rows_to_merge.append(columns[4] + str(dummy_start_row-1) + ":"+columns[4] + str(dummy_start_row-1 + acc_dummy_count))
                            break
                        
                        ii += 1
                        
                if len(cameras["optics_list"]) == 0:
                    last_row_optics = last_row_optics + 1
                optics_count_no_dummy = len(cameras["optics_list"])
                dummy_count = check_for_dummy()
                if last_row_optics != (last_row_optics + dummy_count):
                    rows_to_merge.append(columns[2] + str(last_row_optics+optics_count_no_dummy-1) + ":" + columns[2] + str(last_row_optics+optics_count_no_dummy-1+dummy_count))

                for optics in cameras["optics_list"]:
                    optics_index = self.station_list[station_index]["camera_list"][camera_index]["optics_list"].index(optics)
                    self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["excel_position"] = columns[2]+str(last_row_optics)
                    last_row_optics = last_row_optics + 1
                cam_inc+=1
                
                
            self.between_station_rows.append(last_row_cam)
            #radek mezera mezi kazdou stanici
            last_row+=1
            last_row_cam+=1
            last_row_optics = last_row_optics + 1

        self.excel_rows_used = last_row_optics
        print("LAST ROW USED: ",self.excel_rows_used)

        if self.xlsx_format:
            columns = ["A","C","E","G"]
            for merges in rows_to_merge:
                if columns[0] in merges:
                    rows_to_merge.append(merges.replace(columns[0] ,"B"))
                elif columns[1] in merges:
                    rows_to_merge.append(merges.replace(columns[1] ,"D"))
                elif columns[2] in merges:
                    rows_to_merge.append(merges.replace(columns[2] ,"F"))
                elif columns[3] in merges:
                    rows_to_merge.append(merges.replace(columns[3] ,"H"))
            line = self.values_start_row -1
            rows_to_merge.append(f"A{line}:B{line}")
            rows_to_merge.append(f"C{line}:D{line}")
            rows_to_merge.append(f"E{line}:F{line}")
            rows_to_merge.append(f"G{line}:H{line}")
            rows_to_merge.append(f"I{line}:J{line}")

        # grafika header:
        rows_to_merge.append("A1:A3")
        rows_to_merge.append("B1:C3")

        # if self.xlsx_format:
            # rows_to_merge.append("B1:J3")
            # rows_to_merge.append("B2:J2")
        # else:
            # rows_to_merge.append("B2:E2")

        return rows_to_merge

    def change_vba_script(self):
        """
        Slouží pro přidávání rozsahu hodnot, uložených v hidden sheetu a alokování k určité buňce
        - současná kapacita pro jeden objekt: 78 (3* abeceda)
        """
        vba_code_range = """"""
        alphabet = string.ascii_uppercase
        i = 0
        ii = 0
        iii = 0
        columns = ["A","B","C","D","E","F","G","H","I"]
        column_letter_st = 0
        column_letter_cam = 3
        column_letter_opt = 6

        for stations in self.station_list:
            cell_with_toggle = stations["excel_position"]
            if i > 25:
                column_letter_st +=1
                i=0
            column = columns[column_letter_st] + alphabet[i:i+1] 
            stations["hidden_values"] = column # pridame jen informaci o nazvu sloupce
            station_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", \"{column + str(4)}\", Cancel, Target"
            vba_code_range += "\n            "+station_vba_code_range_row
            i+=1

            for cameras in stations["camera_list"]:
                cell_with_toggle = cameras["excel_position"]
                if ii > 25:
                    column_letter_cam +=1
                    ii=0
                column = columns[column_letter_cam] + alphabet[ii:ii+1] 
                cameras["hidden_values"] = column # pridame jen informaci o nazvu sloupce
                camera_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", \"{column + str(4)}\", Cancel, Target"
                vba_code_range += "\n            "+camera_vba_code_range_row
                ii+=1

                for optics in cameras["optics_list"]:
                    cell_with_toggle = optics["excel_position"]
                    if iii > 25:
                        column_letter_opt += 1
                        iii=0
                    column = columns[column_letter_opt] + alphabet[iii:iii+1]
                    optics["hidden_values"] = column # pridame jen informaci o nazvu sloupce
                    optics_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", \"{column + str(4)}\", Cancel, Target"
                    vba_code_range += "\n            "+optics_vba_code_range_row
                    iii+=1
        i = 0
        ii = 0
        columns = ["J","K","L","M","N","O","P","Q"]
        column_letter_controller = 0
        column_letter_acc = 3
        for controllers in self.controller_list:
            try:
                for controller_positions in controllers["excel_position"]: 
                    cell_with_toggle = controller_positions
                    if i > 25:
                        column_letter_controller +=1
                        i=0
                    column = columns[column_letter_controller] + alphabet[i:i+1] 
                    try:
                        controllers["hidden_values"].append(column) # pridame jen informaci o nazvu sloupce
                    except Exception:
                        controllers["hidden_values"] = [column]
                        
                    controller_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", \"{column + str(4)}\", Cancel, Target"
                    vba_code_range += "\n            "+controller_vba_code_range_row
                    i+=1
                    for accessories in controllers["accessory_list"]:
                        for acc_positions in accessories["excel_position"]:
                            cell_with_toggle = acc_positions
                            if ii > 25:
                                column_letter_acc +=1
                                ii=0
                            column = columns[column_letter_acc] + alphabet[ii:ii+1]
                            try:
                                accessories["hidden_values"].append(column) # pridame jen informaci o nazvu sloupce
                            except Exception:
                                accessories["hidden_values"] = [column]
                            acc_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", \"{column + str(4)}\", Cancel, Target"
                            vba_code_range += "\n            "+acc_vba_code_range_row
                            ii+=1
            except Exception: # the station with this controller was deleted
                pass

        vba_code = f"""
        Private Sub Workbook_SheetBeforeRightClick(ByVal Sh As Object, ByVal Target As Range, Cancel As Boolean)
            {vba_code_range}
        End Sub

        Private Sub ToggleCell(ByVal targetCell As Range, ByVal text1Ref As String, ByVal text2Ref As String, ByVal toggleStatusRef As String,ByVal rowHeightRef as String, ByRef Cancel As Boolean, ByVal clickedCell As Range)
            ' Read text values from hidden worksheet
            Dim text1 As String
            Dim text2 As String
            text1 = Worksheets("HiddenSheet").Range(text1Ref).Value
            text2 = Worksheets("HiddenSheet").Range(text2Ref).Value
            ActiveSheet.Unprotect

            ' Read toggle status from hidden worksheet
            Dim toggle_status As Integer
            toggle_status = Worksheets("HiddenSheet").Range(toggleStatusRef).Value

            Dim row_height As Integer
            row_height = Worksheets("HiddenSheet").Range(rowHeightRef).Value

            ' Check if the right-clicked cell is the target cell
            If Not Intersect(clickedCell, targetCell) Is Nothing Then
                ' Toggle the cell value
                If toggle_status = 1 Then
                    Worksheets("HiddenSheet").Range(text1Ref).Value = targetCell.Value
                    targetCell.Value = text2
                    toggle_status = 0

                    If targetCell.Height < row_height Then
                        targetCell.RowHeight = row_height
                    End If

                Else
                    Worksheets("HiddenSheet").Range(text2Ref).Value = targetCell.Value
                    targetCell.Value = text1
                    toggle_status = 1

                    On Error GoTo ErrorHandler ' Start error handling
                    targetCell.Rows.AutoFit

                End If

                ' Update toggle status on hidden worksheet
                Worksheets("HiddenSheet").Range(toggleStatusRef).Value = toggle_status
                ' Cancel the default right-click menu
                Cancel = True
            End If
            Exit Sub

        ErrorHandler: ' Define what to do if an error occurs
            MsgBox "An error occurred: " & Err.Description, vbExclamation
            Exit Sub

        End Sub
        """
        return vba_code

    def format_cells(self,ws):
        bold_font_white = Font(bold=True,size=20,color="ffffff") # ffffff = bílá!
        bold_font = Font(bold=True,size=12)
        regular_font = Font(bold=False,size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 35
        ws.row_dimensions[3].height = 35

        # Nadpis
        top_header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        if self.project_name == None or self.project_name == "":
            ws["B1"] = "Přehled kamerového vybavení"
        else:
            ws["B1"] = f"Přehled kamerového vybavení\nprojektu: {self.project_name}"
        ws["B1"].alignment = Alignment(horizontal = "center", vertical = "center",wrap_text=True)
        ws["B1"].font = Font(bold=True,size=25)
        ws["B1"].fill = top_header_fill
        # doplnujici info v hlavičce:
        info_letter1 = "D"
        info_letter2 = "E"
        for i in range(0,3):
            ws[info_letter1+str(i+1)].fill = top_header_fill
            ws[info_letter1+str(i+1)].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True,shrink_to_fit=True,justifyLastLine=True)
            ws[info_letter1+str(i+1)].font = bold_font
            ws[info_letter2+str(i+1)].fill = top_header_fill
            ws[info_letter2+str(i+1)].alignment = Alignment(horizontal = "center", vertical = "center",wrap_text=True,shrink_to_fit=True,justifyLastLine=True)
            ws[info_letter2+str(i+1)].font = regular_font

            ws["A"+str(i+1)].border = thin_border
            ws["B"+str(i+1)].border = thin_border
            ws["C"+str(i+1)].border = thin_border
            ws[info_letter1+str(i+1)].border = thin_border
            ws[info_letter2+str(i+1)].border = thin_border

        comment_text = "Pravým klikem na buňky v tabulce zobrazíte podrobnosti"
        comment_author = "TRIMAZKON"
        comment = Comment(comment_text, comment_author)
        if not self.xlsx_format:
            ws['B1'].comment = comment
        
        current_date = datetime.now().date()
        date_string = current_date.strftime("%d.%m.%Y")
        ws[info_letter2 + "3"] = date_string
        header_fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
        
        for columns in self.used_columns:
            ws.column_dimensions[columns].width = self.excel_column_width
            for i in range((self.values_start_row-1),self.excel_rows_used-1): # formát všech zaplněných buněk
                cell = ws[columns + str(i)]
                # if self.xlsx_format:
                #   cell.alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True,shrink_to_fit=True,justifyLastLine=True)
                # else:
                cell.alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)
                cell.border = thin_border

                if i == (self.values_start_row-1): # nadpisy sloupců
                    cell.font = bold_font_white
                    cell.alignment = Alignment(horizontal = "center", vertical = "center")
                    cell.fill = header_fill
                else:
                    cell.font = regular_font
                
                # Názvy stanic:
                if columns == "A" and i != (self.values_start_row-1):
                    cell.font = bold_font

        # fill the empty rows between stations:
        fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
        for rows in self.between_station_rows:
            for columns in self.used_columns:
                if self.xlsx_format:
                    cell = ws[columns + str(rows)]
                    cell.fill = fill
                else:
                    cell = ws[columns + str(rows)]
                    # fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
                    cell.fill = fill

    def fill_values(self,wb):
        """
        vepíše hodnoty
        - xlsx první sloupec - v gui prvně viditelné informace
        - xlsm první informace
        """
        
        def write_to_inventory(device_list,device):
            """
            - camera_list:
            - optics_list:
            - lights_list:
            - controller_list
            - accessory_list
            """
            if device == "" or device == None:
                return
            item_found = False
            for items in self.inventory_list[str(device_list)]:
                if items["name"] == device:
                    items["count"] +=1
                    item_found = True
                    break

            if not item_found:
                self.inventory_list[str(device_list)].append({"name": device,"count":1})
        
        light_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        ws = wb.active
        columns = ["B","C","D","E"]
        if self.xlsx_format:
            columns = ["C","E","G","I"]

        for stations in self.station_list:
            excel_cell = stations["excel_position"]
            ws[excel_cell] = stations["name"]

            if len(stations["camera_list"]) == 0:
                excel_cell = columns[0] + stations["excel_position"][1:]
                ws[excel_cell] = ""
            for cameras in stations["camera_list"]:
                write_to_inventory("camera_list",cameras["type"])
                excel_cell = cameras["excel_position"]
                ws[excel_cell] = cameras["type"]
                if str(cameras["controller_color"]) != "":
                    try:
                        color_modified = str(cameras["controller_color"])[1:]
                        controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                        ws[excel_cell].fill = controller_fill
                    except Exception as e:
                        print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                        pass
                
                if len(cameras["optics_list"]) == 0:
                    excel_cell = columns[1] + cameras["excel_position"][1:]
                    ws[excel_cell] = ""
                for optics in cameras["optics_list"]:
                    excel_cell = optics["excel_position"]
                    try:
                        ws[excel_cell] = optics["type"]
                        if "light_status" in optics:
                            ws[excel_cell].fill = light_fill
                            write_to_inventory("lights_list",optics["type"])
                        else:
                            write_to_inventory("optics_list",optics["type"])

                    except AttributeError:
                        pass

        for controllers in self.controller_list:
            write_to_inventory("controller_list",controllers["type"])
            for acc in controllers["accessory_list"]:
                write_to_inventory("accessory_list",acc["type"])
            try:
                for position in controllers["excel_position"]:
                    excel_cell = str(position)
                    ws[excel_cell] = controllers["type"]
                    if str(controllers["color"]) != "" and str(controllers["color"]) != "#212121":
                        try:
                            color_modified = str(controllers["color"])[1:] # without hashtag
                            controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                            ws[excel_cell].fill = controller_fill
                        except Exception as e:
                            print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                            pass

                    if len(controllers["accessory_list"]) == 0:
                        excel_cell = columns[3] + position[1:]
                        ws[excel_cell] = ""
                    for accessories in controllers["accessory_list"]:
                        for acc_positions in accessories["excel_position"]:
                            # print("accessories excel position: ",accessories["excel_position"])
                            excel_cell = acc_positions
                            ws[excel_cell] = accessories["type"]
            except Exception: # the station with this controller was deleted
                pass       
        self.format_cells(ws)

    def fill_hidden_sheet_values(self,wb):
        """
        Provede vytvoření skrytého listu, kam ukládá toggle hodnoty a aktuální stav přepnutí\n
        Rozdělení:
        - Vždy tři hodnoty
            - toggle první hodnota (název/ typ)
            - toggle druhá hodnota (doplňující informace)
            - stav togglu (přepnutí 0-1)
        - stanice: AZ - CZ
        - kamery: DZ - FZ
        - optika: GZ - IZ
        - kontrolery: JZ - LZ
        - příslušenství: MZ - OZ
        """
        def get_string_rows(input_string):
            rows_splitted = []
            rows_splitted = input_string.split("\n")
            #potreba ty prazdna mista ponechat, kdyz je vlozeno hodne odsazení
            # cleaned_data = [x for x in rows_splitted if x]
            return len(rows_splitted)

        def calculate_new_cell_height(max_rows):
            height_of_one_row = 15
            if max_rows == 0:
                return height_of_one_row
            else:
                return max_rows*height_of_one_row
            
        ws = wb.create_sheet("HiddenSheet")
        ws.sheet_state = 'hidden'
        for stations in self.station_list:
            excel_cell = stations["hidden_values"]
            ws[excel_cell + str(1)] = str(stations["name"])
            ws[excel_cell + str(2)] = str(stations["inspection_description"])
            station_number_of_rows = get_string_rows(str(stations["inspection_description"]))
            ws[excel_cell + str(3)] = 1 # toggle status... default: 1
            new_cell_height = calculate_new_cell_height(station_number_of_rows)
            ws[excel_cell + str(4)] = new_cell_height

            for cameras in stations["camera_list"]:
                excel_cell = cameras["hidden_values"]
                ws[excel_cell + str(1)] = cameras["type"]

                detail_info_cam = Fill_details.camera(cameras)
                ws[excel_cell + str(2)] = detail_info_cam[0]
                ws[excel_cell + str(3)] = 1
                camera_rows = get_string_rows(str(detail_info_cam[0]))
                new_cell_height = calculate_new_cell_height(camera_rows)
                ws[excel_cell + str(4)] = new_cell_height
                
                for optics in cameras["optics_list"]:
                    excel_cell = optics["hidden_values"]
                    ws[excel_cell + str(1)] = optics["type"]
                    detail_info_opt = Fill_details.optics(optics)
                    ws[excel_cell + str(2)] = detail_info_opt
                    ws[excel_cell + str(3)] = 1
                    optic_rows = get_string_rows(str(detail_info_opt))
                    new_cell_height = calculate_new_cell_height(optic_rows)
                    ws[excel_cell + str(4)] = new_cell_height

        for controllers in self.controller_list:
            detail_info_cont = Fill_details.controller(controllers)
            controller_num_of_rows = get_string_rows(detail_info_cont)
            new_cont_cell_height = calculate_new_cell_height(controller_num_of_rows)

            for controller_positions in controllers["hidden_values"]:
                excel_cell = controller_positions
                ws[excel_cell + str(1)] = controllers["type"]
                ws[excel_cell + str(2)] = detail_info_cont
                ws[excel_cell + str(3)] = 1 # toggle status... default: 1
                ws[excel_cell + str(4)] = new_cont_cell_height

                for accessories in controllers["accessory_list"]:
                    detail_info_acc = Fill_details.accessory(accessories)
                    acc_rows = get_string_rows(detail_info_acc)
                    new_acc_cell_height = calculate_new_cell_height(acc_rows)

                    for acc_positions in accessories["hidden_values"]:
                        excel_cell = acc_positions
                        ws[excel_cell + str(1)] = accessories["type"]
                        ws[excel_cell + str(2)] = detail_info_acc
                        ws[excel_cell + str(3)] = 1 # toggle status... default: 1
                        ws[excel_cell + str(4)] = new_acc_cell_height

    def fill_xlsx_column(self,wb):
        def get_string_rows(input_string):
            rows_splitted = []
            rows_splitted = input_string.split("\n")
            # cleaned_data = [x for x in rows_splitted if x]
            return len(rows_splitted)

        def calculate_new_cell_height(max_rows,line_to_be_expanded:int):
            height_of_one_row = 15
            if max_rows == 0:
                return
            try:
                current_cell_height = ws.row_dimensions[line_to_be_expanded].height
                if current_cell_height == None:
                    ws.row_dimensions[line_to_be_expanded].height = max_rows*height_of_one_row

                elif int(current_cell_height) < max_rows*height_of_one_row:
                    ws.row_dimensions[line_to_be_expanded].height = max_rows*height_of_one_row
            except Exception as e:
                print(e)
                
        ws = wb.active
        columns = ["D","F","H","J"]
        for stations in self.station_list:
            excel_cell = str(stations["excel_position"])
            excel_cell = excel_cell.replace("A","B")
            station_cell = str(excel_cell)
            ws[excel_cell] = stations["inspection_description"]
            station_number_of_rows = get_string_rows(str(stations["inspection_description"]))
            # calculate_new_cell_height(station_number_of_rows,int(station_cell[1:]))

            if len(stations["camera_list"]) == 0:
                excel_cell = columns[0] + stations["excel_position"][1:]
                ws[excel_cell] = ""

            camera_num_of_rows = 0
            optics_num_of_rows = 0
            for cameras in stations["camera_list"]:
                excel_cell = cameras["excel_position"]
                excel_cell = excel_cell.replace("C","D")
                detail_info_cam = Fill_details.camera(cameras)
                if detail_info_cam[1] != None:
                    ws[excel_cell].fill = detail_info_cam[1]
                ws[excel_cell] = detail_info_cam[0]
                camera_rows = get_string_rows(detail_info_cam[0])
                # calculate_new_cell_height(camera_rows,int(excel_cell[1:]))

                if len(cameras["optics_list"]) == 0:
                    excel_cell = columns[1] + cameras["excel_position"][1:]
                    ws[excel_cell] = ""

                for optics in cameras["optics_list"]:
                    excel_cell = optics["excel_position"]
                    excel_cell = excel_cell.replace("E","F")
                    detail_info_opt = Fill_details.optics(optics)
                    try:
                        ws[excel_cell] = str(detail_info_opt)
                    except Exception as e:
                        pass
                    optic_rows = get_string_rows(detail_info_opt)
                    # calculate_new_cell_height(optic_rows,int(excel_cell[1:]))

                    optics_num_of_rows += optic_rows
                camera_num_of_rows += camera_rows

            max_rows = max(station_number_of_rows,camera_num_of_rows,optics_num_of_rows)
            calculate_new_cell_height(max_rows,int(station_cell[1:]))

        for controllers in self.controller_list:
            controller_num_of_rows = 0
            acc_num_of_rows = 0
            acc_rows_received = False
            detail_info_cont = Fill_details.controller(controllers)
            controller_num_of_rows = get_string_rows(detail_info_cont)

            for position in controllers["excel_position"]:
                excel_cell = str(position)
                excel_cell = excel_cell.replace("G","H")
                ws[excel_cell] = detail_info_cont

                if len(controllers["accessory_list"]) == 0:
                    excel_cell = columns[3] + position[1:]
                    ws[excel_cell] = ""
                for accessory in controllers["accessory_list"]:
                    detail_info_acc = Fill_details.accessory(accessory)
                    if not acc_rows_received:
                        acc_rows = get_string_rows(detail_info_acc)
                        acc_num_of_rows += acc_rows

                    for acc_positions in accessory["excel_position"]:
                        excel_cell = acc_positions
                        excel_cell = excel_cell.replace("I","J")
                        ws[excel_cell] = detail_info_acc
                acc_rows_received = True
                max_rows = max(controller_num_of_rows,acc_num_of_rows)
                calculate_new_cell_height(max_rows,int(position[1:]))
    
    def fill_images(self,wb):
        for station in self.station_list:
            if "image_list" in station:
                ws = wb.create_sheet(str(station["name"])+" - foto")
                num_of_images = 0
                for image_paths in station["image_list"]:
                    num_of_images +=1
                    try:
                        if image_paths.endswith("/"):
                            image_paths = image_paths[:-1]
                        image = Image(image_paths)
                        if num_of_images > 1:
                            ws.add_image(image,"A"+str(num_of_images*10))
                        else:
                            ws.add_image(image,"A"+str(num_of_images))
                    except Exception as e:
                        print(f"Obrázek {image_paths} se nepodařilo exportovat. {e}")

    def create_inventory(self,wb):
        fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
        light_gray = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        bold_font_white = Font(bold=True,size=20,color="ffffff") # ffffff = bílá!
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws = wb.create_sheet("Kusovník")
        ws["A"+str(1)] = "Kamery"
        ws["D"+str(1)] = "Objektivy"
        ws["G"+str(1)] = "Světla"
        ws["J"+str(1)] = "Kontrolery"
        ws["M"+str(1)] = "Příslušenství"
        device_columns = ["A","D","G","J","M"]
        count_columns = ["B","E","H","K","N"]
        for i in range(0,len(device_columns)):
            ws.merge_cells(device_columns[i] + "1:" + count_columns[i]+"1")

        column_increment = 0
        list_names = ["camera_list","optics_list","lights_list","controller_list","accessory_list"]
        for items in list_names:
            row_increment = 3
            for subitems in self.inventory_list[items]:
                ws[str(device_columns[column_increment])+str(row_increment)] = subitems["name"]
                ws[str(device_columns[column_increment])+str(row_increment)].border = thin_border
                ws[str(count_columns[column_increment])+str(row_increment)] = subitems["count"]
                ws[str(count_columns[column_increment])+str(row_increment)].border = thin_border
                row_increment+=1
            column_increment+=1

        for columns in count_columns:
            ws[columns+"1"].border = thin_border
            ws[columns+"2"].border = thin_border
            ws[columns+"2"] = "počet [ks]"
            ws[columns+"2"].fill = light_gray
            ws[columns+"2"].alignment = Alignment(horizontal = "center", vertical = "center")

        for columns in device_columns:
            ws.column_dimensions[columns].width = 30
            cell = ws[columns+"1"]
            cell.border = thin_border
            cell.font = bold_font_white
            cell.alignment = Alignment(horizontal = "center", vertical = "center")
            cell.fill = fill

            ws[columns+"2"] = "typ"
            ws[columns+"2"].border = thin_border
            ws[columns+"2"].fill = light_gray
            ws[columns+"2"].alignment = Alignment(horizontal = "center", vertical = "center")

    def main(self):
        wb = Workbook() #vytvorit novy excel, prepsat...
        if ".xlsm" in self.excel_file_name:
            self.init_objects()
            rows_to_merge = self.get_cells_to_merge()
            self.make_header(wb)
            self.fill_images(wb)
            # try:
            if os.path.exists(self.temp_excel_file_name):
                os.remove(self.temp_excel_file_name)
            # kličky aby se uložilo vba:
            wb.save(filename=self.temp_excel_file_name)
            wb2 = load_workbook(filename=self.temp_excel_file_name, keep_vba=True)
            wb2.save(self.temp_excel_file_name)
            wb.close()
            wb2.close()
            wb = load_workbook(filename=self.temp_excel_file_name, read_only=False, keep_vba=True)
            self.merge_cells(wb,merge_list=rows_to_merge)
            self.fill_values(wb)
            new_vba_code = self.change_vba_script()
            self.fill_hidden_sheet_values(wb)
            self.create_inventory(wb)
            wb.save(self.temp_excel_file_name)
            wb.close()
            attempt = self.update_sheet_vba_code(new_code=new_vba_code)
            if attempt == False:
                error_message1 = f"Nejprve prosím zavřete soubor {self.excel_file_name}"
                Tools.add_colored_line(self.main_console,f"Nejprve prosím zavřete soubor {self.excel_file_name}","red",None,True)
                return error_message1
            elif attempt == "rights_error":
                error_message2 = f"Nemáte nastavená potřebná práva v excelu pro makra (VBA)"
                Tools.add_colored_line(self.main_console,f"Nemáte nastavená potřebná práva v excelu pro makra (VBA)","red",None,True)
                self.opened_window = ToplevelWindow.excel_manual_window(self.root,self.app_icon_path)
                return error_message2
            else:
                Tools.add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně exportován","green",None,True)
                os.startfile(self.excel_file_name)
                return True
                
        elif ".xlsx" in self.excel_file_name:
            self.used_columns = ["A","B","C","D","E","F","G","H","I","J"]
            self.xlsx_format = True
            self.init_objects()
            rows_to_merge = self.get_cells_to_merge()
            self.make_header(wb)
            self.merge_cells(wb,merge_list=rows_to_merge)
            # try:

            self.fill_values(wb)
            self.create_inventory(wb)
            self.fill_xlsx_column(wb)
            self.fill_images(wb)
            wb.save(self.excel_file_name)
            wb.close()
            Tools.add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně exportován","green",None,True)
            os.startfile(self.excel_file_name)
            return True

            # except Exception as e:
            #     error_message = f"Nejprve prosím zavřete soubor {self.excel_file_name}, chyba: {e}"
            #     Tools.add_colored_line(self.main_console,f"Nejprve prosím zavřete soubor {self.excel_file_name}, chyba: {e}","red",None,True)
            #     return error_message # returns the failure information to main gui
                # wb.close()

# download = download_database.database(database_filename)
# Catalogue_gui(root,download.output)
if testing:
    # Catalogue_gui(root,"testing - stahování vypnuto","","max",database_filename,"excel_testing","xml_testing",0,"xlsx","","fast")
    Catalogue_gui(root,"testing - stahování vypnuto","","max",r"C:/Users/jakub.hlavacek.local/Desktop/JHV/Work/TRIMAZKON/")
    root.mainloop()