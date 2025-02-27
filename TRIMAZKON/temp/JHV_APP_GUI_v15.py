import customtkinter
import os
import time
from openpyxl import load_workbook
from PIL import Image, ImageTk
import Sorting_option_v5 as Trideni
import Deleting_option_v1 as Deleting
import Converting_option_v3 as Converting
import catalogue_maker_v5 as Catalogue
import sharepoint_download as download_database
import IP_setting_v4 as IP_setting
import trimazkon_tray_v2 as trimazkon_tray
import string_database
from tkinter import filedialog
import tkinter as tk
import threading
import shutil
import sys
import ctypes
import win32pipe, win32file, pywintypes, psutil
import subprocess
from win32api import *
from win32gui import *
import win32con
import struct

testing = False

trimazkon_tray_exe_name = "trimazkon_tray_v2.exe"
global_recources_load_error = False
exe_path = sys.executable
exe_name = os.path.basename(exe_path)
print("exe name: ",exe_name)
if testing:
    exe_name = "trimazkon_test.exe"

class Tools:
    @classmethod
    def path_check(cls,path_raw,only_repair = None):
        path=path_raw
        backslash = "\\"
        if backslash[0] in path:
            newPath = path.replace(backslash[0], '/')
            path = newPath
        if path.endswith('/') == False:
            newPath = path + "/"
            path = newPath
        #oprava mezery v nazvu
        path = r"{}".format(path)
        if not os.path.exists(path) and only_repair == None:
            return False
        else:
            return path

    @classmethod
    def get_all_app_processes(cls):
        pid_list = []
        num_of_apps = 0
        for process in psutil.process_iter(['pid', 'name']):
            # if process.info['name'] == "TRIMAZKON_test.exe":
            if process.info['name'] == exe_name:
                pid_list.append(process.info['pid'])
                num_of_apps+=1
        
        return [num_of_apps,pid_list]

    @classmethod
    def check_runing_app_duplicity(cls):
        """
        Spočte procesy a názvem aplikace, pokud je jich více, jak 2 je již aplikace spuštěná
        - v top případě zajistí aby se nenačítalo gui a pouze zajistí odeslání paramterů pro image browser
        """
        found_processes = Tools.get_all_app_processes()
        if found_processes[0] > 2:
            return True
        else:
            return False

    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)
    
    @classmethod
    def read_config_data(cls): # Funkce vraci data z config_TRIMAZKON.
        """
        Funkce vrací data z konfiguračního souboru config_TRIMAZKON.xlsx

        data jsou v pořadí:

        0 supported_formats_sorting\n
        1 supported_formats_deleting\n
        2 path_repaired\n
        3 files_to_keep\n
        4 cutoff_date\n
        5 prefix_function\n
        6 prefix_camera\n
        7 maximalized\n
        8 max_pallets\n
        9 static_dirs_names\n
        10 sorting_safe_mode\n
        11 image_browser_setting [checkbox, increment, movement]\n
        12 show_changelog\n
        13 image_film\n
        14 num_of_IB_film_images\n
        15 default sharepoint database filename\n
        16 default excel filnename prefix for catalogue output\n
        17 default xml filename for catalogue output \n
        18 default subwindow behavior status in catalogue menu\n
        19 default format of catalogue export\n
        20 default path catalogue\n
        21 app zoom\n
        22 app zoom checkbox\n
        23 render mode\n
        24 establish tray icon in startup\n
        """
        def filter_unwanted_chars(to_filter_data, directory = False,even_space=False):
            unwanted_chars = ["\n","\"","\'","[","]"]
            if directory:
                unwanted_chars = ["\n","\"","\'","[","]","\\","/"]
            if even_space:
                unwanted_chars.append(" ")
            filtered_data = ""
            for letters in to_filter_data:
                if letters not in unwanted_chars:
                    filtered_data += letters
            return filtered_data

        def load_default_values():
            dir_names = default_setting_parameters[9:16]
            image_browser_param = default_setting_parameters[17:20]
            output_array = [default_setting_parameters[0],
                            default_setting_parameters[1],
                            default_setting_parameters[2],
                            default_setting_parameters[3],
                            default_setting_parameters[4],
                            default_setting_parameters[5],
                            default_setting_parameters[6],
                            default_setting_parameters[7],
                            default_setting_parameters[8],
                            dir_names,
                            default_setting_parameters[16],
                            image_browser_param,
                            default_setting_parameters[20],
                            default_setting_parameters[21],
                            default_setting_parameters[22],
                            default_setting_parameters[23],
                            default_setting_parameters[24],
                            default_setting_parameters[25],
                            default_setting_parameters[26],
                            default_setting_parameters[27],
                            default_setting_parameters[28],
                            default_setting_parameters[29],
                            default_setting_parameters[30],
                            default_setting_parameters[31],
                            default_setting_parameters[32],
                            ]
            
            print("read intern database (default values)",output_array,len(output_array))
            return output_array

        def insert_new_excel_param(wb,ws,row,param,text):
            """
            Zakládá nové, chybějící parametry
            """
            ws['A' + str(row)] = text
            ws['B' + str(row)] = param
            print(f'inserting new parameter to excel: {text}: {param}')
            wb.save(initial_path+config_filename)

        global global_recources_load_error
        config_filename = "config_TRIMAZKON.xlsx"
        setting_list_name = "Settings_recources"
        default_setting_parameters = string_database.default_setting_database_param
        default_labels = string_database.default_setting_database

        if os.path.exists(initial_path+config_filename):
            try:
                cutoff_date = ["","",""]
                supported_formats_sorting = []
                supported_formats_deleting = []
                wb = load_workbook(initial_path+config_filename)
                ws = wb[setting_list_name]

                # checking possibly missing parameters
                value_check = ws['B' + str(30)].value
                if value_check is None or str(value_check) == "":
                    insert_new_excel_param(wb,ws,row=30,param=default_setting_parameters[29],text=default_labels[29])
                value_check = ws['B' + str(31)].value
                if value_check is None or str(value_check) == "":
                    insert_new_excel_param(wb,ws,row=31,param=default_setting_parameters[30],text=default_labels[30])
                value_check = ws['B' + str(32)].value
                if value_check is None or str(value_check) == "":
                    insert_new_excel_param(wb,ws,row=32,param=default_setting_parameters[31],text=default_labels[31])
                value_check = ws['B' + str(33)].value
                if value_check is None or str(value_check) == "":
                    insert_new_excel_param(wb,ws,row=33,param=default_setting_parameters[32],text=default_labels[32])

                read_formats1 = str(ws['B' + str(1)].value)
                read_formats1 = filter_unwanted_chars(read_formats1,even_space=True)
                found_formats = read_formats1.split(",") 
                for items in found_formats:
                    supported_formats_sorting.append(str(items))

                read_formats2 = str(ws['B' + str(2)].value)
                read_formats2 = filter_unwanted_chars(read_formats2,even_space=True)
                found_formats = read_formats2.split(",")
                for items in found_formats:
                    supported_formats_deleting.append(str(items))
                
                inserted_path = str(ws['B' + str(3)].value)
                path_repaired = Tools.path_check(inserted_path)
                if path_repaired == False:
                    path_repaired = default_setting_parameters[2]

                files_to_keep = int(default_setting_parameters[3])
                files_to_keep_raw = str(ws['B' + str(4)].value)
                if files_to_keep_raw.isdigit():
                    files_to_keep = int(files_to_keep_raw)

                cutoff_date_raw = str(ws['B' + str(5)].value)
                cutoff_date_filtered = filter_unwanted_chars(cutoff_date_raw,even_space=True)
                if "." in cutoff_date_filtered:
                    cutoff_date = cutoff_date_filtered.split(".")
                elif "," in cutoff_date_filtered:
                    cutoff_date = cutoff_date_filtered.split(",")
                
                prefix_function_raw = str(ws['B' + str(6)].value)
                prefix_function_filtered = filter_unwanted_chars(prefix_function_raw,directory=True)
                if prefix_function_filtered != "":
                    prefix_function = prefix_function_filtered
                else:
                    prefix_function = default_setting_parameters[5]

                prefix_camera_raw = str(ws['B' + str(7)].value)
                prefix_camera_filtered = filter_unwanted_chars(prefix_camera_raw,directory=True)
                if prefix_camera_filtered != "":
                    prefix_camera = prefix_camera_filtered
                else:
                    prefix_camera = default_setting_parameters[6]

                maximalized = str(ws['B' + str(8)].value)
                if maximalized != "ano":
                    maximalized = "ne"

                max_pallets = int(default_setting_parameters[8])
                max_pallets_raw = str(ws['B' + str(9)].value)
                if max_pallets_raw.isdigit():
                    max_pallets = int(max_pallets_raw)

                static_dirs_names = []
                for i in range(10,17):
                    dir_name_raw = str(ws['B' + str(i)].value)
                    dir_name_filtered = filter_unwanted_chars(dir_name_raw,directory=True)
                    if dir_name_filtered == "":
                        dir_name_filtered = default_setting_parameters[i-1]
                    static_dirs_names.append(dir_name_filtered)

                safe_mode = str(ws['B' + str(17)].value)
                if safe_mode != "ne":
                    safe_mode = "ano"

                image_browser_param = [default_setting_parameters[17],default_setting_parameters[18],default_setting_parameters[19]] #default
                chosen_option = str(ws['B' + str(18)].value)
                if chosen_option.isdigit():
                    image_browser_param[0] = int(chosen_option)
                zoom_step = str(ws['B' + str(19)].value)
                if zoom_step.isdigit():
                    image_browser_param[1] = int(zoom_step)
                drag_step = str(ws['B' + str(20)].value)
                if drag_step.isdigit():
                    image_browser_param[2] = int(drag_step)
                
                show_changelog = str(ws['B' + str(21)].value)
                if show_changelog != "ano":
                    show_changelog = "ne"

                image_film = str(ws['B' + str(22)].value)
                if image_film != "ano":
                    image_film = "ne"

                num_of_IB_film_images = int(default_setting_parameters[22])
                num_of_IB_film_images_raw = str(ws['B' + str(23)].value)
                if num_of_IB_film_images_raw.isdigit():
                    num_of_IB_film_images = int(num_of_IB_film_images_raw)

                default_sharepoint_database_filename = str(ws['B' + str(24)].value)
                default_sharepoint_database_filename = filter_unwanted_chars(default_sharepoint_database_filename,directory=True)
                default_catalogue_excel_filename = str(ws['B' + str(25)].value)
                default_catalogue_excel_filename = filter_unwanted_chars(default_catalogue_excel_filename,directory=True)
                default_catalogue_xml_filename = str(ws['B' + str(26)].value)
                default_catalogue_xml_filename = filter_unwanted_chars(default_catalogue_xml_filename,directory=True)

                default_catalogue_subwindow_behavior_status = int(default_setting_parameters[26])
                default_catalogue_subwindow_behavior_status_raw = str(ws['B' + str(27)].value)
                if default_catalogue_subwindow_behavior_status_raw.isdigit():
                    default_catalogue_subwindow_behavior_status = int(default_catalogue_subwindow_behavior_status_raw)

                default_catalogue_export_extension = str(ws['B' + str(28)].value)
                default_catalogue_export_extension = filter_unwanted_chars(default_catalogue_export_extension,directory=True,even_space=True)

                default_path_catalogue = str(ws['B' + str(29)].value)
                default_path_catalogue = Tools.path_check(default_path_catalogue)
                if default_path_catalogue == False:
                    default_path_catalogue = default_setting_parameters[28]
                catalogue_save_data = [default_sharepoint_database_filename,default_catalogue_excel_filename,default_catalogue_xml_filename,default_catalogue_subwindow_behavior_status,default_catalogue_export_extension,default_path_catalogue]

                for i in range(0,len(catalogue_save_data)):
                    if catalogue_save_data[i] == "":
                        catalogue_save_data[i] = default_setting_parameters[23+i]

                app_zoom = int(default_setting_parameters[29])
                app_zoom_raw = str(ws['B' + str(30)].value)
                if app_zoom_raw.isdigit():
                    app_zoom = int(app_zoom_raw)

                app_zoom_checkbox = str(ws['B' + str(31)].value)
                if app_zoom_checkbox != "ano":
                    app_zoom_checkbox = "ne"

                render_mode = str(ws['B' + str(32)].value)
                if render_mode != "precise":
                    render_mode = "fast"

                tray_startup_status = str(ws['B' + str(33)].value)
                if tray_startup_status != "ano":
                    tray_startup_status = "ne"

                global_recources_load_error = False
                output_array = [supported_formats_sorting,
                                supported_formats_deleting,
                                path_repaired,
                                files_to_keep,
                                cutoff_date,
                                prefix_function,
                                prefix_camera,
                                maximalized,
                                max_pallets,
                                static_dirs_names,
                                safe_mode,
                                image_browser_param,
                                show_changelog,
                                image_film,
                                num_of_IB_film_images,
                                catalogue_save_data[0],
                                catalogue_save_data[1],
                                catalogue_save_data[2],
                                catalogue_save_data[3],
                                catalogue_save_data[4],
                                catalogue_save_data[5],
                                app_zoom,
                                app_zoom_checkbox,
                                render_mode,
                                tray_startup_status,
                                ]
                
                print("read config",output_array,len(output_array))
                wb.close()
                return output_array

            except Exception as e:
                print(f"Nejdřív zavřete soubor {config_filename} Chyba: {e}")   
                print("Budou načteny defaultní hodnoty")
                global_recources_load_error = True
                output_array = load_default_values()
                return output_array
        else:
            print(f"Chybí konfigurační soubor {config_filename}")
            print("Budou načteny defaultní hodnoty")
            global_recources_load_error = True
            output_array = load_default_values()
            return output_array
            
    @classmethod
    def read_text_file_data_temp(cls): # Funkce vraci data z textoveho souboru Recources.txt
        """
        !!!!TEMP!!!! nez se prejde na novy standard s config\n

        Funkce vraci data z textoveho souboru Recources.txt

        data jsou v poradi:

        0 supported_formats_sorting\n
        1 supported_formats_deleting\n
        2 path_repaired\n
        3 files_to_keep\n
        4 cutoff_date\n
        5 prefix_function\n
        6 prefix_camera\n
        7 maximalized\n
        8 max_pallets\n
        9 static_dirs_names\n
        10 sorting_safe_mode\n
        11 image_browser_setting [checkbox, increment, movement]\n
        12 show_changelog\n
        13 image_film\n
        14 num_of_IB_film_images\n
        15 default sharepoint database filename\n
        16 default excel filnename prefix for catalogue output\n
        17 default xml filename for catalogue output \n
        18 default subwindow behavior status in catalogue menu\n
        19 default format of catalogue export\n
        20 default path catalogue\n
        21 app zoom\n
        """

        if os.path.exists(initial_path+'Recources.txt'):
            cutoff_date = ["","",""]
            with open(initial_path+'Recources.txt','r',encoding='utf-8',errors='replace') as recources_file:
                Lines = recources_file.readlines()
            supported_formats_sorting = []
            supported_formats_deleting = []
            unwanted_chars = ["\n","\"","[","]"]
            for chars in unwanted_chars:
                if chars in Lines[2]:
                    Lines[2] = Lines[2].replace(chars,"")
                if chars in Lines[4]:
                    Lines[4] = Lines[4].replace(chars,"")
                
            found_formats = Lines[2].split(",") 
            for items in found_formats:
                supported_formats_sorting.append(str(items))
            found_formats = Lines[4].split(",")
            for items in found_formats:
                supported_formats_deleting.append(str(items))
            
            inserted_path = Lines[6].replace("\n","")
            inserted_path = str(inserted_path)

            path_repaired = Tools.path_check(inserted_path)

            Lines[8] = Lines[8].replace("\n","")
            files_to_keep = int(Lines[8])
            
            Lines[10] = Lines[10].replace("\n","")
            cutoffdate_splitted = Lines[10].split(".")
            i=0
            for items in cutoffdate_splitted:
                i+=1
                cutoff_date[i-1] = items

            Lines[12] = Lines[12].replace("\n","")
            Lines[12] = Lines[12].replace("\"","")
            Lines[12] = Lines[12].replace("/","")
            if str(Lines[12]) != "":
                prefix_function = Lines[12]
            else:
                prefix_function = "Func_"

            Lines[14] = Lines[14].replace("\n","")
            Lines[14] = Lines[14].replace("\"","")
            Lines[14] = Lines[14].replace("/","")
            if str(Lines[14]) != "":
                prefix_camera = Lines[14]
            else:
                prefix_camera = "Cam_"

            #spoustet v maximalizovanem okne?
            Lines[16] = Lines[16].replace("\n","")
            if str(Lines[16]) != "":
                maximalized = Lines[16]
            else:
                maximalized = "ne"
            #maximalni pocet palet v obehu
            Lines[18] = Lines[18].replace("\n","")
            if str(Lines[18]) != "":
                max_pallets = int(Lines[18])
            else:
                max_pallets = 55

            # cteme nekolik nazvu slozek:
            static_dirs_names = []
            for i in range(20,34,2):
                Lines[i] = Lines[i].replace("\n","")
                Lines[i] = Lines[i].replace("\"","")
                Lines[i] = Lines[i].replace("/","")
                static_dirs_names.append(Lines[i])
                
            #bezpecny mod?
            Lines[34] = Lines[34].replace("\n","")
            if str(Lines[34]) != "":
                safe_mode = Lines[34]
            else:
                safe_mode = "ano"

            #image browser parametry
            image_browser_param = [1,10,200] #default
            Lines[36] = Lines[36].replace("\n","")
            Lines[37] = Lines[37].replace("\n","")
            Lines[38] = Lines[38].replace("\n","")
            if (str(Lines[36]) != "") and Lines[36].isdigit():
                image_browser_param[0] = int(Lines[36])
            if (str(Lines[37]) != "") and Lines[37].isdigit():
                image_browser_param[1] = int(Lines[37])
            if (str(Lines[38]) != "") and Lines[38].isdigit():
                image_browser_param[2] = int(Lines[38])
                
            Lines[40] = Lines[40].replace("\n","")
            if Lines[40] != "ano":
                show_changelog = "ne"
            else:
                show_changelog = Lines[40]
            
            Lines[42] = Lines[42].replace("\n","")
            if Lines[42] != "ano":
                image_film = "ne"
            else:
                image_film = Lines[42]

            Lines[44] = Lines[44].replace("\n","")
            if Lines[44].isdigit():
                num_of_IB_film_images = int(Lines[44])
            else:
                num_of_IB_film_images = 6 #default

            try:
                default_sharepoint_database_filename = Lines[46].replace("\n","")
                default_catalogue_excel_filename = Lines[48].replace("\n","")
                default_catalogue_xml_filename = Lines[50].replace("\n","")
                Lines[52] = Lines[52].replace("\n","")
                if Lines[52].isdigit():
                    default_catalogue_subwindow_behavior_status = int(Lines[52])
                else:
                    default_catalogue_subwindow_behavior_status = False
                default_catalogue_export_extension = Lines[54].replace("\n","")
                default_path_catalogue = Lines[56].replace("\n","")
                catalogue_save_data = [default_sharepoint_database_filename,default_catalogue_excel_filename,default_catalogue_xml_filename,default_catalogue_subwindow_behavior_status,default_catalogue_export_extension,default_path_catalogue]
                for i in range(0,len(catalogue_save_data)):
                    if catalogue_save_data[i] == "":
                        catalogue_save_data[i] = False
            except Exception:
                catalogue_save_data = [False]*6

            # Lines[58] = Lines[58].replace("\n","")
            # if Lines[58].isdigit():
            #     app_zoom = int(Lines[58])
            # else:
            #     app_zoom = 100 #default

            output_array = [supported_formats_sorting,supported_formats_deleting,path_repaired,files_to_keep,cutoff_date,
                    prefix_function,prefix_camera,maximalized,max_pallets,static_dirs_names,safe_mode,image_browser_param,
                    show_changelog,image_film,num_of_IB_film_images,catalogue_save_data[0],catalogue_save_data[1],catalogue_save_data[2],
                    catalogue_save_data[3],catalogue_save_data[4],catalogue_save_data[5]]
            return output_array
        else:
            print("Chybí konfigurační soubor Recources.txt")
            return [False]*26

    @classmethod
    def save_to_config(cls,input_data,which_parameter): # Funkce zapisuje data do souboru config_TRIMAZKON.xlsx
        """
        Funkce zapisuje data do textoveho souboru Recources.txt

        vraci vystupni zpravu: report

        which_parameter je bud: 
        
        1 add_supported_sorting_formats\n
        2 add_supported_deleting_formats\n
        3 pop_supported_sorting_formats\n
        4 pop_supported_deleting_formats\n
        5 default_path\n
        6 default_files_to_keep\n
        7 default_cutoff_date\n
        8 new_default_prefix_func\n
        9 new_default_prefix_cam\n
        10 maximalized\n
        11 pallets_set\n
        12 new_default_static_dir_name\n
        13 sorting_safe_mode\n
        14 image_browser_param_set\n
        15 show_change_log\n
        16 image_film\n
        17 num_of_IB_film_images\n
        18 catalogue_data\n
        19 app_zoom\n
        20 app_zoom_checkbox\n
        21 render_mode\n
        22 tray_icon_startup\n
        """

        def filter_unwanted_chars(to_filter_data, directory = False,formats = False):
            unwanted_chars = ["\n","\"","\'","[","]"]
            if directory:
                unwanted_chars = ["\n","\"","\'","[","]","\\","/"]
            if formats:
                unwanted_chars = ["\n","\"","\'","[","]"," ","."]

            filtered_data = ""
            for letters in to_filter_data:
                if letters not in unwanted_chars:
                    filtered_data += letters
            return filtered_data
        
        config_filename = "config_TRIMAZKON.xlsx"
        setting_list_name = "Settings_recources"
        parameter_row_mapping = {
            "add_supported_sorting_formats": 1,
            "add_supported_deleting_formats": 2,
            "pop_supported_sorting_formats": 1,
            "pop_supported_deleting_formats": 2,
            "default_path": 3,
            "default_files_to_keep": 4,
            "default_cutoff_date": 5,
            "new_default_prefix_func": 6,
            "new_default_prefix_cam": 7,
            "maximalized": 8,
            "pallets_set": 9,
            "new_default_static_dir_name": [10,11,12,13,14,15,16],
            "sorting_safe_mode": 17,
            "image_browser_param_set": [18,19,20],
            "show_change_log": 21,
            "image_film": 22,
            "num_of_IB_film_images": 23,
            "catalogue_data": [24,25,26,27,28,29],
            "app_zoom": 30,
            "app_zoom_checkbox": 31,
            "render_mode": 32,
            "tray_icon_startup": 33,
            }
        
        if os.path.exists(initial_path + config_filename):
            wb = load_workbook(initial_path+config_filename)
            ws = wb[setting_list_name]
            formats_changes = False
            report = ""
            supported_formats_sorting = []
            found_formats = str(ws['B' + str(1)].value)
            found_formats = filter_unwanted_chars(found_formats,formats=True)
            found_formats = found_formats.split(",")
            supported_formats_sorting = found_formats
            supported_formats_deleting = []
            found_formats = str(ws['B' + str(2)].value)
            found_formats = filter_unwanted_chars(found_formats,formats=True)
            found_formats = found_formats.split(",")
            supported_formats_deleting = found_formats

            if which_parameter == "add_supported_sorting_formats":
                corrected_input = filter_unwanted_chars(str(input_data),formats=True)
                if str(corrected_input) not in supported_formats_sorting:
                    supported_formats_sorting.append(str(corrected_input))
                    report = (f"Byl přidán formát: \"{corrected_input}\" do podporovaných formátů pro možnosti třídění")
                    formats_changes = True
                else:
                    report =  (f"Formát: \"{corrected_input}\" je již součástí podporovaných formátů možností třídění")
            
            elif which_parameter == "add_supported_deleting_formats":
                corrected_input = filter_unwanted_chars(str(input_data),formats=True)
                if str(corrected_input) not in supported_formats_deleting:
                    supported_formats_deleting.append(str(corrected_input))
                    report =  (f"Byl přidán formát: \"{corrected_input}\" do podporovaných formátů pro možnosti mazání")
                    formats_changes = True
                else:
                    report =  (f"Formát: \"{corrected_input}\" je již součástí podporovaných formátů možností mazání")
                
            elif which_parameter == "pop_supported_sorting_formats":
                poped = 0
                found = False
                range_to = len(supported_formats_sorting)
                for i in range(0,range_to):
                    if i < range_to:
                        if str(input_data) == supported_formats_sorting[i] and len(str(input_data)) == len(supported_formats_sorting[i]):
                            supported_formats_sorting.pop(i)
                            poped+=1
                            range_to = range_to - poped
                            report =  (f"Z podporovaných formátů možností třídění byl odstraněn formát: \"{input_data}\"")
                            formats_changes = True
                            found = True
                if found == False:
                    report =  (f"Formát: \"{input_data}\" nebyl nalezen v podporovaných formátech možností třídění, nemůže tedy být odstraněn")
                
            elif which_parameter == "pop_supported_deleting_formats":
                poped = 0
                found = False
                range_to = len(supported_formats_deleting)
                for i in range(0,range_to):
                    if i < range_to:
                        if str(input_data) == supported_formats_deleting[i] and len(str(input_data)) == len(supported_formats_deleting[i]):
                            supported_formats_deleting.pop(i)
                            poped+=1
                            range_to = range_to - poped
                            report =  (f"Z podporovaných formátů možností mazání byl odstraněn formát: \".{input_data}\"")
                            formats_changes = True
                            found = True
                if found == False:
                    report =  (f"Formát: \"{input_data}\" nebyl nalezen v podporovaných formátech možností mazání, nemůže tedy být odstraněn")
            
            elif which_parameter == "default_path":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
                report = (f"Základní cesta přenastavena na: {str(input_data)}")
            
            elif which_parameter == "default_files_to_keep":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
            
            elif which_parameter == "default_cutoff_date":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data[0])+"."+str(input_data[1])+"."+str(input_data[2])

            elif which_parameter == "new_default_prefix_func":
                input_filtered = filter_unwanted_chars(str(input_data),directory=True)
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_filtered)
                report = (f"Základní prefix názvu složek pro třídění podle funkce přenastaven na: {str(input_filtered)}")

            elif which_parameter == "new_default_prefix_cam":
                input_filtered = filter_unwanted_chars(str(input_data),directory=True)
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_filtered)
                report = (f"Základní prefix názvu složek pro třídění podle kamery přenastaven na: {str(input_filtered)}")

            elif which_parameter == "maximalized":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "pallets_set":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "new_default_static_dir_name":
                input_data_splitted = str(input_data).split(" | ")
                input_data = input_data_splitted[0]
                increment = int(input_data_splitted[1])
                cells_with_names = parameter_row_mapping.get(which_parameter)
                excel_row = cells_with_names[increment]
                input_filtered = filter_unwanted_chars(str(input_data),directory=True)
                ws['B' + str(excel_row)] = str(input_filtered)
            
            elif which_parameter == "sorting_safe_mode":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "image_browser_param_set":
                cells_with_params = parameter_row_mapping.get(which_parameter)
                for i in range(0,len(cells_with_params)):
                    if input_data[i] != None:
                        ws['B' + str(cells_with_params[i])] = str(input_data[i])

            elif which_parameter == "show_change_log":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = "ne"

            elif which_parameter == "image_film":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
            
            elif which_parameter == "num_of_IB_film_images":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "catalogue_data":
                cells_with_params = parameter_row_mapping.get(which_parameter)
                for i in range(0,len(cells_with_params)):
                    if input_data[i] != None:
                        ws['B' + str(cells_with_params[i])] = str(input_data[i])

            elif which_parameter == "app_zoom":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
            
            elif which_parameter == "app_zoom_checkbox":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "render_mode":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "tray_icon_startup":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            if formats_changes:
                #navraceni poli zpet do stringu radku:
                sorting_formats_string = ""
                for items in supported_formats_sorting:
                    if sorting_formats_string  == "":
                        sorting_formats_string += str(items)
                    else:
                        sorting_formats_string += ("," + str(items))
                
                deleting_formats_string = ""  
                for items in supported_formats_deleting:
                    if deleting_formats_string == "":
                        deleting_formats_string += str(items)
                    else:
                        deleting_formats_string += ("," + str(items))

                ws['B' + str(1)] = str(sorting_formats_string)
                ws['B' + str(2)] = str(deleting_formats_string)

            wb.save(initial_path+config_filename)
            wb.close()
            return report
        
        else:
            print("Chybí konfigurační soubor config_TRIMAZKON.xlsx (nelze ukládat změny)")
            return "Chybí konfigurační soubor config_TRIMAZKON.xlsx (nelze ukládat změny)"
   
    @classmethod
    def browseDirectories(cls,visible_files,start_path=None): # Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat
        """
        Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat

        Vstupní data:

        0: visible_files = "all" / "only_dirs"\n
        1: start_path = None -optimalni, docasne se ulozi posledni nastavena cesta v exploreru

        Výstupní data:

        0: výstupní chybová hlášení
        1: opravená cesta
        2: nazev vybraneho souboru (option: all)
        """
        corrected_path = ""
        output= ""
        name_of_selected_file = ""
        text_file_data = Tools.read_config_data()
        if start_path == None:
            start_path = str(text_file_data[2]) #defaultni cesta
        else: # byla zadana docasna cesta pro explorer
            checked_path = Tools.path_check(start_path)
            if checked_path == False:
                output = "Změněná dočasná základní cesta pro explorer již neexistuje"
                start_path = str(text_file_data[2]) #defaultni cesta
            else:
                start_path = checked_path

        if start_path != False:
            if not os.path.exists(start_path):
                start_path = ""
                output="Konfigurační soubor obsahuje neplatnou cestu"

        else:
            output="Chybí konfigurační soubor config_TRIMAZKON.xlsx s počáteční cestou...\n"
            start_path=""

        # pripad vyberu files, aby byly viditelne
        if visible_files == "all":
            if(start_path != ""):
                foldername_path = filedialog.askopenfile(initialdir = start_path,title = "Klikněte na soubor v požadované cestě")
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:           
                foldername_path = filedialog.askopenfile(initialdir = "/",title = "Klikněte na soubor v požadované cestě")
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"

        # pripad vyberu slozek
        if visible_files == "only_dirs":
            if(start_path != ""):
                path_to_directory = filedialog.askdirectory(initialdir = start_path, title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:
                path_to_directory = filedialog.askdirectory(initialdir = "/", title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"

        check = Tools.path_check(path_to_directory)
        corrected_path = check
        return [output,corrected_path,name_of_selected_file]

    @classmethod
    def add_colored_line(cls,text_widget, text, color,font=None,delete_line = None,no_indent=None):
        """
        Vloží řádek do console
        """
        try:
            text_widget.configure(state=tk.NORMAL)
            if font == None:
                font = ("Arial",16)
            if delete_line != None:
                text_widget.delete("current linestart","current lineend")
                text_widget.tag_configure(color, foreground=color,font=font)
                text_widget.insert("current lineend",text, color)
            else:
                text_widget.tag_configure(color, foreground=color,font=font)
                if no_indent:
                    text_widget.insert(tk.END,text+"\n", color)
                else:
                    text_widget.insert(tk.END,"    > "+ text+"\n", color)

            text_widget.configure(state=tk.DISABLED)
        except Exception as e:
            print(f"Error při psaní do konzole: {e}")

    @classmethod
    def save_path(cls,console,path_entered):
        path_given = path_entered
        path_checked = Tools.path_check(path_given)
        if path_checked != False and path_checked != "/":
            console_input = Tools.save_to_config(path_checked,"default_path")
            Tools.add_colored_line(console,console_input,"green",None,True)
        elif path_checked != "/":
            Tools.add_colored_line(console,f"Zadaná cesta: {path_given} nebyla nalezena, nebude tedy uložena","red",None,True)
        elif path_checked == "/":
            Tools.add_colored_line(console,"Nebyla vložena žádná cesta k souborům","red",None,True)

    @classmethod
    def clear_console(cls,text_widget,from_where=None):
        """
        Vymaže celou consoli
        """
        if from_where == None:
            from_where = 1.0
        text_widget.configure(state=tk.NORMAL)
        text_widget.delete(from_where, tk.END)
        text_widget.configure(state=tk.DISABLED)

    @classmethod
    def check_config_file(cls,config_filename,setting_list_name):
        """
        TEMPORARY - kdyby někdo měl starou verzi
        - kontroluje zda chybi excel s novym nazvem, pak kontroluje jesli excel obsahuje novy list Settings_recources
        - přejmenuje si to config excel
        - pretahne si to data z recources.txt do spolecneho excelu pod novy list
        """
        def move_to_temp_dir(filename:str):
            if not os.path.exists(initial_path + "TRIMAZKON_uz_nechce_vzal_si_data"):
                os.mkdir(initial_path + "TRIMAZKON_uz_nechce_vzal_si_data" + "/")
            print(f"moving old config file {filename}")
            shutil.move(initial_path +filename, initial_path + "TRIMAZKON_uz_nechce_vzal_si_data" + "/"+filename)
        
        if os.path.exists(initial_path+'saved_addresses_2.xlsx'):
            if not os.path.exists(initial_path+config_filename):
                shutil.copy(initial_path+'saved_addresses_2.xlsx',initial_path+config_filename)
                print("copying old excel config file")
            move_to_temp_dir('saved_addresses_2.xlsx')

        try:
            wb = load_workbook(initial_path+config_filename)
            if not setting_list_name in wb.sheetnames:
                user_config_file_data = []
                ws = wb.create_sheet(title=setting_list_name)
                wb.save(initial_path+config_filename)
                if os.path.exists(initial_path+'Recources.txt'):
                    # copy user settings
                    recources_data = Tools.read_text_file_data_temp()
                    for i in range(0,len(recources_data)):
                        if i == 9 or i == 11:
                            for subdata in recources_data[i]:
                                user_config_file_data.append(subdata)
                        else:
                            user_config_file_data.append(recources_data[i])
                    print("using user settings from Recources.txt")
                    move_to_temp_dir('Recources.txt')        
                else:
                    user_config_file_data = string_database.default_setting_database_param
                        
                print(user_config_file_data,len(user_config_file_data))
                user_config_file_data_labels = string_database.default_setting_database
                
                for i in range(1,len(user_config_file_data)+1):
                    ws['A' + str(i)] = str(user_config_file_data_labels[i-1])
                    ws['B' + str(i)] = str(user_config_file_data[i-1])
                wb.save(initial_path+config_filename)

            # renaming sheet with a typo in name:
            if "ip_adress_fav_list" in wb.sheetnames:
                ws = wb["ip_adress_fav_list"]
                ws.title = "ip_address_fav_list"
                wb.save(initial_path+config_filename)
            wb.close()
        except Exception as e:
            print(f"Nejdřív zavřete soubor {config_filename} Chyba: {e}")

    @classmethod
    def check_task_existence_in_TS(cls,taskname):
        process = subprocess.Popen(f'schtasks /query /tn \"{taskname}\" /v /fo LIST',
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        try:
            stdout_str = stdout.decode('utf-8')
            stderr_str = stderr.decode('utf-8')
            data = str(stdout_str)
            error_data = str(stderr_str)
        except UnicodeDecodeError:
            try:
                stdout_str = stdout.decode('cp1250')
                stderr_str = stderr.decode('cp1250')
                data = str(stdout_str)
                error_data = str(stderr_str)
            except UnicodeDecodeError:
                data = str(stdout)
                error_data = str(stderr)
        if "ERROR" in error_data:
            return False
        else:
            return True
    
    @classmethod
    def tray_startup_cmd(cls):
        """
        už není potřebné - aplikace tray není součástí celého exe - je zvlášť
        - protože musela souběžně běžet hlavní aplikace trimazkon kvuli pristupu k appdata temp složce
        """
        def is_process_running(process_name):
            """
            Check if a process with the given name is running.
            :param process_name: Name of the process to check (e.g., 'name.exe')
            :return: True if the process is running, False otherwise
            """
            for process in psutil.process_iter(['name']):
                try:
                    if process.info['name'] and process_name.lower() in process.info['name'].lower():
                        return True
                except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                    pass
            return False
        
        if is_process_running(trimazkon_tray_exe_name):
            return
        # resource_path = Tools.resource_path(trimazkon_tray_exe_name)
        # resource_path = Tools.resource_path(exe_name)
        # print("calling process: ",subexe_path + " " + initial_path + " run_tray")
        cmd_command = initial_path+"/"+trimazkon_tray_exe_name + " run_tray"
        print("calling process: ",cmd_command)
        subprocess.Popen(cmd_command, shell=True, text=True, creationflags=subprocess.CREATE_NO_WINDOW)

    @classmethod
    def establish_startup_tray(cls):
        """
        Sets the startup task of switching on the tray application icon
        - if it doesnt exist already
        """
        task_name = "TRIMAZKON_startup_tray_setup"
        task_presence = Tools.check_task_existence_in_TS(task_name)
        print("task presence: ",task_presence)
        if not task_presence:
            # path_app_location = str(initial_path+exe_name) # predelno na tray exe aplikaci zvlášť
            path_app_location = str(initial_path + trimazkon_tray_exe_name)
            # task_command = "\"" + path_app_location + " tray_startup_call" + "\" /sc onlogon"
            # task_command = "/c start \"" + path_app_location + " " + initial_path + " run_tray" + "\" /sc onlogon"
            # resource_path = Tools.resource_path(trimazkon_tray_exe_name)
            # resource_path = Tools.resource_path(exe_name)
            task_command = "\"" + path_app_location + " run_tray" + "\" /sc onlogon"
            process = subprocess.Popen(f"schtasks /Create /TN {task_name} /TR {task_command}",
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            
            stdout, stderr = process.communicate()
            output_message = "out"+str(stdout) +"err"+str(stderr)
            print(output_message)
            if "Access is denied" in output_message:
                return "need_access"
            
            Tools.tray_startup_cmd() # init sepnutí po prvním zavedení tasku
            # def long_execution():
            #     tray_startup_cmd() # init sepnutí po prvním zavedení tasku
            # tray_thread = threading.Thread(target=long_execution)
            # tray_thread.start()
    
    @classmethod
    def call_again_as_admin(cls,input_flag:str,window_title,main_title):
        def run_as_admin():# Vyžádání admin práv: nefunkční ve vscode
            def is_admin():
                try:
                    return ctypes.windll.shell32.IsUserAnAdmin()
                except:
                    return False
            if not is_admin():
                ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join([input_flag]), None, 1)
                sys.exit()

        def close_prompt(child_root):
            child_root.grab_release()
            child_root.destroy()

        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(app_icon)))
        # child_root.geometry(f"620x150+{300}+{300}")  
        child_root.title(window_title)
        label_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        proceed_label = customtkinter.CTkLabel(master = label_frame,text = main_title,font=("Arial",25),anchor="w",justify="left")
        proceed_label.pack(pady=5,padx=10,anchor="w",side = "left")
        button_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_yes =    customtkinter.CTkButton(master = button_frame,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin())
        button_no =     customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
        button_no       .pack(pady = 5, padx = 10,anchor="e",side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="e",side="right")
        label_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        button_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)

        child_root.update()
        child_root.update_idletasks()
        # child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{300}+{300}")
        child_root.focus()
        child_root.focus_force()
        child_root.grab_set()

    @classmethod
    def remove_task_from_TS(cls,name_of_task):
        cmd_command = f"schtasks /Delete /TN {name_of_task} /F"
        # subprocess.call(cmd_command,shell=True,text=True)

        process = subprocess.Popen(cmd_command,
                                   stdout=subprocess.PIPE,
                                   stderr=subprocess.PIPE,
                                   creationflags=subprocess.CREATE_NO_WINDOW)
            
        stdout, stderr = process.communicate()
        output_message = "out"+str(stdout) +"err"+str(stderr)
        print(output_message)
        if "Access is denied" in output_message:
            return "need_access"

def get_init_path():
    initial_path = Tools.path_check(os.getcwd())
    if len(sys.argv) > 1: #spousteni pres cmd (kliknuti na obrazek) nebo task scheduler - mazání
        raw_path = str(sys.argv[0])
        initial_path = Tools.path_check(raw_path,True)
        initial_path_splitted = initial_path.split("/")
        initial_path = ""
        for i in range(0,len(initial_path_splitted)-2):
            initial_path += str(initial_path_splitted[i])+"/"
        print("SYSTEM: ",sys.argv)

    return initial_path

initial_path = get_init_path()
print("init path: ",initial_path)

class WindowsBalloonTip:
    """
    Windows system notification (balloon tip).
    """
    _class_registered = False  # Ensures window class is registered only once

    def __init__(self, title, msg, app_icon):
        message_map = {
            win32con.WM_DESTROY: self.OnDestroy,
        }

        hinst = GetModuleHandle(None)
        class_name = "PythonTaskbar"
        try:
            if not WindowsBalloonTip._class_registered:
                # Register the Window class once
                wc = WNDCLASS()
                wc.hInstance = hinst
                wc.lpszClassName = class_name
                wc.lpfnWndProc = message_map
                RegisterClass(wc)
                WindowsBalloonTip._class_registered = True  # Mark as registered
        except Exception:
            wc = WNDCLASS()
            wc.hInstance = hinst
            wc.lpszClassName = class_name
            wc.lpfnWndProc = message_map
            RegisterClass(wc)

        # Create a new window (without re-registering the class)
        style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
        self.hwnd = CreateWindow(class_name, "Taskbar", style, 
                                 0, 0, win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT, 
                                 0, 0, hinst, None)

        UpdateWindow(self.hwnd)

        # Load icon
        icon_flags = win32con.LR_LOADFROMFILE | win32con.LR_DEFAULTSIZE
        try:
            hicon = LoadImage(hinst, app_icon, win32con.IMAGE_ICON, 0, 0, icon_flags)
        except:
            hicon = LoadIcon(0, win32con.IDI_APPLICATION)

        # Display notification
        # flags = win32gui.NIF_ICON | win32gui.NIF_MESSAGE | win32gui.NIF_TIP
        flags = NIF_ICON | NIF_MESSAGE | NIF_TIP
        nid = (self.hwnd, 0, flags, win32con.WM_USER+20, hicon, "tooltip")
        Shell_NotifyIcon(NIM_ADD, nid)

        Shell_NotifyIcon(NIM_MODIFY, 
                         (self.hwnd, 0, NIF_INFO, win32con.WM_USER+20,
                          hicon, "Balloon tooltip", msg, 200, title))

        # time.sleep(10)  # Display the notification for 10 seconds
        # self.cleanup()

    def cleanup(self):
        """ Removes the notification icon and destroys the window. """
        nid = (self.hwnd, 0)
        Shell_NotifyIcon(NIM_DELETE, nid)
        DestroyWindow(self.hwnd)

    def OnDestroy(self, hwnd, msg, wparam, lparam):
        """ Handles window destruction. """
        self.cleanup()
        PostQuitMessage(0)  # Terminate the app.

def deleting_via_cmd():
    print("deleting system entry: ",sys.argv)
    task_name = str(sys.argv[2])
    deleting_path = str(sys.argv[3])
    max_days = int(sys.argv[4])
    files_to_keep = int(sys.argv[5])
    cutoff_date = Deleting.get_cutoff_date(days=max_days)
    text_file_data = Tools.read_config_data()
    supported_formats_deleting = text_file_data[1]
    list_of_folder_names = text_file_data[9]
    to_delete_folder_name = list_of_folder_names[2]

    del_instance = Deleting.whole_deleting_function(
        deleting_path,
        more_dirs=False,
        del_option=1,
        files_to_keep=files_to_keep,
        cutoff_date_given=cutoff_date,
        supported_formats=supported_formats_deleting,
        testing_mode=False,
        to_delete_folder_name=to_delete_folder_name
        )
    output_data = del_instance.main()
    output_message = f"|||Datum provedení: {output_data[3]}||Zkontrolováno: {output_data[0]} souborů||Starších: {output_data[1]} souborů||Smazáno: {output_data[2]} souborů"
    output_message_clear = f"Provedeno: {output_data[3]}\nZkontrolováno: {output_data[0]} souborů\nStarších: {output_data[1]} souborů\nSmazáno: {output_data[2]} souborů"
    print(output_message)

    trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path)
    trimazkon_tray_instance.save_new_log(task_name,output_message)

    icon_path = Tools.resource_path('images/logo_TRIMAZKON.ico')
    WindowsBalloonTip("Bylo provedeno automatické mazání",
                        str(output_message_clear),
                        icon_path)
    # try:
    #     notification.notify(title="Bylo provedeno automatické mazání",
    #                         message=str(output_message_clear))
    #                         # app_name="TRIMAZKON",
    #                         # app_icon=icon_path)
    # except Exception:
    #     pass

    # icon_path = Tools.resource_path('images/logo_TRIMAZKON.ico')
    # notification.notify(title="Bylo provedeno automatické mazání",
    #                     message=str(output_message_clear),
    #                     # app_name="TRIMAZKON",
    #                     app_icon=icon_path)
    
    # subexe_path = Tools.resource_path(trimazkon_tray_exe_name)
    # subprocess.run(subexe_path + " " + initial_path + " save_new_log "+task_name+" "+output_message,
    #                 creationflags=subprocess.CREATE_NO_WINDOW)
    return output_message_clear

load_gui=True
print(sys.argv)
if len(sys.argv) > 1:
    if sys.argv[1] == "deleting":
        deleting_output_message = deleting_via_cmd()
        load_gui = False
        sys.exit(f"0: {deleting_output_message}")
        
    # elif sys.argv[1] == "tray_startup_call":
    #     tray_startup_cmd()
    #     load_gui = False
    #     sys.exit(0)

class system_pipeline_communication: # vytvoření pipeline serveru s pipe názvem TRIMAZKON_pipe_ + pid (id systémového procesu)
    """
    aby bylo možné posílat běžící aplikaci parametry:
    - mám otevřené okno ip setting - kliknu na obrázek - jen pošlu parametry
    """
    def __init__(self,exe_name,no_server = False):
        self.root = None #define later (to prevend gui loading when 2 apps opened)
        self.current_pid = None
        self.exe_name = exe_name
        if self.exe_name == None or self.exe_name == "":
            self.exe_name = "TRIMAZKON.exe"
        self.current_pid = os.getpid()
        if not no_server:
            # self.start_server()
            run_server_background = threading.Thread(target=self.start_server,)
            run_server_background.start()

    def server(self,pipe_input):
        """
        Endless loop listening for commands
        """
        pipe_name = fr'\\.\pipe\{pipe_input}'
        while True:
            print(f"Waiting for a TRIMAZKON to connect on {pipe_name}...") 
            pipe = win32pipe.CreateNamedPipe(
                pipe_name,
                win32pipe.PIPE_ACCESS_DUPLEX,
                win32pipe.PIPE_TYPE_MESSAGE | win32pipe.PIPE_READMODE_MESSAGE | win32pipe.PIPE_WAIT,
                1,
                512,
                512,
                0,
                None
            )

            win32pipe.ConnectNamedPipe(pipe, None)
            print("TRIMAZKON connected.")

            try:
                while True:
                    hr, data = win32file.ReadFile(pipe, 64 * 1024)
                    received_data = data.decode()
                    print(f"Received: {received_data}")
                    self.root.after(0,menu.command_landed,received_data)

            except pywintypes.error as e:
                if e.args[0] == 109:  # ERROR_BROKEN_PIPE
                    print("TRIMAZKON disconnected.")
            finally:
                # Close the pipe after disconnection
                win32file.CloseHandle(pipe)
            # Loop back to wait for new client connections

    def client(self,pipe_name_given,command,parameters):
        pipe_name = fr'\\.\pipe\{pipe_name_given}'
        handle = win32file.CreateFile(
            pipe_name,
            win32file.GENERIC_READ | win32file.GENERIC_WRITE,
            0,
            None,
            win32file.OPEN_EXISTING,
            0,
            None
        )
        if "Open image browser" in command:
            message = str(parameters[0]) + ",," + str(parameters[1])
            win32file.WriteFile(handle, message.encode())
            print("Message sent.")

    def start_server(self):
        self.pipe_name = f"TRIMAZKON_pipe_{self.current_pid}"
        running_server = threading.Thread(target=self.server, args=(self.pipe_name,), daemon=True)
        running_server.start()
        time.sleep(0.5)  # Wait for the server to start

    def call_checking(self,command,parameters):
        """
        for every found process with name of an application: send given command
        """
        checking = Tools.get_all_app_processes()
        print("SYSTEM application processes: ",checking)
        # if it is running more then one application, execute (root + self.root)
        if checking[0]>2:
            pid_list = checking[1]
            # try to send command to every process which has application name
            for pids in pid_list:
                if pids != self.current_pid:
                    try:
                        pipe_name = f"TRIMAZKON_pipe_{pids}"
                        self.client(pipe_name,command,parameters)
                    except Exception:
                        pass
            return True
        else:
            return False

if load_gui:
    # pipeline_duplex = system_pipeline_communication(exe_name)# Establishment of pipeline server for duplex communication between running applications
    app_running_status = Tools.check_runing_app_duplicity()
    print("already opened app status: ",app_running_status)
    if len(sys.argv) > 1: # VÝJIMKA: pukud nové spuštění s admin právy načti i gui...
        if sys.argv[0] == sys.argv[1]:
            app_running_status = False

    if not app_running_status:
        pipeline_duplex = system_pipeline_communication(exe_name)# Establishment of pipeline server for duplex communication between running applications
        app_icon = Tools.resource_path('images/logo_TRIMAZKON.ico')
        customtkinter.set_appearance_mode("dark")
        customtkinter.set_default_color_theme("dark-blue")
        root=customtkinter.CTk()
        root.geometry("1200x900")
        root.title("TRIMAZKON v_4.2.1")
        root.wm_iconbitmap(Tools.resource_path(app_icon))

    else:# předání parametrů v případě spuštění obrázkem (základní obrázkový prohlížeč)
        pipeline_duplex_instance = system_pipeline_communication(exe_name,no_server=True) # pokud už je aplikace spuštěná nezapínej server, trvá to...
        if len(sys.argv) > 1:
            raw_path = str(sys.argv[1]) #klik na spusteni trimazkonu s admin právy
            if sys.argv[0] != sys.argv[1]: # pokud se nerovnají jedná se nejspíše o volání základního prohlížeče obrázků (spuštění kliknutím na obrázek...)
                IB_as_def_browser_path=Tools.path_check(raw_path,True)
                IB_as_def_browser_path_splitted = IB_as_def_browser_path.split("/")
                IB_as_def_browser_path = ""
                for i in range(0,len(IB_as_def_browser_path_splitted)-2):
                    IB_as_def_browser_path += IB_as_def_browser_path_splitted[i]+"/"
                selected_image = IB_as_def_browser_path_splitted[len(IB_as_def_browser_path_splitted)-2]
                pipeline_duplex_instance.call_checking(f"Open image browser starting with image: {IB_as_def_browser_path}, {selected_image}",[IB_as_def_browser_path,selected_image])

def set_zoom(zoom_factor):
    try:
        root.after(0, lambda: customtkinter.set_widget_scaling(zoom_factor / 100))
        # customtkinter.set_widget_scaling(zoom_factor / 100)
    except Exception as e:
        print(f"error with zoom scaling: {e}")
    
    root.tk.call('tk', 'scaling', zoom_factor / 100)

class main_menu:
    def __init__(self,root):
        self.root = root
        pipeline_duplex.root = self.root # předání rootu do pipeline_duplex až ve chvílí, kdy je jasné, že aplikace není vícekrát spuštěná:
        config_filename = "config_TRIMAZKON.xlsx"
        setting_list_name = "Settings_recources"
        Tools.check_config_file(config_filename,setting_list_name)
        self.data_read_in_txt = Tools.read_config_data()
        self.database_downloaded = False
        self.ib_running = False
        self.run_as_admin = False
        #init spínání tray podle nastavení
        if self.data_read_in_txt[24] == "ano":
            task_success = Tools.establish_startup_tray()
            if str(task_success) == "need_access":
                self.run_as_admin = True
        else: # když nezaškrtnuto aut. spouštění ujisti se, že není nastavené - potřeba taky admin
            if Tools.check_task_existence_in_TS("TRIMAZKON_startup_tray_setup"):
                Tools.remove_task_from_TS("TRIMAZKON_startup_tray_setup")
        
    def clear_frames(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        # for frames in self.list_of_menu_frames:
        #     frames.pack_forget()
        #     frames.grid_forget()
        #     frames.destroy()
        # self.list_of_menu_frames = []
        
    def call_sorting_option(self):
        self.clear_frames()
        self.root.unbind("<f>")
        Sorting_option(self.root)

    def call_view_option(self,path_given = None,selected_image = ""):
        self.clear_frames()
        self.root.unbind("<f>")
        self.IB_class = Image_browser(self.root,path_given,selected_image)
        self.ib_running = True

    def call_ip_manager(self):
        self.clear_frames()
        self.root.unbind("<f>")
        IP_manager(self.root)
    
    def call_catalogue_maker(self):
        self.clear_frames()
        self.root.unbind("<f>")
        Catalogue_maker(self.root)

    def call_advanced_option(self,success_message = None):
        self.clear_frames()
        self.root.unbind("<f>")
        Advanced_option(self.root,tray_setting_status_message=success_message)

    def fill_changelog(self,change_log):
        # Iterate through each <string> element and print its text
        for string_element in string_database.change_log_list:
            change_log.insert("current lineend",string_element + "\n")
        change_log.see(tk.END)

    def command_landed(self,command):
        """
        tato funkce přijímá příkazy z pipeline serveru
        """
        print("received in menu: ",command)
        params = command.split(",,")
        print("Image browser running status: ",self.ib_running)
        if self.ib_running == False:
            for widget in self.root.winfo_children():
                widget.destroy()
            self.root.unbind("<Button-1>")
            self.call_view_option(params[0],params[1])
        else:
            print("previous path: ",self.IB_class.image_browser_path)
            print("previous path: ",self.IB_class.IB_as_def_browser_path)
            print("previous image: ",self.IB_class.selected_image)
            print("new path: ",params[0])
            print("new image: ",params[1])

            for widget in self.root.winfo_children():
                widget.destroy()
            self.root.unbind("<Button-1>")
            self.call_view_option(params[0],params[1])

    def menu(self,initial=False,catalogue_downloaded = False,zoom_disable = False): # Funkce spouští základní menu při spuštění aplikace (MAIN)
        """
        Funkce spouští základní menu při spuštění aplikace (MAIN)

        -obsahuje 3 rámce:

        list_of_menu_frames = [frame_with_buttons,frame_with_logo,frame_with_buttons_right]
        """
        self.ib_running = False
        if self.data_read_in_txt[7] == "ano":
            self.root.after(0, lambda:self.root.state('zoomed')) # max zoom, porad v okne
            
        if self.data_read_in_txt[22] == "ne" and initial: # pokud není využito nastavení windows
            try:
                root.after(0, lambda: set_zoom(int(self.data_read_in_txt[21])))
            except Exception as e:
                print("error with menu scaling")

        frame_with_logo = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        # logo = customtkinter.CTkImage(Image.open(initial_path+"images/logo.png"),size=(1200, 100))
        logo = customtkinter.CTkImage(Image.open(Tools.resource_path("images/logo.png")),size=(1200, 100))
        image_logo = customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        
        frame_with_buttons_right = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        frame_with_buttons = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        frame_with_logo.pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        image_logo.pack()
        frame_with_buttons_right.pack(pady=0,padx=0,fill="both",expand=True,side = "right")
        frame_with_buttons.pack(pady=0,padx=0,fill="both",expand=True,side = "left")
        
        IB_as_def_browser_path = None
        # self.list_of_menu_frames = [frame_with_buttons,frame_with_logo,frame_with_buttons_right]
        
        manage_images =         customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Obrázky (správa)", command = lambda: self.call_sorting_option(),font=("Arial",25,"bold"))
        viewer_button =         customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Prohlížeč obrázků", command = lambda: self.call_view_option(),font=("Arial",25,"bold"))
        ip_setting_button =     customtkinter.CTkButton(master= frame_with_buttons, width= 400,height=100, text = "IP setting", command = lambda: self.call_ip_manager(),font=("Arial",25,"bold"))
        catalogue_button =      customtkinter.CTkButton(master= frame_with_buttons, width= 400,height=100, text = "Katalog", command = lambda: self.call_catalogue_maker(),font=("Arial",25,"bold"))
        advanced_button =       customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Nastavení", command = lambda: self.call_advanced_option(),font=("Arial",25,"bold"))
        change_log_label =      customtkinter.CTkLabel(master=frame_with_buttons_right, width= 600,height=50,font=("Arial",24,"bold"),text="Seznam posledně provedených změn: ")
        change_log =            customtkinter.CTkTextbox(master=frame_with_buttons_right, width= 600,height=550,fg_color="#212121",font=("Arial",20),border_color="#636363",border_width=3,corner_radius=0)
        resources_load_error =  customtkinter.CTkLabel(master=frame_with_buttons_right, width= 600,height=50,font=("Arial",24,"bold"),text="Nepodařilo se načíst konfigurační soubor (config_TRIMAZKON.xlsx)",text_color="red")
        manage_images.          pack(pady =(105,0), padx=20,side="top",anchor="e")
        viewer_button.          pack(pady = (10,0), padx=20,side="top",anchor="e")
        ip_setting_button.      pack(pady = (10,0), padx=20,side="top",anchor="e")
        catalogue_button.       pack(pady = (10,0), padx=20,side="top",anchor="e")
        advanced_button.        pack(pady = (10,0), padx=20,side="top",anchor="e")
        change_log_label.       pack(pady = (50,5), padx=20,side="top",anchor="w")
        change_log.             pack(pady =0,       padx=20,side="top",anchor="w")
        if global_recources_load_error:
            resources_load_error.pack(pady = (5,5), padx=20,side="top",anchor="w")

        self.fill_changelog(change_log)
        
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
            self.root.update()
        
        self.root.bind("<f>",maximalize_window)
        # initial promenna aby se to nespoustelo porad do kola pri navratu do menu (system argumenty jsou stále uložené v aplikaci)
        if len(sys.argv) > 1 and initial == True:
            raw_path = str(sys.argv[1])
            #klik na spusteni trimazkonu s admin právy
            if sys.argv[0] == sys.argv[1]:
                self.call_ip_manager()
            elif sys.argv[1] == "settings_tray":
                self.call_advanced_option(success_message="Automatické spouštění úspěšně nastaveno")
            elif sys.argv[1] == "settings_tray_del":
                self.call_advanced_option(success_message="Automatické spouštění úspěšně odstraněno")
            elif sys.argv[1] != "admin_menu" and sys.argv[1] != "trigger_by_tray": # pokud se nerovnají jedná se nejspíše o volání základního prohlížeče obrázků (spuštění kliknutím na obrázek...)
                IB_as_def_browser_path=Tools.path_check(raw_path,True)
                IB_as_def_browser_path_splitted = IB_as_def_browser_path.split("/")
                IB_as_def_browser_path = ""
                for i in range(0,len(IB_as_def_browser_path_splitted)-2):
                    IB_as_def_browser_path += IB_as_def_browser_path_splitted[i]+"/"
                root.update()
                self.root.update()
                selected_image = IB_as_def_browser_path_splitted[len(IB_as_def_browser_path_splitted)-2]
                self.call_view_option(IB_as_def_browser_path,selected_image)
        
        if self.run_as_admin:
            self.root.after(1000, lambda: Tools.call_again_as_admin("admin_menu","Upozornění","Aplikace vyžaduje práva pro nastavení aut. spouštění na pozadí\n\n- přejete si znovu spustit aplikaci, jako administrátor?"))

        self.root.mainloop()

class Image_browser: # Umožňuje procházet obrázky a přitom například vybrané přesouvat do jiné složky
    """
    Umožňuje procházet obrázky a přitom například vybrané přesouvat do jiné složky

    - umožňuje: měnit rychlost přehrávání, přiblížení, otočení obrázku
    - reaguje na klávesové zkratky
    """
    def __init__(self,root,IB_as_def_browser_path = None,selected_image = "",path_given = "",params_given = None):
        self.root = root
        self.path_given = path_given
        self.IB_as_def_browser_path = IB_as_def_browser_path
        self.all_images = []
        self.increment_of_image = 0
        self.state = "stop"
        self.rotation_angle = 0.0
        text_file_data = Tools.read_config_data()
        self.config_path = text_file_data[2]
        list_of_dir_names = text_file_data[9]
        self.copy_dir = list_of_dir_names[5]
        self.move_dir = list_of_dir_names[6]
        self.chosen_option = text_file_data[11][0]
        self.zoom_increment = text_file_data[11][1]
        self.drag_increment = text_file_data[11][2]
        self.number_of_film_images = text_file_data[14]
        self.image_browser_path = ""
        self.unbind_list = []
        self.image_extensions = ['.jpg', '.jpeg', '.jpe', '.jif', '.jfif', '.jfi',
                    '.png', '.gif', '.bmp', '.tiff', '.tif', '.ico', '.webp',
                    '.raw', '.cr2', '.nef', '.arw', '.dng', ".ifz"]
        self.previous_zoom = 1
        self.selected_image = selected_image
        self.path_for_explorer = None
        self.temp_bmp_folder = "temp_bmp"
        self.ifz_located = None
        self.converted_images = []
        self.increment_of_ifz_image = 0
        self.default_path = ""
        self.previous_height = 0
        self.previous_width = 0
        if text_file_data[13] == "ne":
            self.image_film = False
        else:
            self.image_film = True
        self.image_queue = [""]*((self.number_of_film_images*2)+1)
        self.flow_direction = ""
        self.ifz_count = 1
        self.count_of_ifz_images_defined = False
        self.name_hide_index = 0
        self.main_image = None
        self.main_image2 = None
        self.drag_option_binded = False
        self.drawing_color = "#000000"
        self.drawing_thickness = 5
        self.draw_mode = "line"
        self.x_growth_multiplier = 0.5
        self.y_growth_multiplier = 0.5
        self.image_dimensions = (0,0)
        self.last_coords = (0,0)
        self.zoom_given = 100
        self.settings_applied = False
        self.loaded_image_status = True
        self.inserted_path_history = []
        self.last_frame_dim = [0,0]
        self.last_resize_time = 0

        if params_given != None:
            print("params given",params_given)
            coords_given = params_given[0]
            zoom_given = params_given[1]
            self.last_coords = coords_given
            self.zoom_given = zoom_given
            self.settings_applied = True
        self.create_widgets()
        
    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu
        """
        #kdyby probihala sekvence obrazku:
        if self.state == "running":
            self.stop()

        list_of_frames = [self.main_frame,self.frame_with_path,self.frame_with_console,self.frame_with_buttons,self.background_frame,self.image_film_frame_center,self.image_film_frame_right,self.image_film_frame_left]
        for frames in list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()

        for keys in self.unbind_list:
            self.root.unbind(keys)
        
        if os.path.exists(self.default_path + self.temp_bmp_folder):
            shutil.rmtree(self.default_path + self.temp_bmp_folder) # vycistit
        menu.menu()
    
    def clear_frame(self,frame):
        try:
            children = frame.winfo_children()
        except Exception:
            return
        for widget in children:
            widget.destroy()

    def get_images(self,path):  # Seznam všech obrázků v podporovaném formátu (včetně cesty)
        """
        Seznam všech obrázků v podporovaném formátu (včetně cesty)
        """
        self.ifz_located = None
        list_of_files_to_view = []
        for files in os.listdir(path):
            files_split = files.split(".")
            if ("."+files_split[len(files_split)-1]) in self.image_extensions:
                list_of_files_to_view.append(path + files)
                if ("."+files_split[len(files_split)-1]) == ".ifz":
                    self.ifz_located = True
        return list_of_files_to_view

    def convert_files(self):
        if self.image_film == False:
            name_of_file = self.all_images[self.increment_of_image].split("/")
            name_of_file = name_of_file[len(name_of_file)-1]
            #vytvoreni temp slozky:
            if os.path.exists(self.default_path + self.temp_bmp_folder):
                shutil.rmtree(self.default_path + self.temp_bmp_folder) # vycistit
                os.mkdir(self.default_path + self.temp_bmp_folder)
            else:
                os.mkdir(self.default_path + self.temp_bmp_folder)

            Converting.whole_converting_function(self.default_path,"bmp",self.temp_bmp_folder,None,True,name_of_file)
            self.converted_images = []
            for files in os.listdir(self.default_path + self.temp_bmp_folder):
                if (self.default_path + self.temp_bmp_folder + "/" + files) not in self.converted_images:
                    self.converted_images.append(self.default_path + self.temp_bmp_folder + "/" + files)

        elif self.image_film == True:
            #uvazuje se, ze konvertovane obrazky budou koncit: _x.bmp, kde x bude nabyvat maximalne hodnot 0-8 nebo bez _x ...
            #1) uprava nazvu pro konvertovani
            names_of_files_to_be_converted = []
            num_of_preload_images = len(self.image_queue)-1
            for i in range(0,len(self.image_queue)):
                image_index = int(self.increment_of_image+i-(num_of_preload_images/2))
                if image_index < 0:
                    image_index = len(self.all_images) + image_index
                elif image_index > len(self.all_images)-1:
                    image_index = 0 + (image_index-len(self.all_images))
                name_of_file = self.all_images[image_index].split("/")
                name_of_file = name_of_file[len(name_of_file)-1]
                names_of_files_to_be_converted.append(name_of_file)
            #2) vytvareni temp slozky
            found_files=[]
            if not os.path.exists(self.default_path + self.temp_bmp_folder):
                os.mkdir(self.default_path + self.temp_bmp_folder)
            else:
                for files in os.listdir(self.default_path + self.temp_bmp_folder):
                    if ".bmp" in files and files[-6:-5] == "_":
                        found_files.append(files[:-6])
                    elif ".bmp" in files:
                        found_files.append(files[:-4])
            #3) check jestli uz neni konvertovane
            to_convert =[]
            for i in range(0,len(names_of_files_to_be_converted)):
                if names_of_files_to_be_converted[i][:-4] not in found_files: #_x (x muze nabyvat max hodnoty 8)
                    to_convert.append(names_of_files_to_be_converted[i])

            #4) volani funkce pro konvertovani
            Converting.whole_converting_function(self.default_path,"bmp",self.temp_bmp_folder,None,True,to_convert)

            #5) definice poctu ifz
            if self.count_of_ifz_images_defined == False:
                self.name_hide_index = 0
                largest_ifz_num = 1
                for files in os.listdir(self.default_path + self.temp_bmp_folder):
                    # urcovani poctu ifz v jednom souboru (vykonat pouze jednou pro vsechny soubory
                    if files[-6:-5] == "_":
                        if files[-5:-4].isdigit():
                            if int(files[-5:-4]) +1 > largest_ifz_num:
                                largest_ifz_num = int(files[-5:-4])+1
                    else:
                        #nenalezeny zadne dalsi ifz podobrazky
                        self.name_hide_index = 2
                self.ifz_count = largest_ifz_num
                self.count_of_ifz_images_defined = True

            #6) mazani nepotrebnych
            for files in os.listdir(self.default_path + self.temp_bmp_folder):
                if (str(files[:(-6+self.name_hide_index)])+".ifz") not in names_of_files_to_be_converted:
                    os.remove(self.default_path + self.temp_bmp_folder + "/" + files)
                    # print("deleting",files)

            #8) plneni pole s kompletni cestou ve spravnem poradi...
            self.converted_images = []
            for i in range(0,len(names_of_files_to_be_converted)):
                matches_made = 0
                for files in os.listdir(self.default_path + self.temp_bmp_folder):
                    if ((self.default_path + self.temp_bmp_folder + "/" + files) not in self.converted_images):
                        if names_of_files_to_be_converted[i] == files[:(-6+self.name_hide_index)]+".ifz":
                            self.converted_images.append(self.default_path + self.temp_bmp_folder + "/" + files)   
                            matches_made += 1
                            #pro zefektivnění:
                            if matches_made == self.ifz_count:
                                break
                             
    def make_image_film_widgets(self):
        def mouse_wheel2(e): # posouvat obrazky
            direction = -e.delta
            if direction < 0:
                self.next_image()
            else:
                self.previous_image()
            
        if len(self.image_queue)>len(self.all_images):
            self.image_queue = [""]*(len(self.all_images))
            if len(self.image_queue) % 2 == 0: #nesmi byt sudé...
                self.image_queue.append("") #kdyztak prictu jeste jeden prvek... append
        else:
            self.image_queue = [""]*((self.number_of_film_images*2)+1)
        
        self.left_labels = [""]*int((len(self.image_queue)-1)/2)
        self.right_labels = [""]*int((len(self.image_queue)-1)/2)
        self.clear_frame(self.image_film_frame_left)
        self.clear_frame(self.image_film_frame_right)

        for i in range(0,len(self.left_labels)):
            self.left_labels[len(self.left_labels)-i-1] = customtkinter.CTkLabel(master = self.image_film_frame_left,text = "")
            self.left_labels[len(self.left_labels)-i-1].bind("<MouseWheel>",mouse_wheel2)
            self.left_labels[len(self.left_labels)-i-1].pack(side = "right",padx=5)
            self.right_labels[i] = customtkinter.CTkLabel(master = self.image_film_frame_right,text = "")
            self.right_labels[i].bind("<MouseWheel>",mouse_wheel2)
            self.right_labels[i].pack(side = "left",padx=5)

    def start(self,path): # Ověřování cesty, init, spuštění
        """
        Ověřování cesty, init, spuštění
        """
        path_found = True
        self.count_of_ifz_images_defined = False
        if path == "" or path == "/": #pripad, ze bylo pouzito tlacitko spusteni manualne vlozene cesty a nebo je chyba v config souboru
            path_found = False
            path = self.path_set.get()
            if path != "":
                check = Tools.path_check(path)
                if check == False:
                    Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" nebyla nalezena","red",None,True)
                else:
                    path = check
                    path_found = True
            else:
                Tools.add_colored_line(self.console,"Nebyla vložena cesta k souborům","red",None,True)

        if os.path.isdir(path) == False: # pokud se nejedna o slozku - je mozne, ze je vlozeny nazev souboru pro zobrazeni, jako prvni
            if os.path.exists(path):
                if path.endswith("/"):
                    path = path[0:len(path)-1]
                path_splitted = path.split("/")
                self.selected_image = path_splitted[len(path_splitted)-1]
                new_path = ""
                i = 0
                for frags in path_splitted:
                    i += 1
                    if i < len(path_splitted):
                        new_path = new_path + frags + "/"

                path = new_path
            else:
                path_found = False
                Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" neobsahuje žádné obrázky","red",None,True)
        #automaticky okamzite otevre prvni z obrazku v dane ceste
        if path_found == True:
            if os.path.exists(path):
                #path = path.replace(" ","")
                while path.endswith(" "):
                    path = path[:len(path)-1]

                if path.endswith("/") == False:
                    path = path + "/"
                    self.path_set.delete("0","300")
                    self.path_set.insert("0",path)
                #oprava mezery v nazvu
                full_path = r"{}".format(path)
                path = full_path
                self.path_for_explorer = path
                self.all_images = self.get_images(path)
                if len(self.all_images) != 0:
                    self.image_browser_path = path
                    path_to_add_to_history = self.image_browser_path
                    # Tools.add_colored_line(self.console,f"Vložena cesta: {path}","green",None,True)
                    if self.image_film == True:
                        self.make_image_film_widgets()
                    if self.ifz_located == None:
                        if self.selected_image == "": 
                            #zobrazit hned prvni obrazek po vlozene ceste
                            self.increment_of_image = 0
                        else:
                            #zobrazit obrazek vybrany v exploreru
                            self.increment_of_image = self.all_images.index(path+self.selected_image)

                        self.view_image(self.increment_of_image)
                        self.current_image_num.configure(text ="/" + str(len(self.all_images)))
                        self.changable_image_num.delete("0","100")
                        self.changable_image_num.insert("0", str(self.increment_of_image+1))
                    else: # Nejprve prevest do bmp formatu
                        if self.selected_image == "": 
                            #zobrazit hned prvni obrazek po vlozene ceste
                            self.increment_of_image = 0
                        else:
                            #zobrazit obrazek vybrany v exploreru
                            self.increment_of_image = self.all_images.index(path+self.selected_image)
                        self.default_path = r"{}".format(path)
                        path_to_add_to_history = self.default_path
                        self.convert_files()
                        center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                        self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image])
                        self.current_image_num.configure(text ="/" + str(len(self.all_images)))
                        self.changable_image_num.delete("0","100")
                        self.changable_image_num.insert("0", str(self.increment_of_image+1))
                        self.current_image_num_ifz.configure(text ="/" + str(self.ifz_count))
                        self.changable_image_num_ifz.delete("0","100")
                        self.changable_image_num_ifz.insert("0", str(self.increment_of_ifz_image+1))
                    if path_to_add_to_history not in self.inserted_path_history:
                        self.inserted_path_history.append(path_to_add_to_history)
                else:
                    Tools.add_colored_line(self.console,"- V zadané cestě nebyly nalezeny obrázky","red",None,True)
            else:
                Tools.add_colored_line(self.console,"- Vložená cesta je neplatná","red",None,True)

    def call_browseDirectories(self): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        if self.path_for_explorer != None:
            output = Tools.browseDirectories("all",self.path_for_explorer)
        else:
            output = Tools.browseDirectories("all")
        if str(output[1]) != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", output[1])
            Tools.add_colored_line(self.console,f"Byla vložena cesta: {output[1]}","white",None,True)
            self.selected_image = output[2]
            
            if os.path.exists(self.default_path + self.temp_bmp_folder):
                shutil.rmtree(self.default_path + self.temp_bmp_folder) # vycistit
            self.increment_of_ifz_image = 0
            self.changable_image_num_ifz.delete("0","100")
            self.changable_image_num_ifz.insert("0",0)
            # self.image_queue = [""]*((self.number_of_film_images*2)+1)
            if len(self.image_queue)>len(self.all_images):
                self.image_queue = [""]*(len(self.all_images))
                if len(self.image_queue) % 2 == 0: #nesmi byt sudé...
                    self.image_queue.append("") #kdyztak prictu jeste jeden prvek... append
            else:
                self.image_queue = [""]*((self.number_of_film_images*2)+1)
            self.converted_images = []
            self.start(output[1])

    def get_frame_dimensions(self): # Vrací aktuální rozměry rámečku
        """
        Vrací aktuální rozměry rámečku
        """
        whole_app_height = self.root._current_height
        whole_app_width = self.root._current_width
        width = whole_app_width
        self.frame_with_path.update_idletasks()
        self.image_film_frame_center.update_idletasks()
        height = whole_app_height-self.frame_with_path._current_height-30
        if self.image_film == True:
            height = height - self.image_film_frame_center._current_height
        return [width, height]

    def calc_current_format(self,width,height,new_window_status = False,frame_dim_given = None): # Přepočítávání rozměrů obrázku do rozměru rámce podle jeho formátu + zooming
        """
        Přepočítávání rozměrů obrázku do rozměru rámce podle jeho formátu

        -vstupními daty jsou šířka a výška obrázku
        -přepočítávání pozicování obrázku a scrollbarů v závislosti na zoomu
        """

        if new_window_status:
            frame_dimensions = frame_dim_given
        else:
            frame_dimensions = self.get_frame_dimensions()
        self.zoom_slider.update_idletasks()
        zoom = self.zoom_slider.get() / 100
        frame_width, frame_height = frame_dimensions
        image_width = width
        image_height = height
        image_ratio = image_width / image_height

        def rescale_image(): # Vmestnani obrazku do velikosti aktualni velikosti ramce podle jeho formatu
            if image_height > image_width:
                new_height = frame_height
                if image_width > frame_width:
                    new_width = int(new_height * image_ratio)
                else:
                    new_width = frame_width
                    if image_width > frame_width:
                        new_width = image_width
                    new_height = int(new_width / image_ratio)

            elif image_height < image_width:
                new_width = frame_width

                if image_height < frame_height:
                    new_height = int(new_width / image_ratio)
                else:
                    new_height = frame_height
                    if image_height < frame_height:
                        new_height = image_height
                    new_width = int(new_height * image_ratio)

            elif image_height == image_width:
                new_height = frame_height
                new_width = new_height

            #doublecheck
            if new_height > frame_height:
                new_height = frame_height
                new_width = int(new_height * image_ratio)
            if new_width > frame_width:
                new_width = frame_width
                new_height = int(new_width / image_ratio)

            return (new_height,new_width)
        
        new_height, new_width = rescale_image()

        if not new_window_status:
            new_height = new_height * zoom
            new_width = new_width * zoom
            self.previous_zoom = zoom
        
        self.main_frame.update()
        self.zoom_grow_x = max(new_width-self.previous_width,self.previous_width-new_width)
        self.zoom_grow_y = max(new_height-self.previous_height,self.previous_height-new_height)
        self.previous_height = new_height
        self.previous_width = new_width

        return [new_width, new_height]
    
    def view_image(self,increment_of_image,direct_path = None,only_refresh=None,reset = False,reload_buffer = False,only_next_ifz = False,in_new_window=False): # Samotné zobrazení obrázku
        """
        Samotné zobrazení obrázku

        -vstupními daty jsou informace o pozici obrázku v poli se všemi obrázky
        -přepočítávání rotace
        """
        
        if not only_next_ifz:
            self.loaded_image_status = False
        
        def corrupted_image_handling():
            with Image.open(Tools.resource_path("images/loading3.png")) as opened_image:
                rotated_image = opened_image.rotate(180,expand=True)
            resized = rotated_image.resize(size=(800,800))
            error_image = ImageTk.PhotoImage(resized)
            if self.main_frame.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                self.main_frame.delete("lower")
                self.main_image = self.main_frame.create_image(0, 0, anchor=tk.NW, image=error_image,tag ="raise")
                self.main_frame.tag_lower(self.main_image)
                self.main_frame.update()

        def check_image_growth_boundaries():
            frame_dimensions = self.get_frame_dimensions()
            nonlocal current_coords
            x_growth = self.zoom_grow_x*self.x_growth_multiplier
            y_growth = self.zoom_grow_y*self.y_growth_multiplier
            minus_x_boundary = frame_dimensions[0] - dimensions[0]
            minus_y_boundary = frame_dimensions[1] - dimensions[1]

            if self.x_growth_multiplier > 0: # tzn. jsme s mysi nalevo 1 až 0.5
                if current_coords[0]+x_growth < 0:
                    x_coords = current_coords[0]+x_growth
                else:
                    x_coords = 0
            else: # napravo, -0.5 až -1
                if current_coords[0]+x_growth < minus_x_boundary:
                    x_coords = minus_x_boundary
                else:
                    x_coords = current_coords[0]+x_growth

            if self.y_growth_multiplier > 0: # tzn. jsme s mysi nahore 1 sž 0.5
                if current_coords[1]+y_growth < 0:
                    y_coords = current_coords[1]+y_growth
                else:
                    y_coords = 0
            else: # -0.5 až -1
                if current_coords[1]+y_growth < minus_y_boundary:
                    y_coords = minus_y_boundary
                else:
                    y_coords = current_coords[1]+y_growth

            return (x_coords, y_coords)

        def make_image_strip(main_image):
            try:
                image_center= customtkinter.CTkImage(main_image,size = ((150,150)))
                if self.image_film_frame_center.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                    self.images_film_center.configure(image = image_center)
                    self.images_film_center.image = image_center
                    self.images_film_center.update_idletasks()
            except Exception as e: 
                print("moc rychle: ",e)
                
            
            if only_refresh == None: # jen pokud rotuju obrazek, aktualizuj prostřední
                def open_image(increment_of_image_given,position):
                    try:
                        if self.ifz_located == None:
                            increment_of_image = increment_of_image_given + position
                            number_of_found_images = len(self.all_images)
                            if increment_of_image < 0:
                                increment_of_image = number_of_found_images + increment_of_image

                            elif increment_of_image > number_of_found_images-1:
                                increment_of_image = 0 + (increment_of_image-number_of_found_images)

                            if len(self.all_images) > abs(increment_of_image):
                                image_to_show = self.all_images[increment_of_image]
                                with Image.open(image_to_show) as current_image:

                                    opened_image = current_image.rotate(self.rotation_angle,expand=True)
                                return opened_image
                            else:
                                return False
                        elif self.ifz_located == True:
                            converted_images_index = (position * self.ifz_count) #+ self.increment_of_ifz_image
                            image_to_show = self.converted_images[converted_images_index + self.increment_of_ifz_image]
                            with Image.open(image_to_show) as current_image:
                                opened_image = current_image.rotate(self.rotation_angle,expand=True)

                            return opened_image
                    except Exception as e:
                        error_message = f"Obrázek: {image_to_show} je poškozen"
                        print(error_message)
                        return False

                if reload_buffer:
                    if len(self.image_queue)>len(self.all_images):
                        self.image_queue = [""]*(len(self.all_images))
                        if len(self.image_queue) % 2 == 0: #nesmi byt sudé...
                            self.image_queue.append("") #kdyztak prictu jeste jeden prvek... append
                    else:
                        self.image_queue = [""]*((self.number_of_film_images*2)+1)

                image_film_dimensions = [80,80]
                half_image_queue = int(len(self.image_queue)/2)

                if "" in self.image_queue: #kdyz jeste nejsou zadne poukladane, preloading
                    #CENTER image preload
                    self.image_queue[half_image_queue] = customtkinter.CTkImage(main_image,size = (image_film_dimensions[0],image_film_dimensions[1]))
                    for i in range(0,half_image_queue): #LEFT
                        current_image = open_image(increment_of_image,-half_image_queue+i)
                        if current_image != False:
                            self.image_queue[i] = customtkinter.CTkImage(current_image,size = (image_film_dimensions[0],image_film_dimensions[1]))
                        
                    for i in range(0,half_image_queue): #RIGHT
                        current_image = open_image(increment_of_image,+i+1)
                        if current_image != False:
                            self.image_queue[i+half_image_queue+1] = customtkinter.CTkImage(current_image,size = (image_film_dimensions[0],image_film_dimensions[1]))
                else:
                    if self.flow_direction == "left":
                        current_image = open_image(increment_of_image,-half_image_queue)
                        if current_image != False:
                            preopened_image = customtkinter.CTkImage(current_image,size = (image_film_dimensions[0],image_film_dimensions[1]))
                            self.image_queue.pop(len(self.image_queue)-1)
                            self.image_queue.insert(0,preopened_image)

                    elif self.flow_direction == "right":
                        current_image = open_image(increment_of_image,half_image_queue)
                        if current_image != False:
                            preopened_image = customtkinter.CTkImage(current_image,size = (image_film_dimensions[0],image_film_dimensions[1]))
                            self.image_queue.append(preopened_image)
                            self.image_queue.pop(0)

                if self.image_film_frame_left.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                    for i in range(0,half_image_queue):
                        try:
                            self.left_labels[i].configure(image = self.image_queue[i],padx=10)
                            self.left_labels[i].image = self.image_queue[i]
                            self.left_labels[i].update_idletasks()
                        except Exception as e: 
                            print("moc rychle: ",e)
                            

                if self.image_film_frame_right.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                    for i in range(0,half_image_queue):
                        try:
                            self.right_labels[i].configure(image = self.image_queue[half_image_queue+i+1],padx=10)
                            self.right_labels[i].image = self.image_queue[half_image_queue+i+1]
                            self.right_labels[i].update_idletasks()
                        except Exception as e:
                            print("moc rychle: ",e)                       

        if len(self.all_images) != 0:
            if direct_path == None:
                image_to_show = self.all_images[increment_of_image]
            else:
                image_to_show = direct_path

            try:
                with Image.open(image_to_show) as opened_image:
                    rotated_image = opened_image.rotate(self.rotation_angle,expand=True)
                    width,height = rotated_image.size
            except Exception as e:
                error_message = f"Obrázek: {image_to_show} je poškozen"
                print(error_message)
                corrupted_image_handling()
                return error_message


            if in_new_window:
                def resize_image(event, label, original_image,frame_given = False):
                    if frame_given:
                        frame_dim = [int(event[0]),int(event[1])]
                    else:
                        frame_dim = [int(event.width),int(event.height)]

                    if frame_dim == self.last_frame_dim:
                        return

                    new_width, new_height = original_image.size
                    dimensions = self.calc_current_format(new_width,new_height,True,frame_dim)
                    

                    resized_image = original_image.resize((int(dimensions[0]),int(dimensions[1])))
                    photo = ImageTk.PhotoImage(resized_image)
                    label.configure(image=photo)
                    label.image = photo  # Keep a reference to avoid garbage collect
                    self.last_frame_dim = [frame_dim[0],frame_dim[1]]
                
                child_root = customtkinter.CTkToplevel()
                child_root.after(200, lambda: child_root.iconbitmap(app_icon))
                child_root.title(image_to_show)
                with Image.open(image_to_show) as opened_image:
                    rotated_image = opened_image.rotate(self.rotation_angle,expand=True)
                photo = ImageTk.PhotoImage(rotated_image)
                label = customtkinter.CTkLabel(child_root, image=photo, text="")
                label.pack(fill="both", expand=True)
                label.image = photo
                child_root.bind("<Configure>", lambda event, window_label = label, window_image = rotated_image: resize_image(event, window_label, window_image))
                child_root.update()
                child_root.update_idletasks()
                child_root.geometry(f"1200x800+{300}+{300}")
                resize_image([1200,800], label, rotated_image,frame_given=True)
                child_root.after(100,child_root.focus_force())
                self.loaded_image_status = True
                return
            
            dimensions = self.calc_current_format(width,height)
            resized = rotated_image.resize(size=(int(dimensions[0]),int(dimensions[1])))
            self.image_dimensions = (int(dimensions[0]),int(dimensions[1]))
            self.tk_image = ImageTk.PhotoImage(resized)
            self.main_frame.itemconfig(self.main_image, image=self.tk_image)
            if self.main_frame.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                try:
                    current_coords = self.main_frame.coords(self.main_image)
                except Exception:
                    reset = True

                if reset:
                    current_coords = [0,0]
                    self.zoom_grow_x=0
                    self.zoom_grow_y=0

                if not self.settings_applied:
                    x_coords, y_coords = check_image_growth_boundaries()
                else:
                    x_coords, y_coords = self.last_coords
                    self.settings_applied = False

                self.main_frame.update_idletasks()
                self.main_frame.delete("lower")
                self.main_image = self.main_frame.create_image(x_coords, y_coords,anchor=tk.NW, image=self.tk_image,tag = "lower")
                self.main_frame.tag_lower(self.main_image)
                self.last_coords = (x_coords,y_coords)
                # self.main_frame.update()

                if self.image_film == True: #refreshujeme pouze stredovy obrazek jinak i okolni
                    # run_background = threading.Thread(target=make_image_strip, args=(rotated_image,),daemon = True)
                    # run_background = threading.Thread(target=make_image_strip, args=(rotated_image,))
                    # run_background.start()
                    make_image_strip(rotated_image)

        self.loaded_image_status = True

    def next_image(self,silent=False,reload_buffer =False): # Další obrázek v pořadí (šipka vpravo)
        """
        Další obrázek v pořadí (šipka vpravo)
        """
        if not self.loaded_image_status:
            return
        load_status = None
        self.flow_direction = "right"
        number_of_found_images = len(self.all_images)
        if number_of_found_images != 0:
            if self.increment_of_image < number_of_found_images -1:
                self.increment_of_image += 1
            else:
                self.increment_of_image = 0
            
            if self.ifz_located == None:
                load_status = self.view_image(self.increment_of_image,reload_buffer = reload_buffer)
            else:
                self.convert_files()
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],reload_buffer = reload_buffer)
            
            if self.main_frame.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                self.current_image_num.configure(text ="/" + str(len(self.all_images)))
                self.changable_image_num.delete("0","100")
                self.changable_image_num.insert("0", str(self.increment_of_image+1))
                if silent == False:
                    if load_status != None:
                        Tools.add_colored_line(self.console,load_status,"orange",None,True)
                    elif self.name_or_path.get() == 1:
                        only_name = str(self.all_images[self.increment_of_image]).split("/")
                        only_name = only_name[int(len(only_name))-1]
                        Tools.add_colored_line(self.console,str(only_name),"white",None,True)
                    else:
                        Tools.add_colored_line(self.console,str(self.all_images[self.increment_of_image]),"white",None,True)

                self.current_image_num_ifz.configure(text ="/" + str(self.ifz_count))
                self.changable_image_num_ifz.delete("0","100")
                self.changable_image_num_ifz.insert("0", str(self.increment_of_ifz_image+1))
    
    def next_ifz_image(self): # Další ifz obrázek v pořadí
        number_of_found_images = self.ifz_count
        if number_of_found_images != 0 and number_of_found_images != 1:
            if self.increment_of_ifz_image < number_of_found_images -1:
                self.increment_of_ifz_image += 1
            else:
                self.increment_of_ifz_image = 0

            center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
            # load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True,reload_buffer=True)
            load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],reload_buffer=True,only_next_ifz = True)
            if self.main_frame.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                self.current_image_num_ifz.configure(text ="/" + str(self.ifz_count))
                self.changable_image_num_ifz.delete("0","100")
                self.changable_image_num_ifz.insert("0", str(self.increment_of_ifz_image+1))
                if load_status != None:
                    Tools.add_colored_line(self.console,load_status,"orange",None,True)
                elif self.name_or_path.get() == 1:
                    only_name = str(self.converted_images[center_image_index + self.increment_of_ifz_image]).split("/")
                    only_name = only_name[int(len(only_name))-1]
                    Tools.add_colored_line(self.console,str(only_name),"white",None,True)
                else:
                    Tools.add_colored_line(self.console,str(self.converted_images[center_image_index + self.increment_of_ifz_image]),"white",None,True)

    def previous_image(self): # Předchozí obrázek v pořadí (šipka vlevo)
        """
        Předchozí obrázek v pořadí (šipka vlevo)
        """
        if not self.loaded_image_status:
            return
        self.flow_direction = "left"
        load_status = None
        number_of_found_images = len(self.all_images)
        if number_of_found_images != 0:
            if self.increment_of_image > 0:
                self.increment_of_image -= 1
            else:
                self.increment_of_image = number_of_found_images -1
            
            if self.ifz_located == None:
                load_status = self.view_image(self.increment_of_image)
            else:
                self.convert_files()
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image])
            self.current_image_num.configure(text = "/" + str(len(self.all_images)))
            self.changable_image_num.delete("0","100")
            self.changable_image_num.insert("0", str(self.increment_of_image+1))
            if load_status != None:
                Tools.add_colored_line(self.console,load_status,"orange",None,True)
            elif self.name_or_path.get() == 1:
                only_name = str(self.all_images[self.increment_of_image]).split("/")
                only_name = only_name[int(len(only_name))-1]
                Tools.add_colored_line(self.console,str(only_name),"white",None,True)
            else:
                Tools.add_colored_line(self.console,str(self.all_images[self.increment_of_image]),"white",None,True)

            self.current_image_num_ifz.configure(text ="/" + str(self.ifz_count))
            self.changable_image_num_ifz.delete("0","100")
            self.changable_image_num_ifz.insert("0", str(self.increment_of_ifz_image+1))

    def previous_ifz_image(self): # předešlý ifz obrázek v pořadí
        number_of_found_images = self.ifz_count
        if number_of_found_images != 0 and number_of_found_images != 1:
            if self.increment_of_ifz_image > 0:
                self.increment_of_ifz_image -= 1
            else:
                self.increment_of_ifz_image = number_of_found_images -1     
            
            center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
            # load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True,reload_buffer=True)
            load_status = self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],reload_buffer=True,only_next_ifz = True)
            if self.main_frame.winfo_exists(): # kdyz se prepina do menu a bezi sekvence
                self.current_image_num_ifz.configure(text ="/" + str(self.ifz_count))
                self.changable_image_num_ifz.delete("0","100")
                self.changable_image_num_ifz.insert("0", str(self.increment_of_ifz_image+1))
                if load_status != None:
                    Tools.add_colored_line(self.console,load_status,"orange",None,True)
                elif self.name_or_path.get() == 1:
                    only_name = str(self.converted_images[self.increment_of_ifz_image]).split("/")
                    only_name = only_name[int(len(only_name))-1]
                    Tools.add_colored_line(self.console,str(only_name),"white",None,True)
                else:
                    Tools.add_colored_line(self.console,str(self.converted_images[self.increment_of_ifz_image]),"white",None,True)
  
    def stop(self):
        self.state = "stop"
        self.button_play_stop.configure(text = "SPUSTIT")
        self.button_play_stop.configure(command = lambda: self.play())

    def play(self):
        def load_image_loop():
            speed=self.speed_slider.get()/100
            calculated_time = 2000-speed*2000 # 1% dela necele 2 sekundy, 100%, nula sekund, maximalni vykon
            self.next_image()
            if self.state != "stop":
                self.main_frame.update_idletasks()
                # if self.ifz_located and int(calculated_time) < 200:
                #     calculated_time = 200
                # if int(calculated_time) < 20:
                #     calculated_time = 20
                self.root.after(int(calculated_time),load_image_loop)

        self.state = "running"
        self.button_play_stop.configure(text = "STOP")
        self.button_play_stop.configure(command = lambda: self.stop())

        thread = threading.Thread(target=load_image_loop)
        thread.start()
  
    def update_speed_slider(self,*args):
        new_value = int(*args)
        self.percent1.configure(text = "")
        self.percent1.configure(text=str(new_value) + " %")

    def update_zoom_slider(self,*args):
        new_value = int(*args)
        self.percent2.configure(text = "")
        self.percent2.configure(text=str(new_value) + " %")
        # update image po zoomu
        if len(self.all_images) != 0:
            if self.ifz_located == True:
                if len(self.converted_images) != 0:
                    center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                    self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True,reset=True)
            else:
                self.view_image(self.increment_of_image,None,True,reset=True)

    def copy_image(self,path): # Tlačítko Kopír., zkopíruje daný obrázek do složky v dané cestě
        """
        Tlačítko Kopír., zkopíruje daný obrázek do složky v dané cestě

        -název složky přednastaven na Kopírované_obrázky
        -vlastnosti obrázku nijak nemění
        """
        image_path = self.all_images[self.increment_of_image]
        image = str(image_path).replace(path,"")
        if not os.path.exists(path + "/" + self.copy_dir):
            os.mkdir(path+ "/" + self.copy_dir)
        if not os.path.exists(path + "/" + self.copy_dir+ "/" + image):
            shutil.copy(path+ "/" + image,path + "/" + self.copy_dir+ "/" + image)
            if self.name_or_path.get() == 1:
                Tools.add_colored_line(self.console,f"Obrázek zkopírován do zvláštní složky: \"{self.copy_dir}\".  ({image})","white",None,True)
            else:
                Tools.add_colored_line(self.console,f"Obrázek zkopírován do zvláštní složky: \"{self.copy_dir}\".  ({image_path})","white",None,True)

        else:
            if self.name_or_path.get() == 1:
                Tools.add_colored_line(self.console,f"Obrázek je již zkopírovaný uvnitř složky: {self.copy_dir}.  ({image})","red",None,True)
            else:
                Tools.add_colored_line(self.console,f"Obrázek je již zkopírovaný uvnitř složky: {self.copy_dir}.  ({image_path})","red",None,True)
                
    def move_image(self): # Tlačítko Přesun., přesune daný obrázek do složky v dané cestě
        """
        Tlačítko Přesun., přesune daný obrázek do složky v dané cestě

        -název složky přednastaven na Přesunuté_obrázky
        """
        image_path = self.all_images[self.increment_of_image]
        image = str(image_path).replace(self.image_browser_path,"")
        if not os.path.exists(self.image_browser_path + "/" + self.move_dir):
            os.mkdir(self.image_browser_path+ "/" + self.move_dir)
        if not os.path.exists(self.image_browser_path + "/" + self.move_dir+ "/" + image):
            shutil.move(self.image_browser_path+ "/" + image,self.image_browser_path + "/" + self.move_dir+ "/" + image)
            if self.name_or_path.get() == 1:
                Tools.add_colored_line(self.console,f"Obrázek přesunut do zvláštní složky: \"{self.move_dir}\".  ({image})","white",None,True)
            else:
                Tools.add_colored_line(self.console,f"Obrázek přesunut do zvláštní složky: \"{self.move_dir}\".  ({image_path})","white",None,True)
            self.all_images.pop(self.increment_of_image) # odstraneni z pole
            self.current_image_num.configure(text ="/" + str(len(self.all_images))) # update maximalniho poctu obrazku
            self.increment_of_image -=1
            self.next_image(True)

    def delete_image(self): # Tlačítko SMAZAT
        image_path = self.all_images[self.increment_of_image]
        image = str(image_path).replace(self.image_browser_path,"")
        if os.path.exists(image_path):
            if self.name_or_path.get() == 1:
                Tools.add_colored_line(self.console,f"Právě byl smazán obrázek: {image}","orange",None,True)
            else:
                Tools.add_colored_line(self.console,f"Právě byl smazán obrázek: {image_path}","orange",None,True)

            os.remove(image_path)
            self.all_images.pop(self.increment_of_image) # odstraneni z pole
            self.current_image_num.configure(text ="/" + str(len(self.all_images))) # update maximalniho poctu obrazku
            self.increment_of_image -=1
            self.next_image(True,reload_buffer=True)

    def rotate_image(self):
        angles = [90.0,180.0,270.0,0.0]
        if self.rotation_angle < 270:
            self.rotation_angle += 90.0
        else:
            self.rotation_angle = 0.0
        if self.ifz_located == True:
            if len(self.converted_images) != 0:
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                # self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True,reset=True)
                self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],reset=True,reload_buffer=True)
        else:
            # self.view_image(self.increment_of_image,None,True,reset=True)
            self.view_image(self.increment_of_image,None,reset=True,reload_buffer=True)
    
    def Reset_all(self): # Vrátí všechny slidery a natočení obrázku do původní polohy
        """
        Vrátí všechny slidery a natočení obrázku do původní polohy
        """
        self.rotation_angle = 0.0
        self.zoom_slider.set(100)
        self.update_zoom_slider(100)
        self.speed_slider.set(100)
        self.update_speed_slider(100)
        if self.ifz_located == True:
            if len(self.converted_images) != 0:
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True,reset=True)
        else:
            self.view_image(self.increment_of_image,None,True,reset=True)
        self.root.update_idletasks()
        self.main_frame.update_idletasks()
    
    def refresh_console_setting(self):
        if self.name_or_path.get() == 1:
            if self.ifz_located == None:
                only_name = str(self.all_images[self.increment_of_image]).split("/")
            else:
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                only_name = str(self.converted_images[center_image_index + self.increment_of_ifz_image]).split("/")

            only_name = only_name[int(len(only_name))-1]
            Tools.add_colored_line(self.console,str(only_name),"white",None,True)

        else:
            if self.ifz_located == None:
                Tools.add_colored_line(self.console,str(self.all_images[self.increment_of_image]),"white",None,True)
            else:
                center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                Tools.add_colored_line(self.console,str(self.converted_images[center_image_index + self.increment_of_ifz_image]),"white",None,True)

    def drawing_option_window(self):
        def close_window(window):
            self.main_frame.unbind("<Button-1>")
            self.main_frame.unbind("<B1-Motion>")
            self.main_frame.unbind("<ButtonRelease-1>")
            self.main_frame.unbind("<Motion>")
            self.switch_drawing_mode()
            window.destroy()
        
        def rgb_to_hex(rgb,one_color = False):
            if not one_color:
                return "#%02x%02x%02x" % rgb
            elif one_color == "red":
                return ("#%02x" % rgb) + "0000"
            elif one_color == "green":
                return "#00" + ("%02x" % rgb) + "00"
            elif one_color == "blue":
                return "#0000" + ("%02x" % rgb)
        
        def hex_to_rgb(hex_color):
            # Remove the '#' character if present
            hex_color = hex_color.lstrip('#')
            # Convert the hex string into RGB tuple
            return list(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        
        def update_color(*args):
            nonlocal current_color_frame
            red = int(color_R.get())
            red_hex = rgb_to_hex(red,one_color="red")
            current_color_val.configure(text = str(red))
            color_R.configure(progress_color = red_hex,button_color = red_hex,button_hover_color = red_hex)

            green = int(color_G.get())
            green_hex = rgb_to_hex(green,one_color="green")
            current_color_val2.configure(text = str(green))
            color_G.configure(progress_color = green_hex,button_color = green_hex,button_hover_color = green_hex)

            blue = int(color_B.get())
            blue_hex = rgb_to_hex(blue,one_color="blue")
            current_color_val3.configure(text = str(blue))
            color_B.configure(progress_color = blue_hex,button_color = blue_hex,button_hover_color = blue_hex)

            current_color_frame.configure(fg_color = rgb_to_hex((red,green,blue)))
            self.drawing_color = rgb_to_hex((red,green,blue))
            line_frame.configure(fg_color = self.drawing_color)

        def color_set(rgb):
            color_R.set(rgb[0])
            color_G.set(rgb[1])
            color_B.set(rgb[2])
            update_color("")

        def update_thickness(*args):
            self.drawing_thickness = int(*args)
            current_thickness.configure(text = str(self.drawing_thickness))
            line_frame.configure(height = self.drawing_thickness)
        
        def switch_draw_mode():
            nonlocal draw_circle
            nonlocal draw_line

            if draw_circle.get() == 1:
                self.draw_mode = "circle"
                draw_line.deselect()
            else:
                self.draw_mode = "line"
                draw_circle.deselect()

        def clear_canvas():
            self.main_frame.delete("drawing")

        def draw_cursor(flag = ""):
            image_center_x = self.image_dimensions[0]/2
            image_center_y = self.image_dimensions[1]/2

            if flag =="full":
                cursor_len_x = image_center_x
                cursor_len_y = image_center_y
            else:
                cursor_len_x = 50
                cursor_len_y = 50

            self.main_frame.create_line(image_center_x-cursor_len_x, image_center_y, image_center_x+cursor_len_x, image_center_y, fill=self.drawing_color,tags="drawing",width=self.drawing_thickness)
            self.main_frame.create_line(image_center_x, image_center_y-cursor_len_y, image_center_x, image_center_y+cursor_len_y, fill=self.drawing_color,tags="drawing",width=self.drawing_thickness)


        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(app_icon))
        window_height = 500
        window_width = 620
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        window.geometry(f"{window_width}x{window_height}+{x+150}+{y+50}")
        window.title("Možnosti malování")
        
        top_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,height=120,fg_color="gray10")
        current_color_frame = customtkinter.CTkFrame(master = top_frame,corner_radius=0,border_width=2,height=100,width=100,fg_color="gray10")
        slider_frame =      customtkinter.CTkFrame(master = top_frame,corner_radius=0,width=500,fg_color="gray10")
        top_frame.          pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        top_frame.          pack_propagate(0)

        slider_frame.       pack(pady=(10,0),padx=(5,0),expand=False,side = "left")
        slider_frame.       pack_propagate(0)
        current_color_frame.pack(pady=(10,0),padx=10,expand=False,side = "left",anchor = "w")

        frame_R =           customtkinter.CTkFrame(master = slider_frame,height=20,corner_radius=0,border_width=0,fg_color="gray10")
        color_label =       customtkinter.CTkLabel(master = frame_R,text = "R: ",justify = "left",font=("Arial",16,"bold"))
        color_R =           customtkinter.CTkSlider(master=frame_R,width=400,height=15,from_=0,to=255,command= lambda e: update_color(e))
        current_color_val = customtkinter.CTkLabel(master = frame_R,text = "0",justify = "left",font=("Arial",16,"bold"))
        color_label.        pack(pady=5,padx=5,expand=False,side = "left")
        color_R.            pack(pady=5,padx=5,expand=False,side = "left")
        current_color_val.  pack(pady=5,padx=5,expand=False,side = "left")
        color_R.set(0.0)
        
        frame_G =           customtkinter.CTkFrame(master = slider_frame,height=20,corner_radius=0,border_width=0,fg_color="gray10")
        color_label =       customtkinter.CTkLabel(master = frame_G,text = "G: ",justify = "left",font=("Arial",16,"bold"))
        color_G =           customtkinter.CTkSlider(master=frame_G,width=400,height=15,from_=0,to=255,command= lambda e: update_color(e))
        current_color_val2 = customtkinter.CTkLabel(master = frame_G,text = "0",justify = "left",font=("Arial",16,"bold"))
        color_label.        pack(pady=5,padx=5,expand=False,side = "left")
        color_G.            pack(pady=5,padx=5,expand=False,side = "left")
        current_color_val2. pack(pady=5,padx=5,expand=False,side = "left")
        color_G.set(0.0)

        frame_B =           customtkinter.CTkFrame(master = slider_frame,height=20,corner_radius=0,border_width=0,fg_color="gray10")
        color_label =       customtkinter.CTkLabel(master = frame_B,text = "B: ",justify = "left",font=("Arial",16,"bold"))
        color_B =           customtkinter.CTkSlider(master=frame_B,width=400,height=15,from_=0,to=255,command= lambda e: update_color(e))
        current_color_val3 = customtkinter.CTkLabel(master = frame_B,text = "0",justify = "left",font=("Arial",16,"bold"))
        color_label.        pack(pady=5,padx=5,expand=False,side = "left")
        color_B.            pack(pady=5,padx=5,expand=False,side = "left")
        current_color_val3. pack(pady=5,padx=5,expand=False,side = "left")
        color_B.set(0.0)

        bottom_frame =      customtkinter.CTkFrame(master = window,corner_radius=0,fg_color="gray10") 
        common_colors =     customtkinter.CTkFrame(master = bottom_frame,corner_radius=0,border_width=0,fg_color="gray10")
        white_button =      customtkinter.CTkButton(master = common_colors,text="",width = 30,height=30,corner_radius=0,border_width=1,fg_color="#FFFFFF",command=lambda: color_set([255,255,255]))
        black_button =      customtkinter.CTkButton(master = common_colors,text="",width = 30,height=30,corner_radius=0,border_width=1,fg_color="#000000",command=lambda: color_set([0,0,0]))
        red_button =        customtkinter.CTkButton(master = common_colors,text="",width = 30,height=30,corner_radius=0,border_width=1,fg_color="#FF0000",command=lambda: color_set([255,0,0]))
        green_button =      customtkinter.CTkButton(master = common_colors,text="",width = 30,height=30,corner_radius=0,border_width=1,fg_color="#00FF00",command=lambda: color_set([0,255,0]))
        blue_button =       customtkinter.CTkButton(master = common_colors,text="",width = 30,height=30,corner_radius=0,border_width=1,fg_color="#0000FF",command=lambda: color_set([0,0,255]))
        white_button.       pack(pady=5,padx=(5,0),expand=False,side = "left")
        black_button.       pack(pady=5,padx=(5,0),expand=False,side = "left")
        red_button.         pack(pady=5,padx=(5,0),expand=False,side = "left")
        green_button.       pack(pady=5,padx=(5,0),expand=False,side = "left")
        blue_button.        pack(pady=5,padx=(5,0),expand=False,side = "left")
        common_colors.      pack(pady=0,padx=0,expand=False,side = "top",fill="x")
        
        shape_checkboxes =  customtkinter.CTkFrame(master = bottom_frame,corner_radius=0,fg_color="gray10") 
        draw_circle =       customtkinter.CTkCheckBox(master = shape_checkboxes, text = "Kruh",command = lambda: switch_draw_mode(),font=("Arial",20))
        draw_line =         customtkinter.CTkCheckBox(master = shape_checkboxes, text = "Osa",command = lambda: switch_draw_mode(),font=("Arial",20))
        draw_circle.        pack(pady=0,padx=5,expand=False,side = "left")
        draw_line.          pack(pady=0,padx=5,expand=False,side = "left")
        shape_checkboxes.   pack(pady=15,padx=5,expand=False,side = "top",fill="x")

        bottom_frame_label = customtkinter.CTkLabel(master = bottom_frame,text = "Nastavení tloušťky čáry:",justify = "left",font=("Arial",18,"bold"),anchor="w")

        thickness_frame =   customtkinter.CTkFrame(master = bottom_frame,corner_radius=0,fg_color="gray10",height=55) 
        thickness =         customtkinter.CTkSlider(master=thickness_frame,width=450,height=15,from_=1,to=50,command= lambda e: update_thickness(e))
        current_thickness = customtkinter.CTkLabel(master = thickness_frame,text = "0",justify = "left",font=("Arial",16,"bold"))
        line_frame =        customtkinter.CTkFrame(master = thickness_frame,corner_radius=0,fg_color="black",height=1,width = 100) 
        thickness.          pack(pady=5,padx=5,expand=False,side = "left")
        current_thickness.  pack(pady=5,padx=5,expand=False,side = "left")
        line_frame.         pack(pady=5,padx=5,expand=False,side = "left")
        cursor_frame =      customtkinter.CTkFrame(master = bottom_frame,corner_radius=0,fg_color="gray10",height=55)
        cursor_button =     customtkinter.CTkButton(master = cursor_frame,text = "Kurzor uprostřed",font=("Arial",22,"bold"),width = 150,height=40,corner_radius=0,command=lambda: draw_cursor())
        cursor_button2 =    customtkinter.CTkButton(master = cursor_frame,text = "Kurzor uprostřed (celý)",font=("Arial",22,"bold"),width = 150,height=40,corner_radius=0,command=lambda: draw_cursor("full"))
        cursor_button.      pack(pady=5,padx=5,expand=True,side = "left",fill="x")
        cursor_button2.     pack(pady=5,padx=5,expand=True,side = "left",fill="x")

        clear_all =         customtkinter.CTkButton(master = bottom_frame,text = "Vyčistit",font=("Arial",22,"bold"),width = 150,height=40,corner_radius=0,command=lambda: clear_canvas())
        frame_R.            pack(pady=0,padx=0,expand=False,side = "top",fill="x")
        frame_G.            pack(pady=0,padx=0,expand=False,side = "top",fill="x")
        frame_B.            pack(pady=0,padx=0,expand=False,side = "top",fill="x")
        bottom_frame.       pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        bottom_frame_label. pack(pady=5,padx=5,expand=False,side = "top",fill="x",anchor = "w")
        thickness_frame.    pack(pady=(5,0),padx=5,expand=False,side = "top",fill="x")
        thickness_frame.    pack_propagate(0)
        cursor_frame.       pack(pady=(20,5),padx=5,expand=False,side = "top",fill="x")
        clear_all.          pack(pady=5,padx=10,expand=False,side = "top",fill="x")
        current_color_frame.configure(fg_color = self.drawing_color)
        previous_color = hex_to_rgb(self.drawing_color)
        color_R.set(previous_color[0])
        color_G.set(previous_color[1])
        color_B.set(previous_color[2])
        draw_line.select()
        thickness.set(self.drawing_thickness)
        update_thickness(self.drawing_thickness)
        update_color("")

        button_exit = customtkinter.CTkButton(master = window,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_exit.pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        # self.root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        window.transient(self.root)
        window.attributes("-topmost", True)
        window.attributes("-disabled", False)
        # window.focus_force()
        # window.grab_set()
        # window.grab_release()
        # window.focus()

    def switch_drawing_mode(self,initial = False):
        """
        Making binds to canvas\n
        Operation:
        - "drawing"
        - "image"
        """
        self.released = False
        self.start_x = 0
        self.start_y = 0

        self.main_frame.unbind("<Button-1>")
        self.main_frame.unbind("<B1-Motion>")
        self.main_frame.unbind("<ButtonRelease-1>")
        self.main_frame.unbind("<Motion>")

        def bind_drawing():
            def on_click(event):
                # Save the start position
                self.start_x = event.x
                self.start_y = event.y

            def on_drag(event):
                # Clear the canvas (if you want to see only the final shape)
                self.main_frame.delete("temp_shape")
                if self.draw_mode == "line":
                    # Draw a line from the start position to the current position
                    self.main_frame.create_line(self.start_x, self.start_y, event.x, event.y, fill=self.drawing_color, tags="temp_shape",width=self.drawing_thickness)
                elif self.draw_mode == "circle":
                    # Draw an oval (circle) based on the start position and current position
                    self.main_frame.create_oval(self.start_x, self.start_y, event.x, event.y, outline=self.drawing_color, tags="temp_shape",width=self.drawing_thickness)

            def on_release(event):
                # Finalize the shape
                self.main_frame.delete("temp_shape")
                if self.draw_mode == "line":
                    # Draw the final line
                    self.main_frame.create_line(self.start_x, self.start_y, event.x, event.y, fill=self.drawing_color,tags="drawing",width=self.drawing_thickness)
                elif self.draw_mode == "circle":
                    # Draw the final circle
                    self.main_frame.create_oval(self.start_x, self.start_y, event.x, event.y, outline=self.drawing_color,tags="drawing",width=self.drawing_thickness)
            
            self.main_frame.bind("<Button-1>", on_click)
            self.main_frame.bind("<B1-Motion>", on_drag)
            self.main_frame.bind("<ButtonRelease-1>", on_release)

        def bind_image_dragging():
            def mouse_clicked(e):
                self.main_frame.focus_set()
                self.released = False
                x,y = e.x,e.y

                def get_direction(e):
                    current_coords = self.main_frame.coords(self.main_image) 
                    option = ""
                    if abs(max(e.x,x)-min(e.x,x)) > abs(max(e.y,y)-min(e.y,y)):
                        option = "horizontal"
                    else:
                        option = "vertical"

                    if option == "horizontal":
                        if e.x > x:
                            #right
                            if (current_coords[0] + self.drag_increment) < 0:
                                self.main_frame.coords(self.main_image,current_coords[0]+self.drag_increment,current_coords[1])
                            else:
                                self.main_frame.coords(self.main_image,0,current_coords[1])
                        else:
                            #left
                            if (current_coords[0] - self.drag_increment) > -(self.tk_image.width()-self.main_frame.winfo_width()):
                                self.main_frame.coords(self.main_image,current_coords[0]-self.drag_increment,current_coords[1])
                            else:
                                self.main_frame.coords(self.main_image,-(self.tk_image.width()-self.main_frame.winfo_width()),current_coords[1])

                    if option == "vertical":
                        if e.y > y:
                            #down
                            if (current_coords[1] + self.drag_increment) < 0:
                                self.main_frame.coords(self.main_image,current_coords[0],current_coords[1]+self.drag_increment)
                            else:
                                self.main_frame.coords(self.main_image,current_coords[0],0)
                        else:
                            #up
                            if (current_coords[1] - self.drag_increment) > -(self.tk_image.height()-self.main_frame.winfo_height()):
                                self.main_frame.coords(self.main_image,current_coords[0],current_coords[1]-self.drag_increment)
                            else:
                                self.main_frame.coords(self.main_image,current_coords[0],-(self.tk_image.height()-self.main_frame.winfo_height()))
                    
                    self.main_frame.update()
                    self.last_coords = self.main_frame.coords(self.main_image) 
                    return

                self.main_frame.bind("<Motion>", get_direction)
                if self.released == True:
                    return

                def end_func(e):
                    self.main_frame.unbind("<Motion>")
                    self.main_frame.unbind("<ButtonRelease-1>")
                    self.released = True
                    return

                self.main_frame.bind("<ButtonRelease-1>",end_func)
            self.main_frame.bind("<Button-1>",mouse_clicked)

        if initial:
            bind_image_dragging()
            self.drag_option_binded = True
            return

        if self.drag_option_binded:
            bind_drawing()
            self.drawing_option_window()
            self.drag_option_binded = False

        elif not self.drag_option_binded:
            bind_image_dragging()
            self.drag_option_binded = True

    def create_widgets(self): # Vytvoření veškerých widgets (MAIN image browseru)
        def call_setting_window():
            if self.ifz_located == True:
                path_to_send = self.all_images[self.increment_of_image]
            else:
                try:
                    path_to_send = self.all_images[self.increment_of_image]
                except Exception:
                    path_to_send = ""
            Advanced_option(self.root,windowed=True,spec_location="image_browser", path_to_remember = path_to_send,last_params = [self.last_coords,self.zoom_slider.get()])
        
        def call_path_context_menu(event):
            def insert_path(path):
                self.path_set.delete("0","200")
                self.path_set.insert("0", path)
            if len(self.inserted_path_history) > 0:
                path_context_menu = tk.Menu(self.root, tearoff=0,fg="white",bg="black")
                for i in range(0,len(self.inserted_path_history)):
                    path_context_menu.add_command(label=self.inserted_path_history[i], command=lambda row_path = self.inserted_path_history[i]: insert_path(row_path),font=("Arial",22,"bold"))
                    if i < len(self.inserted_path_history)-1:
                        path_context_menu.add_separator()
                        
                path_context_menu.tk_popup(context_menu_button.winfo_rootx(),context_menu_button.winfo_rooty()+30)

        def call_start():
            self.changable_image_num.delete("0","100")
            self.changable_image_num.insert("0",str(0))
            self.changable_image_num_ifz.delete("0","100")
            self.changable_image_num_ifz.insert("0",str(0))
            self.selected_image = ""
            self.current_image_num_ifz.configure(text = "/0")
            self.current_image_num.configure(text = "/0")
            self.start(self.path_set.get())

        self.frame_with_path =          customtkinter.CTkFrame(master=self.root,height = 200,corner_radius=0)
        menu_button  =                  customtkinter.CTkButton(master = self.frame_with_path, width = 100,height=30, text = "MENU", command = lambda: self.call_menu(),font=("Arial",16,"bold"))
        context_menu_button  =          customtkinter.CTkButton(master = self.frame_with_path, width = 50,height=30, text = "V",font=("Arial",16,"bold"),corner_radius=0,fg_color="#505050")
        self.path_set =                 customtkinter.CTkEntry(master = self.frame_with_path,width = 680,height=30,placeholder_text="Zadejte cestu k souborům (kde se soubory přímo nacházejí)",corner_radius=0)
        manual_path  =                  customtkinter.CTkButton(master = self.frame_with_path, width = 90,height=30,text = "Otevřít", command = lambda: call_start(),font=("Arial",16,"bold"))
        tree         =                  customtkinter.CTkButton(master = self.frame_with_path, width = 120,height=30,text = "EXPLORER", command = self.call_browseDirectories,font=("Arial",16,"bold"))
        button_save_path =              customtkinter.CTkButton(master = self.frame_with_path,width=100,height=30, text = "Uložit cestu", command = lambda: Tools.save_path(self.console,self.path_set.get()),font=("Arial",16,"bold"))        
        button_open_setting =           customtkinter.CTkButton(master = self.frame_with_path,width=30,height=30, text = "⚙️", command = lambda: call_setting_window(),font=("",16))
        button_drawing =                customtkinter.CTkButton(master = self.frame_with_path,width=30,height=30, text = "Malování", command = lambda: self.switch_drawing_mode(),font=("Arial",16,"bold"))
        menu_button.                    pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        context_menu_button.            pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        self.path_set.                  pack(pady = (5,0),padx =(0,0),side="left",anchor = "w")
        manual_path.                    pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        tree.                           pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        button_save_path.               pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        button_open_setting.            pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        button_drawing.                 pack(pady = (5,0),padx =(5,0),side="left",anchor = "w")
        self.frame_with_path.           pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        self.frame_with_console =       customtkinter.CTkFrame(master=self.root,height = 200,corner_radius=0)
        self.name_or_path =             customtkinter.CTkCheckBox(master = self.frame_with_console,font=("Arial",16), text = "Název/cesta",command= lambda: self.refresh_console_setting())
        self.console =                  tk.Text(self.frame_with_console, wrap="none", height=0, width=180,background="black",font=("Arial",14),state=tk.DISABLED)
        self.name_or_path.              pack(pady = (5,0),padx =10,anchor = "w",side="left")
        self.console.                   pack(pady = (5,0),padx =10,anchor = "w",side="left")
        self.frame_with_console.        pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        self.frame_with_buttons =       customtkinter.CTkFrame(master=self.root,height = 200,corner_radius=0)
        button_back  =                  customtkinter.CTkButton(master = self.frame_with_buttons, width = 20,height=30,text = "<", command = self.previous_image,font=("Arial",16,"bold"))
        self.changable_image_num =      customtkinter.CTkEntry(master = self.frame_with_buttons,width=45,justify = "left",font=("Arial",16,"bold"))
        self.changable_image_num.delete("0","100")
        self.changable_image_num.insert("0",0)
        self.current_image_num =        customtkinter.CTkLabel(master = self.frame_with_buttons,text = "/0",justify = "left",font=("Arial",16,"bold"))
        button_next  =                  customtkinter.CTkButton(master = self.frame_with_buttons, width = 20,height=30,text = ">", command = self.next_image,font=("Arial",16,"bold"))
        self.button_play_stop  =        customtkinter.CTkButton(master = self.frame_with_buttons, width = 90,height=30,text = "SPUSTIT", command = self.play,font=("Arial",16,"bold"))
        button_copy  =                  customtkinter.CTkButton(master = self.frame_with_buttons, width = 80,height=30,text = "Kopír.", command = lambda: self.copy_image(self.image_browser_path),font=("Arial",16,"bold"))
        rotate_button =                 customtkinter.CTkButton(master = self.frame_with_buttons, width = 80,height=30,text = "OTOČIT", command =  lambda: self.rotate_image(),font=("Arial",16,"bold"))
        speed_label  =                  customtkinter.CTkLabel(master = self.frame_with_buttons,text = "Rychlost:",justify = "left",font=("Arial",12))
        self.speed_slider =             customtkinter.CTkSlider(master = self.frame_with_buttons,width=120,from_=1,to=100,command= self.update_speed_slider)
        self.percent1 =                 customtkinter.CTkLabel(master = self.frame_with_buttons,text = "%",justify = "left",font=("Arial",12))
        zoom_label   =                  customtkinter.CTkLabel(master = self.frame_with_buttons,text = "ZOOM:",justify = "left",font=("Arial",12))
        self.zoom_slider =              customtkinter.CTkSlider(master = self.frame_with_buttons,width=120,from_=100,to=500,command= self.update_zoom_slider)
        self.percent2 =                 customtkinter.CTkLabel(master = self.frame_with_buttons,text = "%",justify = "left",font=("Arial",12))
        reset_button =                  customtkinter.CTkButton(master = self.frame_with_buttons, width = 80,height=30,text = "RESET", command = lambda: self.Reset_all(),font=("Arial",16,"bold"))
        ifz_label =                     customtkinter.CTkLabel(master = self.frame_with_buttons,text = "IFZ:",justify = "left",font=("Arial",12))
        button_back_ifz  =              customtkinter.CTkButton(master = self.frame_with_buttons, width = 20,height=30,text = "<", command = self.previous_ifz_image,font=("Arial",16,"bold"))
        self.changable_image_num_ifz =  customtkinter.CTkEntry(master = self.frame_with_buttons,width=20,justify = "left",font=("Arial",16,"bold"))
        self.changable_image_num_ifz.delete("0","100")
        self.changable_image_num_ifz.insert("0",1)
        self.current_image_num_ifz =    customtkinter.CTkLabel(master = self.frame_with_buttons,text = "/0",justify = "left",font=("Arial",16,"bold"))
        button_next_ifz  =              customtkinter.CTkButton(master = self.frame_with_buttons, width = 20,height=30,text = ">", command = self.next_ifz_image,font=("Arial",16,"bold"))
        button_move =                   customtkinter.CTkButton(master = self.frame_with_buttons, width = 80,height=30,text = "Přesun.", command =  lambda: self.move_image(),font=("Arial",16,"bold"))
        button_delete =                 customtkinter.CTkButton(master = self.frame_with_buttons, width = 80,height=30,text = "SMAZAT", command =  lambda: self.delete_image(),font=("Arial",16,"bold"))
        button_back.                    pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        self.changable_image_num.       pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        self.current_image_num.         pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        button_next.                    pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        self.button_play_stop.          pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        button_copy.                    pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        rotate_button.                  pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        speed_label.                    pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.speed_slider.              pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.percent1.                  pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        zoom_label.                     pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.zoom_slider.               pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.percent2.                  pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        reset_button.                   pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        ifz_label.                      pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        button_back_ifz.                pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.changable_image_num_ifz.   pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        self.current_image_num_ifz.     pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        button_next_ifz.                pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        button_move.                    pack(pady = (5,0),padx =(10,0),anchor = "w",side="left")
        button_delete.                  pack(pady = (5,0),padx =(5,0),anchor = "w",side="left")
        self.frame_with_buttons.        pack(pady=0,padx=(0,0),fill="x",expand=False,side = "top")
        self.background_frame =         customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.main_frame =               tk.Canvas(master=self.background_frame,bg="black",highlightthickness=0)
        self.image_film_frame_left =    customtkinter.CTkFrame(master=self.root,height = 100,corner_radius=0)
        self.image_film_frame_center =  customtkinter.CTkFrame(master=self.root,height = 100,width = 200,corner_radius=0)
        self.image_film_frame_right =   customtkinter.CTkFrame(master=self.root,height = 100,corner_radius=0)
        self.background_frame.          pack(pady=(10,0),padx=5,ipadx=10,ipady=10,fill="both",expand=True,side = "top")
        if self.image_film == True:
            self.image_film_frame_left. pack(pady=5,expand=True,side = "left",fill="x")
            self.image_film_frame_center.pack(pady=5,padx=10,expand=False,side = "left",anchor = "center")
            self.image_film_frame_right.pack(pady=5,expand=True,side = "left",fill="x")
            self.images_film_center =   customtkinter.CTkLabel(master = self.image_film_frame_center,text = "")
            self.images_film_center.    pack()
        self.main_frame.                pack(pady=0,padx=5,ipadx=10,ipady=10,fill="both",expand=True,side = "bottom",anchor= "center")
        self.name_or_path.select()
        context_menu_button.bind("<Button-1>", call_path_context_menu)

        def jump_to_image(e):
            if self.changable_image_num.get().isdigit():
                inserted_value = int(self.changable_image_num.get())
                if inserted_value >= 1 and inserted_value <= int(len(self.all_images)):
                    self.increment_of_image = inserted_value-1
                    if self.ifz_located == True:
                        self.convert_files()
                        center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                        self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image])
                    else:
                        self.view_image(self.increment_of_image)
            self.changable_image_num.delete("0","100")
            self.changable_image_num.insert("0",self.increment_of_image+1)
            self.root.focus_set() # unfocus    
        self.changable_image_num.bind("<Return>",jump_to_image)

        # nastaveni defaultnich hodnot slideru
        self.zoom_slider.set(self.zoom_given)
        self.update_zoom_slider(self.zoom_given)
        self.speed_slider.set(100)
        self.update_speed_slider(100)
        self.switch_drawing_mode(initial=True)

        def focused_entry_widget():
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return True
            else:
                return False
            
        def focused_control_widget():
            currently_focused = str(self.root.focus_get())
            if ".!text" in currently_focused:
                return True
            else:
                return False

        # KEYBOARD BINDING
        def rotate_button_hover(e):
            if int(self.rotation_angle==270):
                rotate_button.configure(text="0°",font=("Arial",15))
            else:
                rotate_button.configure(text=str(int(self.rotation_angle)+90)+"°",font=("Arial",15))
            rotate_button.update_idletasks()
            return
                
        def rotate_button_hover_leave(e):
            rotate_button.configure(text="OTOČIT",font=("Arial",16,"bold"))
        rotate_button.bind("<Enter>",rotate_button_hover)
        rotate_button.bind("<Button-1>",rotate_button_hover)
        rotate_button.bind("<Leave>",rotate_button_hover_leave)

        def move_button_hover(e):
            button_move.configure(text="\"M\"",font=("Arial",16,"bold"))
            button_move.update_idletasks()
            return
                
        def move_button_hover_leave(e):
            button_move.configure(text="Přesun.",font=("Arial",16,"bold"))

        button_move.bind("<Enter>",move_button_hover)
        button_move.bind("<Button-1>",move_button_hover)
        button_move.bind("<Leave>",move_button_hover_leave)

        def copy_button_hover(e):
            button_copy.configure(text="\"C\"",font=("Arial",16,"bold"))
            button_copy.update_idletasks()
            return
                
        def copy_button_hover_leave(e):
            button_copy.configure(text="Kopír.",font=("Arial",16,"bold"))

        button_copy.bind("<Enter>",copy_button_hover)
        button_copy.bind("<Button-1>",copy_button_hover)
        button_copy.bind("<Leave>",copy_button_hover_leave)

        def pressed_space(e):
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            if self.state != "stop":
                self.state = "stop"
                self.stop()
            else:
                self.state = "running"
                self.play()
        self.root.bind("<space>",pressed_space)
        self.unbind_list.append("<space>")

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")

        def pressed_left(e):
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.previous_image()
        self.root.bind("<Left>",pressed_left)
        self.unbind_list.append("<Left>")

        def pressed_right(e):
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.next_image()
        self.root.bind("<Right>",pressed_right)
        self.unbind_list.append("<Right>")

        def pressed_up(e):
            if focused_entry_widget() or self.ifz_located == None: # pokud nabindovany znak neni vepisovan do entry widgetu nebo nejsou ifz
                return
            self.next_ifz_image()
        self.root.bind("<Up>",pressed_up)
        self.unbind_list.append("<Up>")

        def pressed_down(e):
            if focused_entry_widget() or self.ifz_located == None: # pokud nabindovany znak neni vepisovan do entry widgetu nebo nejsou ifz
                return
            self.previous_ifz_image()
        self.root.bind("<Down>",pressed_down)
        self.unbind_list.append("<Down>")

        def pressed_copy(e):
            if focused_entry_widget() or focused_control_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.copy_image(self.image_browser_path)
        self.root.bind("<c>",pressed_copy)
        self.unbind_list.append("<c>")

        def pressed_move(e):
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.move_image()
        self.root.bind("<m>",pressed_move)
        self.unbind_list.append("<m>")
        
        def pressed_rotate(e):
            rotate_button_hover(e) #update uhlu zobrazovanem na tlacitku
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.rotate_image()
        self.root.bind("<r>",pressed_rotate)
        self.unbind_list.append("<r>")

        def pressed_delete(e):
            if focused_entry_widget(): # pokud nabindovany znak neni vepisovan do entry widgetu
                return
            self.delete_image()
        self.root.bind("<Delete>",pressed_delete)
        self.unbind_list.append("<Delete>")

        def call_refresh(e):
            self.Reset_all()
        self.root.bind("<F5>",lambda e: call_refresh(e))
        self.unbind_list.append("<F5>")

        #Funkce kolecka mysi: priblizovat nebo posouvat vpred/ vzad
        def mouse_wheel1(e): # priblizovat
            direction = -e.delta

            self.main_frame.update_idletasks()
            self.zoom_slider.update_idletasks()
            frame_dim = self.get_frame_dimensions()
            frame_width = frame_dim[0]
            frame_height = frame_dim[1]

            if e.x <= frame_width/2: # pokud větší, tak budeme růst obrázku přičítat
                self.x_growth_multiplier = 1-((e.x/(frame_width/2)))/2
            elif e.x >= frame_width/2: # pokud větší, tak budeme růst obrázku odečítat
                self.x_growth_multiplier = -((e.x/(frame_width/2))/2)
                if direction > 0:
                    self.x_growth_multiplier = self.x_growth_multiplier *(-1)

            if e.y <= frame_height/2: # pokud větší, tak budeme růst obrázku přičítat
                self.y_growth_multiplier = 1-((e.y/(frame_height/2)))/2
            elif e.y >= frame_height/2: # pokud větší, tak budeme růst obrázku odečítat
                self.y_growth_multiplier = -((e.y/(frame_height/2))/2)
                if direction > 0:
                    self.y_growth_multiplier = self.y_growth_multiplier *(-1)
            
            if direction < 0:
                direction = "in"
                new_value = self.zoom_slider.get()+self.zoom_increment
                if self.zoom_slider._to >= new_value:
                    self.zoom_slider.set(new_value)
                    self.percent2.configure(text=str(int(new_value)) + " %")
                else:
                    self.zoom_slider.set(self.zoom_slider._to) # pro pripad, ze by zbyvalo mene nez 5 do maxima 
            else:
                direction = "out"
                new_value = self.zoom_slider.get()-self.zoom_increment
                if self.zoom_slider._from_ <= new_value:
                    self.zoom_slider.set(new_value)
                    self.percent2.configure(text=str(int(new_value)) + " %")
                else:
                    self.zoom_slider.set(self.zoom_slider._from_) # pro pripad, ze by zbyvalo vice nez 5 do minima
            
            if len(self.all_images) != 0: # update zobrazeni
                if self.ifz_located == True:
                    center_image_index = int((len(self.image_queue)-1)/2) * self.ifz_count
                    self.view_image(None,self.converted_images[center_image_index + self.increment_of_ifz_image],True)
                    #self.view_image(None,self.converted_images[self.increment_of_ifz_image])  
                else:
                    self.view_image(self.increment_of_image,None,True)
        
        def mouse_wheel2(e): # posouvat obrazky
            direction = -e.delta
            if direction < 0:
                self.previous_image()
            else:
                self.next_image()

        self.main_frame.bind("<MouseWheel>",mouse_wheel1)
        self.frame_with_path.bind("<MouseWheel>",mouse_wheel2)
        self.console.bind("<MouseWheel>",mouse_wheel2)
        if self.image_film == True:
            self.image_film_frame_left.bind("<MouseWheel>",mouse_wheel2)
            self.image_film_frame_center.bind("<MouseWheel>",mouse_wheel2)
            self.image_film_frame_right.bind("<MouseWheel>",mouse_wheel2)
            self.images_film_center.bind("<MouseWheel>",mouse_wheel2)
        self.unbind_list.append("<MouseWheel>")
        
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)
        self.unbind_list.append("<f>")

        def save_path_enter_btn(e):
            self.root.focus_set()
            self.start(self.path_set.get())
        self.path_set.bind("<Return>",save_path_enter_btn)

        def open_path():
            checked_path = Tools.path_check(self.path_for_explorer)
            if checked_path != False and checked_path != "/" and checked_path != "":
                if os.path.exists(checked_path):
                    os.startfile(checked_path)

        def show_context_menu(event):
            context_menu.tk_popup(event.x_root, event.y_root)

        self.main_frame.bind("<Button-3>", show_context_menu)
        context_menu = tk.Menu(self.root, tearoff=0,fg="white",bg="black")
        context_menu.add_command(label="Otevřít cestu", command=lambda: open_path(),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Otevřít v novém okně", command=lambda: self.view_image(self.increment_of_image,in_new_window=True),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Malovat", command=lambda: self.switch_drawing_mode(),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Otočit", command=lambda: self.rotate_image(),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Kopírovat", command=lambda: self.copy_image(self.image_browser_path),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Přesunout", command=lambda: self.move_image(),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Reset", command=lambda: self.Reset_all(),font=("Arial",22,"bold"))
        context_menu.add_separator()
        context_menu.add_command(label="Smazat", command=lambda: self.delete_image(),font=("Arial",22,"bold"))

        #kdyz je vyuzit TRIMAZKON, jako vychozi prohlizec obrazku
        if self.IB_as_def_browser_path != None:
            self.path_set.delete("0","200")
            self.path_set.insert("0", self.IB_as_def_browser_path)
            Tools.add_colored_line(self.console,"TRIMAZKON, jako výchozí prohlížeč!","white",None,True)
            self.root.update_idletasks()
            self.image_browser_path = self.IB_as_def_browser_path
            self.start(self.IB_as_def_browser_path)

        #hned na zacatku to vleze do defaultni slozky
        elif self.path_given != "":
            self.path_set.delete("0","200")
            self.path_set.insert("0", self.path_given)
            Tools.add_colored_line(self.console,"Nastavené změny uloženy","green",None,True)
            self.root.update_idletasks()
            self.image_browser_path = self.path_given
            self.start(self.path_given)
        else:
            path = self.config_path
            if path != "/" and path != False:
                self.path_set.delete("0","200")
                self.path_set.insert("0", path)
                Tools.add_colored_line(self.console,"Byla vložena cesta z konfiguračního souboru","white",None,True)
                self.root.update_idletasks()
                self.image_browser_path = path
                self.start(path)
            else:
                Tools.add_colored_line(self.console,"Konfigurační soubor obsahuje neplatnou cestu k souborům\n(můžete vložit v pokročilém nastavení)","orange",None,True)

class Advanced_option: # Umožňuje nastavit základní parametry, které ukládá do textového souboru
    """
    Umožňuje nastavit základní parametry, které ukládá do textového souboru
    """
    def __init__(self,root,windowed=None,spec_location=None,path_to_remember = None,last_params = None,tray_setting_status_message = None):
        self.spec_location = spec_location
        self.path_to_remember = path_to_remember
        self.ib_last_params = last_params
        self.windowed = windowed
        self.root = root
        self.tray_setting_status_message = tray_setting_status_message
        self.unbind_list = []
        self.drop_down_prefix_dir_names_list = []
        self.drop_down_static_dir_names_list = []
        self.default_displayed_prefix_dir = "cam"
        self.default_displayed_static_dir = 0
        self.submenu_option = "default_path"
        self.text_file_data = Tools.read_config_data()
        default_dir_names = string_database.default_setting_database_param
        self.default_dir_names = [" (default: "+ default_dir_names[9] + ")",
                                  " (default: "+ default_dir_names[10] + ")",
                                  " (default: "+ default_dir_names[11] + ")",
                                  " (default: "+ default_dir_names[12] + ")",
                                  " (default: "+ default_dir_names[13] + ")",
                                  " (default: "+ default_dir_names[14] + ")",
                                  " (default: "+ default_dir_names[15] + ")"
                                  ]
        
        self.creating_advanced_option_widgets()
    
    def set_zoom(self,zoom_factor):
        try:
            root.after(0,customtkinter.set_widget_scaling(zoom_factor / 100))
        except Exception as e:
            print(f"error with zoom scaling: {e}")
        
        root.tk.call('tk', 'scaling', zoom_factor / 100)

    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu
        """
        self.list_of_frames = [self.top_frame,
                               self.bottom_frame_default_path,
                               self.menu_buttons_frame]
        for frames in self.list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()
        
        for binds in self.unbind_list:
            self.root.unbind(binds)
        menu.menu(zoom_disable = True)

    def clear_frame(self,frame): # Smaže widgets na daném framu
        """
        Smaže widgets na daném framu
        """
        try:
            children = frame.winfo_children()
        except Exception:
            return
        for widget in children:
            widget.destroy()

    def maximalized(self): # Nastavení základního spouštění (v okně/ maximalizované)
        option = self.checkbox_maximalized.get()
        if option == 1:
            Tools.save_to_config("ano","maximalized")
        else:
            Tools.save_to_config("ne","maximalized")
    
    def tray_startup_setup(self,main_console): # Nastavení základního spouštění (v okně/ maximalizované)
        option = self.tray_checkbox.get()
        if option == 1:
            Tools.save_to_config("ano","tray_icon_startup")
            new_task_success = Tools.establish_startup_tray()
            if str(new_task_success) == "need_access":
                Tools.call_again_as_admin("settings_tray","Upozornění","Aplikace vyžaduje práva pro nastavení aut. spouštění na pozadí\n\n- přejete si znovu spustit aplikaci, jako administrátor?")
                main_console.configure(text = "Jsou vyžadována admin práva",text_color="red")
            else:
                # Tools.establish_startup_tray()
                main_console.configure(text = "Automatické spouštění úspěšně nastaveno",text_color="green")

        else:
            Tools.save_to_config("ne","tray_icon_startup")
            remove_task_success = Tools.remove_task_from_TS("TRIMAZKON_startup_tray_setup")
            if str(remove_task_success) == "need_access":
                Tools.call_again_as_admin("settings_tray_del","Upozornění","Aplikace vyžaduje práva pro odstranění aut. spouštění na pozadí\n\n- přejete si znovu spustit aplikaci, jako administrátor?")
                main_console.configure(text = "Jsou vyžadována admin práva",text_color="red")
            else:
                main_console.configure(text = "Automatické spouštění úspěšně odstraněno",text_color="green")

    def set_safe_mode(self): # Nastavení základního spouštění (v okně/ maximalizované)
        option = self.checkbox_safe_mode.get()
        if option == 1:
            Tools.save_to_config("ano","sorting_safe_mode")
        else:
            Tools.save_to_config("ne","sorting_safe_mode")

    def refresh_main_window(self):
        self.clear_frame(self.root)
        self.clear_frame(self.current_root)
        self.current_root.destroy()
        if self.spec_location == "image_browser":
            Image_browser(root=self.root,path_given=self.path_to_remember,params_given=self.ib_last_params)
        elif self.spec_location == "converting_option":
            Converting_option(self.root)
        elif self.spec_location == "deleting_option":
            Deleting_option(self.root)
        elif self.spec_location == "sorting_option":
            Sorting_option(self.root)

    def setting_widgets(self,exception=False,main_console_text = "",main_console_text_color = "white",submenu_option = None): # samotné možnosti úprav parametrů uložených v config souboru
        """
        Nabídka možností úprav

        0 = default_path
        1 = set_folder_names
        2 = set_default_parametres
        3 = set_supported_formats
        4 = set_image_browser_setting

        """

        if self.tray_setting_status_message != None:
            main_console_text = self.tray_setting_status_message
            main_console_text_color = "green"

        self.clear_frame(self.bottom_frame_default_path)
        text_file_data = Tools.read_config_data()
        if exception == False:
            cutoff_date = text_file_data[4]
        else:
            cutoff_date = exception
        
        files_to_keep = text_file_data[3]
        default_prefix_func=text_file_data[5]
        default_prefix_cam =text_file_data[6]
        self.drop_down_prefix_dir_names_list = [(str(default_prefix_cam)+" (pro třídění podle č. kamery)"),(str(default_prefix_func)+" (pro třídění podle č. funkce)")]
        default_max_num_of_pallets=text_file_data[8]
        self.drop_down_static_dir_names_list = text_file_data[9]
        # pridani defaultniho nazvu pred zmenami do drop down menu
        for i in range(0,len(self.drop_down_static_dir_names_list)):
            self.drop_down_static_dir_names_list[i] += self.default_dir_names[i]

        row_index = 0

        for buttons in self.option_buttons:
            buttons.configure(fg_color = "black")

        def call_browseDirectories(): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if select_by_dir.get() == 1:
                output = Tools.browseDirectories("only_dirs")
            else:
                output = Tools.browseDirectories("all")
            if str(output[1]) != "/":
                self.path_set.delete("0","200")
                self.path_set.insert("0", output[1])
                console_input = Tools.save_to_config(output[1],"default_path") # hlaska o nove vlozene ceste
                default_path_insert_console.configure(text="")
                default_path_insert_console.configure(text = "Aktuálně nastavená základní cesta k souborům: " + str(output[1]),text_color="white")
                main_console.configure(text="")
                main_console.configure(text=console_input,text_color="green")
            else:
                main_console.configure(text = str(output[0]),text_color="red")

        def save_path():
            path_given = str(self.path_set.get())
            path_checked = Tools.path_check(path_given)
            if path_checked != False and path_checked != "/":
                console_input = Tools.save_to_config(path_checked,"default_path")
                default_path_insert_console.configure(text="")
                default_path_insert_console.configure(text = "Aktuálně nastavená základní cesta k souborům: " + str(path_checked),text_color="white")
                main_console.configure(text="")
                main_console.configure(text=console_input,text_color="green")
            elif path_checked != "/":
                main_console.configure(text="")
                main_console.configure(text=f"Zadaná cesta: {path_given} nebyla nalezena, nebude tedy uložena",text_color="red")
            elif path_checked == "/":
                main_console.configure(text="")
                main_console.configure(text="Nebyla vložena žádná cesta k souborům",text_color="red")
        
        def select_path_by_file():
            select_by_file.select()
            select_by_dir.deselect()

        def select_path_by_dir():
            select_by_dir.select()
            select_by_file.deselect()

        def set_default_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(cutoff_date[1]))
                        if int(cutoff_date[0]) > max_days_in_month:
                            cutoff_date[0] = str(max_days_in_month)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Měsíc: " + str(input_month) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení měsíce jste nezadali číslo",text_color="red")

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        cutoff_date[0] = int(input_day)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Den: " + str(input_day) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení dne jste nezadali číslo",text_color="red")

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        cutoff_date[2] = int(input_year) + 2000
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    elif len(str(input_year)) == 4:
                        cutoff_date[2] = int(input_year)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Rok: " + str(input_year) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení roku jste nezadali číslo",text_color="red")

            Tools.save_to_config(cutoff_date,"default_cutoff_date")
            self.setting_widgets(False, main_console._text,main_console._text_color,submenu_option="set_default_parametres")

        def set_files_to_keep():
            nonlocal main_console
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    files_to_keep = int(input_files_to_keep)
                    Tools.save_to_config(files_to_keep,"default_files_to_keep")
                    main_console.configure(text="")
                    main_console.configure(text="Základní počet ponechaných starších souborů nastaven na: " + str(files_to_keep),text_color="green")
                    console_files_to_keep.configure(text = "Aktuálně nastavené minimum: "+str(files_to_keep),text_color="white")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="Mimo rozsah",text_color="red")
            else:
                main_console.configure(text="")
                main_console.configure(text="Nazadali jste číslo",text_color="red")

            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_default_parametres")

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                cutoff_date[i-1]=items
            main_console.configure(text="")
            main_console.configure(text="Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)",text_color="orange")
            self.setting_widgets(cutoff_date,main_console._text,main_console._text_color,submenu_option="set_default_parametres")

        def set_new_default_prefix(which_folder):
            report = ""
            inserted_prefix = str(set_new_def_prefix.get()).replace(" ","")
            if len(inserted_prefix) != 0:
                if inserted_prefix != str(default_prefix_cam) and inserted_prefix != str(default_prefix_func):
                    if which_folder == "cam":
                        report = Tools.save_to_config(inserted_prefix,"new_default_prefix_cam")
                        self.default_displayed_prefix_dir = "cam"
                    if which_folder == "func":
                        report = Tools.save_to_config(inserted_prefix,"new_default_prefix_func")
                        self.default_displayed_prefix_dir = "func"
                    main_console.configure(text="")
                    main_console.configure(text=report,text_color="green")
                    self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_folder_names") # refresh
                else:
                    main_console.configure(text="")
                    main_console.configure(text = "Zadané jméno je již zabrané",text_color="red")
            else:
                main_console.configure(text="")
                main_console.configure(text = "Nutný alespoň jeden znak",text_color="red")
                
        def change_prefix_dir(*args):
            if str(*args) == str(self.drop_down_prefix_dir_names_list[1]):
                button_save_new_def_prefix.configure(command=lambda: set_new_default_prefix("func"))
                set_new_def_prefix.configure(placeholder_text = str(default_prefix_func))
                set_new_def_prefix.delete("0","100")
                set_new_def_prefix.insert("0", str(default_prefix_func))
            elif str(*args) == str(self.drop_down_prefix_dir_names_list[0]):
                button_save_new_def_prefix.configure(command=lambda: set_new_default_prefix("cam"))
                set_new_def_prefix.configure(placeholder_text = str(default_prefix_cam))
                set_new_def_prefix.delete("0","100")
                set_new_def_prefix.insert("0", str(default_prefix_cam))

        def set_new_default_dir_name():
            inserted_new_name = str(set_new_def_folder_name.get()).replace(" ","")
            report = ["Základní název složky pro nepáry (soubory nezastoupenými všemi nalezenými formáty) změněn na: ",
                      "Základní název složky pro nalezené dvojice změněn na: ",
                      "Základní název složky se soubory, které jsou určené ke smazání změněn na: ",
                      "Základní název složky pro soubory převedené do .bmp formátu změněn na: ",
                      "Základní název složky pro soubory převedené do .jpg formátu změněn na: ",
                      "Základní název složky pro zkopírované obrázky z prohlížeče obrázků změněn na: ",
                      "Základní název složky pro přesunuté obrázky z prohlížeče obrázků změněn na: "]
            colisions = 0
            
            if len(inserted_new_name) != 0:
                for i in range(0,len(self.drop_down_static_dir_names_list)):
                    neme_list_without_suffix = str(self.drop_down_static_dir_names_list[i]).replace(str(self.default_dir_names[i]),"")
                    if inserted_new_name == neme_list_without_suffix:
                        colisions += 1        
                if colisions == 0:
                    #zjistujeme, co mame navolene
                    for i in range(0,len(self.drop_down_static_dir_names_list)):
                        
                        if str(drop_down_static_dir_names.get()) == str(self.drop_down_static_dir_names_list[i]):
                            # zasilame k zapsani informaci o nastavenem vstupu a soucasne i pozici v poli
                            Tools.save_to_config(inserted_new_name+" | "+str(i),"new_default_static_dir_name")
                            self.default_displayed_static_dir = i
                            neme_list_without_suffix = inserted_new_name.replace(str(self.default_dir_names[i]),"")
                            main_console.configure(text="")
                            main_console.configure(text=report[i]+neme_list_without_suffix,text_color="green")
                            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_folder_names") # refresh
                else:
                    main_console.configure(text="")
                    main_console.configure(text = "Zadané jméno je již zabrané",text_color="red")
            else:
                main_console.configure(text="")
                main_console.configure(text = "Nutný alespoň jeden znak",text_color="red")
                
        def change_static_dir(*args):
            for i in range(0,len(self.drop_down_static_dir_names_list)):
                if str(self.drop_down_static_dir_names_list[i]) == str(*args):
                    neme_list_without_suffix = str(self.drop_down_static_dir_names_list[i]).replace(str(self.default_dir_names[i]),"")
                    set_new_def_folder_name.configure(placeholder_text = neme_list_without_suffix)
                    set_new_def_folder_name.delete("0","100")
                    set_new_def_folder_name.insert("0", neme_list_without_suffix)
        
        def add_format(which_operation):
            if which_operation == 0:
                new_format = str(formats_set.get())
                if new_format !="":
                    main_console_text_add = Tools.save_to_config(new_format,"add_supported_sorting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_add,text_color="white")
                    
            if which_operation == 1:
                new_format = str(formats_deleting_input.get())
                if new_format !="":
                    main_console_text_add = Tools.save_to_config(new_format,"add_supported_deleting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_add,text_color="white")
            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_supported_formats")

        def pop_format(which_operation):
            if which_operation == 0:
                format_to_delete = str(formats_set.get())
                if format_to_delete !="":
                    main_console_text_pop = Tools.save_to_config(format_to_delete,"pop_supported_sorting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_pop,text_color="white")
            if which_operation == 1:
                format_to_delete = str(formats_deleting_input.get())
                if format_to_delete !="":
                    main_console_text_pop = Tools.save_to_config(format_to_delete,"pop_supported_deleting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_pop,text_color="white")

            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_supported_formats")

        def set_max_num_of_pallets():
            nonlocal main_console
            input_1 = set_max_pallets.get()
            if input_1.isdigit() == False:
                main_console.configure(text="")
                main_console.configure(text = "Nezadali jste číslo",text_color="red")
            elif int(input_1) <1:
                main_console.configure(text="")
                main_console.configure(text = "Mimo rozsah",text_color="red")
            else:
                main_console.configure(text="")
                main_console.configure(text = f"Počet palet nastaven na: {input_1}",text_color="green")
                Tools.save_to_config(input_1,"pallets_set")
        
        def update_zoom_increment_slider(*args):
            if self.text_file_data[11][1] != int(*args):
                label_IB4.configure(text = str(int(*args)) + " %")
                self.text_file_data[11][1] = int(*args)

        def update_drag_movement_slider(*args):
            if self.text_file_data[11][2] != int(*args):
                label_IB6.configure(text = str(int(*args)) + " px")
                self.text_file_data[11][2] = int(*args)

        def on_off_image_film():
            if switch_image_film.get() == 1:
                Tools.save_to_config("ano","image_film")
            else:
                Tools.save_to_config("ne","image_film")

        def change_image_film_number(*args):
            input_number = int(*args)
            num_of_image_film_images.configure(text = str(input_number) + " obrázků na každé straně")

        def manage_app_zoom(*args):
            app_zoom_percent.configure(text = str(int(*args)) + " %")

        def windows_zoom_setting():
            def get_screen_dpi():
                user32 = ctypes.windll.user32
                user32.SetProcessDPIAware()  # Make sure the process is DPI aware
                hdc = user32.GetDC(0)
                dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # 88 is the index for LOGPIXELSX
                return dpi

            if checkbox_app_zoom.get() == 1:
                Tools.save_to_config("ano","app_zoom_checkbox")
                current_dpi = get_screen_dpi()
                if current_dpi == 96:
                    set_zoom(100)
                elif current_dpi == 120:
                    set_zoom(125)
                elif current_dpi == 144:
                    set_zoom(150)
                app_zoom_slider.configure(state = "disabled",button_color = "gray50",button_hover_color = "gray50")
            else:
                app_zoom_slider.configure(state = "normal",button_color = "#3a7ebf",button_hover_color = "#3a7ebf")
                Tools.save_to_config("ne","app_zoom_checkbox")
                set_zoom(int(app_zoom_slider.get()))

        if submenu_option == "default_path":
            self.option_buttons[0].configure(fg_color="#212121")
            row_index = 1
            first_option_frame =        customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            self.checkbox_maximalized = customtkinter.CTkCheckBox(master = first_option_frame,height=40,text = "Spouštět v maximalizovaném okně",command = lambda: self.maximalized(),font=("Arial",22,"bold"))
            first_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")

            tray_option_frame =         customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            self.tray_checkbox =        customtkinter.CTkCheckBox(master = tray_option_frame,height=40,text = "Spouštět TRIMAZKON na pozadí (v systémové nabídce \"tray_icons\") při zapnutí systému Windows?",command = lambda: self.tray_startup_setup(main_console),font=("Arial",22,"bold"))
            tray_option_frame.          pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")

            current_zoom = self.text_file_data[21]
            new_option_frame =          customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            new_option_frame.           pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            zomm_app_label =            customtkinter.CTkLabel(master = new_option_frame,height=20,text = "Nastavte celkové přiblížení aplikace:",justify = "left",font=("Arial",22,"bold"))
            checkbox_app_zoom =         customtkinter.CTkCheckBox(master = new_option_frame,height=40,text = "Použít nastavení Windows",command = lambda: windows_zoom_setting(),font=("Arial",22,"bold"))
            app_zoom_slider =           customtkinter.CTkSlider(master = new_option_frame,width=300,height=15,from_=60,to=200,number_of_steps= 14,command = lambda e: manage_app_zoom(e))
            app_zoom_percent =          customtkinter.CTkLabel(master= new_option_frame,height=20,text = str(current_zoom) + " %",justify = "left",font=("Arial",20))
            zomm_app_label.             grid(column =0,row=0,sticky = tk.W,pady =(10,10),padx=10)
            app_zoom_slider.            grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=10)
            app_zoom_percent.           grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=320)
            checkbox_app_zoom.          grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=400)

            second_option_frame =        customtkinter.CTkFrame(    master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            second_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label5 =                    customtkinter.CTkLabel(     master = second_option_frame,height=40,text = "Nastavte základní cestu k souborům při spuštění:",justify = "left",font=("Arial",22,"bold"))
            explorer_settings_label =   customtkinter.CTkLabel(     master = second_option_frame,height=40,text = "Nastavení EXPLORERU: ",justify = "left",font=("Arial",20,"bold"))
            select_by_dir =             customtkinter.CTkCheckBox(  master = second_option_frame,height=40,text = "Vybrat cestu zvolením složky",font=("Arial",20),command = lambda: select_path_by_dir())
            select_by_file =            customtkinter.CTkCheckBox(  master = second_option_frame,height=40,text = "Vybrat cestu zvolením souboru (jsou viditelné při vyhledávání)",font=("Arial",20),command = lambda: select_path_by_file())
            self.path_set =             customtkinter.CTkEntry(     master = second_option_frame,width=800,height=40,font=("Arial",20),placeholder_text="")
            button_save5 =              customtkinter.CTkButton(    master = second_option_frame,width=100,height=40, text = "Uložit", command = lambda: save_path(),font=("Arial",22,"bold"))
            button_explorer =           customtkinter.CTkButton(    master = second_option_frame,width=100,height=40, text = "EXPLORER", command = lambda: call_browseDirectories(),font=("Arial",22,"bold"))
            default_path_insert_console=customtkinter.CTkLabel(     master = second_option_frame,height=40,text ="",justify = "left",font=("Arial",22),text_color="white")
            console_frame =             customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.              pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =              customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =            customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.             pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =   customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            self.checkbox_maximalized.  grid(column =0,row=row_index-1,sticky = tk.W,pady =20,padx=10)
            self.tray_checkbox.         grid(column =0,row=row_index-1,sticky = tk.W,pady =20,padx=10)
            label5.                     grid(column =0,row=row_index,sticky = tk.W,pady =(5,0),padx=10)
            explorer_settings_label.    grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            select_by_dir .             grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=250)
            select_by_file.             grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=550)
            self.path_set.              grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
            button_save5.               grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=815)
            button_explorer.            grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=920)
            default_path_insert_console.grid(column =0,row=row_index+3,sticky = tk.W,pady =10,padx=10)
            main_console.               grid(column =0,row=row_index+4,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.    pack(pady =5,padx=10,anchor = "e")
            select_by_dir.select()

            def save_path_enter_btn(e):
                save_path()
                self.current_root.focus_set()
            self.path_set.bind("<Return>",save_path_enter_btn)

            app_zoom_slider.set(self.text_file_data[21])
            if self.text_file_data[22] == "ano":
                checkbox_app_zoom.select()
                windows_zoom_setting()

            def slider_released(e):
                """
                save after the slider is released - it still opening and closing excel otherwise
                """
                if not checkbox_app_zoom.get() == 1:
                    current_zoom = int(app_zoom_slider.get())
                    Tools.save_to_config(current_zoom,"app_zoom")
                    self.set_zoom(current_zoom)

            app_zoom_slider.bind("<ButtonRelease-1>",lambda e: slider_released(e))

            if text_file_data[2] != False and text_file_data[2] != "/":
                default_path_insert_console.configure(text="Aktuálně nastavená základní cesta k souborům: " + str(text_file_data[2]),text_color="white")
                self.path_set.configure(placeholder_text=str(text_file_data[2]))
                self.path_set.delete("0","200")
                self.path_set.insert("0", str(text_file_data[2]))
            else:
                default_path_insert_console.configure(text="Aktuálně nastavená základní cesta k souborům v konfiguračním souboru je neplatná",text_color="red")
                self.path_set.configure(placeholder_text="Není nastavena žádná základní cesta")
            
            if text_file_data[7] == "ano":
                self.checkbox_maximalized.select()
            else:
                self.checkbox_maximalized.deselect()

            if text_file_data[24] == "ano":
                self.tray_checkbox.select()
            else:
                self.tray_checkbox.deselect()

        if submenu_option == "set_folder_names":
            self.option_buttons[1].configure(fg_color="#212121")
            #upravovani prefixu slozek, default: pro trideni podle kamer
            first_option_frame =            customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            first_option_frame.             pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_folder_prefixes           = customtkinter.CTkLabel(master = first_option_frame,height=40,text = "1. Vyberte prefix složky, u které chcete změnit základní název:",justify = "left",font=("Arial",22,"bold"))
            drop_down_dir_names             = customtkinter.CTkOptionMenu(master = first_option_frame,width=290,height=40,values=self.drop_down_prefix_dir_names_list,font=("Arial",20),command= change_prefix_dir)
            set_new_def_prefix              = customtkinter.CTkEntry(master = first_option_frame,width=200,height=40,font=("Arial",20), placeholder_text= str(default_prefix_cam))
            button_save_new_def_prefix      = customtkinter.CTkButton(master = first_option_frame,width=100,height=40, text = "Uložit", command = lambda: set_new_default_prefix("cam"),font=("Arial",22,"bold"))
            label_folder_prefixes.          grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            set_new_def_prefix.             grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
            button_save_new_def_prefix.     grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=215)
            drop_down_dir_names.            grid(column =0,row=row_index+3,sticky = tk.W,pady =5,padx=10)

            set_new_def_prefix.insert("0", str(default_prefix_cam))
            def prefix_enter_btn(e):
                if str(drop_down_dir_names.get()) == str(self.drop_down_prefix_dir_names_list[0]):
                    set_new_default_prefix("cam")
                elif str(drop_down_dir_names.get()) == str(self.drop_down_prefix_dir_names_list[1]):
                    set_new_default_prefix("func")
                self.current_root.focus_set()
            set_new_def_prefix.bind("<Return>",prefix_enter_btn)
            #nastaveni defaultniho vyberu z drop-down menu
            if self.default_displayed_prefix_dir == "cam":
                change_prefix_dir(self.drop_down_prefix_dir_names_list[0])
                drop_down_dir_names.set(self.drop_down_prefix_dir_names_list[0])
            elif self.default_displayed_prefix_dir == "func":
                change_prefix_dir(self.drop_down_prefix_dir_names_list[1])
                drop_down_dir_names.set(self.drop_down_prefix_dir_names_list[1])
            
            #widgets na nastaveni jmen statickych slozek
            second_option_frame =           customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            second_option_frame.            pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_folder_name_change        = customtkinter.CTkLabel(master = second_option_frame,height=40,text = "2. Vyberte složku, u které chcete změnit základní název",justify = "left",font=("Arial",22,"bold"))
            set_new_def_folder_name         = customtkinter.CTkEntry(master = second_option_frame,width=200,height=40,font=("Arial",20), placeholder_text= str(default_prefix_func))
            button_save_new_name            = customtkinter.CTkButton(master = second_option_frame,width=100,height=40, text = "Uložit", command = lambda: set_new_default_dir_name(),font=("Arial",22,"bold"))
            drop_down_static_dir_names      = customtkinter.CTkOptionMenu(master = second_option_frame,width=290,height=40,values=self.drop_down_static_dir_names_list,font=("Arial",20),command= change_static_dir)
            console_frame                   = customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.                  pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console                    = customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =                customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =           customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            label_folder_name_change.       grid(column =0,row=row_index+4,sticky = tk.W,pady =10,padx=10)
            set_new_def_folder_name.        grid(column =0,row=row_index+5,sticky = tk.W,pady =0,padx=10)
            button_save_new_name.           grid(column =0,row=row_index+5,sticky = tk.W,pady =0,padx=215)
            drop_down_static_dir_names.     grid(column =0,row=row_index+6,sticky = tk.W,pady =5,padx=10)
            main_console.                   grid(column =0,row=row_index+7,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.        pack(pady =5,padx=10,anchor = "e")

            drop_down_increment = self.default_displayed_static_dir
            corrected_default_input = str(self.drop_down_static_dir_names_list[drop_down_increment]).replace(str(self.default_dir_names[drop_down_increment]),"")
            set_new_def_folder_name.insert("0", corrected_default_input)
            def static_dir_enter_btn(e):
                set_new_default_dir_name()
                self.current_root.focus_set()
            set_new_def_folder_name.bind("<Return>",static_dir_enter_btn)
            #nastaveni defaultniho vyberu z drop-down menu
            drop_down_static_dir_names.set(self.drop_down_static_dir_names_list[drop_down_increment])

        if submenu_option == "set_default_parametres":
            self.option_buttons[2].configure(fg_color="#212121")
            #widgets na nastaveni zakladniho poctu palet v obehu
            first_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            first_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_pallets =                     customtkinter.CTkLabel(master = first_option_frame,height=50,text = "1. Nastavte základní maximální počet paletek v oběhu:",justify = "left",font=("Arial",22,"bold"))
            set_max_pallets =                   customtkinter.CTkEntry(master = first_option_frame,width=100,height=50,font=("Arial",20), placeholder_text= str(default_max_num_of_pallets))
            button_save_max_num_of_pallets =    customtkinter.CTkButton(master = first_option_frame,width=100,height=50, text = "Uložit", command = lambda: set_max_num_of_pallets(),font=("Arial",22,"bold"))
            label_pallets.                      grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            set_max_pallets.                    grid(column =0,row=row_index+2,sticky = tk.W,pady =(0,10),padx=10)
            button_save_max_num_of_pallets.     grid(column =0,row=row_index+2,sticky = tk.W,pady =(0,10),padx=115)
            def new_max_pallets_enter_btn(e):
                set_max_num_of_pallets()
                self.current_root.focus_set()
            set_max_pallets.bind("<Return>",new_max_pallets_enter_btn)

            #widgets na nastaveni zakladniho poctu files_to_keep
            files_to_keep_console_text ="Aktuálně nastavené minimum: "+str(files_to_keep)
            second_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            second_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_files_to_keep =               customtkinter.CTkLabel(master = second_option_frame,height=50,text = "2. Nastavte základní počet ponechaných souborů, vyhodnocených jako starších:",justify = "left",font=("Arial",22,"bold"))
            files_to_keep_set =                 customtkinter.CTkEntry(master = second_option_frame,width=100,height=50,font=("Arial",20), placeholder_text= files_to_keep)
            button_save2 =                      customtkinter.CTkButton(master = second_option_frame,width=100,height=50, text = "Uložit", command = lambda: set_files_to_keep(),font=("Arial",22,"bold"))
            console_files_to_keep=              customtkinter.CTkLabel(master = second_option_frame,height=50,text =files_to_keep_console_text,justify = "left",font=("Arial",22))
            label_files_to_keep.                grid(column =0,row=row_index+3,sticky = tk.W,pady =10,padx=10)
            files_to_keep_set.                  grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
            button_save2.                       grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=115)
            console_files_to_keep.              grid(column =0,row=row_index+5,sticky = tk.W,pady =(0,10),padx=10)
            def files_to_keep_enter_btn(e):
                set_files_to_keep()
                self.current_root.focus_set()
            files_to_keep_set.bind("<Return>",files_to_keep_enter_btn)
            
            #widgets na nastaveni zakladniho dne
            third_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            third_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_set_default_date =            customtkinter.CTkLabel(master = third_option_frame,height=50,text = "3. Nastavte základní datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",22,"bold"))
            set_day =                           customtkinter.CTkEntry(master = third_option_frame,width=40,height=50,font=("Arial",20), placeholder_text= cutoff_date[0])
            sep1 =                              customtkinter.CTkLabel(master = third_option_frame,height=50,width=10,text = ".",font=("Arial",22))
            set_month =                         customtkinter.CTkEntry(master = third_option_frame,width=40,height=50,font=("Arial",20), placeholder_text= cutoff_date[1])
            sep2 =                              customtkinter.CTkLabel(master = third_option_frame,height=50,width=10,text = ".",font=("Arial",22))
            set_year =                          customtkinter.CTkEntry(master = third_option_frame,width=60,height=50,font=("Arial",20), placeholder_text= cutoff_date[2])
            button_save_date =                  customtkinter.CTkButton(master = third_option_frame,width=100,height=50, text = "Uložit", command = lambda: set_default_cutoff_date(),font=("Arial",22,"bold"))
            insert_button =                     customtkinter.CTkButton(master = third_option_frame,width=285,height=50, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",22,"bold"))
            console_frame =                     customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.                      pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =                      customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =                    customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.                     pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =               customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            label_set_default_date.             grid(column =0,row=row_index+6,sticky = tk.W,pady =10,padx=10)
            set_day.                            grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=10)
            sep1.                               grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=55)
            set_month.                          grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=70)
            sep2.                               grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=115)
            set_year.                           grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=130)
            button_save_date.                   grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=195)
            insert_button.                      grid(column =0,row=row_index+8,sticky = tk.W,pady =5,padx=10)
            main_console.                       grid(column =0,row=row_index+9,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.            pack(pady =5,padx=10,anchor = "e")

            def new_date_enter_btn(e):
                set_default_cutoff_date()
                self.current_root.focus_set()
            set_day.bind("<Return>",new_date_enter_btn)
            set_month.bind("<Return>",new_date_enter_btn)
            set_year.bind("<Return>",new_date_enter_btn)

        if submenu_option == "set_supported_formats":
            self.option_buttons[3].configure(fg_color="#212121")
            #widgets pro nastavovani podporovanych formatu
            supported_formats_deleting = "Aktuálně nastavené podporované formáty pro možnosti mazání: " + str(text_file_data[1])
            first_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            first_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_supported_formats_deleting =  customtkinter.CTkLabel(master = first_option_frame,height=50,text = "1. Nastavte podporované formáty pro možnosti: MAZÁNÍ:",justify = "left",font=("Arial",22,"bold"))
            formats_deleting_input =            customtkinter.CTkEntry(master = first_option_frame,height=50,font=("Arial",20),width=200)
            button_save4 =                      customtkinter.CTkButton(master = first_option_frame,width=50,height=50, text = "Uložit", command = lambda: add_format(1),font=("Arial",22,"bold"))
            button_pop2 =                       customtkinter.CTkButton(master = first_option_frame,width=70,height=50, text = "Odebrat", command = lambda: pop_format(1),font=("Arial",22,"bold"))
            console_bottom_frame_4=             customtkinter.CTkLabel(master = first_option_frame,height=50,text =supported_formats_deleting,justify = "left",font=("Arial",22))
            label_supported_formats_deleting.   grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            formats_deleting_input.             grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
            button_save4.                       grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=215)
            button_pop2.                        grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=290)
            console_bottom_frame_4.             grid(column =0,row=row_index+3,sticky = tk.W,pady =0,padx=10)

            supported_formats_sorting = "Aktuálně nastavené podporované formáty pro možnosti třídění: " + str(text_file_data[0])
            second_option_frame =               customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            second_option_frame.                pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label3 =                            customtkinter.CTkLabel(master = second_option_frame,height=50,text = "2. Nastavte podporované formáty pro možnosti: TŘÍDĚNÍ:",justify = "left",font=("Arial",22,"bold"))
            formats_set =                       customtkinter.CTkEntry(master = second_option_frame,width=200,height=50,font=("Arial",20))
            button_save3 =                      customtkinter.CTkButton(master = second_option_frame,width=50,height=50, text = "Uložit", command = lambda: add_format(0),font=("Arial",22,"bold"))
            button_pop =                        customtkinter.CTkButton(master = second_option_frame,width=70,height=50, text = "Odebrat", command = lambda: pop_format(0),font=("Arial",22,"bold"))
            console_bottom_frame_3=             customtkinter.CTkLabel(master = second_option_frame,height=50,text =supported_formats_sorting,justify = "left",font=("Arial",22))
            console_frame =                     customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.                      pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =                      customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =                    customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.                     pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =               customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            label3.                             grid(column =0,row=row_index+4,sticky = tk.W,pady =10,padx=10)
            formats_set.                        grid(column =0,row=row_index+5,sticky = tk.W,pady =0,padx=10)
            button_save3.                       grid(column =0,row=row_index+5,sticky = tk.W,pady =0,padx=215)
            button_pop.                         grid(column =0,row=row_index+5,sticky = tk.W,pady =0,padx=290)
            console_bottom_frame_3.             grid(column =0,row=row_index+6,sticky = tk.W,pady =0,padx=10)
            main_console.                       grid(column =0,row=row_index+7,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.            pack(pady =5,padx=10,anchor = "e")

            def add_or_rem_formats(e):
                self.current_root.focus_set()
            formats_deleting_input.bind("<Return>",add_or_rem_formats)
            formats_set.bind("<Return>",add_or_rem_formats)

        if submenu_option == "set_image_browser_setting":
            self.option_buttons[4].configure(fg_color="#212121")
            text_increment = str(self.text_file_data[11][1]) + " %"
            text_movement = str(self.text_file_data[11][2]) + " px"
            text_image_film = str(self.text_file_data[14]) + " obrázků na každé straně"
            second_option_frame =       customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            second_option_frame.        pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_IB3 =                 customtkinter.CTkLabel(master = second_option_frame,height=20,text = "1. Nastavte o kolik procent se navýší přiblížení jedním krokem kolečka myši:",justify = "left",font=("Arial",22,"bold"))
            zoom_increment_set =        customtkinter.CTkSlider(master=second_option_frame,width=300,height=15,from_=5,to=100,number_of_steps= 19,command= update_zoom_increment_slider)
            label_IB4 =                 customtkinter.CTkLabel(master = second_option_frame,height=20,text = text_increment,justify = "left",font=("Arial",20))
            third_option_frame =        customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            third_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_IB5 =                 customtkinter.CTkLabel(master = third_option_frame,height=20,text = "2. Nastavte velikost kroku při posouvání levým tlačítkem myši:",justify = "left",font=("Arial",22,"bold"))
            zoom_movement_set =         customtkinter.CTkSlider(master=third_option_frame,width=300,height=15,from_=10,to=100,number_of_steps= 18,command= update_drag_movement_slider)
            label_IB6 =                 customtkinter.CTkLabel(master = third_option_frame,height=20,text = text_movement,justify = "left",font=("Arial",20))
            forth_option_frame =        customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            forth_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_image_film =          customtkinter.CTkLabel(master = forth_option_frame,height=20,text = "3. Upravte nastavení filmu obrázků:",justify = "left",font=("Arial",22,"bold"))
            switch_image_film =         customtkinter.CTkCheckBox(master = forth_option_frame, text = "Zapnuto",command = lambda: on_off_image_film(),font=("Arial",20))
            num_of_image_film_images_slider = customtkinter.CTkSlider(master=forth_option_frame,width=300,height=15,from_=1,to=15,command= change_image_film_number)
            num_of_image_film_images =  customtkinter.CTkLabel(master = forth_option_frame,height=20,text = text_image_film,justify = "left",font=("Arial",20))
            console_frame =             customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.              pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =              customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =            customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.             pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =   customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            label_IB3.                  grid(column =0,row=row_index+6,sticky = tk.W,pady =10,padx=10)
            zoom_increment_set.         grid(column =0,row=row_index+7,sticky = tk.W,pady =(10,20),padx=10)
            label_IB4.                  grid(column =0,row=row_index+7,sticky = tk.W,pady =(10,20),padx=320)
            label_IB5.                  grid(column =0,row=row_index+8,sticky = tk.W,pady =10,padx=10)
            zoom_movement_set.          grid(column =0,row=row_index+9,sticky = tk.W,pady =(10,20),padx=10)
            label_IB6.                  grid(column =0,row=row_index+9,sticky = tk.W,pady =(10,20),padx=320)
            label_image_film.           grid(column =0,row=row_index+10,sticky = tk.W,pady =10,padx=10)
            switch_image_film.          grid(column =0,row=row_index+11,sticky = tk.W,pady =0,padx=10)
            num_of_image_film_images_slider.grid(column =0,row=row_index+12,sticky = tk.W,pady =(10,20),padx=10)
            num_of_image_film_images.   grid(column =0,row=row_index+12,sticky = tk.W,pady =(10,20),padx=320)
            main_console.               grid(column =0,row=row_index+13,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.    pack(pady =5,padx=10,anchor = "e")

            zoom_increment_set.set(self.text_file_data[11][1])
            zoom_movement_set.set(self.text_file_data[11][2])
            num_of_image_film_images_slider.set(self.text_file_data[14])

            def slider_released(e,parameter):
                """
                save after the slider is released - it still opening and closing excel otherwise
                """
                if parameter == "zoom_increment":
                    Tools.save_to_config([None,int(zoom_increment_set.get()),None],"image_browser_param_set")
                elif parameter == "zoom_move":
                    Tools.save_to_config([None,None,int(zoom_movement_set.get())],"image_browser_param_set")
                elif parameter == "IB_image_num":
                    Tools.save_to_config(int(num_of_image_film_images_slider.get()),"num_of_IB_film_images")

            zoom_increment_set.bind("<ButtonRelease-1>",lambda e: slider_released(e,"zoom_increment"))
            zoom_movement_set.bind("<ButtonRelease-1>",lambda e: slider_released(e,"zoom_move"))
            num_of_image_film_images_slider.bind("<ButtonRelease-1>",lambda e: slider_released(e,"IB_image_num"))
            if self.text_file_data[13] == "ano":
                switch_image_film.select()  

    def creating_advanced_option_widgets(self): # Vytváří veškeré widgets (advance option MAIN)
        if self.windowed:
            self.current_root=customtkinter.CTkToplevel()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            self.current_root.geometry(f"1250x900+{x+200}+{y+200}")
            self.current_root.title("Pokročilá nastavení")
            self.current_root.after(200, lambda: self.current_root.iconbitmap(Tools.resource_path(app_icon)))
        else:
            self.current_root = self.root
        self.bottom_frame_default_path   = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,border_width = 0)
        self.top_frame                   = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,border_width = 0)
        self.menu_buttons_frame          = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,fg_color="#636363",height=50,border_width = 0)
        self.top_frame.                 pack(pady=(2.5,0),padx=5,fill="x",expand=False,side = "top")
        self.menu_buttons_frame.        pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.bottom_frame_default_path. pack(pady=(0,2.5),padx=5,fill="both",expand=True,side = "bottom")
        
        label0          = customtkinter.CTkLabel(master = self.top_frame,height=20,text = "Nastavte požadované parametry (nastavení bude uloženo i po vypnutí aplikace): ",justify = "left",font=("Arial",22,"bold"))
        main_menu_button =  customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "MENU",                  command =  lambda: self.call_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options0 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Základní nastavení",    command =  lambda: self.setting_widgets(submenu_option="default_path"),font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        options1 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Názvy složek",          command =  lambda: self.setting_widgets(submenu_option="set_folder_names"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options2 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Počáteční parametry",   command =  lambda: self.setting_widgets(submenu_option="set_default_parametres"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options3 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Podporované formáty",   command =  lambda: self.setting_widgets(submenu_option="set_supported_formats"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options4 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Prohlížeč obrázků",     command =  lambda: self.setting_widgets(submenu_option="set_image_browser_setting"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        label0.             grid(column = 0,row=0,sticky = tk.W,pady =10,padx=10)
        shift_const = 210
        if not self.windowed:
            main_menu_button.grid(column = 0,row=0,pady = (10,0),padx =10,sticky = tk.W)
            shift_const = 0
        options0.           grid(column = 0,row=0,pady = (10,0),padx =220-shift_const,sticky = tk.W)
        options1.           grid(column = 0,row=0,pady = (10,0),padx =430-shift_const,sticky = tk.W)
        options2.           grid(column = 0,row=0,pady = (10,0),padx =640-shift_const,sticky = tk.W)
        options3.           grid(column = 0,row=0,pady = (10,0),padx =850-shift_const,sticky = tk.W)
        options4.           grid(column = 0,row=0,pady = (10,0),padx =1070-shift_const,sticky = tk.W)
        self.option_buttons = [options0,options1,options2,options3,options4]

        if self.windowed and not global_recources_load_error:
            if self.spec_location == "image_browser":
                self.setting_widgets(submenu_option="set_image_browser_setting")
            else:
                self.setting_widgets(submenu_option="default_path")
        elif not global_recources_load_error:
            self.setting_widgets(submenu_option="default_path")
        elif global_recources_load_error:
            error_label = customtkinter.CTkLabel(master = self.bottom_frame_default_path,height=20,text = "Nepodařilo se načíst konfigurační soubor config_TRIMAZKON.xlsx (nastavení se nemá kam uložit)",justify = "left",font=("Arial",22,"bold"),text_color="red")
            error_label.grid(column = 0,row=0,pady = (10,0),padx =20,sticky = tk.W)
            options0.configure(state = "disabled")
            options1.configure(state = "disabled")
            options2.configure(state = "disabled")
            options3.configure(state = "disabled")
            options4.configure(state = "disabled")


        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.current_root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.current_root._current_width) > 1200:
                self.current_root.after(0, lambda:self.current_root.state('normal'))
                self.current_root.geometry("1250x900")
            else:
                self.current_root.after(0, lambda:self.current_root.state('zoomed'))
        self.current_root.bind("<f>",maximalize_window)
        self.unbind_list.append("<f>")

        def unfocus_widget(e):
            self.current_root.focus_set()
        self.current_root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")

        if self.windowed:
            self.current_root.update()
            self.current_root.update_idletasks()
            self.current_root.focus_force()
            self.current_root.focus()
            # click outside the window - kill it
            self.root.bind("<Button-1>",lambda e: self.current_root.destroy())

class Converting_option: # Spouští možnosti konvertování typu souborů
    """
    Spouští možnosti konvertování typu souborů

    -Spouští přes příkazový řádek command, který je vykonáván v externí aplikaci s dll knihovnami
    """
    def __init__(self,root):
        self.root = root
        self.text_file_data = Tools.read_config_data()
        list_of_folder_names = self.text_file_data[9]
        self.bmp_folder_name = list_of_folder_names[3]
        self.jpg_folder_name = list_of_folder_names[4]
        self.temp_path_for_explorer = None
        self.create_convert_option_widgets()
    
    def call_extern_function(self,list_of_frames,function:str): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu\n
        function:
        - menu
        - sorting
        - deleting
        - (converting)
        """
        for frames in list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()

        if function == "menu":
            menu.menu()
        elif function == "sorting":
            Sorting_option(self.root)
        elif function == "deleting":
            Deleting_option(self.root)

    def convert_files(self,path): # zde se volá externí script
        selected_format = "bmp"
        if self.checkbox_bmp.get() == 1:
            selected_format = "bmp"
        if self.checkbox_jpg.get() == 1:
            selected_format = "jpg"

        def trigger_progress_bar(interval):
            for i in range(1, 101):
                time.sleep(interval/100)
                self.loading_bar.set(value = i/100)
                self.root.update_idletasks()
                root.update_idletasks()
                self.bottom_frame2.update_idletasks()

        def call_converting_main(whole_instance):
            whole_instance.main()
        running_program = Converting.whole_converting_function(
            path,
            selected_format,
            self.bmp_folder_name,
            self.jpg_folder_name
        )

        run_background = threading.Thread(target=call_converting_main, args=(running_program,))
        run_background.start()

        completed =False
        condition_met = False
        previous_len = 0
        while not running_program.finish or completed == False:
            time.sleep(0.05)
            if running_program.processing_time != 0 and not condition_met:
                new_row = "Očekávaná doba procesu: " + str(running_program.processing_time) + " s"
                Tools.add_colored_line(self.console,str(new_row),"white")
                run_background_loading = threading.Thread(target=trigger_progress_bar(running_program.processing_time))
                run_background_loading.start()
                condition_met = True
            if int(len(running_program.output)) > previous_len:
                new_row = str(running_program.output[previous_len])
                if "Konvertování bylo dokončeno" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"green",("Arial",15,"bold"))
                elif "cesta neobsahuje" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"red",("Arial",15,"bold"))
                else:
                    Tools.add_colored_line(self.console,str(new_row),"white")
                self.console.update_idletasks()
                self.root.update_idletasks()
                previous_len +=1

            if running_program.finish and (int(len(running_program.output)) == previous_len):
                completed = True
            
        self.console.update_idletasks()
        run_background.join()

    def start(self):# Ověřování cesty, init, spuštění
        """
        Ověřování cesty, init, spuštění
        """
        Tools.clear_console(self.console)
        self.console.update_idletasks()
        if self.checkbox_bmp.get()+self.checkbox_jpg.get() == 0:
            Tools.add_colored_line(self.console,"Nevybrali jste žádný formát, do kterého se má konvertovat :-)","red")
        else:
            path = self.path_set.get() 
            if path != "":
                check = Tools.path_check(path)
                if check == False:
                    Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" nebyla nalezena","red")
                else:
                    path = check
                    Tools.add_colored_line(self.console,f"Probíhá konvertování souborů v cestě: {path}","white")
                    self.console.update_idletasks()
                    self.root.update_idletasks()
                    self.convert_files(path)
            else:
                Tools.add_colored_line(self.console,"Nebyla vložena cesta k souborům","red")

    def call_browseDirectories(self): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        if self.temp_path_for_explorer == None:
            output = Tools.browseDirectories("all")
        else:
            output = Tools.browseDirectories("all",self.temp_path_for_explorer)

        if str(output[1]) != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", output[1])
            Tools.add_colored_line(self.console,f"Byla vložena cesta: {output[1]}","green")
            self.temp_path_for_explorer = output[1]
        else:
            Tools.add_colored_line(self.console,str(output[0]),"red")

    def selected_bmp(self):
        self.checkbox_jpg.deselect()
        self.label.configure(text=f"Konvertované soubory budou vytvořeny uvnitř separátní složky: \"{self.bmp_folder_name}\"\nPodporované formáty: .ifz\nObsahuje-li .ifz soubor více obrázků, budou uloženy v následující syntaxi:\nxxx_0.bmp, xxx_1.bmp ...")
    
    def selected_jpg(self):
        self.checkbox_bmp.deselect()
        self.label.configure(text=f"Konvertované soubory budou vytvořeny uvnitř separátní složky: \"{self.jpg_folder_name}\"\nPodporované formáty: .ifz\nObsahuje-li .ifz soubor více obrázků, budou uloženy v následující syntaxi:\nxxx_0.bmp, xxx_1.bmp ...")

    def create_convert_option_widgets(self):  # Vytváří veškeré widgets (convert option MAIN)
        frame_with_logo =       customtkinter.CTkFrame(master=self.root,corner_radius=0)
        logo =                  customtkinter.CTkImage(Image.open(Tools.resource_path("images/logo.png")),size=(1200, 100))
        image_logo =            customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        frame_with_logo.        pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        image_logo.pack()
        frame_with_cards =      customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=100)
        self.frame_path_input = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.bottom_frame2 =    customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        self.bottom_frame1 =    customtkinter.CTkFrame(master=self.root,height = 80,corner_radius=0)
        frame_with_cards.       pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        self.frame_path_input.  pack(pady=5,padx=5,fill="both",expand=False,side = "top")
        self.bottom_frame2.     pack(pady=5,padx=5,fill="both",expand=True,side = "bottom")
        self.bottom_frame1.     pack(pady=0,padx=5,fill="x",expand=False,side = "bottom")

        list_of_frames = [self.frame_path_input,self.bottom_frame1,self.bottom_frame2,frame_with_cards,frame_with_logo]
        shift_const = 250
        menu_button =       customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "MENU",                  command =  lambda: self.call_extern_function(list_of_frames,function="menu"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        sorting_button =    customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Třídění souborů",      command =  lambda: self.call_extern_function(list_of_frames,function="sorting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        deleting_button =   customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Mazání souborů",        command =  lambda: self.call_extern_function(list_of_frames,function="deleting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        converting_button = customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Konvertování souborů",
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        menu_button.        grid(column = 0,row=0,pady = (10,0),padx =260-shift_const,sticky = tk.W)
        sorting_button.     grid(column = 0,row=0,pady = (10,0),padx =520-shift_const,sticky = tk.W)
        deleting_button.    grid(column = 0,row=0,pady = (10,0),padx =780-shift_const,sticky = tk.W)
        converting_button.  grid(column = 0,row=0,pady = (10,0),padx =1040-shift_const,sticky = tk.W)

        self.checkbox_bmp =     customtkinter.CTkCheckBox(master = self.bottom_frame1, text = "Konvertovat do formátu .bmp",command=self.selected_bmp,font=("Arial",16,"bold"))
        self.checkbox_jpg =     customtkinter.CTkCheckBox(master = self.bottom_frame1, text = "Konvertovat do formátu .jpg",command=self.selected_jpg,font=("Arial",16,"bold"))
        self.checkbox_bmp.      pack(pady =10,padx=10,anchor ="w")
        self.checkbox_jpg.      pack(pady =10,padx=10,anchor ="w")
        self.path_set =         customtkinter.CTkEntry(master = self.frame_path_input,font=("Arial",18),placeholder_text="Zadejte cestu k souborům určeným ke konvertování (kde se soubory přímo nacházejí)")
        tree         =          customtkinter.CTkButton(master = self.frame_path_input, width = 180,text = "EXPLORER", command = self.call_browseDirectories,font=("Arial",20,"bold"))
        button_open_setting =   customtkinter.CTkButton(master = self.frame_path_input,width=30,height=30, text = "⚙️", command = lambda: Advanced_option(self.root,windowed=True,spec_location="converting_option"),font=("Arial",16))
        self.path_set.          pack(pady = 12,padx = (10,0), anchor ="w",side="left",fill="both",expand=True)
        tree.                   pack(pady = 12,padx = 10,anchor ="w",side="left")
        button_open_setting.    pack(pady = 12,padx = (0,10),anchor ="w",side="left")
        self.label   =          customtkinter.CTkLabel(master = self.bottom_frame2,text = f"Konvertované soubory budou vytvořeny uvnitř separátní složky: \"{self.bmp_folder_name}\"\nPodporované formáty: .ifz\nObsahuje-li .ifz soubor více obrázků, budou uloženy v následující syntaxi:\nxxx_0.bmp, xxx_1.bmp ...",justify = "left",font=("Arial",18,"bold"))
        button  =               customtkinter.CTkButton(master = self.bottom_frame2, text = "KONVERTOVAT", command = self.start,font=("Arial",20,"bold"))
        self.loading_bar =      customtkinter.CTkProgressBar(master = self.bottom_frame2, mode='determinate',width = 800,height =20,progress_color="green",corner_radius=0)
        self.console =          tk.Text(self.bottom_frame2, wrap="word", height=20, width=1200,background="black",font=("Arial",16),state=tk.DISABLED)
        self.label.             pack(pady =10,padx=10)
        button.                 pack(pady =20,padx=10)
        button.                 _set_dimensions(300,60)
        self.loading_bar.       pack(pady = 5,padx = 5)
        self.loading_bar.       set(value = 0)
        self.console.           pack(pady =10,padx=10)
        self.checkbox_bmp.select()
        recources_path = self.text_file_data[2]
        if recources_path != False and recources_path != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", str(recources_path))
            Tools.add_colored_line(self.console,"Byla vložena cesta z konfiguračního souboru","white")
        else:
            Tools.add_colored_line(self.console,"Konfigurační soubor obsahuje neplatnou cestu k souborům (můžete vložit v pokročilém nastavení)","orange")
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)

        def unfocus_widget(e):
            self.root.focus_set()
        self.path_set.bind("<Return>",unfocus_widget)

class Deleting_option: # Umožňuje mazat soubory podle nastavených specifikací
    """
    Umožňuje mazat soubory podle nastavených specifikací

    -obsahuje i režim testování, kde soubory pouze přesune do složky ke smazání
    -umožňuje procházet více subsložek
    
    """
    @classmethod
    def confirm_window(cls,prompt_message):
        selected_option = False
        def selected_yes(child_root):# Vyžádání admin práv: nefunkční ve vscode
            child_root.grab_release()
            child_root.destroy()
            nonlocal selected_option
            selected_option = True

        def close_prompt(child_root):
            child_root.grab_release()
            child_root.destroy()
            nonlocal selected_option
            selected_option = False
            
        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(app_icon)))
        # child_root.geometry(f"620x150+{300}+{300}")  
        child_root.title("Upozornění (první spuštění aplikace)")
        label_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        proceed_label = customtkinter.CTkLabel(master = label_frame,text = prompt_message,font=("Arial",25),anchor="w",justify="left")
        proceed_label.pack(pady=5,padx=10,anchor="w",side = "left")
        button_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_yes =   customtkinter.CTkButton(master = button_frame,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: selected_yes(child_root))
        button_no =    customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
        button_no      .pack(pady = 5, padx = 10,anchor="e",side="right")
        button_yes     .pack(pady = 5, padx = 10,anchor="e",side="right")
        label_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        button_frame   .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        child_root.update()
        child_root.update_idletasks()
        # child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{300}+{300}")
        child_root.focus()
        child_root.focus_force()
        child_root.grab_set()
        child_root.wait_window()
        return selected_option

    def __init__(self,root):
        self.root = root
        self.unbind_list = []
        self.text_file_data = Tools.read_config_data()
        self.supported_formats_deleting = self.text_file_data[1]
        self.files_to_keep = self.text_file_data[3]
        self.cutoff_date = self.text_file_data[4]
        list_of_folder_names = self.text_file_data[9]
        self.to_delete_folder_name = list_of_folder_names[2]
        self.console_frame_right_1_text = "","white"
        self.console_frame_right_2_text = "","white"
        self.config_filename = "config_TRIMAZKON.xlsx"
        self.temp_path_for_explorer = None
        self.create_deleting_option_widgets()
 
    def call_extern_function(self,list_of_frames,function:str): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu\n
        function:
        - menu
        - sorting
        - (deleting)
        - converting
        """
        
        for frames in list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()
        
        for binds in self.unbind_list:
            self.root.unbind(binds)

        if function == "menu":
            menu.menu()
        elif function == "sorting":
            Sorting_option(self.root)
        elif function == "converting":
            Converting_option(self.root)

    def start(self):# Ověřování cesty, init, spuštění
        """
        Ověřování cesty, init, spuštění
        """
        if self.checkbox.get()+self.checkbox2.get()+self.checkbox3.get() == 0:
            Tools.add_colored_line(self.console,"Nevybrali jste žádný způsob mazání","red")
            self.info.configure(text = "")

        else:
            path = self.path_set.get() 
            if path != "":
                check = Tools.path_check(path)
                if check == False:
                    Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" nebyla nalezena","red")
                else:
                    path = check
                    if self.checkbox_testing.get() != 1:
                        if self.checkbox6.get() == 1 and self.checkbox3.get() != 1: # sublozky u adresaru
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání souborů v cestě:\n{path}\na procházet přitom i SUBSLOŽKY?"
                        elif self.checkbox3.get() == 1:
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání ADRESÁŘŮ v cestě:\n{path}"
                        else:
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání souborů v cestě:\n{path}"
                        # confirm = tk.messagebox.askokcancel("Potvrzení", confirm_prompt_msg)
                        confirm = Deleting_option.confirm_window(confirm_prompt_msg)
                    else: # pokud je zapnut rezim testovani
                        confirm = True

                    if confirm == True:
                        Tools.add_colored_line(self.console,"- Provádím navolené možnosti mazání v cestě: " + str(path) + "\n","orange")
                        self.console.update_idletasks()
                        self.root.update_idletasks()
                        self.del_files(path)
                    else:
                        Tools.add_colored_line(self.console,"Zrušeno uživatelem","red")
            else:
                Tools.add_colored_line(self.console,"Nebyla vložena cesta k souborům","red")

    def del_files(self,path): # zde se volá externí script: Deleting
        testing_mode = True
        del_option = 0
        if self.checkbox.get() == 1:
            del_option = 1
        if self.checkbox2.get() == 1:
            del_option = 2
        if self.checkbox3.get() == 1:
            del_option = 3
        if self.checkbox6.get() == 1:
            self.more_dirs = True
        else:
            self.more_dirs = False
        if self.checkbox_testing.get() == 1:
            testing_mode = True
        else:
            testing_mode = False

        def call_deleting_main(whole_instance):
            whole_instance.main()

        running_deleting = Deleting.whole_deleting_function(
            path,
            self.more_dirs,
            del_option,
            self.files_to_keep,
            self.cutoff_date,
            self.supported_formats_deleting,
            testing_mode,
            self.to_delete_folder_name
            )

        run_del_background = threading.Thread(target=call_deleting_main, args=(running_deleting,))
        run_del_background.start()

        completed = False
        previous_len = 0

        while not running_deleting.finish or completed == False:
            time.sleep(0.05)
            if int(len(running_deleting.output)) > previous_len:
                new_row = str(running_deleting.output[previous_len])
                if "Mazání dokončeno" in new_row or "Zkontrolováno" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"green",("Arial",15,"bold"))
                elif "Chyba" in new_row or "Nebyly nalezeny" in new_row or "- zrušeno" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"red",("Arial",15,"bold"))
                elif "Smazalo by se" in new_row or "Smazáno souborů" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"orange",("Arial",15,"bold"))
                else:
                    Tools.add_colored_line(self.console,str(new_row),"white")
                self.console.update_idletasks()
                self.root.update_idletasks()
                previous_len +=1

            if running_deleting.finish and (int(len(running_deleting.output)) == previous_len):
                completed = True
        
        run_del_background.join()

    def call_browseDirectories(self): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        if self.checkbox6.get() == 1: # pokud je zvoleno more_dirs v exploreru pouze slozky...
            output = Tools.browseDirectories("only_dirs",self.temp_path_for_explorer)
        else:
            output = Tools.browseDirectories("all",self.temp_path_for_explorer)
        if str(output[1]) != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", output[1])
            Tools.add_colored_line(self.console,f"Byla vložena cesta: {output[1]}","green")
            self.temp_path_for_explorer = output[1]
        else:
            Tools.add_colored_line(self.console,str(output[0]),"red")

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def selected(self,clear:bool): # První možnost mazání, od nejstarších
        """
        Nastavení widgets pro první možnost mazání

        -vstup: console text a barva textu

        -Budou smazány soubory starší než nastavené datum, přičemž bude ponechán nastavený počet souborů, vyhodnocených, jako starších\n
        -Podporované formáty jsou uživatelem nastavené a uložené v textovém souboru
        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        #self.console.configure(text = "")
        Tools.clear_console(self.console)
        self.checkbox.select()
        self.checkbox2.deselect()
        self.checkbox3.deselect()
        self.info.configure(text = f"- Budou smazány soubory starší než nastavené datum, přičemž bude ponechán nastavený počet souborů, vyhodnocených, jako starších\n(Ponechány budou všechny novější než nastavené datum a k tomu bude ponecháno: {self.files_to_keep} starších souborů)\nPodporované formáty: {self.supported_formats_deleting}\n\n",
                            font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            # if set_month.get() == self.cutoff_date[1] and set_day.get() == self.cutoff_date[0] and set_day.get() == self.cutoff_date[2]
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení měsíce jste nezadali číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení dne jste nezadali číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení roku jste nezadali číslo","red"

            self.selected(False)

        def set_files_to_keep():
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)
                    self.console_frame_right_2_text = "Počet ponechaných starších souborů nastaven na: " + str(self.files_to_keep),"green"
                else:
                    self.console_frame_right_2_text = "Mimo rozsah","red"
            else:
                self.console_frame_right_2_text = "Nazadali jste číslo","red"

            self.selected(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)","orange"

            self.selected(False)

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)

        def set_max_days():
            new_cutoff = Deleting.get_cutoff_date(int(max_days_entry.get()))
            set_day.insert(0,new_cutoff[0])
            set_month.insert(0,new_cutoff[1])
            set_year.insert(0,new_cutoff[2])
            set_cutoff_date()

        console_1_text, console_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        label1      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",16))
        set_day     = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month   = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year    = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1 = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        max_days_entry = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        max_days_label = customtkinter.CTkLabel(master = self.frame_right,text = "dní",font=("Arial",16))
        max_days_save = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_max_days(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1 = customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_1_text,justify = "left",font=("Arial",18),text_color=console_1_color)
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=150)
        max_days_entry.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=260)
        max_days_label.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=320)
        max_days_save.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=350)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =5,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        max_days_entry.insert(0,Deleting.get_max_days(self.cutoff_date))

        console_2_text, console_2_color = self.console_frame_right_2_text
        label2          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte počet ponechaných souborů, vyhodnocených jako starších:",justify = "left",font=("Arial",16))
        files_to_keep_set = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.files_to_keep)
        button_save2    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_files_to_keep(),font=("Arial",18,"bold"))
        console_frame_right_2 = customtkinter.CTkLabel(master = self.frame_right,height=30,text =console_2_text,justify = "left",font=("Arial",18),text_color=console_2_color)
        label2.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)
        files_to_keep_set.grid(column =0,row=6,sticky = tk.W,pady =0,padx=10)
        button_save2.grid(column =0,row=6,sticky = tk.W,pady =0,padx=60)
        console_frame_right_2.grid(column =0,row=7,sticky = tk.W,pady =0,padx=10)
        def new_FTK_enter_btn(e):
            set_files_to_keep()
        files_to_keep_set.bind("<Return>",new_FTK_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů
        
    def selected2(self,clear:bool): # Druhá možnost mazání, mazání všech starých, redukce nových
        """
        Nastavení widgets pro druhou možnost mazání

        -Budou smazány VŠECHNY soubory starší než nastavené datum, přičemž budou redukovány i soubory novější\n
        -Souborů, vyhodnocených, jako novější, bude ponechán nastavený počet\n
        -(vhodné při situacích rychlého pořizování velkého množství fotografií, kde je potřebné ponechat nějaké soubory pro referenci)\n
        -Podporované formáty jsou uživatelem nastavené a uložené v textovém souboru
        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        Tools.clear_console(self.console)
        self.checkbox.deselect()
        self.checkbox2.select()
        self.checkbox3.deselect()
        self.info.configure(text = f"- Budou smazány VŠECHNY soubory starší než nastavené datum, přičemž budou redukovány i soubory novější\n(Ošetřeno: pokud se v dané cestě nacházejí pouze starší soubory, než nastavené datum, zruší se mazání)\n- Souborů, vyhodnocených, jako novější, než nastavené datum, bude ponecháno: {self.files_to_keep}\n(vhodné při situacích rychlého pořizování velkého množství fotografií, kde je potřebné ponechat nějaké soubory pro referenci)\nPodporované formáty: {self.supported_formats_deleting}",font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"            
            self.selected2(False)

        def set_files_to_keep():
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)
                    self.console_frame_right_2_text = "Počet ponechaných starších souborů nastaven na: " + str(self.files_to_keep),"green"
                else:
                    self.console_frame_right_2_text = "Mimo rozsah","red"
            else:
                self.console_frame_right_2_text = "Nazadali jste číslo","red"

            self.selected2(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)","orange"
            self.selected2(False)

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)

        def set_max_days():
            new_cutoff = Deleting.get_cutoff_date(int(max_days_entry.get()))
            set_day.insert(0,new_cutoff[0])
            set_month.insert(0,new_cutoff[1])
            set_year.insert(0,new_cutoff[2])
            set_cutoff_date()

        console_frame_right_1_text, console_frame_right_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        label1      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",16))
        set_day     = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month   = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year    = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1 = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        max_days_entry = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        max_days_label = customtkinter.CTkLabel(master = self.frame_right,text = "dní",font=("Arial",16))
        max_days_save = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_max_days(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1=customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_frame_right_1_text,justify = "left",font=("Arial",18),text_color=console_frame_right_1_color)
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=150)
        max_days_entry.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=260)
        max_days_label.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=320)
        max_days_save.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=350)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =5,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        max_days_entry.insert(0,Deleting.get_max_days(self.cutoff_date))
        
        console_frame_right_2_text, console_frame_right_2_color = self.console_frame_right_2_text
        label2          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte počet ponechaných novějších souborů:",justify = "left",font=("Arial",16))
        files_to_keep_set = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.files_to_keep)
        button_save2    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_files_to_keep(),font=("Arial",18,"bold"))
        console_frame_right_2=customtkinter.CTkLabel(master = self.frame_right,height=30,text =console_frame_right_2_text,justify = "left",font=("Arial",18),text_color=console_frame_right_2_color)
        label2.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)
        files_to_keep_set.grid(column =0,row=6,sticky = tk.W,pady =0,padx=10)
        button_save2.grid(column =0,row=6,sticky = tk.W,pady =0,padx=60)
        console_frame_right_2.grid(column =0,row=7,sticky = tk.W,pady =0,padx=10)
        def new_FTK_enter_btn(e):
            set_files_to_keep()
        files_to_keep_set.bind("<Return>",new_FTK_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů
   
    def selected3(self,clear:bool): # Třetí možnost mazání, mazání datumových adresářů
        """
        Nastavení widgets pro třetí možnost mazání

        Budou smazány VŠECHNY adresáře (včetně všech subadresářů), které obsahují v názvu podporovaný formát datumu a jsou vyhodnoceny,jako starší než nastavené datum\n
        -Podporované datumové formáty jsou ["YYYYMMDD","DDMMYYYY","YYMMDD"] a podporované datumové separátory: [".","/","_"]

        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        Tools.clear_console(self.console)
        self.checkbox2.deselect()
        self.checkbox3.select()
        self.checkbox.deselect()
        self.info.configure(text = f"- Budou smazány VŠECHNY adresáře (včetně všech subadresářů), které obsahují v názvu podporovaný formát datumu a jsou vyhodnoceny,\njako starší než nastavené datum\nPodporované datumové formáty: {Deleting.supported_date_formats}\nPodporované separátory datumu: {Deleting.supported_separators}",font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

                        
            self.selected3(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny adresáře vyhodnoceny, jako starší!)","orange"

            self.selected3(False) #refresh

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

        console_frame_right_1_text, console_frame_right_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        images2         = customtkinter.CTkLabel(master = self.frame_right,text = "")
        label1          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení datumu v názvu adresářů, jako staršího:",justify = "left",font=("Arial",16))
        set_day         = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1            = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month       = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2            = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year        = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1 = customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_frame_right_1_text,justify = "left",font=("Arial",18),text_color=console_frame_right_1_color)
        directories     = customtkinter.CTkImage(Image.open(Tools.resource_path("images/directories.png")),size=(240, 190))
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        images2.grid(column =0,row=row_index,sticky = tk.W,pady =15,padx=600,rowspan=5)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=140)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =0,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        images2.configure(image = directories)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů

    def selected6(self): # checkbox s dotazem procházet subsložky
        """
        checkbox s dotazem procházet subsložky
        """
        if self.checkbox6.get() == 1:
            if self.checkbox3.get() == 1:
                self.info2.configure(text = "- Pro tuto možnost třídění není tato funkce podporována",font=("Arial",16,"bold"),text_color="white")
            else:
                self.info2.configure(text = "- VAROVÁNÍ: Máte spuštěné možnosti mazání obrázkových souborů i ve všech subsložkách vložené cesty (max:6 subsložek)",font=("Arial",16,"bold"),text_color="yellow")
        else:
            self.info2.configure(text = "")

    def save_new_task(self):
        def call_browse_directories():
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if os.path.exists(str(operating_path.get())):
                output = Tools.browseDirectories("only_dirs",start_path=str(operating_path.get()))
            else:
                output = Tools.browseDirectories("only_dirs")
            if str(output[1]) != "/":
                operating_path.delete(0,300)
                operating_path.insert(0, str(output[1]))
                Tools.add_colored_line(console,"Byla vložena cesta pro vykonávání úkolu","green",None,True)
            print(output[0])
            window.focus()
            window.focus_force()

        def save_task_to_config():
            if check_entry("",hour_format=True,input_char=str(frequency_entry.get())) == False:
                return
            def get_task_name(current_tasks):
                names_taken = []
                new_task_name = "TRIMAZKON_task_xx"
                for tasks in current_tasks:
                    names_taken.append(tasks[0])
                for i in range(1,100):
                    name_suggestion = "TRIMAZKON_task_" + str(i)
                    if not name_suggestion in names_taken:
                        new_task_name = name_suggestion
                        break
                return new_task_name

            def set_up_task_in_ts():
                def check_freq_format(freq_input):
                    input_splitted = freq_input.split(":")
                    if len(str(input_splitted[0])) == 1:
                        corrected = "0"+str(input_splitted[0]) +":"+ str(input_splitted[1])
                        return corrected
                    else:
                        return freq_input
                        
                name_of_task = new_task[0]
                repaired_freq_param = check_freq_format(str(new_task[4]))
                path_app_location = str(initial_path+"/"+exe_name) 
                # task_command = "/c start \""+ path_app_location+ " deleting " + name_of_task + " " + str(new_task[1]) + " " + str(new_task[2]) + " " + str(new_task[3]) + "\" /SC DAILY /ST " + repaired_freq_param
                task_command = "\""+ path_app_location+ " deleting " + name_of_task + " " + str(new_task[1]) + " " + str(new_task[2]) + " " + str(new_task[3]) + "\" /SC DAILY /ST " + repaired_freq_param
                process = subprocess.Popen(f"schtasks /Create /TN {name_of_task} /TR {task_command}",
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                stdout, stderr = process.communicate()
                try:
                    stdout_str = stdout.decode('utf-8')
                    data = str(stdout_str)
                except UnicodeDecodeError:
                    try:
                        stdout_str = stdout.decode('cp1250')
                        data = str(stdout_str)
                    except UnicodeDecodeError:
                        data = str(stdout)
                output_message = "out"+str(stdout) +"err"+str(stderr)
                print(output_message)
                if "SUCCESS" in stdout_str:
                    os.startfile("taskschd.msc")
                    return True
                else:
                    return False

            current_tasks = trimazkon_tray_instance.read_config()
            wb = load_workbook(initial_path + self.config_filename)
            ws = wb["Task_settings"]
            # subexe_path = Tools.resource_path(trimazkon_tray_exe_name)
            # process_output = subprocess.run(subexe_path + " " + initial_path + " read_config",
            #                 creationflags=subprocess.CREATE_NO_WINDOW,
            #                 stdout=subprocess.PIPE,
            #                 text = True)
            # current_tasks = list(process_output.stdout)
            if len(current_tasks) == 0:
                current_tasks = []
            print("current tasks: ",current_tasks)
            new_task_name = get_task_name(current_tasks)
            new_task = [new_task_name,operating_path.get(),self.older_then_entry.get(),minimum_file_entry.get(),frequency_entry.get(),str(Deleting.get_current_date()[2]),""]
            current_tasks.insert(0,new_task)

            row_to_print = 1
            for tasks in current_tasks:
                ws['A' + str(row_to_print)] = tasks[0] # název tasku
                ws['B' + str(row_to_print)] = tasks[1] # cesta vykonavani
                ws['C' + str(row_to_print)] = tasks[2] # max days
                ws['D' + str(row_to_print)] = tasks[3] # min left
                ws['E' + str(row_to_print)] = tasks[4] # frequency
                ws['F' + str(row_to_print)] = tasks[5] # datum přidání tasku
                ws['G' + str(row_to_print)] = tasks[6] # log mazání (pocet smazanych,datum,seznam smazanych)
                row_to_print +=1
            try:
                success_status = set_up_task_in_ts()
                if success_status:
                    Tools.add_colored_line(console,"Nový úkol byl uložen a zaveden do task scheduleru","green",None,True)
                    wb.save(self.config_filename)
                else:
                    Tools.add_colored_line(console,"Neočekávaná chyba, nepovedlo se nastavit nový úkol","red",None,True)
                wb.close()  

            except Exception as e:
                Tools.add_colored_line(console,f"Prosím zavřete konfigurační soubor ({e})","red",None,True)
                wb.close()

        def refresh_cutoff_date():
            self.older_then_entry.update()
            self.older_then_entry.update_idletasks()
            try:
                cutoffdate_list = Deleting.get_cutoff_date(int(self.older_then_entry.get()))
                new_date = "(starší než: "+str(cutoffdate_list[0])+"."+str(cutoffdate_list[1])+"."+str(cutoffdate_list[2])+")"
                if older_then_label3.cget("text") != new_date:
                    older_then_label3.configure(text = new_date)
            except Exception:
                pass

        def check_entry(event,number=False,hour_format=False,input_char=None,flag=""):
            if flag == "cutoff":
                window.after(100, lambda: refresh_cutoff_date())
            if event != "":
                if event.keysym == "BackSpace" or event.keysym == "Return":
                    return

            if number:
                if not event.char.isdigit():
                    Tools.add_colored_line(console,"Vkládejte pouze čísla","red",None,True)
                    event.widget.insert(tk.INSERT,"")
                    return "break"  # Stop the event from inserting the original character
                
            elif hour_format:
                if not ":" in input_char:
                    Tools.add_colored_line(console,"Neplatný formát času, chybí separátor (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif len(input_char.split(":")) != 2:
                    Tools.add_colored_line(console,"Neplatný formát času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif len(str(input_char.split(":")[1])) != 2:
                    Tools.add_colored_line(console,"Neplatný formát času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif not input_char.split(":")[0].isdigit() or not input_char.split(":")[1].isdigit():
                    Tools.add_colored_line(console,"Neplatné znaky u času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif int(input_char.split(":")[0]) > 23 or int(input_char.split(":")[0]) < 0 or int(input_char.split(":")[1]) > 59 or int(input_char.split(":")[1]) < 0:
                    Tools.add_colored_line(console,"Neplatný formát času, mimo rozsah (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                
        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(app_icon))
        window.title("Nastavení nového úkolu")
        trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path)

        parameter_frame = customtkinter.CTkFrame(master = window,corner_radius=0)
        path_label = customtkinter.CTkLabel(master = parameter_frame,text = "Zadejte cestu, kde bude úkol spouštěn:",font=("Arial",22,"bold"))
        path_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        operating_path = customtkinter.CTkEntry(master = path_frame,font=("Arial",20),height=50,corner_radius=0)
        explorer_btn = customtkinter.CTkButton(master = path_frame,text = "...",font=("Arial",22,"bold"),width = 40,height=50,corner_radius=0,command=lambda: call_browse_directories())
        path_label.pack(pady = (10,0),padx = (10,0),side="top",anchor="w")
        operating_path.pack(pady = (10,0),padx = (10,0),side="left",anchor="w",expand = True,fill="x")
        explorer_btn.pack(pady = (10,0),padx = (0,10),side="left",anchor="w")
        path_frame.pack(side="top",anchor="w",fill="x",expand = True)

        older_then_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        older_then_label = customtkinter.CTkLabel(master = older_then_frame,text = "Odstanit soubory starší než:",font=("Arial",22,"bold"))
        self.older_then_entry = customtkinter.CTkEntry(master = older_then_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        older_then_label2 = customtkinter.CTkLabel(master = older_then_frame,text = "dní",font=("Arial",22,"bold"))
        older_then_label3 = customtkinter.CTkLabel(master = older_then_frame,text = "",font=("Arial",22,"bold"))
        older_then_label.pack(pady = (10,0),padx = (10,10),side="left")
        self.older_then_entry.pack(pady = (10,0),padx = (0,0),side="left")
        older_then_label2.pack(pady = (10,0),padx = (10,0),side="left")
        older_then_label3.pack(pady = (10,0),padx = (10,10),side="left")
        older_then_frame.pack(side="top",fill="x",anchor="w")
        self.older_then_entry.bind("<Key>",lambda e: check_entry(e,number=True,flag="cutoff"))

        minimum_file_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        minimum_file_label = customtkinter.CTkLabel(master = minimum_file_frame,text = "Ponechat souborů:",font=("Arial",22,"bold"))
        minimum_file_entry = customtkinter.CTkEntry(master = minimum_file_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        minimum_file_label.pack(pady = (10,0),padx = (10,10),side="left")
        minimum_file_entry.pack(pady = (10,0),padx = (0,10),side="left")
        minimum_file_frame.pack(side="top",fill="x",anchor="w")
        minimum_file_entry.bind("<Key>",lambda e: check_entry(e,number=True))

        frequency_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        frequency_label = customtkinter.CTkLabel(master = frequency_frame,text = "Frekvence: denně, ",font=("Arial",22,"bold"))
        frequency_entry = customtkinter.CTkEntry(master = frequency_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        frequency_label2 = customtkinter.CTkLabel(master = frequency_frame,text = "hodin (př.: 0:00, 6:00, 14:30)",font=("Arial",22,"bold"))
        frequency_label.pack(pady = (10,0),padx = (10,10),side="left",anchor="w")
        frequency_entry.pack(pady = (10,0),padx = (0,0),side="left",anchor="w")
        frequency_label2.pack(pady = (10,0),padx = (10,10),side="left",anchor="w")
        frequency_frame.pack(side="top",fill="x",anchor="w")
        console = tk.Text(parameter_frame, wrap="none", height=0, width=30,background="black",font=("Arial",22),state=tk.DISABLED)
        console.pack(pady = 10,padx =10,side="top",anchor="w",fill="x")

        button_frame =   customtkinter.CTkFrame(master = window,corner_radius=0)
        show_tasks_btn = customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zobrazit nastavené úkoly", command =  lambda: trimazkon_tray_instance.show_all_tasks(toplevel=True),font=("Arial",20,"bold"),corner_radius=0)
        # show_tasks_btn = customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zobrazit nastavené úkoly", command =  lambda: call_show_all_tasks(),font=("Arial",20,"bold"),corner_radius=0)
        save_task_btn =  customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Uložit nový úkol", command =  lambda: save_task_to_config(),font=("Arial",20,"bold"),corner_radius=0)
        cancel_btn =  customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zavřít", command =  lambda: window.destroy(),font=("Arial",20,"bold"),corner_radius=0)
        cancel_btn.   pack(pady=10,padx=(10,10),side="right",anchor="e")
        save_task_btn.   pack(pady=10,padx=(10,0),side="right",anchor="e")
        show_tasks_btn.  pack(pady=10,padx=(10,0),side="right",anchor="e")
        parameter_frame.pack(side="top",fill="both")
        button_frame.pack(side="top",fill="x")
        operating_path.insert("0",self.path_set.get())
        max_days = Deleting.get_max_days(self.cutoff_date)
        self.older_then_entry.insert("0",max_days)
        minimum_file_entry.insert("0",self.files_to_keep)
        frequency_entry.insert("0","12:00")
        refresh_cutoff_date()
        window.update()
        window.update_idletasks()
        window_width = window.winfo_width()
        if window_width < 1200:
            window_width = 1200
        window.geometry(f"{window_width}x{window.winfo_height()}")
        window.after(100,window.focus_force())
        window.focus()

    def create_deleting_option_widgets(self):  # Vytváří veškeré widgets (delete option MAIN)
        #definice ramcu
        frame_with_logo =       customtkinter.CTkFrame(master=self.root,corner_radius=0)
        logo =                  customtkinter.CTkImage(Image.open(Tools.resource_path("images/logo.png")),size=(1200, 100))
        image_logo =            customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        frame_with_logo.        pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        image_logo.pack()
        frame_with_cards =      customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=100)
        self.frame_path_input = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.bottom_frame2 =    customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        self.bottom_frame1 =    customtkinter.CTkFrame(master=self.root,height = 80,corner_radius=0)
        checkbox_frame =        customtkinter.CTkFrame(master=self.root,width=400,height = 150,corner_radius=0)
        self.frame_right =      customtkinter.CTkFrame(master=self.root,corner_radius=0,height = 150)
        frame_with_cards.       pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        self.frame_path_input.  pack(pady=5,padx=5,fill="both",expand=False,side = "top")
        self.bottom_frame2.     pack(pady=0,padx=5,fill="both",expand=True,side = "bottom")
        self.bottom_frame1.     pack(pady=5,padx=5,fill="x",expand=False,side = "bottom")
        checkbox_frame.         pack(pady=0,padx=5,fill="y",expand=False,side="left")
        self.frame_right.       pack(pady=0,padx=0,fill="both",expand=True,side="right")
        self.frame_with_checkboxes = checkbox_frame
        list_of_frames = [self.frame_path_input,self.bottom_frame1,self.bottom_frame2,self.frame_right,self.frame_with_checkboxes,frame_with_cards,frame_with_logo]

        shift_const = 250
        menu_button =       customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "MENU",                  command =  lambda: self.call_extern_function(list_of_frames,function="menu"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        sorting_button =    customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Třídění souborů",      command =  lambda: self.call_extern_function(list_of_frames,function="sorting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        deleting_button =   customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Mazání souborů",
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        converting_button = customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Konvertování souborů",  command =  lambda: self.call_extern_function(list_of_frames,function="converting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        menu_button.        grid(column = 0,row=0,pady = (10,0),padx =260-shift_const,sticky = tk.W)
        sorting_button.     grid(column = 0,row=0,pady = (10,0),padx =520-shift_const,sticky = tk.W)
        deleting_button.    grid(column = 0,row=0,pady = (10,0),padx =780-shift_const,sticky = tk.W)
        converting_button.  grid(column = 0,row=0,pady = (10,0),padx =1040-shift_const,sticky = tk.W)
        
        # menu_button =           customtkinter.CTkButton(master = self.frame_path_input, width = 180, text = "MENU", command = lambda: self.call_menu(),font=("Arial",20,"bold"))
        self.path_set    =      customtkinter.CTkEntry(master = self.frame_path_input,font=("Arial",18),placeholder_text="Zadejte cestu k souborům z kamery (kde se přímo nacházejí soubory nebo datumové složky)")
        tree        =           customtkinter.CTkButton(master = self.frame_path_input, width = 180,text = "EXPLORER", command = self.call_browseDirectories,font=("Arial",20,"bold"))
        button_save_path =      customtkinter.CTkButton(master = self.frame_path_input,width=50,text = "Uložit cestu", command = lambda: Tools.save_path(self.console,self.path_set.get()),font=("Arial",20,"bold"))
        button_open_setting =   customtkinter.CTkButton(master = self.frame_path_input,width=30,height=30, text = "⚙️", command = lambda: Advanced_option(self.root,windowed=True,spec_location="deleting_option"),font=("Arial",16))
        # menu_button.            pack(pady =12,padx=10,anchor ="w",side = "left")
        self.path_set.          pack(pady = 12,padx =(10,0),anchor ="w",side = "left",fill="both",expand=True)
        tree.                   pack(pady = 12,padx =10,anchor ="w",side = "left")
        button_save_path.       pack(pady = 12,padx =0,anchor ="w",side = "left")
        button_open_setting.    pack(pady = 12,padx =10,anchor = "w",side = "left")

        self.checkbox  =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Mazání souborů starších než: určité datum",command = lambda: self.selected(True))
        self.checkbox2 =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Redukce novějších, mazání souborů starších než: určité datum",command = lambda: self.selected2(True))
        self.checkbox3 =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Mazání adresářů s názvem ve formátu určitého datumu",command = lambda: self.selected3(True))
        self.checkbox.          pack(pady =10,padx=10,anchor ="w")
        self.checkbox2.         pack(pady =10,padx=10,anchor ="w")
        self.checkbox3.         pack(pady =10,padx=10,anchor ="w")

        self.checkbox6 =        customtkinter.CTkCheckBox(master = self.bottom_frame1, text = "Procházet subsložky? (max:6)",command = self.selected6,font=("Arial",16,"bold"))
        self.info2 =            customtkinter.CTkLabel(   master = self.bottom_frame1,text = "",font=("Arial",16,"bold"))
        self.checkbox_testing = customtkinter.CTkCheckBox(master = self.bottom_frame1, text = f"Režim TESTOVÁNÍ (Soubory vyhodnocené ke smazání se pouze přesunou do složky s názvem: \"{self.to_delete_folder_name}\")",font=("Arial",16,"bold"))
        self.checkbox6.         grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        self.info2.             grid(column =0,row=0,sticky = tk.W,pady =5,padx=280)
        self.checkbox_testing.  grid(column =0,row=1,sticky = tk.W,pady =5,padx=10)
        self.info =             customtkinter.CTkLabel(master = self.bottom_frame2,text = "",font=("Arial",16,"bold"))
        execution_btn_frame =   customtkinter.CTkFrame(master=self.bottom_frame2,corner_radius=0)
        button =                customtkinter.CTkButton(master = execution_btn_frame,width = 300,height = 60,text = "SPUSTIT", command = self.start,font=("Arial",20,"bold"))
        create_task_btn =       customtkinter.CTkButton(master = execution_btn_frame,width = 300,height = 60,text = "Nastavit aut. spouštění",command = lambda: self.save_new_task(),font=("Arial",20,"bold"))
        button.                 pack(pady=10,padx=(10,0),side="left",anchor="w")
        create_task_btn.        pack(pady=10,padx=(10,0),side="left",anchor="w")
        self.console =          tk.Text(self.bottom_frame2, wrap="word", height=20, width=1200,background="black",font=("Arial",16),state=tk.DISABLED)
        self.info.              pack(pady = 12,padx =10,anchor="w",side = "top")
        execution_btn_frame.    pack(pady =20,padx=10,side = "top",anchor="n")
        self.console.           pack(pady =10,padx=10,side = "top")
        #default:
        self.checkbox.select()
        self.checkbox_testing.select()
        self.selected(False)

        if global_recources_load_error:
            create_task_btn.configure(state = "disabled")

        recources_path = self.text_file_data[2]
        if recources_path != False and recources_path != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", str(recources_path))
            Tools.add_colored_line(self.console,"Byla vložena cesta z konfiguračního souboru","white")
        else:
            Tools.add_colored_line(self.console,"Konfigurační soubor obsahuje neplatnou cestu k souborům (můžete vložit v pokročilém nastavení)","orange")
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)
        self.unbind_list.append("<f>")

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")
        self.path_set.bind("<Return>",unfocus_widget)

class Sorting_option: # Umožňuje nastavit možnosti třídění souborů
    """
    Umožňuje nastavit možnosti třídění souborů

    -možnosti s ukázkou očekávané syntaxe názvu souboru i s visualizací\n
    -umožňuje operace s ID daného obrázku\n
    -umožňuje hledání chybějících ID v řadě za sebou (na lince několik palet s výrobkem)
    """
    def __init__(self,root):
        self.root = root
        self.aut_detect_num_of_pallets = True
        self.by_which_ID_num = ""   
        self.more_dirs = False
        self.unbind_list = []
        self.text_file_data = Tools.read_config_data()
        self.supported_formats_sorting = self.text_file_data[0]
        self.prefix_func = self.text_file_data[5]
        self.prefix_Cam = self.text_file_data[6]
        self.max_num_of_pallets = self.text_file_data[8]
        self.safe_mode = self.text_file_data[10]
        list_of_folder_names = self.text_file_data[9]
        self.nok_folder_name = list_of_folder_names[0]
        self.pairs_folder_name = list_of_folder_names[1]
        self.sort_inside_pair_folder = True
        self.temp_path_for_explorer = None
        self.original_image = Image.open(Tools.resource_path("images/loading3.png"))
        self.original_image = self.original_image.resize((300, 300))
        self.angle = 0

        self.create_sorting_option_widgets()

    def start(self):# Ověřování cesty, init, spuštění
        """
        Ověřování cesty, init, spuštění
        """
        Tools.clear_console(self.console)
        if self.checkbox.get()+self.checkbox2.get()+self.checkbox3.get()+self.checkbox4.get()+self.checkbox5.get() == 0:
            Tools.add_colored_line(self.console,"Nevybrali jste žádný způsob třídění :-)","red")
            nothing = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nothing.png")),size=(1, 1))
            self.images.configure(image = nothing)
            self.name_example.configure(text = "")

        else:
            path = self.path_set.get() 
            if path != "":
                check = Tools.path_check(path)
                if check == False:
                    Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" nebyla nalezena","red")
                else:
                    path = check
                    Tools.add_colored_line(self.console,"- Provádím nastavenou možnost třídění v cestě: "+str(path)+"\n","orange")

                    self.console.update_idletasks()
                    self.root.update_idletasks()
                    self.sort_files(path)
            else:
                Tools.add_colored_line(self.console,"Nebyla vložena cesta k souborům","red")

    def sort_files(self,path): # Volání externího scriptu
        selected_sort = 0
        self.loading_bar.set(value = 0)
        ignore_pairs = False

        only_one_subfolder = False
        if self.checkbox.get() == 1:
            selected_sort = 1
        if self.checkbox2.get() == 1:
            selected_sort = 2
        if self.checkbox3.get() == 1:
            selected_sort = 3
        if self.checkbox4.get() == 1:
            selected_sort = 4
        if self.checkbox5.get() == 1:
            selected_sort = 5
        if self.checkbox6.get() == 1:
            self.more_dirs = True
        else:
            self.more_dirs = False
            if self.one_subfolder.get() == 1:
                self.more_dirs = True
                only_one_subfolder = True
        try:
            if self.checkbox_ignore_pairs.get() == 1 and selected_sort == 2:
                ignore_pairs = True
        except Exception: # the checkbox was not even loaded once...
            pass

        if self.checkbox_safe_mode.get() == 1:
            self.safe_mode = "ne"
        else:
            self.safe_mode = "ano"

        popup = tk.Toplevel(master=root)
        popup.attributes('-topmost', True)
        geometry_string = "1000x1000+" + str(int(root.winfo_screenwidth()/2)-500)+ "+" + str(int(root.winfo_screenheight()/2)-500)
        popup.geometry(str(geometry_string))
        popup.label = tk.Label(popup)
        image_ = ImageTk.PhotoImage(self.original_image)
        popup.label.config(image=image_)  # Update label's image
        popup.label.image = image_  # Keep a reference to the image to prevent garbage collection
        popup.label.place(x=0, y=0, relwidth=1, relheight=1)
        popup.wm_attributes("-transparentcolor","white")# trick to force the bg transparent
        popup.config(bg= 'white')
        popup.label.config(bg= 'white')
        popup.overrideredirect(True)# hide the frame of a window
        popup.update()
        popup.label.update()

        def rotate_image():
            self.angle += 10  # Adjust the rotation speed as needed
            rotated_image = self.original_image.rotate(self.angle)
            image_ = ImageTk.PhotoImage(rotated_image)
            popup.label.config(image=image_)  # Update label's image
            popup.label.image = image_
            popup.update()

        def call_trideni_main(whole_instance):
            whole_instance.main()

        running_program = Trideni.whole_sorting_function(
            path,
            selected_sort,
            self.more_dirs,
            self.max_num_of_pallets,
            self.by_which_ID_num,
            self.prefix_func,
            self.prefix_Cam,
            self.supported_formats_sorting,
            self.aut_detect_num_of_pallets,
            self.nok_folder_name,
            self.pairs_folder_name,
            self.safe_mode,
            self.sort_inside_pair_folder,
            only_one_subfolder,
            ignore_pairs
        )

        run_background = threading.Thread(target=call_trideni_main, args=(running_program,))
        run_background.start()

        output_list_increment = 0
        completed = False
        output_text2 = ""
        previous_console2_text = []
        previous_progres = 0

        while not running_program.finish or completed == False:
            time.sleep(0.05)
            rotate_image()
            #progress bar:
            if running_program.progress != previous_progres:
                self.loading_bar.set(value = running_program.progress/100)
                self.root.update_idletasks()
                root.update_idletasks()
                self.frame5.update_idletasks()

            if len(running_program.output_list) > output_list_increment:
                for i in range(0,len(running_program.output_list[output_list_increment])):
                    new_row = str(running_program.output_list[output_list_increment][i])
                    if "bylo dokončeno" in new_row or "byla dokončena" in new_row:
                        Tools.add_colored_line(self.console,str(new_row),"green",("Arial",16,"bold"))
                    elif "Chyba" in new_row or "Třídění ukončeno" in new_row or "Celkový počet duplikátů" in new_row:
                        Tools.add_colored_line(self.console,str(new_row),"red",("Arial",16,"bold"))
                    elif "Nepáry" in new_row:
                        Tools.add_colored_line(self.console,str(new_row),"orange",("Arial",16,"bold"))
                    elif "OK soubory" in new_row:
                        Tools.add_colored_line(self.console,str(new_row),"green",("Arial",16,"bold"))
                    else:
                        Tools.add_colored_line(self.console,str(new_row),"white")
                    self.console.update_idletasks()
                self.root.update_idletasks()
                output_list_increment+=1

            if running_program.output_console2 != previous_console2_text and len(running_program.output_console2) != 0:
                for i in range(0,len(running_program.output_console2)):
                    output_text2 = output_text2 + running_program.output_console2[i] + "\n"
                if output_text2 != "":
                    if "Chyba" in output_text2:
                        self.console2.configure(text = output_text2,text_color="red")
                    else:
                        self.console2.configure(text = output_text2,text_color="green")
                self.console2.update_idletasks()
                previous_console2_text = running_program.output_console2
                output_text2 = ""

            if running_program.finish and len(running_program.output_list) == output_list_increment and running_program.output_console2 == previous_console2_text and int(self.loading_bar.get())== 1:
                completed = True
        
        self.console.update_idletasks()
        run_background.join()
        popup.destroy()

    def clear_frame(self,frame): # mazání widgets v daném framu
        for widget in frame.winfo_children():
            widget.destroy()

    def selected(self): #Třídit podle typu souboru
        """
        Nastavení widgets pro třídění podle typu souboru (základní)
        """
        self.clear_frame(self.frame6)
        self.view_image(1)
        #self.console.configure(text = "")
        Tools.clear_console(self.console)
        self.checkbox2.deselect()
        self.checkbox3.deselect()
        self.checkbox4.deselect()
        self.checkbox5.deselect()

        label_fill = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=self.height_of_frame6+15,text = "",justify = "left",font=("Arial",12))
        label_fill.grid(column =0,row=0,pady =0,padx=10)
        
    def selected2(self): #Třídit podle čísla funkce (ID)
        """
        Nastavení widgets pro třídění podle čísla funkce
        """
        self.clear_frame(self.frame6)
        self.view_image(2)
        Tools.clear_console(self.console)
        self.checkbox.deselect()
        self.checkbox3.deselect()
        self.checkbox4.deselect()
        self.checkbox5.deselect()

        def set_prefix():
            input_1 = str(prefix_set.get()).replace(" ","")
            if len(input_1) != 0:
                if input_1 != self.prefix_Cam:
                    console_frame6_1.configure(text = f"Prefix nastaven na: {input_1}",text_color="green")
                    self.prefix_func = input_1
                    prefix_set.delete("0","100")
                    prefix_set.insert("0", input_1)
                else:
                    console_frame6_1.configure(text = "Jméno zabrané pro třídění podle kamer",text_color="red")
            else:
                console_frame6_1.configure(text = "Nutný alespoň jeden znak",text_color="red")

        label1          = customtkinter.CTkLabel(master = self.frame6,text = "Nastavte prefix adresářů:",justify = "left",font=("Arial",16))
        prefix_set      = customtkinter.CTkEntry(master = self.frame6,width=150,font=("Arial",16), placeholder_text= self.prefix_func)
        button_save1    = customtkinter.CTkButton(master = self.frame6,width=50, text = "Uložit", command = lambda: set_prefix(),font=("Arial",18,"bold"))
        console_frame6_1 = customtkinter.CTkLabel(master = self.frame6,text = " ",justify = "left",font=("Arial",18))
        label1.grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        prefix_set.grid(column =0,row=1,sticky = tk.W,pady =0,padx=10)
        button_save1.grid(column =0,row=1,sticky = tk.W,pady =0,padx=160)
        console_frame6_1.grid(column =0,row=2,sticky = tk.W,pady =0,padx=10)
        prefix_set.insert("0", str(self.prefix_func))
        def prefix_enter_btn(e):
            set_prefix()
        prefix_set.bind("<Return>",prefix_enter_btn)
        checkbox_advance = customtkinter.CTkCheckBox(master = self.frame6,font=("Arial",16), text = "Pokročilá nastavení",command = self.selected2_advance)
        label_fill         = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=167,text = "",justify = "left",font=("Arial",16))
        checkbox_advance.grid(column =0,row=4,sticky = tk.W,pady =10,padx=10)
        label_fill.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)

    def selected2_advance(self): # Třídit podle určitého čísla v ID
        """
        Nastavení widgets pro třídění podle určitého čísla v ID
        """
        self.clear_frame(self.frame6)

        def set_which_num_of_ID():
            input3=num_set.get()
            if input3.isdigit():
                if int(input3) > 0:
                    self.by_which_ID_num = int(input3)
                    console_frame6_1.configure(text = f"Řídit podle {self.by_which_ID_num}. čísla v ID",text_color="white")
                else:
                    console_frame6_1.configure(text = "Mimo rozsah",text_color="red")
                    self.by_which_ID_num = ""
            else:
                console_frame6_1.configure(text = "Nezadali jste číslo",text_color="red")
                self.by_which_ID_num = ""

        label1           = customtkinter.CTkLabel(master = self.frame6,height=60,
                                        text = "Podle kterého čísla v ID se řídit:\nvolte první = 1 atd.\nprázdné = celé ID (aut. detekce)",
                                        justify = "left",font=("Arial",16))
        num_set          = customtkinter.CTkEntry(master = self.frame6,height=30,width=150,font=("Arial",16), placeholder_text= self.by_which_ID_num)
        button_save1     = customtkinter.CTkButton(master = self.frame6,height=30,width=50, text = "Uložit", command = lambda: set_which_num_of_ID(),font=("Arial",18,"bold"))
        console_frame6_1 = customtkinter.CTkLabel(master = self.frame6,height=30,text = " ",justify = "left",font=("Arial",18))
        self.checkbox_ignore_pairs = customtkinter.CTkCheckBox(master = self.frame6,height=30,font=("Arial",16), text = "Ignorovat páry (Třídit pouze podle id)")
        button_back      = customtkinter.CTkButton(master = self.frame6,width=100,height=30, text = "Zpět", command = self.selected2,font=("Arial",18,"bold"))
        label_fill          = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=115,text = "",justify = "left",font=("Arial",16))
        label1.grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        num_set.grid(column =0,row=1,sticky = tk.W,pady =0,padx=10)
        button_save1.grid(column =0,row=1,sticky = tk.W,pady =0,padx=160)
        console_frame6_1.grid(column =0,row=2,sticky = tk.W,pady =5,padx=10)  
        self.checkbox_ignore_pairs.grid(column =0,row=3,sticky = tk.W,pady =(0,10),padx=10)
        button_back.grid(column =0,row=5,sticky = tk.W,pady =10,padx=10)
        label_fill.grid(column =0,row=6,sticky = tk.W,pady =0,padx=10)
        def which_id_num_enter_btn(e):
            set_which_num_of_ID()
        num_set.bind("<Return>",which_id_num_enter_btn)
        
    def selected3(self): #Třídit podle čísla kamery
        """
        Nastavení widgets pro třídění podle čísla kamery
        """
        self.clear_frame(self.frame6)
        Tools.clear_console(self.console)
        self.view_image(3)   
        self.checkbox.deselect()
        self.checkbox2.deselect()
        self.checkbox4.deselect()
        self.checkbox5.deselect()

        def set_prefix():
            input_1 = str(prefix_set.get()).replace(" ","")
            if len(input_1) != 0:
                if input_1 != self.prefix_func:
                    console_frame6_1.configure(text = f"Prefix nastaven na: {input_1}",text_color="green")
                    self.prefix_Cam = input_1
                    prefix_set.delete("0","100")
                    prefix_set.insert("0", input_1)
                else:
                    console_frame6_1.configure(text = "Jméno zabrané pro třídění podle funkce",text_color="red")
            else:
                console_frame6_1.configure(text = "Nutný alespoň jeden znak",text_color="red")

        label1       = customtkinter.CTkLabel(master = self.frame6,height=20,text = "Nastavte prefix adresářů:",justify = "left",font=("Arial",16))
        prefix_set   = customtkinter.CTkEntry(master = self.frame6,height=30,width=150,font=("Arial",16), placeholder_text= self.prefix_Cam)
        button_save1 = customtkinter.CTkButton(master = self.frame6,height=30,width=50, text = "Uložit", command = lambda: set_prefix(),font=("Arial",18,"bold"))
        console_frame6_1 = customtkinter.CTkLabel(master = self.frame6,height=30,text = " ",justify = "left",font=("Arial",18))
        label_fill       = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=205,text = "",justify = "left",font=("Arial",16))
        label1.grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        prefix_set.grid(column =0,row=1,sticky = tk.W,pady =0,padx=10)
        button_save1.grid(column =0,row=1,sticky = tk.W,pady =0,padx=160)
        console_frame6_1.grid(column =0,row=2,sticky = tk.W,pady =5,padx=10)
        label_fill.grid(column =0,row=3,pady =0,sticky = tk.W,padx=10)
        prefix_set.insert("0", str(self.prefix_Cam))
        def prefix_enter_btn(e):
            set_prefix()
        prefix_set.bind("<Return>",prefix_enter_btn)

    def selected4(self): #Třídit podle obojího (funkce, kamery)
        """
        Nastavení widgets pro třídění podle funkce i čísla kamery
        """
        self.clear_frame(self.frame6)
        #self.console.configure(text = "")
        Tools.clear_console(self.console)
        self.view_image(4)
        self.checkbox.deselect()
        self.checkbox2.deselect()
        self.checkbox3.deselect()
        self.checkbox5.deselect()

        label_fill = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=self.height_of_frame6+15,text = "",justify = "left",font=("Arial",12))
        label_fill.grid(column =0,row=0,pady =0,padx=10)

    def selected5(self): #hledani paru
        """
        Nastavení widgets pro hledání dvakrát vyfocených výrobků za sebou se stejným ID

        - nalezené dvojice nakopíruje do složky
        """
        self.clear_frame(self.frame6)
        Tools.clear_console(self.console)
        self.view_image(5)
        self.checkbox.deselect()
        self.checkbox2.deselect()
        self.checkbox3.deselect()
        self.checkbox4.deselect()
        
        def set_max_pallet_num():
            input_1 = pallets_set.get()
            if input_1.isdigit() == False:
                console_frame6_1.configure(text = "Nezadali jste číslo",text_color="red")
            elif int(input_1) <1:
                console_frame6_1.configure(text = "Mimo rozsah",text_color="red")
            else:
                console_frame6_1.configure(text = f"Počet palet nastaven na: {input_1}",text_color="green")
                self.max_num_of_pallets = input_1
                
        def set_aut_detect():
            if checkbox_aut_detect.get() == 1:
                self.aut_detect_num_of_pallets = True
            else:
                self.aut_detect_num_of_pallets = False

        def set_sorting_pair_folder():
            if sort_pair_folder.get() == 1:
                self.sort_inside_pair_folder = True
            else:
                self.sort_inside_pair_folder = False

        label1              = customtkinter.CTkLabel(master = self.frame6,height=20,text = "Nastavte počet palet v oběhu:",justify = "left",font=("Arial",16))
        pallets_set         = customtkinter.CTkEntry(master = self.frame6,width=150,height=30, placeholder_text= self.max_num_of_pallets,font=("Arial",16))
        button_save1        = customtkinter.CTkButton(master = self.frame6,width=50,height=30, text = "Uložit", command = lambda: set_max_pallet_num(),font=("Arial",18,"bold"))
        label_aut_detect    = customtkinter.CTkLabel(master = self.frame6,height=60,text = "Možnost aut. detekce:\n(případ, že v cestě nechybí nejvyšší ID)",justify = "left",font=("Arial",16))
        checkbox_aut_detect = customtkinter.CTkCheckBox(master = self.frame6,height=30,font=("Arial",16), text = "Automatická detekce",command=set_aut_detect)
        sort_pair_folder    = customtkinter.CTkCheckBox(master = self.frame6,height=90,font=("Arial",16), text = "Třídit uvnitř složky s páry\npodle typu souboru",command = lambda: set_sorting_pair_folder())
        console_frame6_1    = customtkinter.CTkLabel(master = self.frame6,height=30,text = " ",justify = "left",font=("Arial",18))
        label_fill              = customtkinter.CTkLabel(master = self.frame6,width=self.width_of_frame6,height=0,text = "",justify = "left",font=("Arial",12))
        label1.grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        pallets_set.grid(column =0,row=1,sticky = tk.W,pady =0,padx=10)
        button_save1.grid(column =0,row=1,sticky = tk.W,pady =0,padx=160)
        console_frame6_1.grid(column =0,row=2,sticky = tk.W,pady =5,padx=10)
        label_aut_detect.grid(column =0,row=3,sticky = tk.W,pady =5,padx=10)
        checkbox_aut_detect.grid(column =0,row=4,sticky = tk.W,pady =0,padx=10)
        sort_pair_folder.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)
        label_fill.grid(column =0,row=6,pady =0,padx=10)
        checkbox_aut_detect.select()
        sort_pair_folder.select()
        self.sort_inside_pair_folder = True

        def max_pallets_num_enter_btn(e):
            set_max_pallet_num()
        pallets_set.bind("<Return>",max_pallets_num_enter_btn)

    def one_subfolder_checked(self): # checkbox na přepínání: procházet/ neprocházet 1 subsložku
        self.checkbox6.deselect()
        if self.one_subfolder.get() == 1:
            if self.checkbox_safe_mode.get()==1:
                dir1sub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/1sub_roz.png")),size=(522, 173))
                self.images2.configure(image =dir1sub)
                self.console2.configure(text = "Zadaná cesta/ 1.složka/ složky se soubory",text_color="white")
            else:
                nodir1sub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/1sub_vol.png")),size=(513, 142))
                self.images2.configure(image =nodir1sub)
                self.console2.configure(text = "Zadaná cesta/ 1.složka/ soubory volně, neroztříděné",text_color="white")
        else:
            if self.checkbox_safe_mode.get()==1:
                dirsnosub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nosub_roz.png")),size=(432, 133))
                self.images2.configure(image =dirsnosub)
                self.console2.configure(text = "Zadaná cesta/ složky se soubory",text_color="white")
            else:
                nodirsnosub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nosub_vol.png")),size=(253, 142))
                self.images2.configure(image =nodirsnosub)
                self.console2.configure(text = "Zadaná cesta/ soubory volně, neroztříděné",text_color="white")
    
    def two_subfolders_checked(self): # checkbox na přepínání: procházet/ neprocházet 2 subsložky
        self.one_subfolder.deselect()
        if self.checkbox6.get() == 1:
            if self.checkbox_safe_mode.get()==1:
                dir2sub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/2sub_roz.png")),size=(553, 111))
                self.images2.configure(image =dir2sub)
                self.console2.configure(text = "Zadaná cesta/ 1.složka/ 2.složka/ složky se soubory",text_color="white")
            else:
                nodir2sub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/2sub_vol.png")),size=(553, 111))
                self.images2.configure(image =nodir2sub)
                self.console2.configure(text = "Zadaná cesta/ 1.složka/ 2.složka/ soubory volně, neroztříděné",text_color="white")
        else:
            if self.checkbox_safe_mode.get()==1:
                dirsnosub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nosub_roz.png")),size=(432, 133))
                self.images2.configure(image =dirsnosub)
                self.console2.configure(text = "Zadaná cesta/ složky se soubory",text_color="white")
            else:
                nodirsnosub = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nosub_vol.png")),size=(253, 142))
                self.images2.configure(image =nodirsnosub)
                self.console2.configure(text = "Zadaná cesta/ soubory volně, neroztříděné",text_color="white")
    
    def safe_mode_checked(self):
        if self.one_subfolder.get() == 1:
            self.one_subfolder_checked()
        else:
            self.two_subfolders_checked()
    
    def view_image(self,which_one): # zobrazení ilustračního obrázku
        """
        zobrazení ilustračního obrázku
        """
        if self.checkbox.get()+self.checkbox2.get()+self.checkbox3.get()+self.checkbox4.get()+self.checkbox5.get() == 0:
            nothing = customtkinter.CTkImage(Image.open(Tools.resource_path("images/nothing.png")),size=(1, 1))
            self.images.configure(image = nothing)
            self.name_example.configure(text = "")
        else:
            if which_one == 1:
                type_24 = customtkinter.CTkImage(Image.open(Tools.resource_path("images/24_type.png")),size=(224, 85))
                self.images.configure(image =type_24)
                self.name_example.configure(text = f"221013_092241_0000000842_21_&Cam1Img  => .Height <=  .bmp\n(Podporované formáty:{self.supported_formats_sorting})")
            if which_one == 2:
                func_24 = customtkinter.CTkImage(Image.open(Tools.resource_path("images/24_func.png")),size=(363, 85))
                self.images.configure(image =func_24)
                self.name_example.configure(text = f"221013_092241_0000000842_  => 21 <=  _&Cam1Img.Height.bmp\n(Podporované formáty:{self.supported_formats_sorting})")
            if which_one == 3:
                cam_24 = customtkinter.CTkImage(Image.open(Tools.resource_path("images/24_cam.png")),size=(437, 85))
                self.images.configure(image =cam_24)
                self.name_example.configure(text = f"221013_092241_0000000842_21_&  => Cam1 <=  Img.Height.bmp\n(Podporované formáty:{self.supported_formats_sorting})")
            if which_one == 4:
                both_24 = customtkinter.CTkImage(Image.open(Tools.resource_path("images/24_both.png")),size=(900, 170))
                self.images.configure(image =both_24)
                self.name_example.configure(text = f"221013_092241_0000000842_  => 21 <=  _&  => Cam1 <=  Img.Height.bmp\n(Podporované formáty:{self.supported_formats_sorting})")
            if which_one == 5:
                PAIRS = customtkinter.CTkImage(Image.open(Tools.resource_path("images/25basic.png")),size=(265, 85))
                self.images.configure(image =PAIRS)
                self.name_example.configure(
                    text = f"Nakopíruje nalezené dvojice souborů do složky s názvem PAIRS\n(např. obsluha vloží dvakrát stejnou paletu po sobě před kameru)\n2023_04_13-07_11_09_xxxx_=> 0020 <=_&Cam2Img.Height.bmp\n(funkce postupuje podle časové známky v názvu souboru, kdy byly soubory pořízeny)\n(Podporované formáty:{self.supported_formats_sorting})")
    
    def call_extern_function(self,list_of_frames,function:str): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu\n
        function:
        - menu
        - (sorting)
        - deleting
        - converting
        """
        for frames in list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()
        
        for binds in self.unbind_list:
            self.root.unbind(binds)

        if function == "menu":
            menu.menu()
        elif function == "deleting":
            Deleting_option(self.root)
        elif function == "converting":
            Converting_option(self.root)

    def call_browseDirectories(self): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        if self.checkbox6.get() == 1 or self.one_subfolder.get() == 1: # pokud je zvoleno more_dirs v exploreru pouze slozky...
            output = Tools.browseDirectories("only_dirs",self.temp_path_for_explorer)
        else:
            output = Tools.browseDirectories("all",self.temp_path_for_explorer)

        if str(output[1]) != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", output[1])
            Tools.add_colored_line(self.console,f"Byla vložena cesta: {output[1]}","green")
            self.temp_path_for_explorer = output[1]
        else:
            Tools.add_colored_line(self.console,str(output[0]),"red")

    def create_sorting_option_widgets(self):  # Vytváří veškeré widgets (sorting option MAIN)
        frame_with_logo =       customtkinter.CTkFrame(master=self.root,corner_radius=0)
        logo =                  customtkinter.CTkImage(Image.open(Tools.resource_path("images/logo.png")),size=(1200, 100))
        image_logo =            customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        frame_with_logo.        pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        image_logo.pack()
        frame_with_cards = customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=100)
        frame2 =        customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.frame5 =   customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        self.frame3 =   customtkinter.CTkFrame(master=self.root,corner_radius=0,width=400,height = 290)
        self.frame4 =   customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        frame_with_cards.pack(pady=0,padx=0,fill="x",expand=False,side = "top")
        frame2.         pack(pady=5,padx=5,fill="both",expand=False,side = "top")
        self.frame5.    pack(pady=0,padx=5,fill="both",expand=True,side = "bottom")
        self.frame3.    pack(pady=5,padx=5,fill="both",expand=False,side="left")
        self.frame4.    pack(pady=5,padx=5,fill="both",expand=True,side="right")

        self.height_of_frame6 = 290
        self.width_of_frame6 = 370
        self.frame6 =   customtkinter.CTkFrame(master=self.root,corner_radius=0,width=self.width_of_frame6 ,height=self.height_of_frame6)
        self.frame6.    pack(pady=5,padx=0,fill="both",expand=False,side = "bottom")
        list_of_frames = [frame2,self.frame3,self.frame4,self.frame5,self.frame6,frame_with_cards,frame_with_logo]
        shift_const = 250
        menu_button =       customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "MENU",                  command =  lambda: self.call_extern_function(list_of_frames,function="menu"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        sorting_button =    customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Třídění souborů",
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        deleting_button =   customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Mazání souborů",        command =  lambda: self.call_extern_function(list_of_frames,function="deleting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        converting_button = customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Konvertování souborů",  command =  lambda: self.call_extern_function(list_of_frames,function="converting"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        menu_button.        grid(column = 0,row=0,pady = (10,0),padx =260-shift_const,sticky = tk.W)
        sorting_button.     grid(column = 0,row=0,pady = (10,0),padx =520-shift_const,sticky = tk.W)
        deleting_button.    grid(column = 0,row=0,pady = (10,0),padx =780-shift_const,sticky = tk.W)
        converting_button.  grid(column = 0,row=0,pady = (10,0),padx =1040-shift_const,sticky = tk.W)

        self.path_set = customtkinter.CTkEntry(master = frame2,font=("Arial",18),placeholder_text="Zadejte cestu k souborům z kamery (kde se nacházejí složky se soubory nebo soubory přímo)")
        tree =          customtkinter.CTkButton(master = frame2, width = 180,text = "EXPLORER", command = self.call_browseDirectories,font=("Arial",20,"bold"))
        button_save_path = customtkinter.CTkButton(master = frame2,width=50,text = "Uložit cestu", command = lambda: Tools.save_path(self.console,self.path_set.get()),font=("Arial",20,"bold"))
        button_open_setting = customtkinter.CTkButton(master = frame2,width=30,height=30, text = "⚙️", command = lambda: Advanced_option(self.root,windowed=True,spec_location="sorting_option"),font=("Arial",16))
        self.path_set.  pack(pady = 12,padx =(10,0),anchor ="w",side="left",fill="both",expand=True)
        tree.           pack(pady = 12,padx =10,anchor ="w",side="left")
        button_save_path.pack(pady = 12,padx =0,anchor ="w",side="left")
        button_open_setting.pack(pady = 12,padx =10,anchor ="w",side="left")

        self.checkbox =  customtkinter.CTkCheckBox(master = self.frame3,font=("Arial",16), text = "Třídit podle typů souborů",command = self.selected)
        self.checkbox2 = customtkinter.CTkCheckBox(master = self.frame3,font=("Arial",16), text = "Třídit podle čísla funkce (ID)",command = self.selected2)
        self.checkbox3 = customtkinter.CTkCheckBox(master = self.frame3,font=("Arial",16), text = "Třídit podle čísla kamery",command = self.selected3)
        self.checkbox4 = customtkinter.CTkCheckBox(master = self.frame3,font=("Arial",16), text = "Třídit podle čísla funkce i kamery",command = self.selected4)
        self.checkbox5 = customtkinter.CTkCheckBox(master = self.frame3,font=("Arial",16), text = "Najít dvojice (soubory se stejným ID, v řadě za sebou)",command = self.selected5)
        self.checkbox.  pack(pady =12,padx=10,anchor ="w")
        self.checkbox2. pack(pady =12,padx=10,anchor ="w")
        self.checkbox3. pack(pady =12,padx=10,anchor ="w")
        self.checkbox4. pack(pady =12,padx=10,anchor ="w")
        self.checkbox5. pack(pady =12,padx=10,anchor ="w")
        self.one_subfolder = customtkinter.CTkCheckBox(master = self.frame4,font=("Arial",16), text = "Projít 1 subsložku?",command = self.one_subfolder_checked)
        self.checkbox6   = customtkinter.CTkCheckBox(master = self.frame4,font=("Arial",16), text = "Projít 2 subsložky?",command = self.two_subfolders_checked)
        self.checkbox_safe_mode = customtkinter.CTkCheckBox(master = self.frame4,font=("Arial",16), text = "Rozbalit poslední složky?",command = self.safe_mode_checked)
        self.images2     = customtkinter.CTkLabel(master = self.frame4,text = "")
        self.console2    = customtkinter.CTkLabel(master = self.frame4,text = " ",font=("Arial",18,"bold"))
        self.console2.  pack(pady =5,padx=10,side=tk.BOTTOM)
        self.images2.   pack(side=tk.BOTTOM)
        self.one_subfolder.pack(pady =10,padx=10,anchor="w",side=tk.LEFT)
        self.checkbox6. pack(pady =10,padx=10,anchor="w",side=tk.LEFT)
        self.checkbox_safe_mode.pack(pady =10,padx=10,anchor="w",side=tk.LEFT)
        self.checkbox_safe_mode.select()
        self.images =       customtkinter.CTkLabel(master = self.frame5,text = "")
        self.name_example = customtkinter.CTkLabel(master = self.frame5,text = "",font=("Arial",18,"bold"))
        button =            customtkinter.CTkButton(master = self.frame5, text = "SPUSTIT", command = self.start,font=("Arial",20,"bold"))
        self.loading_bar =  customtkinter.CTkProgressBar(master = self.frame5, mode='determinate',width = 800,height =20,progress_color="green",corner_radius=0)
        self.console =      tk.Text(self.frame5, wrap="word", height=20, width=1200,background="black",font=("Arial",16),state=tk.DISABLED)
        self.images.        pack()
        self.name_example.  pack(pady = 12,padx =10)
        button.             pack(pady =12,padx=10)
        button.             _set_dimensions(300,60)
        self.loading_bar.   pack(pady = 5,padx = 5)
        self.loading_bar.   set(value = 0)
        self.console.       pack(pady =10,padx=10)

        #default nastaveni:
        self.checkbox.select()
        self.selected()
        self.view_image(1)
        self.two_subfolders_checked()
        #predvyplneni cesty pokud je platna v configu
        recources_path = self.text_file_data[2]
        if recources_path != False and recources_path != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", str(recources_path))
            Tools.add_colored_line(self.console,"Byla vložena cesta z konfiguračního souboru","white")
        else:
            Tools.add_colored_line(self.console,"Konfigurační soubor obsahuje neplatnou cestu k souborům (můžete vložit v pokročilém nastavení)","orange")

        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")
        self.path_set.bind("<Return>",unfocus_widget)

class IP_manager: # Umožňuje nastavit možnosti třídění souborů
    """
    Umožňuje měnit statickou IPv4 adresu a spravovat síťové disky

    - pracuje s excelovým souborem, kam ukládá data o projektech a o nastavení\n
    - umožňuje projekty doplňovat poznámkami\n
    - umožňuje odpojit síťový disk\n
    - umožňuje namountit síťový disk a trvale jej přidat do windows exploreru\n
    - poskytuje informaci o aktuální statické ip adrese u daného interfacu\n
    - poskytuje informaci o současně připojených síťových discích\n
    - poskytuje informaci o namountěných offline síťových discích\n
    - vše je ošetřeno timeoutem\n
    """
    def __init__(self,root):
        self.root = root
        self.create_IP_manager_widgets()
    
    def callback(self):
        menu.menu()

    def create_IP_manager_widgets(self):
        if root.wm_state() == "zoomed":
            current_window_size = "max"
        else:
            current_window_size = "min"

        app_data = Tools.read_config_data()
        zoom_factor = app_data[21] 

        # IP_setting.IP_assignment(self.root,self.callback,current_window_size,initial_path,zoom_factor)
        IP_setting.main(self.root,self.callback,current_window_size,initial_path,zoom_factor)

class Catalogue_maker: # Umožňuje nastavit možnosti třídění souborů
    """
    Umožňuje sestavit katalog produktů k projektu
    - ten následné vyexportovat do excelu (.xlm/ .xls)
    - rozpracovaný projekt je možné uložit do souboru .xml
    - databázi produktů stahuje automaticky při spuštění z sharepointu (po restartu celé app)
    """
    def __init__(self,root):
        self.root = root
        self.database_downloaded = menu.database_downloaded
        # automatic download bypass:
        if testing:
            self.database_downloaded = True 
        text_file_data = Tools.read_config_data()
        self.database_filename = text_file_data[15]
        self.default_excel_filename = text_file_data[16]
        self.default_xml_file_name = text_file_data[17]
        self.default_subwindow_status = text_file_data[18]
        self.default_export_extension = text_file_data[19]
        self.default_path = text_file_data[20]
        self.default_render_mode = text_file_data[23]
        self.create_catalogue_maker_widgets()

    def callback(self,data_to_save):
        print("received data: ",data_to_save)
        render_mode = data_to_save[6]
        Tools.save_to_config(render_mode,"render_mode")
        data_to_save.pop(6)
        Tools.save_to_config(data_to_save,"catalogue_data")
        menu.menu()

    def create_catalogue_maker_widgets(self):
        if root.wm_state() == "zoomed":
            current_window_size = "max"
        else:
            current_window_size = "min"
        
        if not self.database_downloaded:
            download = download_database.database(self.database_filename)
            input_message = str(download.output)
            menu.database_downloaded = True
        else:
            input_message = "Datábáze se stáhne znovu až po restartu TRIMAZKONU"
        
        Catalogue.Catalogue_gui(self.root,
                                input_message,
                                self.callback,
                                current_window_size,
                                self.database_filename,
                                self.default_excel_filename,
                                self.default_xml_file_name,
                                self.default_subwindow_status,
                                self.default_export_extension,
                                self.default_path,
                                self.default_render_mode)

if load_gui:
    if not app_running_status:
        menu = main_menu(root)
        menu.menu(initial=True)

