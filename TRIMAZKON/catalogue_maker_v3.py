import customtkinter
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl import load_workbook
import xlwings as xw
import string
from PIL import Image as PILImage
from datetime import datetime
from tkinter import filedialog
import os
import xml.etree.ElementTree as ET
# import sharepoint_download as download_database
import sys
import threading

# customtkinter.set_appearance_mode("dark")
# customtkinter.set_default_color_theme("dark-blue")
# root=customtkinter.CTk()
# root.geometry("1200x900")
# root.title("Catalogue maker v3.0")
# database_filename  = "Sharepoint_databaze.xlsx"
# root.state('zoomed')

def add_colored_line(text_widget, text, color,font=None,delete_line = None):
    """
    Vloží řádek do console
    """
    text_widget.configure(state=tk.NORMAL)
    if font == None:
        font = ("Arial",16)
    if delete_line != None:
        text_widget.delete("current linestart","current lineend")
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert("current lineend",text, color)
    else:
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert(tk.END,"    > "+ text+"\n", color)

    text_widget.configure(state=tk.DISABLED)

def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

app_icon_path = resource_path('images\\logo_TRIMAZKON.ico')

def path_check(path_raw,only_repair = None):
    path=path_raw
    backslash = "\\"
    if backslash[0] in path:
        newPath = path.replace(backslash[0], '/')
        path = newPath
    if path.endswith('/') == False:
        newPath = path + "/"
        path = newPath
    #oprava mezery v nazvu
    path = r"{}".format(path)
    if not os.path.exists(path) and only_repair == None:
        return False
    else:
        return path

def browseDirectories(visible_files,start_path=None): # Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat
    """
    Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat

    Vstupní data:

    0: visible_files = "all" / "only_dirs"\n
    1: start_path = None -optimalni, docasne se ulozi posledni nastavena cesta v exploreru

    Výstupní data:

    0: výstupní chybová hlášení
    1: opravená cesta
    2: nazev vybraneho souboru (option: all)
    """
    corrected_path = ""
    output= ""
    name_of_selected_file = ""
    start_path = resource_path(os.getcwd())
    start_path = path_check(start_path)

    # pripad vyberu files, aby byly viditelne
    if visible_files == "all":
        if(start_path != ""):
            foldername_path = filedialog.askopenfile(initialdir = start_path,title = "Klikněte na soubor v požadované cestě")
            path_to_directory= ""
            if foldername_path != None:
                path_to_file = str(foldername_path.name)
                path_to_file_split = path_to_file.split("/")
                i=0
                for parts in path_to_file_split:
                    i+=1
                    if i<len(path_to_file_split):
                        if i == 1:
                            path_to_directory = path_to_directory + parts
                        else:
                            path_to_directory = path_to_directory +"/"+ parts
                    else:
                        name_of_selected_file = parts
            else:
                output = "Přes explorer nebyla vložena žádná cesta"
        else:           
            foldername_path = filedialog.askopenfile(initialdir = "/",title = "Klikněte na soubor v požadované cestě")
            path_to_directory= ""
            if foldername_path != None:
                path_to_file = str(foldername_path.name)
                path_to_file_split = path_to_file.split("/")
                i=0
                for parts in path_to_file_split:
                    i+=1
                    if i<len(path_to_file_split):
                        if i == 1:
                            path_to_directory = path_to_directory + parts
                        else:
                            path_to_directory = path_to_directory +"/"+ parts
                    else:
                        name_of_selected_file = parts
            else:
                output = "Přes explorer nebyla vložena žádná cesta"

    # pripad vyberu slozek
    if visible_files == "only_dirs":
        if(start_path != ""):
            path_to_directory = filedialog.askdirectory(initialdir = start_path, title = "Vyberte adresář")
            if path_to_directory == None or path_to_directory == "":
                output = "Přes explorer nebyla vložena žádná cesta"
        else:
            path_to_directory = filedialog.askdirectory(initialdir = "/", title = "Vyberte adresář")
            if path_to_directory == None or path_to_directory == "":
                output = "Přes explorer nebyla vložena žádná cesta"

    check = path_check(path_to_directory)
    corrected_path = check
    return [output,corrected_path,name_of_selected_file]

def strip_lines_to_fit(text):
    number_of_chars = 0
    text_splitted = text.split(" ")
    new_string = ""
    max_num_of_chars_one_line = 32
    for items in text_splitted:
        if items != "\n":
            number_of_chars += len(items)
            if number_of_chars > max_num_of_chars_one_line:
                new_string += "\n" + str(items) + " "
                number_of_chars = len(items)
            else:
                new_string += str(items) + " "

    return new_string

class Save_prog_metadata:
    def __init__(self,console,controller_database=[],station_list=[],project_name="",xml_file_path=""):
        self.controller_database = controller_database
        self.station_list = station_list
        self.project_name = project_name
        self.main_console = console
        self.xml_file_path = xml_file_path
        # self.xml_file_name = str(self.project_name) + "_metadata_catalogue.xml"

        # self.store_xml_data()
        # print(self.read_xml_data(self.xml_file_name))

    def store_xml_data(self):
        # KONTROLERY ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        root1 = ET.Element("metadata")
        controller_list = ET.SubElement(root1, "controllers")
        for controllers in self.controller_database:
            controller = ET.SubElement(controller_list, "controller")
            for contr_key, contr_value in controllers.items():
                if contr_key == "accessory_list":
                    accessories = ET.SubElement(controller, "accessory_list")
                    for accessory in contr_value:
                        accessory_element = ET.SubElement(accessories, "accessory")
                        for acc_key, acc_value in accessory.items():
                            acc_child = ET.SubElement(accessory_element, acc_key)
                            acc_child.text = str(acc_value)  # Ensure value is a string
                else:
                    contr_child = ET.SubElement(controller, contr_key)
                    contr_child.text = str(contr_value)

        # STANICE ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        station_list = ET.SubElement(root1, "station_list")
        for stations in self.station_list:
            station_element = ET.SubElement(station_list, "station")
            for key, value in stations.items():
                if key == "camera_list":
                    cameras = ET.SubElement(station_element, "camera_list")
                    for camera in value:
                        camera_element = ET.SubElement(cameras, "camera")
                        for cam_key, cam_value in camera.items():
                            if cam_key == "optics_list":
                                optics = ET.SubElement(camera_element, "optics_list")
                                for optic in cam_value:
                                    optic_element = ET.SubElement(optics, "optic")
                                    for opt_key, opt_value in optic.items():
                                        opt_child = ET.SubElement(optic_element, opt_key)
                                        opt_child.text = str(opt_value)  # Ensure value is a string
                            else:
                                cam_child = ET.SubElement(camera_element, cam_key)
                                cam_child.text = str(cam_value)  # Ensure value is a string
                else:
                    child = ET.SubElement(station_element, key)
                    child.text = str(value)  # Ensure value is a string
            
        # NÁZEV PROJEKTU ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        project_name = ET.SubElement(root1,"project_name")
        # root3 = ET.Element("project_name")
        if self.project_name == None:
            self.project_name = ""
        project_name.text = str(self.project_name)

        # ULOŽENÍ ----------------------------------------------------------------------------------------------------------------------------------------------------------------
        tree1 = ET.ElementTree(root1)
        try:
            tree1.write(self.xml_file_path, encoding="utf-8", xml_declaration=True)
            add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně uložen","green",None,True)
        except Exception as e:
            add_colored_line(self.main_console,f"Neočekávaná chyba {e}","red",None,True)

    def read_xml_data(self,file_path):
        stations = self.read_stations_xml(file_path)
        controllers = self.read_controllers(file_path)
        project_name = self.read_project_name(file_path)
        print("stations, controllers, project_name: ",[stations,controllers,project_name])
        return [stations,controllers,project_name]

    def read_controllers(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        controllers = []
        controller_list = root.find("controllers")
        for controller in controller_list.findall("controller"):
            # controller_data = {child.tag: child.text for child in controller}
            controller_data = {}
            for child in controller:
                if child.tag == "accessory_list":
                    accessory_list = []
                    for accessory in child.findall("accessory"):
                        # accessory_data = {acc_child.tag: acc_child.text for acc_child in accessory}
                        accessory_data = {}
                        for acc_child in accessory:
                            if acc_child.text is not None:
                                accessory_data[acc_child.tag] = acc_child.text
                            else:
                                accessory_data[acc_child.tag] = ""
                        accessory_list.append(accessory_data)
                    controller_data[child.tag] = accessory_list
                else:                        
                    if child.text is not None:
                        controller_data[child.tag] = child.text
                    else:
                        controller_data[child.tag] = ""

            controllers.append(controller_data)
        
        return controllers
     
    def read_stations_xml(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        stations = []
        station_list = root.find("station_list")
        for station in station_list.findall("station"):
            station_data = {}
            for child in station:
                if child.tag == "camera_list":
                    camera_list = []
                    for camera in child.findall("camera"):
                        camera_data = {}
                        for cam_child in camera:
                            if cam_child.tag == "optics_list":
                                optics_list = []
                                for optic in cam_child.findall("optic"):
                                    optic_data = {}
                                    for opt_child in optic:
                                        if opt_child.text is not None:
                                            optic_data[opt_child.tag] = opt_child.text
                                        else:
                                            optic_data[opt_child.tag] = ""
                                        # optic_data[opt_child.tag] = opt_child.text
                                    optics_list.append(optic_data)
                                camera_data[cam_child.tag] = optics_list
                            else:
                                if cam_child.text is not None:
                                    camera_data[cam_child.tag] = cam_child.text
                                else:
                                    camera_data[cam_child.tag] = ""
                                # camera_data[cam_child.tag] = cam_child.text
                        camera_list.append(camera_data)
                    station_data[child.tag] = camera_list
                else:
                    if child.text is not None:
                        station_data[child.tag] = child.text
                    else:
                        station_data[child.tag] = ""
                        
            stations.append(station_data)
        return stations

    def read_project_name(self,file_path):
        tree = ET.parse(file_path)
        root = tree.getroot()
        project_name = ""
        
        project_name_element = root.find("project_name")
        if project_name_element is not None:
            project_name = project_name_element.text
        else:
            project_name = ""
        
        return project_name
    
class ToplevelWindow:
    def __init__(self,root,controller_databases = [[],[]],callback = None,custom_controller_database = [],accessory_databases=[[],[],[]]):
        self.controller_database = controller_databases[0]
        self.controller_notes_database = controller_databases[1]
        self.custom_controller_database = custom_controller_database
        self.accessory_database = accessory_databases[0]
        self.whole_accessory_database = accessory_databases[1]
        self.accessory_notes_database = accessory_databases[2]
        self.root = root
        self.callback_function = callback
        self.x = self.root.winfo_rootx()
        self.y = self.root.winfo_rooty()
        self.controller_color_list = [
            "",
            "#1E90FF",  # Dodger Blue
            "#32CD32",  # Lime Green
            "#FF4500",  # Orange Red
            "#8A2BE2",  # Blue Violet
            "#00CED1",  # Dark Turquoise
            "#DC143C",  # Crimson
            "#FF6347",  # Tomato
            "#FF69B4",  # Hot Pink
            "#7FFF00",  # Chartreuse
            "#FFD700"  # Gold
        ]
        self.controller_color_pointer = 0
        self.accessory_database_pointer = 0
        self.one_segment_width = 450

    def excel_manual_window(self):
        #1824x805
        # window.geometry(f"912x402+{self.x+100}+{self.y+100}")
        window = customtkinter.CTkToplevel()
        window.geometry(f"1200x580+{self.x+100}+{self.y+200}")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Manual")

        manual_frame =  customtkinter.CTkFrame(master=window,corner_radius=0,height=100,fg_color="#212121")
        manual_frame    .pack(pady=0,padx=0,expand=False,side = "right",anchor="e",ipady = 10,ipadx = 10)
        manual =        customtkinter.CTkImage(PILImage.open(resource_path("images/excel_manual.png")),size=(1200,520))
        manual_label =  customtkinter.CTkLabel(master = manual_frame,text = "",image =manual,bg_color="#212121")
        manual_label    .pack(pady=0,padx=0,expand=True)
        button_exit =   customtkinter.CTkButton(master = manual_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: window.destroy())
        button_exit     .pack(pady=10,padx=10,expand=True,side = "bottom",anchor = "e")

        self.root.bind("<Button-1>",lambda e: window.destroy())
        window.grab_set()
        window.focus_force()

    def new_controller_window(self,childroot,controller = None,edit = False,accessory_index =0,only_accessory=False):
        """
        vrací:
        - zvolený kontroler z databáze
        - název (pojmenování) kontoleru
        - barva kontroleru
        - ftp adresa
        - jméno uživatele
        - heslo uživatele
        - accessory list, příslušenství ke kontroleru
        - poznámky ke kontroleru
        """
        def save_contoller():
            save_changes()
            # output = [controller_entry.get(),controller_name_entry.get(),IP_adress_entry.get(),username_entry.get(),password_entry.get()]

            notes = notes_input.get("1.0", tk.END)
            try:
                color_chosen = controller_color.cget("fg_color")
            except Exception:
                color_chosen = ""
            print("chosen color: ",color_chosen)
            output = [controller_entry.get(),controller_name_entry.get(),color_chosen,IP_adress_entry.get(),username_entry.get(),password_entry.get(),controller["accessory_list"],notes]
            # window.destroy()
            close_window(window)
            # self.root.focus_force()
            self.callback_function(output)

        def check_used_colors():
            used_colors=[]
            print("controller database input",self.custom_controller_database)
            for items in self.custom_controller_database:
                if items["color"] != "": # možnost nezvolit žádnou barvu
                    if items["color"] in self.controller_color_list:
                        if controller != None:
                            if not items["color"] == controller["color"]:
                                used_colors.append(items["color"])
                                self.controller_color_list.pop(self.controller_color_list.index(items["color"]))
                        elif not edit:
                            used_colors.append(items["color"])
                            self.controller_color_list.pop(self.controller_color_list.index(items["color"]))
            print("used colors: ",used_colors)
        check_used_colors()

        def switch_color():
            self.controller_color_pointer += 1
            if self.controller_color_pointer > len(self.controller_color_list)-1:
                self.controller_color_pointer = 0

            new_color = self.controller_color_list[self.controller_color_pointer]
            if new_color != "":
                controller_color.configure(fg_color = new_color,hover_color = new_color)
            else:
                controller_color.configure(fg_color= "#212121",hover_color="#212121")

        def close_window(window):
            if not edit:
                try:
                    childroot.unbind("<Button-1>")
                except Exception:
                    pass
            window.destroy()

        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window_height = 870
        window.geometry(f"{self.one_segment_width}x{window_height}+{self.x+150}+{self.y+5}")
        if edit:
            current_name = controller["name"]
            current_type = controller["type"]
            current_ip = controller["ip"]
            current_username = controller["username"]
            current_password = controller["password"]
            current_color = controller["color"]
            current_notes = controller["notes"]

            window.title(f"Editování kontroleru: {current_name} ({current_type})")
        else:
            window.title("Nový kontroler")
            controller = {
                "accessory_list": []
            }
        
        def save_changes():
            try:
                controller["accessory_list"][accessory_index]["type"] = hw_type_entry.get()
                controller["accessory_list"][accessory_index]["description"] = notes_input3.get("1.0", tk.END)
            except IndexError:
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    new_accessory = {
                    "type": hw_type_entry.get(),
                    "description":notes_input3.get("0.0", "end"),
                    }
                    controller["accessory_list"].append(new_accessory)
            except TypeError: # pokud je jako index vložen None
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    new_accessory = {
                    "type": hw_type_entry.get(),
                    "description":notes_input3.get("0.0", "end"),
                    }
                    controller["accessory_list"].append(new_accessory)
            
            print("acc_list --------- ",controller["accessory_list"])

        def next_accessory():
            nonlocal accessory_index
            accessory_index += 1
            if accessory_index < len(controller["accessory_list"]):
                accessory_index -= 1
                save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                accessory_index += 1
                # initial_prefill() # prefill s novým indexem - index se prenese i do ukládání

            else: # TLACITKO +:
                # program nedopusti pridani noveho accessory pokud neni alespon vyplnen typ nebo poznamka
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    accessory_index -= 1
                    save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                    accessory_index += 1
                else:
                    accessory_index -= 1
            initial_prefill() # prefill s novým indexem

        def previous_accessory():
            nonlocal accessory_index
            accessory_index -= 1
            if accessory_index > -1:
                if hw_type_entry.get() != "" or notes_input3.get("0.0", "end") != "\n":
                    accessory_index += 1
                    save_changes() # ulozit zmeny pri prepinani jeste u predesle stanice
                    accessory_index -= 1
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                accessory_index += 1

        def switch_database_section(operation,database,widget_menu):
            """
            mění hodnotu pointeru na pole hodnot v option menu
            """
            if operation == "next":
                self.accessory_database_pointer +=1
                if self.accessory_database_pointer > len(database)-1:
                    self.accessory_database_pointer = 0
            elif operation == "prev":
                self.accessory_database_pointer -=1
                if self.accessory_database_pointer < 0:
                    self.accessory_database_pointer = len(database)-1
            
            widget_menu.configure(values = database[self.accessory_database_pointer])
            widget_menu._open_dropdown_menu()

        def import_notes(operation = ""):
            notes_string = ""
            if operation == "controller":
                current_controller = controller_entry.get()
                if current_controller != "":
                    controller_notes = str(self.controller_notes_database[self.controller_database.index(current_controller)])
                    if controller_notes != "":
                        notes_string = notes_string + controller_notes + "\n\n"
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",notes_string)
            else:
                current_accessory = hw_type_entry.get()
                if current_accessory != "":
                    notes_string = notes_string + str(self.accessory_notes_database[self.whole_accessory_database.index(current_accessory)])
                notes_input3.delete("1.0",tk.END)
                notes_input3.insert("1.0",notes_string)


        # KONTROLER ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        main_frame =                customtkinter.CTkFrame(master = window,corner_radius=0,border_width=3)
        controller_frame =          customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_width=3)
        controller_type =           customtkinter.CTkLabel(master = controller_frame,text = "Typ kontroleru: ",font=("Arial",20,"bold"))
        controller_entry =          customtkinter.CTkOptionMenu(master = controller_frame,font=("Arial",22),dropdown_font=("Arial",22),values=self.controller_database,corner_radius=0,height=50)
        controller_name =           customtkinter.CTkLabel(master = controller_frame,text = "Název (interní označení): ",font=("Arial",20,"bold"))
        controller_name_entry =     customtkinter.CTkEntry(master = controller_frame,font=("Arial",20),corner_radius=0,height=50)
        controller_color =          customtkinter.CTkButton(master = controller_frame,corner_radius=0,text="Podbarvení kontroleru",font=("Arial",20,"bold"),height=50,command=lambda:switch_color())
        IP_adress =                 customtkinter.CTkLabel(master = controller_frame,text = "IP adresa: ",font=("Arial",20,"bold"))
        IP_adress_entry =           customtkinter.CTkEntry(master = controller_frame,font=("Arial",20),corner_radius=0,height=50)
        username =                  customtkinter.CTkLabel(master = controller_frame,text = "Jméno: ",font=("Arial",20,"bold"))
        username_entry =            customtkinter.CTkEntry(master = controller_frame,font=("Arial",20),corner_radius=0,height=50)
        password =                  customtkinter.CTkLabel(master = controller_frame,text = "Heslo: ",font=("Arial",20,"bold"))
        password_entry =            customtkinter.CTkEntry(master = controller_frame,font=("Arial",20),corner_radius=0,height=50,placeholder_text="*******")
        note_label_frame =         customtkinter.CTkFrame(master = controller_frame,corner_radius=0)
        note_label =               customtkinter.CTkLabel(master = note_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes_btn =         customtkinter.CTkButton(master = note_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("controller"))
        note_label                 .pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes_btn           .pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        notes_input =              customtkinter.CTkTextbox(master = controller_frame,font=("Arial",22),corner_radius=0)

        controller_type.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        controller_entry.           pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        controller_name.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        controller_name_entry.      pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        controller_color.           pack(pady=0,padx=10,side =      "top",anchor = "w",fill="x")
        IP_adress.                  pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        IP_adress_entry.            pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        username.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        username_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        password.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
        password_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill ="x")
        note_label_frame.           pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        notes_input.                pack(pady = 5, padx = 10,expand=True,side="top",fill="both")

        selected_color = self.controller_color_list[self.controller_color_pointer]
        if selected_color != "":
            controller_color.configure(fg_color=selected_color,hover_color=selected_color)
        else:
            controller_color.configure(fg_color= "#212121",hover_color="#212121")

        # PŘÍSLUŠENSTVÍ ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        accessory_frame =           customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_width=3)
        counter_frame_acc =         customtkinter.CTkFrame(master = accessory_frame,corner_radius=0,fg_color="transparent")
        button_prev_acc =           customtkinter.CTkButton(master = counter_frame_acc,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_accessory())
        counter_acc =               customtkinter.CTkLabel(master = counter_frame_acc,text = "0/0",font=("Arial",22,"bold"))
        button_next_acc =           customtkinter.CTkButton(master = counter_frame_acc,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_accessory())
        button_prev_acc             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        counter_acc                 .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_acc             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        accessory_label =           customtkinter.CTkLabel(master = accessory_frame,text = "Příslušenství:",font=("Arial",22,"bold"))
        hw_type =                   customtkinter.CTkLabel(master = accessory_frame,text = "Zařízení:",font=("Arial",22,"bold"))
        option_menu_frame_acc =     customtkinter.CTkFrame(master = accessory_frame,corner_radius=0)
        hw_type_entry =             customtkinter.CTkOptionMenu(master = option_menu_frame_acc,font=("Arial",22),dropdown_font=("Arial",22),height=50,values=self.accessory_database[self.accessory_database_pointer],corner_radius=0)
        button_prev_section_acc =   customtkinter.CTkButton(master = option_menu_frame_acc,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                            command=lambda: switch_database_section("prev",self.accessory_database,hw_type_entry))
        button_next_section_acc =   customtkinter.CTkButton(master = option_menu_frame_acc,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                            command=lambda: switch_database_section("next",self.accessory_database,hw_type_entry))
        hw_type_entry               .pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left",fill="x")
        button_prev_section_acc     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_section_acc     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        note3_label_frame =         customtkinter.CTkFrame(master = accessory_frame,corner_radius=0)
        note3_label =               customtkinter.CTkLabel(master = note3_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes3_btn =         customtkinter.CTkButton(master = note3_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes())
        note3_label                 .pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes3_btn           .pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        notes_input3 =              customtkinter.CTkTextbox(master = accessory_frame,font=("Arial",22),corner_radius=0)
        counter_frame_acc           .pack(pady=(10,0),padx=3,anchor="n",expand=False,side = "top")
        accessory_label             .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
        hw_type                     .pack(pady= 5 ,padx=10,anchor="w",expand=False,side = "top")
        option_menu_frame_acc       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        # hw_type_entry               .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        note3_label_frame           .pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        notes_input3                .pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        main_frame                  .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="top")
        if not only_accessory:
            window.geometry(f"{2*self.one_segment_width}x{window_height}+{self.x+150}+{self.y+5}")
            controller_frame            .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
        accessory_frame             .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
        bottom_frame =              customtkinter.CTkFrame(master=window,corner_radius=0)
        bottom_frame                .pack(pady = 0, padx = 0,fill="x",anchor="s",expand=True,side="bottom")
        button_save =               customtkinter.CTkButton(master = bottom_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_contoller())
        button_exit =               customtkinter.CTkButton(master = bottom_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        button_save                 .pack(pady=10,padx=(0,10),expand=False,side = "right",anchor="e")
        button_exit                 .pack(pady=10,padx=(0,10),expand=False,side = "right",anchor="e")

        def filter_text_input(text):
            """
            - removes extra new empty lines
            """
            legit_rows = []
            legit_notes = ""
            rows = text.split("\n")
            for i in range(0,len(rows)):
                if rows[i].replace(" ","") != "":
                    legit_rows.append(rows[i])

            for i in range(0,len(legit_rows)): 
                if i == len(legit_rows)-1:
                    legit_notes = legit_notes + legit_rows[i]
                else:
                    legit_notes = legit_notes + legit_rows[i]+ "\n"
            return legit_notes

        if edit:
            IP_adress_entry.insert(0,str(current_ip))
            controller_entry.set(str(current_type))
            controller_name_entry.insert(0,str(current_name))
            username_entry.insert(0,str(current_username))
            password_entry.insert(0,str(current_password))
            if str(current_color) != "":
                controller_color.configure(fg_color = str(current_color))
            else:
                controller_color.configure(fg_color= "#212121",hover_color="#212121")

            notes_input.delete("1.0",tk.END)
            notes_input.insert("1.0",filter_text_input(str(current_notes)))
        else:
            IP_adress_entry.insert(0,"192.168.000.000")
            controller_name_entry.insert(0,"Kontroler " + str(len(self.custom_controller_database)+1) + " ")
            childroot.bind("<Button-1>",lambda e: close_window(window))

        def refresh_counters():
            nonlocal accessory_index
            nonlocal counter_acc
            if len(controller["accessory_list"]) == 0:
                addition = 0
            else:
                addition = 1
            try:
                counter_acc_state = str(accessory_index+1) + "/" + str(len(controller["accessory_list"]))
                counter_acc.configure(text = counter_acc_state)
            except Exception:
                pass

        def refresh_button_appearance():
            nonlocal accessory_index
            nonlocal button_prev_acc
            nonlocal button_next_acc

            def config_buttons(button_left,button_right,index,max_array_value):
                if index ==0:
                    button_left.configure(text = "",fg_color = "#636363")
                else:
                    button_left.configure(text = "<",fg_color = "#636363")

                if index == max_array_value:
                    button_right.configure(text = "+",fg_color = "green")
                else:
                    button_right.configure(text = ">",fg_color = "#636363")
            # pokud není accessory:
            try:
                config_buttons(button_prev_acc,button_next_acc,accessory_index,len(controller["accessory_list"])-1)
            except Exception as e:
                print(f"chyba při nastavování vzhledu tlačítek - accessory: {e}")

        def initial_prefill():
            nonlocal hw_type_entry
            nonlocal notes_input3
            nonlocal accessory_index

            try:
                if str(controller["accessory_list"][accessory_index]["type"]) in self.whole_accessory_database:
                    hw_type_entry.set(str(controller["accessory_list"][accessory_index]["type"]))
                else:
                    hw_type_entry.set("")
                notes_input3.delete("1.0",tk.END)
                notes_input3.insert("1.0",filter_text_input(str(controller["accessory_list"][accessory_index]["description"])))
            except TypeError: # pokud je v indexu None - defaultně nastavit index na nulu:
                try:
                    accessory_index = 0
                    if str(controller["accessory_list"][accessory_index]["type"]) in self.whole_accessory_database:   
                        hw_type_entry.set(str(controller["accessory_list"][accessory_index]["type"]))
                    else:
                        hw_type_entry.set("")
                    notes_input3.delete("1.0",tk.END)
                    notes_input3.insert("1.0",filter_text_input(str(controller["accessory_list"][accessory_index]["description"])))
                except IndexError: #případ, že není accessory
                    hw_type_entry.set("")
                    notes_input3.delete("1.0",tk.END)
            except IndexError: #případ, že není accessory
                hw_type_entry.set("")
                notes_input3.delete("1.0",tk.END)

            refresh_counters()
            refresh_button_appearance()

        initial_prefill()
        self.root.bind("<Button-1>",lambda e: close_window(window))
        global opened_subwindow
        opened_subwindow = window
        window.update()
        window.update_idletasks()
        window.focus_force()
        window.focus()

    def save_prog_options_window(self,main_console,station_list,project_name,callback,callback_save_last_file,last_file = None,last_path = "",default_xml_file_name="_metadata_catalogue",default_path = ""):
        """
        okno s možnostmi uložení rozdělaného projektu
        """
        window = customtkinter.CTkToplevel()
        window.geometry(f"1015x350+{self.x+200}+{self.y+50}")
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window.title("Možnosti uložení projektu")
        subwindow = ""

        def confirm_window(final_path):
            nonlocal subwindow
            def call_save(final_path):
                nonlocal subwindow
                save_file(final_path)
                subwindow.destroy()
            subwindow = customtkinter.CTkToplevel()
            subwindow.after(200, lambda: subwindow.iconbitmap(app_icon_path))
            subwindow.geometry(f"900x110+{self.x+200}+{self.y+50}")
            subwindow.title("Potvrdit přepsání souboru")
            main_frame = customtkinter.CTkFrame(master = subwindow,corner_radius=0)
            export_label = customtkinter.CTkLabel(master = main_frame,text = "V zadané cestě se soubor s tímto názvem již vyskytuje, přepsat?",font=("Arial",22,"bold"),justify = "left",text_color="orange")
            button_frame = customtkinter.CTkFrame(master = main_frame,corner_radius=0)
            button_yes = customtkinter.CTkButton(master = button_frame,text = "Ano",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command = lambda: call_save(final_path))
            button_no = customtkinter.CTkButton(master = button_frame,text = "Ne",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command = lambda: subwindow.destroy())
            main_frame.pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
            export_label.pack(pady = 0, padx = 0,fill="x",anchor="w",expand=False,side="top")
            button_frame.pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="bottom")
            button_yes.pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
            button_no.pack(pady = 10, padx = (10,0),expand=False,side="right",anchor = "e")
            subwindow.update()
            subwindow.update_idletasks()
            subwindow.focus_force()
            subwindow.focus()

        def close_window(window):
            nonlocal subwindow
            try:
                if subwindow.winfo_exists():
                    subwindow.destroy()
            except Exception:
                pass
            window.destroy()

        def create_path(path_inserted):
            nonlocal export_name
            nonlocal default_xml_file_name
            file_name = export_name.get()
            if file_name =="":
                file_name = default_xml_file_name
            path = path_inserted + file_name + ".xml"
            print(path)
            return path

        def save_file(final_path):
            nonlocal window
            # ukladani posledne zadaneho nazvu souboru:
            nonlocal export_name
            nonlocal export_path
            path_inserted = export_path.get()
            file_name = export_name.get()
            callback_save_last_file(file_name,path_inserted)
            # samotne ukladani vsech dat:
            save_prog = Save_prog_metadata(station_list=station_list,project_name=project_name,controller_database=self.custom_controller_database,console=console,xml_file_path=final_path)
            save_prog.store_xml_data()
            add_colored_line(main_console,f"Data úspěšně uložena do: {final_path}","green",None,True)
            close_window(window)

        def call_save_file(window):
            nonlocal console
            nonlocal export_path
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                final_path = create_path(path_inserted)
                if os.path.exists(final_path):
                    confirm_window(final_path)
                else:
                    save_file(final_path)
            else:
                add_colored_line(console,"Zadaná cesta pro uložení je neplatná","red",None,True)
        
        def call_load_file(window):
            nonlocal console
            nonlocal export_path
            nonlocal export_name
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                final_path = create_path(path_inserted)
                save_prog = Save_prog_metadata(station_list=station_list,project_name=project_name,controller_database=self.custom_controller_database,console=console)
                try:
                    received_data = save_prog.read_xml_data(final_path)
                    add_colored_line(main_console,f"Data úspěšně nahrána z: {final_path}","green",None,True)
                    callback(received_data)
                    # ulozit posledně načtený soubor
                    file_name = export_name.get()
                    callback_save_last_file(file_name,path_inserted)
                    window.destroy()
                except Exception:
                    add_colored_line(main_console,f"Soubor .xml je neplatný: {final_path}","red",None,True)
                    window.destroy()
            else:
                add_colored_line(console,f"V zadané cestě nebyl nalezen soubor .xml s názvem {export_name.get()}","red",None,True)

        def call_browse_directories(what_search):
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if what_search == "only_dirs":
                output = browseDirectories(what_search)
                if str(output[1]) != "/":
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            else:
                output = browseDirectories(what_search)
                if str(output[1]) != "/":
                    export_name.delete(0,300)
                    name_without_extension = str(output[2])[:-4]
                    export_name.insert(0, name_without_extension)
                    export_path.delete(0,300)
                    export_path.insert(0, str(output[1]))
                    add_colored_line(console,"Byla vložena cesta a název souboru","green",None,True)
            print(output[0])

            window.focus_force()
            window.focus()

        def search_for_xmls(path):
            found_files = []
            for files in os.listdir(path):
                if ".xml" in files:
                    if not files in found_files:
                        found_files.append(files)
            return found_files

        def save_current_path():
            path_inserted = str(export_path.get())
            if path_inserted.replace(" ","") != "":
                checked_path = path_check(path_inserted)
                if checked_path != False:
                    callback_save_last_file(None,None,checked_path)

        export_frame =          customtkinter.CTkFrame(master = window,corner_radius=0)
        export_label =          customtkinter.CTkLabel(master = export_frame,text = "Zadejte název souboru:",font=("Arial",22,"bold"))
        export_name_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0)
        export_name =           customtkinter.CTkEntry(master = export_name_frame,font=("Arial",20),width=730,height=50,corner_radius=0)
        explorer_btn_name =     customtkinter.CTkButton(master = export_name_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("all"))
        format_entry =          customtkinter.CTkOptionMenu(master = export_name_frame,font=("Arial",22),dropdown_font=("Arial",22),values=[".xml"],width=200,height=50,corner_radius=0)
        export_name             .pack(pady = 5, padx = 10,anchor="w",fill="x",expand=True,side="left")
        format_entry            .pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        explorer_btn_name       .pack(pady = 5, padx = 0,anchor="e",expand=False,side="right")
        export_label2 =         customtkinter.CTkLabel(master = export_frame,text = "Zadejte cestu, kam soubor uložit:",font=("Arial",22,"bold"))
        export_path_frame =     customtkinter.CTkFrame(master = export_frame,corner_radius=0)
        export_path =           customtkinter.CTkEntry(master = export_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        explorer_btn =          customtkinter.CTkButton(master = export_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories("only_dirs"))
        save_path_btn =         customtkinter.CTkButton(master = export_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: save_current_path())
        export_path             .pack(pady = 5, padx = (10,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn           .pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        explorer_btn            .pack(pady = 5, padx = (10,0),anchor="e",expand=False,side="right")
        console =               tk.Text(export_frame, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
        button_load =           customtkinter.CTkButton(master = export_frame,text = "Nahrát",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_load_file(window))
        button_save =           customtkinter.CTkButton(master = export_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_save_file(window))
        button_exit =           customtkinter.CTkButton(master = export_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        export_frame            .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        export_label            .pack(pady=(15,5),padx=10,anchor="w",expand=False,side="top")
        export_name_frame       .pack(expand=True,side="top",anchor="n",fill="x")
        export_label2           .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        export_path_frame       .pack(expand=True,side="top",anchor="n",fill="x")
        console                 .pack(expand=True,side="top",anchor="n",fill="x")
        button_load             .pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        button_save             .pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        button_exit             .pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")

        checked_last_path = path_check(last_path)
        default_path = path_check(default_path)
        if checked_last_path != False and checked_last_path != None and checked_last_path.replace(" ","") != "" and checked_last_path.replace(" ","") != "/":
            initial_path = str(checked_last_path)
            export_path.insert("0",resource_path(str(checked_last_path)))
            add_colored_line(console,"Byla vložena posledně zvolená cesta","green",None,True)

        elif default_path != False and default_path != None and default_path.replace(" ","") != "" and default_path.replace(" ","") != "/":
            initial_path = str(default_path)
            export_path.insert("0",resource_path(str(default_path)))
            add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)
        else:
            initial_path = path_check(os.getcwd())
            export_path.insert("0",resource_path(str(initial_path)))

        found_xmls = search_for_xmls(initial_path)
        # posledni ulozeny/ nacteny soubor:
        if last_file != None and last_file.replace(" ","") != "":
            export_name.insert("0",str(last_file))
        # první soubor nalezeny ve slozce:
        elif len(found_xmls) > 0:
            export_name.insert("0",str(found_xmls[0].replace(".xml","")))
            print("nalezené soubory xml: ",found_xmls)
        # default název + název projektu:
        else:
            export_name.insert("0",str(project_name) + default_xml_file_name)

        self.root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        window.focus_force()
        window.focus()

    def setting_window(self,default_excel_name,default_xml_name,window_status,callback,default_database_filename):
        def close_window(window):
            window.destroy()

        def save_changes():
            def filter_input(data):
                forbidden_formats = [".","xml","xlsm","xlsx"]
                for formats in forbidden_formats:
                    data = data.replace(formats,"")
                return data

            if checkbox.get() == 1:
                window_status = 1
            else:
                window_status = 0
            default_excel_name = filter_input(str(excel_name_label_entry.get()))
            default_xml_name = filter_input(str(xml_name_label_entry.get()))
            default_database_filename = filter_input(str(default_database_name_entry.get()))

            input_data = [default_excel_name,default_xml_name,window_status,default_database_filename]
            callback(input_data)
            close_window(window)

        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(app_icon_path))
        window_height = 500
        window_width = 700
        window.geometry(f"{window_width}x{window_height}+{self.x+150}+{self.y+50}")
        window.title("Nastavení")

        main_frame =                customtkinter.CTkFrame(master = window,corner_radius=0)
        option1_frame =             customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1)
        checkbox =                  customtkinter.CTkCheckBox(master = option1_frame, text = "Okna editování otevírat maximalizované",font=("Arial",22,"bold"))#,command=lambda: save_new_behav_notes()
        checkbox.                   pack(pady = 20, padx = 10,anchor="w")
        option2_frame =             customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1)
        xml_name_label =            customtkinter.CTkLabel(master = option2_frame,text = "Nastavte základní název pro ukládání rozpracovaného projektu:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        xml_name_frame =            customtkinter.CTkFrame(master = option2_frame,corner_radius=0)
        xml_name_label_entry =      customtkinter.CTkEntry(master = xml_name_frame,font=("Arial",20),corner_radius=0)
        xml_extension_label =       customtkinter.CTkLabel(master = xml_name_frame,text = ".xml",font=("Arial",22,"bold"),justify = "left",anchor="w")
        xml_extension_label.        pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        xml_name_label_entry.       pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        xml_name_label.             pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        xml_name_frame.             pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        option3_frame =             customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1)
        excel_name_label =          customtkinter.CTkLabel(master = option3_frame,text = "Nastavte základní název pro ukládání excelu:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        excel_name_frame =          customtkinter.CTkFrame(master = option3_frame,corner_radius=0)
        excel_name_label_entry =    customtkinter.CTkEntry(master = excel_name_frame,font=("Arial",20),corner_radius=0)
        excel_extension_label =     customtkinter.CTkLabel(master = excel_name_frame,text = ".xlsm/ .xlsx",font=("Arial",22,"bold"),justify = "left",anchor="w")
        excel_extension_label.      pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        excel_name_label_entry.     pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        excel_name_label.           pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        excel_name_frame.           pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        option4_frame =                 customtkinter.CTkFrame(master = main_frame,corner_radius=0,border_color="#505050",border_width=1)
        default_database_name =         customtkinter.CTkLabel(master = option4_frame,text = "Nastavte základní název souboru databáze produktů:",font=("Arial",22,"bold"),justify = "left",anchor="w")
        default_database_name_warning = customtkinter.CTkLabel(master = option4_frame,text = "(název se musí shodovat s názvem souboru na sharepointu)",font=("Arial",22),justify = "left",anchor="w",text_color="orange")
        default_database_name_frame =   customtkinter.CTkFrame(master = option4_frame,corner_radius=0)
        default_database_name_entry =   customtkinter.CTkEntry(master = default_database_name_frame,font=("Arial",20),corner_radius=0)
        database_extension_label =      customtkinter.CTkLabel(master = default_database_name_frame,text = ".xlsx",font=("Arial",22,"bold"),justify = "left",anchor="w")
        database_extension_label.       pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        default_database_name_entry.    pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        default_database_name.          pack(pady = (10,0), padx = 10,fill="x",anchor="w",side="top")
        default_database_name_warning.  pack(pady = 0, padx = 10,fill="x",anchor="w",side="top")
        default_database_name_frame.    pack(pady = 10, padx = 10,fill="x",anchor="w",side="top")
        button_save =               customtkinter.CTkButton(master = main_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_changes())
        button_exit =               customtkinter.CTkButton(master = main_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(window))
        main_frame.                 pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 10,ipadx=10)
        option1_frame.              pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option2_frame.              pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option3_frame.              pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        option4_frame.              pack(pady = 0, padx = 0,fill="x",anchor="n",expand=False,side="top")
        button_save.                pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        button_exit.                pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")

        excel_name_label_entry.insert(0,str(default_excel_name))
        xml_name_label_entry.insert(0,str(default_xml_name))
        default_database_name_entry.insert(0,str(default_database_filename.replace(".xlsx","")))

        if window_status == 1:
            checkbox.select()

        self.root.bind("<Button-1>",lambda e: close_window(window))
        window.update()
        window.update_idletasks()
        window.focus_force()
        window.focus()

class Catalogue_gui:
    def __init__(self,root,download_status,callback_function,window_size,database_filename,default_excel_name,default_xml_name,default_subwindow_status,default_file_extension,default_path):
        self.root = root
        self.default_path = default_path
        self.download_status = download_status
        self.callback = callback_function
        self.default_database_filename = database_filename
        self.path_for_callback = None
        if window_size == "max":
            self.root.state('zoomed')
            # root.state('zoomed')
            self.root.update()
        else:
            self.root.state('normal')
            self.root.geometry("1600x900")
            self.root.update()
        self.station_list = []
        self.default_block_width = 400
        self.format_list = ["xlsm","xlsx"]
        self.favourite_format = "xlsm"
        try:
            if default_file_extension in self.format_list:
                self.favourite_format = default_file_extension
            elif default_file_extension.replace(".","") in self.format_list:
                self.favourite_format = default_file_extension.replace(".","")
        except Exception:
            pass
        self.current_block_id = "00"
        self.controller_object_list = []
        self.custom_controller_drop_list = [""]
        self.chosen_manufacturer = "Omron"
        self.last_selected_widget = ""
        self.controller_database = []
        self.controller_notes_database = []
        self.camera_type_database = [["no data"]]
        self.camera_notes_database = []
        self.whole_camera_type_database = []
        self.camera_cable_database = [["no data"]]
        self.cable_notes_database = []
        self.whole_camera_cable_database = []
        self.optics_database = [["no data"]]
        self.whole_optics_database =[]
        self.optics_notes_database = []
        self.accessory_database = [["no data"]]
        self.accessory_notes_database = []
        self.whole_accessory_database = []
        self.last_xml_filename = ""
        self.last_path_input = ""
        self.default_excel_filename = default_excel_name
        self.default_xml_file_name = default_xml_name
        self.default_subwindow_status = default_subwindow_status # 0 = minimalized, 1 = maximalized
        self.read_database()
        self.create_main_widgets()

    def close_window(self,window):
        window.update_idletasks()
        window.destroy()
    
    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused:
            return True
        else:
            return False

    def read_database(self):
        """
        Stahuje aktuální databázi do adresáře
        - 1. controller_database, controller_notes_database
        - 2. camera_database
        - 3. optics_database
        - 4. accessory_database
        """
        self.download_database_console_input = []
        if self.chosen_manufacturer == "Omron":
            column_index = 1
        else:
            column_index = 3

        if "chyba" in self.download_status or "nepodařilo" in self.download_status:
            text_color = "red"
        else:
            text_color = "green"
        self.download_database_console_input.append(self.download_status)
        self.download_database_console_input.append(text_color)

        sharepoint_database_path = resource_path(path_check(os.getcwd()) + self.default_database_filename)

        self.camera_database_pointer = 0
        self.optics_database_pointer = 0
        self.camera_cable_database_pointer = 0
        self.accessory_database_pointer = 0
        load_failed = False
        try:
            wb = load_workbook(filename=sharepoint_database_path)
        except Exception:
            load_failed = True
            self.download_database_console_input = []
            self.download_database_console_input.append(f"Chyba - selhalo načtení databáze produktů ({sharepoint_database_path})")
            self.download_database_console_input.append("red")

        if not load_failed:
            def fill_lists(wb,name_of_excel_sheet:str,empty_option = True):
                """
                - Vrací seznam produktů přečtecńých z databáze
                    - 0 = celá, kompletní databáze produktů
                    - 1 = databáze rozdělená podle *** ([[]])
                    - 2 = druhý parametr - nyní poznámky - kompletní databáze
                """
                nonlocal column_index
                database_section = [""]
                section_database = []
                whole_database = [""]
                notes_database = [""]
                if not empty_option:
                    whole_database = []
                    notes_database = []
                ws = wb[name_of_excel_sheet]
                # první parametr - typ produktu
                row_count = 0
                for row in ws.iter_rows(min_row=2,min_col=column_index, max_col=column_index,values_only=True):
                    if (row[0] is not None or str(row[0]) != "None") and "***" not in str(row[0]): 
                        database_section.append(str(row[0]))
                        whole_database.append(str(row[0]))
                    elif "***" in str(row[0]):
                        section_database.append(database_section)
                        database_section = []
                    row_count +=1
                if database_section != []:
                    section_database.append(database_section)

                # druhý parametr, dolplňující poznámky
                for row in ws.iter_rows(min_row=2,max_row=row_count+1,min_col=column_index+1, max_col=column_index+1,values_only=True):
                    if row[0] is not None and str(row[0]) != "None": 
                        notes_database.append(str(row[0]))
                    else:
                        notes_database.append("")
                    
                
                return [whole_database,section_database,notes_database]
            
            # KONTROLERY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kontrolery",empty_option = False)
            self.controller_database = read_database[0]
            self.controller_notes_database = read_database[2]
            # KAMERY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kamery",empty_option = True)
            self.whole_camera_type_database = read_database[0]
            self.camera_type_database = read_database[1]
            self.camera_notes_database = read_database[2]
            # KABELY ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Kabely",empty_option = True)
            self.whole_camera_cable_database = read_database[0]
            self.camera_cable_database = read_database[1]
            self.cable_notes_database = read_database[2]
            # OPTIKA ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Optika",empty_option = True)
            self.whole_optics_database = read_database[0]
            self.optics_database = read_database[1]
            self.optics_notes_database = read_database[2]
            # PŘÍSLUŠENSTVÍ ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
            read_database = fill_lists(wb,"Přislušenství",empty_option = True)
            self.whole_accessory_database = read_database[0]
            self.accessory_database = read_database[1]
            self.accessory_notes_database = read_database[2]
            wb.close()
        
    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
        """
        self.clear_frame(self.root)
        if not  self.default_database_filename.endswith(".xlsx"):
            default_database_name_w_extension = self.default_database_filename + ".xlsx"
        else:
            default_database_name_w_extension = self.default_database_filename
        self.callback([self.default_excel_filename,self.default_xml_file_name,self.default_subwindow_status,self.favourite_format,self.path_for_callback,default_database_name_w_extension])

    def switch_widget_info(self,args,widget_tier,widget):
        if len(widget_tier) == 2: #01-99 stanice
            station_index = int(widget_tier[:2])
            if widget._text != str(self.station_list[station_index]["name"]):
                widget.configure(text=str(self.station_list[station_index]["name"]),font = ("Arial",25,"bold"))
            else:
                notes_raw = str(self.station_list[station_index]["inspection_description"])
                description = strip_lines_to_fit(notes_raw)
                # widget.configure(text=str(self.station_list[station_index]["inspection_description"]),font = ("Arial",25))
                widget.configure(text=str(description),font = ("Arial",25))
        
        elif len(widget_tier) == 3: # c01-c99 kontolery
            details = ""
            controller_index = int(widget_tier[1:3])
            if widget._text == str(self.controller_object_list[controller_index]["type"]):
                whole_name = self.controller_object_list[controller_index]["detailed_name"]
                ip_address = self.controller_object_list[controller_index]["ip"]
                username = self.controller_object_list[controller_index]["username"]
                password = self.controller_object_list[controller_index]["password"]
                notes_raw = str(self.controller_object_list[controller_index]["notes"])
                description = strip_lines_to_fit(notes_raw)

                details = whole_name + "\n"
                if not description == "":
                    details = details + description + "\n\n"

                if not ip_address == "" and not ip_address == "192.168.000.000":
                    details = details + ip_address + "\n"
                if not username == "":
                    details = details + "Jméno: " + username + "\n"
                if not password == "":
                    details = details + "Heslo: " + password + "\n"
                

                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=str(self.controller_object_list[controller_index]["type"]),font = ("Arial",25,"bold"))

        elif len(widget_tier) == 4: # 0101-9999 kamery
            station_index = int(widget_tier[:2])
            details = ""
            camera_index = int(widget_tier[2:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["type"]):
                notes_raw = str(self.station_list[station_index]["camera_list"][camera_index]["description"])
                description = strip_lines_to_fit(notes_raw)
                cable = str(self.station_list[station_index]["camera_list"][camera_index]["cable"])
                if cable != "":
                    details = details + "Kabel: " + str(self.station_list[station_index]["camera_list"][camera_index]["cable"])+ "\n"
                details = details + description
                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["type"]),font = ("Arial",25,"bold"))

        elif len(widget_tier) == 5: # c0101-c9999 prislusenstvi kontroleru
            details = ""
            controller_index = int(widget_tier[1:3])
            accessory_index = int(widget_tier[3:5])

            if widget._text == str(self.controller_object_list[controller_index]["accessory_list"][accessory_index]["type"]):
                notes_raw = str(self.controller_object_list[controller_index]["accessory_list"][accessory_index]["description"])
                description = strip_lines_to_fit(notes_raw)
                widget.configure(text=description,font = ("Arial",25))
            else:
                widget.configure(text=str(self.controller_object_list[controller_index]["accessory_list"][accessory_index]["type"]),font = ("Arial",25,"bold"))
            
        elif len(widget_tier) == 6: # 010101-999999 optika
            details = ""
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"]):
                alternative = str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["alternative"])
                if alternative != "":
                    details = "Alternativa: " +  alternative + "\n"

                notes_raw = str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["description"])
                description = strip_lines_to_fit(notes_raw)
                details = details + description
                widget.configure(text=details,font = ("Arial",25))
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"]),font = ("Arial",25,"bold"))

    def select_block(self,args,widget_tier,widget):
        """
        - vyvoláno levým klikem
        - Vkládá widget tier do vyhledávače
        - mění názvy tlačítek v závislosti na nakliknutém zařízení
        """
        self.current_block_id = str(widget_tier)
        if len(widget_tier) == 2: #01-99 stanice
            self.new_device.configure(state = tk.NORMAL)
            self.new_device.configure(text="Nová kamera")
            self.edit_device.configure(text="Editovat stanici")
            self.del_device.configure(text = "Odebrat stanici")
        elif len(widget_tier) == 4: # 0101-9999 kamery
            self.new_device.configure(state = tk.NORMAL)
            self.new_device.configure(text="Nová optika")
            self.edit_device.configure(text="Editovat kameru")
            self.del_device.configure(text = "Odebrat kameru")
        elif len(widget_tier) == 6: # 010101-999999 optika
            self.new_device.configure(text="",state = tk.DISABLED)
            self.edit_device.configure(text="Editovat optiku")
            self.del_device.configure(text = "Odebrat optiku")

        elif len(widget_tier) == 3: # c01-c99 kontrolery
            self.new_device.configure(state = tk.NORMAL)
            self.new_device.configure(text="Nové příslušenství")
            self.edit_device.configure(text="Editovat kontroler")
            self.del_device.configure(text = "Odebrat kontroler")
        elif len(widget_tier) == 5: # c0101-c9999 prislusenstvi kontroleru
            self.new_device.configure(text="",state = tk.DISABLED)
            self.edit_device.configure(text="Editovat příslušenství")
            self.del_device.configure(text = "Odebrat příslušenství")

        if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
            self.last_selected_widget.configure(border_color="#636363")
        self.last_selected_widget = widget
        widget.configure(border_color="white")

    def make_block(self,master_widget,height,width,fg_color,text,side,dummy_block = False,tier = ""):
        if dummy_block:
            dummy_block_widget =    customtkinter.CTkFrame(master=master_widget,corner_radius=0,height=height,width =width-10,fg_color="#212121")
            dummy_block_widget.     pack(pady = 0,padx =0,expand = False,side = side,anchor="w")
            return dummy_block
        else:
            block_widget =    customtkinter.CTkFrame(master=master_widget,corner_radius=0,fg_color=fg_color,height=height,width =width,border_width= 2,border_color="#636363")
            block_widget.     pack(pady = (0,0),padx =0,expand = False,side = side,anchor="w")
            block_name =      customtkinter.CTkLabel(master = block_widget,text = text,font=("Arial",25,"bold"),width=block_widget.cget("width")-15,height=block_widget.cget("height")-15,justify = "left",anchor="w")
            block_name.       pack(pady = 5,padx =5)
            # block_name.       pack(pady = 0,padx =0)
            block_widget.bind("<Button-3>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
            block_name.bind("<Button-3>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
            block_widget.bind("<Button-1>",lambda e, widget_tier=tier,widget = block_widget: self.select_block(e, widget_tier,widget))
            block_name.bind("<Button-1>",lambda e, widget_tier=tier,widget = block_widget: self.select_block(e, widget_tier,widget))
            return block_widget
        
    def make_new_object(self,which_one,object_to_edit = None,cam_index = None,optic_index = None):
        """
        which_one:
        - station
        - camera
        - optic
        - accessory
        """
        if which_one == "station":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            
            camera = {
                "type": "",
                "controller": "",
                "controller_color": "",
                "cable": "",
                "optics_list": [optic],
                "description": "",
            }
            station = {
                "name": "Název stanice",
                "inspection_description": "- popis inspekce",
                "camera_list": [camera],
            }

            return station
        
        elif which_one == "camera":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            camera = {
                "type": "",
                "controller": "",
                "controller_color": "",
                "cable": "",
                "optics_list": [optic],
                "description": "",
            }

            object_to_edit["camera_list"].append(camera)
            return object_to_edit
        
        elif which_one == "optic":
            optic = {
                "type": "",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }

            object_to_edit["camera_list"][cam_index]["optics_list"].append(optic)
            return object_to_edit
        
        elif which_one == "accessory":
            accessory = {
                "type": "",
                "description":"",
            }

            object_to_edit["accessory_list"].append(accessory)
            return object_to_edit

    def manage_widgets(self,args,widget_tier,btn,open_edit = True):
        if btn == "add_line": # nova stanice
            new_station = self.make_new_object("station")
            self.station_list.append(new_station)
            self.make_project_widgets()
            self.edit_object("",widget_tier,new_station=True)
            return
        
        if len(widget_tier) == 2: #01-99 stanice
            if btn == "add_object": # nova kamera ke stanici 0101-9999 kamery
                station_index = int(widget_tier[:2])
                station_with_new_camera = self.make_new_object("camera",object_to_edit = self.station_list[station_index])
                self.station_list[station_index] = station_with_new_camera
                self.make_project_widgets()
                if open_edit:
                    self.edit_object("",widget_tier,new_station=False)

        elif len(widget_tier) == 3: #c01-c99 kontrolery - tzn. nove prislusenstvi ke kontroleru
            if btn == "add_object":
                controller_index = int(widget_tier[1:3])
                controller_with_new_accessories = self.make_new_object("accessory",object_to_edit = self.controller_object_list[controller_index])
                self.controller_object_list[controller_index] = controller_with_new_accessories
                # self.make_project_widgets()
                if open_edit:
                    self.edit_object("",widget_tier,new_station=False)


        
        elif len(widget_tier) == 4: # 0101-9999 kamery, nove bude pridano: 010101-999999 optika
            if btn == "add_object": # nova optika kamery
                station_index = int(widget_tier[:2])
                camera_index = int(widget_tier[2:])
                camera_with_new_optics = self.make_new_object("optic",object_to_edit = self.station_list[station_index],cam_index = camera_index)
                self.station_list[station_index] = camera_with_new_optics
                self.make_project_widgets()
                if open_edit:
                    self.edit_object("",widget_tier,new_station=False)

        print(widget_tier)


        # print("STATION_LIST: ",self.station_list)
    
    def confirm_delete(self,to_del_object):
        def make_decision(decision):
            if decision == True:
                self.station_list.pop(to_del_object)
            self.make_project_widgets()
            close_window(child_root)

        def close_window(window):
            self.root.unbind("<Button-1>")
            # window.quit()
            window.destroy()
        
        # child_root=customtkinter.CTk()
        child_root = customtkinter.CTkToplevel()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"650x130+{x+80}+{y+80}")
        child_root.after(200, lambda: child_root.iconbitmap(app_icon_path))
        child_root.title("Upozornění")

        proceed_label = customtkinter.CTkLabel(master = child_root,text = "Opravdu si přejete odstranit celou stanici a všechna zařízení k ní připojená?",font=("Arial",18))
        button_yes =    customtkinter.CTkButton(master = child_root,text = "Pokračovat",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: make_decision(True))
        button_no =     customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  make_decision(False))

        proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top")
        button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")

        self.root.bind("<Button-1>",lambda e: close_window(child_root))
        child_root.update()
        child_root.update_idletasks()
        child_root.focus_force()
        child_root.focus()
        # child_root.mainloop()
        
    def delete_block(self,args,widget_tier):
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            print("deleting",self.station_list[station_index])
            self.confirm_delete(station_index)
            return
        
        elif len(widget_tier) == 4: # 0101-9999 kamery
            camera_index = int(widget_tier[2:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index])
            self.station_list[station_index]["camera_list"].pop(camera_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"].pop(optic_index)
            
        elif len(widget_tier) == 8: # 01010101-99999999 prislusenstvi
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:6])
            accessory_index = int(widget_tier[6:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"].pop(accessory_index)
        
        #refresh
        self.make_project_widgets()

    def edit_object_gui_new(self,object:str,station_index,camera_index = None,optics_index = None,accessory_index = None,controller_index = None,all_parameters = False):
        """
        Object:
        - station
        - camera
        - optics
        """

        def save_changes(no_window_shut = False):
            if object == "station" or all_parameters:
                self.station_list[station_index]["name"] = new_name.get()
                self.station_list[station_index]["inspection_description"] = new_description.get("1.0", tk.END)

            if object == "camera" or all_parameters:
                self.station_list[station_index]["camera_list"][camera_index]["type"] = camera_type_entry.get()
                self.station_list[station_index]["camera_list"][camera_index]["controller"] = controller_entry.get()
                current_controller = controller_entry.get()
                controller_index = None
                if str(current_controller).replace(" ","") != "":
                    for controllers in self.controller_object_list:
                        if (controllers["name"]+"("+controllers["type"]+")").replace(" ","") == str(current_controller).replace(" ",""):
                            controller_index = self.controller_object_list.index(controllers)
                self.station_list[station_index]["camera_list"][camera_index]["controller_index"] = controller_index
                self.station_list[station_index]["camera_list"][camera_index]["cable"] = cam_cable_menu.get()
                self.station_list[station_index]["camera_list"][camera_index]["description"] = notes_input.get("1.0", tk.END)
                
            if object == "optics" or "camera" or all_parameters:
                # Pokud je zadán manuálně, upřednostni
                if manual_optics_input.get().replace(" ","") != "":
                    optic_type = manual_optics_input.get()
                else:
                    optic_type = optic_type_entry.get()
                self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"] = optic_type
                self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"] = alternative_entry.get()
                self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"] = notes_input2.get("1.0", tk.END)

            self.make_project_widgets() #refresh
            if not no_window_shut:
                # child_root.grab_release()
                self.close_window(child_root)

        def next_station():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            camera_index = 0
            optics_index = 0
            accessory_index = 0
            station_index += 1
            if station_index < len(self.station_list):
                station_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                station_index += 1
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # TLACITKO +:
                station_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                close_window(child_root)
                if station_index < 10:
                    widget_tier = "0" + str(station_index)
                else:
                    widget_tier = str(station_index)
                self.manage_widgets("",widget_tier,"add_line")

        def previous_station():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            camera_index = 0
            optics_index = 0
            accessory_index = 0
            station_index -= 1
            if station_index > -1:
                station_index += 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                station_index -= 1
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                station_index += 1
            
        def next_camera():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            # optics_index = 0
            # accessory_index = 0
            camera_index += 1
            if camera_index < len(self.station_list[station_index]["camera_list"]):
                camera_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                camera_index += 1
                optics_index = 0
                accessory_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání

            else: # TLACITKO +:
                camera_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                camera_index += 1
                optics_index = 0
                accessory_index = 0
                if station_index < 10:
                    widget_tier_st = "0" + str(station_index)
                else:
                    widget_tier_st = str(station_index)

                print("camera st widget tier",widget_tier_st)
                self.manage_widgets("",widget_tier_st,"add_object",open_edit=False)
                initial_prefill() # prefill s novým indexem 

        def previous_camera():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            # optics_index = 0
            # accessory_index = 0
            camera_index -= 1
            if camera_index > -1:
                camera_index += 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                camera_index -= 1
                optics_index = 0
                accessory_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                camera_index += 1

        def next_optic():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            optics_index += 1
            if optics_index < len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"]):
                optics_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                optics_index += 1
                accessory_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání

            else: # TLACITKO +:
                optics_index -= 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                optics_index += 1
                
                if station_index < 10:
                    widget_tier_st = "0" + str(station_index)
                else:
                    widget_tier_st = str(station_index)
                if camera_index < 10:
                    widget_tier_cam = "0" + str(camera_index)
                else:
                    widget_tier_cam = str(camera_index)
                widget_tier = widget_tier_st + widget_tier_cam
                self.manage_widgets("",widget_tier,"add_object",open_edit=False)
                initial_prefill() # prefill s novým indexem 

        def previous_optic():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            optics_index -= 1
            if optics_index > -1:
                optics_index += 1
                save_changes(no_window_shut=True) # ulozit zmeny pri prepinani jeste u predesle stanice
                optics_index -= 1
                accessory_index = 0
                initial_prefill() # prefill s novým indexem - index se prenese i do ukládání
            else: # aby to neslo zase odznovu:
                optics_index += 1

        def close_window(child_root):
            try:
                if opened_subwindow.winfo_exists():
                    self.close_window(opened_subwindow)
            except Exception:
                pass
            
            self.root.unbind("<Button-1>")
            child_root.destroy()

        def callback_new_controller(new_controller_data):
            print("saving new controller: ",new_controller_data)
            new_controller = {
                "type": new_controller_data[0],
                "name": new_controller_data[1],
                "color": new_controller_data[2],
                "ip": new_controller_data[3],
                "username": new_controller_data[4],
                "password": new_controller_data[5],
                "accessory_list": new_controller_data[6],
                "notes": new_controller_data[7],
            }
            print("Nový kontroler------ ",new_controller)
            self.controller_object_list.append(new_controller)
            new_drop_option = f"{new_controller['name']} ({new_controller['type']})"
            self.custom_controller_drop_list.append(new_drop_option)
            controller_entry.configure(values = self.custom_controller_drop_list)
            controller_entry.set(new_drop_option)
            controller_opt_menu_color("",only_color=new_controller["color"])
            child_root.focus_force()

        def import_notes(which):
            """
            - camera (Kontroler, Kamera, Kabel)
            - optics (Objektiv, Alternativní)
            - accessory
            """
            if which == "camera":
                # current_controller = controller_entry.get()
                current_camera = camera_type_entry.get()
                current_cable = cam_cable_menu.get()
                notes_string = ""
                # if current_controller != "":
                #     controller_type = current_controller.split("(")[1]
                #     controller_notes = str(self.controller_notes_database[self.controller_database.index(controller_type[:-1])])
                #     if controller_notes != "":
                #         notes_string = notes_string + "Kontroler: " + controller_notes + "\n\n"
                if current_camera != "":
                    camera_notes = str(self.camera_notes_database[self.whole_camera_type_database.index(current_camera)])
                    if camera_notes != "":
                        notes_string = notes_string + "Kamera: " + camera_notes + "\n\n"
                if current_cable != "":
                    cable_notes = str(self.cable_notes_database[self.whole_camera_cable_database.index(current_cable)]) 
                    if cable_notes != "":
                        notes_string = notes_string + "Kabel: " + cable_notes + "\n\n"
                
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",notes_string)
            
            elif which == "optics":
                current_optics = optic_type_entry.get()
                current_alternative = alternative_entry.get()
                notes_string = ""
                if current_optics != "":
                    optic_notes = str(self.optics_notes_database[self.whole_optics_database.index(current_optics)])
                    if optic_notes !="":
                        notes_string = notes_string + "Objektiv: " + optic_notes + "\n\n"
                if current_alternative != "":
                    alternative_notes = str(self.optics_notes_database[self.whole_optics_database.index(current_alternative)])
                    if alternative_notes != "":
                        notes_string = notes_string + "Alternativní: " + alternative_notes + "\n\n"
                
                notes_input2.delete("1.0",tk.END)
                notes_input2.insert("1.0",notes_string)

        def call_new_controller_gui():
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_new_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            window.new_controller_window(child_root)

        def switch_database_section(operation,database,widget_menu,menu):
            """
            mění hodnotu pointeru na pole hodnot v option menu
            - vstupní hodnoty, menu:
                - camera_type
                - cable_type
                - optic
                - optic_alternative
            """
            if menu == "camera_type":
                if operation == "next":
                    self.camera_database_pointer +=1
                    if self.camera_database_pointer > len(database)-1:
                        self.camera_database_pointer = 0
                elif operation == "prev":
                    self.camera_database_pointer -=1
                    if self.camera_database_pointer < 0:
                        self.camera_database_pointer = len(database)-1
                
                widget_menu.configure(values = database[self.camera_database_pointer])
                widget_menu._open_dropdown_menu()

            elif menu == "cable_type":
                if operation == "next":
                    self.camera_cable_database_pointer +=1
                    if self.camera_cable_database_pointer > len(database)-1:
                        self.camera_cable_database_pointer = 0
                elif operation == "prev":
                    self.camera_cable_database_pointer -=1
                    if self.camera_cable_database_pointer < 0:
                        self.camera_cable_database_pointer = len(database)-1
            
                widget_menu.configure(values = database[self.camera_cable_database_pointer])
                widget_menu._open_dropdown_menu()
            
            elif menu == "optic":
                if operation == "next":
                    self.optics_database_pointer +=1
                    if self.optics_database_pointer > len(database)-1:
                        self.optics_database_pointer = 0
                elif operation == "prev":
                    self.optics_database_pointer -=1
                    if self.optics_database_pointer < 0:
                        self.optics_database_pointer = len(database)-1
            
                widget_menu.configure(values = database[self.optics_database_pointer])
                widget_menu._open_dropdown_menu()

        def controller_opt_menu_color(*args,only_color = False):
            if not only_color:
                current_controller = str(*args)
                if str(current_controller).replace(" ","") != "":
                    for controllers in self.controller_object_list:
                        if controllers["color"] != "":
                            if (controllers["name"]+"("+controllers["type"]+")").replace(" ","") == str(current_controller).replace(" ",""):
                                controller_entry.configure(fg_color = controllers["color"])
                                break
                else:
                    controller_entry.configure(fg_color = "#636363")
            else:
                controller_entry.configure(fg_color = str(only_color))

        child_root = customtkinter.CTkToplevel()
        # STANICE ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        station_frame =             customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)
        station_name_label =        customtkinter.CTkLabel(master = station_frame,text = "Název stanice:",font=("Arial",22,"bold"))
        name_frame =                customtkinter.CTkFrame(master = station_frame,corner_radius=0)
        button_prev_st =            customtkinter.CTkButton(master = name_frame,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_station())
        new_name =                  customtkinter.CTkEntry(master = name_frame,font=("Arial",22),width=300,height=50,corner_radius=0)
        button_next_st =            customtkinter.CTkButton(master = name_frame,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_station())
        button_prev_st              .pack(pady = 5, padx = 0,anchor="w",expand=False,side="left")
        new_name                    .pack(pady = 5, padx = 0,anchor="w",expand=True,side="left",fill="x")
        button_next_st              .pack(pady = 5, padx = 0,anchor="w",expand=False,side="left")
        inspection_description =    customtkinter.CTkLabel(master = station_frame,text = "Popis inspekce:",font=("Arial",22,"bold"))
        new_description =           customtkinter.CTkTextbox(master = station_frame,font=("Arial",22),width=300,height=220,corner_radius=0)
        station_name_label          .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
        name_frame                  .pack(pady = 5, padx = 5,anchor="w",expand=False,side="top",fill="x")
        inspection_description      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        new_description             .pack(pady = 5, padx = 10,expand=True,side="top",fill="both")

        # KAMERY ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        camera_frame =              customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)
        counter_frame_cam =         customtkinter.CTkFrame(master = camera_frame,corner_radius=0,fg_color="transparent")
        button_prev_cam =           customtkinter.CTkButton(master = counter_frame_cam,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_camera())
        counter_cam =               customtkinter.CTkLabel(master = counter_frame_cam,text = "0/0",font=("Arial",22,"bold"))
        button_next_cam =           customtkinter.CTkButton(master = counter_frame_cam,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_camera())
        button_prev_cam             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        counter_cam                 .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_cam             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")

        camera_type =               customtkinter.CTkLabel(master = camera_frame,text = "Typ kamery:",font=("Arial",22,"bold"))
        option_menu_frame_cam =     customtkinter.CTkFrame(master = camera_frame,corner_radius=0)
        camera_type_entry =         customtkinter.CTkOptionMenu(master = option_menu_frame_cam,font=("Arial",22),dropdown_font=("Arial",22),width = 300,height=50,values=self.camera_type_database[self.camera_database_pointer],corner_radius=0)
        button_prev_section_cam =   customtkinter.CTkButton(master = option_menu_frame_cam,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                            command=lambda: switch_database_section("prev",self.camera_type_database,camera_type_entry,"camera_type"))
        button_next_section_cam =   customtkinter.CTkButton(master = option_menu_frame_cam,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                            command=lambda: switch_database_section("next",self.camera_type_database,camera_type_entry,"camera_type"))
        camera_type_entry           .pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left",fill="x")
        button_prev_section_cam     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_section_cam     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")

        cam_cable =                   customtkinter.CTkLabel(master = camera_frame,text = "Kabel ke kameře:",font=("Arial",22,"bold"))
        option_menu_frame_cable =     customtkinter.CTkFrame(master = camera_frame,corner_radius=0)
        cam_cable_menu =              customtkinter.CTkOptionMenu(master = option_menu_frame_cable,font=("Arial",22),dropdown_font=("Arial",22),width = 300,height=50,values=self.camera_cable_database[self.camera_cable_database_pointer],corner_radius=0)
        button_prev_section_cable =   customtkinter.CTkButton(master = option_menu_frame_cable,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("prev",self.camera_cable_database,cam_cable_menu,"cable_type"))
        button_next_section_cable =   customtkinter.CTkButton(master = option_menu_frame_cable,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("next",self.camera_cable_database,cam_cable_menu,"cable_type"))
        cam_cable_menu                .pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left",fill="x")
        button_prev_section_cable     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_section_cable     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")

        controller =                customtkinter.CTkLabel(master = camera_frame,text = "Kontroler:",font=("Arial",22,"bold"))
        controller_frame =          customtkinter.CTkFrame(master = camera_frame,corner_radius=0)
        controller_entry =          customtkinter.CTkOptionMenu(master = controller_frame,font=("Arial",22),dropdown_font=("Arial",22),width=280,height=50,values=self.custom_controller_drop_list,corner_radius=0,fg_color="#212121",command=controller_opt_menu_color)
        new_controller =            customtkinter.CTkButton(master = controller_frame,text = "Přidat",font=("Arial",22,"bold"),width = 80,height=50,corner_radius=0,command=lambda: call_new_controller_gui())
        controller_entry.           pack(pady = 5, padx = (10,0),anchor="w",expand=True,side="left",fill="x")
        new_controller.             pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        note_label_frame =          customtkinter.CTkFrame(master = camera_frame,corner_radius=0)
        note_label =                customtkinter.CTkLabel(master = note_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes_btn =          customtkinter.CTkButton(master = note_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("camera"))
        note_label.                 pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes_btn.           pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        # notes_input_controller =    customtkinter.CTkTextbox(master = camera_frame,font=("Arial",22),height=100,corner_radius=0)
        notes_input =               customtkinter.CTkTextbox(master = camera_frame,font=("Arial",22),corner_radius=0)
        counter_frame_cam           .pack(pady=(10,0),padx= 3,anchor="n",expand=False,side="top")
        camera_type                 .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_cam       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        cam_cable                   .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_cable     .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        controller                  .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        controller_frame            .pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        new_controller              .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        note_label_frame            .pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        notes_input                 .pack(pady = (5,0), padx = 10,expand=True,side="top",fill="both")

        # OPTIKA --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        if "" in self.optics_database:
            self.optics_database.pop(self.optics_database.index(""))
        optics_frame =               customtkinter.CTkFrame(master = child_root,corner_radius=0,border_width=3)

        counter_frame_optics =      customtkinter.CTkFrame(master = optics_frame,corner_radius=0,fg_color="transparent")
        button_prev_opt =           customtkinter.CTkButton(master = counter_frame_optics,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: previous_optic())
        counter_opt =               customtkinter.CTkLabel(master = counter_frame_optics,text = "0/0",font=("Arial",22,"bold"))
        button_next_opt =           customtkinter.CTkButton(master = counter_frame_optics,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,command=lambda: next_optic())
        button_prev_opt             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        counter_opt                 .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_opt             .pack(pady = 0, padx = (5,0),anchor="w",expand=False,side="left")

        optic_type =                 customtkinter.CTkLabel(master = optics_frame,text = "Typ objektivu:",font=("Arial",22,"bold"))
        option_menu_frame_optic =    customtkinter.CTkFrame(master = optics_frame,corner_radius=0)
        optic_type_entry =           customtkinter.CTkOptionMenu(master = option_menu_frame_optic,font=("Arial",22),dropdown_font=("Arial",22),width=300,height=50,values=self.optics_database[self.optics_database_pointer],corner_radius=0)
        button_prev_section_optic =  customtkinter.CTkButton(master = option_menu_frame_optic,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("prev",self.optics_database,optic_type_entry,"optic"))
        button_next_section_optic =  customtkinter.CTkButton(master = option_menu_frame_optic,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("next",self.optics_database,optic_type_entry,"optic"))
        optic_type_entry             .pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left",fill="x")
        button_prev_section_optic    .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_section_optic    .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        manual_optics_input =        customtkinter.CTkEntry(master = optics_frame,font=("Arial",22),width=305,height=50,corner_radius=0,placeholder_text="manuálně")

        alternative_type =                  customtkinter.CTkLabel(master = optics_frame,text = "Alternativa:",font=("Arial",22,"bold"))
        option_menu_frame_alternative =     customtkinter.CTkFrame(master = optics_frame,corner_radius=0)
        alternative_entry =                 customtkinter.CTkOptionMenu(master = option_menu_frame_alternative,font=("Arial",22),dropdown_font=("Arial",22),width=300,height=50,values=self.optics_database[self.optics_database_pointer],corner_radius=0)
        button_prev_section_alternative =   customtkinter.CTkButton(master = option_menu_frame_alternative,text = "<",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("prev",self.optics_database,alternative_entry,"optic"))
        button_next_section_alternative =   customtkinter.CTkButton(master = option_menu_frame_alternative,text = ">",font=("Arial",22,"bold"),width = 30,height=50,corner_radius=0,
                                                              command=lambda: switch_database_section("next",self.optics_database,alternative_entry,"optic"))
        alternative_entry                   .pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left",fill="x")
        button_prev_section_alternative     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_next_section_alternative     .pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        
        note2_label_frame =                  customtkinter.CTkFrame(master = optics_frame,corner_radius=0)
        note2_label =                        customtkinter.CTkLabel(master = note2_label_frame,text = "Poznámky:",font=("Arial",22,"bold"))
        import_notes2_btn =                  customtkinter.CTkButton(master = note2_label_frame,text = "Import z databáze",font=("Arial",22,"bold"),width = 100,height=30,corner_radius=0,command=lambda: import_notes("optics"))
        note2_label.                         pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        import_notes2_btn.                   pack(pady = 5, padx = (10,0),anchor="w",expand=False,side="left")
        notes_input2 =                      customtkinter.CTkTextbox(master = optics_frame,font=("Arial",22),width=300,height=200,corner_radius=0)
        counter_frame_optics                .pack(pady=(10,0),padx=3,anchor="n",expand=False,side = "top")
        optic_type                          .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_optic             .pack(pady = (5,0), padx = 10,anchor="w",expand=False,side="top",fill="x")
        manual_optics_input                 .pack(pady = 0, padx = 15,anchor="w",expand=False,side="top",fill="x")
        alternative_type                    .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
        option_menu_frame_alternative       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top",fill="x")
        note2_label_frame                   .pack(pady = 0, padx = 3,anchor="w",expand=False,side="top",fill="x")
        notes_input2                        .pack(pady = 5, padx = 10,expand=True,side="top",fill="both")
        
        def refresh_counters():
            nonlocal station_index
            nonlocal optics_index
            nonlocal camera_index
            nonlocal counter_cam
            nonlocal counter_opt

            try:
                counter_cam_state = str(camera_index+1) + "/" + str(len(self.station_list[station_index]["camera_list"]))
                counter_cam.configure(text = counter_cam_state)
            except Exception:
                pass
            try:
                counter_opt_state = str(optics_index+1) + "/" + str(len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"]))
                counter_opt.configure(text = counter_opt_state)
            except Exception:
                pass

        def refresh_button_appearance():
            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal button_prev_st
            nonlocal button_next_st
            nonlocal button_prev_cam
            nonlocal button_next_cam
            nonlocal button_prev_opt
            nonlocal button_next_opt

            def config_buttons(button_left,button_right,index,max_array_value):
                if index ==0:
                    button_left.configure(text = "",fg_color = "#636363")
                else:
                    button_left.configure(text = "<",fg_color = "#636363")

                if index == max_array_value:
                    button_right.configure(text = "+",fg_color = "green")
                else:
                    button_right.configure(text = ">",fg_color = "#636363")

            config_buttons(button_prev_st,button_next_st,station_index,len(self.station_list)-1)
            config_buttons(button_prev_cam,button_next_cam,camera_index,len(self.station_list[station_index]["camera_list"])-1)
            config_buttons(button_prev_opt,button_next_opt,optics_index,len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"])-1)

        def initial_prefill():
            def filter_text_input(text):
                """
                - removes extra new empty lines
                """
                legit_rows = []
                legit_notes = ""
                rows = text.split("\n")
                for i in range(0,len(rows)):
                    if rows[i].replace(" ","") != "":
                        legit_rows.append(rows[i])

                for i in range(0,len(legit_rows)): 
                    if i == len(legit_rows)-1:
                        legit_notes = legit_notes + legit_rows[i]
                    else:
                        legit_notes = legit_notes + legit_rows[i]+ "\n"
                return legit_notes

            nonlocal station_index
            nonlocal camera_index
            nonlocal optics_index
            nonlocal accessory_index
            new_name.delete(0,300)
            new_name.insert(0,str(self.station_list[station_index]["name"]))
            new_description.delete("1.0",tk.END)
            new_description.insert("1.0",filter_text_input(str(self.station_list[station_index]["inspection_description"])))
            # initial prefill - camera:
            try:
                if str(self.station_list[station_index]["camera_list"][camera_index]["type"]) in self.whole_camera_type_database:
                    camera_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["type"]))
                if str(self.station_list[station_index]["camera_list"][camera_index]["controller"]) in self.custom_controller_drop_list:
                    controller_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["controller"]))
                if str(self.station_list[station_index]["camera_list"][camera_index]["cable"]) in self.whole_camera_cable_database:
                    cam_cable_menu.set(str(self.station_list[station_index]["camera_list"][camera_index]["cable"]))
                
                notes_input.delete("1.0",tk.END)
                notes_input.insert("1.0",filter_text_input(str(self.station_list[station_index]["camera_list"][camera_index]["description"])))
            except TypeError:
                camera_index = 0
                if len(self.station_list[station_index]["camera_list"]) > 0:
                    if str(self.station_list[station_index]["camera_list"][camera_index]["type"]) in self.whole_camera_type_database:
                        camera_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["type"]))
                    if str(self.station_list[station_index]["camera_list"][camera_index]["controller"]) in self.custom_controller_drop_list:
                        controller_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["controller"]))
                    if str(self.station_list[station_index]["camera_list"][camera_index]["cable"]) in self.whole_camera_cable_database:
                        cam_cable_menu.set(str(self.station_list[station_index]["camera_list"][camera_index]["cable"]))

                    notes_input.delete("1.0",tk.END)
                    notes_input.insert("1.0",filter_text_input(str(self.station_list[station_index]["camera_list"][camera_index]["description"])))

            # initial prefill - optics:
            try:
                if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]) not in self.whole_optics_database:
                    manual_optics_input.delete(0,300)
                    manual_optics_input.insert(0,str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
                if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]) in self.whole_optics_database:
                    optic_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
                if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]) in self.whole_optics_database:
                    alternative_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]))
                else:
                    alternative_entry.set("")
                notes_input2.delete("1.0",tk.END)
                notes_input2.insert("1.0",filter_text_input(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"])))
            except TypeError:
                optics_index = 0
                if len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"]) > 0:
                    if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]) not in self.whole_optics_database:
                        manual_optics_input.delete(0,300)
                        manual_optics_input.insert(0,str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
                    if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]) in self.whole_optics_database:
                        optic_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
                    if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]) in self.whole_optics_database:
                        alternative_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]))
                    else:
                        alternative_entry.set("")
                    notes_input2.delete("1.0",tk.END)
                    notes_input2.insert("1.0",filter_text_input(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"])))

            refresh_counters()
            refresh_button_appearance()
            
        initial_prefill()
        controller_opt_menu_color(controller_entry.get())

        button_frame =  customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_frame    .pack(pady = 0, padx = 0,fill="x",anchor="s",expand=False,side="bottom")
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        one_segment_width = 450
        height = 850
        child_root.after(200, lambda: child_root.iconbitmap(app_icon_path))

        if object == "station":
            # child_root.geometry(f"420x450+{x+80}+{y+80}")
            width = 3*one_segment_width
            child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování stanice: " + str(self.station_list[station_index]["name"]))
            station_frame   .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            camera_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        elif object == "camera":
            width = 2*one_segment_width
            child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování kamery: " + str(self.station_list[station_index]["camera_list"][camera_index]["type"]))
            camera_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        elif object == "optics":
            width = one_segment_width
            child_root.geometry(f"{width}x{height}+{x+100}+{y+30}")
            child_root.title("Editování optiky: " + str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
            optics_frame    .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left",ipady = 3,ipadx = 3)

        button_save =   customtkinter.CTkButton(master = button_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: save_changes())
        button_exit =   customtkinter.CTkButton(master = button_frame,text = "Zavřít",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(child_root))
        button_save     .pack(pady = 10, padx = 10,anchor="e",expand=False,side="right")
        button_exit     .pack(pady = 10, padx = 10,anchor="e",expand=True,side="right")

        if self.default_subwindow_status == 1:
            child_root.state('zoomed')
        child_root.update()
        child_root.update_idletasks()
        child_root.focus_force()
        child_root.focus()

    def edit_object(self,args,widget_tier,new_station = False):
        def callback_edited_controller(new_controller_data):
            nonlocal controller_index
            self.controller_object_list[controller_index]["type"] = new_controller_data[0]
            self.controller_object_list[controller_index]["name"] = new_controller_data[1]
            self.controller_object_list[controller_index]["color"] = new_controller_data[2]
            self.controller_object_list[controller_index]["ip"] = new_controller_data[3]
            self.controller_object_list[controller_index]["username"] = new_controller_data[4]
            self.controller_object_list[controller_index]["password"] = new_controller_data[5]
            self.controller_object_list[controller_index]["accessory_list"] = new_controller_data[6]
            self.controller_object_list[controller_index]["notes"] = new_controller_data[7]
            # refresh dropdownlist:
            self.custom_controller_drop_list = []
            for controllers in self.controller_object_list:
                new_drop_option = f"{controllers['name']} ({controllers['type']})"
                self.custom_controller_drop_list.append(new_drop_option)
            # self.make_controller_widgets()
            self.make_project_widgets()
            self.root.focus_force()

        if len(widget_tier) == 2: #01-99 stanice
            station_index = int(widget_tier[:2])
            print("editing",self.station_list[station_index])
            # self.edit_object_gui_new("station",station_index,all_parameters=True)
            if new_station:
                self.edit_object_gui_new("station",(len(self.station_list)-1),all_parameters=True)
            else:
                # kdyz nova kamera, chci rovnou editovat tu nově přidanou
                current_cam_count = len(self.station_list[station_index]["camera_list"])
                camera_index = 0
                if current_cam_count > 0:
                    camera_index = current_cam_count-1
                self.edit_object_gui_new("station",station_index,camera_index,all_parameters=True)

        elif len(widget_tier) == 4: # 0101-9999 kamery
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:])
            print("editing",self.station_list[station_index]["camera_list"][camera_index])
            # kdyz edituju, chci prejit na posledni optiku...
            current_optics_count = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"])
            optic_index = 0
            if current_optics_count > 0:
                optic_index = current_optics_count-1
            self.edit_object_gui_new("camera",station_index,camera_index,optic_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            station_index = int(widget_tier[:2])
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("editing",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.edit_object_gui_new("optics",station_index,camera_index,optic_index)
        
        elif len(widget_tier) == 3: # c01-c99 kontolery
            controller_index = int(widget_tier[1:3])
            current_acc_count = len(self.controller_object_list[controller_index]["accessory_list"])
            accessory_index = 0
            if current_acc_count > 0:
                accessory_index = current_acc_count-1
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_edited_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            window.new_controller_window(childroot=None,controller=self.controller_object_list[controller_index],edit=True,accessory_index=accessory_index)

        elif len(widget_tier) == 5: # c0101-c9999 prislusenstvi kontoleru
            controller_index = int(widget_tier[1:3])
            accessory_index = int(widget_tier[3:5])
            window = ToplevelWindow(self.root,[self.controller_database,self.controller_notes_database],callback_edited_controller,self.controller_object_list,[self.accessory_database,self.whole_accessory_database,self.accessory_notes_database])
            window.new_controller_window(childroot=None,controller=self.controller_object_list[controller_index],edit=True,accessory_index = accessory_index,only_accessory=True)
        
        #refresh
        self.make_project_widgets()

    def export_to_excel(self,path_with_name):
        save_excel_class = Save_excel(self.root,station_list = self.station_list,project_name = self.project_name_input.get(),excel_name=path_with_name,controller_list=self.controller_object_list,console = self.main_console)
        output = save_excel_class.main()
        return

    def export_option_window(self):
        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(app_icon_path))
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"1000x350+{x+200}+{y+100}")
        child_root.title("Možnosti exportování souboru")

        def get_excel_path():
            nonlocal export_path
            nonlocal export_name
            nonlocal format_entry
            name_inserted = export_name.get()
            path_inserted = export_path.get()
            path_inserted = resource_path(path_inserted)
            if path_inserted.replace(" ","") == "":
                return None
            else:
                print("Cesta pro export: ",path_inserted + name_inserted + "." + format_entry.get())
                return path_inserted + name_inserted + "." + format_entry.get()

        def call_save_file(child_root):
            nonlocal console
            nonlocal export_path
            path_inserted = export_path.get()
            if os.path.exists(path_inserted):
                excel_path_with_name = get_excel_path()
                if os.path.exists(excel_path_with_name): # kontrola souboru se stejným názvem
                    nonlocal click_count
                    nonlocal previous_path
                    click_count += 1
                    add_colored_line(console,f"Cesta již obsahuje soubor se stejným názvem, při druhém kliknutí na \"Uložit\" bude přepsán","orange",None,True)
                    if click_count > 1 and previous_path == excel_path_with_name: # když podruhé a nebyla změněna cesta
                        self.favourite_format = str(format_entry.get())
                        self.last_path_input = path_inserted
                        self.export_to_excel(excel_path_with_name)
                        close_window(child_root)
                    elif click_count > 1 and previous_path != excel_path_with_name:
                        click_count =1
                    previous_path = excel_path_with_name
                else:
                    self.favourite_format = str(format_entry.get())
                    self.last_path_input = path_inserted
                    self.export_to_excel(excel_path_with_name)
                    close_window(child_root)
            else:
                add_colored_line(console,"Zadaná cesta pro uložení je neplatná","red",None,True)

        def close_window(child_root):
            try:
                self.root.unbind("<Button-1>")
            except Exception:
                pass
            # child_root.quit()
            child_root.destroy()

        def call_browse_directories():
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            output = browseDirectories("only_dirs")
            if str(output[1]) != "/":
                export_path.delete(0,300)
                export_path.insert(0, str(output[1]))
                add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            print(output[0])
            # child_root.grab_set()
            # child_root.update()
            # child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()

        def save_current_path():
            path_inserted = str(export_path.get())
            if path_inserted.replace(" ","") != "":
                checked_path = path_check(path_inserted)
                if checked_path != False:
                    self.last_path_input = path_inserted
                    self.path_for_callback = path_inserted

        click_count = 0
        previous_path = ""
        export_frame =      customtkinter.CTkFrame(master = child_root,corner_radius=0)
        export_label =      customtkinter.CTkLabel(master = export_frame,text = "Zadejte název souboru:",font=("Arial",22,"bold"))
        export_name_frame = customtkinter.CTkFrame(master = export_frame,corner_radius=0)
        export_name =       customtkinter.CTkEntry(master = export_name_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        format_entry =      customtkinter.CTkOptionMenu(master = export_name_frame,font=("Arial",22),dropdown_font=("Arial",22),width=200,height=50,values=self.format_list,corner_radius=0)
        export_name         .pack(pady = 5, padx = (10,0),anchor="w",fill="x",expand=True,side="left")
        format_entry        .pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        export_label2 =      customtkinter.CTkLabel(master = export_frame,text = "Zadejte cestu, kam soubor uložit:",font=("Arial",22,"bold"))
        export_path_frame = customtkinter.CTkFrame(master = export_frame,corner_radius=0)
        export_path =       customtkinter.CTkEntry(master = export_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        save_path_btn =     customtkinter.CTkButton(master = export_path_frame,text = "💾",font=("",22),width = 50,height=50,corner_radius=0,command=lambda: save_current_path())
        explorer_btn =      customtkinter.CTkButton(master = export_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories())
        export_path         .pack(pady = 5, padx = (10,0),anchor="w",fill="x",expand=True,side="left")
        save_path_btn       .pack(pady = 5, padx = 10,anchor="e",expand=False,side="right")
        explorer_btn        .pack(pady = 5, padx = (10,0),anchor="e",expand=False,side="right")
        console =           tk.Text(export_frame, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
        button_save =       customtkinter.CTkButton(master = export_frame,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: call_save_file(child_root))
        button_exit =       customtkinter.CTkButton(master = export_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(child_root))

        export_frame        .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        export_label        .pack(pady=(15,5),padx=10,anchor="w",expand=False,side="top")
        export_name_frame   .pack(expand=True,side="top",anchor="n",fill="x")
        export_label2       .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        export_path_frame   .pack(expand=True,side="top",anchor="n",fill="x")
        console             .pack(expand=True,side="top",anchor="n",fill="x")
        button_save         .pack(pady = 10, padx = 10,expand=False,side="right",anchor = "e")
        button_exit         .pack(pady = 10, padx = 10,expand=True,side="right",anchor = "e")

        excel_filename = self.default_excel_filename
        if str(self.project_name_input.get().replace(" ","")) != "":
            excel_filename = self.default_excel_filename + "_projekt_" + str(self.project_name_input.get())
        export_name.insert("0",excel_filename)

        default_path = path_check(self.default_path)
        checked_last_path = path_check(self.last_path_input)
        if checked_last_path != False and checked_last_path != None and checked_last_path.replace(" ","") != "" and checked_last_path.replace(" ","") != "/":
            initial_path = resource_path(checked_last_path)
            add_colored_line(console,"Byla vložena poslední zvolená cesta","green",None,True)
        elif default_path != False and default_path != None and default_path.replace(" ","") != "" and default_path.replace(" ","") != "/":
            initial_path = resource_path(default_path)
            add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)
        else:
            initial_path = resource_path(path_check(os.getcwd()))

        export_path.insert("0",str(initial_path))
        format_entry.set(self.favourite_format)

        self.root.bind("<Button-1>",lambda e: close_window(child_root))
        child_root.update()
        child_root.update_idletasks()
        child_root.focus()
        child_root.focus_force()
    
    def load_metadata_callback(self,input_data):
        self.station_list = input_data[0]
        print("loaded station list: ",self.station_list)

        self.controller_object_list = input_data[1]
        if str(input_data[2]) != "None":
            self.project_name_input.delete(0,300)
            self.project_name_input.insert(0,str(input_data[2]))

        self.custom_controller_drop_list = [""]
        for controllers in self.controller_object_list:
            new_drop_option = f"{controllers['name']} ({controllers['type']})"
            self.custom_controller_drop_list.append(new_drop_option)

        self.make_project_widgets()

    def call_save_metadata_gui(self):
        def callback_save_last_input(filename,path_inserted,path_to_save=None):
            if filename != None:
                self.last_xml_filename = filename
            if path_inserted != None:
                self.last_path_input = path_inserted
            if path_to_save != None:
                self.last_path_input = path_to_save
                self.path_for_callback = path_to_save

        window = ToplevelWindow(self.root,custom_controller_database=self.controller_object_list)
        window.save_prog_options_window(self.main_console,self.station_list,self.project_name_input.get(),self.load_metadata_callback,callback_save_last_input,self.last_xml_filename,self.last_path_input,self.default_xml_file_name,self.default_path)

    def create_main_widgets(self):
        def call_manage_widgets(button):
            widget_tier = ""
            widget_tier = self.current_block_id
            if button == "add_line":
                if widget_tier != "":
                    if len(widget_tier) > 2: # pokud je nakliknuteho neco jiného než stanice - přidej novou pod poslední
                        next_st_widget_tier = len(self.station_list)
                        if next_st_widget_tier < 10:
                            next_st_widget_tier = "0" + str(next_st_widget_tier)
                        self.current_block_id = str(next_st_widget_tier)
                        self.manage_widgets("",str(next_st_widget_tier),btn=button)
                        return

                    self.manage_widgets("",widget_tier,btn=button)
                    return
                
            elif widget_tier != "":
                self.manage_widgets("",widget_tier,btn=button)
                return
            
            add_colored_line(self.main_console,f"Nejprve zvolte pro co zařízení přidat","red",None,True)
        
        def call_edit_object():
            widget_tier = ""
            widget_tier = self.current_block_id
            if widget_tier != "":
                self.edit_object("",widget_tier)
            else:
                add_colored_line(self.main_console,f"Nejprve zvolte zařízení pro editaci","red",None,True)

        def call_delete_object():
            widget_tier = ""
            widget_tier = self.current_block_id
            if widget_tier != "":
                self.delete_block("",widget_tier)
            else:
                add_colored_line(self.main_console,f"Nejprve zvolte zařízení pro odebrání","red",None,True)

        def switch_manufacturer():
            if self.chosen_manufacturer == "Omron":
                manufacturer_logo =             customtkinter.CTkImage(PILImage.open(resource_path("images/keyence_logo.png")),size=(240, 50))
                self.chosen_manufacturer = "Keyence"
                self.camera_database_pointer = 0
                self.optics_database_pointer = 0
                self.camera_cable_database_pointer = 0
                self.accessory_database_pointer = 0
                switch_manufacturer_image.configure(image = manufacturer_logo)
                self.read_database()
            elif self.chosen_manufacturer == "Keyence":
                manufacturer_logo =             customtkinter.CTkImage(PILImage.open(resource_path("images/omron_logo.png")),size=(240, 50))
                self.chosen_manufacturer = "Omron"
                self.camera_database_pointer = 0
                self.optics_database_pointer = 0
                self.camera_cable_database_pointer = 0
                self.accessory_database_pointer = 0
                switch_manufacturer_image.configure(image = manufacturer_logo)
                self.read_database()

        def call_setting_window():
            def apply_changes_callback(input_data):
                if input_data[0] != "":
                    self.default_excel_filename = input_data[0]
                if input_data[1] != "":
                    self.default_xml_file_name = input_data[1]
                if input_data[2] != "":
                    self.default_subwindow_status = input_data[2]
                if input_data[3] != "":
                    self.default_database_filename = input_data[3]
            window = ToplevelWindow(self.root)
            window.setting_window(self.default_excel_filename,self.default_xml_file_name,self.default_subwindow_status,apply_changes_callback,self.default_database_filename)

        self.clear_frame(self.root)
        main_header =               customtkinter.CTkFrame(master=self.root,corner_radius=0,height=100)
        console_frame=              customtkinter.CTkFrame(master=self.root,corner_radius=0,height=50)
        main_header_row0 =          customtkinter.CTkFrame(master=main_header,corner_radius=0,height=100,fg_color="#636363")
        main_header                 .pack(pady=0,padx=5,expand=False,fill="x",side = "top",ipady = 10,ipadx = 10,anchor="w")
        main_header_row0            .pack(pady=0,padx=0,expand=True,fill="x",side = "top",anchor="w")
        image_frame =               customtkinter.CTkFrame(master=main_header,corner_radius=0,height=100,fg_color="#212121")
        image_frame                 .pack(pady=0,padx=0,expand=False,side = "right",anchor="e",ipady = 10,ipadx = 10)
        logo =                      customtkinter.CTkImage(PILImage.open(resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =                customtkinter.CTkLabel(master = image_frame,text = "",image =logo,bg_color="#212121")
        image_logo                  .pack(pady=0,padx=0,expand=True)
        main_header_row1 =          customtkinter.CTkFrame(master=main_header,corner_radius=0,height=100,fg_color="#212121")
        main_header_row2 =          customtkinter.CTkFrame(master=main_header,corner_radius=0,height=100,fg_color="#212121")
        main_header_row1            .pack(pady=0,padx=0,expand=True,fill="x",side = "top",anchor="w")
        main_header_row2            .pack(pady=(5,0),padx=0,expand=True,fill="x",side = "top",anchor="w")
        console_frame               .pack(pady=0,padx=0,fill="x",expand=False,side = "top")

        main_menu_button =          customtkinter.CTkButton(master = main_header_row0, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        main_menu_button            .pack(pady = (10,0),padx =(20,0),anchor = "s",side = "left")

        # self.search_input =         customtkinter.CTkEntry(master = main_header,font=("Arial",20),width=250,height=50,placeholder_text="Zvolený blok",corner_radius=0)
        self.project_name_input =   customtkinter.CTkEntry(master = main_header_row1,font=("Arial",20),width=250,height=50,placeholder_text="Název projektu",corner_radius=0)
        new_station =               customtkinter.CTkButton(master = main_header_row1,text = "Nová stanice",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command= lambda: call_manage_widgets("add_line"))
        self.new_device =           customtkinter.CTkButton(master = main_header_row1,text = "Nová kamera",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command= lambda: call_manage_widgets("add_object"))
        self.edit_device =          customtkinter.CTkButton(master = main_header_row1,text = "Editovat stanici",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command= lambda: call_edit_object())
        self.del_device =           customtkinter.CTkButton(master = main_header_row1,text = "Odebrat stanici",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command= lambda: call_delete_object())

        # self.search_input           .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        self.project_name_input     .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        new_station                 .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        self.new_device             .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        self.edit_device            .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        self.del_device             .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")

        export_button =                 customtkinter.CTkButton(master = main_header_row2,text = "Exportovat",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command=lambda:self.export_option_window())
        switch_manufacturer_frame =     customtkinter.CTkFrame(master = main_header_row2,corner_radius=0)
        switch_manufacturer_btn =       customtkinter.CTkButton(master=switch_manufacturer_frame,text="Změnit výrobce:",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,command=lambda:switch_manufacturer())
        manufacturer_logo =             customtkinter.CTkImage(PILImage.open(resource_path("images/omron_logo.png")),size=(240, 50))
        switch_manufacturer_image =     customtkinter.CTkLabel(master = switch_manufacturer_frame,text = "",image=manufacturer_logo)
        save_button =                   customtkinter.CTkButton(master = main_header_row2,text = "Uložit/ Nahrát",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,
                                                               command=lambda:self.call_save_metadata_gui())
        button_settings =               customtkinter.CTkButton(master = main_header_row2, width = 50,height=50,text="⚙️",command =  lambda: call_setting_window(),font=("",22),corner_radius=0)
        
        switch_manufacturer_btn         .pack(pady = 0, padx = 0,anchor="w",side="left")
        switch_manufacturer_image       .pack(pady = 0, padx = (10,0),anchor="w",side="left")
        export_button                   .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        switch_manufacturer_frame       .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        save_button                     .pack(pady = 0, padx = (20,0),anchor="w",expand=False,side="left")
        button_settings                 .pack(pady = 0, padx = (10,0),anchor="w",expand=False,side="left")
        self.main_console =             tk.Text(console_frame, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
        self.main_console               .pack(pady = 10, padx = (10,0),anchor="w",expand=True,side="bottom")
        # self.search_input.insert("0","00")

        column_labels =             customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        self.project_tree =         customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        column_labels               .pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree           .pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        stations_column_header =    customtkinter.CTkLabel(master = column_labels,text = "Stanice",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        camera_column_header =      customtkinter.CTkLabel(master = column_labels,text = "Kamera",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        optics_column_header =      customtkinter.CTkLabel(master = column_labels,text = "Objektiv",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        controller_column_header =  customtkinter.CTkLabel(master = column_labels,text = "Kontrolery",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        accessory_column_header =   customtkinter.CTkLabel(master = column_labels,text = "Příslušenství",font=("Arial",25,"bold"),bg_color="#212121",width=self.default_block_width-35,height=50)
        stations_column_header      .pack(pady=(15,0),padx=15,expand=False,side = "left")
        camera_column_header        .pack(pady=(15,0),padx=15,expand=False,side = "left")
        optics_column_header        .pack(pady=(15,0),padx=15,expand=False,side = "left")
        controller_column_header    .pack(pady=(15,0),padx=15,expand=False,side = "left")
        accessory_column_header     .pack(pady=(15,0),padx=15,expand=False,side = "left")
        
        self.project_column =   customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)
        self.camera_column =    customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)    
        self.optic_column =     customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)    
        self.controller_column = customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)
        self.accessory_column =  customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)
        self.project_column     .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.camera_column      .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.optic_column       .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.controller_column  .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.accessory_column   .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        # self.station_list.append(self.make_new_object("station"))
        self.make_project_widgets()
        add_colored_line(self.main_console,self.download_database_console_input[0],self.download_database_console_input[1],None,True)
        
        def unfocus_entry(e):
            self.root.focus_set()
        # self.root.bind("<Button-1>",lambda e:unfocus_entry(e))
        self.project_name_input.bind("<Leave>",lambda e:unfocus_entry(e))
        
        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1600:
                self.root.state('normal')
                self.root.geometry("1600x900")
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')

        self.root.bind("<f>",lambda e: maximalize_window(e))
        self.root.mainloop()
    
    def check_widget_growth(self,widget:str,station_index=None,camera_index=None,optics_index=None,controller_index = None):
        """
        widget:
        - station
        - camera
        - optics xxxxxx
        - controller
        """
        station_widget_row_growth = 0
        camera_widget_row_growth = 0
        default_widget_height = 50

        if widget == "station":
            number_of_blocks = 0

            for camera in self.station_list[station_index]["camera_list"]:
                station_optics_count = 0
                accessory_count = 0

                # station_optic_count += len(camera["optics_list"])
                station_optics_count = len(camera["optics_list"])
                try:
                    try:
                        controller_index = int(camera["controller_index"])
                    except Exception:
                        controller_index = camera["controller_index"]

                    accessory_count = len(self.controller_object_list[controller_index]["accessory_list"])
                except Exception as e:
                    # neni prirazen kontroler - index je v tomto pripade None
                    pass
                if accessory_count > station_optics_count:
                    number_of_blocks += accessory_count
                else:
                    number_of_blocks += station_optics_count

            if number_of_blocks>0:
                station_widget_row_growth = ((number_of_blocks*default_widget_height)-default_widget_height)    

            self.station_list[station_index]["row_count"] = number_of_blocks
            return station_widget_row_growth
            
        elif widget == "camera":
            accessory_count = 0
            optics_count = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"])
            try:
                accessory_count = len(self.controller_object_list[controller_index]["accessory_list"])
            except Exception as e:
                # neni prirazen kontroler
                pass

            if accessory_count > optics_count:
                number_of_blocks = accessory_count
            else:
                number_of_blocks = optics_count

            if number_of_blocks>0:
                camera_widget_row_growth = ((number_of_blocks*default_widget_height)-default_widget_height)
                self.station_list[station_index]["camera_list"][camera_index]["row_count"] = number_of_blocks

            return camera_widget_row_growth
        
        # elif widget == "optics":
        #     station_accessory_count = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"])
        #     if station_accessory_count>0:
        #         station_widget_growth_accessory = ((station_accessory_count*default_widget_height)-default_widget_height)
            
        #     self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["row_count"] = station_accessory_count
        #     return station_widget_growth_accessory

        # elif widget == "controller":
        #     controller_accessory_count = len(self.controller_object_list[controller_index]["accessory_list"])
        #     if controller_accessory_count>0:
        #         controller_widget_growth_accessory = ((controller_accessory_count*default_widget_height)-default_widget_height)
            
        #     self.controller_object_list[controller_index]["row_count"] = controller_accessory_count
        #     return controller_widget_growth_accessory
         
    def make_project_widgets(self):
        self.clear_frame(self.project_column)
        self.clear_frame(self.camera_column)
        self.clear_frame(self.optic_column)
        self.clear_frame(self.controller_column)
        self.clear_frame(self.accessory_column)
        default_height = 55

        # creating stations ------------------------------------------------------------------------------------------------------------------------------
        for i in range(0,len(self.station_list)):
            station_name = self.station_list[i]["name"]
            if i < 10:
                station_tier =  "0" + str(i) #01-99 
            else:
                station_tier =  str(i) #01-99

            station_camera_list = self.station_list[i]["camera_list"]
            camera_count = len(station_camera_list)

            station_widget_growth = self.check_widget_growth("station",station_index=i)
            station_widget = self.make_block(master_widget=self.project_column,height=default_height+station_widget_growth,width=self.default_block_width,fg_color="#181818",side = "top",text=station_name,tier=station_tier)
            # creating cameras ------------------------------------------------------------------------------------------------------------------------------
            if camera_count == 0:
                dummy_cam =         self.make_block(master_widget=self.camera_column,height=default_height-5+station_widget_growth,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                dummy_opt =         self.make_block(master_widget=self.optic_column,height=default_height-5+station_widget_growth,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                dummy_controller =  self.make_block(master_widget=self.controller_column,height=default_height-5+station_widget_growth,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                dummy_acc =         self.make_block(master_widget=self.accessory_column,height=default_height-5+station_widget_growth,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)

            for x in range(0,camera_count):
                camera_type = station_camera_list[x]["type"]
                try:
                    controller_index = int(station_camera_list[x]["controller_index"])
                    controller_color = self.controller_object_list[controller_index]["color"]
                    if controller_color == "": # nebyla zvolena žádná barva
                        controller_color = "#181818"
                except Exception: # pokud ke kamere neni prirazen zadny kontroler
                    controller_color = "#181818"
                    controller_index = None

                station_camera_optic_list = station_camera_list[x]["optics_list"]
                optic_count = len(station_camera_optic_list)
                if x < 10:
                    camera_tier =  station_tier + "0" + str(x) #0101-9999
                else:    
                    camera_tier =  station_tier + str(x) #0101-9999

                camera_widget_growth = self.check_widget_growth("camera",station_index=i,camera_index=x,controller_index=controller_index)
                camera_widget = self.make_block(master_widget=self.camera_column,height=default_height+camera_widget_growth,width=self.default_block_width,fg_color=controller_color,side = "top",text=camera_type,tier = camera_tier)

                # CONTROLLERS and accessories ---------------------------------------------------------------------------------------------------------------------------------------
                
                accessory_count = 0
                if controller_index != None:
                    if controller_index < 10:
                        controller_tier =  "c0" + str(controller_index) #c01-c99 
                    else:
                        controller_tier =  "c" + str(controller_index) #c01-c99
                    self.controller_object_list[controller_index]["detailed_name"] = str(self.controller_object_list[controller_index]["name"]) + "(" + str(self.controller_object_list[controller_index]["type"]) + ")"
                    # controller_widget_growth = self.check_widget_growth("controller",controller_index=controller_index)
                    controller_color = self.controller_object_list[controller_index]["color"]
                    if controller_color == "": # nebyla zvolena žádná barva
                        controller_color = "#181818"
                    self.make_block(master_widget=self.controller_column,height=default_height+camera_widget_growth,width=self.default_block_width,fg_color=controller_color,
                                    side = "top",text=self.controller_object_list[controller_index]["type"],tier = controller_tier)
                    
                    accessory_list = self.controller_object_list[controller_index]["accessory_list"]
                    accessory_count = len(accessory_list)
                    for x in range(0,accessory_count):
                        accessory_type = accessory_list[x]["type"]
                        if x < 10:
                            accessory_tier =  controller_tier + "0" + str(x) #c0101-c9999
                        else:
                            accessory_tier =  controller_tier + str(x) #c0101-c9999
                        
                        self.make_block(master_widget=self.accessory_column,height=default_height,width=self.default_block_width,fg_color="#181818",side = "top",text=accessory_type,tier = accessory_tier)
                    if accessory_count == 0:
                        dummy_acc = self.make_block(master_widget=self.accessory_column,height=default_height+camera_widget_growth-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                else:
                    dummy_controller = self.make_block(master_widget=self.controller_column,height=default_height+camera_widget_growth-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                    dummy_acc = self.make_block(master_widget=self.accessory_column,height=default_height+camera_widget_growth-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)

                # creating optics ------------------------------------------------------------------------------------------------------------------------------

                cycle_start_value_opt = 0
                if optic_count == 0:
                    dummy_opt = self.make_block(master_widget=self.optic_column,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                    cycle_start_value_opt = 1

                for y in range(0,optic_count):
                    optic_type = station_camera_optic_list[y]["type"]
                    if y < 10:
                        optic_tier =  camera_tier + "0" + str(y) #010101-999999
                    else:
                        optic_tier =  camera_tier + str(y) #010101-999999

                    optic_widget = self.make_block(master_widget=self.optic_column,height=default_height,width=self.default_block_width,fg_color="#181818",side = "top",text=optic_type,tier=optic_tier)

                # doplneni prazdnych mist
                if (optic_count-accessory_count) < 0:
                    for _ in range(cycle_start_value_opt,-(optic_count-accessory_count)):
                        dummy_opt = self.make_block(master_widget=self.optic_column,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)
                
                if (optic_count-accessory_count) > 0 and accessory_count > 0:
                    for _ in range(0,(optic_count-accessory_count)):
                        dummy_acc = self.make_block(master_widget=self.accessory_column,height=default_height-5,width=self.default_block_width,fg_color="#181818",side = "top",text="",dummy_block=True)

class Save_excel:
    def __init__(self,root,station_list,project_name,excel_name,controller_list,console):
        self.root = root
        self.main_console = console
        self.project_name = project_name
        self.station_list = station_list
        self.controller_list = controller_list
        self.values_start_row = 4
        # self.excel_file_name = "Katalog_kamerového_vybavení.xlsm"
        self.excel_file_name = excel_name
        if self.excel_file_name == None:
            self.excel_file_name = "Katalog_kamerového_vybavení.xlsm"
        self.temp_excel_file_name = self.excel_file_name[:-5] + "_temp.xlsm"
        self.excel_rows_used = 0
        self.used_columns = ["A","B","C","D","E"]
        self.excel_column_width=50
        self.between_station_rows = []
        self.xlsx_format = False
        # self.main()

    def make_header(self,wb):
        ws = wb["Sheet"]
        if self.xlsx_format:
            ws["A3"] = "Stanice"
            ws["C3"] = "Kamera"
            ws["E3"] = "Optika"
            ws["G3"] = "Kontrolery"
            ws["I3"] = "Příslušenství"
        else:
            ws["A3"] = "Stanice"
            ws["B3"] = "Kamera"
            ws["C3"] = "Optika"
            ws["D3"] = "Kontrolery"
            ws["E3"] = "Příslušenství"

        image = Image(resource_path("images/jhv_logo2.png"))
        ws.add_image(image,"A1")
   
    def init_objects(self):
        """
        The excel_position and the hidden_values parameters of objects needs to be inited
        - only for objects containing more locations (mentioned above)
        - case of exporting again with some changes made
        """
        for controllers in self.controller_list:
            controllers["excel_position"] = []
            controllers["hidden_values"] = []
            for accessories in controllers["accessory_list"]:
                accessories["excel_position"] = []
                accessories["hidden_values"] = []

    def merge_cells(self,wb,merge_list:str):
        """
        cell range format: A1:A2
        """
        
        ws = wb.active
        for merge in merge_list:
            ws.merge_cells(merge)

    def update_sheet_vba_code(self,new_code):
        try:
            unsuccessfull = False
            app = xw.App(visible=False)
            wb = app.books.open(self.temp_excel_file_name)
            vb_project = wb.api.VBProject
            # vb_project.VBComponents.Add(1) # musi se pridat prazdny modul...
            code_module = vb_project.VBComponents("ThisWorkbook").CodeModule
            code_module.DeleteLines(1, code_module.CountOfLines)
            code_module.AddFromString(new_code)
            try:
                wb.save(self.excel_file_name)
            except Exception:
                unsuccessfull = True
            wb.close()
            app.quit()

            if os.path.exists(self.temp_excel_file_name): # nutná operace (vyuzivat temp soubor) kvůli zapisování vba
                os.remove(self.temp_excel_file_name)
            
            if unsuccessfull:
                return False
        except Exception as e:
            print("chyba: nejsou povolena práva na makra")
            wb.close()
            app.quit()
            if os.path.exists(self.temp_excel_file_name): # nutná operace (vyuzivat temp soubor) kvůli zapisování vba
                os.remove(self.temp_excel_file_name)
            return "rights_error"

    def get_cells_to_merge(self):
        """
        skládá se pole, které buňky je potřeba spojit\n
        + ukládají se pozice pro danou hodnotu v excelu k danému objektu pod key argument: excel position
        """
        last_row = self.values_start_row
        last_row_cam = self.values_start_row
        last_row_optics = self.values_start_row
        last_row_accessory = self.values_start_row
        rows_to_merge = []
        columns = ["A","B","C","D","E"]
        if self.xlsx_format:
            columns = ["A","C","E","G","I"]

        def check_for_dummy(last_cam=False):
            #fill the rest with dummies:
            addition = 0
            if last_cam:
                addition = 1
            if int(cameras["row_count"]) > len(cameras["optics_list"]):
                cam_row_count = (int(cameras["row_count"]) - len(cameras["optics_list"])) + addition
                for _ in range(0,cam_row_count):
                    dummy_opt = {
                        "type": "",
                        "alternative": "",
                        "excel_position": columns[2]+str(last_row_optics),
                        "description": ""
                    }
                    self.station_list[station_index]["camera_list"][camera_index]["optics_list"].append(dummy_opt)

        for stations in self.station_list:
            station_index = self.station_list.index(stations)
            if stations["row_count"] > 1:
                self.station_list[station_index]["excel_position"] = columns[0]+str(last_row)
                rows_to_merge.append(columns[0] + str(last_row) + ":" + columns[0] + str(last_row + int(stations["row_count"]) - 1))
                last_row = last_row + (stations["row_count"])
            else:
                self.station_list[station_index]["excel_position"] = columns[0]+str(last_row)
                last_row = last_row + 1

            if len(stations["camera_list"]) == 0:
                last_row_cam = last_row_cam + 1
                last_row_optics = last_row_optics + 1
                # last_row_accessory = last_row_accessory + 1
            cam_inc = 0
            for cameras in stations["camera_list"]:
                camera_index = self.station_list[station_index]["camera_list"].index(cameras)
                row_before_addition = last_row_cam
                if int(cameras["row_count"]) > 1:
                    self.station_list[station_index]["camera_list"][camera_index]["excel_position"] = columns[1]+str(last_row_cam)
                    rows_to_merge.append(columns[1] + str(last_row_cam) + ":"+columns[1] + str(last_row_cam + int(cameras["row_count"]) - 1))
                    # kontrolery maji stejný merge, pocet vsech radku ulozen v kamere (i kdyz je vetsi pocet prislusenstvi nez objektivu ke kamere)
                    rows_to_merge.append(columns[3] + str(last_row_cam) + ":"+columns[3] + str(last_row_cam + int(cameras["row_count"]) - 1))
                    last_row_cam = last_row_cam + (cameras["row_count"])
                else:
                    self.station_list[station_index]["camera_list"][camera_index]["excel_position"] = columns[1]+str(last_row_cam)
                    last_row_cam = last_row_cam + 1

                if cameras["controller"] != "":
                    ii = 0
                    for controllers in self.controller_list:
                        if str((controllers["name"]+"("+controllers["type"]+")")).replace(" ","") == str(cameras["controller"]).replace(" ",""):
                            controller_index = ii
                            # stejny kontroler muze byt soucasti vice kamer, proto pole s excel_position
                            try:
                                if not str(columns[3]+str(row_before_addition)) in self.controller_list[controller_index]["excel_position"]:
                                    self.controller_list[controller_index]["excel_position"].append(str(columns[3]+str(row_before_addition)))
                            except Exception:
                                self.controller_list[controller_index]["excel_position"] = [(columns[3]+str(row_before_addition))]

                            if len(controllers["accessory_list"]) == 0:
                                last_row_accessory = last_row_accessory + 1
                            iii = 0
                            for accessories in controllers["accessory_list"]:
                                # nejprve hledam index abych mohl prepisovat přímo celý objekt
                                # accessory_index = self.controller_list[controller_index]["accessory_list"].index(accessories)
                                accessory_index = iii
                                try:
                                    if not (columns[4]+str(row_before_addition+iii)) in self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"]:
                                        self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"].append(columns[4]+str(row_before_addition+iii))
                                except Exception:
                                    self.controller_list[controller_index]["accessory_list"][accessory_index]["excel_position"] = [columns[4]+str(row_before_addition+iii)]
                                last_row_accessory = last_row_accessory + 1
                                iii+=1
                            break
                        ii += 1


                if len(cameras["optics_list"]) == 0:
                    last_row_optics = last_row_optics + 1
                
                # print("cam inc",cam_inc,"length list",len(stations["camera_list"]))
                # if cam_inc == len(stations["camera_list"])-1:
                #     check_for_dummy(last_cam = True)
                # else:
                check_for_dummy()

                for optics in cameras["optics_list"]:
                    optics_index = self.station_list[station_index]["camera_list"][camera_index]["optics_list"].index(optics)
                    self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["excel_position"] = columns[2]+str(last_row_optics)
                    last_row_optics = last_row_optics + 1

                cam_inc+=1

                
            self.between_station_rows.append(last_row_cam)
            #radek mezera mezi kazdou stanici
            last_row+=1
            last_row_cam+=1
            last_row_optics = last_row_optics + 1

        self.excel_rows_used = last_row_optics
        print("LAST ROW USED: ",self.excel_rows_used)

        if self.xlsx_format:
            columns = ["A","C","E","G"]
            for merges in rows_to_merge:
                if columns[0] in merges:
                    rows_to_merge.append(merges.replace(columns[0] ,"B"))
                elif columns[1] in merges:
                    rows_to_merge.append(merges.replace(columns[1] ,"D"))
                elif columns[2] in merges:
                    rows_to_merge.append(merges.replace(columns[2] ,"F"))
                elif columns[3] in merges:
                    rows_to_merge.append(merges.replace(columns[3] ,"H"))
            line = self.values_start_row -1
            rows_to_merge.append(f"A{line}:B{line}")
            rows_to_merge.append(f"C{line}:D{line}")
            rows_to_merge.append(f"E{line}:F{line}")
            rows_to_merge.append(f"G{line}:H{line}")
            rows_to_merge.append(f"I{line}:J{line}")
        # grafika header:
        rows_to_merge.append("A1:A2")
        if self.xlsx_format:
            rows_to_merge.append("B1:J1")
            rows_to_merge.append("B2:J2")
        else:
            rows_to_merge.append("B1:E1")
            rows_to_merge.append("B2:E2")

        return rows_to_merge

    def change_vba_script(self):
        """
        Slouží pro přidávání rozsahu hodnot, uložených v hidden sheetu a alokování k určité buňce
        - současná kapacita pro jeden objekt: 78 (3* abeceda)
        """
        vba_code_range = """"""
        alphabet = string.ascii_uppercase
        i = 0
        ii = 0
        iii = 0
        columns = ["A","B","C","D","E","F","G","H","I"]
        column_letter_st = 0
        column_letter_cam = 3
        column_letter_opt = 6

        for stations in self.station_list:
            cell_with_toggle = stations["excel_position"]
            if i > 25:
                column_letter_st +=1
                i=0
            column = columns[column_letter_st] + alphabet[i:i+1] 
            stations["hidden_values"] = column # pridame jen informaci o nazvu sloupce
            station_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", Cancel, Target"
            vba_code_range += "\n            "+station_vba_code_range_row
            i+=1

            for cameras in stations["camera_list"]:
                cell_with_toggle = cameras["excel_position"]
                if ii > 25:
                    column_letter_cam +=1
                    ii=0
                column = columns[column_letter_cam] + alphabet[ii:ii+1] 
                cameras["hidden_values"] = column # pridame jen informaci o nazvu sloupce
                camera_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", Cancel, Target"
                vba_code_range += "\n            "+camera_vba_code_range_row
                ii+=1

                for optics in cameras["optics_list"]:
                    cell_with_toggle = optics["excel_position"]
                    if iii > 25:
                        column_letter_opt += 1
                        iii=0
                    column = columns[column_letter_opt] + alphabet[iii:iii+1] 
                    optics["hidden_values"] = column # pridame jen informaci o nazvu sloupce
                    optics_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", Cancel, Target"
                    vba_code_range += "\n            "+optics_vba_code_range_row
                    iii+=1
        i = 0
        ii = 0
        columns = ["J","K","L","M","N","O","P","Q"]
        column_letter_controller = 0
        column_letter_acc = 3
        for controllers in self.controller_list:
            try:
                for controller_positions in controllers["excel_position"]: 
                    cell_with_toggle = controller_positions
                    if i > 25:
                        column_letter_controller +=1
                        i=0
                    column = columns[column_letter_controller] + alphabet[i:i+1] 
                    try:
                        controllers["hidden_values"].append(column) # pridame jen informaci o nazvu sloupce
                    except Exception:
                        controllers["hidden_values"] = [column]
                        
                    controller_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", Cancel, Target"
                    vba_code_range += "\n            "+controller_vba_code_range_row
                    i+=1
                    for accessories in controllers["accessory_list"]:
                        for acc_positions in accessories["excel_position"]:
                            cell_with_toggle = acc_positions
                            if ii > 25:
                                column_letter_acc +=1
                                ii=0
                            column = columns[column_letter_acc] + alphabet[ii:ii+1]
                            try:
                                accessories["hidden_values"].append(column) # pridame jen informaci o nazvu sloupce
                            except Exception:
                                accessories["hidden_values"] = [column]
                            acc_vba_code_range_row = f"ToggleCell Range(\"Sheet!{cell_with_toggle}\"), \"{column + str(1)}\", \"{column + str(2)}\", \"{column + str(3)}\", Cancel, Target"
                            vba_code_range += "\n            "+acc_vba_code_range_row
                            ii+=1
            except Exception: # the station with this controller was deleted
                pass

        vba_code = f"""
        Private Sub Workbook_SheetBeforeRightClick(ByVal Sh As Object, ByVal Target As Range, Cancel As Boolean)
            {vba_code_range}
        End Sub

        Private Sub ToggleCell(ByVal targetCell As Range, ByVal text1Ref As String, ByVal text2Ref As String, ByVal toggleStatusRef As String, ByRef Cancel As Boolean, ByVal clickedCell As Range)
            ' Read text values from hidden worksheet
            Dim text1 As String
            Dim text2 As String
            text1 = Worksheets("HiddenSheet").Range(text1Ref).Value
            text2 = Worksheets("HiddenSheet").Range(text2Ref).Value

            ' Read toggle status from hidden worksheet
            Dim toggle_status As Integer
            toggle_status = Worksheets("HiddenSheet").Range(toggleStatusRef).Value

            ' Check if the right-clicked cell is the target cell
            If Not Intersect(clickedCell, targetCell) Is Nothing Then
                ' Toggle the cell value
                If toggle_status = 1 Then
                    Worksheets("HiddenSheet").Range(text1Ref).Value = targetCell.Value
                    targetCell.Value = text2
                    toggle_status = 0
                Else
                    Worksheets("HiddenSheet").Range(text2Ref).Value = targetCell.Value
                    targetCell.Value = text1
                    toggle_status = 1
                End If

                ' Update toggle status on hidden worksheet
                Worksheets("HiddenSheet").Range(toggleStatusRef).Value = toggle_status
                ' Cancel the default right-click menu
                Cancel = True
            End If
        End Sub

        """
        return vba_code

    def format_cells(self,ws):
        bold_font_white = Font(bold=True,size=20,color="ffffff") # ffffff = bílá!
        bold_font = Font(bold=True,size=12)
        regular_font = Font(bold=False,size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 65

        # cell = f"Projekt: {self.project_name}"
        top_header_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
        ws["B1"] = "Přehled kamerového vybavení"
        ws["B1"].alignment = Alignment(horizontal = "left", vertical = "center")
        ws["B1"].font = Font(bold=True,size=25)
        comment_text = "Pravým klikem na buňky v tabulce zobrazíte podrobnosti"
        comment_author = "TRIMAZKON"
        comment = Comment(comment_text, comment_author)
        if not self.xlsx_format:
            ws['B1'].comment = comment
        # ws['B1'].fill = top_header_fill
        
        current_date = datetime.now().date()
        date_string = current_date.strftime("%d.%m.%Y")
        ws["B2"] = f"Projekt: {self.project_name}\nDatum: {date_string}"
        ws["B2"].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)
        ws["B2"].font = Font(bold=True,size=20)
        # ws['B2'].fill = top_header_fill
        
        for columns in self.used_columns:
            for i in range(3,self.excel_rows_used-1):
                ws.column_dimensions[columns].width = self.excel_column_width
                cell = ws[columns + str(i)]
                # cell.alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)
                cell.alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True,shrink_to_fit=True,justifyLastLine=True)
                cell.border = thin_border

                if i == 3: # nadpisy sloupců
                    header_fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
                    cell.font = bold_font_white
                    cell.alignment = Alignment(horizontal = "center", vertical = "center")
                    cell.fill = header_fill
                else:
                    cell.font = regular_font
                
                # Názvy stanic:
                if columns == "A" and i != 3:
                    cell.font = bold_font

        # fill the empty rows between stations:
        for rows in self.between_station_rows:
            for columns in self.used_columns:
                if self.xlsx_format:
                    cell = ws[columns + str(rows)]
                    fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
                    cell.fill = fill
                else:
                    cell = ws[columns + str(rows)]
                    fill = PatternFill(start_color="636363", end_color="636363", fill_type="solid")
                    cell.fill = fill

    def fill_values(self,wb):
        """
        vepíše hodnoty
        - xlsx první sloupec - v gui prvně viditelné informace
        - xlsm první informace
        """
        ws = wb.active
        columns = ["B","C","D","E"]
        if self.xlsx_format:
            columns = ["C","E","G","I"]

        for stations in self.station_list:
            excel_cell = stations["excel_position"]
            ws[excel_cell] = stations["name"]

            if len(stations["camera_list"]) == 0:
                excel_cell = columns[0] + stations["excel_position"][1:]
                ws[excel_cell] = ""
            for cameras in stations["camera_list"]:
                excel_cell = cameras["excel_position"]
                ws[excel_cell] = cameras["type"]
                if str(cameras["controller_color"]) != "":
                    try:
                        color_modified = str(cameras["controller_color"])[1:]
                        controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                        ws[excel_cell].fill = controller_fill
                    except Exception as e:
                        print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                        pass
                
                if len(cameras["optics_list"]) == 0:
                    excel_cell = columns[1] + cameras["excel_position"][1:]
                    ws[excel_cell] = ""
                for optics in cameras["optics_list"]:
                    excel_cell = optics["excel_position"]
                    ws[excel_cell] = optics["type"]

        for controllers in self.controller_list:
            try:
                for position in controllers["excel_position"]:
                    excel_cell = str(position)
                    ws[excel_cell] = controllers["type"]
                    if str(controllers["color"]) != "" and str(controllers["color"]) != "#212121":
                        try:
                            color_modified = str(controllers["color"])[1:] # without hashtag
                            controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                            ws[excel_cell].fill = controller_fill
                        except Exception as e:
                            print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                            pass

                    if len(controllers["accessory_list"]) == 0:
                        excel_cell = columns[3] + position[1:]
                        ws[excel_cell] = ""
                    for accessories in controllers["accessory_list"]:
                        for acc_positions in accessories["excel_position"]:
                            # print("accessories excel position: ",accessories["excel_position"])
                            excel_cell = acc_positions
                            ws[excel_cell] = accessories["type"]
            except Exception: # the station with this controller was deleted
                pass       
        self.format_cells(ws)

    def fill_hidden_sheet_values(self,wb):
        """
        Provede vytvoření skrytého listu, kam ukládá toggle hodnoty a aktuální stav přepnutí\n
        Rozdělení:
        - Vždy tři hodnoty
            - toggle první hodnota (název/ typ)
            - toggle druhá hodnota (doplňující informace)
            - stav togglu (přepnutí 0-1)
        - stanice: AZ - CZ
        - kamery: DZ - FZ
        - optika: GZ - IZ
        - kontrolery: JZ - LZ
        - příslušenství: MZ - OZ
        """

        ws = wb.create_sheet("HiddenSheet")
        ws.sheet_state = 'hidden'

        for stations in self.station_list:
            excel_cell = stations["hidden_values"]
            ws[excel_cell + str(1)] = stations["name"]
            ws[excel_cell + str(2)] = stations["inspection_description"]
            ws[excel_cell + str(3)] = 1 # toggle status... default: 1

            for cameras in stations["camera_list"]:
                excel_cell = cameras["hidden_values"]
                ws[excel_cell + str(1)] = cameras["type"]
                detail_info = ""
                if str(cameras["cable"]) != "":
                    detail_info = detail_info + "Kabel: " + str(cameras["cable"])+ "\n"

                ws[excel_cell + str(2)] = detail_info + str(cameras["description"])
                ws[excel_cell + str(3)] = 1
                
                for optics in cameras["optics_list"]:
                    excel_cell = optics["hidden_values"]
                    ws[excel_cell + str(1)] = optics["type"]
                    detail_info = ""
                    if str(optics["alternative"]) != "":
                        detail_info = "Alternativa: " + str(optics["alternative"]) + "\n"
                    detail_info = detail_info + str(optics["description"])
                    ws[excel_cell + str(2)] = detail_info
                    ws[excel_cell + str(3)] = 1

        for controllers in self.controller_list:
            for controller_positions in controllers["hidden_values"]:
                excel_cell = controller_positions
                ws[excel_cell + str(1)] = controllers["type"]
                details = controllers["detailed_name"] + "\n"
                if not controllers["notes"] == "":
                    details = details + controllers["notes"] + "\n\n"
                if not controllers["ip"] == "" and not controllers["ip"] == "192.168.000.000":
                    details = details + controllers["ip"] + "\n"
                if not controllers["username"] == "":
                    details = details + "Jméno: " + controllers["username"] + "\n"
                if not controllers["password"] == "":
                    details = details + "Heslo: " + controllers["password"] + "\n"
                ws[excel_cell + str(2)] = details
                ws[excel_cell + str(3)] = 1 # toggle status... default: 1

                for accessories in controllers["accessory_list"]:
                    for acc_positions in accessories["hidden_values"]:
                        excel_cell = acc_positions
                        ws[excel_cell + str(1)] = accessories["type"]
                        ws[excel_cell + str(2)] = accessories["description"]
                        ws[excel_cell + str(3)] = 1 # toggle status... default: 1

    def fill_xlsx_column(self,wb):
        ws = wb.active
        columns = ["D","F","H","J"]
        for stations in self.station_list:
            excel_cell = str(stations["excel_position"])
            excel_cell = excel_cell.replace("A","B")
            ws[excel_cell] = stations["inspection_description"]
            ws[excel_cell].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)
            if len(stations["camera_list"]) == 0:
                excel_cell = columns[0] + stations["excel_position"][1:]
                ws[excel_cell] = ""
            for cameras in stations["camera_list"]:
                excel_cell = cameras["excel_position"]
                excel_cell = excel_cell.replace("C","D")
                detail_info = ""
                if str(cameras["controller_color"]) != "":
                    try:
                        color_modified = str(cameras["controller_color"])[1:]
                        controller_fill = PatternFill(start_color=color_modified, end_color=color_modified, fill_type="solid")
                        ws[excel_cell].fill = controller_fill
                    except Exception as e:
                        print(f"chyba pri nastavovani barvy kontroleru pri exportu: {e}")
                        pass
                if str(cameras["cable"]) != "":
                    detail_info = detail_info + "Kabel: " + str(cameras["cable"])+ "\n"

                ws[excel_cell] = detail_info + str(cameras["description"])
                ws[excel_cell].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)
                
                if len(cameras["optics_list"]) == 0:
                    excel_cell = columns[1] + cameras["excel_position"][1:]
                    ws[excel_cell] = ""
                for optics in cameras["optics_list"]:
                    excel_cell = optics["excel_position"]
                    excel_cell = excel_cell.replace("E","F")
                    detail_info = ""
                    if str(optics["alternative"]) != "":
                        detail_info = "Alternativa: " + str(optics["alternative"]) + "\n\n"
                    ws[excel_cell] = detail_info + str(optics["description"])
                    ws[excel_cell].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)

        for controllers in self.controller_list:
            for position in controllers["excel_position"]:
                excel_cell = str(position)
                excel_cell = excel_cell.replace("G","H")
                detail_info = controllers["detailed_name"] + "\n"

                if str(controllers["ip"]) != "" and str(controllers["ip"]) != "192.168.000.000":
                    detail_info += str(controllers["ip"]) + "\n"
                if str(controllers["username"]) != "":
                    detail_info += str(controllers["username"]) + "\n"
                if str(controllers["password"]) != "":
                    detail_info += str(controllers["password"]) + "\n\n"
                if str(controllers["notes"]) != "":
                    detail_info += str(controllers["notes"]) + "\n"
                ws[excel_cell] = detail_info
                ws[excel_cell].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)


                if len(controllers["accessory_list"]) == 0:
                    # excel_cell = columns[3] + controllers["excel_position"][1:]
                    excel_cell = columns[3] + position[1:]
                    ws[excel_cell] = ""
                for accessory in controllers["accessory_list"]:
                    for acc_positions in accessory["excel_position"]:
                        excel_cell = acc_positions
                        excel_cell = excel_cell.replace("I","J")
                        ws[excel_cell] = accessory["description"]
                        ws[excel_cell].alignment = Alignment(horizontal = "left", vertical = "center",wrap_text=True)

    def main(self):
        wb = Workbook() #vytvorit novy excel, prepsat...
        if ".xlsm" in self.excel_file_name:
            self.init_objects()
            rows_to_merge = self.get_cells_to_merge()
            self.make_header(wb)
            # try:
            if os.path.exists(self.temp_excel_file_name):
                os.remove(self.temp_excel_file_name)
            # kličky aby se uložilo vba:
            wb.save(filename=self.temp_excel_file_name)
            wb2 = load_workbook(filename=self.temp_excel_file_name, keep_vba=True)
            wb2.save(self.temp_excel_file_name)
            wb.close()
            wb2.close()
            wb = load_workbook(filename=self.temp_excel_file_name, read_only=False, keep_vba=True)
            self.merge_cells(wb,merge_list=rows_to_merge)
            self.fill_values(wb)
            new_vba_code = self.change_vba_script()
            self.fill_hidden_sheet_values(wb)
            wb.save(self.temp_excel_file_name)
            wb.close()
            attempt = self.update_sheet_vba_code(new_code=new_vba_code)
            if attempt == False:
                add_colored_line(self.main_console,f"Nejprve prosím zavřete soubor {self.excel_file_name}","red",None,True)
            elif attempt == "rights_error":
                add_colored_line(self.main_console,f"Nemáte nastavená potřebná práva v excelu pro makra (VBA)","red",None,True)
                window = ToplevelWindow(self.root)
                window.excel_manual_window()
            else:
                add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně exportován","green",None,True)
                os.startfile(self.excel_file_name)
                return
            # except Exception as e:
            #     add_colored_line(self.main_console,f"Neočekávaná chyba {e}","red",None,True)
            #     wb.close()
                
        elif ".xlsx" in self.excel_file_name:
            self.used_columns = ["A","B","C","D","E","F","G","H","I","J"]
            self.xlsx_format = True
            self.init_objects()
            rows_to_merge = self.get_cells_to_merge()
            self.make_header(wb)

            # try:
            self.merge_cells(wb,merge_list=rows_to_merge)
            self.fill_values(wb)
            self.fill_xlsx_column(wb)
            wb.save(self.excel_file_name)
            wb.close()
            add_colored_line(self.main_console,f"Projekt {self.project_name} byl úspěšně exportován","green",None,True)
            os.startfile(self.excel_file_name)
            return
            # except Exception as e:
            #     add_colored_line(self.main_console,f"Nejprve prosím zavřete soubor {self.excel_file_name}, chyba: {e}","red",None,True)
            #     wb.close()

# download = download_database.database(database_filename)
# Catalogue_gui(root,download.output)
# Catalogue_gui(root,"testing - stahování vypnuto","","max",database_filename=database_filename)
# root.mainloop()