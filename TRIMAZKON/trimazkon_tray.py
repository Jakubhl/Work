
from pystray import Icon, Menu, MenuItem
from PIL import Image, ImageDraw
from openpyxl import load_workbook
import threading
import customtkinter


class tray_app_service:
    def __init__(self,app_icon,task_list,deletion_log):
        self.app_icon = app_icon
        self.config_filename = "config_TRIMAZKON.xlsx"
        # self.main()
        

    
    def read_config(self):
        wb = load_workbook(self.config_filename,read_only=True)
        ws = wb["task_settings"]
        all_tasks = []
        self.task_log_list = []
        for row in ws.iter_rows(values_only=True):
            row_array = []
            for items in row[:6]:
                if items is not None:
                    row_array.append(str(items))

            if len(row_array) > 1:
                all_tasks.insert(0,row_array[0])
                self.task_log_list.insert(0,[row_array[0],row_array[1]])
            elif len(row_array) > 0:
                all_tasks.insert(0,row_array[0])
        
        wb.close()
        return all_tasks
    
    def show_all_tasks(self,toplevel=False):
        if not toplevel:
            child_root = customtkinter.CTk()
        else:
            child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
        child_root.title("Seznam nastavených úkolů (task scheduler)")

        main_frame = customtkinter.CTkFrame(master=child_root,corner_radius=0)
        # main_frame = customtkinter.CTkScrollableFrame(master=child_root,corner_radius=0)
        all_tasks = self.read_config()
        i=0
        for tasks in all_tasks:
            task_name = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=0,height= 50,fg_color="#636363")
            task_name_text = customtkinter.CTkLabel(master=task_name,text = "Úkol "+str(i+1),font=("Arial",20,"bold"),anchor="w")
            task_name_text.pack(pady=(5,1),padx=10,anchor="w")
            task_name.pack(pady=(10,0),padx=5,side="top",anchor="w",fill="x")


            task_frame = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=3,height= 50,border_color="#636363")
            param1_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=0,height= 50)
            param1_subframe1 = customtkinter.CTkFrame(master=param1_frame,corner_radius=0,border_width=2,height= 50,width=250)
            param1_label = customtkinter.CTkLabel(master=param1_subframe1,text = "Čas spuštění (denně): ",font=("Arial",20,"bold"),anchor="w")
            param1_label.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param1_subframe2 = customtkinter.CTkFrame(master=param1_frame,corner_radius=0,border_width=2,height= 50)
            param1_label2 = customtkinter.CTkLabel(master=param1_subframe2,text = "12:00",font=("Arial",20),anchor="w")
            param1_label2.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param1_subframe1.pack(side="left")
            param1_subframe1.pack_propagate(0)
            param1_subframe2.pack(side="left",fill="x",expand=True)
            param1_frame.pack(pady=(3,0),padx=3,fill="x",side="top")

            param2_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=1,height= 50)
            param2_subframe1 = customtkinter.CTkFrame(master=param2_frame,corner_radius=0,border_width=2,height= 50,width=250)
            param2_label = customtkinter.CTkLabel(master=param2_subframe1,text = "Pracuje v: ",font=("Arial",20,"bold"),anchor="w")
            param2_label.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param2_subframe2 = customtkinter.CTkFrame(master=param2_frame,corner_radius=0,border_width=2,height= 50)
            param2_label2 = customtkinter.CTkLabel(master=param2_subframe2,text = "images/logo_TRIMAZKON.icoimages/logo_TRIMAZKON.icoimages/logo_TRIMAZKON.ico",font=("Arial",20),anchor="w")
            param2_label2.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param2_subframe1.pack(side="left")
            param2_subframe1.pack_propagate(0)
            param2_subframe2.pack(side="left",fill="x",expand=True)
            param2_frame.pack(pady=(0,0),padx=3,fill="x",side="top")

            param3_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=1,height= 50)
            param3_label = customtkinter.CTkLabel(master=param3_frame,text = "Nastavení: ",font=("Arial",20,"bold"),anchor="w")
            param3_label2 = customtkinter.CTkLabel(master=param3_frame,text = "starší než:  30 dní, minimum = 1000 souborů",font=("Arial",20),anchor="w")
            param3_label.pack(pady=10,padx=(10,0),anchor="w",side="left")
            param3_label2.pack(pady=10,padx=(10,0),anchor="w",side="left")
            param3_frame.pack(pady=(0,3),padx=3,fill="x",side="top")
            
            task_frame.pack(pady=(0,0),padx=5,fill="x",side="top")
            i+=1

        if len(all_tasks) == 0:
            task_label = customtkinter.CTkLabel(master=main_frame,text = "Nejsou nastaveny žádné úkoly...",font=("Arial",22,"bold"),anchor="w")
            task_label.pack(pady=10,padx=10,fill="x",side="top",anchor="w")
            child_root.after(2000, lambda: child_root.destroy())
        # main_frame.update()
        # main_frame.update_idletasks()
        main_frame.pack(fill="both",side="top")
        child_root.update()
        child_root.update_idletasks()
        child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()+10}")
        child_root.mainloop()


    def show_task_log(self):
        child_root = customtkinter.CTk()
        child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
        child_root.title("Záznam o vymazaných souborech")

        main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0)
        self.read_config()

        i=0
        for logs in self.task_log_list:
            task_frame = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=2,height= 50)
            task_label = customtkinter.CTkLabel(master=task_frame,text = str(logs[0]),font=("Arial",20,"bold"),anchor="w")
            task_label2 = customtkinter.CTkLabel(master=task_frame,text = str(logs[1]),font=("Arial",20,"bold"),anchor="w")
            task_label.pack(pady=(10,0),padx=10,anchor="w",side="top")
            task_label2.pack(pady=(10,0),padx=10,anchor="w",side="top")
            if i == 0:
                task_frame.pack(pady=(10,0),padx=10,fill="x",side="top")
            else:
                task_frame.pack(pady=0,padx=10,fill="x",side="top")
            i+=1
        if len(self.task_log_list) == 0:
            task_label = customtkinter.CTkLabel(master=main_frame,text = "Nebyl nalezen žádný záznam",font=("Arial",22,"bold"),anchor="w")
            task_label.pack(pady=10,padx=10,fill="x",side="top",anchor="w")
            child_root.after(2000, lambda: child_root.destroy())
        main_frame.pack(expand=False,fill="x",side="top")
        child_root.update()
        child_root.update_idletasks()
        child_root.geometry(f"{600}x{child_root.winfo_height()+10}")
        child_root.mainloop()


    def quit_application(self, item):
        self.icon.stop()

    def only_refresh(self):
        #read
        #create menu
        # self.icon.menu = Menu(
        # MenuItem('New Action', lambda: print("Action triggered"))
        # )
        all_tasks = self.read_config()
        print("all tasks:",str(all_tasks))

    # Create a menu
    def create_menu(self):
        self.menu = Menu(MenuItem('Zobrazit nastavené úkoly', self.show_all_tasks),
                         MenuItem('Záznamy o mazání', self.show_task_log),
                         MenuItem('Ukončit', self.quit_application))

    def main(self):
        def create_image():
            image = Image.open(self.app_icon)
            return image
        
        self.create_menu()
        self.icon = Icon(
            "TRIMAZKON_tooltip",
            create_image(),
            "TRIMAZKON",
            self.menu
        )

        # Run the tray icon
        self.icon.run()

tray_app_service('images/logo_TRIMAZKON.ico',[],[])
# CREATING TASK:
# name_of_task = "dailyscript_test"
# path_to_app = r"C:\Users\jakub.hlavacek.local\Desktop\JHV\Work\TRIMAZKON\pipe_server\untitled2.py"
# cmd_command = f"schtasks /Create /TN {name_of_task} /TR {path_to_app} /SC DAILY /ST 09:35"
# connection_status = subprocess.call(cmd_command,shell=True,text=True)

#DELETING TASK:
# name_of_task = "dailyscript_test"
# cmd_command = f"schtasks /Delete /TN {name_of_task} /F"
# connection_status = subprocess.call(cmd_command,shell=True,text=True)