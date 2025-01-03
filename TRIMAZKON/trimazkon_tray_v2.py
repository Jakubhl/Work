
from pystray import Icon, Menu, MenuItem
from PIL import Image, ImageDraw
from openpyxl import load_workbook
import customtkinter
import tkinter as tk
import pyperclip
import os
import subprocess
import sys

class Tools:
    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)
    
    @classmethod
    def path_check(cls,path_raw,only_repair = None):
        path=path_raw
        backslash = "\\"
        if backslash[0] in path:
            newPath = path.replace(backslash[0], '/')
            path = newPath
        if path.endswith('/') == False:
            newPath = path + "/"
            path = newPath
        #oprava mezery v nazvu
        path = r"{}".format(path)
        if not os.path.exists(path) and only_repair == None:
            return False
        else:
            return path

class tray_app_service:
    def __init__(self,initial_path,resource_app_path =None):
        if resource_app_path == None:
            self.app_icon = initial_path + 'images/logo_TRIMAZKON.ico'
        else:
            self.app_icon = resource_app_path + 'images/logo_TRIMAZKON.ico'

        self.config_filename = "config_TRIMAZKON.xlsx"
        self.initial_path = initial_path
        self.main_app_exe_name = "TRIMAZKON.exe"
        self.config_sheet_name = "Task_settings" 
        
    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            if widget.winfo_exists():
                widget.unbind("<Enter>")
                widget.unbind("<Leave>")
                widget.unbind("<Return>")
                widget.unbind("<Button-1>")
                widget.unbind("<Button-3>")
                widget.unbind("<Double-1>")
                widget.unbind("<MouseWheel>")
                widget.destroy()
    
    def read_config(self):
        wb = load_workbook(self.initial_path +self.config_filename,read_only=True)
        if not self.config_sheet_name in wb.sheetnames:
            wb.close()
            wb = load_workbook(self.initial_path + self.config_filename)
            ws = wb.create_sheet(self.config_sheet_name)
            ws.sheet_state = 'hidden'
            wb.save(self.initial_path + self.config_filename)
            wb.close()
            return []
        
        ws = wb[self.config_sheet_name]
        all_tasks = []
        self.task_log_list = []
        for row in ws.iter_rows(values_only=True):
            row_array = []
            for items in row:
                if items is not None:
                    row_array.append(str(items))

            if len(row_array) < 7:
                row_array.append("")
            elif len(row_array) == 7:
                # self.task_log_list.append(row_array[6])
                self.task_log_list.insert(0,row_array[6])
            if len(row_array) > 1:
                all_tasks.insert(0,row_array)
        
        wb.close()
        return all_tasks

    def save_new_log(self,task_name:str,new_log:str): #musim mit na vstupu nazev tasku abych ho mohl najit a prepsat to u nej
        wb = load_workbook(self.initial_path +self.config_filename)
        ws = wb[self.config_sheet_name]
        self.check_task_existence()
        current_tasks = self.read_config()
        row_to_print = 1
        if len(current_tasks) > 0:
            row_to_print = len(current_tasks)
        for tasks in current_tasks:
            if str(tasks[0]) == task_name:
                ws['G' + str(row_to_print)] = tasks[6] + new_log # log mazání (pocet smazanych,datum,seznam smazanych)
                break
            row_to_print -=1
        try:
            wb.save(self.initial_path +self.config_filename)
            wb.close()
        except Exception as e:
            print(e)
            wb.close()
            return False
        
    def delete_log(self,task_name:str,childroot): #musim mit na vstupu nazev tasku abych ho mohl najit a prepsat to u nej
        wb = load_workbook(self.initial_path +self.config_filename)
        ws = wb[self.config_sheet_name]
        self.check_task_existence()
        current_tasks = self.read_config()
        row_to_print = 1
        if len(current_tasks) > 0:
            row_to_print = len(current_tasks)
        for tasks in current_tasks:
            if str(tasks[0]) == task_name:
                ws['G' + str(row_to_print)] = ""
                break
            row_to_print -=1
        try:
            wb.save(self.initial_path +self.config_filename)
            wb.close()
            self.show_task_log(root_given=childroot)
        except Exception as e:
            print(e)
            wb.close()
            return False
        
    def save_task_to_config(self,current_tasks):
        def clear_document(wb,ws):
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
            try:
                wb.save(self.initial_path +self.config_filename)
            except Exception:
                pass

        wb = load_workbook(self.initial_path +self.config_filename)
        ws = wb[self.config_sheet_name]
        clear_document(wb,ws)
        row_to_print = 1
        print("current_tasks",current_tasks)
        for tasks in current_tasks:
            ws['A' + str(row_to_print)] = tasks[0] # nazev tasku
            ws['B' + str(row_to_print)] = tasks[1] # cesta vykonavani
            ws['C' + str(row_to_print)] = tasks[2] # max days
            ws['D' + str(row_to_print)] = tasks[3] # min left
            ws['E' + str(row_to_print)] = tasks[4] # frequency
            ws['F' + str(row_to_print)] = tasks[5] # datum pridani tasku
            ws['G' + str(row_to_print)] = tasks[6] # log mazání (pocet smazanych,datum,seznam smazanych)
            row_to_print +=1
        try:
            wb.save(self.initial_path +self.config_filename)
            wb.close()
        except Exception as e:
            print(e)
            wb.close()
            return False

    def delete_task(self,id,root):
        def delete_from_scheduler():
            name_of_task = all_tasks[id][0]
            cmd_command = f"schtasks /Delete /TN {name_of_task} /F"
            subprocess.call(cmd_command,shell=True,text=True)

        self.check_task_existence()
        all_tasks = self.read_config()
        delete_from_scheduler()
        all_tasks.pop(id)
        
        status = self.save_task_to_config(all_tasks)
        if status != False:
            # root.destroy()
            self.show_all_tasks(root_given=root)

    def show_context_menu(self,root,event,widget,id):
        self.check_task_existence()
        all_tasks = self.read_config()
        context_menu = tk.Menu(root,tearoff=0,fg="white",bg="black",font=("Arial",20,"bold"))
        preset_font=("Arial",18,"bold")
        path = all_tasks[id][1]

        if widget == "path":
            context_menu.add_command(label="Otevřít cestu",font=preset_font, command=lambda: os.startfile(path))
            context_menu.add_separator()
            context_menu.add_command(label="Kopírovat cestu",font=preset_font, command=lambda: pyperclip.copy(path))

        elif widget == "time" or widget == "settings" or widget == "name":
            name_of_task = all_tasks[id][0]
            path_app_location = str(self.initial_path+"/"+self.main_app_exe_name) 
            task_command = path_app_location + " deleting " + name_of_task + " " + str(all_tasks[id][1]) + " " + str(all_tasks[id][2]) + " " + str(all_tasks[id][3])
            context_menu.add_command(label="Vykonat úkol",font=preset_font,command=lambda: subprocess.call(task_command,shell=True,text=True))
            context_menu.add_separator()
            context_menu.add_command(label="Upravit úkol",font=preset_font,command=lambda: os.startfile("taskschd.msc"))
            context_menu.add_separator()
            context_menu.add_command(label="Odstranit úkol",font=preset_font,command=lambda: self.delete_task(id,root))
            context_menu.add_separator()
            context_menu.add_command(label="Zobrazit historii mazání",font=preset_font,command=lambda: self.show_task_log(True,task_given=all_tasks[id][0]))

        elif widget == "del_log":
            context_menu.add_command(label="Otevřít cestu",font=preset_font, command=lambda: os.startfile(path))
            context_menu.add_separator()
            context_menu.add_command(label="Kopírovat cestu",font=preset_font, command=lambda: pyperclip.copy(path))
            context_menu.add_separator()
            context_menu.add_command(label="Vymazat historii mazání",font=preset_font, command=lambda: self.delete_log(task_name=all_tasks[id][0],childroot=root))

        context_menu.tk_popup(event.x_root, event.y_root)

    def check_task_existence(self,task_given = None):
        def check_task_status(taskname):
            process = subprocess.Popen(f'schtasks /query /tn \"{taskname}\" /v /fo LIST',
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            try:
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                data = str(stdout_str)
                error_data = str(stderr_str)
            except UnicodeDecodeError:
                try:
                    stdout_str = stdout.decode('cp1250')
                    stderr_str = stderr.decode('cp1250')
                    data = str(stdout_str)
                    error_data = str(stderr_str)
                except UnicodeDecodeError:
                    data = str(stdout)
                    error_data = str(stderr)
            if "ERROR" in error_data:
                return False
            else:
                return True
                
        if task_given != None:
            check_result = check_task_status(task_given)
            return check_result
        
        all_tasks = self.read_config()
        non_existent_tasks = []
        for i in range(0,len(all_tasks)):
            taskname = str(all_tasks[i][0])
            task_presence = check_task_status(taskname)
            if not task_presence:
                non_existent_tasks.append(taskname)
        
        if len(non_existent_tasks) > 0:
            for deleted_tasks in non_existent_tasks:
                for tasks in all_tasks:
                    if deleted_tasks == tasks[0]:
                        all_tasks.pop(all_tasks.index(tasks))
                        break
            self.save_task_to_config(all_tasks)

    def show_all_tasks(self,toplevel=False,root_given = False):
        if root_given != False:
            child_root = root_given
            self.clear_frame(child_root)
        else:
            if not toplevel:
                child_root = customtkinter.CTk()
            else:
                child_root = customtkinter.CTkToplevel()
            child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
            child_root.title("Seznam nastavených úkolů (task scheduler)")

        # main_frame = customtkinter.CTkFrame(master=child_root,corner_radius=0)
        main_frame = customtkinter.CTkScrollableFrame(master=child_root,corner_radius=0)
        self.check_task_existence()
        all_tasks = self.read_config()
        print("all_tasks: ",all_tasks)
        i=0
        for tasks in all_tasks:
            task_name = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=0,height= 50,fg_color="#636363")
            task_name_text = customtkinter.CTkLabel(master=task_name,text = "Úkol "+str(i+1) + f" (scheduler název: {tasks[0]})",font=("Arial",20,"bold"),anchor="w")
            task_date_accessed = customtkinter.CTkLabel(master=task_name,text = f"Přidáno: {tasks[5]}",font=("Arial",20),anchor="e")
            task_name_text.pack(pady=(5,1),padx=10,anchor="w",side="left")
            task_date_accessed.pack(pady=(5,1),padx=10,anchor="e",side="right")
            task_name.pack(pady=(10,0),padx=5,side="top",fill="x")
            task_name.bind("<Button-3>",lambda e,widget = "name",id=i: self.show_context_menu(child_root,e,widget,id))
            task_name_text.bind("<Button-3>",lambda e,widget = "name",id=i: self.show_context_menu(child_root,e,widget,id))

            task_frame = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=3,height= 50,border_color="#636363")
            param1_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=0,height= 50)
            param1_subframe1 = customtkinter.CTkFrame(master=param1_frame,corner_radius=0,border_width=2,height= 50,width=230)
            param1_label = customtkinter.CTkLabel(master=param1_subframe1,text = "Čas spuštění (denně): ",font=("Arial",20,"bold"),anchor="w")
            param1_label.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param1_subframe2 = customtkinter.CTkFrame(master=param1_frame,corner_radius=0,border_width=2,height= 50)
            param1_label2 = customtkinter.CTkLabel(master=param1_subframe2,text = str(tasks[4]),font=("Arial",20),anchor="w")
            param1_label2.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param1_subframe1.pack(side="left")
            param1_subframe1.pack_propagate(0)
            param1_subframe2.pack(side="left",fill="x",expand=True)
            param1_frame.pack(pady=(3,0),padx=3,fill="x",side="top")
            param1_label2.bind("<Button-3>",lambda e,widget = "time",id=i: self.show_context_menu(child_root,e,widget,id))
            param1_label.bind("<Button-3>",lambda e,widget = "time",id=i: self.show_context_menu(child_root,e,widget,id))
            param1_subframe1.bind("<Button-3>",lambda e,widget = "time",id=i: self.show_context_menu(child_root,e,widget,id))
            param1_subframe2.bind("<Button-3>",lambda e,widget = "time",id=i: self.show_context_menu(child_root,e,widget,id))

            param2_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=1,height= 50)
            param2_subframe1 = customtkinter.CTkFrame(master=param2_frame,corner_radius=0,border_width=2,height= 50,width=230)
            param2_label = customtkinter.CTkLabel(master=param2_subframe1,text = "Pracuje v: ",font=("Arial",20,"bold"),anchor="w")
            param2_label.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param2_subframe2 = customtkinter.CTkFrame(master=param2_frame,corner_radius=0,border_width=2,height= 50)
            param2_label2 = customtkinter.CTkLabel(master=param2_subframe2,text = str(tasks[1]),font=("Arial",20),anchor="w")
            param2_label2.pack(pady=10,padx=(10,3),anchor="w",side="left")
            param2_subframe1.pack(side="left")
            param2_subframe1.pack_propagate(0)
            param2_subframe2.pack(side="left",fill="x",expand=True)
            param2_frame.pack(pady=(0,0),padx=3,fill="x",side="top")
            param2_label2.bind("<Button-3>",lambda e,widget = "path",id=i: self.show_context_menu(child_root,e,widget,id))
            param2_subframe2.bind("<Button-3>",lambda e,widget = "path",id=i: self.show_context_menu(child_root,e,widget,id))
            param2_subframe1.bind("<Button-3>",lambda e,widget = "path",id=i: self.show_context_menu(child_root,e,widget,id))

            param3_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=1,height= 50)
            param3_label = customtkinter.CTkLabel(master=param3_frame,text = "Nastavení: ",font=("Arial",20,"bold"),anchor="w")
            param3_label2 = customtkinter.CTkLabel(master=param3_frame,text = f"starší než: {tasks[2]} dní, minimum = {tasks[3]} souborů",font=("Arial",20),anchor="w")
            param3_label.pack(pady=10,padx=(10,0),anchor="w",side="left")
            param3_label2.pack(pady=10,padx=(10,0),anchor="w",side="left")
            param3_frame.pack(pady=(0,3),padx=3,fill="x",side="top")
            param3_label.bind("<Button-3>",lambda e,widget = "settings",id=i: self.show_context_menu(child_root,e,widget,id))
            param3_label2.bind("<Button-3>",lambda e,widget = "settings",id=i: self.show_context_menu(child_root,e,widget,id))
            param3_frame.bind("<Button-3>",lambda e,widget = "settings",id=i: self.show_context_menu(child_root,e,widget,id))
            task_frame.pack(pady=(0,0),padx=5,fill="x",side="top")
            i+=1

        if len(all_tasks) == 0:
            task_label = customtkinter.CTkLabel(master=main_frame,text = "Nejsou nastaveny žádné úkoly...",font=("Arial",22,"bold"),anchor="w")
            task_label.pack(pady=10,padx=10,fill="x",side="top",anchor="w")
            child_root.after(2000, lambda: child_root.destroy())
        # main_frame.pack(fill="both",side="top")
        main_frame.pack(fill="both",side="top",expand=True)
        child_root.update()
        child_root.update_idletasks()
        # child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()+10}")
        child_root.geometry(f"{1200}x{800}")
        child_root.focus_force()
        child_root.focus()
        child_root.mainloop()

    def show_task_log(self,specify_task=False,task_given = None,root_given = False):
        if not root_given:
            child_root = customtkinter.CTk()
            child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
            child_root.title("Záznam o vymazaných souborech")
        else:
            child_root = root_given
            self.clear_frame(child_root)
        main_frame = customtkinter.CTkScrollableFrame(master=child_root,corner_radius=0)
        self.check_task_existence()
        current_tasks = self.read_config()

        def hide_details(task,given_task_frame,button):
            self.clear_frame(given_task_frame)
            button.configure(text="v")
            given_task_frame.configure(height=0)
            given_task_frame.update()
            given_task_frame.update_idletasks()
            button.unbind("<Button-1>")
            button.bind("<Button-1>",lambda e,tasks = task, log_frame = given_task_frame, button_details = button: show_details(tasks,log_frame,button_details))
            
        def show_details(task,given_task_frame,button,get_log_count = False):
            all_task_logs = task[6].split("|||")
            all_task_logs.pop(0) #nultá pozice v poli vždy prázdná
            if get_log_count:
                return len(all_task_logs)
            for logs in all_task_logs:
                log_data = logs.split("||")
                print(log_data)
                label_data = str(log_data[0])+"\n"+str(log_data[1])+"\n"+str(log_data[2])+"\n"+str(log_data[3])
                log_frame = customtkinter.CTkFrame(master=given_task_frame,corner_radius=0,border_width=2)
                log_text = customtkinter.CTkLabel(master=log_frame,text = label_data,font=("Arial",20),anchor="w",justify="left")
                log_text.pack(pady=(10,5),padx=10,anchor="w",side="top")
                log_frame.pack(pady=(0),padx=0,fill="x",anchor="w",side="top")

            button.configure(text="^")
            button.unbind("<Button-1>")
            button.bind("<Button-1>",lambda e,tasks = task, log_frame = given_task_frame, button_details = button: hide_details(tasks,log_frame,button_details))
        
        i=0
        for tasks in current_tasks:
            if specify_task:
                if tasks[0] != task_given:
                    i+=1
                    continue #preskoč když se nejedná o hledaný specifický task
            task_frame = customtkinter.CTkFrame(master=main_frame,corner_radius=0,border_width=0)
            header_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=0,fg_color="#636363")
            task_name_text = customtkinter.CTkLabel(master=header_frame,text = "Úkol "+str(i+1) + f" (scheduler název: {tasks[0]}), přidáno: {tasks[5]}",font=("Arial",20,"bold"),anchor="w",justify="left")
            empty_log_frame = customtkinter.CTkFrame(master=task_frame,corner_radius=0,border_width=0,height=0)
            button_details = customtkinter.CTkButton(master = header_frame,text = "v",font=("Arial",40,"bold"),width = 50,height=50,corner_radius=0,fg_color="#505050")
            button_details.bind("<Button-1>",lambda e,task = tasks, log_frame = empty_log_frame, button = button_details: show_details(task,log_frame,button))
            task_name_text.pack(pady=(5,1),padx=10,anchor="w",side="left")
            button_details.pack(pady=(5),padx=5,anchor="e",side="right")
            header_frame.pack(pady=0,padx=0,anchor="w",side="top",fill="x")
            empty_log_frame.pack(pady=(0),padx=0,side="top",anchor="w",fill="x",expand = True)
            task_frame.pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill="x",expand = True)
            button_details.pack_propagate(0)
            header_frame.bind("<Button-3>",lambda e,widget = "del_log",id=i: self.show_context_menu(child_root,e,widget,id))
            task_name_text.bind("<Button-3>",lambda e,widget = "del_log",id=i: self.show_context_menu(child_root,e,widget,id))
            i+=1
            if specify_task:
                show_details(tasks,empty_log_frame,button_details) #rovnou otevřít (zobrazit detaily)

            if show_details(tasks,None,None,get_log_count=True) == 0:
                button_details.configure(state="disabled")

        if len(self.task_log_list) == 0:
            log_text = customtkinter.CTkLabel(master=main_frame,text = "Nebyl nalezen žádný záznam",font=("Arial",22,"bold"),anchor="w")
            log_text.pack(pady=10,padx=10,fill="x",side="top",anchor="w")
            child_root.after(2000, lambda: child_root.destroy())

        main_frame.pack(fill="both",side="top",expand=True)
        child_root.update()
        child_root.update_idletasks()
        child_root.geometry(f"{1200}x{800}")
        child_root.mainloop()

    def create_menu(self):
        def call_main_app():
            command = "\"" + self.initial_path + self.main_app_exe_name + "\""
            command = command.replace("/","\\")
            # subprocess.call(command,shell=True,text=True)
            subprocess.Popen(command, shell=True, text=True, creationflags=subprocess.CREATE_NO_WINDOW)

        self.menu = Menu(MenuItem('Spustit aplikaci TRIMAZKON', lambda: call_main_app()),
                         MenuItem('Zobrazit nastavené úkoly', lambda: self.show_all_tasks()),
                         MenuItem('Záznamy o mazání', lambda: self.show_task_log()),
                         MenuItem('Vypnout', lambda: self.quit_application()))

    def quit_application(self):
        self.icon.stop()
        try:
            sys.exit()
        except Exception:
            pass

    def main(self):
        def create_image():
            image = Image.open(self.app_icon)
            return image
        
        self.create_menu()
        self.icon = Icon(
            "TRIMAZKON_tooltip",
            create_image(),
            "TRIMAZKON",
            self.menu
        )
        self.icon.run() # Run the tray icon

def handle_system_arguments():
    def get_init_path(raw_path):
        """
        removes the last item with application from path
        """
        initial_path = Tools.path_check(raw_path,True)
        initial_path_splitted = initial_path.split("/")
        initial_path = ""
        for i in range(0,len(initial_path_splitted)-2):
            initial_path += str(initial_path_splitted[i])+"/"
        print("SYSTEM: ",sys.argv)

        return initial_path
    
    initial_path = Tools.path_check(sys.argv[1])
    resource_app_path = get_init_path(sys.argv[0])
    tray_app_instance = tray_app_service(initial_path,resource_app_path)
    print(sys.argv, len(sys.argv))
    if str(sys.argv[2]) == "run_tray":
        tray_app_instance.main()
        sys.exit()

    elif str(sys.argv[2]) == "check_task_existence":
        task_given = str(sys.argv[3])
        output_status = tray_app_instance.check_task_existence(task_given=task_given)
        print("output check task existance status: ",output_status)
        sys.exit(output_status)
        # return output_status

    elif str(sys.argv[2]) == "save_new_log":
        task_name = str(sys.argv[3])
        new_log = str(sys.argv[4])
        tray_app_instance.save_new_log(task_name,new_log)

    elif str(sys.argv[2]) == "read_config":
        output_data = tray_app_instance.read_config()
        print("output read_config: ",output_data)
        sys.exit(output_data)
        # return output_data
    
    elif str(sys.argv[2]) == "show_all_tasks":
        tray_app_instance.show_all_tasks(toplevel=True)

# handle_system_arguments() # uncoment when making exe file

# inst = tray_app_service(r"C:\Users\jakub.hlavacek.local\Desktop\JHV\Work\TRIMAZKON/")
# inst.main()

# CREATING TASK:
# name_of_task = "dailyscript_test"
# path_to_app = r"C:\Users\jakub.hlavacek.local\Desktop\JHV\Work\TRIMAZKON\pipe_server\untitled2.py"
# cmd_command = f"schtasks /Create /TN {name_of_task} /TR {path_to_app} /SC DAILY /ST 09:35"
# connection_status = subprocess.call(cmd_command,shell=True,text=True)

#DELETING TASK:
# name_of_task = "dailyscript_test"
# cmd_command = f"schtasks /Delete /TN {name_of_task} /F"
# connection_status = subprocess.call(cmd_command,shell=True,text=True)
