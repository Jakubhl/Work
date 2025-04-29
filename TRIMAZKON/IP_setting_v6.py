import customtkinter
import tkinter as tk
from openpyxl import load_workbook
from openpyxl import Workbook
import subprocess
import os
import re
import time
import threading
import psutil
import socket
from PIL import Image
import sys
import ctypes
import winreg
import win32net
import copy
import pyperclip
import json

testing_mode = False

if testing_mode:
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    root=customtkinter.CTk()
    root.geometry("1200x900")
    root.title("ip_setting - testing")
    root.state('zoomed')

class Tools:
    config_json_filename = "jhv_IP.json"
    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        # if hasattr(sys, '_MEIPASS'):
        #     return os.path.join(sys._MEIPASS, relative_path)
        # return os.path.join(os.path.abspath("."), relative_path)
        BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.abspath(".")
        return os.path.join(BASE_DIR, relative_path)
    
    @classmethod
    def add_colored_line(cls,text_widget, text, color,font=None,delete_line = None):
        """
        Vloží řádek do console
        """
        text_widget.configure(state=tk.NORMAL)
        if font == None:
            font = ("Arial",16)
        if delete_line != None:
            text_widget.delete("current linestart","current lineend")
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert("current lineend",text, color)
        else:
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert(tk.END,"    > "+ text+"\n", color)

        text_widget.configure(state=tk.DISABLED)
    
    @classmethod
    def read_json_config(cls,config_file_path): # Funkce vraci data z configu
        """
        Funkce vrací data z konfiguračního souboru

        \nIP SETTINGS\n
        - default_ip_interface
        - favorite_ip_window_status
        - disk_or_ip_window
        - default_window_size
        - init_disk_refresh
        - editable_notes
        - disk_persistent
        - auto_order_when_edit
        - ask_to_delete

        - edited_project_bin
        - deleted_project_bin
        \n
        - [0] = default connection option (0/1)
        - [1] = show favourite ip as default (0/1)
        - [2] = show disk environment as default (0/1)
        - [3] = last set widnow size  (0/1/2 - 2 is the narrow and long one)
        - [4] = check disk statutes automatically status (0/1)
        - [5] = editable/ non-editable notes (0/1)
        - [6] = persistent/ non-persistent disk (0/1)
        - [7] = shift edited project on top status (0/1)
        - [8] = delete - pop up window main window (110), when edit (101)
        """

        if os.path.exists(config_file_path):
            try:
                output_data = []
                with open(config_file_path, "r") as file:
                    output_data = json.load(file)

                return output_data["ip_settings"]

            except Exception as e:
                print(f"Nejdřív zavřete soubor {cls.config_json_filename} Chyba: {e}")   
                print("Budou načteny defaultní hodnoty")
                return
        else:
            print(f"Chybí konfigurační soubor {cls.config_json_filename}")
            return
        
    @classmethod
    def save_to_json_config(cls,which_parameter,input_data,config_file_path,language_force = "cz",which_settings="ip_settings"): # Funkce zapisuje data do souboru configu
        """
        Funkce zapisuje data do konfiguračního souboru

        vraci vystupni zpravu: report

        \nIP_SETTINGS\n
        - default_ip_interface
        - favorite_ip_window_status
        - disk_or_ip_window
        - default_window_size
        - init_disk_refresh
        - editable_notes
        - disk_persistent
        - auto_order_when_edit
        - ask_to_delete

        - edited_project_bin
        - deleted_project_bin
        """

        def get_input_data_format():
            if isinstance(input_data,list):
                return input_data
            elif isinstance(input_data,str):
                return str(input_data)
            elif isinstance(input_data,int):
                return int(input_data)
        
        if os.path.exists(config_file_path):
            with open(config_file_path, "r") as file:
                config_data = json.load(file)

            config_data[which_settings][which_parameter] = get_input_data_format()
                              
            with open(config_file_path, "w") as file:
                json.dump(config_data, file, indent=4)
        
        else:
            print("Chybí konfigurační soubor (nelze ukládat změny)")
            return
   
    @classmethod
    def clear_frame(cls,frame):
        frame.update()
        frame.update_idletasks()
        for widget in frame.winfo_children():
            if widget.winfo_exists():
                widget.unbind("<Enter>")
                widget.unbind("<Leave>")
                widget.unbind("<Return>")
                widget.unbind("<Button-1>")
                widget.unbind("<Button-3>")
                widget.unbind("<Double-1>")
                widget.unbind("<MouseWheel>")
                widget.destroy()
    
    @classmethod
    def get_legit_notes(cls,notes):
        notes_legit_rows = []
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    @classmethod
    def focused_entry_widget(cls,root):
        currently_focused = str(root.focus_get())
        if ".!ctkentry" in currently_focused or ".!ctktextbox" in currently_focused:
            return True
        else:
            return False

    @classmethod
    def get_none_count(cls,array_given):
        none_count = 0
        for items in array_given:
            if items == None:
                none_count += 1
        return none_count

    @classmethod
    def path_check(cls,path_raw,only_repair = None):
        path=path_raw
        backslash = "\\"
        if backslash[0] in path:
            newPath = path.replace(backslash[0], '/')
            path = newPath
        if path.endswith('/') == False:
            newPath = path + "/"
            path = newPath
        #oprava mezery v nazvu
        path = r"{}".format(path)
        if not os.path.exists(path) and only_repair == None:
            return False
        else:
            return path

    @classmethod
    def browseDirectories(cls,visible_files,start_path=None,file_type = [("All files", "*.*")]): # Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat
        """
        Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat

        Vstupní data:

        0: visible_files = "all" / "only_dirs"\n
        1: start_path = None -optimalni, docasne se ulozi posledni nastavena cesta v exploreru

        Výstupní data:

        0: výstupní chybová hlášení
        1: opravená cesta
        2: nazev vybraneho souboru (option: all)
        """
        corrected_path = ""
        output= ""
        name_of_selected_file = ""

        # if start_path == None:
        #     start_path = Tools.read_json_config()["app_settings"]["default_path"] #defaultni cesta
        # else: # byla zadana docasna cesta pro explorer
        start_path = Tools.path_check(start_path)
            # if checked_path == False:
            #     output = "Změněná dočasná základní cesta pro explorer již neexistuje"
            #     start_path = Tools.read_json_config()["app_settings"]["default_path"] #defaultni cesta
            # else:
            #     start_path = checked_path

        if start_path != False:
            if not os.path.exists(start_path):
                start_path = ""
                output="Konfigurační soubor obsahuje neplatnou cestu"

        else:
            output="Chybí konfigurační soubor config_TRIMAZKON.xlsx s počáteční cestou...\n"
            start_path=""

        # pripad vyberu files, aby byly viditelne
        if visible_files == "all":
            if(start_path != ""):
                foldername_path = tk.filedialog.askopenfile(initialdir = start_path,
                                                            title = "Klikněte na soubor v požadované cestě",
                                                            filetypes=file_type)
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:           
                foldername_path = tk.filedialog.askopenfile(initialdir = "/",
                                                            title = "Klikněte na soubor v požadované cestě",
                                                            filetypes=file_type)
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"

        # pripad vyberu slozek
        if visible_files == "only_dirs":
            if(start_path != ""):
                path_to_directory = tk.filedialog.askdirectory(initialdir = start_path, title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:
                path_to_directory = tk.filedialog.askdirectory(initialdir = "/", title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"

        check = Tools.path_check(path_to_directory)
        corrected_path = check
        return [output,corrected_path,name_of_selected_file]

    @classmethod
    def import_option_window(cls,root,app_icon_path,default_path,callback,ip_env = False,setting_window=None):
        child_root = customtkinter.CTkToplevel(fg_color="#212121")
        child_root.after(200, lambda: child_root.iconbitmap(app_icon_path))
        child_root.title("Možnosti exportování souboru")

        def get_excel_path():
            path_inserted = import_path.get()
            path_inserted = Tools.resource_path(path_inserted)
            if path_inserted.replace(" ","") == "":
                return None
            else:
                if path_inserted.endswith(".xlsx"):
                    return path_inserted
                else:
                    return path_inserted + ".xlsx"

        def close_window(child_root):
            try:
                root.unbind("<Button-1>")
            except Exception:
                pass
            child_root.destroy()

        def call_browse_directories():
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            output = Tools.browseDirectories("all",str(import_path.get()),file_type=[("xlsx files", "*.xlsx"),("All files", "*.*")])
            if str(output[1]) != "/":
                import_path.delete(0,300)
                import_path.insert(0, str(output[1])+str(output[2]))
                Tools.add_colored_line(console,"Byla vložena cesta pro uložení","green",None,True)
            print(output[0])

            if setting_window != None:
                setting_window.focus()
                setting_window.focus_force()
            child_root.focus()
            child_root.focus_force()

        def load_data():
            path_to_send = get_excel_path()
            if os.path.exists(path_to_send) and path_to_send.endswith(".xlsx"):
                if ip_env:
                    callback(path_to_send,all_data_checkbox.get())
                else:
                    callback(path_to_send)
                child_root.destroy()
            else:
                Tools.add_colored_line(console,"Neplatná cesta k souboru (hledaný je .xlsx soubor)","red",None,True)

        import_frame =      customtkinter.CTkFrame(master = child_root,corner_radius=0,fg_color="#212121")
        import_label =      customtkinter.CTkLabel(master = import_frame,text = "Zadejte cestu, kam soubor uložit:",font=("Arial",22,"bold"))
        import_path_frame = customtkinter.CTkFrame(master = import_frame,corner_radius=0,fg_color="#212121")
        import_path =       customtkinter.CTkEntry(master = import_path_frame,font=("Arial",20),width=780,height=50,corner_radius=0)
        explorer_btn =      customtkinter.CTkButton(master = import_path_frame,text = "...",font=("Arial",22,"bold"),width = 50,height=50,corner_radius=0,command=lambda: call_browse_directories())
        import_path         .pack(pady = 5, padx = (10,0),anchor="w",fill="x",expand=True,side="left")
        explorer_btn        .pack(pady = 5, padx = (0,10),anchor="e",expand=False,side="right")
        all_data_checkbox = customtkinter.CTkCheckBox(master= import_frame,text = "Načíst rovnou i adresy disků?",font=("Arial",22,"bold"))
        console =           tk.Text(import_frame, wrap="none", height=0, width=30,background="black",font=("Arial",22),state=tk.DISABLED)
        button_save =       customtkinter.CTkButton(master = import_frame,text = "Načíst",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: load_data())
        button_exit =       customtkinter.CTkButton(master = import_frame,text = "Zrušit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,command=lambda: close_window(child_root))

        import_frame        .pack(pady = 0, padx = 0,fill="both",anchor="n",expand=True,side="left")
        import_label       .pack(pady=(10,5),padx=10,anchor="w",expand=False,side="top")
        import_path_frame   .pack(expand=True,side="top",anchor="n",fill="x")
        if ip_env:
            all_data_checkbox.pack(pady = 5, padx = (10),anchor="w",side="top")
        console             .pack(padx = 5,expand=True,side="top",anchor="n",fill="x")
        button_exit         .pack(pady = 10, padx = (5,10),expand=False,side="right",anchor = "e")
        button_save         .pack(pady = 10, padx = 5,expand=False,side="right",anchor = "e")

        if os.path.exists(default_path):
            import_path.insert("0",str(default_path))
            Tools.add_colored_line(console,"Byla vložena uložená cesta z konfiguračního souboru","green",None,True)

        root.bind("<Button-1>",lambda e: close_window(child_root))
        child_root.update()
        child_root.update_idletasks()
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{x+200}+{y+100}")
        child_root.focus()
        child_root.focus_force()

    @classmethod
    def get_unique_name(cls,project_list,project_name):
        project_name_list = []
        found_count=0
        for projects in project_list:
            project_name_list.append(str(projects['name']))
            if str(projects['name']) == str(project_name):
                found_count +=1
        i=0
        modified_project_name = project_name
        if modified_project_name in project_name_list:
            while modified_project_name in project_name_list:
                i+=1
                modified_project_name = str(project_name)+" ("+str(i)+")"

        return modified_project_name
    
    @classmethod
    def get_project_index(cls,project_list,project_name):
        name_list = []
        for projects in project_list:
            name_list.append(str(projects['name']))

        try:
            found_index = name_list.index(project_name)
            return found_index
        except ValueError:
            return None
        
    @classmethod
    def found_project_name(cls,project_list,project_name):
        for projects in project_list:
            if str(projects['name']) == str(project_name):
                return projects
        return False
    
class main:
    class DM_tools:  
        @classmethod
        def check_network_drive_status(cls,drive_path):
            checking_done = False
            status = False
            try:
                # Attempt to access a file or directory on the network drive
                drive_path = drive_path[0:3]
                def call_subprocess():
                    nonlocal checking_done
                    nonlocal status
                    if os.path.exists(drive_path):
                        os.listdir(drive_path)
                        status = True
                        checking_done = True
                        return
                    else:
                        status = False
                        checking_done = True
                        return
                    
                run_background = threading.Thread(target=call_subprocess,)
                run_background.start()

                time_start = time.time()
                while checking_done==False:
                    time.sleep(0.05)
                    if time.time() - time_start > 1:
                        print("terminated due to runtime error")
                        return False
                
                if status == True:
                    return True
                else:
                    return False

            except FileNotFoundError:
                return False
            except OSError:
                return False
            
        @classmethod
        def list_mapped_disks(cls,whole_format=None):
            remote_drives = []
            try:
                reg_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Network')
                for i in range(0, winreg.QueryInfoKey(reg_key)[0]):
                    drive_letter = winreg.EnumKey(reg_key, i)
                    if whole_format:
                        remote_drives.append(drive_letter + ':')
                    else:
                        remote_drives.append(drive_letter)
                winreg.CloseKey(reg_key)
            except Exception as e:
                print("Exception occurred: ", e)

            print("persistent disks: ",remote_drives)
                    
            return remote_drives

        @classmethod
        def list_non_persistent_disks(cls):
            non_persistent_drives = []
            try:
                # Enumerate network connections
                level = 1  # Level 1 provides the 'ui1_flags' information
                connections, _, _ = win32net.NetUseEnum(None, level)
                for i in range(0,len(connections)):
                    non_persistent_drives.append(connections[i]["local"]) 
            except Exception as e:
                print("Exception occurred: ", e)

            print("non-persistent disks: ",non_persistent_drives)
            
            return non_persistent_drives

        @classmethod
        def save_excel_data_disk(cls,
                                excel_file_path,
                                project_list,
                                wb_given = None):
            if wb_given == None:
                workbook = load_workbook(excel_file_path)
            else:
                workbook = wb_given

            worksheet = workbook["disk_list"]
         
            for i in range(0,len(project_list)):
                notes = str(project_list[i]["notes"])
                if notes == None or notes.replace(" ","") == "":
                    notes = ""
                worksheet['A' + str(i+1)] = str(project_list[i]['name']) #A = nazev projektu
                worksheet['B' + str(i+1)] = str(project_list[i]['disk_letter'])  #B = písmeno disku, označení...
                worksheet['C' + str(i+1)] = str(project_list[i]['ftp'])  #C = ftp adresa
                worksheet['D' + str(i+1)] = str(project_list[i]['user'])  #D = uživatelské jméno
                worksheet['E' + str(i+1)] = str(project_list[i]['password'])  #E = heslo
                worksheet['F' + str(i+1)] = notes #F = poznamky

            workbook.save(filename = excel_file_path)
            if wb_given == None:
                workbook.close()
        
        @classmethod
        def read_excel_data(cls,excel_file_path):
            """
            returns obj:
            - name
            - disk_letter
            - ftp
            - user
            - password
            - notes
            """
            workbook = load_workbook(excel_file_path,read_only=True)
            worksheet = workbook["disk_list"]
            project_list = []

            for row in worksheet.iter_rows(values_only=True):
                row_array = []
                for items in row[:6]:
                    if items is None:
                        row_array.append("")
                    else:
                        row_array.append(str(items))

                project_object = {
                    'name':row_array[0],
                    "disk_letter":row_array[1],
                    "ftp":row_array[2],
                    "user":row_array[3],
                    "password":row_array[4],
                    "notes":row_array[5],
                }
                project_list.append(project_object)
            workbook.close()
            return project_list

    class IP_tools:
        @classmethod
        def save_excel_data(cls,
                            excel_file_path,
                            project_list,
                            wb_given = None):
            if wb_given == None:
                workbook = load_workbook(excel_file_path)
            else:
                workbook = wb_given

            excel_worksheet = "ip_address_list"
            worksheet = workbook[excel_worksheet]
            
            for i in range(0,len(project_list)):
                notes = str(project_list[i]["notes"])
                if notes == None or notes.replace(" ","") == "":
                    notes = ""
                worksheet['A' + str(i+1)] = str(project_list[i]['name']) #A = nazev projektu
                worksheet['B' + str(i+1)] = str(project_list[i]["ip"]) #B = ip adresa
                worksheet['C' + str(i+1)] = str(project_list[i]["mask"]) #C = maska
                worksheet['D' + str(i+1)] = notes #D = poznamky
                worksheet['E' + str(i+1)] = str(project_list[i]["fav_status"]) #E = oblibenost

            workbook.save(filename=excel_file_path)
            if wb_given == None:
                workbook.close()

        @classmethod
        def read_excel_data(cls,excel_file_path):
            """
            returns obj:
            - name
            - ip
            - mask
            - notes
            - fav_status
            """
            excel_worksheet = "ip_address_list"
            workbook = load_workbook(excel_file_path,read_only=True)
            worksheet = workbook[excel_worksheet]
            project_list = []
            for row in worksheet.iter_rows(values_only=True):
                row_array = []
                
                for items in row[:4]:
                    if items is None:
                        row_array.append("")
                    else:
                        row_array.append(str(items))

                if "1" in str(row[4:5]):
                    row_array.append(str(1))
                else:
                    row_array.append(str(0))

                project_object = {
                    'name':row_array[0],
                    "ip":row_array[1],
                    "mask":row_array[2],
                    "notes":row_array[3],
                    "fav_status":row_array[4],
                }
                project_list.append(project_object)

            workbook.close()
            # print(project_list)
            return project_list

        @classmethod
        def get_current_ip_list(cls,connection_option_list:list):
            def get_current_ip_address(interface_name):
                # Get network interfaces and their addresses
                addresses = psutil.net_if_addrs()
                # Check if the specified interface exists
                if interface_name in addresses:
                    addr_count = 0
                    # print(f"Adresy interfacu: {interface_name}")
                    for addr in addresses[interface_name]:
                        # prvni AF_INET je pridelena automaticky, druha je privatni, nastavena DHCP
                        if addr.family == socket.AF_INET:  # IPv4 address
                            if addr_count == 1:
                                return addr.address
                            addr_count +=1
                            remembered_addr = addr
                    if addr_count == 1:
                        if "AddressFamily.AF_INET:" in str(remembered_addr):
                            return remembered_addr.address
                        else:
                            return "Nenalezeno"
                else:
                    return "Nenalezeno"
            current_address_list = []
            for items in connection_option_list:
                found_address = get_current_ip_address(items)
                current_address_list.append(found_address)
            return current_address_list
        
        @classmethod
        def fill_interfaces(cls):
            """
            Vrátí:
            - seznam interfaců [0]
            - seznam připojených interfaců [1]
            """
            process = subprocess.Popen("netsh interface show interface",
                                                        stdout=subprocess.PIPE,
                                                        stderr=subprocess.PIPE,
                                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            stdout_str = stdout.decode("cp852",errors="ignore")
            print(stdout_str)
            data = str(stdout_str)
            lines = data.strip().splitlines()
            interfaces = []
            interface_statuses =[]
            # Process each line starting from the second line
            for line in lines[2:]:  # Skip the first two lines (headers and separator)
                # Split the line based on spaces
                values = line.split()
                if len(values) >= 4:  # Ensure there are enough columns
                    # Combine the last columns into Interface Name, handle spaces in the name
                    interface_name = ' '.join(values[3:])
                    interface_statuses.append(values[1])
                    interfaces.append(interface_name)

            print("interface list: ",interfaces)
            print("status: ", interface_statuses)
            connected_interfaces =[]
            for i in range(0,len(interface_statuses)):
                if interface_statuses[i] != "Odpojen" and interface_statuses[i] != "Odpojeno" and interface_statuses[i] != "Disconnected":
                    if interfaces[i] not in connected_interfaces:
                        connected_interfaces.append(interfaces[i])
            print("online: ", connected_interfaces)
            
            return [interfaces,connected_interfaces]

        @classmethod
        def get_ipv4_addresses(cls):
            """
            returns list of ipv4 addresses of all interfaces
            """
            process = subprocess.Popen("ipconfig",
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            result2 = "" 
            try:
                stdout_str = stdout.decode('cp1250')
                result2 = str(stdout_str)
                
            except Exception as e:
                print("chyba ",e)

            # Regular expression to match the IPv4 address
            ipv4_pattern = re.compile(r'IPv4 Address[.\s]*: ([\d.]+)')
            ipv4_addresses = []
            lines = result2.splitlines()
            current_interface = None
            # Iterate over each line to find interface names and IPv4 addresses
            for line in lines:
                if line.strip():
                    # Detect interface name
                    if line[0].isalpha():
                        current_interface = line.strip()
                    else:
                        # Detect IPv4 address for the current interface
                        match = ipv4_pattern.search(line)
                        if match and current_interface:
                            ipv4_addresses.append(current_interface)
                            ipv4_addresses.append(match.group(1))
            return ipv4_addresses

        @classmethod
        def check_DHCP(cls,interface):
            process = subprocess.Popen(f'netsh interface ip show config name="{interface}"',
                                                        stdout=subprocess.PIPE,
                                                        stderr=subprocess.PIPE,
                                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            try:
                stdout_str = stdout.decode('utf-8')
                output_data = str(stdout_str)
            except UnicodeDecodeError:
                try:
                    stdout_str = stdout.decode('cp1250')
                    output_data = str(stdout_str)
                except UnicodeDecodeError:
                    output_data = str(stdout)

            output_data_lines = output_data.split("\n")
            for lines in output_data_lines:
                if "DHCP enabled" in lines and "Yes" in lines:
                    print(f"{interface} DHCP: yes")
                    return True
                elif "DHCP povoleno" in lines and "Ano" in lines:
                    print(f"{interface} DHCP: yes")
                    return True
            print(f"{interface} DHCP: no")

        @classmethod
        def get_favourite_ips_addr(cls,excel_file_path):
            all_projects = cls.read_excel_data(excel_file_path)
            fav_ip_list = []
            for projects in all_projects:
                if str(projects["fav_status"]) == "1":
                    fav_ip_list.append(str(projects["ip"])+" | "+str(projects["name"]))

            return fav_ip_list

        @classmethod
        def change_to_DHCP(cls,interface,interface_ip,callback_function):
            def delay_the_refresh():
                new_addr = cls.get_current_ip_list([interface])[0]
                i = 0
                while new_addr == previous_addr or new_addr == None:
                    print(f"Čekám, až windows provede změny: {7-i} s...")
                    time.sleep(1)
                    new_addr = cls.get_current_ip_list([interface])[0]
                    print("current addr: ",new_addr)
                    if i > 6:
                        output_message = f"Chyba, u {interface} se nepodařilo změnit ip adresu na DHCP (pro nastavování odpojených interfaců spusťte aplikaci jako administrátor)"
                        callback_function(output_message)
                        return
                    i+=1
                    
                output_message = f"IPv4 adresa interfacu: {interface} úspěšně přenastavena na DHCP (automatickou)"
                callback_function(output_message)
                return
            
            output_message = ""
            if not cls.check_DHCP(interface):
                if interface != None or interface != "":
                    previous_addr = interface_ip
                    try:
                        # Construct the netsh command
                        netsh_command = f"netsh interface ipv4 set address name=\"{interface}\" source=dhcp"
                        print(f"calling: {netsh_command}")
                        powershell_command = [
                            'powershell.exe',
                            '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                        ]
                        process = subprocess.Popen(powershell_command,
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
                        
                        stdout, stderr = process.communicate()
                        stdout_str = stdout.decode('utf-8')
                        stderr_str = stderr.decode('utf-8')
                        if stderr_str:
                            print(f"Error occurred: {stderr_str}")
                        else:
                            print(f"Command executed successfully:\n{stdout_str}")                
                        
                        run_background = threading.Thread(target=delay_the_refresh,)
                        run_background.start()

                    except Exception as e:
                        output_message = f"Exception occurred: {str(e)}"
                        callback_function(output_message)
                else:
                    output_message = "Nebyl zvolen žádný interface"
                    callback_function(output_message)
            else:
                output_message = f"{interface} má již nastavenou DHCP"
                callback_function(output_message)

        @classmethod
        def change_computer_ip(cls,ip_given,interface_given,interface_ip_given,online_address_list,callback_function):
            """
            button_row - index, kde se nachazi ip a maska v poli: self.all_rows
            """
            def make_sure_ip_changed():
                def call_subprocess():
                    nonlocal output_message
                    try:
                        if ip == interface_ip_given:
                            output_message = f"Pro interface {interface_given} je již tato adresa ({ip}) nastavena"
                            callback_function(output_message)
                            return
                        
                        elif ip in online_address_list:
                            output_message = "Chyba, adresa je již používána pro jiný interface"
                            callback_function(output_message)
                            return
                        
                        win_change_ip_time = 7
                        for i in range(0,win_change_ip_time):
                            print(f"Čekám, až windows provede změny: {7-i} s...")
                            # Tools.add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                            current_interface_ip = cls.get_current_ip_list([interface_given])[0]
                            print(ip,current_interface_ip)
                            if ip == current_interface_ip: # někdy dříve než 7 sekund...
                                break
                            time.sleep(1)

                        if ip == current_interface_ip:
                            # Tools.add_colored_line(self.main_console,f"IPv4 adresa u {interface_given} byla přenastavena na: {ip}","green",None,True)
                            output_message = f"IPv4 adresa u {interface_given} byla přenastavena na: {ip}"
                            callback_function(output_message)
                        else:
                            output_message = "Chyba, neplatná adresa nebo daný inteface odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťte aplikaci jako administrátor)"
                            callback_function(output_message)
                    except Exception:
                        pass
                
                run_background = threading.Thread(target=call_subprocess,)
                run_background.start()

            def connected_interface(interface,ip,mask):
                """
                Když jsou vyžadována admin práva, tato funkce ověří, zda není daný interface připojen nebo součástí zařízení a zkusí znovu
                """
                nonlocal output_message
                try:
                    # Construct the netsh command
                    netsh_command = f"netsh interface ip set address \"{interface}\" static {ip} {mask}"
                    powershell_command = [
                        'powershell.exe',
                        '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                    ]
                    process = subprocess.Popen(powershell_command,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
                    
                    stdout, stderr = process.communicate()
                    stdout_str = stdout.decode('utf-8')
                    stderr_str = stderr.decode('utf-8')
                    if stderr_str:
                        print(f"Error occurred: {stderr_str}")
                        output_message = "Chyba, nebyla poskytnuta práva (dejte ANO)"
                        callback_function(output_message)
                    else:
                        print(f"Command executed successfully:\n{stdout_str}")
                        make_sure_ip_changed()

                except Exception as e:
                    print(f"Exception occurred: {str(e)}")

            ip = ip_given
            mask = "255.255.255.0"
            interface_name = interface_given
            output_message = ""
            powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
            try:
                process = subprocess.Popen(['powershell.exe', '-Command', powershell_command],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                stdout, stderr =process.communicate()
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                print(f"VÝSTUP Z IP SETTING: {str(stdout_str)}")

                if len(str(stdout_str)) > 7:
                    raise subprocess.CalledProcessError(1, powershell_command, stdout_str)
                if stderr_str:
                    raise subprocess.CalledProcessError(1, powershell_command, stderr_str)
                make_sure_ip_changed()

            except subprocess.CalledProcessError as e:
                if "Run as administrator" in str(stdout_str) or "pustit jako správce" in str(stdout_str):
                    output_message = "Chyba, tato funkce musí být spuštěna s administrátorskými právy"
                    # callback_function(output_message)
                    connected_interface(interface_name,ip,mask)# trigger powershell potvrzení:
                elif "Invalid address" in str(stdout_str) or "Adresa není platná" in str(stdout_str):
                    output_message = "Chyba, neplatná IP adresa"
                    callback_function(output_message)
                else:
                    output_message = "Chyba, Nemáte tuto adresu již nastavenou pro jiný interface?\n(nebo daný interface na tomto zařízení neexistuje)"
                    callback_function(output_message)
            except Exception as e:
                output_message = f"Nastala neočekávaná chyba: {e}"
                callback_function(output_message)
            
            return output_message

        @classmethod
        def manual_ip_setting(cls,app_icon_path,output_callback):
            window = customtkinter.CTkToplevel()
            window.after(200, lambda: window.iconbitmap(Tools.resource_path(app_icon_path)))
            # self.opened_window = window
            window.title("Manuální nastavení IPv4 adresy")

            def check_ip_and_mask(input_value):
                input_splitted = input_value.split(".")
                if len(input_splitted) == 4:
                    return input_value
                else:
                    return False

            def call_ip_change():
                interface = str(select_interface.get())
                interface_ip = cls.get_current_ip_list([interface])[0]
                online_interfaces = cls.fill_interfaces()[1]
                online_addresses = cls.get_current_ip_list(online_interfaces)

                if "DHCP" in str(select_mode.get()):
                    def console_callback(msg):
                        Tools.add_colored_line(manual_console,msg,"white",None,True)
                        output_callback(msg)
                    cls.change_to_DHCP(interface,interface_ip,console_callback)
                    # window.destroy()
                    return
                
                ip_input = ip_address_entry.get()
                mask_input = mask_entry.get()
                ip_checked = check_ip_and_mask(ip_input)
                mask_checked = check_ip_and_mask(mask_input)
                errors = 0
                if ip_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná IP adresa","red",None,True)
                    errors += 1
                if mask_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná maska","red",None,True)
                    errors += 1

                if errors == 0:
                    # self.change_computer_ip(0,force_params=[ip_input,mask_input])
                    def console_callback(msg):
                        Tools.add_colored_line(manual_console,msg,"white",None,True)
                        output_callback(msg)
                    cls.change_computer_ip(ip_input,interface,interface_ip,online_addresses,console_callback)
                    # window.destroy()

            def call_option_change(*args):
                nonlocal ip_address_entry
                # self.interface_drop_options.set(str(*args))
                # self.option_change(*args)
                ip_address_entry.delete(0,300)
                ip_address_entry.insert(0,cls.get_current_ip_list([select_interface.get()])[0])
                check_interface_status()

            def switch_manual_dhcp(*args):
                nonlocal ip_address_entry
                nonlocal mask_entry
                if "DHCP" in str(*args):
                    ip_address_entry.configure(state = "disabled",text_color = "gray32")
                    mask_entry.configure(state = "disabled",text_color = "gray32")
                else:
                    ip_address_entry.configure(state = "normal",text_color = "gray84")
                    mask_entry.configure(state = "normal",text_color = "gray84")

            def check_interface_status(online_list = False):
                if online_list == False:
                    online_list = cls.fill_interfaces()[1]

                found = False
                for items in online_list:
                    if items == str(select_interface.get()):
                        found = True
                        select_interface.configure(fg_color = "green",button_color = "green")
                        interface_status.configure(text = "Online")
                        break

                if not found:
                    select_interface.configure(fg_color = "red",button_color = "red")
                    interface_status.configure(text = "Offline")

            interface_label =       customtkinter.CTkLabel(master = window,text = "Manuálně nastavit IPv4 adresu pro: ",font=("Arial",20,"bold"))
            interface_frame =       customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            select_interface =      customtkinter.CTkOptionMenu(master = interface_frame,width=320,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,command= lambda args:  call_option_change(args))
            interface_status =      customtkinter.CTkLabel(master = interface_frame,text = "",font=("Arial",20,"bold"))
            select_interface.       pack(pady=(10,0),padx=10,side = "left",anchor = "w",fill="x",expand=True)
            interface_status.       pack(pady=(10,0),padx=10,side = "right",anchor = "e")
            mode_label =            customtkinter.CTkLabel(master = window,text = "Způsob nastavení: ",font=("Arial",20,"bold"))
            select_mode =           customtkinter.CTkOptionMenu(master = window,width=400,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,values = ["manuálně","automaticky (DHCP)"],command= lambda args: switch_manual_dhcp(args))
            ip_address =            customtkinter.CTkLabel(master = window,text = "IPv4 adresa: ",font=("Arial",20,"bold"))
            ip_address_entry =      customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            mask =                  customtkinter.CTkLabel(master = window,text = "IPv4 maska: ",font=("Arial",20,"bold"))
            mask_entry =            customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            manual_console =        tk.Text(window, wrap="none", height=0, width=36,background="black",font=("Arial",14),state=tk.DISABLED)
            buttons_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            save_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Nastavit", command = lambda: call_ip_change(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Zrušit", command = lambda: window.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            interface_label.        pack(pady=(10,0),padx=10,side = "top",anchor = "w",expand = False)
            interface_frame.        pack(pady=(0),padx=0,side = "top",anchor = "w",fill="x")
            mode_label.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            select_mode.            pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill="x")
            ip_address.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            ip_address_entry.       pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill="x")
            mask.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            mask_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill="x")
            manual_console.         pack(pady=(10,0),padx=10,side = "top",anchor = "w",fill="x")
            exit_button.            pack(pady=0,padx=(10,0),side = "right",anchor = "w")
            save_button.            pack(pady=0,padx=0,side = "right",anchor = "e")
            buttons_frame.          pack(pady=(10),padx=10,side = "bottom",anchor = "e",fill="x")
        
            online_list = cls.fill_interfaces()[1]
            all_interfaces = cls.fill_interfaces()[0]
            select_interface.configure(values = all_interfaces)
            select_interface.set(all_interfaces[0])
            ip_address_entry.insert(0,cls.get_current_ip_list([all_interfaces[0]])[0])
            mask_entry.insert(0,"255.255.255.0")
            check_interface_status(online_list)
            window.update()
            window.update_idletasks()
            window.geometry(f"{600}x{500}")
            window.focus_force()
            window.focus()

    class ToolTip:
        def __init__(self, widget, text, root,delay =0.2,add_delay = False):
            self.widget = widget
            self.text = text
            self.root = root
            self.tip_window = None
            self.long_task_called = False
            self.delay = delay # in seconds
            self.add_delay = add_delay
            self.bind_it()

        def bind_it(self):
            self.widget.bind("<Enter>",lambda e,widget = self.widget: self.really_entering(e,widget))
            self.widget.bind("<Leave>",lambda e,widget = self.widget: self.really_leaving(e,widget))
            self.widget.bind("<Button-1>",lambda e: self.just_destroy(e))

        def just_destroy(self,e,unbind=False):
            try:
                if unbind:
                    self.widget.unbind("<Enter>")
                    self.widget.unbind("<Leave>")
                    self.widget.unbind("<Button-1>")
                # self.tip_window.destroy()
                self.root.after(0,self.tip_window.destroy)
            except Exception as ee:
                pass

        def really_entering(self,e,widget):
            if self.tip_window != None or self.long_task_called:
                return

            def show_tooltip():
                # x = widget.winfo_rootx() + 50
                # y = widget.winfo_rooty() + int(widget.winfo_height())/2
                self.widget.master.update_idletasks()
                x = self.widget.winfo_rootx()+self.widget._current_width
                y = self.widget.winfo_rooty()+self.widget._current_height/2
                self.tip_window = customtkinter.CTkLabel(
                    self.root,
                    text=self.text,
                    font=("Arial", 20),
                    text_color="black",
                    bg_color= "white"

                )
                # self.tip_window.place(x=x,y=y)
                self.tip_window.place(x=x,y=y)
            show_tooltip()
            self.tip_window.bind("<Leave>",lambda e,widget = self.widget: self.really_leaving(e,widget))

        
        def really_leaving(self,e,widget):
            if self.tip_window == None or self.long_task_called:
                return
            def long_task():
                time_start = time.time()
                while 1:
                    time.sleep(0.05)
                    if time.time() - time_start > self.delay:
                        break
                try:
                    self.root.after(0,self.tip_window.destroy)
                    # self.tip_window.destroy()
                except Exception as e1:
                    print("error1")
                self.tip_window = None
                self.long_task_called = False

            x = widget.winfo_width()-1
            y = widget.winfo_height()-1
            if (e.x < 1 or e.x > x) or (e.y<1 or e.y > y):
                # if self.tip_window != None and not self.long_task_called:
                if self.add_delay:
                    tooltip_thread = threading.Thread(target=long_task,)
                    self.long_task_called = True
                    tooltip_thread.start()
                else:
                    try:
                        # self.tip_window.destroy()
                        self.root.after(0,self.tip_window.destroy)
                    except Exception as e2:
                        print("error2")
                    self.tip_window = None

    def __init__(self,root,menu_callback_function,window_mode,initial_path,zoom_factor,config_filename,without_gui =False):
        self.root = root
        self.menu_callback = menu_callback_function
        self.initial_path = initial_path
        self.window_mode = window_mode
        self.zoom_factor = zoom_factor
        self.show_favourite_ip = False
        self.config_filename_path = initial_path + config_filename
        self.excel_file_path = initial_path + "TRIMAZKON_address_list.xlsx"
        # self.excel_file_path = initial_path + "config_TRIMAZKON.xlsx"
        self.app_icon = Tools.resource_path('images\\logo_TRIMAZKON.ico')
        self.default_environment = "ip"
        if not without_gui:
            self.check_default_env()
            self.check_excel_presence()
            if self.default_environment == "disk":
                self.Disk_management_gui(self)
            else:
                self.IP_assignment(self)
        else:
            self.check_excel_presence()

    def check_default_env(self):
        try:
            default_environment = Tools.read_json_config(self.config_filename_path)["disk_or_ip_window"]
            if int(default_environment) == 1:
                self.default_environment = "disk"

            def_show_favourite = Tools.read_json_config(self.config_filename_path)["favorite_ip_window_status"]
            if int(def_show_favourite) == 1:
                self.show_favourite_ip = True
            else:
                self.show_favourite_ip = False
        
        except Exception as e:
            print(f"Nejprve zavřete soubor {self.config_filename_path} Chyba: {e}")
        
    def check_excel_presence(self):
        try:
            workbook = load_workbook(self.excel_file_path)
            workbook.save(self.excel_file_path) #check if it is opened currently
            workbook.close()
            return
    
        except Exception as e:
            print(f"Nejprve zavřete soubor {self.excel_file_path} Chyba: {e}")
            if "Errno 13" in str(e):
                self.default_environment = "config_load_error"
            else:
                workbook = Workbook()
                ws = workbook.active
                ws.title = "ip_address_list"
                workbook.create_sheet(title="ip_address_fav_list")
                workbook.create_sheet(title="disk_list")
                workbook.save(self.excel_file_path) #check if it is opened currently
                workbook.close()

    class Disk_management_gui:
        def __init__(self,parent):
            self.parent_instance = parent
            self.root = parent.root
            self.menu_callback = parent.menu_callback
            self.window_mode = parent.window_mode
            self.excel_file_path = parent.excel_file_path
            self.app_icon = parent.app_icon
            self.config_filename_path = parent.config_filename_path
            self.initial_path = parent.initial_path
            self.all_project_list = []
            self.last_managed_project = None
            self.last_project_id = ""
            self.opened_window = ""
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.last_inserted_password = ""
            self.changed_notes_disk = []
            self.selected_list_disk = []
            self.remember_to_change_back = []
            self.deleted_projects_bin = []
            self.edited_projects_bin = []
            self.notes_frame_height = 50
            read_parameters = Tools.read_json_config(self.config_filename_path)
            if read_parameters != None:
                if read_parameters["default_window_size"] == 2:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                self.default_disk_status_behav = read_parameters["init_disk_refresh"]
                self.default_note_behav = read_parameters["editable_notes"]
                self.mapping_condition = read_parameters["disk_persistent"]
                if read_parameters["auto_order_when_edit"] == 1:
                    self.make_edited_project_first = True
                else:
                    self.make_edited_project_first = False
                self.deletion_behav = read_parameters["ask_to_delete"]
            else:
                self.default_disk_status_behav = 0
                self.default_note_behav = 0
                self.mapping_condition = 0
                self.make_edited_project_first = True
                self.deletion_behav = 100

            self.manage_bin("read_sheet")
            self.create_widgets_disk(init=True)

        def call_menu(self): # Tlačítko menu (konec, návrat do menu)
            """
            Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
            """

            Tools.clear_frame(self.main_widgets)
            Tools.clear_frame(self.root)
            # self.root.unbind("<f>")
            self.root.unbind("<Escape>")
            self.root.unbind("<F5>")
            self.root.unbind("<Button-1>")
            self.root.unbind("<Control_L>")
            self.root.unbind("<Control-Button-1>")
            self.root.unbind("<KeyRelease-Control_L>")
            self.root.unbind("<Delete>")
            self.root.update()
            self.root.update_idletasks()
            self.menu_callback()
        
        def clicked_on_project(self,project,widget,textbox = "",flag = ""):
            """
            flag = notes:
            - při nakliknutí poznámky zůstanou expandnuté a při kliku na jinou je potřeba předchozí vrátit zpět
            flag = unfocus:
            - při kliku mimo se odebere focus z nakliknutých widgetů
            """
            def on_leave_entry(last_selected_textbox):
                """
                při kliku na jiný widget:
                - upraví text pouze na první řádek
                """
                new_height = self.notes_frame_height
                last_selected_textbox.configure(state = "normal")
                if "\n" in self.last_managed_project["notes"]:
                    notes_rows = self.last_managed_project["notes"].split("\n")
                    first_row = notes_rows[0]
                    last_selected_textbox.delete("1.0",tk.END)
                    last_selected_textbox.insert(tk.END,str(first_row))
                    last_selected_textbox.configure(height = new_height-10) #notes
                if self.default_note_behav == 0:
                    last_selected_textbox.configure(state = "disabled")

            if flag == "unfocus":
                try:
                    if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                        if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox)
                            self.last_selected_textbox = ""
                            self.last_selected_notes_widget = ""

                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        self.last_selected_widget.configure(border_color="#636363")
                        self.last_selected_widget = ""

                    for frame in self.remember_to_change_back:
                        if frame.winfo_exists(): 
                            frame.configure(border_color="#636363")
                    self.selected_list_disk = []
                    self.remember_to_change_back = []
                    self.last_managed_project = None

                except Exception as e:
                    print("chyba při odebírání focusu: ",e)
                return

            if project == None:
                return
            print("clicked project: ",project['name'])
            self.search_input.delete("0","300")
            self.search_input.insert("0",str(project['name']))
            # only if it is not pressed againt the same:
            if widget != self.last_selected_widget:
                try:
                    if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                        on_leave_entry(self.last_selected_textbox)
                    else:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""

                    if flag == "notes":
                        self.last_selected_textbox = textbox
                        self.last_selected_notes_widget = widget
                        widget.focus_set()
                    else:
                        # init the values:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""
                        widget.focus_set()

                except Exception as e:
                    print("chyba s navracenim framu do puvodniho formatu",e)
                
                try:
                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        if len(self.selected_list_disk) == 0 and not self.control_pressed:
                            self.last_selected_widget.configure(border_color="#636363")

                            if self.last_selected_widget in self.remember_to_change_back:
                                self.remember_to_change_back.pop(self.remember_to_change_back.index(self.last_selected_widget))

                        # pokud došlo k další interakci s jiným widgeten
                        elif not self.control_pressed:
                            for frame in self.remember_to_change_back:
                                if frame.winfo_exists(): 
                                    frame.configure(border_color="#636363")
                            self.selected_list_disk = []
                            self.remember_to_change_back = []

                    self.last_selected_widget = widget
                    widget.configure(border_color="white")

                    if not widget in self.remember_to_change_back:
                        self.remember_to_change_back.append(widget)
                    print("remember: ", self.remember_to_change_back)

                except Exception as e:
                    print("chyba pri zmene fucusu",e)
                    pass

                self.last_managed_project = project

        def refresh_disk_statuses(self,silent=True):
            online_disks = []
            offline_disks = []
            # self.refresh_btn.configure(text = "🔄",font=("",25))
            # self.refresh_btn.update()
            # self.refresh_btn.update_idletasks()

            def refresh_thread():
                mapped_disks = main.DM_tools.list_mapped_disks(whole_format = True)
                non_persistant_disks = main.DM_tools.list_non_persistent_disks()
                for y in range(0,len(self.disk_letter_frame_list)):
                    param_frame = self.disk_letter_frame_list[y]
                    param_frame.configure(fg_color = "black") # <= init

                    for i in range(0,len(non_persistant_disks)):
                        if non_persistant_disks[i][0:1] == str(self.all_project_list[y][1]):
                            drive_status = main.DM_tools.check_network_drive_status(non_persistant_disks[i])
                            if drive_status == True:
                                online_disks.append(non_persistant_disks[i][0:1])
                                param_frame.configure(fg_color = "#00CED1")
                            else:
                                offline_disks.append(non_persistant_disks[i][0:1])
                                param_frame.configure(fg_color = "red")

                    for i in range(0,len(mapped_disks)):
                        if mapped_disks[i][0:1] == str(self.all_project_list[y]["disk_letter"]):
                            drive_status = main.DM_tools.check_network_drive_status(mapped_disks[i])
                            if drive_status == True:
                                online_disks.append(mapped_disks[i][0:1])
                                param_frame.configure(fg_color = "green")
                            else:
                                offline_disks.append(mapped_disks[i][0:1])
                                try:
                                    param_frame.configure(fg_color = "red")
                                except Exception:
                                    pass
                if len(mapped_disks) == 0 and len(non_persistant_disks) == 0 and silent == False:
                    Tools.add_colored_line(self.main_console,f"Nejsou namapované žádné disky","red",None,True)
                elif silent == False:
                    if len(online_disks) != 0 and len(offline_disks) != 0:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: online: {list(set(online_disks))}, offline: {list(set(offline_disks))}","white",None,True)
                    elif len(online_disks) == 0 and  len(offline_disks) != 0:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: offline: {list(set(offline_disks))}","white",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: online: {list(set(online_disks))}","white",None,True)

                # self.refresh_btn.configure(text = "Refresh statusů",font=("Arial",20,"bold"))
            
            run_backgroung = threading.Thread(target=refresh_thread,)
            run_backgroung.start()

        def manage_bin(self,flag="",project=None,new_edited_name=None):
            """
            flag:\n
            - read_sheet
            - save_project_disk
            - load_deleted_disk
            - save_edited_disk
            - load_edited_disk
            """
            max_stored_deletions = 5
            max_stored_edits = 10

            def read_sheet():
                config_data = Tools.read_json_config(self.config_filename_path)
                try:
                    self.deleted_projects_bin = config_data["deleted_project_bin_disk"]
                except Exception:
                    Tools.save_to_json_config("deleted_project_bin_disk",self.deleted_projects_bin,self.config_filename_path)

                # self.edited_projects_bin = config_data["edited_project_bin_disk"]
                Tools.save_to_json_config("edited_project_bin_disk",[],self.config_filename_path) #vymazat historii editu při zapnutí
                    
            def save_project_disk():# saving after deleting:
                if project == None:
                    return
                config_data = Tools.read_json_config(self.config_filename_path)
                self.deleted_projects_bin = config_data["deleted_project_bin_disk"]
                self.undo_button.configure(state = "normal")
                self.deleted_projects_bin.insert(0,project)

                if len(self.deleted_projects_bin) > max_stored_deletions:
                    self.deleted_projects_bin.pop()
                Tools.save_to_json_config("deleted_project_bin_disk",self.deleted_projects_bin,self.config_filename_path)

            def save_edited_disk():# saving after editing:
                if project == None or new_edited_name == None:
                    return
                self.undo_edit.configure(state = "normal")
                config_data = Tools.read_json_config(self.config_filename_path)
                self.edited_projects_bin = config_data["edited_project_bin_disk"]
                project["new_name"] = new_edited_name
                print("\nSAVING: ",project)
                self.edited_projects_bin.insert(0,project)
                if len(self.edited_projects_bin) > max_stored_edits:
                    self.edited_projects_bin.pop()
                Tools.save_to_json_config("edited_project_bin_disk",self.edited_projects_bin,self.config_filename_path)

            def load_deleted_disk():
                """
                adds new project from history and deletes the history
                """
                config_data = Tools.read_json_config(self.config_filename_path)
                self.deleted_projects_bin = config_data["deleted_project_bin_disk"]
                project_to_load = self.deleted_projects_bin[0]
                self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
                modified_project_name = Tools.get_unique_name(self.all_project_list,project_to_load['name'])
                project_to_load["name"] = modified_project_name
                self.deleted_projects_bin.pop(0)
                if len(self.deleted_projects_bin) ==0:
                    self.undo_button.configure(state = "disabled")

                Tools.save_to_json_config("deleted_project_bin_disk",self.deleted_projects_bin,self.config_filename_path)
                self.all_project_list.insert(0,project_to_load)
                main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)
                Tools.add_colored_line(self.main_console,f"Projekt: {project_to_load['name']} byl úspěšně obnoven","green",None,True)
                self.make_project_cells_disk()

            def load_edited_disk():
                config_data = Tools.read_json_config(self.config_filename_path)
                self.edited_projects_bin = config_data["edited_project_bin_disk"]
                project_to_load = self.edited_projects_bin[0]
                print("project to load: ",project_to_load)
                old_project_name = str(project_to_load['name'])
                current_project_name = str(project_to_load["new_name"])
                self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
                project_index = Tools.get_project_index(self.all_project_list,current_project_name)

                if project_index == None:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {current_project_name} nenalezeno, nelze ho tedy obnovit","red",None,True)
                    
                self.edited_projects_bin.pop(0)
                if len(self.edited_projects_bin) ==0:
                    self.undo_edit.configure(state = "disabled")
                Tools.save_to_json_config("edited_project_bin_disk",self.edited_projects_bin,self.config_filename_path)
                if project_index == None: #let it to be deleted... no use, corrupted
                    return
                
                print(project_to_load,"\n",self.all_project_list[project_index])
                self.all_project_list[project_index]["name"] = str(project_to_load["name"])
                self.all_project_list[project_index]["disk_letter"] = str(project_to_load["disk_letter"])
                self.all_project_list[project_index]["ftp"] = str(project_to_load["ftp"])
                self.all_project_list[project_index]["user"] = str(project_to_load["user"])
                self.all_project_list[project_index]["password"] = str(project_to_load["password"])
                self.all_project_list[project_index]["notes"] = str(project_to_load["notes"])

                # self.all_project_list.insert(0,project_to_load)
                main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)
                if old_project_name != current_project_name:
                    Tools.add_colored_line(self.main_console,f"U projektu: {old_project_name} (původně: {current_project_name}) byly odebrány provedené změny","green",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"U projektu: {old_project_name} byly odebrány provedené změny","green",None,True)
                self.make_project_cells_disk()

            mapping_logic = {
                "read_sheet": read_sheet,
                "save_project_disk": save_project_disk,
                "load_deleted_disk": load_deleted_disk,
                "save_edited_disk": save_edited_disk,
                "load_edited_disk": load_edited_disk,
            }

            output = mapping_logic[flag]()

            self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
            return output
        
        def check_given_input(self,given_data = None,search_flag=False):
            """
            Fills all parameters of last project
            """
            if given_data == None:
                given_data = self.search_input.get()
            if given_data == "":
                found = None
                return found
            found = False

            for i in range(0,len(self.all_project_list)):
                if search_flag:
                    if str(given_data) in str(self.all_project_list[i]['name']):
                        self.last_managed_project = self.all_project_list[i]
                        found = True
                else:
                    if given_data == self.all_project_list[i]['name']:
                        self.last_managed_project = self.all_project_list[i]
                        found = True
            return found  

        def refresh_explorer(self,refresh_disk=None):
            """
            Resetuje windows explorer přes cmd
            refresh_disk = udelat nove všechni widgets (make_project_cells_disk())
            """
            refresh_explorer="taskkill /f /im explorer.exe"
            subprocess.run(refresh_explorer, shell=True)
            refresh_explorer="start explorer.exe"
            subprocess.run(refresh_explorer, shell=True)
            if refresh_disk:
                self.make_project_cells_disk(disk_statuses=True)

        def delete_disk_option_menu(self):
            def delete_disk(child_root):
                drive_letter = str(self.drive_letter_input.get())
                if len(str(self.DL_manual_entry.get())) > 0:
                    drive_letter = str(self.DL_manual_entry.get())
                
                delete_command = "net use " + drive_letter +": /del"
                process = subprocess.Popen(delete_command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
                stdout, stderr= process.communicate()
                if "Is it OK to continue disconnecting and force them closed?" in stdout:
                    Tools.add_colored_line(self.main_console,f"Disk je právě používán, nejprve jej zavřete","red",None,True)
                    child_root.destroy()
                else:
                    self.refresh_explorer()
                    Tools.add_colored_line(self.main_console,f"Disky s označením {drive_letter} byly odpojeny","orange",None,True)
                    self.refresh_disk_statuses()
                    child_root.destroy()

            child_root = customtkinter.CTkToplevel(fg_color="#212121")
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"450x190+{x+250}+{y+200}")
            child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
            child_root.title("Odpojování síťového disku")
            
            found_drive_letters=[]
            for i in range(0,len(self.all_project_list)):
                if not self.all_project_list[i]["disk_letter"] in found_drive_letters:
                    found_drive_letters.append(self.all_project_list[i]["disk_letter"])

            mapped_disks = main.DM_tools.list_mapped_disks()
            non_persistent_disks = main.DM_tools.list_non_persistent_disks()
            for disk in non_persistent_disks:
                if not disk in mapped_disks:
                    mapped_disks.append(disk)

            for i in range(0,len(mapped_disks)):
                if not mapped_disks[i] in found_drive_letters:
                    found_drive_letters.append(mapped_disks[i])

            label =                     customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte disk nebo vyhledejte manuálně: ",font=("Arial",20,"bold"))
            self.drive_letter_input =   customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=found_drive_letters,corner_radius=0)
            self.DL_manual_entry =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manuálně")
            btn_frame =                 customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,fg_color="#212121")
            del_button =                customtkinter.CTkButton(master = btn_frame, width = 200,height=40,text = "Odpojit", command = lambda: delete_disk(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
            exit_button =               customtkinter.CTkButton(master = btn_frame, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button.                pack(pady=5,padx=(0,10),anchor = "e",side="right")
            del_button.                 pack(pady=5,padx=(0,10),anchor = "e",side="right")
            btn_frame.                  pack(pady=0,padx=0,fill="x",side = "bottom")
            label.                      pack(pady=5,padx=(10),anchor = "n",side="top")
            self.drive_letter_input.    pack(pady=5,padx=(10),anchor = "n",side="top")
            self.DL_manual_entry.       pack(pady=5,padx=(10),anchor = "n",side="top")

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
            print("disk letter list: ",found_drive_letters)
            try:
                self.drive_letter_input.set(self.last_managed_project["disk_letter"])
            except Exception:
                pass

        def delete_project_disk(self,wanted_project=None,silence=None,flag=""):
            if "!ctktextbox" in str(self.root.focus_get()):
                return
            
            project_found = False
            name_list = []

            def check_multiple_projects(window):
                nonlocal wanted_project
                nonlocal name_list
                nonlocal project_found

                if len(self.selected_list_disk) > 1:
                    for names in name_list:
                        print(names)
                        project_found = False
                        self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
                        proceed(names,window,True)
                            
                    Tools.add_colored_line(self.main_console,f"Byly úspěšně odstraněny tyto projekty: {name_list}","orange",None,True)
                    try:
                        self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                    except Exception as e:
                        print("chyba, refresh po mazani")
                else:
                    proceed(wanted_project,window)

            def proceed(wanted_project, window = True, multiple_status = False):
                nonlocal project_found
                nonlocal child_root
                deleted_project = None
                if wanted_project == None:
                    self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
                    wanted_project = str(self.search_input.get())
                workbook = load_workbook(self.excel_file_path)
                worksheet = workbook["disk_list"]

                project_to_delete = Tools.found_project_name(self.all_project_list,wanted_project)
                if project_to_delete == False:
                    Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    return
                
                self.manage_bin(flag="save_project_disk",project=project_to_delete)
                project_index = Tools.get_project_index(self.all_project_list,project_to_delete['name'])
                worksheet.delete_rows(project_index+1)
                workbook.save(self.excel_file_path)
                workbook.close()
                project_found = True
                deleted_project = project_to_delete

                workbook.close()
                if silence==None and not multiple_status:
                    if project_found:
                        Tools.add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)    
                        self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                    elif wanted_project.replace(" ","") == "":
                        Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)

                elif project_found and not multiple_status:  # zresetuj i v pripade silence...
                    self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...

                if window and child_root.winfo_exists():
                    child_root.grab_release()
                    child_root.destroy()

                return deleted_project

            if self.last_managed_project is None:
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif str(self.last_managed_project['name']).replace(" ","") == "":
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif wanted_project == None:
                wanted_project = self.last_managed_project['name']
            
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                check_multiple_projects(False)
                return

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Upozornění")
            proceed_label_text = f"Opravdu si přejete odstranit projekt {self.last_managed_project['name']}?"
            # if flag == "context_menu":
            #     self.selected_list_disk = []
            if len(self.selected_list_disk) > 1:
                for projects in self.selected_list_disk:
                    if str(projects['name']) not in name_list:
                        name_list.append(str(projects['name']))
                proceed_label_text = f"Opravdu si přejete odstranit vybrané projekty:\n{name_list}?"
            proceed_label = customtkinter.CTkLabel(master = child_root,text = proceed_label_text,font=("Arial",22,"bold"),justify = "left",anchor="w")
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: check_multiple_projects(True))
            button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
            proceed_label   .pack(pady=(15,0),padx=10,side = "top",fill="x")
            button_no       .pack(pady = 5, padx = 10,anchor="w",side="right",expand = False)
            button_yes      .pack(pady = 5, padx = 10,anchor="w",side="right",expand = False)
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
            child_root.update()
            child_root.update_idletasks()
            child_root.grab_set()
            child_root.focus()
            child_root.focus_force()
            child_root.wait_window()
            return project_found

        def add_new_project_disk(self,edit = None,init_copy=False,childroot_given = None):
            def mouse_wheel_change(e):
                if -e.delta < 0:
                    switch_up()
                else:
                    switch_down()

            def copy_previous_project():
                try:
                    if self.last_managed_project['name'] == "":
                        Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
                        return
                    self.last_inserted_password = str(self.last_managed_project['password'])
                    self.name_input.delete("0","300")
                    self.name_input.insert("0",str(self.last_managed_project['name']))
                    self.disk_letter_input.delete("0","300")
                    self.disk_letter_input.insert("0",str(self.last_managed_project['disk_letter']))
                    self.FTP_adress_input.delete("0","300")
                    self.FTP_adress_input.insert("0",str(self.last_managed_project['ftp']))
                    self.username_input.delete("0","300")
                    self.username_input.insert("0",str(self.last_managed_project['user']))
                    self.password_input.delete("0","300")
                    self.password_input.insert("0",str(len(self.last_managed_project['password'])*"*"))
                    self.notes_input.delete("1.0",tk.END)
                    self.notes_input.insert(tk.END,str(self.last_managed_project['notes']))

                except TypeError:
                    Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)

            def switch_up(force_index = None):
                if force_index != None:
                    project_index = force_index
                else:
                    project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])

                another_project_id = project_index
                another_project_id -= 1
                if another_project_id < 0:
                    another_project_id = len(self.all_project_list)-1
                    
                self.last_managed_project = self.all_project_list[another_project_id]
                copy_previous_project()
                refresh_title()

            def switch_down():
                project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                previous_project_id = project_index
                previous_project_id += 1
                if previous_project_id > len(self.all_project_list)-1:
                    previous_project_id = 0

                self.last_managed_project = self.all_project_list[previous_project_id]
                copy_previous_project()
                refresh_title()

            def del_project():
                nonlocal child_root
                project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                result = self.delete_project_disk()
                if result:
                    switch_up(project_index)
                else:
                    print("aborted")

                child_root.focus()
                child_root.focus_force()
                child_root.grab_set()
            
            def refresh_title():
                if edit == None:
                    child_root.title("Nový projekt")
                else:
                    child_root.title("Editovat projekt: "+ str(self.last_managed_project['name']))

            def show_password():
                if str(self.password_input.get()) == str(len(self.last_inserted_password)*"*"):
                    self.password_input.configure(state = "normal")
                    self.password_input.delete("0","300")
                    self.password_input.insert("0",self.last_inserted_password)
                else:
                    self.last_inserted_password = str(self.password_input.get())
                    self.password_input.delete("0","300")
                    self.password_input.insert("0",str(len(self.last_inserted_password)*"*"))
                    self.password_input.configure(state = "disabled")

            def save_project(add_next = False):
                project_name = str(self.name_input.get())
                disk_letter =  str(self.disk_letter_input.get())
                ftp_address =  str(self.FTP_adress_input.get())
                username =     str(self.username_input.get())
                password =     str(self.last_inserted_password)
                notes = Tools.get_legit_notes(self.notes_input.get("1.0", tk.END))
                errors = 0

                if project_name.replace(" ","") == "":
                    Tools.add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
                    errors += 1
                elif disk_letter.replace(" ","") == "":
                    Tools.add_colored_line(self.console,f"Nezadali jste písmeno disku","red",None,True)
                    errors += 1
                elif ftp_address.replace(" ","") == "":
                    Tools.add_colored_line(self.console,f"Nezadali jste adresu","red",None,True)
                    errors += 1
                if errors>0:
                    return
                
                self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)

                if edit:
                    print("last_managed project:", self.last_managed_project)
                    currently_edited_project_id = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                    if currently_edited_project_id == None:
                        Tools.add_colored_line(self.main_console,f"Projekt {self.last_managed_project['name']} nenalezen","red",None,True)
                        return
                    
                    if self.all_project_list[currently_edited_project_id]['name'] != project_name:
                        project_name = Tools.get_unique_name(self.all_project_list,project_name)
                    self.all_project_list[currently_edited_project_id]['name'] = project_name
                    self.all_project_list[currently_edited_project_id]["disk_letter"] = disk_letter
                    self.all_project_list[currently_edited_project_id]["ftp"] = ftp_address
                    self.all_project_list[currently_edited_project_id]["user"] = username
                    self.all_project_list[currently_edited_project_id]["password"] = password
                    self.all_project_list[currently_edited_project_id]["notes"] = notes
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)

                    if self.last_managed_project['name'] != project_name:
                        status_text = f"Projekt: {self.last_managed_project['name']} (nově: {project_name}) úspěšně pozměněn"
                    else:
                        status_text = f"Projekt: {self.last_managed_project['name']} úspěšně pozměněn"
                    Tools.add_colored_line(self.main_console,status_text,"green",None,True)
                    
                    self.manage_bin(flag="save_edited_disk",project=self.last_managed_project,new_edited_name=project_name)
                    if self.make_edited_project_first:
                        self.make_project_first_disk(purpouse="silent",make_cells=False,project=self.all_project_list[currently_edited_project_id],input_entry_bypass=project_name)
                else:
                    project_name = Tools.get_unique_name(self.all_project_list,project_name)
                    self.all_project_list.insert(0,{
                        'name':project_name,
                        "disk_letter":disk_letter,
                        "ftp":ftp_address,
                        "user":username,
                        "password":password,
                        "notes":notes,
                    })
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)
                    Tools.add_colored_line(self.main_console,f"Nový projekt: {project_name} byl úspěšně přidán","green",None,True)

                self.make_project_cells_disk()
                if add_next:
                    Tools.clear_frame(child_root)
                    self.root.after(10,self.add_new_project_disk(childroot_given=child_root))
                else:
                    child_root.destroy()


            if childroot_given == None:
                child_root = customtkinter.CTkToplevel(fg_color="#212121")
            else:
                child_root = childroot_given
            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            refresh_title()
            top_main_frame =            customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,fg_color="#212121")
            top_left_frame =            customtkinter.CTkFrame(master=top_main_frame,corner_radius=0,border_width=2,fg_color="#212121")
            top_right_frame =           customtkinter.CTkFrame(master=top_main_frame,corner_radius=0,border_width=2,fg_color="#212121")
            project_name =              customtkinter.CTkLabel(master = top_left_frame, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
            project_selection_label =   customtkinter.CTkLabel(master = top_right_frame, width = 200,height=30,text = "Přepnout projekt: ",font=("Arial",20,"bold"))
            project_switch_frame =      customtkinter.CTkFrame(master=top_right_frame,corner_radius=0,border_width=0,height=140,width=80)
            project_up =                customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↑",command= lambda: switch_up())
            project_down =              customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↓",command= lambda: switch_down())
            project_up                  .pack(pady=(0,5),padx=5,side = "top",fill="x")
            project_down                .pack(pady=0,padx=5,side = "top",fill="x")
            project_switch_frame.       bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_up.                 bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_down.               bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            copy_check =                customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: copy_previous_project())
            save_and_add_next =         customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Uložit a přidat další",command= lambda: save_project(add_next = True))
            del_project_btn =           customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Smazat tento projekt",command= lambda: del_project(),fg_color="red")
            self.name_input =           customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            disk_letter =               customtkinter.CTkLabel(master = top_left_frame,height=30,text = "Písmeno disku: ",font=("Arial",20,"bold"))
            self.disk_letter_input =    customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            FTP_adress =                customtkinter.CTkLabel(master = top_left_frame,height=30,text = "FTP adresa: ",font=("Arial",20,"bold"))
            self.FTP_adress_input =     customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=500,height=30,corner_radius=0)
            user =                      customtkinter.CTkLabel(master = top_left_frame,height=30,text = "Uživatelské jméno: ",font=("Arial",20,"bold"))
            self.username_input =       customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            password =                  customtkinter.CTkLabel(master = top_left_frame,height=30,text = "Heslo: ",font=("Arial",20,"bold"))
            pwd_frame =                 customtkinter.CTkFrame(master=top_left_frame,corner_radius=0,border_width=0,fg_color="#212121")
            self.password_input =       customtkinter.CTkEntry(master = pwd_frame,font=("Arial",20),width=170,height=30,corner_radius=0)
            show_pass_btn =             customtkinter.CTkButton(master = pwd_frame,font=("Arial",15),width=30,height=30,corner_radius=0,text="👁",command= lambda: show_password())
            notes_label =               customtkinter.CTkLabel(master = child_root,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
            self.notes_input =          customtkinter.CTkTextbox(master = child_root,font=("Arial",20),height=260,corner_radius=0)
            self.console =              tk.Text(child_root, wrap="none", height=0,background="black",font=("Arial",14),state=tk.DISABLED)
            buttons_frame =             customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,fg_color="#212121")
            save_button =               customtkinter.CTkButton(master = buttons_frame, width = 200,height=40,text = "Uložit", command = lambda: save_project(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =               customtkinter.CTkButton(master = buttons_frame, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            project_name.pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            if edit != True:
                copy_check.         pack(pady = (10),padx =(10),anchor="w",side="top")
                save_and_add_next.  pack(pady = (0,10),padx =(10),anchor="w",side="top")
            if edit:
                project_selection_label.pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
                project_switch_frame.pack(pady=10,padx=10,anchor="w",side = "top",fill="x")
                del_project_btn.pack(pady = (10),padx =(10),anchor="s",side="left",fill="x")
            self.name_input.        pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            disk_letter.            pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.disk_letter_input. pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            FTP_adress.             pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.FTP_adress_input.  pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            user.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.username_input.    pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            password.               pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.password_input.    pack(anchor="w",side="left",fill="x",expand=True)
            show_pass_btn.          pack(anchor="w",side="left")
            pwd_frame.              pack(pady = (10),padx =(5),anchor="w",side="top",fill="x",expand=True)
            top_left_frame.         pack(anchor="w",side="left",fill="both",expand = True)
            top_right_frame.        pack(anchor="w",side="left",fill="y",expand =False,ipadx=2,ipady=2)
            top_main_frame.         pack(anchor="w",side="top",fill="both",expand = True)
            notes_label.            pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.notes_input.       pack(pady = (10,0),padx =(5),anchor="w",side="top",fill="both",expand=True)
            self.console.           pack(pady = (10,0),padx =(5),anchor="w",side="top",fill="x",expand=False)
            exit_button.            pack(pady = (10,0),padx =(5,0),anchor="e",side="right")
            save_button.            pack(pady = (10,0),padx =(5,0),anchor="e",side="right")
            buttons_frame.          pack(pady = (0,10),padx =(0,10),anchor="w",side="top",fill="x",expand=False)

            if init_copy: # kopírovat pro vytvoreni noveho projektu, neni edit...
                copy_previous_project()
                self.password_input.configure(state = "disabled")
            elif edit == None:
                self.last_inserted_password = ""
                self.disk_letter_input.delete("0","300")
                self.disk_letter_input.insert("0","P")
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0","\\\\192.168.000.000\\")
                self.username_input.delete("0","300")
                self.password_input.delete("0","300")
                if str(self.search_input.get()).replace(" ","") != "":
                    self.name_input.delete("0","300")
                    self.name_input.insert("0",str(self.search_input.get()))
            else:
                copy_previous_project()
                self.password_input.configure(state = "disabled")

            def update_last_password():
                if self.password_input.cget("state") == "normal":
                    self.last_inserted_password = str(self.password_input.get())
            self.password_input.bind("<Key>",lambda e: update_last_password())

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def make_project_first_disk(self,purpouse = None,make_cells = True,input_entry_bypass = None,project= None,upwards=False,downwards=False):
            def check_position():
                project_index = Tools.get_project_index(self.all_project_list,project['name'])
                prev_pos = project_index
                max_position = len(self.all_project_list)
                if upwards:
                    position = prev_pos -1
                elif downwards:
                    position = prev_pos +1

                if position < 0:
                    position = max_position-1
                elif position > max_position-1:
                    position = 0
                return position
            
            if purpouse == "search":
                result = self.check_given_input(input_entry_bypass,search_flag=True)
                if result == True:
                    self.search_input.delete("0","300")
                    self.search_input.insert("0",str(self.last_managed_project['name']))
            else:
                result = self.check_given_input(input_entry_bypass)
            self.remember_to_change_back = []
            self.last_selected_widget = ""

            if result == True: #zmena poradi
                if project == None:
                    project = self.last_managed_project

                if downwards or upwards:
                    position = check_position()
                else:
                    position = 0

                if len(self.all_project_list) > 0:
                    project_index = Tools.get_project_index(self.all_project_list,project['name'])
                    self.all_project_list.pop(project_index)

                self.all_project_list.insert(position,project)
                main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)

                if make_cells:
                    self.make_project_cells_disk()
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} nalezen","green",None,True)
                elif purpouse != "silent":
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} přesunut na začátek","green",None,True)
            elif result == None and purpouse != "silent":
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            elif purpouse != "silent":
                Tools.add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

        def map_disk(self,project):
            Drive_letter = str(project["disk_letter"])
            ftp_adress = str(project["ftp"])
            user = str(project["user"])
            password = str(project["password"])

            delete_command = "net use " + Drive_letter + ": /del"
            subprocess.run(delete_command, shell=True)
            if self.mapping_condition == 1:
                persistant_status = " /persistent:yes"
            else:
                persistant_status =  " /persistent:no"

            if user != "" or password != "":
                second_command = "net use " + Drive_letter + ": " + ftp_adress + " " + password + " /user:" + user + persistant_status
            else:
                second_command = "net use " + Drive_letter + ": " + ftp_adress + persistant_status
            print("calling: ",second_command)


            def call_subprocess():
                nonlocal connection_status
                """process = subprocess.Popen(second_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                stdout, stderr = process.communicate()
                self.connection_status = process.returncode
                print("STDOUT:", stdout)
                print("STDERR:", stderr)
                print("Return Code:", self.connection_status)"""
                connection_status = subprocess.call(second_command,shell=True,text=True)
            connection_status = None
            run_background = threading.Thread(target=call_subprocess,)
            run_background.start()

            time_start = time.time()
            while connection_status==None:
                time.sleep(0.05)
                if time.time() - time_start > 3:
                    print("terminated due to runtime error")
                    break

            if connection_status == 0:
                Tools.add_colored_line(self.main_console,f"Disk úspěšně připojen","green",None,True)
                self.refresh_explorer()
                self.refresh_disk_statuses()

                def open_explorer(path):
                    if os.path.exists(path):
                        os.startfile(path)
                    else:
                        print(f"The path {path} does not exist.")

                open_explorer(Drive_letter + ":\\")
            else:
                Tools.add_colored_line(self.main_console,f"Připojení selhalo (ixon? musí být zvolena alespoň 1 složka na disku...)","red",None,True)

        def show_context_menu(self,event,project,flag=""):
            context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="#202020",activebackground="#606060")
            self.last_managed_project = project
            # self.check_given_input(given_data=self.all_project_list[first_index][0])
            
            if flag == "button":
                context_menu.add_command(label="Namapovat",font=("Arial",22,"bold"),command=lambda: self.map_disk(project))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat FTP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(str(project["ftp"])))
                context_menu.add_separator()
                context_menu.add_command(label="Editovat",font=("Arial",22,"bold"),command=lambda: self.add_new_project_disk(edit=True))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat projekt",font=("Arial",22,"bold"),command=lambda: self.add_new_project_disk(init_copy=True))
                context_menu.add_separator()
                context_menu.add_command(label="Přesunout na začátek",font=("Arial",22,"bold"),command=lambda: self.make_project_first_disk(input_entry_bypass=str(project["name"])))
                context_menu.add_separator()
                context_menu.add_command(label="Odstranit",font=("Arial",22,"bold"),command=lambda: self.delete_project_disk(flag="context_menu"))
            elif flag == "ftp_frame":
                context_menu.add_command(label="Kopírovat FTP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(str(project["ftp"])))
            elif flag == "disk_letter":
                context_menu.add_command(label="Refresh",font=("Arial",22,"bold"), command=lambda: self.refresh_disk_statuses(silent=False))
                context_menu.add_separator()
                context_menu.add_command(label="Odpojit disk",font=("Arial",22,"bold"),command=lambda: self.delete_disk_option_menu())

            context_menu.tk_popup(event.x_root, event.y_root)

        def make_project_cells_disk(self,no_read = False,disk_statuses = False,init=False):
            def opened_window_check():
                if self.opened_window == "":
                    return False
                try:
                    if self.opened_window.winfo_exists():
                        return True
                    else:
                        return False
                except Exception as e:
                    print(e)
                    return False

            def filter_text_input(text):
                legit_rows = []
                legit_notes = ""
                rows = text.split("\n")
                for i in range(0,len(rows)):
                    if rows[i].replace(" ","") != "":
                        legit_rows.append(rows[i])

                for i in range(0,len(legit_rows)): 
                    if i == len(legit_rows)-1:
                        legit_notes = legit_notes + legit_rows[i]
                    else:
                        legit_notes = legit_notes + legit_rows[i]+ "\n"
                return legit_notes
            
            def save_changed_notes(notes,project):
                workbook = load_workbook(self.excel_file_path)
                def save_to_workbook(notes,row,excel_worksheet):
                    nonlocal workbook
                    worksheet = workbook[excel_worksheet]
                    worksheet['D' + str(row+1)] = notes

                self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)
                project_index = Tools.get_project_index(self.all_project_list,project['name'])
                save_to_workbook(notes,project_index,"disk_list")
                self.all_project_list[project_index]["notes"] = notes
                workbook.save(filename=self.excel_file_path)
                workbook.close()

            def on_enter_entry(widget,project):
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        tolerance = 5
                        if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance: # if the height is not 50 then it means it is expanded already
                            filtered_input = filter_text_input(project["notes"])
                            project["notes"] = filtered_input
                            addition = self.notes_frame_height
                            if "\n" in project["notes"]:
                                notes_rows = project["notes"].split("\n")
                                if len(notes_rows) > 1:
                                    expanded_dim = addition + (len(notes_rows)-1) * 25
                                    # widget[0].configure(height = expanded_dim)
                                    widget[1].configure(state = "normal")
                                    widget[1].configure(height = expanded_dim-10)
                                    if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                                        widget[1].delete("1.0",tk.END)
                                        widget[1].insert(tk.END,str(project["notes"]))

                    if self.default_note_behav == 0:
                        widget[1].configure(state = "disabled")

            def on_leave_entry(widget,project):
                if not opened_window_check():
                    notes_before = filter_text_input(str(project["notes"]))
                    notes_after = filter_text_input(str(widget[1].get("1.0",tk.END)))
                    if str(widget[1]) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        widget[1].configure(state = "normal")
                        if notes_before != notes_after:
                            self.changed_notes_disk = [project["notes"],notes_before]
                            self.undo_edit.configure(state = "normal")
                            self.manage_bin(flag="save_edited_disk",project=project,new_edited_name=project['name'])
                            project["notes"] = notes_after
                            save_changed_notes(notes_after,project)

                        if "\n" in project["notes"]:
                            notes_rows = project["notes"].split("\n")
                            first_row = notes_rows[0]
                            widget[1].delete("1.0",tk.END)
                            widget[1].insert(tk.END,str(first_row))
                        
                        if self.default_note_behav == 0:
                            widget[1].configure(state = "disabled")
                        self.root.focus_set() # unfocus widget
                    else:
                        # jinak pouze ulož změny
                        if notes_before != notes_after:
                            self.manage_bin(flag="save_edited_disk",project=project,new_edited_name=project['name'])
                            project["notes"] = notes_after
                            self.changed_notes_disk = [project["notes"],notes_before]
                            self.undo_edit.configure(state = "normal")
                            save_changed_notes(notes_after,project)
                        self.root.focus_set() # unfocus widget

                    tolerance = 5
                    if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance:
                        return
                    if not opened_window_check():
                        if str(widget[0]) != str(self.last_selected_notes_widget):
                            widget[1].configure(state = "normal")
                            new_height = self.notes_frame_height
                            # widget[0].configure(height = new_height)
                            widget[1].configure(height = new_height-10)
                            if self.default_note_behav == 0:
                                widget[1].configure(state = "disabled")

            def add_row_return(widget):
                addition = widget[0]._current_height
                expanded_dim = addition + 24
                # widget[0].configure(height = expanded_dim)
                widget[1].configure(height = expanded_dim-10)

            if not no_read:
                self.all_project_list = main.DM_tools.read_excel_data(self.excel_file_path)

            Tools.clear_frame(self.project_tree)
            if self.default_disk_status_behav == 1:
                disk_statuses = True
            column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0,width = 50)
            column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column4 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column2_header =    customtkinter.CTkLabel(master = column2,text = "💾",font=("",22)) #💿
            column3_header =    customtkinter.CTkLabel(master = column3,text = "ftp adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column4_header =    customtkinter.CTkLabel(master = column4,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column2_header.     pack(padx = (12,0),side = "top",anchor = "w")
            column3_header.     pack(padx = (5,0),side = "top",anchor = "w",expand = False)
            column4_header.     pack(padx = (5,0),side = "top",anchor = "w")
            self.disk_letter_frame_list = []

            for projects in self.all_project_list:
                btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                button =    customtkinter.CTkButton(master = btn_frame,width=200,height=40,text = projects["name"],font=("Arial",20,"bold"),corner_radius=0,command=lambda widget = btn_frame, project = projects: self.clicked_on_project(project,widget))
                button.     pack(padx =5,pady = 5, fill= "x")
                btn_frame.  pack(side = "top",anchor = "w",expand = False)
                # button.     bind("<Button-1>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                button.     bind("<Double-1>",lambda e,project = projects: self.map_disk(project))
                button.     bind("<Button-3>",lambda e,project = projects: self.show_context_menu(e,project,flag="button"))

                param_frame =   customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                parameter =     customtkinter.CTkLabel(master = param_frame,text = projects["disk_letter"],font=("Arial",20,"bold"),width = 40,height=40)
                parameter.      pack(padx = (5,5),pady = 5)
                param_frame.    pack(side = "top")
                self.disk_letter_frame_list.append(param_frame)
                param_frame.    bind("<Button-1>",lambda e,widget = param_frame, project = projects: self.clicked_on_project(project,widget))
                parameter.      bind("<Button-1>",lambda e,widget = param_frame, project = projects: self.clicked_on_project(project,widget))
                parameter.      bind("<Button-3>",lambda e, project = projects: self.show_context_menu(e,project,flag="disk_letter"))

                param_frame2 =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                parameter2 =     customtkinter.CTkLabel(master = param_frame2,text = projects["ftp"],font=("Arial",20,"bold"),justify='left',anchor = "w",width = 300,height=40)
                parameter2.      pack(padx = (10,5),pady = 5,anchor = "w",fill="x")
                param_frame2.    pack(side = "top",fill="x",expand = False)
                param_frame2.    bind("<Button-1>",lambda e,widget = param_frame2, project = projects: self.clicked_on_project(project,widget))
                parameter2.      bind("<Button-1>",lambda e,widget = param_frame2, project = projects: self.clicked_on_project(project,widget))
                parameter2.      bind("<Button-3>",lambda e, project = projects: self.show_context_menu(e,project,flag="ftp_frame"))

                notes_frame =   customtkinter.CTkFrame(master=column4,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),height=40,corner_radius=0,fg_color="black")
                notes.          pack(padx =5,pady = 5,anchor="w",fill="x")
                notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, project = projects: self.clicked_on_project(project,widget,textbox_widget,flag="notes"))
                notes.          bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, project = projects: self.clicked_on_project(project,widget,textbox_widget,flag="notes"))

                project_notes = str(projects["notes"])
                if "\n" in project_notes:
                    notes_rows = project_notes.split("\n")
                    first_row = notes_rows[0]
                    notes.delete("1.0",tk.END)
                    notes.insert(tk.END,str(first_row))
                else:
                    notes.insert(tk.END,project_notes)
                
                notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],project=projects: on_enter_entry(widget,project))
                notes.bind("<Leave>",lambda e, widget = [notes_frame,notes],project=projects: on_leave_entry(widget,project))
                notes.bind("<Return>",lambda e, widget = [notes_frame,notes]: add_row_return(widget))

                if self.default_note_behav == 0:
                    notes.configure(state = "disabled")

                if projects == self.last_managed_project: # případ že posouvám s projektem nahoru/ dolů/ top (zvíraznit selectnuté)
                    self.selected_list_disk.append(projects)
                    self.last_selected_widget = btn_frame
                    btn_frame.configure(border_color="white")
                    self.remember_to_change_back.append(btn_frame)
                    param_frame.configure(border_color="white")
                    self.remember_to_change_back.append(param_frame)
                    param_frame2.configure(border_color="white")
                    self.remember_to_change_back.append(param_frame2)
                    notes_frame.configure(border_color="white")
                    self.remember_to_change_back.append(notes_frame)

            column1.pack(fill="both",expand=False,side = "left")
            column2.pack(fill="both",expand=False,side = "left")
            column3.pack(fill="both",expand=False,side = "left")
            column4.pack(fill="both",expand=True, side = "left")
            self.project_tree.update()
            self.project_tree.update_idletasks()
            if len(self.all_project_list) > 0:
                try:
                    self.notes_frame_height = int(notes_frame._current_height)
                except Exception:
                    pass
            try:
                self.project_tree._parent_canvas.yview_moveto(0.0)
            except Exception:
                pass
            if disk_statuses:
                if init:
                    self.refresh_disk_statuses(False)
                else:
                    self.refresh_disk_statuses()

        def setting_window(self):
            def save_new_behav_disk():
                nonlocal checkbox2
                if int(checkbox2.get()) == 0:
                    self.default_disk_status_behav = 0
                    Tools.save_to_json_config("init_disk_refresh",0,self.config_filename_path)
                elif int(checkbox2.get()) == 1:
                    self.default_disk_status_behav = 1
                    Tools.save_to_json_config("init_disk_refresh",1,self.config_filename_path)
                    self.make_project_cells_disk(no_read=True)

            def save_new_behav_notes():
                nonlocal checkbox
                if int(checkbox.get()) == 0:
                    self.default_note_behav = 0
                    Tools.save_to_json_config("editable_notes",0,self.config_filename_path)
                    self.make_project_cells_disk()

                elif int(checkbox.get()) == 1:
                    self.default_note_behav = 1
                    Tools.save_to_json_config("editable_notes",1,self.config_filename_path)
                    self.make_project_cells_disk()

            def save_new_disk_map_cond():
                nonlocal checkbox3
                if int(checkbox3.get()) == 0:
                    self.mapping_condition = 0
                    Tools.save_to_json_config("disk_persistent",0,self.config_filename_path)
                elif int(checkbox3.get()) == 1:
                    self.mapping_condition = 1
                    Tools.save_to_json_config("disk_persistent",1,self.config_filename_path)

            def change_make_first_behav():
                nonlocal checkbox4
                if int(checkbox4.get()) == 0:
                    self.make_edited_project_first = False
                    Tools.save_to_json_config("auto_order_when_edit",0,self.config_filename_path)
                elif int(checkbox4.get()) == 1:
                    self.make_edited_project_first = True
                    Tools.save_to_json_config("auto_order_when_edit",1,self.config_filename_path)

            def delete_behav():
                if int(checkbox5.get()) == 0 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 100
                    Tools.save_to_json_config("ask_to_delete",self.deletion_behav,self.config_filename_path)
                elif int(checkbox5.get()) == 0 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 101
                    Tools.save_to_json_config("ask_to_delete",self.deletion_behav,self.config_filename_path)
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 110
                    Tools.save_to_json_config("ask_to_delete",self.deletion_behav,self.config_filename_path)
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 111
                    Tools.save_to_json_config("ask_to_delete",self.deletion_behav,self.config_filename_path)

            def load_old_config():
                def callback_with_path(path_given):
                    try:
                        disk_rows = main.DM_tools.read_excel_data(path_given)
                        main.DM_tools.save_excel_data_disk(self.excel_file_path,disk_rows)    
                        self.make_project_cells_disk()
                        Tools.add_colored_line(self.main_console,"Seznam adres disků ze souboru úspěšně nahrán a uložen","green",None,True)
                            
                    except Exception as e:
                        Tools.add_colored_line(self.main_console,f"Nepodařilo se načíst data z externího souboru: {e}","red",None,True)

                Tools.import_option_window(self.root,self.app_icon,self.initial_path,callback_with_path,setting_window=child_root)

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()

            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Nastavení")
            main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "Chování poznámek (editovatelné/ needitovatelné):",font=("Arial",20,"bold"))
            checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "Přímo zapisovat a ukládat do poznámek na úvodní obrazovce",font=("Arial",16,"bold"),command=lambda: save_new_behav_notes())
            label.          pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox.       pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame2 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label2 =        customtkinter.CTkLabel(master = main_frame2, width = 100,height=40,text = "Chování při vstupu do menu \"Síťové disky\":",font=("Arial",20,"bold"))
            checkbox2 =     customtkinter.CTkCheckBox(master = main_frame2, text = "Při spuštění aktualizovat statusy disků",font=("Arial",16,"bold"),command=lambda: save_new_behav_disk())
            label2.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox2.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame3 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label3 =        customtkinter.CTkLabel(master = main_frame3, width = 100,height=40,text = "Nastavení mapování disků:",font=("Arial",20,"bold"))
            checkbox3 =     customtkinter.CTkCheckBox(master = main_frame3, text = "Automaticky připojovat po restartu PC",font=("Arial",16,"bold"),command=lambda: save_new_disk_map_cond())
            frame_drive1 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
            drive_color1 =  customtkinter.CTkFrame(master=frame_drive1,corner_radius=0,width = 30,height = 30,fg_color="green")
            drive_label1 =  customtkinter.CTkLabel(master = frame_drive1, width = 100,height=40,text = "= disk je online, persistentní (po vypnutí bude znovu načten)",font=("Arial",18))
            drive_color1.   pack(pady = (2,0),padx=10,side="left",anchor = "w")
            drive_label1.   pack(pady = (2,0),padx=0,side="left",anchor = "w")
            frame_drive2 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
            drive_color2 =  customtkinter.CTkFrame(master=frame_drive2,corner_radius=0,width = 30,height = 30,fg_color="#00CED1")
            drive_label2 =  customtkinter.CTkLabel(master = frame_drive2, width = 100,height=40,text = "= disk je online, nepersistentní (bude odpojen po vypnutí)",font=("Arial",18))
            drive_color2.   pack(pady = (2,0),padx=10,side="left",anchor = "w")
            drive_label2.   pack(pady = (2,0),padx=0,side="left",anchor = "w")
            label3.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox3.      pack(pady = 10,padx=10,side="top",anchor = "w")
            frame_drive1.   pack(pady = 0,padx=(2,5),side="top",anchor = "w",fill="x")
            frame_drive2.   pack(pady = 0,padx=(2,5),side="top",anchor = "w",fill="x")
            
            main_frame4 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label4 =        customtkinter.CTkLabel(master = main_frame4, width = 100,height=40,text = "Nastavení chování při editaci projektů:",font=("Arial",20,"bold"))
            checkbox4 =     customtkinter.CTkCheckBox(master = main_frame4, text = "Automaticky přesouvat editovaný projekt na začátek",font=("Arial",16,"bold"),command=lambda: change_make_first_behav())
            label4.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox4.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame5 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label5 =        customtkinter.CTkLabel(master = main_frame5, width = 100,height=40,text = "Odvolit dotazování při mazání:",font=("Arial",20,"bold"))
            checkbox5 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit v hlavním okně",font=("Arial",16,"bold"),command=lambda: delete_behav())
            checkbox6 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit při editu",font=("Arial",16,"bold"),command=lambda: delete_behav())
            label5.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox5.      pack(pady = 0,padx=10,side="top",anchor = "w")
            checkbox6.      pack(pady = (5,5),padx=10,side="top",anchor = "w")

            load_config_frame = customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            load_config_label = customtkinter.CTkLabel(master = load_config_frame, width = 100,height=40,text = "Načíst seznam adres disků (z jiného konfiguračního souboru)",font=("Arial",20,"bold"))
            config_btn_frame =  customtkinter.CTkFrame(master=load_config_frame,corner_radius=0,fg_color="#212121")
            button_load =       customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Zvolit soubor",command = lambda:load_old_config(),font=("Arial",20,"bold"),corner_radius=0)
            button_open =       customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Otevřít aktuální",command = lambda: os.startfile(self.excel_file_path),font=("Arial",20,"bold"),corner_radius=0)
            open_path =         customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Otevřít složku",command = lambda: os.startfile(self.initial_path),font=("Arial",20,"bold"),corner_radius=0)
            load_config_label.  pack(pady = (10,0),padx=10,side="top",anchor = "w")
            button_load.        pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            button_open.        pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            open_path.          pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            config_btn_frame.   pack(pady = 2,padx=2,side="top",fill="x",anchor = "w")

            close_frame =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
            button_close =  customtkinter.CTkButton(master = close_frame, width = 150,height=40,text = "Zavřít",command = child_root.destroy,font=("Arial",20,"bold"),corner_radius=0)
            button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame2.    pack(expand=False,fill="x",side="top")
            main_frame3.    pack(expand=False,fill="x",side="top")
            main_frame4.    pack(expand=False,fill="x",side="top")
            main_frame5.    pack(expand=False,fill="x",side="top")
            load_config_frame.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")

            if self.default_note_behav == 1:
                checkbox.select()
            
            if self.make_edited_project_first:
                checkbox4.select()

            if self.default_disk_status_behav == 1:
                checkbox2.select()

            if self.mapping_condition == 1:
                checkbox3.select()

            if self.deletion_behav == 110 or self.deletion_behav == 111:
                checkbox5.select()
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                checkbox6.select()

            self.opened_window = child_root
            child_root.update()
            child_root.update_idletasks()
            child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{x+350}+{y+180}")
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def sort_by_alphabet(self):
            project_names_array=[]
            for projects in self.all_project_list:
                project_names_array.append(projects['name'])
            project_names_sorted = sorted(project_names_array)
            whole_projects_sorted = []
            for names in project_names_sorted:
                for projects in self.all_project_list:
                    if projects['name'] == names:
                        whole_projects_sorted.append(projects)
                        break
            
            self.all_project_list = copy.deepcopy(whole_projects_sorted)
            main.DM_tools.save_excel_data_disk(self.excel_file_path,self.all_project_list)
            self.make_project_cells_disk()
            Tools.add_colored_line(self.main_console,f"Projekty úspěsně seřazeny podle abecedy","green",None,True)

        def create_widgets_disk(self,init=None):
            Tools.clear_frame(self.root)
            def edit_project():
                result = self.check_given_input()
                if result == True:
                    self.add_new_project_disk(True)
                elif result == None:
                    Tools.add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
        
            if init:
                if self.window_mode == "max":
                    Tools.save_to_json_config("default_window_size",1,self.config_filename_path)
                else:
                    Tools.save_to_json_config("default_window_size",0,self.config_filename_path)
                    
            Tools.clear_frame(self.root)
            self.selected_list_disk = []
            self.control_pressed = False
            Tools.save_to_json_config("disk_or_ip_window",1,self.config_filename_path)
            top_frame =                     customtkinter.CTkFrame(master=self.root,corner_radius=0,border_width=0,fg_color="#212121")
            top_left_frame =                customtkinter.CTkFrame(master=top_frame,corner_radius=0,border_width=0,fg_color="#212121")
            top_right_frame =               customtkinter.CTkFrame(master=top_frame,corner_radius=0,border_width=0,fg_color="#212121")
            menu_cards =                    customtkinter.CTkFrame(master=top_left_frame,corner_radius=0,fg_color="#636363",height=50)
            self.main_widgets =             customtkinter.CTkFrame(master=top_left_frame,corner_radius=0)
            self.project_tree =             customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
            logo =                          customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
            image_logo =                    customtkinter.CTkLabel(master = top_right_frame,text = "",image =logo,bg_color="#212121")
            main_menu_button =              customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_all_ip =          customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: main.IP_assignment(self.parent_instance,fav_w_called=False),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_favourite_ip =    customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: main.IP_assignment(self.parent_instance,fav_w_called=True),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_disk =            customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
            first_row_frame =               customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,fg_color="#212121")
            project_label =                 customtkinter.CTkLabel(master = first_row_frame,height=40,text = "Projekt: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.search_input =             customtkinter.CTkEntry(master = first_row_frame,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
            # button_search =                 customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first_disk("search"),font=("Arial",20,"bold"),corner_radius=0)
            search_icon =                   customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/SearchWhite.png")),size=(32, 32)),bg_color="#212121")
            search_icon.bind("<Enter>",lambda e: search_icon._image.configure(size=(36,36)))
            search_icon.bind("<Leave>",lambda e: search_icon._image.configure(size=(32,32)))
            search_icon.bind("<Button-1>",lambda e: self.make_project_first_disk("search"))
            # button_add =                    customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project_disk(),font=("Arial",20,"bold"),corner_radius=0)
            new_project_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/green_plus.png")),size=(32, 32)),bg_color="#212121")
            new_project_icon.bind("<Enter>",lambda e: new_project_icon._image.configure(size=(36,36)))
            new_project_icon.bind("<Leave>",lambda e: new_project_icon._image.configure(size=(32,32)))
            new_project_icon.bind("<Button-1>",lambda e: self.add_new_project_disk())
            button_remove =                 customtkinter.CTkButton(master = first_row_frame, width = 100,height=40,text = "Smazat", command =  lambda: self.delete_project_disk(flag="main_menu"),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_button =              customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_deleted_disk"),font=("",28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_edit =                   customtkinter.CTkButton(master = first_row_frame, width = 110,height=40,text = "Editovat",command =  lambda: edit_project(),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_edit =                customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_edited_disk"),font=("",28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_make_first =             customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "🔝",command =  lambda: self.make_project_first_disk(),font=(None,30),corner_radius=0)
            move_upwards =                  customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↑",command =  lambda: self.make_project_first_disk(purpouse="silent",upwards=True),font=(None,25),corner_radius=0)
            move_downwards =                customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↓",command =  lambda: self.make_project_first_disk(purpouse="silent",downwards=True),font=(None,25),corner_radius=0)
            sort_alphabet =                 customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "A↑",command =  lambda: self.sort_by_alphabet(),font=(None,25),corner_radius=0)
            button_settings =               customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(),font=("",22),corner_radius=0)
            # second_row_frame =              customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,fg_color="#212121")
            # delete_disk =                   customtkinter.CTkButton(master = second_row_frame, width = 250,height=40,text = "Odpojit síťový disk",command =  lambda: self.delete_disk_option_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
            unplug_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/unplug.png")),size=(32, 32)),bg_color="#212121")
            unplug_icon.bind("<Enter>",lambda e: unplug_icon._image.configure(size=(36,36)))
            unplug_icon.bind("<Leave>",lambda e: unplug_icon._image.configure(size=(32,32)))
            unplug_icon.bind("<Button-1>",lambda e: self.delete_disk_option_menu())
            # reset =                         customtkinter.CTkButton(master = second_row_frame, width = 200,height=40,text = "Reset exploreru",command = lambda: self.refresh_explorer(refresh_disk=True),font=("Arial",20,"bold"),corner_radius=0)
            reset_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/reset.png")),size=(32, 32)),bg_color="#212121")
            reset_icon.bind("<Enter>",lambda e: reset_icon._image.configure(size=(36,36)))
            reset_icon.bind("<Leave>",lambda e: reset_icon._image.configure(size=(32,32)))
            reset_icon.bind("<Button-1>",lambda e: self.refresh_explorer(refresh_disk=True))
            # self.refresh_btn =              customtkinter.CTkButton(master = second_row_frame, width = 200,height=40,text = "Refresh statusů",command = lambda: self.refresh_disk_statuses(silent=False),font=("Arial",20,"bold"),corner_radius=0)
            refresh_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/refresh.png")),size=(32, 32)),bg_color="#212121")
            refresh_icon.bind("<Enter>",lambda e: refresh_icon._image.configure(size=(36,36)))
            refresh_icon.bind("<Leave>",lambda e: refresh_icon._image.configure(size=(32,32)))
            refresh_icon.bind("<Button-1>",lambda e: self.refresh_disk_statuses(silent=False))
            as_admin_label =                customtkinter.CTkLabel(master = first_row_frame,text = "",font=("Arial",20,"bold"))
            third_row_frame =               customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#212121")
            self.main_console =             tk.Text(third_row_frame, wrap="none", height=0,background="black",font=("Arial",22),state=tk.DISABLED)
            project_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.search_input.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            # button_search.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            search_icon.                    pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            # button_add.                     pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            new_project_icon.               pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            button_remove.                  pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            self.undo_button.               pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_edit.                    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_edit.                 pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_make_first.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_upwards.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_downwards.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            sort_alphabet.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_settings.                pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            # delete_disk.                    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            unplug_icon.                    pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            # reset.                          pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            reset_icon.                     pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            # self.refresh_btn.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            refresh_icon.                   pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            as_admin_label.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.main_console.              pack(pady = (0,0),padx =(5,0),anchor="w",side="top",fill="x",expand=False)
            first_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            # second_row_frame.               pack(pady=0,padx=0,fill="x",side = "top")
            main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_all_ip.           pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_favourite_ip.     pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            image_logo.                     pack(anchor = "e",side = "top",ipadx = 20,ipady = 20,expand=False)
            menu_cards.                     pack(pady=0,padx=5,fill="x",expand=False,side = "top")
            self.main_widgets.              pack(pady=0,padx=0,fill="both",side = "top")
            top_left_frame.                 pack(pady=0,padx=0,fill="x",side = "left",expand=True,anchor="n")
            top_right_frame.                pack(pady=0,padx=0,fill="y",side = "right",expand=False)
            top_frame.                      pack(pady=0,padx=0,fill="x",side = "top")
            third_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            self.project_tree.              pack(pady=5,padx=5,fill="both",expand=True,side = "top")

            main.ToolTip(new_project_icon," Nový projekt ",self.root)
            main.ToolTip(search_icon," Vyhledat projekt ",self.root)
            main.ToolTip(reset_icon," Reset exploreru ",self.root)
            main.ToolTip(unplug_icon," Odpojit síťový disk ",self.root)
            main.ToolTip(refresh_icon," Refresh statusu disků ",self.root)
            main.ToolTip(button_make_first," Přesunout projekt na začátek ",self.root)
            main.ToolTip(self.undo_button," Vrátit poslední smazaný projekt ",self.root)
            main.ToolTip(self.undo_edit," Vrátit poslední změnu ",self.root)
            main.ToolTip(move_upwards," Posunout o pozici výše ",self.root)
            main.ToolTip(move_downwards," Posunout o pozici níže ",self.root)
            main.ToolTip(sort_alphabet," Seřadit podle abecedy ",self.root)
            main.ToolTip(button_settings," Nastavení ",self.root)

            config_data = Tools.read_json_config(self.config_filename_path)
            if len(config_data["edited_project_bin_disk"])>0:
                self.undo_edit.configure(state = "normal")
            else:
                self.undo_edit.configure(state = "disabled")
            if len(config_data["deleted_project_bin_disk"])>0:
                self.undo_button.configure(state = "normal")
            else:
                self.undo_button.configure(state = "disabled")
            
            def is_admin():
                try:
                    return ctypes.windll.shell32.IsUserAnAdmin()
                except:
                    return False
                
            if is_admin():
                as_admin_label.configure(text = "Aplikace je spuštěna, jako administrátor\n(mapovat disky lze pouze na uživatelském účtu)",text_color = "orange")

            def maximalize_window(e):
                self.root.update_idletasks()
                self.root.update()
                current_width = int(self.root.winfo_width())
                # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
                if Tools.focused_entry_widget(self.root): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                    return
                if int(current_width) > 1200:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                    Tools.save_to_json_config("default_window_size",2,self.config_filename_path)
                elif int(current_width) ==260:
                    self.root.geometry("1200x900")
                    Tools.save_to_json_config("default_window_size",0,self.config_filename_path)
                else:
                    self.root.state('zoomed')
                    Tools.save_to_json_config("default_window_size",1,self.config_filename_path)
                self.root.update_idletasks()
                self.root.update()
            self.root.bind("<f>",lambda e: maximalize_window(e))

            def unfocus_widget(e):
                self.root.focus_set()
            self.root.bind("<Escape>",unfocus_widget)
            self.search_input.bind("<Return>",unfocus_widget)

            def call_search(e):
                self.make_project_first_disk("search")
            self.search_input.bind("<Return>",call_search)

            def call_refresh(e):
                # self.refresh_explorer(refresh_disk=True)
                self.refresh_disk_statuses(silent=False)
            self.root.bind("<F5>",lambda e: call_refresh(e))

            def call_unfocus(e):
                widget = str(e.widget)
                if not ".!ctkscrollableframe" in widget and not ".!ctktoplevel" in widget and not ".!ctkbutton" in widget:
                    #odebrat focus
                    self.clicked_on_project(None,None,None,flag="unfocus")
            self.root.bind("<Button-1>",call_unfocus,"+")

            def control_button(status):
                self.control_pressed = status
                if status == True:
                    if self.last_managed_project is None:
                        return
                    if not self.last_managed_project in self.selected_list_disk:
                        self.selected_list_disk.append(self.last_managed_project)

            def multi_select():
                if self.last_managed_project is None:
                    return
                if not self.last_managed_project in self.selected_list_disk:
                    self.selected_list_disk.append(self.last_managed_project)
                    # print("selected_list - ",self.selected_list_disk)

            self.root.bind("<Control_L>",lambda e: control_button(True))
            self.root.bind("<Control-Button-1>",lambda e: multi_select())
            self.root.bind("<KeyRelease-Control_L>",lambda e: control_button(False))
            self.root.bind("<Delete>",lambda e: self.delete_project_disk(flag="main_menu"))
            self.root.update()
            self.make_project_cells_disk(disk_statuses=False,init=True)
            # self.root.mainloop()

    class IP_assignment: # Umožňuje měnit statickou IP
        """
        Umožňuje měnit nastavení statických IP adres
        """  
        
        def __init__(self,parent,fav_w_called=None):
            self.parent_instance = parent
            self.root = parent.root
            self.menu_callback = parent.menu_callback
            self.window_mode = parent.window_mode
            self.initial_path = parent.initial_path
            self.app_icon = parent.app_icon
            self.excel_file_path = parent.excel_file_path
            self.config_filename_path = parent.config_filename_path
            self.last_managed_project = None
            self.make_project_favourite = False
            self.connection_option_list = []
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.opened_window = ""
            self.ip_frame_list = []
            self.selected_list = []
            self.remember_to_change_back = []
            self.control_pressed = False
            self.edited_project_name = None
            self.deleted_projects_bin = []
            self.edited_projects_bin = []
            self.changed_notes = []
            self.notes_frame_height = 50

            read_parameters = Tools.read_json_config(self.config_filename_path)
            if read_parameters != None:
                self.default_connection_option = read_parameters["default_ip_interface"]
                if read_parameters["favorite_ip_window_status"] == 1:
                    self.show_favourite = True
                else:
                    self.show_favourite = False
                if read_parameters["default_window_size"] == 2:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                self.default_note_behav = read_parameters["editable_notes"]
                if read_parameters["auto_order_when_edit"] == 1:
                    self.make_edited_project_first = True
                else:
                    self.make_edited_project_first = False
                self.deletion_behav = read_parameters["ask_to_delete"]
            else:
                self.default_connection_option = 0
                self.show_favourite = False
                self.default_note_behav = 0
                self.make_edited_project_first = True
                self.deletion_behav = 100
            
            if parent.default_environment == "config_load_error":
                self.create_widgets(init=True,excel_load_error=True)
            elif fav_w_called == None:
                self.manage_bin("read_sheet")
                self.create_widgets(fav_status=self.show_favourite,init=True)
            else:
                self.manage_bin("read_sheet")
                self.create_widgets(fav_status=fav_w_called,init=True)

        def call_menu(self): # Tlačítko menu (konec, návrat do menu)
            """
            Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
            """

            Tools.clear_frame(self.main_widgets)
            Tools.clear_frame(self.root)
            # self.root.unbind("<f>")
            # self.root.unbind("<Escape>")
            # self.root.unbind("<F5>")
            # self.root.unbind("<Button-1>")
            # self.root.unbind("<Control_L>")
            # self.root.unbind("<Control-Button-1>")
            # self.root.unbind("<KeyRelease-Control_L>")
            # self.root.unbind("<Delete>")
            self.root.update()
            self.root.update_idletasks()
            self.menu_callback()
    
        def manage_bin(self,flag="",project=None,new_edited_name = None):
            """
            flag:\n
            - read_sheet
            - save_project_ip
            - load_deleted_ip
            - save_edited_ip
            - load_edited_ip
            """
            max_stored_deletions = 5
            max_stored_edits = 10

            def read_sheet():
                config_data = Tools.read_json_config(self.config_filename_path)
                try:
                    self.deleted_projects_bin = config_data["deleted_project_bin"]
                except Exception:
                    Tools.save_to_json_config("deleted_project_bin",self.deleted_projects_bin,self.config_filename_path)

                # self.edited_projects_bin = config_data["edited_project_bin"]
                Tools.save_to_json_config("edited_project_bin",[],self.config_filename_path) #vymazat historii editu při zapnutí
                    
            def save_project_ip():# saving after deleting:
                if project == None:
                    return
                config_data = Tools.read_json_config(self.config_filename_path)
                self.deleted_projects_bin = config_data["deleted_project_bin"]
                self.undo_button.configure(state = "normal")
                self.deleted_projects_bin.insert(0,project)

                if len(self.deleted_projects_bin) > max_stored_deletions:
                    self.deleted_projects_bin.pop()
                Tools.save_to_json_config("deleted_project_bin",self.deleted_projects_bin,self.config_filename_path)

            def save_edited_ip():# saving after editing:
                if project == None or new_edited_name == None:
                    return
                self.undo_edit.configure(state = "normal")
                config_data = Tools.read_json_config(self.config_filename_path)
                self.edited_projects_bin = config_data["edited_project_bin"]
                project["new_name"] = new_edited_name
                print("\nSAVING: ",project)
                self.edited_projects_bin.insert(0,project)
                if len(self.edited_projects_bin) > max_stored_edits:
                    self.edited_projects_bin.pop()
                Tools.save_to_json_config("edited_project_bin",self.edited_projects_bin,self.config_filename_path)

            def load_deleted_ip():
                """
                adds new project from history and deletes the history
                """
                config_data = Tools.read_json_config(self.config_filename_path)
                self.deleted_projects_bin = config_data["deleted_project_bin"]
                project_to_load = self.deleted_projects_bin[0]
                self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                modified_project_name = Tools.get_unique_name(self.all_project_list,project_to_load['name'])
                project_to_load["name"] = modified_project_name
                self.deleted_projects_bin.pop(0)
                if len(self.deleted_projects_bin) ==0:
                    self.undo_button.configure(state = "disabled")

                Tools.save_to_json_config("deleted_project_bin",self.deleted_projects_bin,self.config_filename_path)
                self.all_project_list.insert(0,project_to_load)
                main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)
                Tools.add_colored_line(self.main_console,f"Projekt: {project_to_load['name']} byl úspěšně obnoven","green",None,True)
                self.make_project_cells()

            def load_edited_ip():
                config_data = Tools.read_json_config(self.config_filename_path)
                self.edited_projects_bin = config_data["edited_project_bin"]
                project_to_load = self.edited_projects_bin[0]
                print("project to load: ",project_to_load)
                old_project_name = str(project_to_load['name'])
                current_project_name = str(project_to_load["new_name"])
                self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                project_index = Tools.get_project_index(self.all_project_list,current_project_name)

                if project_index == None:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {current_project_name} nenalezeno, nelze ho tedy obnovit","red",None,True)
                    
                self.edited_projects_bin.pop(0)
                if len(self.edited_projects_bin) ==0:
                    self.undo_edit.configure(state = "disabled")
                Tools.save_to_json_config("edited_project_bin",self.edited_projects_bin,self.config_filename_path)
                if project_index == None: #let it to be deleted... no use, corrupted
                    return
                
                print(project_to_load,"\n",self.all_project_list[project_index])
                self.all_project_list[project_index]["name"] = str(project_to_load["name"])
                self.all_project_list[project_index]["ip"] = str(project_to_load["ip"])
                self.all_project_list[project_index]["mask"] = str(project_to_load["mask"])
                self.all_project_list[project_index]["notes"] = str(project_to_load["notes"])
                self.all_project_list[project_index]["fav_status"] = str(project_to_load["fav_status"])

                # self.all_project_list.insert(0,project_to_load)
                main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)
                if old_project_name != current_project_name:
                    Tools.add_colored_line(self.main_console,f"U projektu: {old_project_name} (původně: {current_project_name}) byly odebrány provedené změny","green",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"U projektu: {old_project_name} byly odebrány provedené změny","green",None,True)
                self.make_project_cells()

            mapping_logic = {
                "read_sheet": read_sheet,
                "save_project_ip": save_project_ip,
                "load_deleted_ip": load_deleted_ip,
                "save_edited_ip": save_edited_ip,
                "load_edited_ip": load_edited_ip,
            }

            output = mapping_logic[flag]()  # This will call the corresponding function
            return output

        def switch_fav_status_new(self,project,wanted_status:str,refresh = False):
            project_index = Tools.get_project_index(self.all_project_list,project['name'])
            project["fav_status"] = str(wanted_status)
            self.all_project_list[project_index] = project
            main.IP_tools.save_excel_data(self.excel_file_path, self.all_project_list)
            if refresh:
                if wanted_status == "1":
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} byl přidán do oblíbených","green",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} byl odebrán z oblíbených","green",None,True)

                self.make_project_cells()

        def delete_project(self,wanted_project=None,silence=None,flag=""):
            if "!ctktextbox" in str(self.root.focus_get()):
                return
        
            project_found = False
            name_list = []

            def check_multiple_projects(window):
                nonlocal wanted_project
                nonlocal name_list
                nonlocal project_found

                if len(self.selected_list) > 1:
                    for names in name_list:
                        print(names)
                        project_found = False
                        self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                        proceed(names,window,True)
                        # print(deleted_project)
                            
                    Tools.add_colored_line(self.main_console,f"Byly úspěšně odstraněny tyto projekty: {name_list}","orange",None,True)
                    try:
                        self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                    except Exception as e:
                        print("chyba, refresh po mazani")
                else:
                    proceed(wanted_project,window)

            def proceed(wanted_project,window = True,multiple_status=False):
                nonlocal project_found
                nonlocal child_root
                deleted_project = None
                if wanted_project == None:
                    self.all_project_list= main.IP_tools.read_excel_data(self.excel_file_path)
                    wanted_project = str(self.search_input.get())
                workbook = load_workbook(self.excel_file_path)
                excel_worksheet = "ip_address_list"
                worksheet = workbook[excel_worksheet]

                project_to_delete = Tools.found_project_name(self.all_project_list,wanted_project)
                if project_to_delete == False:
                    Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    return
                if not self.show_favourite:
                    # když mažu z oblíbených, tak neukládám historii
                    self.manage_bin(flag="save_project_ip",project=project_to_delete)
                    project_index = Tools.get_project_index(self.all_project_list,project_to_delete['name'])
                    worksheet.delete_rows(project_index+1)
                    workbook.save(self.excel_file_path)
                    workbook.close()
                    project_found = True
                    deleted_project = project_to_delete
                else: #jen změnit status u oblíbených
                    self.switch_fav_status_new(project_to_delete,"0")
                    Tools.add_colored_line(self.main_console,f"Projekt {wanted_project} byl odebrán z oblíbených","orange",None,True)
                    self.make_project_cells()
                    return
                
                if silence == None and not multiple_status:
                    if project_found:
                        Tools.add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                        self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                    elif wanted_project.replace(" ","") == "":
                        Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)

                elif project_found and not multiple_status:  # zresetuj i v pripade silence...
                    self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...

                if window and child_root.winfo_exists():
                    child_root.grab_release()
                    child_root.destroy()

                return deleted_project
            
            if self.last_managed_project is None:
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif str(self.last_managed_project['name']).replace(" ","") == "":
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif wanted_project == None:
                wanted_project = self.last_managed_project['name']

            if self.deletion_behav == 101 or self.deletion_behav == 111 or self.show_favourite == True: #pokud odvolené dotazování nebo jsme v oblíbených
                check_multiple_projects(False)
                return
            
            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Upozornění")
            proceed_label_text = f"Opravdu si přejete odstranit projekt {self.last_managed_project['name']}?"

            # if flag == "context_menu":
            #     self.selected_list = []
            if len(self.selected_list) > 1:
                for projects in self.selected_list:
                    if str(projects['name']) not in name_list:
                        name_list.append(str(projects['name']))
                proceed_label_text = f"Opravdu si přejete odstranit vybrané projekty:\n{name_list}?"
                
            proceed_label = customtkinter.CTkLabel(master = child_root,text = proceed_label_text,font=("Arial",22,"bold"),justify = "left",anchor="w")
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: check_multiple_projects(True))
            button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
            proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top",anchor="w")
            button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
            child_root.update()
            child_root.update_idletasks()
            child_root.grab_set()
            child_root.focus()
            child_root.focus_force()
            child_root.wait_window()
            return project_found

        def add_new_project(self,edit = None,init_copy = False,childroot_given = None):
            def mouse_wheel_change(e):
                if -e.delta < 0:
                    switch_up()
                else:
                    switch_down()

            def copy_previous_project():
                try:
                    if self.last_managed_project['name'] == "":
                        Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
                        return
                    self.name_input.delete("0","300")
                    self.name_input.insert("0",str(self.last_managed_project['name']))
                    self.IP_adress_input.delete("0","300")
                    self.IP_adress_input.insert("0",str(self.last_managed_project["ip"]))
                    self.mask_input.delete("0","300")
                    self.mask_input.insert("0",str(self.last_managed_project["mask"]))
                    self.notes_input.delete("1.0",tk.END)
                    self.notes_input.insert(tk.END,str(self.last_managed_project["notes"]))
                except TypeError:
                    Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)

            def switch_up(force_index = None):
                if force_index != None:
                    project_index = force_index
                else:
                    project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])

                another_project_id = project_index
                another_project_id -= 1
                if another_project_id < 0:
                    another_project_id = len(self.all_project_list)-1
                    
                self.last_managed_project = self.all_project_list[another_project_id]
                copy_previous_project()
                refresh_favourite_status()
                refresh_title()

            def switch_down():
                project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                previous_project_id = project_index
                previous_project_id += 1
                if previous_project_id > len(self.all_project_list)-1:
                    previous_project_id = 0

                self.last_managed_project = self.all_project_list[previous_project_id]
                copy_previous_project()
                refresh_favourite_status()
                refresh_title()

            def del_project():
                nonlocal child_root
                project_index = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                result = self.delete_project(wanted_project=self.last_managed_project['name'])
                print(result)
                if result:
                    switch_up(project_index)
                else:
                    print("aborted")

                child_root.focus()
                child_root.focus_force()
                child_root.grab_set()

            def refresh_favourite_status():
                if self.last_managed_project is None: # nový projekt
                    fav_status = "0"
                else:
                    fav_status = str(self.last_managed_project["fav_status"])

                if fav_status == "1":
                    self.make_project_favourite = True #init hodnota
                    self.make_fav_label.configure(text = "Oblíbený",font=("Arial",22))
                    self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
                else:
                    self.make_project_favourite = False #init hodnota
                    self.make_fav_label.configure(text = "Neoblíbený",font=("Arial",22))
                    self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")

            def refresh_title():
                if edit:
                    child_root.title("Editovat projekt: "+ str(self.last_managed_project['name']))
                else:
                    child_root.title("Nový projekt")

            def make_favourite_toggle_via_edit(e):
                def do_favourite():
                    self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
                    self.make_fav_label.configure(text = "Oblíbený")
                
                def unfavourite():
                    self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")
                    self.make_fav_label.configure(text = "Neoblíbený")

                if self.make_project_favourite:
                    self.make_project_favourite = False
                    unfavourite()
                else:
                    self.make_project_favourite = True
                    do_favourite()

            def save_project(add_next = False):
                def check_ip_and_mask(input_value):
                    input_splitted = input_value.split(".")
                    if len(input_splitted) == 4:
                        return input_value
                    else:
                        return False

                project_name = str(self.name_input.get())
                IP_adress = str(self.IP_adress_input.get())
                IP_adress = check_ip_and_mask(IP_adress)
                mask = str(self.mask_input.get())
                mask = check_ip_and_mask(mask)
                notes = Tools.get_legit_notes(self.notes_input.get("1.0", tk.END))
                errors = 0
                if project_name.replace(" ","") == "":
                    Tools.add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
                    errors += 1
                if IP_adress == False and errors == 0:
                    Tools.add_colored_line(self.console,f"Neplatná IP adresa","red",None,True)
                    errors += 1
                if mask == False and errors == 0:
                    Tools.add_colored_line(self.console,f"Neplatná maska","red",None,True)
                    errors += 1
                if errors>0:
                    return
                
                self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                fav_status = "0"
                if self.make_project_favourite:
                    fav_status = "1"

                if edit:
                    print("last_managed project:", self.last_managed_project)
                    currently_edited_project_id = Tools.get_project_index(self.all_project_list,self.last_managed_project['name'])
                    if currently_edited_project_id == None:
                        Tools.add_colored_line(self.main_console,f"Projekt {self.last_managed_project['name']} nenalezen","red",None,True)
                        return
                    
                    if self.all_project_list[currently_edited_project_id]['name'] != project_name:
                        project_name = Tools.get_unique_name(self.all_project_list,project_name)
                    self.all_project_list[currently_edited_project_id]['name'] = project_name
                    self.all_project_list[currently_edited_project_id]["ip"] = IP_adress
                    self.all_project_list[currently_edited_project_id]["mask"] = mask
                    self.all_project_list[currently_edited_project_id]["notes"] = notes
                    self.all_project_list[currently_edited_project_id]["fav_status"] = fav_status
                    main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)
                    if fav_status != self.last_managed_project["fav_status"]: #pokud doslo ke zmenene statusu
                        if self.make_project_favourite:
                            if self.last_managed_project['name'] != project_name:
                                status_text = f"Projekt: {self.last_managed_project['name']} (nově: {project_name}) úspěšně pozměněn a přidán do oblíbených"
                            else:
                                status_text = f"Projekt: {self.last_managed_project['name']} úspěšně pozměněn a přidán do oblíbených"
                            Tools.add_colored_line(self.main_console,status_text,"green",None,True)
                        else:
                            if self.last_managed_project['name'] != project_name:
                                status_text = f"Projekt: {self.last_managed_project['name']} (nově: {project_name}) úspěšně pozměněn a odebrán z oblíbených"
                            else:
                                status_text = f"Projekt: {self.last_managed_project['name']} úspěšně pozměněn a odebrán z oblíbených"
                            Tools.add_colored_line(self.main_console,status_text,"green",None,True)
                    else:
                        if self.last_managed_project['name'] != project_name:
                            status_text = f"Projekt: {self.last_managed_project['name']} (nově: {project_name}) úspěšně pozměněn"
                        else:
                            status_text = f"Projekt: {self.last_managed_project['name']} úspěšně pozměněn"
                        Tools.add_colored_line(self.main_console,status_text,"green",None,True)
                    
                    
                    self.manage_bin(flag="save_edited_ip",project=self.last_managed_project,new_edited_name=project_name)
                    if self.make_edited_project_first:
                        # self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                        self.make_project_first(purpouse="silent",make_cells=False,project=self.all_project_list[currently_edited_project_id],input_entry_bypass=project_name)
                else:
                    project_name = Tools.get_unique_name(self.all_project_list,project_name)
                    self.all_project_list.insert(0,{
                        'name':project_name,
                        "ip":IP_adress,
                        "mask":mask,
                        "notes":notes,
                        "fav_status":fav_status,
                    })
                    main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)
                    Tools.add_colored_line(self.main_console,f"Nový projekt: {project_name} byl úspěšně přidán","green",None,True)

                self.make_project_cells()
                if add_next:
                    Tools.clear_frame(child_root)
                    self.root.after(10,self.add_new_project(childroot_given=child_root))
                else:
                    child_root.destroy()

            if childroot_given == None:
                child_root = customtkinter.CTkToplevel(fg_color="#212121")
            else:
                child_root = childroot_given

            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            refresh_title()
            top_main_frame =        customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,fg_color="#212121")
            top_left_frame =        customtkinter.CTkFrame(master=top_main_frame,corner_radius=0,border_width=2,fg_color="#212121")
            top_right_frame =       customtkinter.CTkFrame(master=top_main_frame,corner_radius=0,border_width=2,fg_color="#212121")
            project_name =          customtkinter.CTkLabel(master = top_left_frame, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
            self.name_input =       customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            project_selection_label = customtkinter.CTkLabel(master = top_right_frame, width = 200,height=30,text = "Přepnout projekt: ",font=("Arial",20,"bold"))
            project_switch_frame =  customtkinter.CTkFrame(master=top_right_frame,corner_radius=0,height=140,width=80)
            project_up =            customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↑",command= lambda: switch_up())
            project_down =          customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↓",command= lambda: switch_down())
            project_up              .pack(pady=(0,5),padx=5,side = "top",fill="x")
            project_down            .pack(pady=0,padx=5,side = "top",fill="x")
            project_switch_frame.   bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_up.             bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_down.           bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            IP_adress =            customtkinter.CTkLabel(master = top_left_frame, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
            self.IP_adress_input = customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            mask =                 customtkinter.CTkLabel(master = top_left_frame, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
            self.mask_input =      customtkinter.CTkEntry(master = top_left_frame,font=("Arial",20),width=200,height=30,corner_radius=0)
            copy_check =           customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: copy_previous_project())
            save_and_add_next =    customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Uložit a přidat další",command= lambda: save_project(add_next = True))
            del_project_btn =      customtkinter.CTkButton(master = top_right_frame,font=("Arial",20),width=250,height=30,corner_radius=0,text="Smazat tento projekt",command= lambda: del_project(),fg_color="red")
            fav_status =           customtkinter.CTkLabel(master = top_left_frame, width = 20,height=30,text = "Status oblíbenosti: ",font=("Arial",20,"bold"))
            fav_frame =            customtkinter.CTkFrame(master=top_left_frame,corner_radius=0,border_width=0,height=50,fg_color="#353535",border_color="#606060")
            self.make_fav_label =  customtkinter.CTkLabel(master = fav_frame, width = 20,height=30)
            self.make_fav_btn =    customtkinter.CTkLabel(master = fav_frame, width = 50,height=50)
            refresh_favourite_status()
            def really_leaving(e,frame):
                x = frame.winfo_width()-1
                y = frame.winfo_height()-1
                if (e.x < 1 or e.x > x) or (e.y<1 or e.y > y):
                    fav_frame.configure(border_width = 0,fg_color = "#353535")

            fav_frame.bind("<Enter>",lambda e: fav_frame.configure(border_width = 2,fg_color = "#404040"))
            fav_frame.bind("<Leave>",lambda e ,frame = fav_frame: really_leaving(e,frame))
            notes_label =          customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
            self.notes_input =     customtkinter.CTkTextbox(master = child_root,font=("Arial",20),height=280,corner_radius=0)
            self.console =         tk.Text(child_root, wrap="none", height=0,background="black",font=("Arial",14),state=tk.DISABLED)
            buttons_frame =        customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,fg_color="#212121")
            save_button =  customtkinter.CTkButton(master = buttons_frame, width = 200,height=40,text = "Uložit", command = lambda: save_project(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =  customtkinter.CTkButton(master = buttons_frame, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)

            project_name.pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            if edit:
                project_selection_label.pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
                project_switch_frame.pack(pady=10,padx=10,anchor="w",side = "top",fill="x")
            else:
                copy_check.         pack(pady = (10),padx =(10),anchor="w",side="top")
                save_and_add_next.  pack(pady = (0,10),padx =(10),anchor="w",side="top")
            self.name_input.        pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            IP_adress.              pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.IP_adress_input.   pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            mask.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.mask_input.        pack(pady = (10,0),padx =(5,5),anchor="w",side="top",fill="x",expand=True)
            fav_status.             pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            if edit and not self.show_favourite:
                del_project_btn.    pack(pady = (10),padx =(10),anchor="s",side="left")
            fav_frame.              pack(pady = (10),padx =(10),anchor="w",side="top",fill="x",expand=True)
            self.make_fav_btn.      pack(pady = (10),padx =(5,0),anchor="w",side="left")
            self.make_fav_label.    pack(pady = (10),padx =(5,0),anchor="w",side="left")
            fav_frame.              bind("<Button-1>",lambda e: make_favourite_toggle_via_edit(e))
            self.make_fav_btn.      bind("<Button-1>",lambda e: make_favourite_toggle_via_edit(e))
            self.make_fav_label.    bind("<Button-1>",lambda e: make_favourite_toggle_via_edit(e))
            top_left_frame.         pack(anchor="w",side="left",fill="both",expand = True)
            top_right_frame.        pack(anchor="e",side="right",fill="y",expand = False,ipadx=2,ipady=2)
            top_main_frame.         pack(anchor="w",side="top",fill="both",expand = True)
            notes_label.            pack(pady = (10,0),padx =(5,0),anchor="w",side="top")
            self.notes_input.       pack(pady = (10,0),padx =(5),anchor="w",side="top",fill="both",expand=True)
            self.console.           pack(pady = (10,0),padx =(5),anchor="w",side="top",fill="x",expand=False)
            exit_button.            pack(pady = (10,0),padx =(5,0),anchor="e",side="right")
            save_button.            pack(pady = (10,0),padx =(5,0),anchor="e",side="right")
            buttons_frame.          pack(pady = (0,10),padx =(0,10),anchor="w",side="top",fill="x",expand=False)
            if edit or init_copy:
                copy_previous_project()
            else:
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0","192.168.000.000")
                self.mask_input.delete("0","300")
                self.mask_input.insert("0","255.255.255.0")
                # if str(self.search_input.get()).replace(" ","") != "":
                #     self.name_input.delete("0","300")
                #     self.name_input.insert("0",str(self.search_input.get()))

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
        
        def make_sure_ip_changed(self,interface_name,ip):
            def run_as_admin():
                # Vyžádání admin práv: nefunkční ve vscode
                def is_admin():
                    try:
                        return ctypes.windll.shell32.IsUserAnAdmin()
                    except:
                        return False
                pid = os.getpid()
                if not is_admin():
                    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(["admin_ip_setting",str(pid)]), None, 1)
                    sys.exit()
            def open_app_as_admin_prompt():
                def close_prompt(child_root):
                    child_root.destroy()
                child_root = customtkinter.CTkToplevel()
                child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
                self.opened_window = child_root
                x = self.root.winfo_rootx()
                y = self.root.winfo_rooty()
                child_root.geometry(f"620x150+{x+300}+{y+300}")  
                child_root.title("Upozornění")
                proceed_label = customtkinter.CTkLabel(master = child_root,text = "Přejete si znovu spustit aplikaci, jako administrátor?",font=("Arial",25))
                button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin())
                button_no =     customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
                proceed_label   .pack(pady=(15,0),padx=10,anchor="w",expand=False,side = "top")
                button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
                button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
                child_root.update()
                child_root.update_idletasks()
                child_root.focus()
                child_root.focus_force()
                self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

            interface_index = self.connection_option_list.index(interface_name)

            def call_subprocess():
                try:

                    if ip == self.current_address_list[interface_index]:
                        Tools.add_colored_line(self.main_console,f"Pro interface {interface_name} je již tato adresa ({ip}) nastavena","orange",None,True)
                        return
                    elif ip in self.current_address_list:
                        Tools.add_colored_line(self.main_console,f"Chyba, adresa je již používána pro jiný interface","red",None,True)
                        return
                    win_change_ip_time = 7
                    for i in range(0,win_change_ip_time):
                        Tools.add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                        self.option_change("",silent=True)
                        if ip == self.current_address_list[interface_index]: # někdy dříve než 7 sekund...
                            break
                        time.sleep(1)

                    self.option_change("",silent=True)
                    if ip == self.current_address_list[interface_index]:
                        Tools.add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla přenastavena na: {ip}","green",None,True)
                        self.refresh_ip_statuses()
                    else:
                        Tools.add_colored_line(self.main_console,f"Chyba, neplatná adresa nebo daný inteface odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťte aplikaci jako administrátor)","red",None,True)
                        open_app_as_admin_prompt()
                except Exception:
                    pass
            
            run_background = threading.Thread(target=call_subprocess,)
            run_background.start()

        def change_to_DHCP(self):
            def delay_the_refresh():
                nonlocal previous_addr
                nonlocal interface_index
                nonlocal interface
                new_addr = self.current_address_list[interface_index]
                i = 0
                while new_addr == previous_addr or new_addr == None:
                    Tools.add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                    time.sleep(1)
                    self.option_change("",silent=True)
                    new_addr = self.current_address_list[interface_index]
                    print("current addr: ",new_addr)
                    i+=1
                    if i > 6:
                        Tools.add_colored_line(self.main_console,f"Chyba, u {interface} se nepodařilo změnit ip adresu (pro nastavování odpojených interfaců spusťte aplikaci jako administrátor)","red",None,True)
                        return
                
                Tools.add_colored_line(self.main_console,f"IPv4 adresa interfacu: {interface} úspěšně přenastavena na DHCP (automatickou)","green",None,True)
                self.refresh_ip_statuses()
                return
            
            interface = str(self.interface_drop_options.get())
            if not main.IP_tools.check_DHCP(interface):
                if interface != None or interface != "":
                    interface_index = self.connection_option_list.index(interface)
                    previous_addr = self.current_address_list[interface_index]
                    print("previous addr: ",previous_addr)
                    try:
                        # Construct the netsh command
                        netsh_command = f"netsh interface ipv4 set address name=\"{interface}\" source=dhcp"
                        print(f"calling: {netsh_command}")
                        powershell_command = [
                            'powershell.exe',
                            '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                        ]
                        process = subprocess.Popen(powershell_command,
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
                        
                        stdout, stderr = process.communicate()
                        stdout_str = stdout.decode('utf-8')
                        stderr_str = stderr.decode('utf-8')
                        if stderr_str:
                            print(f"Error occurred: {stderr_str}")
                        else:
                            print(f"Command executed successfully:\n{stdout_str}")                
                        
                        run_background = threading.Thread(target=delay_the_refresh,)
                        run_background.start()

                    except Exception as e:
                        print(f"Exception occurred: {str(e)}")
                else:
                    Tools.add_colored_line(self.main_console,"Nebyl zvolen žádný interface","red",None,True)
            else:
                connected_interfaces = self.refresh_interfaces()
                if interface in connected_interfaces:
                    Tools.add_colored_line(self.main_console,f"{interface} má již nastavenou DHCP","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Chyba, {interface} je odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťte aplikaci jako administrátor)","red",None,True)

        def change_computer_ip(self,project,force_params = []):
            """
            input - force_params = [ip,mask]
            """
            def connected_interface(interface,ip,mask):
                """
                Když jsou vyžadována admin práva, tato funkce ověří, zda není daný interface připojen nebo součástí zařízení a zkusí znovu
                """
                try:
                    # Construct the netsh command
                    netsh_command = f"netsh interface ip set address \"{interface}\" static {ip} {mask}"
                    powershell_command = [
                        'powershell.exe',
                        '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                    ]
                    process = subprocess.Popen(powershell_command,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
                    
                    stdout, stderr = process.communicate()
                    stdout_str = stdout.decode('utf-8')
                    stderr_str = stderr.decode('utf-8')
                    if stderr_str:
                        print(f"Error occurred: {stderr_str}")
                        Tools.add_colored_line(self.main_console,f"Chyba, nebyla poskytnuta práva (dejte ANO)","red",None,True)
                    else:
                        print(f"Command executed successfully:\n{stdout_str}")
                        self.make_sure_ip_changed(interface_name,ip)

                except Exception as e:
                    print(f"Exception occurred: {str(e)}")

            if len(force_params) > 0:
                ip = force_params[0]
                mask = force_params[1]
            else:
                ip = str(project["ip"])
                mask = str(project["mask"])

            # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
            interface_name = str(self.interface_drop_options.get())
            powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
            try:
                process = subprocess.Popen(['powershell.exe', '-Command', powershell_command],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                stdout, stderr =process.communicate()
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                print(f"VÝSTUP Z IP SETTING: {str(stdout_str)}")

                if len(str(stdout_str)) > 7:
                    raise subprocess.CalledProcessError(1, powershell_command, stdout_str)
                if stderr_str:
                    raise subprocess.CalledProcessError(1, powershell_command, stderr_str)

                self.make_sure_ip_changed(interface_name,ip)

            except subprocess.CalledProcessError as e:
                if "Run as administrator" in str(stdout_str) or "pustit jako správce" in str(stdout_str):
                    Tools.add_colored_line(self.main_console,f"Chyba, tato funkce musí být spuštěna s administrátorskými právy","red",None,True)
                    # trigger powershell potvrzení:
                    connected_interface(interface_name,ip,mask)
                elif "Invalid address" in str(stdout_str) or "Adresa není platná" in str(stdout_str):
                    Tools.add_colored_line(self.main_console,f"Chyba, neplatná IP adresa","red",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Chyba, Nemáte tuto adresu již nastavenou pro jiný interface? (nebo daný interface na tomto zařízení neexistuje)","red",None,True)
            except Exception as e:
                # Handle any other exceptions that may occur
                Tools.add_colored_line(self.main_console, f"Nastala neočekávaná chyba: {e}", "red", None, True)

        def check_given_input(self,given_data = None,search_flag=False):
            """
            Fills all parameters of last project
            """
            if given_data == None:
                given_data = self.search_input.get()
            if given_data == "":
                found = None
                return found
            found = False

            for i in range(0,len(self.all_project_list)):
                if search_flag:
                    if str(given_data) in str(self.all_project_list[i]['name']):
                        self.last_managed_project = self.all_project_list[i]
                        found = True
                else:
                    if given_data == self.all_project_list[i]['name']:
                        self.last_managed_project = self.all_project_list[i]
                        found = True
            return found    

        def clicked_on_project(self,project,widget,textbox = "",flag = ""):
            """
            flag = notes:
            - při nakliknutí poznámky zůstanou expandnuté a při kliku na jinou je potřeba předchozí vrátit zpět
            flag = unfocus:
            - při kliku mimo se odebere focus z nakliknutých widgetů
            """
            def on_leave_entry(last_selected_textbox):
                """
                při kliku na jiný widget:
                - upraví text pouze na první řádek
                """
                last_selected_textbox.configure(state = "normal")
                if "\n" in self.last_managed_project["notes"]:
                    notes_rows = self.last_managed_project["notes"].split("\n")
                    first_row = notes_rows[0]
                    last_selected_textbox.delete("1.0",tk.END)
                    last_selected_textbox.insert(tk.END,str(first_row))
                    last_selected_textbox.configure(height = 40) #notes
                if self.default_note_behav == 0:
                    last_selected_textbox.configure(state = "disabled")

            if flag == "unfocus":
                try:
                    if str(self.last_selected_notes_widget) != "" and self.last_selected_notes_widget.winfo_exists():
                        if str(self.last_selected_textbox) != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox)
                            self.last_selected_textbox = ""
                            self.last_selected_notes_widget = ""

                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        self.last_selected_widget.configure(border_color="#636363")
                        self.last_selected_widget = ""

                    for frame in self.remember_to_change_back:
                        if frame.winfo_exists(): 
                            frame.configure(border_color="#636363")
                    self.selected_list = []
                    self.remember_to_change_back = []
                    self.last_managed_project = None

                except Exception as e:
                    print("chyba při odebírání focusu: ",e)
                return

            if project == None:
                return
            
            print("clicked project: ",project['name'])
            self.search_input.delete("0","300")
            self.search_input.insert("0",str(project['name']))
            # only if it is not pressed againt the same:
            if widget != self.last_selected_widget:
                try:
                    if str(self.last_selected_textbox) != "" and self.last_selected_textbox.winfo_exists():
                        on_leave_entry(self.last_selected_textbox)
                    else:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""

                    if flag == "notes":
                        self.last_selected_textbox = textbox
                        self.last_selected_notes_widget = widget
                        widget.focus_set()
                    else:
                        # init the values:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""
                        widget.focus_set()
                except Exception as e:
                    print("chyba s navracenim framu do puvodniho formatu: ",e)
                
                try:
                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        if len(self.selected_list) == 0 and not self.control_pressed:
                            self.last_selected_widget.configure(border_color="#636363")
                            if self.last_selected_widget in self.remember_to_change_back:
                                self.remember_to_change_back.pop(self.remember_to_change_back.index(self.last_selected_widget))
                            
                            self.selected_list = []
                            self.remember_to_change_back = []

                        # pokud došlo k další interakci s jiným widgeten
                        elif not self.control_pressed:
                            for frame in self.remember_to_change_back:
                                if frame.winfo_exists(): 
                                    frame.configure(border_color="#636363")
                            self.selected_list = []
                            self.remember_to_change_back = []

                    self.last_selected_widget = widget
                    widget.configure(border_color="white")

                    if not widget in self.remember_to_change_back:
                        self.remember_to_change_back.append(widget)
                    print("remember: ", self.remember_to_change_back)

                except Exception as e:
                    print("chyba pri zmene fucusu",e)
                    pass

                self.last_managed_project = project

        def refresh_ip_statuses(self):
            def unbind_connected_ip(widget,frame):
                widget.unbind("<Enter>")
                frame.unbind("<Enter>")
                widget.unbind("<Leave>")
                frame.unbind("<Leave>")
                frame.configure(fg_color = "black")

            def on_enter(e,interface,widget):
                widget.configure(text = interface)   

            def on_leave(e,ip,widget,frame):
                widget.configure(text = ip)
                if ip not in self.current_address_list:
                    unbind_connected_ip(widget,frame)

            for i in range(0,len(self.ip_frame_list)):
                ip_addr = self.all_project_list[i]["ip"]
                ip_frame = self.ip_frame_list[i][0]
                parameter = self.ip_frame_list[i][1]
                if ip_addr in self.current_address_list:
                    ip_frame.   configure(fg_color = "green") 
                    ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                    ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
                    parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                    parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
                else:
                    unbind_connected_ip(parameter,ip_frame)

        def show_context_menu(self,event,project,flag=""):
            context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="#202020",activebackground="#606060")
            self.last_managed_project = project
            delete_label = "Odstranit"
            if self.show_favourite:
                delete_label = "Odebrat z oblíbených"
            
            if flag == "button":
                context_menu.add_command(label="Nastavit",font=("Arial",22),command=lambda: self.change_computer_ip(project))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat IP adresu",font=("Arial",22), command=lambda: pyperclip.copy(str(project["ip"])))
                context_menu.add_separator()
                context_menu.add_command(label="Editovat",font=("Arial",22),command=lambda: self.add_new_project(True,project))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat projekt",font=("Arial",22),command=lambda: self.add_new_project(init_copy=True))
                context_menu.add_separator()
                context_menu.add_command(label="Přesunout na začátek",font=("Arial",22),command=lambda: self.make_project_first(input_entry_bypass=str(project['name'])))
                context_menu.add_separator()
                if not self.show_favourite:
                    if project["fav_status"] == "0":
                        context_menu.add_command(label="Přidat do oblíbených",font=("Arial",22),command=lambda: self.switch_fav_status_new(project,"1",refresh = True))
                        context_menu.add_separator()
                    elif project["fav_status"] == "1":
                        context_menu.add_command(label="Odebrat z oblíbených",font=("Arial",22),command=lambda: self.switch_fav_status_new(project,"0",refresh = True))
                        context_menu.add_separator()
                context_menu.add_command(label=delete_label,font=("Arial",22),command=lambda: self.delete_project(flag="context_menu"))

            elif flag == "ip_frame":
                context_menu.add_command(label="Kopírovat IP adresu",font=("Arial",22), command=lambda: pyperclip.copy(str(project["ip"])))


            context_menu.tk_popup(event.x_root, event.y_root)

        def make_project_cells(self,no_read = None):
            Tools.clear_frame(self.project_tree)

            def opened_window_check():
                if self.opened_window == "":
                    return False
                try:
                    if self.opened_window.winfo_exists():
                        return True
                    else:
                        return False
                except Exception as err:
                    print(err)
                    return False
                
            def on_enter(interface,widget):
                widget.configure(text = interface)   

            def on_leave(ip,widget,frame):
                widget.configure(text = ip)
                if ip not in self.current_address_list:
                    unbind_connected_ip(widget,frame)

            def unbind_connected_ip(widget,frame):
                widget.unbind("<Enter>")
                frame.unbind("<Enter>")
                widget.unbind("<Leave>")
                frame.unbind("<Leave>")
                frame.configure(fg_color = "black")

            def filter_text_input(text):
                legit_rows = []
                legit_notes = ""
                rows = text.split("\n")
                for i in range(0,len(rows)):
                    if rows[i].replace(" ","") != "":
                        legit_rows.append(rows[i])

                for i in range(0,len(legit_rows)): 
                    if i == len(legit_rows)-1:
                        legit_notes = legit_notes + legit_rows[i]
                    else:
                        legit_notes = legit_notes + legit_rows[i]+ "\n"
                return legit_notes
            
            def save_changed_notes(notes,project):
                workbook = load_workbook(self.excel_file_path)
                def save_to_workbook(notes,row,excel_worksheet):
                    nonlocal workbook
                    worksheet = workbook[excel_worksheet]
                    worksheet['D' + str(row+1)] = notes

                self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)
                project_index = Tools.get_project_index(self.all_project_list,project['name'])
                save_to_workbook(notes,project_index,"ip_address_list")
                self.all_project_list[project_index]["notes"] = notes
                workbook.save(filename=self.excel_file_path)
                workbook.close()

            def on_enter_entry(widget,project):
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        tolerance = 5
                        if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance: # if the height is not default then it means it is expanded already
                            filtered_input = filter_text_input(project["notes"])
                            project["notes"] = filtered_input
                            addition = self.notes_frame_height
                            if "\n" in project["notes"]:
                                notes_rows = project["notes"].split("\n")
                                if len(notes_rows) > 1:
                                    expanded_dim = addition + (len(notes_rows)-1) * 25
                                    # widget[0].configure(height = expanded_dim)
                                    widget[1].configure(state = "normal")
                                    widget[1].configure(height = expanded_dim-10)
                                    if str(widget[1]) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                                        widget[1].delete("1.0",tk.END)
                                        widget[1].insert(tk.END,str(project["notes"]))

                    if self.default_note_behav == 0:
                        widget[1].configure(state = "disabled")

            def on_leave_entry(widget,project):
                """
                při opuštění widgetu cursorem:
                - upraví text pouze na první řádek
                - uloží změny
                """

                if not opened_window_check():
                    notes_widget = widget[1]
                    notes_before = filter_text_input(str(project["notes"]))
                    notes_after = filter_text_input(str(notes_widget.get("1.0",tk.END)))
                    if str(notes_widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        notes_widget.configure(state = "normal")
                        if notes_before != notes_after:
                            self.changed_notes = [project["notes"],notes_before]
                            self.undo_edit.configure(state = "normal")
                            self.manage_bin(flag="save_edited_ip",project=project,new_edited_name=project['name'])
                            project["notes"] = notes_after
                            save_changed_notes(notes_after,project)

                        if "\n" in project["notes"]:
                            notes_rows = project["notes"].split("\n")
                            first_row = notes_rows[0]
                            notes_widget.delete("1.0",tk.END)
                            notes_widget.insert(tk.END,str(first_row))

                        if self.default_note_behav == 0:
                            notes_widget.configure(state = "disabled")
                        self.root.focus_set() # unfocus widget
                    else:
                        # jinak pouze ulož změny (když je dvakrát nakliknuto to samé)
                        if notes_before != notes_after:
                            self.manage_bin(flag="save_edited_ip",project=project,new_edited_name=project['name'])
                            project["notes"] = notes_after
                            self.changed_notes = [project["notes"],notes_before]
                            self.undo_edit.configure(state = "normal")
                            save_changed_notes(notes_after,project)
                        self.root.focus_set() # unfocus widget

                    tolerance = 5
                    if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance:
                        return
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        widget[1].configure(state = "normal")
                        new_height = self.notes_frame_height
                        # widget[0].configure(height = new_height) #frame
                        widget[1].configure(height = new_height-10) #notes
                        if self.default_note_behav == 0:
                            widget[1].configure(state = "disabled")
                        
            def add_row_return(widget):
                addition = widget[0]._current_height
                expanded_dim = addition + 24
                # widget[0].configure(height = expanded_dim)
                widget[1].configure(height = expanded_dim-10)
           
            if no_read == None:
                self.all_project_list = main.IP_tools.read_excel_data(self.excel_file_path)

            column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column2_header =    customtkinter.CTkLabel(master = column2,text = "IPv4 adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column3_header =    customtkinter.CTkLabel(master = column3,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column2_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column3_header.     pack(padx = (5,0),side = "top",anchor = "w")

            self.ip_frame_list = []
            for projects in self.all_project_list:
                if self.show_favourite and projects["fav_status"] == "0":
                    continue
                btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)# frame s názvem projektu
                button =    customtkinter.CTkButton(master = btn_frame,width = 200,height=40,text = str(projects['name']),font=("Arial",20,"bold"),corner_radius=0,command=lambda widget = btn_frame, project = projects: self.clicked_on_project(project,widget))
                button.     pack(padx =5,pady = 5, fill= "x")
                btn_frame.  pack(side = "top",anchor = "w",expand = False,fill= "x")
                # button.     bind("<Button-1>",lambda e,widget = btn_frame, project = projects: self.clicked_on_project(project,widget))
                button.     bind("<Double-1>",lambda e,project = projects: self.change_computer_ip(project))
                button.     bind("<Button-3>",lambda e, project = projects: self.show_context_menu(e,project,flag="button"))
                if str(projects["fav_status"]) == "1":
                    button.configure(fg_color = "#1E90FF")

                ip_addr = str(projects["ip"])
                ip_frame =  customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2) # frame s ip adresou
                parameter = customtkinter.CTkLabel(master = ip_frame,text = ip_addr,height=40,width = 250,font=("Arial",20,"bold"),justify='left',anchor = "w")
                parameter.  pack(padx = (10,5),pady = 5)
                ip_frame.   pack(side = "top")
                ip_frame.   bind("<Button-1>",lambda e,widget = ip_frame, project = projects: self.clicked_on_project(project,widget))
                parameter.  bind("<Button-1>",lambda e,widget = ip_frame, project = projects: self.clicked_on_project(project,widget))
                parameter.  bind("<Button-3>",lambda e, project = projects: self.show_context_menu(e,project,flag="ip_frame"))

                self.ip_frame_list.append([ip_frame,parameter])
                if ip_addr in self.current_address_list:
                    ip_frame.   configure(fg_color = "green")
                    ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                    ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))
                    parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                    parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))

                notes_frame =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)# frame s poznamkami...
                notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),corner_radius=0,fg_color="black",height=40)
                notes.          pack(padx =5,pady = 5,anchor="w",fill="x")
                notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, project = projects, textbox_widget = notes: self.clicked_on_project(project,widget,textbox_widget,flag="notes"))
                notes.          bind("<Button-1>",lambda e,widget = notes_frame, project = projects, textbox_widget = notes: self.clicked_on_project(project,widget,textbox_widget,flag="notes"))

                project_notes = str(projects['notes'])
                if "\n" in project_notes:
                    notes_rows = project_notes.split("\n")
                    first_row = notes_rows[0]
                    notes.delete("1.0",tk.END)
                    notes.insert(tk.END,str(first_row))
                else:
                    notes.insert(tk.END,project_notes)

                notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],project=projects: on_enter_entry(widget,project))
                notes.bind("<Leave>",lambda e, widget = [notes_frame,notes],project=projects: on_leave_entry(widget,project))
                notes.bind("<Return>",lambda e, widget = [notes_frame,notes]:add_row_return(widget))

                if self.default_note_behav == 0:
                    notes.configure(state = "disabled")

                if projects == self.last_managed_project: # případ že posouvám s projektem nahoru/ dolů/ top (zvíraznit selectnuté)
                    self.selected_list.append(projects)
                    self.last_selected_widget = btn_frame
                    btn_frame.configure(border_color="white")
                    self.remember_to_change_back.append(btn_frame)
                    ip_frame.configure(border_color="white")
                    self.remember_to_change_back.append(ip_frame)
                    notes_frame.configure(border_color="white")
                    self.remember_to_change_back.append(notes_frame)
            
            column1.pack(fill="both",expand=False,side = "left")
            column2.pack(fill="both",expand=False,side = "left")
            column3.pack(fill="both",expand=True, side = "left")
            self.project_tree.update()
            self.project_tree.update_idletasks()
            if len(self.all_project_list) > 0:
                try:
                    self.notes_frame_height = int(notes_frame._current_height)
                except Exception:
                    pass
            try:
                self.project_tree._parent_canvas.yview_moveto(0.0)
            except Exception:
                pass
        
        def edit_project(self):
            result = self.check_given_input()
            if result == True:
                self.add_new_project(True)
            elif result == None:
                Tools.add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            else:
                Tools.add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
            
        def option_change(self,args,only_console = False,silent = False):
            """
            Volá get_current_ip_list(), aktualizuje současně nastavené adresy (self.current_address_list)
            - only console: vypíše do konzole aktuální připojení
            - silent: nevypisuje do konzole
            """
            if not only_console:
                try:
                    self.default_connection_option = self.connection_option_list.index(self.interface_drop_options.get())
                except ValueError as e:
                    print(f"Error: {e}")
                    self.default_connection_option = 0

                #pamatovat si naposledy zvoleny zpusob pripojeni:
                Tools.save_to_json_config("default_ip_interface",int(self.default_connection_option),self.config_filename_path)
                self.current_address_list = main.IP_tools.get_current_ip_list(self.connection_option_list)
                if self.static_label2.winfo_exists():
                    self.static_label2.configure(text=self.current_address_list[self.default_connection_option])

                dhcp_status = main.IP_tools.check_DHCP(self.interface_drop_options.get())
                if dhcp_status == True:
                    # print(self.button_dhcp.cget("fg_color"))
                    self.button_dhcp.configure(fg_color = "green")
                else:
                    self.button_dhcp.configure(fg_color = "#1f538d")
                    # self.button_dhcp.configure(fg_color = "#3a7ebf")

            if not silent:
                # ziskat data o aktualnim pripojeni
                current_connection = main.IP_tools.get_ipv4_addresses()
                message = ""
                for items in current_connection:
                    message = message + items + " "
                if message == "":
                    message = "nenalezeno"
                Tools.add_colored_line(self.main_console,f"Současné připojení: {message}","white",None,True)

        def make_project_first(self,purpouse=None,make_cells = True,project = None, input_entry_bypass = None,upwards=False,downwards=False):
            """
            purpouse:
            - search
            - silent
            """
            def check_position():
                project_index = Tools.get_project_index(self.all_project_list,project['name'])
                prev_pos = project_index
                max_position = len(self.all_project_list)
                if upwards:
                    position = prev_pos -1
                elif downwards:
                    position = prev_pos +1

                if position < 0:
                    position = max_position-1
                elif position > max_position-1:
                    position = 0
                return position
            if purpouse == "search":
                result = self.check_given_input(input_entry_bypass,search_flag=True)
                if result == True:
                    self.search_input.delete("0","300")
                    self.search_input.insert("0",str(self.last_managed_project['name']))
            else:
                result = self.check_given_input(input_entry_bypass)
            self.remember_to_change_back = []
            self.last_selected_widget = ""

            if result == True: #zmena poradi
                if project == None:
                    project = self.last_managed_project

                if downwards or upwards:
                    position = check_position()
                else:
                    position = 0

                if len(self.all_project_list) > 0:
                    project_index = Tools.get_project_index(self.all_project_list,project['name'])
                    self.all_project_list.pop(project_index)

                self.all_project_list.insert(position,project)
                main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)

                if make_cells:
                    self.make_project_cells()
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} nalezen","green",None,True)
                elif purpouse != "silent":
                    Tools.add_colored_line(self.main_console,f"Projekt {project['name']} přesunut na začátek","green",None,True)
            elif result == None and purpouse != "silent":
                print("nevlozeno id")
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            elif purpouse != "silent":
                Tools.add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)
                print("projekt nenalezen")

        def show_favourite_toggle(self,keep_search_input = False,determine_status = None): # hlavni prepinaci tlacitko oblibene/ neoblibene
            if self.show_favourite and (determine_status == None or determine_status == "all"):
                self.show_favourite = False
                window_status = 0
                self.last_selected_widget = ""
                self.last_selected_notes_widget = ""
                self.last_selected_textbox = ""    
                self.ip_frame_list = []
                if keep_search_input == False:
                    self.search_input.delete("0","300")
                    self.search_input.configure(placeholder_text="Název projektu")
                    self.make_project_cells()
                else:
                    self.make_project_cells()
                self.button_remove_main.configure(command = lambda: self.delete_project(flag="main_menu"))
                Tools.save_to_json_config("favorite_ip_window_status",window_status,self.config_filename_path)
                self.button_switch_favourite_ip. configure(fg_color="black")
                self.button_switch_all_ip.       configure(fg_color="#212121")
                self.button_remove_main.         configure(text="Smazat")

            elif self.show_favourite == False and (determine_status == None or determine_status == "fav"):
                # favourite window
                self.show_favourite = True
                window_status = 1
                self.last_selected_widget = ""
                self.last_selected_notes_widget = ""
                self.last_selected_textbox = ""
                self.ip_frame_list = []
                if keep_search_input == False:
                    self.search_input.delete("0","300")
                    self.search_input.configure(placeholder_text="Název projektu")
                    self.make_project_cells()
                else:
                    self.make_project_cells()
                # self.button_remove_main.configure(command = lambda: self.switch_fav_status("with_refresh"))
                Tools.save_to_json_config("favorite_ip_window_status",window_status,self.config_filename_path)
                self.button_switch_favourite_ip. configure(fg_color="#212121")
                self.button_switch_all_ip.       configure(fg_color="black")
                self.button_remove_main.         configure(text="Odebrat")

        def refresh_interfaces(self,all = False):
            """
            - All parametr refreshne i statusy ip adres
            """
            interfaces_data = main.IP_tools.fill_interfaces()
            self.connection_option_list = interfaces_data[0]
            self.interface_drop_options.configure(values = self.connection_option_list)
            online_list_text = ""
            if len(interfaces_data[1]) > 0:
                for data in interfaces_data[1]:
                    online_list_text = online_list_text + str(data) +", "
                online_list_text = online_list_text[:-2] # odebrat čárku s mezerou

            self.online_list.configure(text=online_list_text)
            if all:
                self.option_change("")
            self.refresh_ip_statuses()

            return interfaces_data[1]

        def setting_window(self):
            def save_new_behav_notes():
                nonlocal checkbox
                if int(checkbox.get()) == 0:
                    self.default_note_behav = 0
                    Tools.save_to_json_config("editable_notes",0,self.config_filename_path)
                    self.make_project_cells()

                elif int(checkbox.get()) == 1:
                    self.default_note_behav = 1      
                    Tools.save_to_json_config("editable_notes",1,self.config_filename_path)
                    self.make_project_cells()

            def change_make_first_behav():
                nonlocal checkbox4
                if int(checkbox4.get()) == 0:
                    self.make_edited_project_first = False
                    Tools.save_to_json_config("auto_order_when_edit",0,self.config_filename_path)
                elif int(checkbox4.get()) == 1:
                    self.make_edited_project_first = True
                    Tools.save_to_json_config("auto_order_when_edit",1,self.config_filename_path)
            
            def delete_behav():
                if int(checkbox5.get()) == 0 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 100
                elif int(checkbox5.get()) == 0 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 101
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 110
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 111
                else:
                    return
                Tools.save_to_json_config("ask_to_delete",self.deletion_behav,self.config_filename_path)

            def load_old_config():
                def callback_with_path(path_given,load_all_data_status):
                    try:
                        all_project_list = main.IP_tools.read_excel_data(path_given) # read from another file
                        main.IP_tools.save_excel_data(self.excel_file_path,all_project_list)
                        self.make_project_cells() # make project cells with loaded data (it reads again...)
                        Tools.add_colored_line(self.main_console,"Seznam ip adres ze souboru úspěšně nahrán a uložen","green",None,True)

                        if int(load_all_data_status) == 1:
                            disk_rows = main.DM_tools.read_excel_data(path_given)
                            main.DM_tools.save_excel_data_disk(self.excel_file_path,disk_rows)    
                            Tools.add_colored_line(self.main_console,"Seznam ip adres a disků ze souboru úspěšně nahrán a uložen","green",None,True)
                            
                    except Exception as e:
                        Tools.add_colored_line(self.main_console,f"Nepodařilo se načíst data z externího souboru: {e}","red",None,True)

                Tools.import_option_window(self.root,self.app_icon,self.initial_path,callback_with_path,ip_env = True,setting_window=child_root)
                
            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Nastavení")
            main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "Chování poznámek (editovatelné/ needitovatelné):",font=("Arial",20,"bold"))
            checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "Přímo zapisovat a ukládat do poznámek na úvodní obrazovce",font=("Arial",16,"bold"),command=lambda: save_new_behav_notes())
            label.          pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox.       pack(pady = 10,padx=10,side="top",anchor = "w")
            
            main_frame4 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label4 =        customtkinter.CTkLabel(master = main_frame4, width = 100,height=40,text = "Nastavení chování při editaci projektů:",font=("Arial",20,"bold"))
            checkbox4 =     customtkinter.CTkCheckBox(master = main_frame4, text = "Automaticky přesouvat editovaný projekt na začátek",font=("Arial",16,"bold"),command=lambda: change_make_first_behav())
            label4.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox4.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame5 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label5 =        customtkinter.CTkLabel(master = main_frame5, width = 100,height=40,text = "Odvolit dotazování při mazání:",font=("Arial",20,"bold"))
            checkbox5 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit v hlavním okně",font=("Arial",16,"bold"),command=lambda: delete_behav())
            checkbox6 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit při editu",font=("Arial",16,"bold"),command=lambda: delete_behav())
            label5.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox5.      pack(pady = 0,padx=10,side="top",anchor = "w")
            checkbox6.      pack(pady = (5,5),padx=10,side="top",anchor = "w")

            load_config_frame = customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            load_config_label = customtkinter.CTkLabel(master = load_config_frame, width = 100,height=40,text = "Načíst seznam adres (z jiného konfiguračního souboru)",font=("Arial",20,"bold"))
            config_btn_frame = customtkinter.CTkFrame(master=load_config_frame,corner_radius=0,fg_color="#212121")
            button_load =       customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Zvolit soubor",command = lambda:load_old_config(),font=("Arial",20,"bold"),corner_radius=0)
            button_open =       customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Otevřít aktuální",command = lambda: os.startfile(self.excel_file_path),font=("Arial",20,"bold"),corner_radius=0)
            open_path =         customtkinter.CTkButton(master = config_btn_frame, width = 150,height=40,text = "Otevřít složku",command = lambda: os.startfile(self.initial_path),font=("Arial",20,"bold"),corner_radius=0)
            load_config_label.  pack(pady = (10,0),padx=10,side="top",anchor = "w")
            button_load.        pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            button_open.        pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            open_path.        pack(pady = (5,10),padx=(10,0),side="left",anchor = "w")
            config_btn_frame.   pack(pady = 2,padx=2,side="top",fill="x",anchor = "w")

            close_frame =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
            button_close =  customtkinter.CTkButton(master = close_frame, width = 150,height=40,text = "Zavřít",command = child_root.destroy,font=("Arial",20,"bold"),corner_radius=0)
            button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame4.    pack(expand=False,fill="x",side="top")
            main_frame5.    pack(expand=False,fill="x",side="top")
            load_config_frame.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")

            if self.default_note_behav == 1:
                checkbox.select()
            if self.make_edited_project_first:
                checkbox4.select()
            if self.deletion_behav == 110 or self.deletion_behav == 111:
                checkbox5.select()
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                checkbox6.select()

            self.opened_window = child_root
            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def manual_ip_setting(self):
            window = customtkinter.CTkToplevel()
            window.after(200, lambda: window.iconbitmap(Tools.resource_path(self.app_icon)))
            self.opened_window = window
            window.title("Manuální nastavení IPv4 adresy")

            def check_ip_and_mask(input_value):
                input_splitted = input_value.split(".")
                if len(input_splitted) == 4:
                    return input_value
                else:
                    return False

            def call_ip_change():
                if "DHCP" in str(select_mode.get()):
                    self.change_to_DHCP()
                    window.destroy()
                    return
                
                ip_input = ip_address_entry.get()
                mask_input = mask_entry.get()
                ip_checked = check_ip_and_mask(ip_input)
                mask_checked = check_ip_and_mask(mask_input)
                errors = 0
                if ip_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná IP adresa","red",None,True)
                    errors += 1
                if mask_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná maska","red",None,True)
                    errors += 1

                if errors == 0:
                    self.change_computer_ip(0,force_params=[ip_input,mask_input])
                    window.destroy()

            def call_option_change(*args):
                nonlocal ip_address_entry
                self.interface_drop_options.set(str(*args))
                self.option_change(*args)
                ip_address_entry.delete(0,300)
                ip_address_entry.insert(0,self.current_address_list[self.default_connection_option])
                check_interface_status()

            def switch_manual_dhcp(*args):
                nonlocal ip_address_entry
                nonlocal mask_entry
                if "DHCP" in str(*args):
                    ip_address_entry.configure(state = "disabled",text_color = "gray32")
                    mask_entry.configure(state = "disabled",text_color = "gray32")
                else:
                    ip_address_entry.configure(state = "normal",text_color = "gray84")
                    mask_entry.configure(state = "normal",text_color = "gray84")

            def check_interface_status(online_list = False):
                if online_list == False:
                    online_list = self.refresh_interfaces()

                found = False
                for items in online_list:
                    if items == str(select_interface.get()):
                        found = True
                        select_interface.configure(fg_color = "green",button_color = "green")
                        interface_status.configure(text = "Online")
                        break

                if not found:
                    select_interface.configure(fg_color = "red",button_color = "red")
                    interface_status.configure(text = "Offline")

            interface_label =       customtkinter.CTkLabel(master = window,text = "Manuálně nastavit IPv4 adresu pro: ",font=("Arial",20,"bold"))
            interface_frame =       customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            select_interface =      customtkinter.CTkOptionMenu(master = interface_frame,width=320,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,command= lambda args:  call_option_change(args))
            interface_status =      customtkinter.CTkLabel(master = interface_frame,text = "",font=("Arial",20,"bold"))
            select_interface.       pack(pady=(10,0),padx=10,side = "left",anchor = "w")
            interface_status.       pack(pady=(10,0),padx=10,side = "left",anchor = "w")
            mode_label =            customtkinter.CTkLabel(master = window,text = "Způsob nastavení: ",font=("Arial",20,"bold"))
            select_mode =           customtkinter.CTkOptionMenu(master = window,width=400,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,values = ["manuálně","automaticky (DHCP)"],command= lambda args: switch_manual_dhcp(args))
            ip_address =            customtkinter.CTkLabel(master = window,text = "IPv4 adresa: ",font=("Arial",20,"bold"))
            ip_address_entry =      customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            mask =                  customtkinter.CTkLabel(master = window,text = "IPv4 maska: ",font=("Arial",20,"bold"))
            mask_entry =            customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            manual_console =        tk.Text(window, wrap="none", height=0, width=36,background="black",font=("Arial",14),state=tk.DISABLED)
            buttons_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            save_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Nastavit", command = lambda: call_ip_change(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Zrušit", command = lambda: window.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            interface_label.        pack(pady=(10,0),padx=10,side = "top",anchor = "w",expand = False)
            interface_frame.        pack(pady=(0),padx=0,side = "top",anchor = "w")
            mode_label.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            select_mode.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            ip_address.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            ip_address_entry.       pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            mask.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            mask_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            manual_console.         pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            exit_button.            pack(pady=0,padx=(10,0),side = "right",anchor = "e")
            save_button.            pack(pady=0,padx=0,side = "right",anchor = "e")
            buttons_frame.          pack(pady=(10),padx=10,side = "bottom",anchor = "e")
            online_list = self.refresh_interfaces()
            select_interface.configure(values = self.connection_option_list)
            select_interface.set(self.interface_drop_options.get())
            ip_address_entry.insert(0,self.current_address_list[self.default_connection_option])
            mask_entry.insert(0,"255.255.255.0")
            check_interface_status(online_list)
            
            self.root.bind("<Button-1>",lambda e: window.destroy(),"+")
            window.update()
            window.update_idletasks()
            window.focus_force()
            window.focus()

        def sort_by_alphabet(self):
            project_names_array=[]
            for projects in self.all_project_list:
                project_names_array.append(projects['name'])
            project_names_sorted = sorted(project_names_array)
            whole_projects_sorted = []
            for names in project_names_sorted:
                for projects in self.all_project_list:
                    if projects['name'] == names:
                        whole_projects_sorted.append(projects)
                        break
            
            self.all_project_list = copy.deepcopy(whole_projects_sorted)
            main.IP_tools.save_excel_data(self.excel_file_path,self.all_project_list)
            self.make_project_cells()
            Tools.add_colored_line(self.main_console,f"Projekty úspěsně seřazeny podle abecedy","green",None,True)

        def create_widgets(self,fav_status = None,init=None,excel_load_error = False):
            if not excel_load_error:
                if init:
                    if self.window_mode == "max":
                        Tools.save_to_json_config("default_window_size",1,self.config_filename_path)
                    else:
                        Tools.save_to_json_config("default_window_size",0,self.config_filename_path)
                if fav_status:
                    self.show_favourite = True
                    Tools.save_to_json_config("favorite_ip_window_status",1,self.config_filename_path)
                if fav_status == False:
                    self.show_favourite = False
                    Tools.save_to_json_config("favorite_ip_window_status",0,self.config_filename_path)
                Tools.save_to_json_config("disk_or_ip_window",0,self.config_filename_path)
            
            Tools.clear_frame(self.root)
            self.control_pressed = False
            top_frame =                         customtkinter.CTkFrame(master=self.root,corner_radius=0,border_width=0,fg_color="#212121")
            top_left_frame =                    customtkinter.CTkFrame(master=top_frame,corner_radius=0,border_width=0,fg_color="#212121")
            top_right_frame =                   customtkinter.CTkFrame(master=top_frame,corner_radius=0,border_width=0,fg_color="#212121")
            menu_cards =                        customtkinter.CTkFrame(master=top_left_frame,corner_radius=0,fg_color="#636363",height=50,border_width=0)
            self.main_widgets =                 customtkinter.CTkFrame(master=top_left_frame,corner_radius=0,border_width=0)
            self.project_tree =                 customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0,border_width=0)
            logo =                              customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
            image_logo =                        customtkinter.CTkLabel(master = top_right_frame,text = "",image =logo,bg_color="#212121")
            main_menu_button =                  customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            self.button_switch_all_ip =         customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: self.show_favourite_toggle(determine_status="all"),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
            self.button_switch_favourite_ip =   customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: self.show_favourite_toggle(determine_status="fav"),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_disk =                customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",command =  lambda: main.Disk_management_gui(parent=self.parent_instance),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            if excel_load_error:
                self.connection_option_list = ["data nenalezena"]
                self.show_favourite = False
                self.button_switch_all_ip.configure(state = "disabled")
                self.button_switch_favourite_ip.configure(state = "disabled")
                button_switch_disk.configure(state = "disabled")

            first_row_frame =           customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,border_width=0,fg_color="#212121")
            project_label =             customtkinter.CTkLabel(master = first_row_frame, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.search_input =         customtkinter.CTkEntry(master = first_row_frame,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
            # button_search =             customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first("search"),font=("Arial",20,"bold"),corner_radius=0)
            search_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/SearchWhite.png")),size=(32, 32)),bg_color="#212121")
            search_icon.bind("<Enter>",lambda e: search_icon._image.configure(size=(36,36)))
            search_icon.bind("<Leave>",lambda e: search_icon._image.configure(size=(32,32)))
            search_icon.bind("<Button-1>",lambda e: self.make_project_first("search"))
            # self.button_add_main =      customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project(),font=("Arial",20,"bold"),corner_radius=0)
            
            new_project_icon =               customtkinter.CTkLabel(master = first_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/green_plus.png")),size=(32, 32)),bg_color="#212121")
            new_project_icon.bind("<Enter>",lambda e: new_project_icon._image.configure(size=(36,36)))
            new_project_icon.bind("<Leave>",lambda e: new_project_icon._image.configure(size=(32,32)))
            new_project_icon.bind("<Button-1>",lambda e: self.add_new_project())
            self.button_remove_main =   customtkinter.CTkButton(master = first_row_frame, width = 100,height=40,text = "Smazat", command =  lambda: self.delete_project(flag="main_menu"),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_button =          customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_deleted_ip"),font=(None,28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_edit_main =          customtkinter.CTkButton(master = first_row_frame, width = 110,height=40,text = "Editovat",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_edit =            customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_edited_ip"),font=(None,28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_make_first =         customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "🔝",command =  lambda: self.make_project_first(),font=(None,30),corner_radius=0)
            move_upwards =              customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↑",command =  lambda: self.make_project_first(purpouse="silent",upwards=True),font=(None,25),corner_radius=0)
            move_downwards =            customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↓",command =  lambda: self.make_project_first(purpouse="silent",downwards=True),font=(None,25),corner_radius=0)
            sort_alphabet =             customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "A↑",command =  lambda: self.sort_by_alphabet(),font=(None,25),corner_radius=0)
            button_settings_behav =     customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(),font=(None,22),corner_radius=0)
            manual_ip_set =             customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="Manuálně",command =  lambda: self.manual_ip_setting(),font=("Arial",20,"bold"),corner_radius=0)
            
            if self.show_favourite:
                self.button_remove_main.            configure(text="Odebrat")
                self.button_switch_favourite_ip.    configure(fg_color="#212121")
                self.button_switch_all_ip.          configure(fg_color="black")
                self.undo_button.                   configure(state = "disabled")
            else:
                self.button_remove_main.            configure(text="Smazat")
                self.button_switch_favourite_ip.    configure(fg_color="black")
                self.button_switch_all_ip.          configure(fg_color="#212121")
                # poznámky mohou být None (delete undo)
                if len(self.deleted_projects_bin) > 0:
                    self.undo_button.configure(state = "normal")
                else:
                    self.undo_button.configure(state = "disabled")

            # edit undo
            config_data = Tools.read_json_config(self.config_filename_path)
            if len(config_data["edited_project_bin"])>0:
                self.undo_edit.configure(state = "normal")
            else:
                self.undo_edit.configure(state = "disabled")
            if len(config_data["deleted_project_bin"])>0:
                self.undo_button.configure(state = "normal")
            else:
                self.undo_button.configure(state = "disabled")

            second_row_frame =              customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,border_width=0,fg_color="#212121")
            connect_label =                 customtkinter.CTkLabel(master = second_row_frame, width = 100,height=40,text = "Připojení: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.interface_drop_options =   customtkinter.CTkOptionMenu(master = second_row_frame,width=200,height=40,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,command=  self.option_change)
            # "⚙️", "⚒", "🔧", "🔩"
            # button_settings =               customtkinter.CTkButton(master = second_row_frame, width = 40,height=40,text="⚒",command =  lambda: self.refresh_interfaces(all=True),font=("",22),corner_radius=0) #refresh interface statusů
            refresh_icon =               customtkinter.CTkLabel(master = second_row_frame,text = "",image =customtkinter.CTkImage(Image.open(Tools.resource_path("images/refresh.png")),size=(32, 32)),bg_color="#212121")
            refresh_icon.bind("<Enter>",lambda e: refresh_icon._image.configure(size=(36,36)))
            refresh_icon.bind("<Leave>",lambda e: refresh_icon._image.configure(size=(32,32)))
            refresh_icon.bind("<Button-1>",lambda e: self.refresh_interfaces(all=True))
            self.button_dhcp =              customtkinter.CTkButton(master = second_row_frame, width = 100,height=40,text = "DHCP",command =  lambda: self.change_to_DHCP(),font=("Arial",20,"bold"),corner_radius=0)
            static_label =                  customtkinter.CTkLabel(master = second_row_frame, height=40,text = "Static:",font=("Arial",20,"bold"))
            self.static_label2 =            customtkinter.CTkLabel(master = second_row_frame,width=200, height=40,text = "",font=("Arial",22,"bold"),bg_color="black")
            online_label =                  customtkinter.CTkLabel(master = second_row_frame, height=40,text = "Online: ",font=("Arial",22,"bold"))
            self.online_list =              customtkinter.CTkLabel(master = second_row_frame, height=40,text = "",font=("Arial",22,"bold"))
            third_row_frame =               customtkinter.CTkFrame(master=self.root,corner_radius=0,border_width=0,fg_color="#212121")
            self.main_console =             tk.Text(third_row_frame, wrap="none", height=0,background="black",font=("Arial",22),state=tk.DISABLED)
            main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            self.button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            self.button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            image_logo.                     pack(anchor = "e",side = "top",ipadx = 20,ipady = 20,expand=False)
            menu_cards.                     pack(pady=0,padx=5,fill="x",expand=False,side = "top")
            project_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.search_input.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            # button_search.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            search_icon.                    pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            # self.button_add_main.           pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            new_project_icon.               pack(pady = (10,0),padx =(10,0),anchor="w",side="left")
            self.button_remove_main.        pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_button.               pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_edit_main.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_edit.                 pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_make_first.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_upwards.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_downwards.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            sort_alphabet.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_settings_behav.          pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            manual_ip_set.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            connect_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.interface_drop_options.    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            # button_settings.                pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            refresh_icon.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.button_dhcp.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            static_label.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.static_label2.             pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            online_label.                   pack(pady = (10,0),padx =(20,0),anchor="w",side="left")
            self.online_list.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.main_console.              pack(pady = (10,0),padx =5,anchor="w",side="left",fill="x",expand=True)
            first_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            second_row_frame.               pack(pady=0,padx=0,fill="x",side = "top")
            self.main_widgets.              pack(pady=0,padx=0,fill="x",side = "top")
            top_left_frame.                 pack(pady=0,padx=0,fill="x",side = "left",expand=True)
            top_right_frame.                pack(pady=0,padx=0,fill="y",side = "right",expand=False)
            top_frame.                      pack(pady=0,padx=0,fill="x",side = "top")
            third_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            self.project_tree.              pack(pady=(0,5),padx=5,fill="both",expand=True,side = "top")

            main.ToolTip(new_project_icon," Nový projekt ",self.root)
            main.ToolTip(search_icon," Vyhledat projekt ",self.root)
            main.ToolTip(refresh_icon," Refresh připojení ",self.root)
            main.ToolTip(button_make_first," Přesunout projekt na začátek ",self.root)
            main.ToolTip(self.undo_button," Vrátit poslední smazaný projekt ",self.root)
            main.ToolTip(self.undo_edit," Vrátit poslední změnu ",self.root)
            main.ToolTip(move_upwards," Posunout o pozici výše ",self.root)
            main.ToolTip(move_downwards," Posunout o pozici níže ",self.root)
            main.ToolTip(sort_alphabet," Seřadit podle abecedy ",self.root)
            main.ToolTip(button_settings_behav," Nastavení ",self.root)
            main.ToolTip(manual_ip_set," Manuální nastavení adresy ",self.root)

            self.refresh_interfaces() # aktualizace hodnot nabídky
            if self.default_connection_option < len(self.connection_option_list):
                self.interface_drop_options.set(self.connection_option_list[self.default_connection_option])# nastavení naposledy zvoleného interfacu
            else:
                self.default_connection_option = 0             
                Tools.save_to_json_config("default_ip_interface",0,self.config_filename_path)

                self.interface_drop_options.set(self.connection_option_list[self.default_connection_option])

            if not excel_load_error:
                self.option_change("")
                self.make_project_cells()
                self.current_address_list = main.IP_tools.get_current_ip_list(self.connection_option_list)
                self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
            else:
                only_name = self.excel_file_path.split("/")
                only_name = only_name[len(only_name)-1]
                Tools.add_colored_line(self.main_console,f"Konfigurační soubor: {only_name} nebyl nalezen nebo je otevřený","red",None,True)

            def maximalize_window(e):
                self.root.update_idletasks()
                current_width = int(self.root.winfo_width())
                # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
                if Tools.focused_entry_widget(self.root): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                    return
                if int(current_width) > 1200:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                    Tools.save_to_json_config("default_window_size",2,self.config_filename_path)
                elif int(current_width) ==260:
                    self.root.geometry("1200x900")
                    Tools.save_to_json_config("default_window_size",0,self.config_filename_path)
                else:
                    self.root.state('zoomed')
                    Tools.save_to_json_config("default_window_size",1,self.config_filename_path)

            self.root.bind("<f>",lambda e: maximalize_window(e))

            def unfocus_widget(e):
                self.root.focus_set()
            self.root.bind("<Escape>",unfocus_widget)
            self.search_input.bind("<Return>",unfocus_widget)

            def call_search(e):
                self.make_project_first("search")
            self.search_input.bind("<Return>",call_search)

            def call_unfocus(e):
                widget = str(e.widget)
                if not ".!ctkscrollableframe" in widget and not ".!ctktoplevel" in widget and not ".!ctkbutton" in widget:
                    #odebrat focus
                    self.clicked_on_project(None,None,None,flag="unfocus")
                    return
                
            self.root.bind("<Button-1>",call_unfocus,"+")

            def control_button(status):
                self.control_pressed = status
                if status == True:
                    if self.last_managed_project is None:
                        return
                    if not self.last_managed_project in self.selected_list:
                        self.selected_list.append(self.last_managed_project)

            def multi_select():
                if self.last_managed_project is None:
                    return
                if not self.last_managed_project in self.selected_list:
                    self.selected_list.append(self.last_managed_project)
                    # print("selected_list - ",self.selected_list)

            self.root.bind("<Control_L>",lambda e: control_button(True))
            self.root.bind("<Control-Button-1>",lambda e: multi_select())
            self.root.bind("<KeyRelease-Control_L>",lambda e: control_button(False))
            self.root.bind("<Delete>",lambda e: self.delete_project(flag="main_menu"))
            # self.root.mainloop()

if testing_mode:
    # IP_assignment(root,"","max",str(os.getcwd())+"\\",100)
    print(str(os.getcwd())+"\\")
    main(root,"","max",str(os.getcwd())+"\\",100,"jhv_IP.json")
    
    root.mainloop()