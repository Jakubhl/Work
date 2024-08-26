import customtkinter
import tkinter as tk
from openpyxl import load_workbook
import subprocess
import os
import re
import time
import threading
import psutil
import socket
# import win32api
# import win32file
from PIL import Image
import sys
import ctypes
import winreg
import win32net
import win32netcon

testing_mode = True
if testing_mode:
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    root=customtkinter.CTk()
    root.geometry("1200x900")
    root.title("ip_setting - testing")

def path_check(path_raw,only_repair = None):
    path=path_raw
    backslash = "\\"
    if backslash[0] in path:
        newPath = path.replace(backslash[0], '/')
        path = newPath
    if path.endswith('/') == False:
        newPath = path + "/"
        path = newPath
    #oprava mezery v nazvu
    path = r"{}".format(path)
    if not os.path.exists(path) and only_repair == None:
        return False
    else:
        return path

def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

def add_colored_line(text_widget, text, color,font=None,delete_line = None):
    """
    Vloží řádek do console
    """
    text_widget.configure(state=tk.NORMAL)
    if font == None:
        font = ("Arial",16)
    if delete_line != None:
        text_widget.delete("current linestart","current lineend")
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert("current lineend",text, color)
    else:
        text_widget.tag_configure(color, foreground=color,font=font)
        text_widget.insert(tk.END,"    > "+ text+"\n", color)

    text_widget.configure(state=tk.DISABLED)
    
def check_network_drive_status(drive_path):
    checking_done = False
    status = False
    try:
        # Attempt to access a file or directory on the network drive
        drive_path = drive_path[0:3]
        def call_subprocess():
            nonlocal checking_done
            nonlocal status
            if os.path.exists(drive_path):
                os.listdir(drive_path)
                checking_done = True
                status = True
                return
            else:
                checking_done = True
                status = False
                return
            
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

        time_start = time.time()
        while checking_done==False:
            time.sleep(0.05)
            if time.time() - time_start > 2:
                print("terminated due to runtime error")
                return False
        
        if status == True:
            return True
        else:
            return False

    except FileNotFoundError:
        return False
    except OSError:
        return False

def list_mapped_disks(whole_format=None):
    remote_drives = []
    try:
        reg_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Network')
        for i in range(0, winreg.QueryInfoKey(reg_key)[0]):
            drive_letter = winreg.EnumKey(reg_key, i)
            if whole_format:
                remote_drives.append(drive_letter + ':')
            else:
                remote_drives.append(drive_letter)
        winreg.CloseKey(reg_key)
    except Exception as e:
        print("Exception occurred: ", e)

    print("persistent disks: ",remote_drives)
    non_persistent_drives = list_non_persistent_disks()
    print("non-persistent disks: ",non_persistent_drives)
    for drives in non_persistent_drives:
        if whole_format:
            remote_drives.append(drives)
        else:
            remote_drives.append(drives[:1])
            
    return remote_drives

def list_non_persistent_disks():
    non_persistent_drives = []
    try:
        # Enumerate network connections
        level = 1  # Level 1 provides the 'ui1_flags' information
        connections, _, _ = win32net.NetUseEnum(None, level)
        for i in range(0,len(connections)):
            non_persistent_drives.append(connections[i]["local"]) 
    except Exception as e:
        print("Exception occurred: ", e)
    return non_persistent_drives

class IP_assignment: # Umožňuje měnit statickou IP a mountit disky
    """
    Umožňuje měnit nastavení statických IP adres
    """

    def __init__(self,root,callback_function,window_mode,initial_path):
        self.initial_path = initial_path
        self.window_mode = window_mode
        self.callback = callback_function
        self.root = root
        self.app_icon = 'images\\logo_TRIMAZKON.ico'
        self.rows_taken = 0
        self.all_rows = []
        self.project_list = []
        self.excel_file_path = initial_path + "saved_addresses_2.xlsx"
        self.last_project_name = ""
        self.last_project_ip = ""
        self.last_project_mask = ""
        self.last_project_notes = ""
        self.last_project_id = ""
        self.last_project_disk_letter = ""
        self.last_project_ftp = ""
        self.last_project_username = ""
        self.last_project_password = ""
        self.managing_disk = False
        self.connection_status = None
        self.make_project_favourite = False
        self.favourite_list = []
        self.default_connection_option = 0
        self.connection_option_list = []
        self.default_disk_status_behav = 0
        self.default_note_behav = 0
        self.mapping_condition = 0
        self.last_selected_widget = ""
        self.last_selected_notes_widget = ""
        self.last_selected_textbox = ""
        self.last_selected_widget_id = 0
        self.opened_window = ""
        self.ip_frame_list = []
        self.disk_letter_frame_list = []
        self.make_edited_project_first = True
        def call_main(what:str):
            try:
                if what == "disk":
                    self.create_widgets_disk(init=True)
                else:
                    self.create_widgets(init=True)
            except Exception as e:
                add_colored_line(self.main_console,f"Neočekávaná chyba: {e}","red",None,True)

        def insert_new_excel_param(wb,ws,param):
            """
            Oveřuje zda konfigurační excel již obsahuje tyto parametry, případně zapíše
            param:
            - (disk_behav) default chování načítání obrazovky s disky
            - (notes_behav) default chování poznámek
            - (mapping_cond) disk persistent - yes/ no
            - (make_first_behav) = chování při editu
            """
            if param == "disk_behav":
                ws['B' + str(6)] = 0
                ws['A' + str(6)] = "aktualizovat statusy disků při vstupu do okna s disky (default)"
                wb.save(self.excel_file_path)
                print('inserting new parameter to excel')
            elif param == "notes_behav":
                ws['B' + str(7)] = 0
                ws['A' + str(7)] = "editovatelné(1)/ needitovatelné(0) poznámky (default)"
                wb.save(self.excel_file_path)
                print('inserting new parameter to excel')
            elif param == "mapping_cond":
                ws['B' + str(8)] = 0
                ws['A' + str(8)] = "disk persistentní - yes(1)/ no(0)"
                wb.save(self.excel_file_path)
                print('inserting new parameter to excel')
            elif param == "make_first_behav":
                ws['B' + str(9)] = 1
                ws['A' + str(9)] = "automaticky přesouvat upravené projekty na začátek"
                wb.save(self.excel_file_path)
                print('inserting new parameter to excel')

        try:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["Settings"]
            # z worksheetu nastavení čtu základní zvolený interface připojení
            # - všechny možné intefaces
            # - defaultní okno zobrazení (oblíbené/ všechny/ disky)
            # - defaultní velikost okna - pamatuje si nejmenší zvolenou
            saved_def_con_option = worksheet['B' + str(1)].value
            self.default_connection_option = int(saved_def_con_option)

            def_show_favourite = worksheet['B' + str(3)].value
            if int(def_show_favourite) == 1:
                self.show_favourite = True
            else:
                self.show_favourite = False

            def_window_size = worksheet['B' + str(5)].value
            if def_window_size == 2:
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")
            
            value_check = worksheet['B' + str(6)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,param="disk_behav")
            else:
                self.default_disk_status_behav = int(worksheet['B' + str(6)].value)

            value_check = worksheet['B' + str(7)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,param="notes_behav")
            else:
                self.default_note_behav = int(worksheet['B' + str(7)].value)

            value_check = worksheet['B' + str(8)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,param="mapping_cond")
            else:
                self.mapping_condition = int(worksheet['B' + str(8)].value)

            value_check = worksheet['B' + str(9)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,param="make_first_behav")
            else:
                excel_value =  int(worksheet['B' + str(9)].value)
                if excel_value == 1:
                    self.make_edited_project_first = True
                else:
                    self.make_edited_project_first = False

            def_show_disk = worksheet['B' + str(4)].value
            workbook.close()

            if int(def_show_disk) == 1:
                call_main("disk")
            else:
                call_main("ip")
                
        except Exception as e:
            self.connection_option_list = ["data nenalezena"]
            self.show_favourite = False
            self.create_widgets(init=True,excel_load_error=True)
            print(f"Nejdřív zavřete soubor {self.excel_file_path} Chyba: {e}")

    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
        """
        self.clear_frame(self.root)
        self.callback()

    def clear_frame(self,frame):
        frame.update()
        frame.update_idletasks()
        for widget in frame.winfo_children():
            if widget.winfo_exists():
                widget.unbind("<Enter>")
                widget.unbind("<Leave>")
                widget.unbind("<Return>")
                widget.unbind("<Button-1>")
                widget.unbind("<Button-3>")
                widget.destroy()

    def fill_interfaces(self):
        """
        Vrátí:
        - seznam interfaců [0]
        - seznam připojených interfaců [1]
        """
        process = subprocess.Popen("netsh interface show interface",
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        try:
            stdout_str = stdout.decode('utf-8')
            data = str(stdout_str)
        except UnicodeDecodeError:
            try:
                stdout_str = stdout.decode('cp1250')
                data = str(stdout_str)
            except UnicodeDecodeError:
                data = str(stdout)
            
        lines = data.strip().splitlines()
        interfaces = []
        interface_statuses =[]
        # Process each line starting from the second line
        for line in lines[2:]:  # Skip the first two lines (headers and separator)
            # Split the line based on spaces
            values = line.split()
            if len(values) >= 4:  # Ensure there are enough columns
                # Combine the last columns into Interface Name, handle spaces in the name
                interface_name = ' '.join(values[3:])
                interface_statuses.append(values[1])
                interfaces.append(interface_name)

        print("interface list: ",interfaces)
        print("status: ", interface_statuses)
        connected_interfaces =[]
        for i in range(0,len(interface_statuses)):
            if interface_statuses[i] != "Odpojen" and interface_statuses[i] != "Odpojeno" and interface_statuses[i] != "Disconnected":
                if interfaces[i] not in connected_interfaces:
                    connected_interfaces.append(interfaces[i])
        print("online: ", connected_interfaces)
        
        return [interfaces,connected_interfaces]

    def read_excel_data(self,force_ws = None):
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        if force_ws != None:
            excel_worksheet = force_ws
        workbook = load_workbook(self.excel_file_path)
        # seznam vsech statickych ip adres
        self.all_rows = []
        self.project_list = []
        self.favourite_list = []
        worksheet = workbook[excel_worksheet]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:4]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            if len(row_array) < 4:
                row_array.append("")
            self.project_list.insert(0,row_array[0])
            self.all_rows.insert(0,row_array)
            for items in row[4:5]:
                self.favourite_list.insert(0,items)
            

        # seznam vsech ftp pripojeni k diskum
        self.disk_all_rows = []
        self.disk_project_list = []  
        worksheet = workbook["disk_list"]
        for row in worksheet.iter_rows(values_only=True):
            row_array = []
            for items in row[:6]:
                if items is None:
                    row_array.append("")
                else:
                    row_array.append(str(items))
            """if len(row_array) < 4:
                row_array.append("")"""
            self.disk_project_list.insert(0,row_array[0])
            self.disk_all_rows.insert(0,row_array)

        # ukladani nastavenych hodnot
        worksheet = workbook["Settings"]
        saved_def_con_option = worksheet['B' + str(1)].value
        self.default_connection_option = int(saved_def_con_option)

        self.default_disk_status_behav = int(worksheet['B' + str(6)].value)
        workbook.close()
                     
    def save_excel_data(self,project_name,IP_adress,mask,notes,only_edit = None,force_row_to_print=None,fav_status = None,force_ws = None):
        workbook = load_workbook(self.excel_file_path)
        if self.show_favourite:
            excel_worksheet = "ip_adress_fav_list"
        else:
            excel_worksheet = "ip_address_list"
        if force_ws != None:
            excel_worksheet = force_ws
        worksheet = workbook[excel_worksheet]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.all_rows)) +1
            if only_edit != None:
                #pouze změna na temtýž řádku
                row_to_print = (len(self.all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        if notes.replace(" ","") == "" or notes == None:
            notes = ""
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = ip adresa
        worksheet['B' + str(row_to_print)] = IP_adress
        #C = maska
        worksheet['C' + str(row_to_print)] = mask
        #D = poznamky
        worksheet['D' + str(row_to_print)] = notes
        #E = oblibenost
        if fav_status != None:
            worksheet['E' + str(row_to_print)] = fav_status

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def save_excel_data_disk(self,project_name,disk_letter,ftp_address,username,password,notes,only_edit = None,force_row_to_print=None):
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["disk_list"]
        # excel je od jednicky...
        if force_row_to_print == None:
            row_to_print = int(len(self.disk_all_rows)) +1
            if only_edit != None:
                #pouze změna na temtýž řádku
                row_to_print = (len(self.disk_all_rows)- self.last_project_id)
        else:
            row_to_print = force_row_to_print
        #A = nazev projektu
        worksheet['A' + str(row_to_print)] = project_name
        #B = písmeno disku, označení...
        worksheet['B' + str(row_to_print)] = disk_letter
        #C = ftp adresa
        worksheet['C' + str(row_to_print)] = ftp_address
        #D = uživatelské jméno
        worksheet['D' + str(row_to_print)] = username
        #E = heslo
        worksheet['E' + str(row_to_print)] = password
        #F = poznamky
        worksheet['F' + str(row_to_print)] = notes

        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def get_notes(self):
        notes_legit_rows = []
        notes = str(self.notes_input.get("1.0", tk.END))
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    def check_ip_and_mask(self,input_value):
        input_splitted = input_value.split(".")
        if len(input_splitted) == 4:
            return input_value
        else:
            return False

    def switch_fav_status(self,operation:str,project_given=None,change_status = False):
        if project_given == None:
            selected_project = str(self.search_input.get())
            if selected_project not in self.project_list:
                add_colored_line(self.main_console,"Nebyl vložen projekt",color="red",font=None,delete_line=True)
                return
            else:
                selected_project = self.all_rows[self.project_list.index(selected_project)]
        else:
            selected_project = project_given
        if self.show_favourite == False:
            if operation == "add_favourite":
                if change_status:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=1)
                self.show_favourite = True
                self.read_excel_data()
                # do tohoto prostředí uložím na začátek
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()
            
            elif operation == "del_favourite":
                if change_status:
                    self.save_excel_data(selected_project[0],selected_project[1],selected_project[2],selected_project[3],True,None,fav_status=0)
                # přepnutí
                self.show_favourite = True
                self.read_excel_data()
                # z tohoto prostředí smažu
                self.delete_project(wanted_project=selected_project[0],silence=True)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()

            elif operation == "rewrite_favourite":
                # přepnutí
                self.show_favourite = True
                self.read_excel_data()
                # nejprve popnu stary projekt, s povodnim jmenem
                # poté insertnu pozměněný
                the_id_to_pop = self.project_list.index(self.last_project_name)
                self.all_rows.pop(the_id_to_pop)
                self.all_rows.insert(0,selected_project)
                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=1)
                # přepnutí zpět
                self.show_favourite = False
                self.read_excel_data()

        elif self.show_favourite:
            # z aktuálního prostředí smažu
            self.delete_project(wanted_project=selected_project[0],silence=True)
            # musim prepnout prostředí jen kvůli změně statusu
            self.show_favourite = False
            self.read_excel_data()
            match_found = False
            for i in range(0,len(self.project_list)):
                if self.project_list[i] == selected_project[0] and len(str(self.project_list[i])) == len(str(selected_project[0])):
                    row_index = self.project_list.index(selected_project[0])
                    match_found = True
            if match_found:
                row = len(self.all_rows) - row_index
                self.save_excel_data(self.all_rows[row_index][0],self.all_rows[row_index][1],self.all_rows[row_index][2],self.all_rows[row_index][3],None,row,fav_status=0)

            # přepnutí zpět
            self.show_favourite = True
            self.read_excel_data()
        
        if operation == "with_refresh":
            add_colored_line(self.main_console,f"Projekt: {selected_project[0]} byl odebrán z oblíbených","green",None,True)
            self.make_project_cells(no_read=True)

    def save_new_project_data(self,child_root,only_edit = None,make_fav=False):

        def get_both_row_indexes(new_project = False):
            """
            - new project = bool - returs the last position of excel row, where the new project takes place\n
            returns array of 2 excel row indexes: (finds matches)\n
            [0] = normal list\n
            [1] = favourite list\n
            - if not found returns "no data"\n
            """
            wanted_project = self.last_project_name
            def find_project_index(wanted_project,new_fav_status):
                index_of_project = "no data"
                try:
                    if new_fav_status != None:
                        self.show_favourite = new_fav_status
                        self.read_excel_data()
                    for i in range(0,len(self.all_rows)):
                        if self.all_rows[i][0] == wanted_project:
                            index_of_project = i
                            break
                    if new_project:
                        return len(self.all_rows)
                    elif index_of_project != "no data":
                        return (len(self.all_rows) - index_of_project)
                    else:
                        return index_of_project
                except Exception as err:
                    print(err)
                    return "no data"

            if self.show_favourite:
                index_of_fav_project = find_project_index(wanted_project,new_fav_status = None)
                index_of_project = find_project_index(wanted_project,new_fav_status = False)
                self.show_favourite = True

            elif not self.show_favourite:
                index_of_project = find_project_index(wanted_project,new_fav_status = None)
                index_of_fav_project = find_project_index(wanted_project,new_fav_status = True)
                self.show_favourite = False

            self.read_excel_data()
            return [index_of_project,index_of_fav_project]

        def switch_database():
            if self.show_favourite:
                self.show_favourite = False
                self.read_excel_data()
            else:
                self.show_favourite = True
                self.read_excel_data()

        project_name = str(self.name_input.get())
        IP_adress = str(self.IP_adress_input.get())
        IP_adress = self.check_ip_and_mask(IP_adress)
        mask = str(self.mask_input.get())
        mask = self.check_ip_and_mask(mask)
        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
            errors += 1
        if project_name in self.project_list and only_edit == None:
            add_colored_line(self.console,f"Jméno je již používané","red",None,True)
            errors +=1

        if IP_adress == False and errors == 0:
            add_colored_line(self.console,f"Neplatná IP adresa","red",None,True)
            errors += 1
        if mask == False and errors == 0:
            add_colored_line(self.console,f"Neplatná maska","red",None,True)
            errors += 1
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()
            # pridavam novy projekt 1: rovnou do oblibených, 2:jen do všech
            if only_edit == None: 
                row_index_list = get_both_row_indexes(new_project=True)
                if make_fav:
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[0]+1,fav_status=1,force_ws="ip_address_list")
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[1]+1,fav_status=1,force_ws="ip_adress_fav_list")
                    add_colored_line(self.main_console,f"Přidán nový oblíbený projekt: {project_name}","green",None,True)
                else:
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[0]+1,fav_status=0,force_ws="ip_address_list")
                    add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)

            elif only_edit:
                # kdyz edituji muze mit projekt jiz prideleny status
                current_fav_status = self.is_project_favourite(self.last_project_id)

                if make_fav and current_fav_status == 0:
                    # zaskrtnuto oblibene + nebyl oblibeny  = ZMENA:
                    row_index_list = get_both_row_indexes(new_project=True)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[1]+1,fav_status=1,force_ws="ip_adress_fav_list")
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=1)

                    if self.last_project_name != project_name:
                        status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn a přidán do oblíbených"
                    else:
                        status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn a přidán do oblíbených"
                    add_colored_line(self.main_console,status_text,"green",None,True)

                    edited_project = [project_name,IP_adress,mask,notes,1]
                    if self.make_edited_project_first:
                        self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)
                
                elif make_fav == False and current_fav_status == 1:
                    # neni zaskrtnuto oblibene + je jiz oblibeny = ZMENA
                    row_index_list = get_both_row_indexes()
                    print("odebran z oblibenych", row_index_list)

                    if row_index_list[0] == "no data" or row_index_list[1] == "no data":
                        add_colored_line(self.main_console,f"Chyba synchronizace (oblíbené <-> všechny). Projekt {self.last_project_name} se nepodařilo pozměnit","red",None,True)
                        child_root.destroy()
                        return

                    if self.show_favourite:
                        self.delete_project(wanted_project=self.last_project_name,silence=True)
                        self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[0],fav_status=0,force_ws="ip_address_list")

                    else:
                        # nejprve smazat z oblíbených:
                        workbook = load_workbook(self.excel_file_path)
                        worksheet = workbook["ip_adress_fav_list"]
                        worksheet.delete_rows(row_index_list[1])
                        workbook.save(self.excel_file_path)
                        # poté uložit změnu statusu do všech:
                        self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=None,fav_status=0)

                    if self.last_project_name != project_name:
                        status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn a odebrán z oblíbených"
                    else:
                        status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn a odebrán z oblíbených"
                    add_colored_line(self.main_console,status_text,"green",None,True)

                    edited_project = [project_name,IP_adress,mask,notes,0]
                    if self.make_edited_project_first:
                        if not self.show_favourite:
                            self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)


                elif make_fav and current_fav_status == 1:
                    # zaskrtnuto oblibene + je jiz oblibeny = BEZ ZMENY
                    #nedoslo ke zmene statusu, ale mohlo dojit ke zmene - proto prepsat v oblibenych
                    row_index_list = get_both_row_indexes()
                    print("pozmenen 1",row_index_list)
                    if row_index_list[0] != "no data":
                        self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[0],fav_status=current_fav_status,force_ws="ip_address_list")
                    if row_index_list[1] != "no data":
                        self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[1],fav_status=current_fav_status,force_ws="ip_adress_fav_list")
                    if self.last_project_name != project_name:
                        status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn"
                    else:
                        status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn"
                    add_colored_line(self.main_console,status_text,"green",None,True)

                    edited_project = [project_name,IP_adress,mask,notes,current_fav_status]
                    if self.make_edited_project_first:
                        self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)
                        # promítnout změny i do druhého menu:
                        switch_database()
                        self.make_project_first(purpouse="silent",make_cells=False,project=edited_project,input_entry_bypass=edited_project[0])
                        switch_database()
                    
                elif make_fav == False and current_fav_status == 0:
                    # neni zaskrtnuto oblibene + nebyl oblibeny = BEZ ZMENY
                    row_index_list = get_both_row_indexes()
                    print("pozmenen 2",row_index_list)
                    self.save_excel_data(project_name,IP_adress,mask,notes,only_edit=True,force_row_to_print=row_index_list[0],fav_status=current_fav_status,force_ws="ip_address_list")
                    
                    if self.last_project_name != project_name:
                        status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn"
                    else:
                        status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn"
                    add_colored_line(self.main_console,status_text,"green",None,True)

                    edited_project = [project_name,IP_adress,mask,notes,current_fav_status]
                    if self.make_edited_project_first:
                        self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)
            child_root.destroy()
            self.make_project_cells()
    
    def save_new_project_data_disk(self,child_root,only_edit = None):
        project_name =  str(self.name_input.get())
        disk_letter =   str(self.disk_letter_input.get())
        ftp_address =   str(self.FTP_adress_input.get())
        username =      str(self.username_input.get())
        password =      str(self.password_input.get())

        notes = self.get_notes()
        errors = 0
        if project_name.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
            errors += 1
        if project_name in self.disk_project_list and only_edit == None:
            add_colored_line(self.console,f"Jméno je již používané","red",None,True)
            errors +=1
        elif disk_letter.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste písmeno disku","red",None,True)
            errors += 1
        elif ftp_address.replace(" ","") == "":
            add_colored_line(self.console,f"Nezadali jste adresu","red",None,True)
            errors += 1
        
        # poznamky nejsou povinne
        if errors ==0:
            self.read_excel_data()
            if only_edit == None:
                self.save_excel_data_disk(project_name,disk_letter,ftp_address,username,password,notes)
            else:
                self.save_excel_data_disk(project_name,disk_letter,ftp_address,username,password,notes,True)
            child_root.destroy()
            if only_edit == None:
                self.make_project_cells_disk()
                add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)
            else: #musi byt proveden reset
                self.make_project_cells_disk()
                add_colored_line(self.main_console,f"Projekt: {project_name} úspěšně pozměněn","green",None,True)

    def delete_project(self,wanted_project=None,silence=None,button_trigger = False):
        project_found = False

        def proceed(window = True):
            nonlocal wanted_project
            nonlocal silence
            nonlocal project_found
            remove_favourite_as_well = False
            if wanted_project == None:
                self.read_excel_data()
                wanted_project = str(self.search_input.get())
            workbook = load_workbook(self.excel_file_path)
            if self.show_favourite:
                excel_worksheet = "ip_adress_fav_list"
            else:
                excel_worksheet = "ip_address_list"
            worksheet = workbook[excel_worksheet]

            for i in range(0,len(self.project_list)):
                if self.project_list[i] == wanted_project and len(str(self.project_list[i])) == len(str(wanted_project)) and project_found == False:
                    row_index = self.project_list.index(wanted_project)
                    worksheet.delete_rows(len(self.all_rows)-row_index)
                    workbook.save(self.excel_file_path)
                    project_found = True
                    #pokud ma status oblibenosti, tak vymazat i z oblibenych:
                    if self.favourite_list[row_index] == 1 and self.show_favourite == False:
                        remove_favourite_as_well = True
                        deleted_project = self.all_rows[row_index]
                    break
            
            workbook.close()
            if silence == None:
                if project_found:
                    add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                    self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                elif wanted_project.replace(" ","") == "":
                    add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                else:
                    add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)
            
            if remove_favourite_as_well:
                self.switch_fav_status("del_favourite",deleted_project)

            if window:
                nonlocal child_root
                child_root.grab_release()
                child_root.destroy()

        if not button_trigger:
            proceed(window=False)
            return

        if self.last_project_name.replace(" ","") == "":
            add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            return
        elif wanted_project == None:
            wanted_project = self.last_project_name

        child_root = customtkinter.CTkToplevel()
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"650x130+{x+80}+{y+150}")
        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        child_root.title("Upozornění")
        proceed_label = customtkinter.CTkLabel(master = child_root,text = f"Opravdu si přejete odstranit projekt {self.last_project_name}?",font=("Arial",22,"bold"),justify = "left",anchor="w")
        button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: proceed())
        button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
        proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top")
        button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")

        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
        child_root.update()
        child_root.update_idletasks()
        child_root.grab_set()
        child_root.focus()
        child_root.focus_force()
        child_root.wait_window()
        return project_found

    def delete_project_disk(self,button_trigger = False):
        wanted_project = str(self.search_input.get())

        def proceed(window = True):
            nonlocal wanted_project
            self.read_excel_data()
            project_found = False
            workbook = load_workbook(self.excel_file_path)
            if wanted_project.replace(" ","") != "":
                for i in range(0,len(self.disk_project_list)):
                    if self.disk_project_list[i] == wanted_project and len(str(self.disk_project_list[i])) == len(str(wanted_project)):
                        row_index = self.disk_project_list.index(wanted_project)
                        worksheet = workbook["disk_list"]
                        worksheet.delete_rows(len(self.disk_all_rows)-row_index)
                        workbook.save(self.excel_file_path)
                        project_found = True
                        break

                if project_found:
                    add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)    
                    self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                elif project_found == False:
                    add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)
            else:
                add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            
            workbook.close()
            if window:
                nonlocal child_root
                child_root.grab_release()
                child_root.destroy()

        if not button_trigger:
            proceed(window=False)
            return

        if wanted_project.replace(" ","") == "":
            add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            return

        child_root = customtkinter.CTkToplevel()
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"650x130+{x+80}+{y+150}")
        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        child_root.title("Upozornění")
        proceed_label = customtkinter.CTkLabel(master = child_root,text = f"Opravdu si přejete odstranit projekt {wanted_project}?",font=("Arial",22,"bold"),justify = "left",anchor="w")
        button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: proceed())
        button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
        proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top")
        button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")

        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
        child_root.update()
        child_root.update_idletasks()
        child_root.grab_set()
        child_root.focus()
        child_root.focus_force()

    def copy_previous_project(self,disk=None):
        if self.last_project_name == "":
            add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
        else:
            self.name_input.delete("0","300")
            self.name_input.insert("0",str(self.last_project_name))
            if disk == None:
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0",str(self.last_project_ip))
                self.mask_input.delete("0","300")
                self.mask_input.insert("0",str(self.last_project_mask))
                self.notes_input.delete("1.0",tk.END)
                self.notes_input.insert(tk.END,str(self.last_project_notes))
            else:
                self.disk_letter_input.delete("0","300")
                self.disk_letter_input.insert("0",str(self.last_project_disk_letter))
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0",str(self.last_project_ftp))
                self.username_input.delete("0","300")
                self.username_input.insert("0",str(self.last_project_username))
                self.password_input.delete("0","300")
                self.password_input.insert("0",str(self.last_project_password))
                self.notes_input.delete("1.0",tk.END)
                self.notes_input.insert(tk.END,str(self.last_project_notes))

    def make_favourite_toggle_via_edit(self,e):
        def do_favourite():
            self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
            self.make_fav_label.configure(text = "Oblíbený ❤️")
        
        def unfavourite():
            self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")
            self.make_fav_label.configure(text = "Neoblíbený")

        if self.make_project_favourite:
            self.make_project_favourite = False
            unfavourite()
        else:
            self.make_project_favourite = True
            do_favourite()

    def add_new_project(self,edit = None):
        def mouse_wheel_change(e):
            if -e.delta < 0:
                switch_up()
            else:
                switch_down()

        def switch_up():
            print("up ",self.last_project_id)
            self.last_project_id -= 1
            if self.last_project_id < 0:
                self.last_project_id = len(self.all_rows)-1
                
            self.check_given_input(given_data=self.all_rows[self.last_project_id][0])
            self.copy_previous_project()
            refresh_favourite_status()

        def switch_down():
            print("down ",self.last_project_id)
            self.last_project_id += 1
            if self.last_project_id > len(self.all_rows)-1:
                self.last_project_id = 0

            self.check_given_input(given_data=self.all_rows[self.last_project_id][0])
            self.copy_previous_project()
            refresh_favourite_status()

        def del_project():
            nonlocal child_root
            result = self.delete_project(button_trigger=True)
            print(result)
            if result:
                print("deleting ", self.all_rows[self.last_project_id][0])
                switch_up()
            else:
                print("aborted")

            child_root.focus()
            child_root.focus_force()
            child_root.grab_set()

        def refresh_favourite_status():
            if self.is_project_favourite(self.last_project_id):
                self.make_project_favourite = True #init hodnota
                self.make_fav_label.configure(text = "Oblíbený ❤️",font=("Arial",22))
                self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
            else:
                self.make_project_favourite = False #init hodnota
                self.make_fav_label.configure(text = "Neoblíbený",font=("Arial",22))
                self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")

        child_root = customtkinter.CTkToplevel()
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"520x750+{x+50}+{y+80}")
        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        if edit:
            child_root.title("Editovat projekt: "+self.last_project_name)
        else:
            child_root.title("Nový projekt")
        
        project_name =    customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
        self.name_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        project_selection_label = customtkinter.CTkLabel(master = child_root, width = 200,height=30,text = "Přepnout projekt: ",font=("Arial",20,"bold"))
        project_switch_frame =  customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=140,width=80)
        project_up =            customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↑",command= lambda: switch_up())
        project_down =          customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↓",command= lambda: switch_down())
        project_switch_frame.   grid_propagate(0)
        project_up              .grid(column = 0,row=0,pady = (5,0),padx =10)
        project_down            .grid(column = 0,row=1,pady = 5,padx =10)
        project_switch_frame.   bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
        project_up.             bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
        project_down.           bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
        IP_adress =            customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
        self.IP_adress_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        mask =                 customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
        self.mask_input =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        copy_check =           customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: self.copy_previous_project())
        del_project_btn =      customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Smazat tento projekt",command= lambda: del_project(),fg_color="red")
        fav_status =           customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Status oblíbenosti: ",font=("Arial",20,"bold"))
        fav_frame =            customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=50,width=200,fg_color="#353535")
        self.make_fav_label =  customtkinter.CTkLabel(master = fav_frame, width = 20,height=30)
        self.make_fav_btn =    customtkinter.CTkLabel(master = fav_frame, width = 50,height=50)
        refresh_favourite_status()
        notes =                customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        self.notes_input =     customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=280)
        self.console =         tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,True,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,None,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
        exit_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)

        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        if edit:
            project_selection_label.grid(column = 0,row=0,padx=265,sticky = tk.W)
            project_switch_frame.   grid(row=1,column=0,padx=320,sticky=tk.W,rowspan=4)
        else:
            copy_check.             grid(column = 0,row=0,pady = 5,padx =240,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        IP_adress.              grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.IP_adress_input.   grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        mask.                   grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        self.mask_input.        grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        fav_status.             grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        if edit:
            del_project_btn.grid(column = 0,row=6,pady = 5,padx =240,sticky = tk.W)
        fav_frame.              grid(column = 0,row=7,padx= 10,sticky=tk.W)
        fav_frame.              grid_propagate(0)
        self.make_fav_btn.      grid(column=0,row=0,pady = 0,padx =0,sticky = tk.W)
        self.make_fav_btn.      bind("<Button-1>",lambda e: self.make_favourite_toggle_via_edit(e))
        self.make_fav_label.    grid(column = 0,row=0,pady = 0,padx =60,sticky = tk.W)
        self.make_fav_label.    bind("<Button-1>",lambda e: self.make_favourite_toggle_via_edit(e))
        notes.                  grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=11,pady = 5,padx =100,sticky = tk.W)
        exit_button.            grid(column = 0,row=11,pady = 5,padx =310,sticky = tk.W)

        if edit:
            self.copy_previous_project()
        else:
            self.IP_adress_input.delete("0","300")
            self.IP_adress_input.insert("0","192.168.000.000")
            self.mask_input.delete("0","300")
            self.mask_input.insert("0","255.255.255.0")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))

        child_root.update()
        child_root.update_idletasks()
        child_root.focus()
        child_root.focus_force()
        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

    def add_new_project_disk(self,edit = None):
        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"520x800+{x+50}+{y+100}")
        if edit == None:
            child_root.title("Nový projekt")
        else:
            child_root.title("Editovat projekt: "+self.last_project_name)

        project_name =              customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
        copy_check =                customtkinter.CTkButton(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: self.copy_previous_project(True))
        self.name_input =           customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        disk_letter =               customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Písmeno disku: ",font=("Arial",20,"bold"))
        self.disk_letter_input =    customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        FTP_adress =                customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
        self.FTP_adress_input =     customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=500,height=30,corner_radius=0)
        user =                      customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Uživatelské jméno: ",font=("Arial",20,"bold"))
        self.username_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        password =                  customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Heslo: ",font=("Arial",20,"bold"))
        self.password_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
        notes =                     customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
        self.notes_input =          customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=260)
        self.console =              tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
        if edit == None:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disk(child_root),font=("Arial",20,"bold"),corner_radius=0)
        else:
            save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disk(child_root,True),font=("Arial",20,"bold"),corner_radius=0)
        exit_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)

        project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        if edit != True:
            copy_check.             grid(column = 0,row=0,pady = 5,padx =230,sticky = tk.W)
        self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        disk_letter.            grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.disk_letter_input. grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        FTP_adress.             grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
        self.FTP_adress_input.  grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
        user.                   grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
        self.username_input.    grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
        password.               grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
        self.password_input.    grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
        notes.                  grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
        self.notes_input.       grid(column = 0,row=11,pady = 5,padx =10,sticky = tk.W)
        self.console.           grid(column = 0,row=12,pady = 5,padx =10,sticky = tk.W)
        save_button.            grid(column = 0,row=13,pady = 5,padx =100,sticky = tk.W)
        exit_button.            grid(column = 0,row=13,pady = 5,padx =310,sticky = tk.W)

        if edit == None:
            self.disk_letter_input.delete("0","300")
            self.disk_letter_input.insert("0","P")
            self.FTP_adress_input.delete("0","300")
            self.FTP_adress_input.insert("0","\\\\192.168.000.000\\")
            self.username_input.delete("0","300")
            self.password_input.delete("0","300")
            if str(self.search_input.get()).replace(" ","") != "":
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.search_input.get()))
        else:
            self.copy_previous_project(True)

        child_root.update()
        child_root.update_idletasks()
        child_root.focus()
        child_root.focus_force()
        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

    def focused_entry_widget(self):
        currently_focused = str(self.root.focus_get())
        if ".!ctkentry" in currently_focused or ".!ctktextbox" in currently_focused:
            return True
        else:
            return False
        
    def make_sure_ip_changed(self,interface_name,ip):
        def run_as_admin():
            # Vyžádání admin práv: nefunkční ve vscode
            def is_admin():
                try:
                    return ctypes.windll.shell32.IsUserAnAdmin()
                except:
                    return False
            if not is_admin():
                ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
                sys.exit()
        def open_app_as_admin_prompt():
            def close_prompt(child_root):
                child_root.destroy()
            child_root = customtkinter.CTkToplevel()
            child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"620x150+{x+300}+{y+300}")  
            child_root.title("Upozornění")
            proceed_label = customtkinter.CTkLabel(master = child_root,text = "Přejete si znovu spustit aplikaci, jako administrátor?",font=("Arial",25))
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin(child_root))
            button_no =     customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
            proceed_label   .pack(pady=(15,0),padx=10,anchor="w",expand=False,side = "top")
            button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        interface_index = self.connection_option_list.index(interface_name)

        def call_subprocess():
            try:

                if ip == self.current_address_list[interface_index]:
                    add_colored_line(self.main_console,f"Pro interface {interface_name} je již tato adresa ({ip}) nastavena","orange",None,True)
                    return
                elif ip in self.current_address_list:
                    add_colored_line(self.main_console,f"Chyba, adresa je již používána pro jiný interface","red",None,True)
                    return
                win_change_ip_time = 7
                for i in range(0,win_change_ip_time):
                    add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                    self.option_change("",silent=True)
                    if ip == self.current_address_list[interface_index]: # někdy dříve než 7 sekund...
                        break
                    time.sleep(1)

                self.option_change("",silent=True)
                if ip == self.current_address_list[interface_index]:
                    add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla přenastavena na: {ip}","green",None,True)
                    self.refresh_ip_statuses()
                else:
                    print("temp ip troubleshooting: ------ ",ip)
                    add_colored_line(self.main_console,f"Chyba, neplatná adresa nebo daný inteface odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)
                    open_app_as_admin_prompt()
            except Exception:
                pass
        
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

    def check_DHCP(self,interface):
        process = subprocess.Popen(f'netsh interface ip show config name="{interface}"',
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        try:
            stdout_str = stdout.decode('utf-8')
            output_data = str(stdout_str)
        except UnicodeDecodeError:
            try:
                stdout_str = stdout.decode('cp1250')
                output_data = str(stdout_str)
            except UnicodeDecodeError:
                output_data = str(stdout)

        output_data_lines = output_data.split("\n")
        for lines in output_data_lines:
            if "DHCP enabled" in lines and "Yes" in lines:
                print(f"{interface} DHCP: yes")
                return True
        print(f"{interface} DHCP: no")

    def change_to_DHCP(self):
        def delay_the_refresh():
            nonlocal previous_addr
            nonlocal interface_index
            nonlocal interface
            new_addr = self.current_address_list[interface_index]
            i = 0
            while new_addr == previous_addr or new_addr == None:
                add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                time.sleep(1)
                self.option_change("",silent=True)
                new_addr = self.current_address_list[interface_index]
                print("current addr: ",new_addr)
                i+=1
                if i > 6:
                    add_colored_line(self.main_console,f"Chyba, u {interface} se nepodařilo změnit ip adresu (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)
                    return
            
            add_colored_line(self.main_console,f"IPv4 adresa interfacu: {interface} úspěšně přenastavena na DHCP (automatickou)","green",None,True)
            self.refresh_ip_statuses()
            return
        
        interface = str(self.drop_down_options.get())
        if not self.check_DHCP(interface):
            if interface != None or interface != "":
                interface_index = self.connection_option_list.index(interface)
                previous_addr = self.current_address_list[interface_index]
                print("previous addr: ",previous_addr)
                try:
                    # Construct the netsh command
                    netsh_command = f"netsh interface ipv4 set address name=\"{interface}\" source=dhcp"
                    print(f"calling: {netsh_command}")
                    powershell_command = [
                        'powershell.exe',
                        '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                    ]
                    process = subprocess.Popen(powershell_command,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
                    
                    stdout, stderr = process.communicate()
                    stdout_str = stdout.decode('utf-8')
                    stderr_str = stderr.decode('utf-8')
                    if stderr_str:
                        print(f"Error occurred: {stderr_str}")
                    else:
                        print(f"Command executed successfully:\n{stdout_str}")                
                    
                    run_background = threading.Thread(target=delay_the_refresh,)
                    run_background.start()

                except Exception as e:
                    print(f"Exception occurred: {str(e)}")
            else:
                add_colored_line(self.main_console,"Nebyl zvolen žádný interface","red",None,True)
        else:
            connected_interfaces = self.refresh_interfaces()
            if interface in connected_interfaces:
                add_colored_line(self.main_console,f"{interface} má již nastavenou DHCP","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Chyba, {interface} je odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)

    def change_computer_ip(self,button_row):
        def connected_interface(interface,ip,mask):
            """
            Když jsou vyžadována admin práva, tato funkce ověří, zda není daný interface připojen nebo součástí zařízení a zkusí znovu
            """
            try:
                # Construct the netsh command
                netsh_command = f"netsh interface ip set address \"{interface}\" static {ip} {mask}"
                powershell_command = [
                    'powershell.exe',
                    '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                ]
                process = subprocess.Popen(powershell_command,
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                
                stdout, stderr = process.communicate()
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                if stderr_str:
                    print(f"Error occurred: {stderr_str}")
                    add_colored_line(self.main_console,f"Chyba, nebyla poskytnuta práva (dejte ANO :))","red",None,True)
                else:
                    print(f"Command executed successfully:\n{stdout_str}")
                    self.make_sure_ip_changed(interface_name,ip)

            except Exception as e:
                print(f"Exception occurred: {str(e)}")

        
        ip = str(self.all_rows[button_row][1])
        mask = str(self.all_rows[button_row][2])
        # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
        interface_name = str(self.drop_down_options.get())
        powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
        try:
            process = subprocess.Popen(['powershell.exe', '-Command', powershell_command],
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr =process.communicate()
            stdout_str = stdout.decode('utf-8')
            stderr_str = stderr.decode('utf-8')
            print(f"VÝSTUP Z IP SETTING: {str(stdout_str)}")

            if len(str(stdout_str)) > 7:
                raise subprocess.CalledProcessError(1, powershell_command, stdout_str)
            if stderr_str:
                raise subprocess.CalledProcessError(1, powershell_command, stderr_str)

            self.make_sure_ip_changed(interface_name,ip)

        except subprocess.CalledProcessError as e:
            if "Run as administrator" in str(stdout_str):
                add_colored_line(self.main_console,f"Chyba, tato funkce musí být spuštěna s administrátorskými právy","red",None,True)
                # trigger powershell potvrzení:
                connected_interface(interface_name,ip,mask)
            elif "Invalid address" in str(stdout_str):
                add_colored_line(self.main_console,f"Chyba, neplatná IP adresa","red",None,True)
            else:
                add_colored_line(self.main_console,f"Chyba, Nemáte tuto adresu již nastavenou pro jiný interface? (nebo daný interface na tomto zařízení neexistuje)","red",None,True)
        except Exception as e:
            # Handle any other exceptions that may occur
            add_colored_line(self.main_console, f"Nastala neočekávaná chyba: {e}", "red", None, True)

    def check_given_input(self,given_data = None):
        if given_data == None:
            given_data = self.search_input.get()
        if given_data == "":
            found = None
            return found
        found = False

        if self.managing_disk == False:
            for i in range(0,len(self.all_rows)):
                if given_data == self.all_rows[i][0]:
                    self.last_project_name =    str(self.all_rows[i][0])
                    self.last_project_ip =      str(self.all_rows[i][1])
                    self.last_project_mask =    str(self.all_rows[i][2])
                    self.last_project_notes =   str(self.all_rows[i][3])
                    self.last_project_id = i
                    found = True
        else:
            for i in range(0,len(self.disk_all_rows)):
                if given_data == self.disk_all_rows[i][0]:
                    self.last_project_name =        str(self.disk_all_rows[i][0])
                    self.last_project_disk_letter = str(self.disk_all_rows[i][1])
                    self.last_project_ftp =         str(self.disk_all_rows[i][2])
                    self.last_project_username =    str(self.disk_all_rows[i][3])
                    self.last_project_password =    str(self.disk_all_rows[i][4])
                    self.last_project_notes =       str(self.disk_all_rows[i][5])
                    self.last_project_id = i
                    found = True
            
        return found    

    def clicked_on_project(self,event,widget_id,widget,textbox = "",flag = ""):
        """
        flag = notes:
        - při nakliknutí poznámky zůstanou expandnuté a při kliku na jinou je potřeba předchozí vrátit zpět
        flag = unfocus:
        - při kliku mimo se odebere focus z nakliknutých widgetů
        """
        print("widget_id",widget_id)
        if widget_id == None:
            return

        def on_leave_entry(widget,row_of_widget):
            """
            při kliku na jiný widget:
            - upraví text pouze na první řádek
            """
            if self.managing_disk:
                widget.configure(state = "normal")
                if "\n" in self.disk_all_rows[row_of_widget][5]:
                    notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                    first_row = notes_rows[0]
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(first_row))
                if self.default_note_behav == 0:
                    widget.configure(state = "disabled")
            else:
                widget.configure(state = "normal")
                if "\n" in self.all_rows[row_of_widget][3]:
                    notes_rows = self.all_rows[row_of_widget][3].split("\n")
                    first_row = notes_rows[0]
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(first_row))
                if self.default_note_behav == 0:
                    widget.configure(state = "disabled")

        def shrink_frame(widget_frame,widget_notes):
            widget_notes.configure(state = "normal")
            new_height = 50
            widget_frame.configure(height = new_height) #frame
            widget_notes.configure(height = new_height-10) #notes
            if self.default_note_behav == 0:
                widget_notes.configure(state = "disabled")

        if flag == "unfocus":
            try:
                if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                    if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                        on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                        shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""

                if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                    self.last_selected_widget.configure(border_color="#636363")
                    self.last_selected_widget = ""
            except Exception as e:
                print("chyba při odebírání focusu: ",e)

            return

        self.search_input.delete("0","300")
        if self.managing_disk == False:
            self.search_input.insert("0",str(self.all_rows[widget_id][0]))
        else:
            self.search_input.insert("0",str(self.disk_all_rows[widget_id][0]))

        self.check_given_input()
        # only if it is not pressed againt the same:
        if widget != self.last_selected_widget:
            try:
                if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                    if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                        on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                        shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                if flag == "notes":
                    self.last_selected_textbox = textbox
                    self.last_selected_notes_widget = widget
                    widget.focus_set()
                else:
                    # init the values:
                    self.last_selected_textbox = ""
                    self.last_selected_notes_widget = ""
                    widget.focus_set()

            except Exception as e:
                print("chyba s navracenim framu do puvodniho formatu",e)
            
            try:
                if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                    self.last_selected_widget.configure(border_color="#636363")
                self.last_selected_widget = widget
                widget.configure(border_color="white")

            except Exception as e:
                print(e)
                pass

            self.last_selected_widget_id = widget_id

    def is_project_favourite(self,array_index):
        try:
            fav_status = int(self.favourite_list[array_index])
            if fav_status == 1:
                return True
            else:
                return False
            
        except Exception:
            return False

    def refresh_ip_statuses(self):
        def unbind_connected_ip(widget,frame):
            widget.unbind("<Enter>")
            frame.unbind("<Enter>")
            widget.unbind("<Leave>")
            frame.unbind("<Leave>")
            frame.configure(fg_color = "black")

        def on_enter(e,interface,widget):
            widget.configure(text = interface)   

        def on_leave(e,ip,widget,frame):
            widget.configure(text = ip)
            if ip not in self.current_address_list:
                unbind_connected_ip(widget,frame)

        for i in range(0,len(self.ip_frame_list)):
            ip_addr = self.all_rows[i][1]
            ip_frame = self.ip_frame_list[i][0]
            parameter = self.ip_frame_list[i][1]
            if ip_addr in self.current_address_list:
                ip_frame.   configure(fg_color = "green") 
                ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
                parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
            else:
                unbind_connected_ip(parameter,ip_frame)

    def make_project_cells(self,no_read = None):

        self.clear_frame(self.project_tree)

        def opened_window_check():
            if self.opened_window == "":
                return False
            try:
                if self.opened_window.winfo_exists():
                    return True
                else:
                    return False
            except Exception as err:
                print(err)
                return False
            
        def on_enter(interface,widget):
            widget.configure(text = interface)   

        def on_leave(ip,widget,frame):
            widget.configure(text = ip)
            if ip not in self.current_address_list:
                unbind_connected_ip(widget,frame)

        def unbind_connected_ip(widget,frame):
            widget.unbind("<Enter>")
            frame.unbind("<Enter>")
            widget.unbind("<Leave>")
            frame.unbind("<Leave>")
            frame.configure(fg_color = "black")

        def filter_text_input(text):
            legit_rows = []
            legit_notes = ""
            rows = text.split("\n")
            for i in range(0,len(rows)):
                if rows[i].replace(" ","") != "":
                    legit_rows.append(rows[i])

            for i in range(0,len(legit_rows)): 
                if i == len(legit_rows)-1:
                    legit_notes = legit_notes + legit_rows[i]
                else:
                    legit_notes = legit_notes + legit_rows[i]+ "\n"
            return legit_notes
        
        def save_changed_notes(notes,row):
            workbook = load_workbook(self.excel_file_path)

            def find_notes_in_whole_list(row,new_fav_status):
                index_of_project = "no data"
                try:
                    wanted_project = self.all_rows[row][0]
                    self.show_favourite = new_fav_status
                    self.read_excel_data()
                    for i in range(0,len(self.all_rows)):
                        if self.all_rows[i][0] == wanted_project:
                            index_of_project = i
                            break
                    return index_of_project
                except Exception as err:
                    print(err)
                    return "no data"

            def save_to_workbook(notes,row,excel_worksheet):
                nonlocal workbook
                worksheet = workbook[excel_worksheet]
                worksheet['D' + str(len(self.all_rows)-row)] = notes

            if self.show_favourite:
                save_to_workbook(notes,row,"ip_adress_fav_list")
                index_of_project = find_notes_in_whole_list(row,new_fav_status = False)
                print("index",index_of_project)
                if str(index_of_project) != "no data":
                    save_to_workbook(notes,index_of_project,"ip_address_list")
                self.show_favourite = True
            else:
                save_to_workbook(notes,row,"ip_address_list")
                index_of_project = find_notes_in_whole_list(row,new_fav_status = True)
                print("index",index_of_project)
                if str(index_of_project) != "no data":
                    save_to_workbook(notes,index_of_project,"ip_adress_fav_list")
                self.show_favourite = False

            workbook.save(filename=self.excel_file_path)
            workbook.close()
            self.read_excel_data()

        def on_enter_entry(widget,row_of_widget):
            if not opened_window_check():
                if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                    widget.configure(state = "normal")
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(self.all_rows[row_of_widget][3]))
                    if self.default_note_behav == 0:
                        widget.configure(state = "disabled")

        def on_leave_entry(widget,row_of_widget):
            """
            při opuštění widgetu cursorem:
            - upraví text pouze na první řádek
            - uloží změny
            """
            if not opened_window_check():
                notes_before = filter_text_input(str(self.all_rows[row_of_widget][3]))
                notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
                if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                    widget.configure(state = "normal")
                    if notes_before != notes_after:
                        self.all_rows[row_of_widget][3] = notes_after
                        save_changed_notes(notes_after,row_of_widget)

                    if "\n" in self.all_rows[row_of_widget][3]:
                        notes_rows = self.all_rows[row_of_widget][3].split("\n")
                        first_row = notes_rows[0]
                        widget.delete("1.0",tk.END)
                        widget.insert(tk.END,str(first_row))

                    if self.default_note_behav == 0:
                        widget.configure(state = "disabled")
                    self.root.focus_set() # unfocus widget
                else:
                    # jinak pouze ulož změny
                    if notes_before != notes_after:
                        self.all_rows[row_of_widget][3] = notes_after
                        save_changed_notes(notes_after,row_of_widget)
                    self.root.focus_set() # unfocus widget
                    
        def shrink_frame(widget):
            if not opened_window_check():
                if str(widget[0]) != str(self.last_selected_notes_widget):
                    widget[1].configure(state = "normal")
                    new_height = 50
                    widget[0].configure(height = new_height) #frame
                    widget[1].configure(height = new_height-10) #notes
                    if self.default_note_behav == 0:
                        widget[1].configure(state = "disabled")

        def expand_frame(widget,row_of_widget):
            if not opened_window_check():
                if str(widget[0]) != str(self.last_selected_notes_widget):
                    # if the height is not 50 then it means it is expanded already
                    if widget[0].winfo_height() == 50:
                        widget[1].configure(state = "normal")
                        filtered_input = filter_text_input(self.all_rows[row_of_widget][3])
                        self.all_rows[row_of_widget][3] = filtered_input
                        addition = widget[0]._current_height
                        if "\n" in self.all_rows[row_of_widget][3]:
                            notes_rows = self.all_rows[row_of_widget][3].split("\n")
                            expanded_dim = addition + (len(notes_rows)-1) * 24
                            widget[0].configure(height = expanded_dim)
                            widget[1].configure(height = expanded_dim-10)
                            if self.default_note_behav == 0:
                                widget[1].configure(state = "disabled")
                        else:
                            if self.default_note_behav == 0:
                                widget[1].configure(state = "disabled")
         
        def add_row_return(widget):
            addition = widget[0]._current_height
            expanded_dim = addition + 24
            widget[0].configure(height = expanded_dim)
            widget[1].configure(height = expanded_dim-10)

        if no_read == None:
            self.read_excel_data()

        column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column2_header =    customtkinter.CTkLabel(master = column2,text = "IPv4 adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column3_header =    customtkinter.CTkLabel(master = column3,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column1.            pack(fill="both",expand=False,side = "left")
        column2.            pack(fill="both",expand=False,side = "left")
        column3.            pack(fill="both",expand=True, side = "left")
        column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
        column2_header.     pack(padx = (5,0),side = "top",anchor = "w")
        column3_header.     pack(padx = (5,0),side = "top",anchor = "w")

        self.ip_frame_list = []
        # y = widgets ve smeru y, x = widgets ve smeru x
        for y in range(0,len(self.all_rows)):
            # na pozici x = 2 je maska, kterou nevypisujeme
            for x in range(0,len(self.all_rows[y])):
                if x == 0: # frame s názvem projektu
                    btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    button =    customtkinter.CTkButton(master = btn_frame,width = 200,height=40,text = self.all_rows[y][x],font=("Arial",20,"bold"),corner_radius=0, command = lambda widget_id = y: self.change_computer_ip(widget_id))
                    button.     pack(padx =5,pady = 5, fill= "x")
                    btn_frame.  pack(side = "top",anchor = "w",expand = False,fill= "x")
                    button.     bind("<Button-1>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    # zkopírovat pravým klikem na button:
                    button.     bind("<Button-3>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))

                    if self.is_project_favourite(y):
                        button.configure(fg_color = "#1E90FF")

                elif x == 1: # frame s ip adresou
                    ip_frame =  customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    parameter = customtkinter.CTkLabel(master = ip_frame,text = self.all_rows[y][x],height=40,width = 250,font=("Arial",20,"bold"),justify='left',anchor = "w")
                    parameter.  pack(padx = (10,5),pady = 5)
                    ip_frame.   pack(side = "top")
                    ip_frame.   bind("<Button-1>",lambda e,widget = ip_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    parameter.  bind("<Button-1>",lambda e,widget = ip_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    ip_addr = self.all_rows[y][x]
                    self.ip_frame_list.append([ip_frame,parameter])
                    if ip_addr in self.current_address_list:
                        ip_frame.   configure(fg_color = "green")
                        ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                        ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))
                        parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                        parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))
                elif x == 3: # frame s poznamkami...
                    notes_frame =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),corner_radius=0,fg_color="black",height=40)
                    notes.          pack(padx =5,pady = 5,anchor="w",fill="x",expand = True)
                    notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                    notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, widget_id = y, textbox_widget = notes: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))
                    notes.          bind("<Button-1>",lambda e,widget = notes_frame, widget_id = y, textbox_widget = notes: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))

                    if "\n" in self.all_rows[y][x]:
                        notes_rows = self.all_rows[y][x].split("\n")
                        first_row = notes_rows[0]
                        notes.delete("1.0",tk.END)
                        notes.insert(tk.END,str(first_row))
                    else:
                        notes.insert(tk.END,str(self.all_rows[y][x]))
                    
                    notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(widget,row))
                    notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(widget))
                    notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(widget,row))
                    notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(widget,row))
                    
                    notes.bind("<Return>",lambda e, widget = [notes_frame,notes]: add_row_return(widget))

                    if self.default_note_behav == 0:
                        notes.configure(state = "disabled")

        self.project_tree.update()
        self.project_tree.update_idletasks()
        self.project_tree._parent_canvas.yview_moveto(0.0)
    
    def refresh_disk_statuses(self):
        self.refresh_btn.configure(text = "🔄",font=("",25))
        self.refresh_btn.update()
        self.refresh_btn.update_idletasks()
        mapped_disks = list_mapped_disks(whole_format = True)
        non_persistant_disks = list_non_persistent_disks()
        for y in range(0,len(self.disk_letter_frame_list)):
            param_frame = self.disk_letter_frame_list[y]
            param_frame.configure(fg_color = "black") # <= init
            for i in range(0,len(mapped_disks)):
                if mapped_disks[i][0:1] == str(self.disk_all_rows[y][1]):
                    drive_status = check_network_drive_status(mapped_disks[i])
                    if drive_status == True:
                        param_frame.configure(fg_color = "green")
                        if mapped_disks[i] in non_persistant_disks:
                            param_frame.configure(fg_color = "#00CED1")
                    else:
                        param_frame.configure(fg_color = "red")
        self.refresh_btn.configure(text = "Refresh statusů",font=("Arial",20,"bold"))

    def make_project_cells_disk(self,no_read = None,disk_statuses = False):
        def opened_window_check():
            if self.opened_window == "":
                return False
            try:
                if self.opened_window.winfo_exists():
                    return True
                else:
                    return False
            except Exception as e:
                print(e)
                return False

        def filter_text_input(text):
            legit_rows = []
            legit_notes = ""
            rows = text.split("\n")
            for i in range(0,len(rows)):
                if rows[i].replace(" ","") != "":
                    legit_rows.append(rows[i])

            for i in range(0,len(legit_rows)): 
                if i == len(legit_rows)-1:
                    legit_notes = legit_notes + legit_rows[i]
                else:
                    legit_notes = legit_notes + legit_rows[i]+ "\n"
            return legit_notes
        
        def save_changed_notes(notes,row):
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["disk_list"]
            worksheet['F' + str(len(self.disk_all_rows)-row)] = notes
            workbook.save(filename=self.excel_file_path)
            workbook.close()

        def on_enter_entry(widget,row_of_widget):
            if not opened_window_check():
                if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                    widget.configure(state = "normal")
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(self.disk_all_rows[row_of_widget][5]))
                    if self.default_note_behav == 0:
                        widget.configure(state = "disabled")

        def on_leave_entry(widget,row_of_widget):
            if not opened_window_check():
                notes_before = filter_text_input(str(self.disk_all_rows[row_of_widget][5]))
                notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
                if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                    if notes_before != notes_after:
                        widget.configure(state = "normal")
                        self.disk_all_rows[row_of_widget][5] = notes_after
                        save_changed_notes(notes_after,row_of_widget)

                    if "\n" in self.disk_all_rows[row_of_widget][5]:
                        notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                        first_row = notes_rows[0]
                        widget.delete("1.0",tk.END)
                        widget.insert(tk.END,str(first_row))
                    
                    if self.default_note_behav == 0:
                        widget.configure(state = "disabled")
                    self.root.focus_set() # unfocus widget
                else:
                    # jinak pouze ulož změny
                    if notes_before != notes_after:
                        self.disk_all_rows[row_of_widget][5] = notes_after
                        save_changed_notes(notes_after,row_of_widget)
                    self.root.focus_set() # unfocus widget

        def shrink_frame(widget):
            if not opened_window_check():
                if str(widget[0]) != str(self.last_selected_notes_widget):
                    widget[1].configure(state = "normal")
                    new_height = 50
                    if isinstance(widget,list):
                        widget[0].configure(height = new_height)
                        widget[1].configure(height = new_height-10)
                    else:
                        widget.configure(height = new_height)
                    if self.default_note_behav == 0:
                        widget[1].configure(state = "disabled")

        def expand_frame(widget,row_of_widget):
            if not opened_window_check():
                if str(widget[0]) != str(self.last_selected_notes_widget):
                    # if the height is not 50 then it means it is expanded already
                    if widget[0].winfo_height() == 50:
                        widget[1].configure(state = "normal")
                        filtered_input = filter_text_input(self.disk_all_rows[row_of_widget][5])
                        self.disk_all_rows[row_of_widget][5] = filtered_input
                        addition = widget[0]._current_height
                        if "\n" in self.disk_all_rows[row_of_widget][5]:
                            notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                            expanded_dim = addition + (len(notes_rows)-1) * 26
                            if isinstance(widget,list):
                                widget[0].configure(height = expanded_dim)
                                widget[1].configure(height = expanded_dim-10)
                            else:
                                widget.configure(height = expanded_dim)
                            if self.default_note_behav == 0:
                                widget[1].configure(state = "disabled")
                        else:
                            if self.default_note_behav == 0:
                                widget[1].configure(state = "disabled")

        def add_row_return(widget):
            addition = widget[0]._current_height
            expanded_dim = addition + 26
            widget[0].configure(height = expanded_dim)
            widget[1].configure(height = expanded_dim-10)

        if no_read == None:
            self.read_excel_data()
        self.clear_frame(self.project_tree)
        if self.default_disk_status_behav == 1:
            disk_statuses = True
        column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0,width = 50)
        column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column4 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
        column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column2_header =    customtkinter.CTkLabel(master = column2,text = "💾",font=("",22)) #💿
        column3_header =    customtkinter.CTkLabel(master = column3,text = "ftp adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column4_header =    customtkinter.CTkLabel(master = column4,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
        column1.            pack(fill="both",expand=False,side = "left")
        column2.            pack(fill="both",expand=False,side = "left")
        column3.            pack(fill="both",expand=False,side = "left")
        column4.            pack(fill="both",expand=True, side = "left")
        column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
        column2_header.     pack(padx = (12,0),side = "top",anchor = "w")
        column3_header.     pack(padx = (5,0),side = "top",anchor = "w",expand = False)
        column4_header.     pack(padx = (5,0),side = "top",anchor = "w")
        self.disk_letter_frame_list = []
        # y = widgets ve smeru y, x = widgets ve smeru x
        for y in range(0,len(self.disk_all_rows)):
            for x in range(0,len(self.disk_all_rows[y])):# x: 0=button, 1=disk_letter, 2=ip, 3=name, 4=password, 5=notes
                if x == 0:
                    btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    button =    customtkinter.CTkButton(master = btn_frame,width=200,height=40,text = self.disk_all_rows[y][x], command = lambda widget_id = y: self.map_disk(widget_id),font=("Arial",20,"bold"),corner_radius=0)
                    button.     pack(padx =5,pady = 5, fill= "x")
                    btn_frame.  pack(side = "top",anchor = "w",expand = False) 
                    btn_frame.  bind("<Button-1>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    button.     bind("<Button-3>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                elif x == 1: # frame s písmenem disku, menší šířka, podbarvení
                    param_frame =   customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    parameter =     customtkinter.CTkLabel(master = param_frame,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),width = 40,height=40)
                    parameter.      pack(padx = (5,5),pady = 5)
                    param_frame.    pack(side = "top")
                    self.disk_letter_frame_list.append(param_frame)
                    param_frame.    bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    parameter.      bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))

                elif x == 2: # frame s ftp adresou
                    param_frame =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    parameter =     customtkinter.CTkLabel(master = param_frame,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),justify='left',anchor = "w",width = 300,height=40)
                    parameter.      pack(padx = (10,5),pady = 5,anchor = "w",fill="x")
                    param_frame.    pack(side = "top",fill="x",expand = False)
                    param_frame.    bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                    parameter.      bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))

                elif x == 5: #frame s poznamkami...
                    notes_frame =   customtkinter.CTkFrame(master=column4,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                    notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),height=40,corner_radius=0,fg_color="black")
                    notes.          pack(padx =5,pady = 5,anchor="w",fill="x")
                    notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                    notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, widget_id = y: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))
                    notes.          bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, widget_id = y: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))

                    if "\n" in self.disk_all_rows[y][x]:
                        notes_rows = self.disk_all_rows[y][x].split("\n")
                        first_row = notes_rows[0]
                        notes.delete("1.0",tk.END)
                        notes.insert(tk.END,str(first_row))
                    else:
                        notes.insert(tk.END,str(self.disk_all_rows[y][x]))
                    
                    notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(widget,row))
                    notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(widget))
                    notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(widget,row))
                    notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(widget,row))

                    notes.bind("<Return>",lambda e, widget = [notes_frame,notes]: add_row_return(widget))

                    if self.default_note_behav == 0:
                        notes.configure(state = "disabled") 
        
        self.project_tree.update()
        self.project_tree.update_idletasks()
        self.project_tree._parent_canvas.yview_moveto(0.0)
        if disk_statuses:
            self.refresh_disk_statuses()     

    def edit_project(self):
        result = self.check_given_input()
        if result == True:
            if self.managing_disk == False:
                self.add_new_project(True)
            else:
                self.add_new_project_disk(True)
        elif result == None:
            add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
        else:
            add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
    
    def refresh_explorer(self,refresh_disk=None):
        """
        refresh_disk = udelat nove všechni widgets (make_project_cells_disk())
        """
        refresh_explorer="taskkill /f /im explorer.exe"
        subprocess.run(refresh_explorer, shell=True)
        refresh_explorer="start explorer.exe"
        subprocess.run(refresh_explorer, shell=True)
        if refresh_disk:
            self.make_project_cells_disk(disk_statuses=True)

    def delete_disk(self,child_root):
        drive_letter = str(self.drive_letter_input.get())
        if len(str(self.DL_manual_entry.get())) > 0:
            drive_letter = str(self.DL_manual_entry.get())
        
        delete_command = "net use " + drive_letter +": /del"
        process = subprocess.Popen(delete_command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
        stdout, stderr= process.communicate()
        if "Is it OK to continue disconnecting and force them closed?" in stdout:
            add_colored_line(self.main_console,f"Disk je právě používán, nejprve jej zavřete","red",None,True)
            child_root.destroy()
        else:
            self.refresh_explorer()
            add_colored_line(self.main_console,f"Disky s označením {drive_letter} byly odpojeny","orange",None,True)
            self.refresh_disk_statuses()
            child_root.destroy()

    def delete_disk_option_menu(self):
        child_root = customtkinter.CTkToplevel()
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"450x190+{x+250}+{y+200}")
        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        child_root.title("Odpojování síťového disku")
        
        found_drive_letters=[]
        for i in range(0,len(self.disk_all_rows)):
            if not self.disk_all_rows[i][1] in found_drive_letters:
                found_drive_letters.append(self.disk_all_rows[i][1])

        mapped_disks = list_mapped_disks()
        for i in range(0,len(mapped_disks)):
            if not mapped_disks[i] in found_drive_letters:
                found_drive_letters.append(mapped_disks[i])

        label =                     customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte disk nebo vyhledejte manuálně: ",font=("Arial",20,"bold"))
        self.drive_letter_input =   customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=found_drive_letters,corner_radius=0)
        self.DL_manual_entry =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manuálně")
        del_button =                customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Odpojit", command = lambda: self.delete_disk(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        exit_button =               customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
        label.                      grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
        self.drive_letter_input.    grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.DL_manual_entry.       grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        del_button.                 grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
        exit_button.                grid(column = 0,row=3,pady = 5,padx =220,sticky = tk.W)
        child_root.update()
        child_root.update_idletasks()
        child_root.focus()
        child_root.focus_force()
        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        print("selected: ",self.last_project_disk_letter)
        print("disk letter list: ",found_drive_letters)
        try:
            self.drive_letter_input.set(self.last_project_disk_letter)
        except Exception:
            pass

    def map_disk(self,button_row):
        Drive_letter = str(self.disk_all_rows[button_row][1])
        ftp_adress = str(self.disk_all_rows[button_row][2])
        user = str(self.disk_all_rows[button_row][3])
        password = str(self.disk_all_rows[button_row][4])

        delete_command = "net use " + Drive_letter + ": /del"
        subprocess.run(delete_command, shell=True)
        if self.mapping_condition == 1:
            persistant_status = " /persistent:yes"
        else:
            persistant_status =  " /persistent:no"

        if user != "" or password != "":
            second_command = "net use " + Drive_letter + ": " + ftp_adress + " " + password + " /user:" + user + persistant_status
        else:
            second_command = "net use " + Drive_letter + ": " + ftp_adress + persistant_status
        print("calling: ",second_command)

        def call_subprocess():
            """process = subprocess.Popen(second_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            stdout, stderr = process.communicate()
            self.connection_status = process.returncode
            print("STDOUT:", stdout)
            print("STDERR:", stderr)
            print("Return Code:", self.connection_status)"""
            self.connection_status = subprocess.call(second_command,shell=True,text=True)
  
        run_background = threading.Thread(target=call_subprocess,)
        run_background.start()

        time_start = time.time()
        while self.connection_status==None:
            time.sleep(0.05)
            if time.time() - time_start > 3:
                print("terminated due to runtime error")
                break

        if self.connection_status == 0:
            add_colored_line(self.main_console,f"Disk úspěšně připojen","green",None,True)
            self.refresh_explorer()
            self.refresh_disk_statuses()

            def open_explorer(path):
                if os.path.exists(path):
                    os.startfile(path)
                else:
                    print(f"The path {path} does not exist.")

            open_explorer(Drive_letter + ":\\")
        else:
             add_colored_line(self.main_console,f"Připojení selhalo (ixon? musí být zvolena alespoň 1 složka na disku...)","red",None,True)

    def get_ipv4_addresses(self):
        process = subprocess.Popen("ipconfig",
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE,
                                    creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        result2 = "" 
        try:
            stdout_str = stdout.decode('cp1250')
            result2 = str(stdout_str)
            
        except Exception as e:
            print("chyba ",e)

        # Regular expression to match the IPv4 address
        ipv4_pattern = re.compile(r'IPv4 Address[.\s]*: ([\d.]+)')
        ipv4_addresses = []
        lines = result2.splitlines()
        current_interface = None
        # Iterate over each line to find interface names and IPv4 addresses
        for line in lines:
            if line.strip():
                # Detect interface name
                if line[0].isalpha():
                    current_interface = line.strip()
                else:
                    # Detect IPv4 address for the current interface
                    match = ipv4_pattern.search(line)
                    if match and current_interface:
                        ipv4_addresses.append(current_interface)
                        ipv4_addresses.append(match.group(1))
        return ipv4_addresses

    def option_change(self,args,only_console = False,silent = False):
        """
        Volá get_current_ip_list(), aktualizuje současně nastavené adresy (self.current_address_list)
        - only console: vypíše do konzole aktuální připojení
        - silent: nevypisuje do konzole
        """
        if not only_console:
            try:
                self.default_connection_option = self.connection_option_list.index(self.drop_down_options.get())
            except ValueError as e:
                print(f"Error: {e}")
                self.default_connection_option = 0

            #pamatovat si naposledy zvoleny zpusob pripojeni:
            self.save_setting_parameter(parameter="change_def_conn_option",status=int(self.default_connection_option))
            self.get_current_ip_list()
            if self.static_label2.winfo_exists():
                self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
        if not silent:
            # ziskat data o aktualnim pripojeni
            current_connection = self.get_ipv4_addresses()
            message = ""
            for items in current_connection:
                message = message + items + " "
            if message == "":
                message = "nenalezeno"
            add_colored_line(self.main_console,f"Současné připojení: {message}","white",None,True)

    def make_project_first(self,purpouse=None,make_cells = True,project = None, input_entry_bypass = None):
        """
        purpouse:
        - search
        - silent
        """
        result = self.check_given_input(input_entry_bypass)

        if result == True:
            #zmena poradi
            if project == None:
                project = self.all_rows[self.last_project_id]
                favourite_status = self.favourite_list[self.last_project_id]
            else:
                favourite_status = project[4]

            self.all_rows.pop(self.last_project_id)
            self.all_rows.insert(0,project)
            self.favourite_list.pop(self.last_project_id)
            self.favourite_list.insert(0,favourite_status)

            for i in range(0,len(self.all_rows)):
                row = (len(self.all_rows)-1)-i
                self.save_excel_data(self.all_rows[i][0],self.all_rows[i][1],self.all_rows[i][2],self.all_rows[i][3],None,row+1,fav_status=self.favourite_list[i])
            if make_cells:
                self.make_project_cells()
            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} nalezen","green",None,True)
            elif purpouse != "silent":
                add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} přesunut na začátek","green",None,True)
        elif result == None and purpouse != "silent":
            print("nevlozeno id")
            if purpouse == "search":
                add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
        elif purpouse != "silent":
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)
            print("projekt nenalezen")

    def make_project_first_disk(self,purpouse = None):
        result = self.check_given_input()
        if result == True:
            #zmena poradi
            project = self.disk_all_rows[self.last_project_id]
            self.disk_all_rows.pop(self.last_project_id)
            self.disk_all_rows.insert(0,project)

            for i in range(0,len(self.disk_all_rows)):
                row = (len(self.disk_all_rows)-1)-i
                
                self.save_excel_data_disk(self.disk_all_rows[i][0],self.disk_all_rows[i][1],self.disk_all_rows[i][2],self.disk_all_rows[i][3],self.disk_all_rows[i][4],self.disk_all_rows[i][5],None,row+1)

            self.make_project_cells_disk()

            if purpouse == "search":
                add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} nalezen","green",None,True)
            else:
                add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} přesunut na začátek","green",None,True)
        elif result == None:
            if purpouse == "search":
                add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
            else:
                add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
        else:
            add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

    def get_current_ip_list(self):
        def get_current_ip_address(interface_name):
        # Get network interfaces and their addresses
            addresses = psutil.net_if_addrs()
            # Check if the specified interface exists
            if interface_name in addresses:
                addr_count = 0
                # print(f"Adresy interfacu: {interface_name}")
                for addr in addresses[interface_name]:
                    # prvni AF_INET je pridelena automaticky, druha je privatni, nastavena DHCP
                    if addr.family == socket.AF_INET:  # IPv4 address
                        if addr_count == 1:
                            return addr.address
                        addr_count +=1
                        remembered_addr = addr
                if addr_count == 1:
                    if "AddressFamily.AF_INET:" in str(remembered_addr):
                        return remembered_addr.address
                    else:
                        return "Nenalezeno"
            else:
                return "Nenalezeno"
        self.current_address_list = []
        for items in self.connection_option_list:
            found_address = get_current_ip_address(items)
            self.current_address_list.append(found_address)

    def save_setting_parameter(self,parameter,status):
        """
        list of parameters:\n

        change_def_conn_option\n
        new_conn_options\n
        change_def_ip_window\n
        change_def_main_window\n
        change_def_window_size\n
        change_def_disk_behav\n
        change_def_notes_behav\n
        change_mapping_cond\n
        change_make_first_behav\n
        """
        workbook = load_workbook(self.excel_file_path)
        worksheet = workbook["Settings"]
        if parameter == "change_def_conn_option":
            row = 1
        elif parameter == "new_conn_options":
            row = 2
        elif parameter == "change_def_ip_window":
            row = 3
        elif parameter == "change_def_main_window":
            row = 4
        elif parameter == "change_def_window_size":
            row = 5
        elif parameter == "change_def_disk_behav":
            row = 6
        elif parameter == "change_def_notes_behav":
            row = 7
        elif parameter == "change_mapping_cond":
            row = 8
        elif parameter == "change_make_first_behav":
            row = 9
        worksheet['B' + str(row)] = status
        workbook.save(filename=self.excel_file_path)
        workbook.close()

    def show_favourite_toggle(self,keep_search_input = False,determine_status = None): # hlavni prepinaci tlacitko oblibene/ neoblibene
        if self.show_favourite and (determine_status == None or determine_status == "all"):
            self.show_favourite = False
            window_status = 0
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.last_selected_widget_id = 0           
            self.ip_frame_list = []
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.search_input.configure(placeholder_text="Název projektu")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            self.button_remove_main.configure(command = lambda: self.delete_project(button_trigger=True))
            self.save_setting_parameter(parameter="change_def_ip_window",status=window_status)
            self.button_switch_favourite_ip. configure(fg_color="black")
            self.button_switch_all_ip.       configure(fg_color="#212121")
            self.button_remove_main.         configure(text="Smazat projekt")

        elif self.show_favourite == False and (determine_status == None or determine_status == "fav"):
            # favourite window
            self.show_favourite = True
            window_status = 1
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.last_selected_widget_id = 0
            self.ip_frame_list = []
            if keep_search_input == False:
                self.search_input.delete("0","300")
                self.search_input.configure(placeholder_text="Název projektu")
                self.make_project_cells()
            else:
                self.read_excel_data()
                self.check_given_input() #check ve druhem prostredi
                self.make_project_cells(no_read=True)
            self.button_remove_main.configure(command = lambda: self.switch_fav_status("with_refresh"))
            self.save_setting_parameter(parameter="change_def_ip_window",status=window_status)
            self.button_switch_favourite_ip. configure(fg_color="#212121")
            self.button_switch_all_ip.       configure(fg_color="black")
            self.button_remove_main.         configure(text="Odebrat projekt")

    def refresh_interfaces(self,all = False):
        """
        - All parametr refreshne i statusy ip adres
        """
        interfaces_data = self.fill_interfaces()
        self.connection_option_list = interfaces_data[0]
        self.drop_down_options.configure(values = self.connection_option_list)
        online_list_text = ""
        if len(interfaces_data[1]) > 0:
            for data in interfaces_data[1]:
                online_list_text = online_list_text + str(data) +", "
            online_list_text = online_list_text[:-2] # odebrat čárku s mezerou

        self.online_list.configure(text=online_list_text)
        if all:
            self.option_change("")

        return interfaces_data[1]

    def setting_window(self,ip_window = False):
        def save_new_behav_disk():
            nonlocal checkbox2
            if int(checkbox2.get()) == 0:
                self.default_disk_status_behav = 0
                self.save_setting_parameter(parameter="change_def_disk_behav",status=0)
            elif int(checkbox2.get()) == 1:
                self.default_disk_status_behav = 1
                self.save_setting_parameter(parameter="change_def_disk_behav",status=1)
                self.make_project_cells_disk(no_read=True)

        def save_new_behav_notes():
            nonlocal checkbox
            nonlocal ip_window
            if int(checkbox.get()) == 0:
                self.default_note_behav = 0
                self.save_setting_parameter(parameter="change_def_notes_behav",status=0)
                if ip_window:
                    self.make_project_cells()
                else:
                    self.make_project_cells_disk()

            elif int(checkbox.get()) == 1:
                self.default_note_behav = 1
                self.save_setting_parameter(parameter="change_def_notes_behav",status=1)
                if ip_window:
                    self.make_project_cells()
                else:
                    self.make_project_cells_disk()

        def save_new_disk_map_cond():
            nonlocal checkbox3
            if int(checkbox3.get()) == 0:
                self.mapping_condition = 0
                self.save_setting_parameter(parameter="change_mapping_cond",status=0)
            elif int(checkbox3.get()) == 1:
                self.mapping_condition = 1
                self.save_setting_parameter(parameter="change_mapping_cond",status=1)

        def change_make_first_behav():
            nonlocal checkbox4
            if int(checkbox4.get()) == 0:
                self.make_edited_project_first = False
                self.save_setting_parameter(parameter="change_make_first_behav",status=0)
            elif int(checkbox4.get()) == 1:
                self.make_edited_project_first = True
                self.save_setting_parameter(parameter="change_make_first_behav",status=1)

        child_root = customtkinter.CTkToplevel()
        self.opened_window = child_root
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        if ip_window:
            child_root.geometry(f"580x280+{x+350}+{y+180}")
        else:
            child_root.geometry(f"620x490+{x+350}+{y+180}")

        child_root.after(200, lambda: child_root.iconbitmap(resource_path(self.app_icon)))
        child_root.title("Nastavení")
        main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
        label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "- Chování poznámek (editovatelné/ needitovatelné)",font=("Arial",20,"bold"))
        checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "Přímo zapisovat a ukládat do poznámek na úvodní obrazovce",font=("Arial",16,"bold"),command=lambda: save_new_behav_notes())
        label.          pack(pady = 10,padx=10,side="top",anchor = "w")
        checkbox.       pack(pady = 10,padx=10,side="top",anchor = "w")

        main_frame2 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
        label2 =        customtkinter.CTkLabel(master = main_frame2, width = 100,height=40,text = "- Chování při vstupu do menu \"Síťové disky\"",font=("Arial",20,"bold"))
        checkbox2 =     customtkinter.CTkCheckBox(master = main_frame2, text = "Při spuštění aktualizovat statusy disků",font=("Arial",16,"bold"),command=lambda: save_new_behav_disk())
        label2.         pack(pady = 10,padx=10,side="top",anchor = "w")
        checkbox2.      pack(pady = 10,padx=10,side="top",anchor = "w")

        main_frame3 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
        label3 =        customtkinter.CTkLabel(master = main_frame3, width = 100,height=40,text = "- Nastavení mapování disků",font=("Arial",20,"bold"))
        checkbox3 =     customtkinter.CTkCheckBox(master = main_frame3, text = "Automaticky připojovat po restartu PC",font=("Arial",16,"bold"),command=lambda: save_new_disk_map_cond())

        frame_drive1 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
        drive_color1 =  customtkinter.CTkFrame(master=frame_drive1,corner_radius=0,width = 30,height = 30,fg_color="green")
        drive_label1 =  customtkinter.CTkLabel(master = frame_drive1, width = 100,height=40,text = "= disk je online, persistentní (po vypnutí bude znovu načten)",font=("Arial",20))
        drive_color1.   pack(pady = (5,0),padx=10,side="left",anchor = "w")
        drive_label1.   pack(pady = (5,0),padx=0,side="left",anchor = "w")
        frame_drive2 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
        drive_color2 =  customtkinter.CTkFrame(master=frame_drive2,corner_radius=0,width = 30,height = 30,fg_color="#00CED1")
        drive_label2 =  customtkinter.CTkLabel(master = frame_drive2, width = 100,height=40,text = "= disk je online, nepersistentní (bude odpojen po vypnutí)",font=("Arial",20))
        drive_color2.   pack(pady = (5,0),padx=10,side="left",anchor = "w")
        drive_label2.   pack(pady = (5,0),padx=0,side="left",anchor = "w")
        label3.         pack(pady = 10,padx=10,side="top",anchor = "w")
        checkbox3.      pack(pady = 10,padx=10,side="top",anchor = "w")
        frame_drive1.   pack(pady = 0,padx=0,side="top",anchor = "w",fill="x")
        frame_drive2.   pack(pady = 0,padx=0,side="top",anchor = "w",fill="x")
        
        main_frame4 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
        label4 =        customtkinter.CTkLabel(master = main_frame4, width = 100,height=40,text = "- Nastavení chování při editaci projektů",font=("Arial",20,"bold"))
        checkbox4 =     customtkinter.CTkCheckBox(master = main_frame4, text = "Automaticky přesouvat editovaný projekt na začátek",font=("Arial",16,"bold"),command=lambda: change_make_first_behav())
        label4.         pack(pady = 10,padx=10,side="top",anchor = "w")
        checkbox4.      pack(pady = 10,padx=10,side="top",anchor = "w")

        close_frame =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
        button_close =  customtkinter.CTkButton(master = close_frame, width = 150,height=40,text = "Zavřít",command = child_root.destroy,font=("Arial",20,"bold"),corner_radius=0)
        button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

        if ip_window:
            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame4.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")
            
        else: #disk window...
            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame2.    pack(expand=False,fill="x",side="top")
            main_frame3.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")

        if self.default_note_behav == 1:
            checkbox.select()
        
        if self.make_edited_project_first:
            checkbox4.select()

        if self.default_disk_status_behav == 1 and ip_window == False:
            checkbox2.select()

        if self.mapping_condition == 1 and ip_window == False:
            checkbox3.select()

        child_root.update()
        child_root.update_idletasks()
        child_root.focus()
        child_root.focus_force()
        self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

    def create_widgets(self,fav_status = None,init=None,excel_load_error = False):
        if not excel_load_error:
            if init:
                if self.window_mode == "max":
                    self.save_setting_parameter(parameter="change_def_window_size",status=1)
                else:
                    self.save_setting_parameter(parameter="change_def_window_size",status=0)
            if fav_status:
                self.show_favourite = True
                self.save_setting_parameter(parameter="change_def_ip_window",status=1)
            if fav_status == False:
                self.show_favourite = False
                self.save_setting_parameter(parameter="change_def_ip_window",status=0)
            
            self.save_setting_parameter(parameter="change_def_main_window",status=0)
        
        self.clear_frame(self.root)
        self.managing_disk = False
        menu_cards =            customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50,border_width=0)
        main_widgets =          customtkinter.CTkFrame(master=self.root,corner_radius=0,border_width=0)
        self.project_tree =     customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0,border_width=0)

        menu_cards.             pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        logo =                  customtkinter.CTkImage(Image.open(resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =            customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
        image_logo.             pack(pady=5,padx=15,expand=True,side = "right",anchor="e")
        main_widgets.           pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.      pack(pady=5,padx=5,fill="both",expand=True,side = "top")

        main_menu_button =                  customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        self.button_switch_all_ip =         customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: self.show_favourite_toggle(determine_status="all"),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        self.button_switch_favourite_ip =   customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: self.show_favourite_toggle(determine_status="fav"),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_disk =                customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",command =  lambda: self.create_widgets_disk(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        if excel_load_error:
            self.button_switch_all_ip.configure(state = "disabled")
            self.button_switch_favourite_ip.configure(state = "disabled")
            button_switch_disk.configure(state = "disabled")

        project_label =             customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =         customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
        button_search =             customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first("search"),font=("Arial",20,"bold"),corner_radius=0)
        self.button_add_main =      customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project(),font=("Arial",20,"bold"),corner_radius=0)
        self.button_remove_main =   customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Smazat projekt", command =  lambda: self.delete_project(button_trigger=True),font=("Arial",20,"bold"),corner_radius=0)
        self.button_edit_main =     customtkinter.CTkButton(master = main_widgets, width = 160,height=40,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
        button_make_first =         customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "Přesunout na začátek",command =  lambda: self.make_project_first(),font=("Arial",20,"bold"),corner_radius=0)
        button_settings_behav =     customtkinter.CTkButton(master = main_widgets, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(ip_window=True),font=("",22),corner_radius=0)
        
        if self.show_favourite:
            self.button_remove_main.         configure(text="Odebrat projekt")
            self.button_switch_favourite_ip. configure(fg_color="#212121")
            self.button_switch_all_ip.       configure(fg_color="black")
        else:
            self.button_remove_main.         configure(text="Smazat projekt")
            self.button_switch_favourite_ip. configure(fg_color="black")
            self.button_switch_all_ip.       configure(fg_color="#212121")

        connect_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "Připojení: ",font=("Arial",20,"bold"))
        self.drop_down_options = customtkinter.CTkOptionMenu(master = main_widgets,width=200,height=40,font=("Arial",20,"bold"),corner_radius=0,command=  self.option_change)
        # "⚙️", "⚒", "🔧", "🔩"
        button_settings =       customtkinter.CTkButton(master = main_widgets, width = 40,height=40,text="⚒",command =  lambda: self.refresh_interfaces(all=True),font=("",22),corner_radius=0) #refresh interface statusů
        button_dhcp =           customtkinter.CTkButton(master = main_widgets, width = 100,height=40,text = "DHCP",command =  lambda: self.change_to_DHCP(),font=("Arial",20,"bold"),corner_radius=0)
        static_label =          customtkinter.CTkLabel(master = main_widgets, height=40,text = "Static:",font=("Arial",20,"bold"))
        self.static_label2 =    customtkinter.CTkLabel(master = main_widgets, height=40,text = "",font=("Arial",22,"bold"),bg_color="black")
        online_label =          customtkinter.CTkLabel(master = main_widgets, height=40,text = "Online: ",font=("Arial",22,"bold"))
        self.online_list =      customtkinter.CTkLabel(master = main_widgets, height=40,text = "",font=("Arial",22,"bold"))
        self.main_console =     tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
        main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        self.button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        self.button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        image_logo.                     pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
        project_label.              grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.          grid(column = 0,row=0,pady = 5,padx =90,sticky = tk.W)
        button_search.              grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        self.button_add_main.       grid(column = 0,row=0,pady = 5,padx =410,sticky = tk.W)
        self.button_remove_main.    grid(column = 0,row=0,pady = 5,padx =565,sticky = tk.W)
        self.button_edit_main.      grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)
        button_make_first.          grid(column = 0,row=0,pady = 5,padx =885,sticky = tk.W)
        button_settings_behav.      grid(column = 0,row=0,pady = 5,padx =1100,sticky = tk.W)
        connect_label.              grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        self.drop_down_options.     grid(column = 0,row=1,pady = 0,padx =110,sticky = tk.W)
        button_settings.            grid(column = 0,row=1,pady = 0,padx =315,sticky = tk.W)
        button_dhcp.                grid(column = 0,row=1,pady = 0,padx =360,sticky = tk.W)
        static_label.               grid(column = 0,row=1,pady = 0,padx =470,sticky = tk.W)
        self.static_label2.         grid(column = 0,row=1,pady = 0,padx =540,sticky = tk.W,ipadx = 10,ipady = 2)
        online_label.               grid(column = 0,row=1,pady = 0,padx =725,sticky = tk.W)
        self.online_list.           grid(column = 0,row=1,pady = 0,padx =805,sticky = tk.W)

        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.refresh_interfaces()
        # aktualizace hodnot nabídky
        if self.default_connection_option < len(self.connection_option_list):
            # nastavení naposledy zvoleného interfacu
            self.drop_down_options.set(self.connection_option_list[self.default_connection_option])
        else:
            self.default_connection_option = 0
            self.save_setting_parameter(parameter="change_def_conn_option",status=0)
            self.drop_down_options.set(self.connection_option_list[self.default_connection_option])

        if not excel_load_error:
            self.option_change("")
            self.make_project_cells()
            self.get_current_ip_list()
            self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
        else:
            only_name = self.excel_file_path.split("/")
            only_name = only_name[len(only_name)-1]
            add_colored_line(self.main_console,f"Konfigurační soubor: {only_name} nebyl nalezen nebo je otevřený","red",None,True)

        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")
                self.save_setting_parameter(parameter="change_def_window_size",status=2)
            elif int(current_width) ==260:
                self.root.geometry("1200x900")
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
            else:
                self.root.state('zoomed')
                self.save_setting_parameter(parameter="change_def_window_size",status=1)

        self.root.bind("<f>",lambda e: maximalize_window(e))

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first("search")
        self.search_input.bind("<Return>",call_search)

        def call_unfocus():
            if not ".!ctkscrollableframe" in str(self.root.focus_get()) and not ".!ctktoplevel" in str(self.root.focus_get()):
                #odebrat focus
                self.clicked_on_project("",None,None,None,flag="unfocus")
        self.root.bind("<Button-1>",lambda e: call_unfocus(),"+")

        self.root.mainloop()

    def create_widgets_disk(self,init=None):
        if init:
            if self.window_mode == "max":
                self.save_setting_parameter(parameter="change_def_window_size",status=1)
            else:
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
        self.clear_frame(self.root)
        self.managing_disk = True
        self.save_setting_parameter(parameter="change_def_main_window",status=1)
        menu_cards =            customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        main_widgets =          customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.project_tree =     customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        menu_cards.             pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        logo =                  customtkinter.CTkImage(Image.open(resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =            customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
        image_logo.             pack(pady=5,padx=15,expand=True,side = "right",anchor="e")
        main_widgets.           pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree.      pack(pady=5,padx=5,fill="both",expand=True,side = "top")

        main_menu_button =              customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_all_ip =          customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: self.create_widgets(fav_status=False),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_favourite_ip =    customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: self.create_widgets(fav_status=True),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        button_switch_disk =            customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",command =  lambda: self.create_widgets_disk(),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        project_label =         customtkinter.CTkLabel(master = main_widgets, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"))
        self.search_input =     customtkinter.CTkEntry(master = main_widgets,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
        button_search =         customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first_disk("search"),font=("Arial",20,"bold"),corner_radius=0)
        button_add =            customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project_disk(),font=("Arial",20,"bold"),corner_radius=0)
        button_remove =         customtkinter.CTkButton(master = main_widgets, width = 150,height=40,text = "Smazat projekt", command =  lambda: self.delete_project_disk(button_trigger=True),font=("Arial",20,"bold"),corner_radius=0)
        button_edit =           customtkinter.CTkButton(master = main_widgets, width = 160,height=40,text = "Editovat projekt",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
        button_make_first =     customtkinter.CTkButton(master = main_widgets, width = 250,height=40,text = "Přesunout na začátek",command =  lambda: self.make_project_first_disk(),font=("Arial",20,"bold"),corner_radius=0)
        button_settings =       customtkinter.CTkButton(master = main_widgets, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(),font=("",22),corner_radius=0)
        delete_disk =           customtkinter.CTkButton(master = main_widgets, width = 250,height=40,text = "Odpojit síťový disk",command =  lambda: self.delete_disk_option_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
        reset =                 customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "Reset exploreru",command =  lambda: self.refresh_explorer(refresh_disk=True),font=("Arial",20,"bold"),corner_radius=0)
        self.refresh_btn =      customtkinter.CTkButton(master = main_widgets, width = 200,height=40,text = "Refresh statusů",command =  lambda: self.refresh_disk_statuses(),font=("Arial",20,"bold"),corner_radius=0)
        as_admin_label =        customtkinter.CTkLabel(master = main_widgets,text = "",font=("Arial",20,"bold"))
        self.main_console =     tk.Text(main_widgets, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)

        main_menu_button.          pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        button_switch_disk.        pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
        image_logo.                pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
        project_label.      grid(column = 0,row=0,pady = 5,padx =0,sticky = tk.W)
        self.search_input.  grid(column = 0,row=0,pady = 5,padx =90,sticky = tk.W)
        button_search.      grid(column = 0,row=0,pady = 5,padx =255,sticky = tk.W)
        button_add.         grid(column = 0,row=0,pady = 5,padx =410,sticky = tk.W)
        button_remove.      grid(column = 0,row=0,pady = 5,padx =565,sticky = tk.W)
        button_edit.        grid(column = 0,row=0,pady = 5,padx =720,sticky = tk.W)
        button_make_first.  grid(column = 0,row=0,pady = 5,padx =885,sticky = tk.W)
        button_settings.    grid(column = 0,row=0,pady = 5,padx =1140,sticky = tk.W)
        delete_disk.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
        reset.              grid(column = 0,row=1,pady = 5,padx =265,sticky = tk.W)
        self.refresh_btn.   grid(column = 0,row=1,pady = 5,padx =470,sticky = tk.W)
        as_admin_label.     grid(column = 0,row=1,pady = 5,padx =675,sticky = tk.W)
        self.main_console.grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
        self.option_change("",only_console=True)
        
        def is_admin():
            try:
                return ctypes.windll.shell32.IsUserAnAdmin()
            except:
                return False
            
        if is_admin():
            as_admin_label.configure(text = "Aplikace je spuštěna, jako administrátor\n(mapovat disky lze pouze na uživatelském účtu)",text_color = "orange")

        def maximalize_window(e):
            self.root.update_idletasks()
            self.root.update()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                self.root.state('normal')
                self.root.geometry(f"260x1000+{0}+{0}")
                self.save_setting_parameter(parameter="change_def_window_size",status=2)
            elif int(current_width) ==260:
                self.root.geometry("1200x900")
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
            else:
                self.root.state('zoomed')
                self.save_setting_parameter(parameter="change_def_window_size",status=1)
            self.root.update_idletasks()
            self.root.update()
        self.root.bind("<f>",lambda e: maximalize_window(e))

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.search_input.bind("<Return>",unfocus_widget)

        def call_search(e):
            self.make_project_first_disk("search")
        self.search_input.bind("<Return>",call_search)

        def call_refresh(e):
            self.refresh_explorer(refresh_disk=True)
        self.root.bind("<F5>",lambda e: call_refresh(e))

        def call_unfocus():
            if not ".!ctkscrollableframe" in str(self.root.focus_get()) and not ".!ctktoplevel" in str(self.root.focus_get()):
                #odebrat focus
                self.clicked_on_project("",None,None,None,flag="unfocus")
        self.root.bind("<Button-1>",lambda e: call_unfocus(),"+")

        self.root.update()
        self.make_project_cells_disk()
        self.root.mainloop()

if testing_mode:
    IP_assignment(root,"","max",str(os.getcwd())+"\\")
    root.mainloop()