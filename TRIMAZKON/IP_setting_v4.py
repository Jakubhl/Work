import customtkinter
import tkinter as tk
from openpyxl import load_workbook
import subprocess
import os
import re
import time
import threading
import psutil
import socket
from PIL import Image
import sys
import ctypes
import winreg
import win32net
import copy
import pyperclip

testing_mode = False
if testing_mode:
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    root=customtkinter.CTk()
    root.geometry("1200x900")
    root.title("ip_setting - testing")
    root.state('zoomed')

class Tools:
    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)
    
    @classmethod
    def add_colored_line(cls,text_widget, text, color,font=None,delete_line = None):
        """
        Vloží řádek do console
        """
        text_widget.configure(state=tk.NORMAL)
        if font == None:
            font = ("Arial",16)
        if delete_line != None:
            text_widget.delete("current linestart","current lineend")
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert("current lineend",text, color)
        else:
            text_widget.tag_configure(color, foreground=color,font=font)
            text_widget.insert(tk.END,"    > "+ text+"\n", color)

        text_widget.configure(state=tk.DISABLED)
    
    @classmethod
    def save_setting_parameter(cls,parameter,status,excel_path):
        """
        list of parameters:\n
        change_def_conn_option\n
        new_conn_options\n
        change_def_ip_window\n
        change_def_main_window\n
        change_def_window_size\n
        change_def_disk_behav\n
        change_def_notes_behav\n
        change_mapping_cond\n
        change_make_first_behav\n
        delete_behav\n
        """

        parameter_row_mapping = {
        "change_def_conn_option": 1,
        "new_conn_options": 2,
        "change_def_ip_window": 3,
        "change_def_main_window": 4,
        "change_def_window_size": 5,
        "change_def_disk_behav": 6,
        "change_def_notes_behav": 7,
        "change_mapping_cond": 8,
        "change_make_first_behav": 9,
        "delete_behav": 10
        }

        row = parameter_row_mapping.get(parameter)
        if row is None:
            print(f"Invalid parameter: {parameter}")
            return
        
        workbook = load_workbook(excel_path)
        worksheet = workbook["Settings"]
        worksheet['B' + str(row)] = status
        workbook.save(filename=excel_path)
        workbook.close()

    @classmethod
    def read_setting_parameters(cls,excel_file_path):
        """
        - [0] = default connection option (0/1)
        - [1] = show favourite ip as default (0/1)
        - [2] = show disk environment as default (0/1)
        - [3] = last set widnow size  (0/1/2 - 2 is the narrow and long one)
        - [4] = check disk statutes automatically status (0/1)
        - [5] = editable/ non-editable notes (0/1)
        - [6] = persistent/ non-persistent disk (0/1)
        - [7] = shift edited project on top status (0/1)
        - [8] = delete - pop up window main window (110), when edit (101)
        """
        def insert_new_excel_param(wb,ws,row,param,text):
            """
            Oveřuje zda konfigurační excel již obsahuje tyto parametry, případně zapíše
            """
            ws['B' + str(row)] = param
            ws['A' + str(row)] = text
            wb.save(excel_file_path)
            print('inserting new parameter to excel')

        try:
            workbook = load_workbook(excel_file_path)
            worksheet = workbook["Settings"]
            saved_def_con_option = worksheet['B' + str(1)].value
            def_show_favourite = worksheet['B' + str(3)].value
            def_show_disk = worksheet['B' + str(4)].value
            def_window_size = worksheet['B' + str(5)].value
            
            value_check = worksheet['B' + str(6)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,row=6,param=0,text="aktualizovat statusy disků při vstupu do okna s disky (default)")
            else:
                default_disk_status_behav = int(worksheet['B' + str(6)].value)

            value_check = worksheet['B' + str(7)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,row=7,param=0,text="editovatelné(1)/ needitovatelné(0) poznámky (default)")
            else:
                default_note_behav = int(worksheet['B' + str(7)].value)

            value_check = worksheet['B' + str(8)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,row=8,param=0,text="disk persistentní - yes(1)/ no(0)")
            else:
                mapping_condition = int(worksheet['B' + str(8)].value)

            value_check = worksheet['B' + str(9)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,row=9,param=1,text="automaticky přesouvat upravené projekty na začátek")
            else:
                excel_value =  int(worksheet['B' + str(9)].value)
                if excel_value == 1:
                    make_edited_project_first = True
                else:
                    make_edited_project_first = False
            
            value_check = worksheet['B' + str(10)].value
            if value_check is None or str(value_check) == "":
                insert_new_excel_param(workbook,worksheet,row=10,param=100,text="statusy odvolby dotazování při mazání")
            else:
                deletion_behav = int(worksheet['B' + str(10)].value)
            workbook.close()

            return [int(saved_def_con_option),
                    int(def_show_favourite),
                    int(def_show_disk),
                    int(def_window_size),
                    default_disk_status_behav,
                    default_note_behav,
                    mapping_condition,
                    make_edited_project_first,
                    deletion_behav]

        except Exception as e:
            print(f"Nejdřív zavřete soubor {excel_file_path} Chyba: {e}")
            return 

    @classmethod
    def clear_frame(cls,frame):
        frame.update()
        frame.update_idletasks()
        for widget in frame.winfo_children():
            if widget.winfo_exists():
                widget.unbind("<Enter>")
                widget.unbind("<Leave>")
                widget.unbind("<Return>")
                widget.unbind("<Button-1>")
                widget.unbind("<Button-3>")
                widget.unbind("<Double-1>")
                widget.unbind("<MouseWheel>")
                widget.pack_forget()
                widget.grid_forget()
                widget.destroy()
    
    @classmethod
    def get_legit_notes(cls,notes):
        notes_legit_rows = []
        notes_rows = notes.split("\n")
        for i in range(0,len(notes_rows)):
            if notes_rows[i].replace(" ","") != "":
                notes_legit_rows.append(notes_rows[i])
        string_for_excel = ""
        for i in range(0,len(notes_legit_rows)):
            if i != len(notes_legit_rows)-1:
                string_for_excel = string_for_excel + str(notes_legit_rows[i]) + "\n"
            else:
                string_for_excel = string_for_excel + str(notes_legit_rows[i])

        return string_for_excel
    
    @classmethod
    def focused_entry_widget(cls,root):
        currently_focused = str(root.focus_get())
        if ".!ctkentry" in currently_focused or ".!ctktextbox" in currently_focused:
            return True
        else:
            return False

    @classmethod
    def get_none_count(cls,array_given):
        none_count = 0
        for items in array_given:
            if items == None:
                none_count += 1
        return none_count

class main:
    class DM_tools:  
        @classmethod
        def check_network_drive_status(cls,drive_path):
            checking_done = False
            status = False
            try:
                # Attempt to access a file or directory on the network drive
                drive_path = drive_path[0:3]
                def call_subprocess():
                    nonlocal checking_done
                    nonlocal status
                    if os.path.exists(drive_path):
                        os.listdir(drive_path)
                        status = True
                        checking_done = True
                        return
                    else:
                        status = False
                        checking_done = True
                        return
                    
                run_background = threading.Thread(target=call_subprocess,)
                run_background.start()

                time_start = time.time()
                while checking_done==False:
                    time.sleep(0.05)
                    if time.time() - time_start > 1:
                        print("terminated due to runtime error")
                        return False
                
                if status == True:
                    return True
                else:
                    return False

            except FileNotFoundError:
                return False
            except OSError:
                return False
            
        @classmethod
        def list_mapped_disks(cls,whole_format=None):
            remote_drives = []
            try:
                reg_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Network')
                for i in range(0, winreg.QueryInfoKey(reg_key)[0]):
                    drive_letter = winreg.EnumKey(reg_key, i)
                    if whole_format:
                        remote_drives.append(drive_letter + ':')
                    else:
                        remote_drives.append(drive_letter)
                winreg.CloseKey(reg_key)
            except Exception as e:
                print("Exception occurred: ", e)

            print("persistent disks: ",remote_drives)
                    
            return remote_drives

        @classmethod
        def list_non_persistent_disks(cls):
            non_persistent_drives = []
            try:
                # Enumerate network connections
                level = 1  # Level 1 provides the 'ui1_flags' information
                connections, _, _ = win32net.NetUseEnum(None, level)
                for i in range(0,len(connections)):
                    non_persistent_drives.append(connections[i]["local"]) 
            except Exception as e:
                print("Exception occurred: ", e)

            print("non-persistent disks: ",non_persistent_drives)
            
            return non_persistent_drives

        @classmethod
        def save_excel_data_disk(cls,
                                excel_file_path,
                                len_of_disk_array,
                                last_project_id,
                                project_name,
                                disk_letter,
                                ftp_address,
                                username,
                                password,
                                notes,
                                only_edit = None,
                                force_row_to_print=None,
                                force_ws = None,
                                wb_given = None):
            if wb_given == None:
                workbook = load_workbook(excel_file_path)
            else:
                workbook = wb_given

            if force_ws == None:
                worksheet = workbook["disk_list"]
            else:
                worksheet = workbook[force_ws]
            # excel je od jednicky...
            if force_row_to_print == None:
                row_to_print = int(len_of_disk_array) +1
                if only_edit != None:
                    #pouze změna na temtýž řádku
                    if last_project_id != "":
                        row_to_print = (len_of_disk_array - last_project_id)
            else:
                row_to_print = force_row_to_print
            #A = nazev projektu
            worksheet['A' + str(row_to_print)] = project_name
            #B = písmeno disku, označení...
            worksheet['B' + str(row_to_print)] = disk_letter
            #C = ftp adresa
            worksheet['C' + str(row_to_print)] = ftp_address
            #D = uživatelské jméno
            worksheet['D' + str(row_to_print)] = username
            #E = heslo
            worksheet['E' + str(row_to_print)] = password
            #F = poznamky
            worksheet['F' + str(row_to_print)] = notes

            workbook.save(filename = excel_file_path)
            if wb_given == None:
                workbook.close()
        
        @classmethod
        def read_excel_data(cls,excel_file_path):
            """
            Returns:
            - [disk_all_rows,disk_project_list,default_disk_status_behav]
            """
            workbook = load_workbook(excel_file_path,read_only=True)
            # seznam vsech ftp pripojeni k diskum
            disk_all_rows = []
            disk_project_list = []  
            worksheet = workbook["disk_list"]
            for row in worksheet.iter_rows(values_only=True):
                row_array = []
                for items in row[:6]:
                    if items is None:
                        row_array.append("")
                    else:
                        row_array.append(str(items))
                disk_project_list.insert(0,row_array[0])
                disk_all_rows.insert(0,row_array)

            # ukladani nastavenych hodnot
            worksheet = workbook["Settings"]
            default_disk_status_behav = int(worksheet['B' + str(6)].value)
            workbook.close()

            return [disk_all_rows,disk_project_list,default_disk_status_behav]

    class IP_tools:
        @classmethod
        def save_excel_data(cls,
                            excel_file_path,
                            len_of_excel_array,
                            last_project_id,
                            show_favorite_status,
                            project_name,
                            IP_adress,
                            mask,
                            notes,
                            only_edit = None,
                            force_row_to_print=None,
                            fav_status = None,
                            force_ws = None,
                            wb_given = None):
            if wb_given == None:
                workbook = load_workbook(excel_file_path)
            else:
                workbook = wb_given

            if show_favorite_status:
                excel_worksheet = "ip_address_fav_list"
            else:
                excel_worksheet = "ip_address_list"
            if force_ws != None:
                excel_worksheet = force_ws
            worksheet = workbook[excel_worksheet]

            # excel je od jednicky...
            if force_row_to_print == None:
                row_to_print = len_of_excel_array +1
                if only_edit != None:
                    #pouze změna na temtýž řádku
                    row_to_print = (len_of_excel_array- last_project_id)
            else:
                row_to_print = force_row_to_print
            if notes == None or notes.replace(" ","") == "":
                notes = ""
            worksheet['A' + str(row_to_print)] = project_name #A = nazev projektu
            worksheet['B' + str(row_to_print)] = IP_adress #B = ip adresa
            worksheet['C' + str(row_to_print)] = mask #C = maska
            worksheet['D' + str(row_to_print)] = notes #D = poznamky
            if fav_status != None:
                if fav_status == True:
                    fav_status = 1
                worksheet['E' + str(row_to_print)] = fav_status #E = oblibenost
            else:
                worksheet['E' + str(row_to_print)] = 0

            workbook.save(filename=excel_file_path)
            if wb_given == None:
                workbook.close()

        @classmethod
        def read_excel_data(cls,excel_file_path,show_favorite_status):
            """
            returns:
            - [0] all_rows_ip list
            - [1] project_list
            - [2] favorite_list
            """
            if show_favorite_status:
                excel_worksheet = "ip_address_fav_list"
            else:
                excel_worksheet = "ip_address_list"
            workbook = load_workbook(excel_file_path,read_only=True)
            # seznam vsech statickych ip adres
            all_rows_ip = []
            project_list = []
            favourite_list = []
            worksheet = workbook[excel_worksheet]
            for row in worksheet.iter_rows(values_only=True):
                row_array = []
                for items in row[:4]:
                    if items is None:
                        row_array.append("")
                    else:
                        row_array.append(str(items))
                if len(row_array) < 4:
                    row_array.append("")
                project_list.insert(0,row_array[0])
                all_rows_ip.insert(0,row_array)
                for items in row[4:5]:
                    favourite_list.insert(0,items)
            workbook.close()
            return [all_rows_ip,project_list,favourite_list]

        @classmethod
        def get_current_ip_list(cls,connection_option_list):
            def get_current_ip_address(interface_name):
            # Get network interfaces and their addresses
                addresses = psutil.net_if_addrs()
                # Check if the specified interface exists
                if interface_name in addresses:
                    addr_count = 0
                    # print(f"Adresy interfacu: {interface_name}")
                    for addr in addresses[interface_name]:
                        # prvni AF_INET je pridelena automaticky, druha je privatni, nastavena DHCP
                        if addr.family == socket.AF_INET:  # IPv4 address
                            if addr_count == 1:
                                return addr.address
                            addr_count +=1
                            remembered_addr = addr
                    if addr_count == 1:
                        if "AddressFamily.AF_INET:" in str(remembered_addr):
                            return remembered_addr.address
                        else:
                            return "Nenalezeno"
                else:
                    return "Nenalezeno"
            current_address_list = []
            for items in connection_option_list:
                found_address = get_current_ip_address(items)
                current_address_list.append(found_address)
            return current_address_list
        
        @classmethod
        def fill_interfaces(cls):
            """
            Vrátí:
            - seznam interfaců [0]
            - seznam připojených interfaců [1]
            """
            process = subprocess.Popen("netsh interface show interface",
                                                        stdout=subprocess.PIPE,
                                                        stderr=subprocess.PIPE,
                                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            try:
                stdout_str = stdout.decode('utf-8')
                data = str(stdout_str)
            except UnicodeDecodeError:
                try:
                    stdout_str = stdout.decode('cp1250')
                    data = str(stdout_str)
                except UnicodeDecodeError:
                    data = str(stdout)
                
            lines = data.strip().splitlines()
            interfaces = []
            interface_statuses =[]
            # Process each line starting from the second line
            for line in lines[2:]:  # Skip the first two lines (headers and separator)
                # Split the line based on spaces
                values = line.split()
                if len(values) >= 4:  # Ensure there are enough columns
                    # Combine the last columns into Interface Name, handle spaces in the name
                    interface_name = ' '.join(values[3:])
                    interface_statuses.append(values[1])
                    interfaces.append(interface_name)

            print("interface list: ",interfaces)
            print("status: ", interface_statuses)
            connected_interfaces =[]
            for i in range(0,len(interface_statuses)):
                if interface_statuses[i] != "Odpojen" and interface_statuses[i] != "Odpojeno" and interface_statuses[i] != "Disconnected":
                    if interfaces[i] not in connected_interfaces:
                        connected_interfaces.append(interfaces[i])
            print("online: ", connected_interfaces)
            
            return [interfaces,connected_interfaces]

        @classmethod
        def get_ipv4_addresses(cls):
            """
            returns list of ipv4 addresses of all interfaces
            """
            process = subprocess.Popen("ipconfig",
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            result2 = "" 
            try:
                stdout_str = stdout.decode('cp1250')
                result2 = str(stdout_str)
                
            except Exception as e:
                print("chyba ",e)

            # Regular expression to match the IPv4 address
            ipv4_pattern = re.compile(r'IPv4 Address[.\s]*: ([\d.]+)')
            ipv4_addresses = []
            lines = result2.splitlines()
            current_interface = None
            # Iterate over each line to find interface names and IPv4 addresses
            for line in lines:
                if line.strip():
                    # Detect interface name
                    if line[0].isalpha():
                        current_interface = line.strip()
                    else:
                        # Detect IPv4 address for the current interface
                        match = ipv4_pattern.search(line)
                        if match and current_interface:
                            ipv4_addresses.append(current_interface)
                            ipv4_addresses.append(match.group(1))
            return ipv4_addresses

        @classmethod
        def check_DHCP(cls,interface):
            process = subprocess.Popen(f'netsh interface ip show config name="{interface}"',
                                                        stdout=subprocess.PIPE,
                                                        stderr=subprocess.PIPE,
                                                        creationflags=subprocess.CREATE_NO_WINDOW)
            stdout, stderr = process.communicate()
            try:
                stdout_str = stdout.decode('utf-8')
                output_data = str(stdout_str)
            except UnicodeDecodeError:
                try:
                    stdout_str = stdout.decode('cp1250')
                    output_data = str(stdout_str)
                except UnicodeDecodeError:
                    output_data = str(stdout)

            output_data_lines = output_data.split("\n")
            for lines in output_data_lines:
                if "DHCP enabled" in lines and "Yes" in lines:
                    print(f"{interface} DHCP: yes")
                    return True
            print(f"{interface} DHCP: no")

        @classmethod
        def is_project_favourite(cls,favourite_list,array_index):
            try:
                fav_status = int(favourite_list[array_index])
                if fav_status == 1:
                    return True
                else:
                    return False
                
            except Exception:
                return False

    def __init__(self,root,menu_callback_function,window_mode,initial_path,zoom_factor):
        self.root = root
        self.menu_callback = menu_callback_function
        self.initial_path = initial_path
        self.window_mode = window_mode
        self.zoom_factor = zoom_factor
        self.show_favourite_ip = False
        self.excel_file_path = initial_path + "config_TRIMAZKON.xlsx"
        self.app_icon = Tools.resource_path('images\\logo_TRIMAZKON.ico')
        self.default_environment = self.check_default_env()
        if self.default_environment == "disk":
            self.Disk_management_gui(self)
        else:
            self.IP_assignment(self)
        # ip_instance = main.IP_assignment(self.root,self.menu_callback,self.window_mode,self.initial_path)
        # disk_instance = main.Disk_management_gui(self.root,self.menu_callback,self.window_mode,self.initial_path)

        # if self.default_environment == "disk":
        #     disk_instance.ip_instance = ip_instance
        #     disk_instance.create_widgets_disk(init=True,disk_instance=disk_instance)

        # elif self.default_environment == "config_load_error":
        #     ip_instance.disk_instance = disk_instance
        #     ip_instance.create_widgets(init=True,excel_load_error=True,ip_instance=ip_instance)     
        # else:
        #     ip_instance.disk_instance = disk_instance
        #     ip_instance.create_widgets(fav_status=self.show_favourite_ip,init=True,ip_instance=ip_instance)

    def check_default_env(self):
        try:
            workbook = load_workbook(self.excel_file_path)
            worksheet = workbook["Settings"]
            def_environment = worksheet['B' + str(4)].value
            if int(def_environment) == 1:
                def_environment = "disk"

            def_show_favourite = worksheet['B' + str(3)].value
            if int(def_show_favourite) == 1:
                self.show_favourite_ip = True
            else:
                self.show_favourite_ip = False

            workbook.save(self.excel_file_path) #check if it is opened currently
            workbook.close()
            return def_environment
        
        except Exception as e:
            print(f"Nejprve zavřete soubor {self.excel_file_path} Chyba: {e}")
            return "config_load_error"

    class Disk_management_gui:
        def __init__(self,parent):
            self.parent_instance = parent
            self.root = parent.root
            self.menu_callback = parent.menu_callback
            self.window_mode = parent.window_mode
            self.excel_file_path = parent.excel_file_path
            self.app_icon = parent.app_icon
            self.disk_all_rows = []
            self.disk_project_list = []
            self.bin_projects = [[None],[None]]
            self.last_project_id = ""
            self.opened_window = ""
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_disk_letter = ""
            self.last_project_ftp = ""
            self.last_project_username = ""
            self.last_project_password = ""
            self.last_inserted_password = ""
            self.last_selected_widget_id = 0
            self.changed_notes_disk = []
            self.selected_list_disk = []
            self.remember_to_change_back = []
            self.notes_frame_height = 50
            read_parameters = Tools.read_setting_parameters(self.excel_file_path)
            if read_parameters != None:
                if read_parameters[3] == 2:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                self.default_disk_status_behav = read_parameters[4]
                self.default_note_behav = read_parameters[5]
                self.mapping_condition = read_parameters[6]
                if read_parameters[7] == 1:
                    self.make_edited_project_first = True
                else:
                    self.make_edited_project_first = False
                self.deletion_behav = read_parameters[8]
            else:
                self.default_disk_status_behav = 0
                self.default_note_behav = 0
                self.mapping_condition = 0
                self.make_edited_project_first = True
                self.deletion_behav = 100

            self.create_widgets_disk(init=True)

        def call_menu(self): # Tlačítko menu (konec, návrat do menu)
            """
            Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
            """

            Tools.clear_frame(self.main_widgets)
            Tools.clear_frame(self.root)
            # self.root.unbind("<f>")
            self.root.unbind("<Escape>")
            self.root.unbind("<F5>")
            self.root.unbind("<Button-1>")
            self.root.unbind("<Control_L>")
            self.root.unbind("<Control-Button-1>")
            self.root.unbind("<KeyRelease-Control_L>")
            self.root.unbind("<Delete>")
            self.root.update()
            self.root.update_idletasks()
            self.menu_callback()
        
        def clicked_on_project(self,event,widget_id,widget,textbox = "",flag = ""):
            """
            flag = notes:
            - při nakliknutí poznámky zůstanou expandnuté a při kliku na jinou je potřeba předchozí vrátit zpět
            flag = unfocus:
            - při kliku mimo se odebere focus z nakliknutých widgetů
            """
            def on_leave_entry(widget,row_of_widget):
                """
                při kliku na jiný widget:
                - upraví text pouze na první řádek
                """
                widget.configure(state = "normal")
                if "\n" in self.disk_all_rows[row_of_widget][5]:
                    notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                    first_row = notes_rows[0]
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(first_row))
                if self.default_note_behav == 0:
                    widget.configure(state = "disabled")

            def shrink_frame(widget_frame,widget_notes):
                widget_notes.configure(state = "normal")
                new_height = self.notes_frame_height
                widget_frame.configure(height = new_height) #frame
                widget_notes.configure(height = new_height-10) #notes
                if self.default_note_behav == 0:
                    widget_notes.configure(state = "disabled")

            if flag == "unfocus":
                try:
                    if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                        if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                            shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                            self.last_selected_textbox = ""
                            self.last_selected_notes_widget = ""

                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        self.last_selected_widget.configure(border_color="#636363")
                        self.last_selected_widget = ""

                    for frame_and_id in self.remember_to_change_back:
                        if frame_and_id[0].winfo_exists(): 
                            frame_and_id[0].configure(border_color="#636363")
                    self.selected_list_disk = []
                    self.remember_to_change_back = []

                except Exception as e:
                    print("chyba při odebírání focusu: ",e)
                return

            if widget_id == None:
                return
            print("widget_id",widget_id)
            self.search_input.delete("0","300")
            self.search_input.insert("0",str(self.disk_all_rows[widget_id][0]))
            self.check_given_input()
            # only if it is not pressed againt the same:
            if widget != self.last_selected_widget:
                try:
                    if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                        if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                            shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                    if flag == "notes":
                        self.last_selected_textbox = textbox
                        self.last_selected_notes_widget = widget
                        widget.focus_set()
                    else:
                        # init the values:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""
                        widget.focus_set()

                except Exception as e:
                    print("chyba s navracenim framu do puvodniho formatu",e)
                
                try:
                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        if len(self.selected_list_disk) == 0 and not self.control_pressed:
                            self.last_selected_widget.configure(border_color="#636363")

                            if [self.last_selected_widget,self.last_selected_widget_id] in self.remember_to_change_back:
                                self.remember_to_change_back.pop(self.remember_to_change_back.index([self.last_selected_widget,self.last_selected_widget_id]))

                        # pokud došlo k další interakci s jiným widgeten
                        elif not self.control_pressed:
                            for frame_and_id in self.remember_to_change_back:
                                if frame_and_id[0].winfo_exists(): 
                                    frame_and_id[0].configure(border_color="#636363")
                            self.selected_list_disk = []
                            self.remember_to_change_back = []

                    self.last_selected_widget = widget
                    widget.configure(border_color="white")

                    if not [widget,widget_id] in self.remember_to_change_back:
                        self.remember_to_change_back.append([widget,widget_id])

                    print("remember: ", self.remember_to_change_back)

                except Exception as e:
                    print("chyba pri zmene fucusu",e)
                    pass

                self.last_selected_widget_id = widget_id

        def refresh_disk_statuses(self,silent=True):
            online_disks = []
            offline_disks = []
            self.refresh_btn.configure(text = "🔄",font=("",25))
            self.refresh_btn.update()
            self.refresh_btn.update_idletasks()

            def refresh_thread():
                mapped_disks = main.DM_tools.list_mapped_disks(whole_format = True)
                non_persistant_disks = main.DM_tools.list_non_persistent_disks()
                for y in range(0,len(self.disk_letter_frame_list)):
                    param_frame = self.disk_letter_frame_list[y]
                    param_frame.configure(fg_color = "black") # <= init

                    for i in range(0,len(non_persistant_disks)):
                        if non_persistant_disks[i][0:1] == str(self.disk_all_rows[y][1]):
                            drive_status = main.DM_tools.check_network_drive_status(non_persistant_disks[i])
                            if drive_status == True:
                                online_disks.append(non_persistant_disks[i][0:1])
                                param_frame.configure(fg_color = "#00CED1")
                            else:
                                offline_disks.append(non_persistant_disks[i][0:1])
                                param_frame.configure(fg_color = "red")

                    for i in range(0,len(mapped_disks)):
                        if mapped_disks[i][0:1] == str(self.disk_all_rows[y][1]):
                            drive_status = main.DM_tools.check_network_drive_status(mapped_disks[i])
                            if drive_status == True:
                                online_disks.append(mapped_disks[i][0:1])
                                param_frame.configure(fg_color = "green")
                            else:
                                offline_disks.append(mapped_disks[i][0:1])
                                try:
                                    param_frame.configure(fg_color = "red")
                                except Exception:
                                    pass
                if len(mapped_disks) == 0 and len(non_persistant_disks) == 0 and silent == False:
                    Tools.add_colored_line(self.main_console,f"Nejsou namapované žádné disky","red",None,True)
                elif silent == False:
                    if len(online_disks) != 0 and len(offline_disks) != 0:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: online: {list(set(online_disks))}, offline: {list(set(offline_disks))}","white",None,True)
                    elif len(online_disks) == 0 and  len(offline_disks) != 0:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: offline: {list(set(offline_disks))}","white",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Namapované disky: online: {list(set(online_disks))}","white",None,True)

                self.refresh_btn.configure(text = "Refresh statusů",font=("Arial",20,"bold"))
            
            run_backgroung = threading.Thread(target=refresh_thread,)
            run_backgroung.start()

        def manage_bin(self,flag="",parameters=[],wb=None):
            """
            First_row in bin worksheet = last deleted ip\n
            Second_row in bin worksheet = last deleted disk\n
            self.bin_projects = [0] deleted projects (disk)\n
            self.bin_projects = [1] edited projects (disk)\n
            flag:\n
            - read_sheet
            - save_project_disk
            - load_deleted_disk
            - save_edited_disk
            - load_edited_disk
            - change_notes_back_disk
            """
            bin_worksheet = "projects_bin2"
            
            def read_sheet():
                wb = load_workbook(self.excel_file_path)
                if not bin_worksheet in wb.sheetnames:
                    ws = wb.create_sheet(title=bin_worksheet)
                    ws.sheet_state = "hidden"
                    wb.save(self.excel_file_path)
                    wb.close()
                    print("adding new bin sheet to excel")
                    return [[None],[None]]
                else:
                    ws = wb[bin_worksheet]
                    row_data_disk = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
                    wb.close()
                    return [list(row_data_disk),[None]] # provedene zmeny v editu pri spusteni programu nanacitam. Jen smazané projekty...
                
            def save_project_disk():
                nonlocal wb
                if wb == None:
                    return False
                main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                        len(self.disk_all_rows),
                                                        self.last_project_id,
                                                        parameters[0],
                                                        parameters[1],
                                                        parameters[2],
                                                        parameters[3],
                                                        parameters[4],
                                                        parameters[5],
                                                        force_row_to_print=2,
                                                        force_ws=bin_worksheet,
                                                        wb_given=wb)
                self.bin_projects[0] = [parameters[0],parameters[1],parameters[2],parameters[3],parameters[4],parameters[5]]
                self.undo_button.configure(state = "normal")
            
            def save_edited_disk():
                self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="load_edited_disk"))
                main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                len(self.disk_all_rows),
                                                self.last_project_id,
                                                parameters[0],
                                                parameters[1],
                                                parameters[2],
                                                parameters[3],
                                                parameters[4],
                                                parameters[5],
                                                force_row_to_print=4,
                                                force_ws=bin_worksheet,
                                                wb_given=wb)
                self.bin_projects[1] = [parameters[0],parameters[1],parameters[2],parameters[3],parameters[4],parameters[5]]

            def load_deleted_disk():
                """
                adds new project from history and deletes the history
                """
                wb = load_workbook(self.excel_file_path)
                ws = wb[bin_worksheet]
                row_data_disk = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

                print("\nrow data disk: ",row_data_disk,"\n")

                project_name = row_data_disk[0]
                if project_name in self.disk_project_list:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {project_name} je již používané, nelze ho tedy obnovit","red",None,True)
                    wb.close()
                    return
                
                self.bin_projects[0] = []
                if len(row_data_disk) <6:
                    notes = ""
                else:
                    notes = row_data_disk[5]
                self.undo_button.configure(state = "disabled")
                ws.delete_rows(2)
                wb.save(self.excel_file_path)
                wb.close()
                main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                len(self.disk_all_rows),
                                                self.last_project_id,
                                                project_name,
                                                row_data_disk[1],
                                                row_data_disk[2],
                                                row_data_disk[3],
                                                row_data_disk[4],
                                                notes)
                Tools.add_colored_line(self.main_console,f"Projekt: {project_name} byl úspěšně obnoven","green",None,True)
                self.make_project_cells_disk()

            def change_notes_back_disk():
                print("loading back: ",self.changed_notes_disk)    
            
                def save_changed_notes(notes,row):
                    workbook = load_workbook(self.excel_file_path)
                    worksheet = workbook["disk_list"]
                    worksheet['F' + str(len(self.disk_all_rows)-row)] = notes
                    workbook.save(filename=self.excel_file_path)
                    workbook.close()
                
                project_name = self.changed_notes_disk[0]
                notes_before = self.changed_notes_disk[1]
                id = None
                for i in range(0,len(self.disk_all_rows)):
                    if self.disk_all_rows[i][0] == project_name:
                        id = i

                save_changed_notes(notes_before,id)
                Tools.add_colored_line(self.main_console,f"Poznámky u projektu: {project_name} byly úspěšně obnoveny","green",None,True)
                self.make_project_cells_disk()

                if Tools.get_none_count(self.bin_projects[1]) < 4 and len(self.bin_projects[1]) == 6:
                    self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="load_edited_disk"))
                else:
                    self.undo_edit.configure(state = "disabled")
                self.changed_notes_disk = []

            def load_edited_disk():
                wb = load_workbook(self.excel_file_path)
                ws = wb[bin_worksheet]
                not_edited_data_disk = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

                print("\nrow data disk: ",not_edited_data_disk,"\n")
                if self.edited_project_name_disk not in self.disk_project_list:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {self.edited_project_name_disk} nenalezeno, nelze ho tedy obnovit","red",None,True)
                    wb.close()
                    return
                
                self.bin_projects[1] = []
                self.undo_edit.configure(state = "disabled")
                ws.delete_rows(3)
                wb.save(self.excel_file_path)
                wb.close()

                param = [not_edited_data_disk[0],not_edited_data_disk[1],not_edited_data_disk[2],not_edited_data_disk[3],not_edited_data_disk[4],not_edited_data_disk[5]]
                for i in range(0,len(param)):
                    if param[i] == None:
                        param[i] = ""

                id = None
                for i in range(0,len(self.disk_all_rows)):
                    if self.edited_project_name_disk == self.disk_all_rows[i][0]:
                        id = i
                
                main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                len(self.disk_all_rows),
                                                self.last_project_id,
                                                param[0],
                                                param[1],
                                                param[2],
                                                param[3],
                                                param[4],
                                                param[5],
                                                force_row_to_print=len(self.disk_all_rows)-id)

                if self.edited_project_name_disk != param[0]:
                    Tools.add_colored_line(self.main_console,f"U projektu: {self.edited_project_name_disk} (původně: {param[0]}) byly odebrány provedené změny","green",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"U projektu: {self.edited_project_name_disk} byly odebrány provedené změny","green",None,True)
                self.make_project_cells_disk()

            mapping_logic = {
                "read_sheet": read_sheet,
                "save_project_disk": save_project_disk,
                "load_deleted_disk": load_deleted_disk,
                "save_edited_disk": save_edited_disk,
                "load_edited_disk": load_edited_disk,
                "change_notes_back_disk": change_notes_back_disk,
            }

            output = mapping_logic[flag]()  # This will call the corresponding function
            return output
        
        def check_given_input(self,given_data = None):
            """
            Fills all parameters of last project
            """
            if given_data == None:
                given_data = self.search_input.get()
            if given_data == "":
                found = None
                return found
            found = False

            for i in range(0,len(self.disk_all_rows)):
                if given_data == self.disk_all_rows[i][0]:
                    self.last_project_name =        str(self.disk_all_rows[i][0])
                    self.last_project_disk_letter = str(self.disk_all_rows[i][1])
                    self.last_project_ftp =         str(self.disk_all_rows[i][2])
                    self.last_project_username =    str(self.disk_all_rows[i][3])
                    self.last_project_password =    str(self.disk_all_rows[i][4])
                    self.last_project_notes =       str(self.disk_all_rows[i][5])
                    self.last_project_id = i
                    found = True
                
            return found    

        def refresh_explorer(self,refresh_disk=None):
            """
            Resetuje windows explorer přes cmd
            refresh_disk = udelat nove všechni widgets (make_project_cells_disk())
            """
            refresh_explorer="taskkill /f /im explorer.exe"
            subprocess.run(refresh_explorer, shell=True)
            refresh_explorer="start explorer.exe"
            subprocess.run(refresh_explorer, shell=True)
            if refresh_disk:
                self.make_project_cells_disk(disk_statuses=True)

        def delete_disk_option_menu(self):
            def delete_disk(child_root):
                drive_letter = str(self.drive_letter_input.get())
                if len(str(self.DL_manual_entry.get())) > 0:
                    drive_letter = str(self.DL_manual_entry.get())
                
                delete_command = "net use " + drive_letter +": /del"
                process = subprocess.Popen(delete_command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
                stdout, stderr= process.communicate()
                if "Is it OK to continue disconnecting and force them closed?" in stdout:
                    Tools.add_colored_line(self.main_console,f"Disk je právě používán, nejprve jej zavřete","red",None,True)
                    child_root.destroy()
                else:
                    self.refresh_explorer()
                    Tools.add_colored_line(self.main_console,f"Disky s označením {drive_letter} byly odpojeny","orange",None,True)
                    self.refresh_disk_statuses()
                    child_root.destroy()

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"450x190+{x+250}+{y+200}")
            child_root.after(200, lambda: child_root.iconbitmap(self.app_icon))
            child_root.title("Odpojování síťového disku")
            
            found_drive_letters=[]
            for i in range(0,len(self.disk_all_rows)):
                if not self.disk_all_rows[i][1] in found_drive_letters:
                    found_drive_letters.append(self.disk_all_rows[i][1])

            mapped_disks = main.DM_tools.list_mapped_disks()
            non_persistent_disks = main.DM_tools.list_non_persistent_disks()
            for disk in non_persistent_disks:
                if not disk in mapped_disks:
                    mapped_disks.append(disk)

            for i in range(0,len(mapped_disks)):
                if not mapped_disks[i] in found_drive_letters:
                    found_drive_letters.append(mapped_disks[i])

            label =                     customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Vyberte disk nebo vyhledejte manuálně: ",font=("Arial",20,"bold"))
            self.drive_letter_input =   customtkinter.CTkOptionMenu(master = child_root,font=("Arial",20),width=200,height=30,values=found_drive_letters,corner_radius=0)
            self.DL_manual_entry =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0,placeholder_text="manuálně")
            del_button =                customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Odpojit", command = lambda: delete_disk(child_root),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
            exit_button =               customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            label.                      grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
            self.drive_letter_input.    grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
            self.DL_manual_entry.       grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
            del_button.                 grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
            exit_button.                grid(column = 0,row=3,pady = 5,padx =220,sticky = tk.W)
            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

            print("selected: ",self.last_project_disk_letter)
            print("disk letter list: ",found_drive_letters)
            try:
                self.drive_letter_input.set(self.last_project_disk_letter)
            except Exception:
                pass

        def delete_project_disk(self,button_trigger = False,wanted_project=None,flag=""):
            project_found = False
            name_list = []

            def check_multiple_projects(window):
                nonlocal wanted_project
                nonlocal name_list
                nonlocal project_found

                if len(self.selected_list_disk) > 1:
                    for names in name_list:
                        print(names)
                        project_found = False
                        self.disk_all_rows, self.disk_project_list, self.default_disk_status_behav = main.DM_tools.read_excel_data(self.excel_file_path)
                        proceed(names,window,True)
                            
                    Tools.add_colored_line(self.main_console,f"Byly úspěšně odstraněny tyto projekty: {name_list}","orange",None,True)
                    try:
                        self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                    except Exception as e:
                        print("chyba, refresh po mazani")
                else:
                    proceed(wanted_project,window)

            def proceed(wanted_project, window = True, multiple_status = False):
                nonlocal project_found
                nonlocal child_root
                # nonlocal wanted_project

                if wanted_project == None:
                    self.disk_all_rows, self.disk_project_list, self.default_disk_status_behav = main.DM_tools.read_excel_data(self.excel_file_path)
                    wanted_project = str(self.search_input.get())
                workbook = load_workbook(self.excel_file_path)
                for i in range(0,len(self.disk_project_list)):
                    if self.disk_project_list[i] == wanted_project and len(str(self.disk_project_list[i])) == len(str(wanted_project)):
                        row_index = self.disk_project_list.index(wanted_project)
                        row_data = self.disk_all_rows[row_index]
                        self.manage_bin(flag="save_project_disk",parameters=row_data,wb=workbook)
                        worksheet = workbook["disk_list"]
                        worksheet.delete_rows(len(self.disk_all_rows)-row_index)
                        workbook.save(self.excel_file_path)
                        project_found = True
                        break

                workbook.close()
                if not multiple_status:
                    if project_found:
                        Tools.add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)    
                        self.make_project_cells_disk() #refresh = cele zresetovat, jine: id, poradi...
                    elif wanted_project.replace(" ","") == "":
                        Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)
                
                if window and child_root.winfo_exists():
                    child_root.grab_release()
                    child_root.destroy()

            if not button_trigger:
                proceed(wanted_project,window=False)
                return

            if flag == "main_menu" or flag == "context_menu":
                if self.deletion_behav == 110 or self.deletion_behav == 111:
                    check_multiple_projects(False)
                    return
                
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                check_multiple_projects(False)
                return

            if self.last_project_name.replace(" ","") == "":
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif wanted_project == None:
                wanted_project = self.last_project_name
            
            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root

            # child_root.geometry(f"650x130+{x+80}+{y+150}")
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Upozornění")
            proceed_label_text = f"Opravdu si přejete odstranit projekt {self.last_project_name}?"
            if flag == "context_menu":
                self.selected_list_disk = []
            if len(self.selected_list_disk) > 1:
                for ids in self.selected_list_disk:
                    name_list.append(self.disk_all_rows[ids][0])
                proceed_label_text = f"Opravdu si přejete odstranit vybrané projekty:\n{name_list}?"
            proceed_label = customtkinter.CTkLabel(master = child_root,text = proceed_label_text,font=("Arial",22,"bold"),justify = "left",anchor="w")
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: check_multiple_projects(True))
            button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
            proceed_label   .pack(pady=(15,0),padx=10,side = "top",fill="x")
            button_no       .pack(pady = 5, padx = 10,anchor="w",side="right",expand = False)
            button_yes      .pack(pady = 5, padx = 10,anchor="w",side="right",expand = False)
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
            child_root.update()
            child_root.update_idletasks()
            child_root.grab_set()
            child_root.focus()
            child_root.focus_force()
            child_root.wait_window()
            return project_found

        def save_new_project_data_disk(self,child_root,only_edit = None):
            project_name = str(self.name_input.get())
            disk_letter =  str(self.disk_letter_input.get())
            ftp_address =  str(self.FTP_adress_input.get())
            username =     str(self.username_input.get())
            # password =     str(self.password_input.get())
            password =     str(self.last_inserted_password)
            notes = Tools.get_legit_notes(self.notes_input.get("1.0", tk.END))
            errors = 0
            if project_name.replace(" ","") == "":
                Tools.add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
                errors += 1
            if project_name in self.disk_project_list and only_edit == None:
                Tools.add_colored_line(self.console,f"Jméno je již používané","red",None,True)
                errors +=1
            elif disk_letter.replace(" ","") == "":
                Tools.add_colored_line(self.console,f"Nezadali jste písmeno disku","red",None,True)
                errors += 1
            elif ftp_address.replace(" ","") == "":
                Tools.add_colored_line(self.console,f"Nezadali jste adresu","red",None,True)
                errors += 1
            
            # poznamky nejsou povinne
            if errors ==0:
                main.DM_tools.read_excel_data(self.excel_file_path)
                if only_edit == None:
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                            len(self.disk_all_rows),
                                                            self.last_project_id,
                                                            project_name,
                                                            disk_letter,
                                                            ftp_address,
                                                            username,
                                                            password,
                                                            notes)
                else:
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                            len(self.disk_all_rows),
                                                            self.last_project_id,
                                                            project_name,
                                                            disk_letter,
                                                            ftp_address,
                                                            username,
                                                            password,
                                                            notes,
                                                            only_edit=True)                
                child_root.destroy()
                if only_edit == None:
                    self.make_project_cells_disk()
                    Tools.add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)
                else: # edit - musi byt proveden reset
                    self.edited_project_name_disk = project_name
                    self.manage_bin("save_edited_disk",parameters=[self.last_project_name,self.last_project_disk_letter,self.last_project_ftp,
                                                                self.last_project_username,self.last_project_password,self.last_project_notes])
                    if self.make_edited_project_first:
                        self.make_project_first_disk(purpouse="silent",make_cells = False)
                    self.make_project_cells_disk()
                    if self.last_project_name != project_name:
                        status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn"
                    else:
                        status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn"
                    Tools.add_colored_line(self.main_console,status_text,"green",None,True)

        def add_new_project_disk(self,edit = None,init_copy=False):
            def mouse_wheel_change(e):
                if -e.delta < 0:
                    switch_up()
                else:
                    switch_down()

            def copy_previous_project():
                if self.last_project_name == "":
                    Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
                    return
                
                self.last_inserted_password = str(self.last_project_password)
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.last_project_name))
                self.disk_letter_input.delete("0","300")
                self.disk_letter_input.insert("0",str(self.last_project_disk_letter))
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0",str(self.last_project_ftp))
                self.username_input.delete("0","300")
                self.username_input.insert("0",str(self.last_project_username))
                self.password_input.delete("0","300")
                self.password_input.insert("0",str(len(self.last_project_password)*"*"))
                self.notes_input.delete("1.0",tk.END)
                self.notes_input.insert(tk.END,str(self.last_project_notes))
            
            def switch_up():
                print("up ",self.last_project_id)
                self.last_project_id -= 1
                if self.last_project_id < 0:
                    self.last_project_id = len(self.disk_all_rows)-1
                    
                self.check_given_input(given_data=self.disk_all_rows[self.last_project_id][0])
                copy_previous_project()
                refresh_title()

            def switch_down():
                print("down ",self.last_project_id)
                self.last_project_id += 1
                if self.last_project_id > len(self.disk_all_rows)-1:
                    self.last_project_id = 0

                self.check_given_input(given_data=self.disk_all_rows[self.last_project_id][0])
                copy_previous_project()
                refresh_title()

            def del_project():
                nonlocal child_root
                result = self.delete_project_disk(button_trigger=True)
                if result:
                    switch_up()
                else:
                    print("aborted")

                child_root.focus()
                child_root.focus_force()
                child_root.grab_set()
            
            def refresh_title():
                if edit == None:
                    child_root.title("Nový projekt")
                else:
                    child_root.title("Editovat projekt: "+self.last_project_name)

            def show_password():
                if str(self.password_input.get()) == str(len(self.last_inserted_password)*"*"):
                    self.password_input.configure(state = "normal")
                    self.password_input.delete("0","300")
                    self.password_input.insert("0",self.last_inserted_password)
                else:
                    self.last_inserted_password = str(self.password_input.get())
                    self.password_input.delete("0","300")
                    self.password_input.insert("0",str(len(self.last_inserted_password)*"*"))
                    self.password_input.configure(state = "disabled")

            child_root = customtkinter.CTkToplevel()
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"520x820+{x+50}+{y+100}")
            refresh_title()
            project_name =              customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
            project_selection_label =   customtkinter.CTkLabel(master = child_root, width = 200,height=30,text = "Přepnout projekt: ",font=("Arial",20,"bold"))
            project_switch_frame =      customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=140,width=80)
            project_up =                customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↑",command= lambda: switch_up())
            project_down =              customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↓",command= lambda: switch_down())
            project_switch_frame        .grid_propagate(0)
            project_up                  .grid(column = 0,row=0,pady = (5,0),padx =10)
            project_down                .grid(column = 0,row=1,pady = 5,padx =10)
            project_switch_frame.       bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_up.                 bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_down.               bind("<MouseWheel>",lambda e: mouse_wheel_change(e))

            copy_check =                customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: copy_previous_project())
            del_project_btn =           customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Smazat tento projekt",command= lambda: del_project(),fg_color="red")
            self.name_input =           customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            disk_letter =               customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Písmeno disku: ",font=("Arial",20,"bold"))
            self.disk_letter_input =    customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            FTP_adress =                customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "ftp adresa: ",font=("Arial",20,"bold"))
            self.FTP_adress_input =     customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=500,height=30,corner_radius=0)
            user =                      customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Uživatelské jméno: ",font=("Arial",20,"bold"))
            self.username_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            password =                  customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Heslo: ",font=("Arial",20,"bold"))
            self.password_input =       customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=170,height=30,corner_radius=0)
            show_pass_btn =             customtkinter.CTkButton(master = child_root,font=("Arial",15),width=30,height=30,corner_radius=0,text="👁",command= lambda: show_password(),fg_color="#505050",hover_color="#404040")
            notes =                     customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
            self.notes_input =          customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=260)
            self.console =              tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
            if edit == None:
                save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disk(child_root),font=("Arial",20,"bold"),corner_radius=0)
            else:
                save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data_disk(child_root,True),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
            if edit != True:
                copy_check.             grid(column = 0,row=10,pady = 5,padx =240,sticky = tk.W)
            if edit:
                project_selection_label.grid(column = 0,row=0,pady = 5,padx =265,sticky = tk.W)
                project_switch_frame.grid(column = 0,row=1,pady = 5,padx =320,sticky = tk.W,rowspan = 4)
                del_project_btn.    grid(column = 0,row=10,pady = 5,padx =240,sticky = tk.W)
            self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
            disk_letter.            grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
            self.disk_letter_input. grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
            FTP_adress.             grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
            self.FTP_adress_input.  grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
            user.                   grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
            self.username_input.    grid(column = 0,row=7,pady = 5,padx =10,sticky = tk.W)
            password.               grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
            self.password_input.    grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
            show_pass_btn.          grid(column = 0,row=9,pady = 5,padx =180,sticky = tk.W)
            notes.                  grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
            self.notes_input.       grid(column = 0,row=11,pady = 5,padx =10,sticky = tk.W)
            self.console.           grid(column = 0,row=12,pady = 5,padx =10,sticky = tk.W)
            save_button.            grid(column = 0,row=13,pady = 5,padx =100,sticky = tk.W)
            exit_button.            grid(column = 0,row=13,pady = 5,padx =310,sticky = tk.W)

            if init_copy: # kopírovat pro vytvoreni noveho projektu, neni edit...
                copy_previous_project()
                self.password_input.configure(state = "disabled")
            elif edit == None:
                self.last_inserted_password = ""
                self.disk_letter_input.delete("0","300")
                self.disk_letter_input.insert("0","P")
                self.FTP_adress_input.delete("0","300")
                self.FTP_adress_input.insert("0","\\\\192.168.000.000\\")
                self.username_input.delete("0","300")
                self.password_input.delete("0","300")
                if str(self.search_input.get()).replace(" ","") != "":
                    self.name_input.delete("0","300")
                    self.name_input.insert("0",str(self.search_input.get()))
            else:
                copy_previous_project()
                self.password_input.configure(state = "disabled")

            def update_last_password():
                if self.password_input.cget("state") == "normal":
                    self.last_inserted_password = str(self.password_input.get())
            self.password_input.bind("<Key>",lambda e: update_last_password())

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def make_project_first_disk(self,purpouse = None,make_cells = True,input_entry_bypass = None,project= None,upwards=False,downwards=False):
            def check_position():
                max_position = len(self.disk_all_rows)
                if upwards:
                    position = self.last_project_id -1
                elif downwards:
                    position = self.last_project_id +1

                if position < 0:
                    position = max_position-1
                elif position > max_position-1:
                    position = 0
                return position
            
            result = self.check_given_input(input_entry_bypass)
            self.remember_to_change_back = []
            self.last_selected_widget = ""

            if result == True:
                #zmena poradi
                if project == None:
                    project = self.disk_all_rows[self.last_project_id]

                if downwards or upwards:
                    position = check_position()
                else:
                    position = 0

                self.disk_all_rows.pop(self.last_project_id)
                self.disk_all_rows.insert(position,project)
                self.last_project_id = position

                for i in range(0,len(self.disk_all_rows)):
                    row = (len(self.disk_all_rows)-1)-i
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                len(self.disk_all_rows),
                                                self.last_project_id,
                                                self.disk_all_rows[i][0],
                                                self.disk_all_rows[i][1],
                                                self.disk_all_rows[i][2],
                                                self.disk_all_rows[i][3],
                                                self.disk_all_rows[i][4],
                                                self.disk_all_rows[i][5],
                                                force_row_to_print=row+1)
                if make_cells:
                    self.make_project_cells_disk()

                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} nalezen","green",None,True)
                elif purpouse != "silent":
                    Tools.add_colored_line(self.main_console,f"Projekt {self.disk_all_rows[0][0]} přesunut na začátek","green",None,True)
            elif result == None and purpouse != "silent":
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            elif purpouse != "silent":
                Tools.add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)

        def map_disk(self,button_row):
            Drive_letter = str(self.disk_all_rows[button_row][1])
            ftp_adress = str(self.disk_all_rows[button_row][2])
            user = str(self.disk_all_rows[button_row][3])
            password = str(self.disk_all_rows[button_row][4])

            delete_command = "net use " + Drive_letter + ": /del"
            subprocess.run(delete_command, shell=True)
            if self.mapping_condition == 1:
                persistant_status = " /persistent:yes"
            else:
                persistant_status =  " /persistent:no"

            if user != "" or password != "":
                second_command = "net use " + Drive_letter + ": " + ftp_adress + " " + password + " /user:" + user + persistant_status
            else:
                second_command = "net use " + Drive_letter + ": " + ftp_adress + persistant_status
            print("calling: ",second_command)


            def call_subprocess():
                nonlocal connection_status
                """process = subprocess.Popen(second_command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                stdout, stderr = process.communicate()
                self.connection_status = process.returncode
                print("STDOUT:", stdout)
                print("STDERR:", stderr)
                print("Return Code:", self.connection_status)"""
                connection_status = subprocess.call(second_command,shell=True,text=True)
            connection_status = None
            run_background = threading.Thread(target=call_subprocess,)
            run_background.start()

            time_start = time.time()
            while connection_status==None:
                time.sleep(0.05)
                if time.time() - time_start > 3:
                    print("terminated due to runtime error")
                    break

            if connection_status == 0:
                Tools.add_colored_line(self.main_console,f"Disk úspěšně připojen","green",None,True)
                self.refresh_explorer()
                self.refresh_disk_statuses()

                def open_explorer(path):
                    if os.path.exists(path):
                        os.startfile(path)
                    else:
                        print(f"The path {path} does not exist.")

                open_explorer(Drive_letter + ":\\")
            else:
                Tools.add_colored_line(self.main_console,f"Připojení selhalo (ixon? musí být zvolena alespoň 1 složka na disku...)","red",None,True)

        def show_context_menu(self,event,first_index,second_index,flag=""):
            """
            - first index (y) = index celeho radku
            - second index (x) = index jednoho parametru
            """
            context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="black",font=("Arial",20,"bold"))
            self.check_given_input(given_data=self.disk_all_rows[first_index][0])
            
            if flag == "button":
                context_menu.add_command(label="Namapovat",font=("Arial",22,"bold"),command=lambda: self.map_disk(first_index))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat FTP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(self.disk_all_rows[first_index][2]))
                context_menu.add_separator()
                context_menu.add_command(label="Editovat",font=("Arial",22,"bold"),command=lambda: self.add_new_project_disk(True))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat projekt",font=("Arial",22,"bold"),command=lambda: self.add_new_project_disk(init_copy=True))
                context_menu.add_separator()
                context_menu.add_command(label="Přesunout na začátek",font=("Arial",22,"bold"),command=lambda: self.make_project_first_disk(input_entry_bypass=self.disk_all_rows[first_index][0]))
                context_menu.add_separator()
                context_menu.add_command(label="Odstranit",font=("Arial",22,"bold"),command=lambda: self.delete_project_disk(button_trigger=True,flag="context_menu"))
            elif flag == "ftp_frame":
                context_menu.add_command(label="Kopírovat FTP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(self.disk_all_rows[first_index][2]))
            elif flag == "disk_letter":
                context_menu.add_command(label="Refresh",font=("Arial",22,"bold"), command=lambda: self.refresh_disk_statuses(silent=False))
                context_menu.add_separator()
                context_menu.add_command(label="Odpojit disk",font=("Arial",22,"bold"),command=lambda: self.delete_disk_option_menu())

            context_menu.tk_popup(event.x_root, event.y_root)

        def make_project_cells_disk(self,no_read = None,disk_statuses = False,init=False):
            def opened_window_check():
                if self.opened_window == "":
                    return False
                try:
                    if self.opened_window.winfo_exists():
                        return True
                    else:
                        return False
                except Exception as e:
                    print(e)
                    return False

            def filter_text_input(text):
                legit_rows = []
                legit_notes = ""
                rows = text.split("\n")
                for i in range(0,len(rows)):
                    if rows[i].replace(" ","") != "":
                        legit_rows.append(rows[i])

                for i in range(0,len(legit_rows)): 
                    if i == len(legit_rows)-1:
                        legit_notes = legit_notes + legit_rows[i]
                    else:
                        legit_notes = legit_notes + legit_rows[i]+ "\n"
                return legit_notes
            
            def save_changed_notes(notes,row):
                workbook = load_workbook(self.excel_file_path)
                worksheet = workbook["disk_list"]
                worksheet['F' + str(len(self.disk_all_rows)-row)] = notes
                workbook.save(filename=self.excel_file_path)
                workbook.close()

            def on_enter_entry(widget,row_of_widget):
                if not opened_window_check():
                    if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        widget.configure(state = "normal")
                        widget.delete("1.0",tk.END)
                        widget.insert(tk.END,str(self.disk_all_rows[row_of_widget][5]))
                        if self.default_note_behav == 0:
                            widget.configure(state = "disabled")

            def on_leave_entry(widget,row_of_widget):
                if not opened_window_check():
                    notes_before = filter_text_input(str(self.disk_all_rows[row_of_widget][5]))
                    notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
                    if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        widget.configure(state = "normal")
                        if notes_before != notes_after:
                            self.disk_all_rows[row_of_widget][5] = notes_after
                            self.changed_notes_disk = [self.disk_all_rows[row_of_widget][0],notes_before]
                            self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="change_notes_back_disk"))
                            save_changed_notes(notes_after,row_of_widget)

                        if "\n" in self.disk_all_rows[row_of_widget][5]:
                            notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                            first_row = notes_rows[0]
                            widget.delete("1.0",tk.END)
                            widget.insert(tk.END,str(first_row))
                        
                        if self.default_note_behav == 0:
                            widget.configure(state = "disabled")
                        self.root.focus_set() # unfocus widget
                    else:
                        # jinak pouze ulož změny
                        if notes_before != notes_after:
                            self.disk_all_rows[row_of_widget][5] = notes_after
                            self.changed_notes_disk = [self.disk_all_rows[row_of_widget][0],notes_before]
                            self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="change_notes_back_disk"))
                            save_changed_notes(notes_after,row_of_widget)
                        self.root.focus_set() # unfocus widget

            def shrink_frame(widget):
                tolerance = 5
                if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance:
                    return
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        widget[1].configure(state = "normal")
                        new_height = self.notes_frame_height
                        widget[0].configure(height = new_height)
                        widget[1].configure(height = new_height-10)
                        if self.default_note_behav == 0:
                            widget[1].configure(state = "disabled")

            def expand_frame(widget,row_of_widget):
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        tolerance = 5
                        if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance: # if the height is not 50 then it means it is expanded already
                            filtered_input = filter_text_input(self.disk_all_rows[row_of_widget][5])
                            self.disk_all_rows[row_of_widget][5] = filtered_input
                            addition = self.notes_frame_height
                            if "\n" in self.disk_all_rows[row_of_widget][5]:
                                notes_rows = self.disk_all_rows[row_of_widget][5].split("\n")
                                if len(notes_rows) > 1:
                                    expanded_dim = addition + (len(notes_rows)-1) * 25
                                    widget[0].configure(height = expanded_dim)
                                    widget[1].configure(state = "normal")
                                    widget[1].configure(height = expanded_dim-10)
                                if self.default_note_behav == 0:
                                    widget[1].configure(state = "disabled")

            def add_row_return(widget):
                addition = widget[0]._current_height
                expanded_dim = addition + 26
                widget[0].configure(height = expanded_dim)
                widget[1].configure(height = expanded_dim-10)

            if no_read == None:
                self.disk_all_rows, self.disk_project_list, self.default_disk_status_behav = main.DM_tools.read_excel_data(self.excel_file_path)

            Tools.clear_frame(self.project_tree)
            if self.default_disk_status_behav == 1:
                disk_statuses = True
            column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0,width = 50)
            column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column4 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column2_header =    customtkinter.CTkLabel(master = column2,text = "💾",font=("",22)) #💿
            column3_header =    customtkinter.CTkLabel(master = column3,text = "ftp adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column4_header =    customtkinter.CTkLabel(master = column4,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column2_header.     pack(padx = (12,0),side = "top",anchor = "w")
            column3_header.     pack(padx = (5,0),side = "top",anchor = "w",expand = False)
            column4_header.     pack(padx = (5,0),side = "top",anchor = "w")
            self.disk_letter_frame_list = []
            # y = widgets ve smeru y, x = widgets ve smeru x
            for y in range(0,len(self.disk_all_rows)):
                for x in range(0,len(self.disk_all_rows[y])):# x: 0=button, 1=disk_letter, 2=ip, 3=name, 4=password, 5=notes
                    if x == 0:
                        btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        button =    customtkinter.CTkButton(master = btn_frame,width=200,height=40,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),corner_radius=0)
                        button.     pack(padx =5,pady = 5, fill= "x")
                        btn_frame.  pack(side = "top",anchor = "w",expand = False)
                        button.     bind("<Button-1>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        button.     bind("<Double-1>",lambda e,widget_id = y: self.map_disk(widget_id))
                        button.     bind("<Button-3>",lambda e, first_index = y, second_index = x: self.show_context_menu(e,first_index,second_index,flag="button"))
                        # button.     bind("<Button-3>",lambda e,widget_id = y: self.map_disk(widget_id))

                    elif x == 1: # frame s písmenem disku, menší šířka, podbarvení
                        param_frame =   customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        parameter =     customtkinter.CTkLabel(master = param_frame,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),width = 40,height=40)
                        parameter.      pack(padx = (5,5),pady = 5)
                        param_frame.    pack(side = "top")
                        self.disk_letter_frame_list.append(param_frame)
                        param_frame.    bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter.      bind("<Button-1>",lambda e,widget = param_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter.      bind("<Button-3>",lambda e, first_index = y, second_index = x: self.show_context_menu(e,first_index,second_index,flag="disk_letter"))

                    elif x == 2: # frame s ftp adresou
                        param_frame2 =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        parameter2 =     customtkinter.CTkLabel(master = param_frame2,text = self.disk_all_rows[y][x],font=("Arial",20,"bold"),justify='left',anchor = "w",width = 300,height=40)
                        parameter2.      pack(padx = (10,5),pady = 5,anchor = "w",fill="x")
                        param_frame2.    pack(side = "top",fill="x",expand = False)
                        param_frame2.    bind("<Button-1>",lambda e,widget = param_frame2, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter2.      bind("<Button-1>",lambda e,widget = param_frame2, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter2.      bind("<Button-3>",lambda e, first_index = y, second_index = x: self.show_context_menu(e,first_index,second_index,flag="ftp_frame"))

                    elif x == 5: #frame s poznamkami...
                        notes_frame =   customtkinter.CTkFrame(master=column4,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),height=40,corner_radius=0,fg_color="black")
                        notes.          pack(padx =5,pady = 5,anchor="w",fill="x")
                        notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                        notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, widget_id = y: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))
                        notes.          bind("<Button-1>",lambda e,widget = notes_frame, textbox_widget = notes, widget_id = y: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))

                        if "\n" in self.disk_all_rows[y][x]:
                            notes_rows = self.disk_all_rows[y][x].split("\n")
                            first_row = notes_rows[0]
                            notes.delete("1.0",tk.END)
                            notes.insert(tk.END,str(first_row))
                        else:
                            notes.insert(tk.END,str(self.disk_all_rows[y][x]))
                        
                        notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(widget,row))
                        notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(widget))
                        notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(widget,row))
                        notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(widget,row))
                        notes.bind("<Return>",lambda e, widget = [notes_frame,notes]: add_row_return(widget))

                        if self.default_note_behav == 0:
                            notes.configure(state = "disabled")

                if y == self.last_project_id: # případ že posouvám s projektem nahoru/ dolů/ top (zvíraznit selectnuté)
                    self.selected_list_disk.append(y)
                    self.last_selected_widget = btn_frame
                    btn_frame.configure(border_color="white")
                    self.remember_to_change_back.append([btn_frame,y])
                    param_frame.configure(border_color="white")
                    self.remember_to_change_back.append([param_frame,y])
                    param_frame2.configure(border_color="white")
                    self.remember_to_change_back.append([param_frame2,y])
                    notes_frame.configure(border_color="white")
                    self.remember_to_change_back.append([notes_frame,y])

            column1.pack(fill="both",expand=False,side = "left")
            column2.pack(fill="both",expand=False,side = "left")
            column3.pack(fill="both",expand=False,side = "left")
            column4.pack(fill="both",expand=True, side = "left")
            self.project_tree.update()
            self.project_tree.update_idletasks()
            self.notes_frame_height = int(notes_frame._current_height)
            try:
                self.project_tree._parent_canvas.yview_moveto(0.0)
            except Exception:
                pass
            if disk_statuses:
                if init:
                    self.refresh_disk_statuses(False)
                else:
                    self.refresh_disk_statuses()

        def setting_window(self):
            def save_new_behav_disk():
                nonlocal checkbox2
                if int(checkbox2.get()) == 0:
                    self.default_disk_status_behav = 0
                    Tools.save_setting_parameter(parameter="change_def_disk_behav",status=0,excel_path=self.excel_file_path)
                elif int(checkbox2.get()) == 1:
                    self.default_disk_status_behav = 1
                    Tools.save_setting_parameter(parameter="change_def_disk_behav",status=1,excel_path=self.excel_file_path)
                    self.make_project_cells_disk(no_read=True)

            def save_new_behav_notes():
                nonlocal checkbox
                if int(checkbox.get()) == 0:
                    self.default_note_behav = 0
                    Tools.save_setting_parameter(parameter="change_def_notes_behav",status=0,excel_path=self.excel_file_path)
                    self.make_project_cells_disk()

                elif int(checkbox.get()) == 1:
                    self.default_note_behav = 1
                    Tools.save_setting_parameter(parameter="change_def_notes_behav",status=1,excel_path=self.excel_file_path)
                    self.make_project_cells_disk()

            def save_new_disk_map_cond():
                nonlocal checkbox3
                if int(checkbox3.get()) == 0:
                    self.mapping_condition = 0
                    Tools.save_setting_parameter(parameter="change_mapping_cond",status=0,excel_path=self.excel_file_path)
                elif int(checkbox3.get()) == 1:
                    self.mapping_condition = 1
                    Tools.save_setting_parameter(parameter="change_mapping_cond",status=1,excel_path=self.excel_file_path)

            def change_make_first_behav():
                nonlocal checkbox4
                if int(checkbox4.get()) == 0:
                    self.make_edited_project_first = False
                    Tools.save_setting_parameter(parameter="change_make_first_behav",status=0,excel_path=self.excel_file_path)
                elif int(checkbox4.get()) == 1:
                    self.make_edited_project_first = True
                    Tools.save_setting_parameter(parameter="change_make_first_behav",status=1,excel_path=self.excel_file_path)

            def delete_behav():
                if int(checkbox5.get()) == 0 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 100
                    Tools.save_setting_parameter(parameter="delete_behav",status=self.deletion_behav,excel_path=self.excel_file_path)
                elif int(checkbox5.get()) == 0 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 101
                    Tools.save_setting_parameter(parameter="delete_behav",status=self.deletion_behav,excel_path=self.excel_file_path)
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 110
                    Tools.save_setting_parameter(parameter="delete_behav",status=self.deletion_behav,excel_path=self.excel_file_path)
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 111
                    Tools.save_setting_parameter(parameter="delete_behav",status=self.deletion_behav,excel_path=self.excel_file_path)

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            # child_root.geometry(f"620x580+{x+350}+{y+180}")

            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Nastavení")
            main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "Chování poznámek (editovatelné/ needitovatelné):",font=("Arial",20,"bold"))
            checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "Přímo zapisovat a ukládat do poznámek na úvodní obrazovce",font=("Arial",16,"bold"),command=lambda: save_new_behav_notes())
            label.          pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox.       pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame2 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label2 =        customtkinter.CTkLabel(master = main_frame2, width = 100,height=40,text = "Chování při vstupu do menu \"Síťové disky\":",font=("Arial",20,"bold"))
            checkbox2 =     customtkinter.CTkCheckBox(master = main_frame2, text = "Při spuštění aktualizovat statusy disků",font=("Arial",16,"bold"),command=lambda: save_new_behav_disk())
            label2.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox2.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame3 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label3 =        customtkinter.CTkLabel(master = main_frame3, width = 100,height=40,text = "Nastavení mapování disků:",font=("Arial",20,"bold"))
            checkbox3 =     customtkinter.CTkCheckBox(master = main_frame3, text = "Automaticky připojovat po restartu PC",font=("Arial",16,"bold"),command=lambda: save_new_disk_map_cond())
            frame_drive1 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
            drive_color1 =  customtkinter.CTkFrame(master=frame_drive1,corner_radius=0,width = 30,height = 30,fg_color="green")
            drive_label1 =  customtkinter.CTkLabel(master = frame_drive1, width = 100,height=40,text = "= disk je online, persistentní (po vypnutí bude znovu načten)",font=("Arial",18))
            drive_color1.   pack(pady = (2,0),padx=10,side="left",anchor = "w")
            drive_label1.   pack(pady = (2,0),padx=0,side="left",anchor = "w")
            frame_drive2 =  customtkinter.CTkFrame(master=main_frame3,corner_radius=0,fg_color="#212121")
            drive_color2 =  customtkinter.CTkFrame(master=frame_drive2,corner_radius=0,width = 30,height = 30,fg_color="#00CED1")
            drive_label2 =  customtkinter.CTkLabel(master = frame_drive2, width = 100,height=40,text = "= disk je online, nepersistentní (bude odpojen po vypnutí)",font=("Arial",18))
            drive_color2.   pack(pady = (2,0),padx=10,side="left",anchor = "w")
            drive_label2.   pack(pady = (2,0),padx=0,side="left",anchor = "w")
            label3.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox3.      pack(pady = 10,padx=10,side="top",anchor = "w")
            frame_drive1.   pack(pady = 0,padx=(2,5),side="top",anchor = "w",fill="x")
            frame_drive2.   pack(pady = 0,padx=(2,5),side="top",anchor = "w",fill="x")
            
            main_frame4 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label4 =        customtkinter.CTkLabel(master = main_frame4, width = 100,height=40,text = "Nastavení chování při editaci projektů:",font=("Arial",20,"bold"))
            checkbox4 =     customtkinter.CTkCheckBox(master = main_frame4, text = "Automaticky přesouvat editovaný projekt na začátek",font=("Arial",16,"bold"),command=lambda: change_make_first_behav())
            label4.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox4.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame5 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label5 =        customtkinter.CTkLabel(master = main_frame5, width = 100,height=40,text = "Odvolit dotazování při mazání:",font=("Arial",20,"bold"))
            checkbox5 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit v hlavním okně",font=("Arial",16,"bold"),command=lambda: delete_behav())
            checkbox6 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit při editu",font=("Arial",16,"bold"),command=lambda: delete_behav())
            label5.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox5.      pack(pady = 0,padx=10,side="top",anchor = "w")
            checkbox6.      pack(pady = (5,5),padx=10,side="top",anchor = "w")

            close_frame =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
            button_close =  customtkinter.CTkButton(master = close_frame, width = 150,height=40,text = "Zavřít",command = child_root.destroy,font=("Arial",20,"bold"),corner_radius=0)
            button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame2.    pack(expand=False,fill="x",side="top")
            main_frame3.    pack(expand=False,fill="x",side="top")
            main_frame4.    pack(expand=False,fill="x",side="top")
            main_frame5.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")

            if self.default_note_behav == 1:
                checkbox.select()
            
            if self.make_edited_project_first:
                checkbox4.select()

            if self.default_disk_status_behav == 1:
                checkbox2.select()

            if self.mapping_condition == 1:
                checkbox3.select()

            if self.deletion_behav == 110 or self.deletion_behav == 111:
                checkbox5.select()
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                checkbox6.select()

            child_root.update()
            child_root.update_idletasks()
            child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{x+350}+{y+180}")
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def sort_by_alphabet(self):
            project_names_array=[]
            for projects in self.disk_all_rows:
                project_names_array.append(projects[0])
            project_names_sorted = sorted(project_names_array)
            whole_projects_sorted = []
            for names in project_names_sorted:
                for projects in self.disk_all_rows:
                    if projects[0] == names:
                        whole_projects_sorted.append(projects)
                        break
            
            self.disk_all_rows = copy.deepcopy(whole_projects_sorted)            
            for i in range(0,len(self.disk_all_rows)):
                    row = (len(self.disk_all_rows)-1)-i
                    main.DM_tools.save_excel_data_disk(self.excel_file_path,
                                                len(self.disk_all_rows),
                                                self.last_project_id,
                                                self.disk_all_rows[i][0],
                                                self.disk_all_rows[i][1],
                                                self.disk_all_rows[i][2],
                                                self.disk_all_rows[i][3],
                                                self.disk_all_rows[i][4],
                                                self.disk_all_rows[i][5],
                                                force_row_to_print=row+1)
            
            self.make_project_cells_disk()
            Tools.add_colored_line(self.main_console,f"Projekty seřazeny podle abecedy","green",None,True)

        def create_widgets_disk(self,init=None):
            Tools.clear_frame(self.root)
            def edit_project():
                result = self.check_given_input()
                if result == True:
                    self.add_new_project_disk(True)
                elif result == None:
                    Tools.add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
        
            if init:
                if self.window_mode == "max":
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=1,excel_path=self.excel_file_path)
                else:
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=0,excel_path=self.excel_file_path)
                    
            Tools.clear_frame(self.root)
            self.selected_list_disk = []
            self.control_pressed = False
            Tools.save_setting_parameter(parameter="change_def_main_window",status=1,excel_path=self.excel_file_path)
            menu_cards =                    customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
            self.main_widgets =             customtkinter.CTkFrame(master=self.root,corner_radius=0)
            self.project_tree =             customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
            logo =                          customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
            image_logo =                    customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
            main_menu_button =              customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_all_ip =          customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: main.IP_assignment(self.parent_instance,fav_w_called=False),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_favourite_ip =    customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: main.IP_assignment(self.parent_instance,fav_w_called=True),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_disk =            customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
            first_row_frame =               customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,fg_color="#212121")
            project_label =                 customtkinter.CTkLabel(master = first_row_frame,height=40,text = "Projekt: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.search_input =             customtkinter.CTkEntry(master = first_row_frame,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
            button_search =                 customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first_disk("search"),font=("Arial",20,"bold"),corner_radius=0)
            button_add =                    customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project_disk(),font=("Arial",20,"bold"),corner_radius=0)
            button_remove =                 customtkinter.CTkButton(master = first_row_frame, width = 100,height=40,text = "Smazat", command =  lambda: self.delete_project_disk(button_trigger=True,flag="main_menu"),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_button =              customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_deleted_disk"),font=("",28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_edit =                   customtkinter.CTkButton(master = first_row_frame, width = 110,height=40,text = "Editovat",command =  lambda: edit_project(),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_edit =                customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_edited_disk"),font=("",28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_make_first =             customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "🔝",command =  lambda: self.make_project_first_disk(),font=(None,30),corner_radius=0)
            move_upwards =                  customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↑",command =  lambda: self.make_project_first_disk(purpouse="silent",upwards=True),font=(None,25),corner_radius=0)
            move_downwards =                customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↓",command =  lambda: self.make_project_first_disk(purpouse="silent",downwards=True),font=(None,25),corner_radius=0)
            sort_alphabet =                 customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "A↑",command =  lambda: self.sort_by_alphabet(),font=(None,25),corner_radius=0)
            button_settings =               customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(),font=("",22),corner_radius=0)
            second_row_frame =              customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,fg_color="#212121")
            delete_disk =                   customtkinter.CTkButton(master = second_row_frame, width = 250,height=40,text = "Odpojit síťový disk",command =  lambda: self.delete_disk_option_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="red")
            reset =                         customtkinter.CTkButton(master = second_row_frame, width = 200,height=40,text = "Reset exploreru",command = lambda: self.refresh_explorer(refresh_disk=True),font=("Arial",20,"bold"),corner_radius=0)
            self.refresh_btn =              customtkinter.CTkButton(master = second_row_frame, width = 200,height=40,text = "Refresh statusů",command = lambda: self.refresh_disk_statuses(silent=False),font=("Arial",20,"bold"),corner_radius=0)
            as_admin_label =                customtkinter.CTkLabel(master = second_row_frame,text = "",font=("Arial",20,"bold"))
            third_row_frame =               customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,fg_color="#212121")
            self.main_console =             tk.Text(third_row_frame, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
            project_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.search_input.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_search.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_add.                     pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_remove.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_button.               pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_edit.                    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_edit.                 pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_make_first.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_upwards.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_downwards.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            sort_alphabet.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_settings.                pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            delete_disk.                    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            reset.                          pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.refresh_btn.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            as_admin_label.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.main_console.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            first_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            second_row_frame.               pack(pady=0,padx=0,fill="x",side = "top")
            third_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_all_ip.           pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_favourite_ip.     pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            image_logo.                     pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
            menu_cards.                     pack(pady=0,padx=5,fill="x",expand=False,side = "top")
            self.main_widgets.              pack(pady=0,padx=0,fill="x",side = "top")
            self.project_tree.              pack(pady=5,padx=5,fill="both",expand=True,side = "top")

            # self.option_change("",only_console=True)
            if Tools.get_none_count(self.bin_projects[1]) < 4 and len(self.bin_projects[1]) == 6:
                self.undo_edit.configure(state = "normal")
            else:
                self.undo_edit.configure(state = "disabled")

            # poznámky, username, password mohou být None
            if Tools.get_none_count(self.bin_projects[0]) < 4 and len(self.bin_projects[0]) > 2:
                self.undo_button.configure(state = "normal")
            else:
                self.undo_button.configure(state = "disabled")
            
            def is_admin():
                try:
                    return ctypes.windll.shell32.IsUserAnAdmin()
                except:
                    return False
                
            if is_admin():
                as_admin_label.configure(text = "Aplikace je spuštěna, jako administrátor\n(mapovat disky lze pouze na uživatelském účtu)",text_color = "orange")

            def maximalize_window(e):
                self.root.update_idletasks()
                self.root.update()
                current_width = int(self.root.winfo_width())
                # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
                if Tools.focused_entry_widget(self.root): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                    return
                if int(current_width) > 1200:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=2,excel_path=self.excel_file_path)
                elif int(current_width) ==260:
                    self.root.geometry("1200x900")
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=0,excel_path=self.excel_file_path)
                else:
                    self.root.state('zoomed')
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=1,excel_path=self.excel_file_path)
                self.root.update_idletasks()
                self.root.update()
            self.root.bind("<f>",lambda e: maximalize_window(e))

            def unfocus_widget(e):
                self.root.focus_set()
            self.root.bind("<Escape>",unfocus_widget)
            self.search_input.bind("<Return>",unfocus_widget)

            def call_search(e):
                self.make_project_first_disk("search")
            self.search_input.bind("<Return>",call_search)

            def call_refresh(e):
                self.refresh_explorer(refresh_disk=True)
            self.root.bind("<F5>",lambda e: call_refresh(e))

            def call_unfocus(e):
                widget = str(e.widget)
                if not ".!ctkscrollableframe" in widget and not ".!ctktoplevel" in widget and not ".!ctkbutton" in widget:
                    #odebrat focus
                    self.clicked_on_project("",None,None,None,flag="unfocus")
            self.root.bind("<Button-1>",call_unfocus,"+")

            def control_button(status):
                self.control_pressed = status
                if status == True:
                    if not self.last_selected_widget_id in self.selected_list_disk:
                        self.selected_list_disk.append(self.last_selected_widget_id)

            def multi_select():
                if not self.last_project_id in self.selected_list_disk:
                    self.selected_list_disk.append(self.last_project_id)
                    print(self.selected_list_disk)

            self.root.bind("<Control_L>",lambda e: control_button(True))
            self.root.bind("<Control-Button-1>",lambda e: multi_select())
            self.root.bind("<KeyRelease-Control_L>",lambda e: control_button(False))
            self.root.bind("<Delete>",lambda e: self.delete_project_disk(button_trigger=True,flag="main_menu"))
            self.root.update()
            self.make_project_cells_disk(disk_statuses=True,init=True)
            # self.root.mainloop()

    class IP_assignment: # Umožňuje měnit statickou IP
        """
        Umožňuje měnit nastavení statických IP adres
        """
        def __init__(self,parent,fav_w_called=None):
            self.parent_instance = parent
            self.root = parent.root
            self.menu_callback = parent.menu_callback
            self.window_mode = parent.window_mode
            self.initial_path = parent.initial_path
            self.app_icon = parent.app_icon
            self.all_rows = []
            self.project_list = []
            self.excel_file_path = parent.excel_file_path
            self.last_project_name = ""
            self.last_project_ip = ""
            self.last_project_mask = ""
            self.last_project_notes = ""
            self.last_project_id = ""
            self.make_project_favourite = False
            self.favourite_list = []
            self.connection_option_list = []
            self.last_selected_widget = ""
            self.last_selected_notes_widget = ""
            self.last_selected_textbox = ""
            self.last_selected_widget_id = 0
            self.opened_window = ""
            self.ip_frame_list = []
            self.selected_list = []
            self.remember_to_change_back = []
            self.control_pressed = False
            self.edited_project_name = None
            self.bin_projects = [[None],[None]]
            self.changed_notes = []
            self.notes_frame_height = 50

            read_parameters = Tools.read_setting_parameters(self.excel_file_path)
            if read_parameters != None:
                self.default_connection_option = read_parameters[0]
                if read_parameters[1] == 1:
                    self.show_favourite = True
                else:
                    self.show_favourite = False
                if read_parameters[3] == 2:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                self.default_note_behav = read_parameters[5]
                if read_parameters[7] == 1:
                    self.make_edited_project_first = True
                else:
                    self.make_edited_project_first = False
                self.deletion_behav = read_parameters[8]
            else:
                self.default_connection_option = 0
                self.show_favourite = False
                self.default_note_behav = 0
                self.make_edited_project_first = True
                self.deletion_behav = 100
            
            if parent.default_environment == "config_load_error":
                self.create_widgets(init=True,excel_load_error=True)
            elif fav_w_called == None:
                self.create_widgets(fav_status=self.show_favourite,init=True)
            else:
                self.create_widgets(fav_status=fav_w_called,init=True)

        def call_menu(self): # Tlačítko menu (konec, návrat do menu)
            """
            Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do hlavního menu trimazkonu
            """

            Tools.clear_frame(self.main_widgets)
            Tools.clear_frame(self.root)
            # self.root.unbind("<f>")
            self.root.unbind("<Escape>")
            self.root.unbind("<F5>")
            self.root.unbind("<Button-1>")
            self.root.unbind("<Control_L>")
            self.root.unbind("<Control-Button-1>")
            self.root.unbind("<KeyRelease-Control_L>")
            self.root.unbind("<Delete>")
            self.root.update()
            self.root.update_idletasks()
            self.menu_callback()
    
        def manage_bin(self,flag="",parameters=[],wb=None):
            """
            First_row in bin worksheet = last deleted ip\n
            Second_row in bin worksheet = last deleted disk\n
            self.bin_projects = [0] deleted projects\n
            self.bin_projects = [1] edited projects\n
            flag:\n
            - read_sheet
            - save_project_ip
            - load_deleted_ip
            - save_edited_ip
            - load_edited_ip
            - change_notes_back
            """
            bin_worksheet = "projects_bin2"
            
            def read_sheet():
                wb = load_workbook(self.excel_file_path)
                if not bin_worksheet in wb.sheetnames:
                    ws = wb.create_sheet(title=bin_worksheet)
                    ws.sheet_state = "hidden"
                    wb.save(self.excel_file_path)
                    wb.close()
                    print("adding new bin sheet to excel")
                    return [[None],[None]]
                else:
                    ws = wb[bin_worksheet]
                    row_data_ip = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                    wb.close()
                    return [list(row_data_ip),[None]] # provedene zmeny v editu pri spusteni programu nanacitam. Jen smazané projekty...
                
            def save_project_ip():
                nonlocal wb
                # saving after deleting:
                if flag == "save_project_ip":
                    excel_row = 1
                    if wb == None:
                        return False
                    self.undo_button.configure(state = "normal")
                    main.IP_tools.save_excel_data(self.excel_file_path,
                                                        len(self.all_rows),
                                                        self.last_project_id,
                                                        self.show_favourite,
                                                        parameters[0],
                                                        parameters[1],
                                                        parameters[2],
                                                        parameters[3],
                                                        force_row_to_print=excel_row,
                                                        force_ws=bin_worksheet,
                                                        wb_given=wb)
                    self.bin_projects[0] = [parameters[0],parameters[1],parameters[2],parameters[3]]
                # saving after editing:
                elif flag == "save_edited_ip":
                    excel_row = 3
                    self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="load_edited_ip"))
                    print("saving changes to excel: ",parameters)
                    main.IP_tools.save_excel_data(self.excel_file_path,
                                                        len(self.all_rows),
                                                        self.last_project_id,
                                                        self.show_favourite,
                                                        parameters[0],
                                                        parameters[1],
                                                        parameters[2],
                                                        parameters[3],
                                                        force_row_to_print=excel_row,
                                                        fav_status=parameters[4],
                                                        force_ws=bin_worksheet,
                                                        wb_given=wb)
                    self.bin_projects[1] = [parameters[0],parameters[1],parameters[2],parameters[3],parameters[4]]

            def change_notes_back():
                print("loading back: ",self.changed_notes)    
            
                def save_changed_notes(notes,row):
                    workbook = load_workbook(self.excel_file_path)

                    def find_notes_in_whole_list(row,new_fav_status):
                        index_of_project = "no data"
                        try:
                            wanted_project = self.all_rows[row][0]
                            self.show_favourite = new_fav_status
                            self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                            for i in range(0,len(self.all_rows)):
                                if self.all_rows[i][0] == wanted_project:
                                    index_of_project = i
                                    break
                            return index_of_project
                        except Exception as err:
                            print(err)
                            return "no data"

                    def save_to_workbook(notes,row,excel_worksheet):
                        nonlocal workbook
                        worksheet = workbook[excel_worksheet]
                        worksheet['D' + str(len(self.all_rows)-row)] = notes

                    if self.show_favourite:
                        save_to_workbook(notes,row,"ip_address_fav_list")
                        #kontrola pro druhé prostředí
                        index_of_project = find_notes_in_whole_list(row,new_fav_status = False)
                        print("index",index_of_project)
                        if str(index_of_project) != "no data":
                            save_to_workbook(notes,index_of_project,"ip_address_list")
                        self.show_favourite = True
                    else:
                        save_to_workbook(notes,row,"ip_address_list")
                        #kontrola pro druhé prostředí
                        index_of_project = find_notes_in_whole_list(row,new_fav_status = True)
                        print("index",index_of_project)
                        if str(index_of_project) != "no data":
                            save_to_workbook(notes,index_of_project,"ip_address_fav_list")
                        self.show_favourite = False

                    workbook.save(filename=self.excel_file_path)
                    workbook.close()
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                
                project_name = self.changed_notes[0]
                notes_before = self.changed_notes[1]
                id = None
                for i in range(0,len(self.all_rows)):
                    if self.all_rows[i][0] == project_name:
                        id = i

                save_changed_notes(notes_before,id)
                Tools.add_colored_line(self.main_console,f"Poznámky u projektu: {project_name} byly úspěšně obnoveny","green",None,True)
                self.make_project_cells()

                if Tools.get_none_count(self.bin_projects[1]) < 2 and len(self.bin_projects[1]) == 5:
                    self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="load_edited_ip"))
                else:
                    self.undo_edit.configure(state = "disabled")
                self.changed_notes = []

            def load_deleted_ip():
                """
                adds new project from history and deletes the history
                """
                wb = load_workbook(self.excel_file_path)
                ws = wb[bin_worksheet]
                row_data_ip = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

                print("\nrow data ip: ",row_data_ip,"\n")

                project_name = row_data_ip[0]
                if project_name in self.project_list:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {project_name} je již používané, nelze ho tedy obnovit","red",None,True)
                    wb.close()
                    return
                
                self.bin_projects[0] = []
                self.undo_button.configure(state = "disabled")
                ws.delete_rows(1)
                wb.save(self.excel_file_path)
                wb.close()
                main.IP_tools.save_excel_data(self.excel_file_path,
                                                    len(self.all_rows),
                                                    self.last_project_id,
                                                    self.show_favourite,
                                                    project_name,
                                                    row_data_ip[1],
                                                    row_data_ip[2],
                                                    row_data_ip[3],
                                                    only_edit=True,
                                                    force_row_to_print=len(self.all_rows)+1,
                                                    fav_status=0,
                                                    force_ws="ip_address_list")
                Tools.add_colored_line(self.main_console,f"Projekt: {project_name} byl úspěšně obnoven","green",None,True)
                self.make_project_cells()

            def load_edited_ip():
                wb = load_workbook(self.excel_file_path)
                ws = wb[bin_worksheet]
                not_edited_data_ip = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

                print("\nrow data ip: ",not_edited_data_ip,"\n")
                if self.edited_project_name not in self.project_list:
                    Tools.add_colored_line(self.main_console,f"Jméno projektu: {self.edited_project_name} nenalezeno, nelze ho tedy obnovit","red",None,True)
                    wb.close()
                    return
                
                self.bin_projects[1] = []
                self.undo_edit.configure(state = "disabled")
                ws.delete_rows(3)
                wb.save(self.excel_file_path)
                wb.close()
                make_fav = False
                if not_edited_data_ip[4] == 1 or not_edited_data_ip[4] == True:
                    make_fav = True
                elif not_edited_data_ip[4] == 0 or not_edited_data_ip[4] == False:
                    make_fav = False

                param = [not_edited_data_ip[0],not_edited_data_ip[1],not_edited_data_ip[2],not_edited_data_ip[3]]
                for i in range(0,len(param)):
                    if param[i] == None:
                        param[i] = ""

                self.last_project_name = self.edited_project_name
                self.save_new_project_data(None,only_edit = True,make_fav=make_fav,bin_manage=True,param=param)
                if self.edited_project_name != not_edited_data_ip[0]:
                    Tools.add_colored_line(self.main_console,f"U projektu: {self.edited_project_name} (původně: {not_edited_data_ip[0]}) byly odebrány provedené změny","green",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"U projektu: {self.edited_project_name} byly odebrány provedené změny","green",None,True)
                self.make_project_cells()

            mapping_logic = {
                "read_sheet": read_sheet,
                "save_project_ip": save_project_ip,
                "load_deleted_ip": load_deleted_ip,
                "save_edited_ip": save_project_ip,
                "load_edited_ip": load_edited_ip,
                "change_notes_back": change_notes_back
            }

            output = mapping_logic[flag]()  # This will call the corresponding function
            return output

        def switch_fav_status(self,operation:str,project_given=None,change_status = False):
            if project_given == None:
                selected_project = str(self.search_input.get())
                if selected_project not in self.project_list:
                    Tools.add_colored_line(self.main_console,"Nebyl vložen projekt",color="red",font=None,delete_line=True)
                    return
                else:
                    selected_project = self.all_rows[self.project_list.index(selected_project)]
            else:
                selected_project = project_given
            if self.show_favourite == False:
                if operation == "add_favourite":
                    if change_status:
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            selected_project[0],
                                                            selected_project[1],
                                                            selected_project[2],
                                                            selected_project[3],
                                                            only_edit=True,
                                                            fav_status=1)
                    self.show_favourite = True
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    # do tohoto prostředí uložím na začátek
                    self.all_rows.insert(0,selected_project)
                    for i in range(0,len(self.all_rows)):
                        row = (len(self.all_rows)-1)-i
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            self.all_rows[i][0],
                                                            self.all_rows[i][1],
                                                            self.all_rows[i][2],
                                                            self.all_rows[i][3],
                                                            force_row_to_print=row+1,
                                                            fav_status=1)
                    # přepnutí zpět
                    self.show_favourite = False
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

                
                elif operation == "del_favourite":
                    if change_status:
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            selected_project[0],
                                                            selected_project[1],
                                                            selected_project[2],
                                                            selected_project[3],
                                                            only_edit=True,
                                                            fav_status=0)
                    # přepnutí
                    self.show_favourite = True
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    # z tohoto prostředí smažu
                    self.delete_project(wanted_project=selected_project[0],silence=True,del_favourite = True)
                    # přepnutí zpět
                    self.show_favourite = False
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

                elif operation == "rewrite_favourite":
                    # přepnutí
                    self.show_favourite = True
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    # nejprve popnu stary projekt, s povodnim jmenem
                    # poté insertnu pozměněný
                    the_id_to_pop = self.project_list.index(self.last_project_name)
                    self.all_rows.pop(the_id_to_pop)
                    self.all_rows.insert(0,selected_project)
                    for i in range(0,len(self.all_rows)):
                        row = (len(self.all_rows)-1)-i
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            self.all_rows[i][0],
                                                            self.all_rows[i][1],
                                                            self.all_rows[i][2],
                                                            self.all_rows[i][3],
                                                            force_row_to_print=row+1,
                                                            fav_status=1)
                    # přepnutí zpět
                    self.show_favourite = False
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

            elif self.show_favourite:
                # z aktuálního prostředí smažu
                self.delete_project(wanted_project=selected_project[0],silence=True)
                # musim prepnout prostředí jen kvůli změně statusu
                self.show_favourite = False
                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                match_found = False
                for i in range(0,len(self.project_list)):
                    if self.project_list[i] == selected_project[0] and len(str(self.project_list[i])) == len(str(selected_project[0])):
                        row_index = self.project_list.index(selected_project[0])
                        match_found = True
                if match_found:
                    row = len(self.all_rows) - row_index
                    main.IP_tools.save_excel_data(self.excel_file_path,
                                                        len(self.all_rows),
                                                        self.last_project_id,
                                                        self.show_favourite,
                                                        self.all_rows[row_index][0],
                                                        self.all_rows[row_index][1],
                                                        self.all_rows[row_index][2],
                                                        self.all_rows[row_index][3],
                                                        force_row_to_print=row,
                                                        fav_status=0)
                # přepnutí zpět
                self.show_favourite = True
                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

            
            if operation == "with_refresh":
                Tools.add_colored_line(self.main_console,f"Projekt: {selected_project[0]} byl odebrán z oblíbených","green",None,True)
                self.make_project_cells(no_read=True)

        def save_new_project_data(self,child_root,only_edit = None,make_fav=False,bin_manage = False,param = []):

            def get_both_row_indexes(new_project = False):
                """
                - new project = bool - returs the last position of excel row, where the new project takes place\n
                returns array of 2 excel row indexes: (finds matches)\n
                [0] = normal list\n
                [1] = favourite list\n
                - if not found returns "no data"\n
                """
                wanted_project = self.last_project_name
                print("\n wanted",wanted_project)
                def find_project_index(wanted_project,new_fav_status):
                    index_of_project = "no data"
                    try:
                        if new_fav_status != None:
                            self.show_favourite = new_fav_status
                            self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                        for i in range(0,len(self.all_rows)):
                            if self.all_rows[i][0] == wanted_project:
                                index_of_project = i
                                break
                        if new_project:
                            return len(self.all_rows)
                        elif index_of_project != "no data":
                            return (len(self.all_rows) - index_of_project)
                        else:
                            return index_of_project
                    except Exception as err:
                        print(err)
                        return "no data"

                if self.show_favourite:
                    index_of_fav_project = find_project_index(wanted_project,new_fav_status = None)
                    index_of_project = find_project_index(wanted_project,new_fav_status = False)
                    self.show_favourite = True

                elif not self.show_favourite:
                    index_of_project = find_project_index(wanted_project,new_fav_status = None)
                    index_of_fav_project = find_project_index(wanted_project,new_fav_status = True)
                    self.show_favourite = False

                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                return [index_of_project,index_of_fav_project]

            def switch_database():
                if self.show_favourite:
                    self.show_favourite = False
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                else:
                    self.show_favourite = True
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

            def save_history_data():
                if not bin_manage:
                    self.edited_project_name = project_name
                    self.manage_bin("save_edited_ip",parameters=[self.last_project_name,self.last_project_ip,self.last_project_mask,self.last_project_notes,previous_fav_status])

            def check_ip_and_mask(input_value):
                input_splitted = input_value.split(".")
                if len(input_splitted) == 4:
                    return input_value
                else:
                    return False

            if param == []:
                project_name = str(self.name_input.get())
                IP_adress = str(self.IP_adress_input.get())
                IP_adress = check_ip_and_mask(IP_adress)
                mask = str(self.mask_input.get())
                mask = check_ip_and_mask(mask)
                notes = Tools.get_legit_notes(self.notes_input.get("1.0", tk.END))
            else:
                project_name = param[0]
                IP_adress = param[1]
                mask = param[2]
                notes = param[3]
                
            errors = 0
            if project_name.replace(" ","") == "":
                Tools.add_colored_line(self.console,f"Nezadali jste jméno projektu","red",None,True)
                errors += 1
            if project_name in self.project_list and only_edit == None:
                Tools.add_colored_line(self.console,f"Jméno je již používané","red",None,True)
                errors +=1

            if IP_adress == False and errors == 0:
                Tools.add_colored_line(self.console,f"Neplatná IP adresa","red",None,True)
                errors += 1
            if mask == False and errors == 0:
                Tools.add_colored_line(self.console,f"Neplatná maska","red",None,True)
                errors += 1
            # poznamky nejsou povinne
            if errors ==0:
                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                # pridavam novy projekt 1: rovnou do oblibených, 2:jen do všech
                if only_edit == None: 
                    row_index_list = get_both_row_indexes(new_project=True)
                    if make_fav:
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            force_row_to_print=row_index_list[0]+1,
                                                            fav_status=1,
                                                            force_ws="ip_address_list")
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            force_row_to_print=row_index_list[1]+1,
                                                            fav_status=1,
                                                            force_ws="ip_address_fav_list")
                        Tools.add_colored_line(self.main_console,f"Přidán nový oblíbený projekt: {project_name}","green",None,True)
                    else:
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            force_row_to_print=row_index_list[0]+1,
                                                            fav_status=0,
                                                            force_ws="ip_address_list")
                        Tools.add_colored_line(self.main_console,f"Přidán nový projekt: {project_name}","green",None,True)

                    if not self.make_edited_project_first:
                        save_history_data()
                    if not bin_manage:
                        if  child_root != None:
                            child_root.destroy()
                        self.make_project_cells()

                elif only_edit:
                    # kdyz edituji muze mit projekt jiz prideleny status
                    if not bin_manage:
                        current_fav_status = main.IP_tools.is_project_favourite(self.favourite_list,self.last_project_id)
                    else:
                        id = 0
                        for i in range(0,len(self.all_rows)):
                            if self.all_rows[i][0] == self.last_project_name:
                                id = i
                        print("last project name", self.last_project_name,"id: ",id)
                        current_fav_status = main.IP_tools.is_project_favourite(self.favourite_list,id)
                        self.last_project_id = id

                    print("current fav status: ",current_fav_status)
                    previous_fav_status = current_fav_status

                    if make_fav and current_fav_status == 0:
                        # zaskrtnuto oblibene + nebyl oblibeny  = ZMENA:
                        row_index_list = get_both_row_indexes(new_project=True)
                        print("pridan do oblibenych", row_index_list)
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            force_row_to_print=row_index_list[1]+1,
                                                            fav_status=1,
                                                            force_ws="ip_address_fav_list")
                        
                        main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            fav_status=1)

                        if self.last_project_name != project_name:
                            status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn a přidán do oblíbených"
                        else:
                            status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn a přidán do oblíbených"
                        Tools.add_colored_line(self.main_console,status_text,"green",None,True)
                        current_fav_status = 1

                        edited_project = [project_name,IP_adress,mask,notes,1]
                        if self.make_edited_project_first and not bin_manage:
                            save_history_data()
                            self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)

                    elif make_fav == False and current_fav_status == 1:
                        # neni zaskrtnuto oblibene + je jiz oblibeny = ZMENA
                        row_index_list = get_both_row_indexes()
                        print("odebran z oblibenych", row_index_list)

                        if row_index_list[0] == "no data" or row_index_list[1] == "no data":
                            Tools.add_colored_line(self.main_console,f"Chyba synchronizace (oblíbené <-> všechny). Projekt {self.last_project_name} se nepodařilo pozměnit","red",None,True)
                            if child_root != None:    
                                child_root.destroy()
                            return

                        if self.show_favourite:
                            self.delete_project(wanted_project=self.last_project_name,silence=True)
                            main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            force_row_to_print=row_index_list[0],
                                                            fav_status=0,
                                                            force_ws="ip_address_list")

                        else:
                            # nejprve smazat z oblíbených:
                            workbook = load_workbook(self.excel_file_path)
                            worksheet = workbook["ip_address_fav_list"]
                            worksheet.delete_rows(row_index_list[1])
                            workbook.save(self.excel_file_path)
                            # poté uložit změnu statusu do všech:
                            main.IP_tools.save_excel_data(self.excel_file_path,
                                                            len(self.all_rows),
                                                            self.last_project_id,
                                                            self.show_favourite,
                                                            project_name,
                                                            IP_adress,
                                                            mask,
                                                            notes,
                                                            only_edit=True,
                                                            fav_status=0)

                        if self.last_project_name != project_name:
                            status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn a odebrán z oblíbených"
                        else:
                            status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn a odebrán z oblíbených"
                        Tools.add_colored_line(self.main_console,status_text,"green",None,True)

                        edited_project = [project_name,IP_adress,mask,notes,0]
                        current_fav_status = 0

                        if self.make_edited_project_first and not bin_manage:
                            save_history_data()
                            if not self.show_favourite:
                                self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)

                    elif make_fav and current_fav_status == 1:
                        # zaskrtnuto oblibene + je jiz oblibeny = BEZ ZMENY
                        #nedoslo ke zmene statusu, ale mohlo dojit ke zmene - proto prepsat v oblibenych
                        row_index_list = get_both_row_indexes()
                        print("pozmenen 1",row_index_list)
                        if row_index_list[0] != "no data":
                            main.IP_tools.save_excel_data(self.excel_file_path,
                                                                len(self.all_rows),
                                                                self.last_project_id,
                                                                self.show_favourite,
                                                                project_name,
                                                                IP_adress,
                                                                mask,
                                                                notes,
                                                                only_edit=True,
                                                                force_row_to_print=row_index_list[0],
                                                                fav_status=current_fav_status,
                                                                force_ws="ip_address_list")
                        if row_index_list[1] != "no data":
                            main.IP_tools.save_excel_data(self.excel_file_path,
                                                                len(self.all_rows),
                                                                self.last_project_id,
                                                                self.show_favourite,
                                                                project_name,
                                                                IP_adress,
                                                                mask,
                                                                notes,
                                                                only_edit=True,
                                                                force_row_to_print=row_index_list[1],
                                                                fav_status=current_fav_status,
                                                                force_ws="ip_address_fav_list")
                        if self.last_project_name != project_name:
                            status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn"
                        else:
                            status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn"
                        Tools.add_colored_line(self.main_console,status_text,"green",None,True)

                        edited_project = [project_name,IP_adress,mask,notes,current_fav_status]
                        if self.make_edited_project_first and not bin_manage:
                            save_history_data()
                            self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)
                            # promítnout změny i do druhého menu:
                            switch_database()
                            self.make_project_first(purpouse="silent",make_cells=False,project=edited_project,input_entry_bypass=edited_project[0])
                            switch_database()
                        
                    elif make_fav == False and current_fav_status == 0:
                        # neni zaskrtnuto oblibene + nebyl oblibeny = BEZ ZMENY
                        row_index_list = get_both_row_indexes()
                        print("pozmenen 2",row_index_list)
                        if row_index_list[0] != "no data":
                            main.IP_tools.save_excel_data(self.excel_file_path,
                                                                len(self.all_rows),
                                                                self.last_project_id,
                                                                self.show_favourite,
                                                                project_name,
                                                                IP_adress,
                                                                mask,
                                                                notes,
                                                                only_edit=True,
                                                                force_row_to_print=row_index_list[0],
                                                                fav_status=current_fav_status,
                                                                force_ws="ip_address_list")
                        
                        if self.last_project_name != project_name:
                            status_text = f"Projekt: {self.last_project_name} (nově: {project_name}) úspěšně pozměněn"
                        else:
                            status_text = f"Projekt: {self.last_project_name} úspěšně pozměněn"
                        Tools.add_colored_line(self.main_console,status_text,"green",None,True)

                        edited_project = [project_name,IP_adress,mask,notes,current_fav_status]
                        if self.make_edited_project_first and not bin_manage:
                            save_history_data()
                            self.make_project_first(purpouse="silent",make_cells=False,project=edited_project)

                    if not self.make_edited_project_first:
                        save_history_data()
                    if not bin_manage:
                        if  child_root != None:
                            child_root.destroy()
                        self.make_project_cells()

        def delete_project(self,wanted_project=None,silence=None,button_trigger = False,flag="",del_favourite=False):
            if "!ctktextbox" in str(self.root.focus_get()):
                return
            
            project_found = False
            name_list = []

            def check_multiple_projects(window):
                nonlocal wanted_project
                nonlocal name_list
                nonlocal project_found

                if len(self.selected_list) > 1:
                    for names in name_list:
                        print(names)
                        project_found = False
                        self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                        proceed(names,window,True)
                        # print(deleted_project)
                            
                    Tools.add_colored_line(self.main_console,f"Byly úspěšně odstraněny tyto projekty: {name_list}","orange",None,True)
                    try:
                        self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                    except Exception as e:
                        print("chyba, refresh po mazani")
                else:
                    proceed(wanted_project,window)

            def proceed(wanted_project,window = True,multiple_status=False):
                # nonlocal wanted_project
                nonlocal silence
                nonlocal project_found
                nonlocal child_root
                deleted_project = None
                remove_favourite_as_well = False
                if wanted_project == None:
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    wanted_project = str(self.search_input.get())
                workbook = load_workbook(self.excel_file_path)
                if self.show_favourite:
                    excel_worksheet = "ip_address_fav_list"
                else:
                    excel_worksheet = "ip_address_list"
                worksheet = workbook[excel_worksheet]

                for i in range(0,len(self.project_list)):
                    if self.project_list[i] == wanted_project and len(str(self.project_list[i])) == len(str(wanted_project)) and project_found == False:
                        row_index = self.project_list.index(wanted_project)
                        row_data = self.all_rows[row_index]
                        if not self.show_favourite:
                            # když mažu z oblíbených, tak neukládám historii
                            self.manage_bin(flag="save_project_ip",parameters=row_data,wb=workbook)
                        worksheet.delete_rows(len(self.all_rows)-row_index)
                        workbook.save(self.excel_file_path)
                        project_found = True
                        deleted_project = self.all_rows[row_index]
                        print("project list",self.project_list)
                        #pokud ma status oblibenosti, tak vymazat i z oblibenych:
                        if self.favourite_list[row_index] == 1 and self.show_favourite == False:
                            remove_favourite_as_well = True
                        break
                
                workbook.close()
                if silence == None and not multiple_status:
                    if project_found:
                        Tools.add_colored_line(self.main_console,f"Projekt {wanted_project} byl odstraněn","orange",None,True)
                        self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...
                    elif wanted_project.replace(" ","") == "":
                        Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                    else:
                        Tools.add_colored_line(self.main_console,f"Zadaný projekt: {wanted_project} nebyl nalezen","red",None,True)

                # zresetuj i v pripade silence...
                elif project_found and not multiple_status and not del_favourite:
                    self.make_project_cells() #refresh = cele zresetovat, jine: id, poradi...

                if remove_favourite_as_well:
                    self.switch_fav_status("del_favourite",deleted_project)

                if window and child_root.winfo_exists():
                    child_root.grab_release()
                    child_root.destroy()

                return deleted_project

            if not button_trigger:
                proceed(wanted_project,window=False)
                # check_multiple_projects(False)
                return

            if flag == "main_menu" or flag == "context_menu":
                if self.deletion_behav == 110 or self.deletion_behav == 111:
                    # proceed(wanted_project,window=False)

                    check_multiple_projects(False)
                    return
                
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                # proceed(wanted_project,window=False)
                check_multiple_projects(False)
                return
            
            if self.last_project_name.replace(" ","") == "":
                Tools.add_colored_line(self.main_console,"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
                return
            elif wanted_project == None:
                wanted_project = self.last_project_name

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Upozornění")
            proceed_label_text = f"Opravdu si přejete odstranit projekt {self.last_project_name}?"
            if flag == "context_menu":
                self.selected_list = []
            if len(self.selected_list) > 1:
                for ids in self.selected_list:
                    name_list.append(self.all_rows[ids][0])
                proceed_label_text = f"Opravdu si přejete odstranit vybrané projekty:\n{name_list}?"
                
            proceed_label = customtkinter.CTkLabel(master = child_root,text = proceed_label_text,font=("Arial",22,"bold"),justify = "left",anchor="w")
            button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda: check_multiple_projects(True))
            button_no =     customtkinter.CTkButton(master = child_root,text = "NE",font=("Arial",20,"bold"),width = 180,height=40,corner_radius=0,command=lambda:  child_root.destroy())
            proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top",anchor="w")
            button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
            child_root.update()
            child_root.update_idletasks()
            child_root.grab_set()
            child_root.focus()
            child_root.focus_force()
            child_root.wait_window()
            return project_found

        def add_new_project(self,edit = None,init_copy = False):
            def mouse_wheel_change(e):
                if -e.delta < 0:
                    switch_up()
                else:
                    switch_down()

            def copy_previous_project():
                if self.last_project_name == "":
                    Tools.add_colored_line(self.console,"Není vybrán žádný projekt","red",None,True)
                    return
                self.name_input.delete("0","300")
                self.name_input.insert("0",str(self.last_project_name))
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0",str(self.last_project_ip))
                self.mask_input.delete("0","300")
                self.mask_input.insert("0",str(self.last_project_mask))
                self.notes_input.delete("1.0",tk.END)
                self.notes_input.insert(tk.END,str(self.last_project_notes))

            def switch_up():
                print("up ",self.last_project_id)
                self.last_project_id -= 1
                if self.last_project_id < 0:
                    self.last_project_id = len(self.all_rows)-1
                    
                self.check_given_input(given_data=self.all_rows[self.last_project_id][0])
                copy_previous_project()
                refresh_favourite_status()
                refresh_title()

            def switch_down():
                print("down ",self.last_project_id)
                self.last_project_id += 1
                if self.last_project_id > len(self.all_rows)-1:
                    self.last_project_id = 0

                self.check_given_input(given_data=self.all_rows[self.last_project_id][0])
                copy_previous_project()
                refresh_favourite_status()
                refresh_title()

            def del_project():
                nonlocal child_root
                result = self.delete_project(button_trigger=True)
                print(result)
                if result:
                    switch_up()
                else:
                    print("aborted")

                child_root.focus()
                child_root.focus_force()
                child_root.grab_set()

            def refresh_favourite_status():
                if main.IP_tools.is_project_favourite(self.favourite_list,self.last_project_id):
                    self.make_project_favourite = True #init hodnota
                    self.make_fav_label.configure(text = "Oblíbený ❤️",font=("Arial",22))
                    self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
                else:
                    self.make_project_favourite = False #init hodnota
                    self.make_fav_label.configure(text = "Neoblíbený",font=("Arial",22))
                    self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")

            def refresh_title():
                if edit:
                    child_root.title("Editovat projekt: "+self.last_project_name)
                else:
                    child_root.title("Nový projekt")

            def make_favourite_toggle_via_edit(e):
                def do_favourite():
                    self.make_fav_btn.configure(text = "🐘",font=("Arial",38),text_color = "pink")
                    self.make_fav_label.configure(text = "Oblíbený ❤️")
                
                def unfavourite():
                    self.make_fav_btn.configure(text = "❌",font=("Arial",28),text_color = "red")
                    self.make_fav_label.configure(text = "Neoblíbený")

                if self.make_project_favourite:
                    self.make_project_favourite = False
                    unfavourite()
                else:
                    self.make_project_favourite = True
                    do_favourite()

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"520x750+{x+50}+{y+80}")
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            refresh_title()
            project_name =    customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Název projektu: ",font=("Arial",20,"bold"))
            self.name_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            project_selection_label = customtkinter.CTkLabel(master = child_root, width = 200,height=30,text = "Přepnout projekt: ",font=("Arial",20,"bold"))
            project_switch_frame =  customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=140,width=80)
            project_up =            customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↑",command= lambda: switch_up())
            project_down =          customtkinter.CTkButton(master = project_switch_frame,font=("Arial",25,"bold"),width=60,height=60,corner_radius=0,text="↓",command= lambda: switch_down())
            project_switch_frame.   grid_propagate(0)
            project_up              .grid(column = 0,row=0,pady = (5,0),padx =10)
            project_down            .grid(column = 0,row=1,pady = 5,padx =10)
            project_switch_frame.   bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_up.             bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            project_down.           bind("<MouseWheel>",lambda e: mouse_wheel_change(e))
            IP_adress =            customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "IP adresa: ",font=("Arial",20,"bold"))
            self.IP_adress_input = customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            mask =                 customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Maska: ",font=("Arial",20,"bold"))
            self.mask_input =      customtkinter.CTkEntry(master = child_root,font=("Arial",20),width=200,height=30,corner_radius=0)
            copy_check =           customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Kopírovat předchozí projekt",command= lambda: copy_previous_project())
            del_project_btn =      customtkinter.CTkButton(master = child_root,font=("Arial",20),width=250,height=30,corner_radius=0,text="Smazat tento projekt",command= lambda: del_project(),fg_color="red")
            fav_status =           customtkinter.CTkLabel(master = child_root, width = 20,height=30,text = "Status oblíbenosti: ",font=("Arial",20,"bold"))
            fav_frame =            customtkinter.CTkFrame(master=child_root,corner_radius=0,border_width=0,height=50,width=200,fg_color="#353535")
            self.make_fav_label =  customtkinter.CTkLabel(master = fav_frame, width = 20,height=30)
            self.make_fav_btn =    customtkinter.CTkLabel(master = fav_frame, width = 50,height=50)
            refresh_favourite_status()
            notes =                customtkinter.CTkLabel(master = child_root, width = 60,height=30,text = "Poznámky: ",font=("Arial",20,"bold"))
            self.notes_input =     customtkinter.CTkTextbox(master = child_root,font=("Arial",20),width=500,height=280)
            self.console =         tk.Text(child_root, wrap="none", height=0, width=45,background="black",font=("Arial",14),state=tk.DISABLED)
            if edit:
                save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,True,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
            else:
                save_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Uložit", command = lambda: self.save_new_project_data(child_root,None,self.make_project_favourite),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =  customtkinter.CTkButton(master = child_root, width = 200,height=40,text = "Zrušit", command = lambda: child_root.destroy(),font=("Arial",20,"bold"),corner_radius=0)

            project_name.           grid(column = 0,row=0,pady = 5,padx =10,sticky = tk.W)
            if edit:
                project_selection_label.grid(column = 0,row=0,padx=265,sticky = tk.W)
                project_switch_frame.   grid(row=1,column=0,padx=320,sticky=tk.W,rowspan=4)
            else:
                copy_check.             grid(column = 0,row=8,pady = 5,padx =240,sticky = tk.W)
            self.name_input.        grid(column = 0,row=1,pady = 5,padx =10,sticky = tk.W)
            IP_adress.              grid(column = 0,row=2,pady = 5,padx =10,sticky = tk.W)
            self.IP_adress_input.   grid(column = 0,row=3,pady = 5,padx =10,sticky = tk.W)
            mask.                   grid(column = 0,row=4,pady = 5,padx =10,sticky = tk.W)
            self.mask_input.        grid(column = 0,row=5,pady = 5,padx =10,sticky = tk.W)
            fav_status.             grid(column = 0,row=6,pady = 5,padx =10,sticky = tk.W)
            if edit:
                del_project_btn.grid(column = 0,row=8,pady = 5,padx =240,sticky = tk.W)
            fav_frame.              grid(column = 0,row=7,padx= 10,sticky=tk.W)
            fav_frame.              grid_propagate(0)
            self.make_fav_btn.      grid(column=0,row=0,pady = 0,padx =0,sticky = tk.W)
            self.make_fav_btn.      bind("<Button-1>",lambda e: make_favourite_toggle_via_edit(e))
            self.make_fav_label.    grid(column = 0,row=0,pady = 0,padx =60,sticky = tk.W)
            self.make_fav_label.    bind("<Button-1>",lambda e: make_favourite_toggle_via_edit(e))
            notes.                  grid(column = 0,row=8,pady = 5,padx =10,sticky = tk.W)
            self.notes_input.       grid(column = 0,row=9,pady = 5,padx =10,sticky = tk.W)
            self.console.           grid(column = 0,row=10,pady = 5,padx =10,sticky = tk.W)
            save_button.            grid(column = 0,row=11,pady = 5,padx =100,sticky = tk.W)
            exit_button.            grid(column = 0,row=11,pady = 5,padx =310,sticky = tk.W)

            if edit:
                copy_previous_project()
            elif init_copy:
                copy_previous_project()
            else:
                self.IP_adress_input.delete("0","300")
                self.IP_adress_input.insert("0","192.168.000.000")
                self.mask_input.delete("0","300")
                self.mask_input.insert("0","255.255.255.0")
                if str(self.search_input.get()).replace(" ","") != "":
                    self.name_input.delete("0","300")
                    self.name_input.insert("0",str(self.search_input.get()))

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")
        
        def make_sure_ip_changed(self,interface_name,ip):
            def run_as_admin():
                # Vyžádání admin práv: nefunkční ve vscode
                def is_admin():
                    try:
                        return ctypes.windll.shell32.IsUserAnAdmin()
                    except:
                        return False
                pid = os.getpid()
                if not is_admin():
                    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(["admin_ip_setting",str(pid)]), None, 1)
                    sys.exit()
            def open_app_as_admin_prompt():
                def close_prompt(child_root):
                    child_root.destroy()
                child_root = customtkinter.CTkToplevel()
                child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
                self.opened_window = child_root
                x = self.root.winfo_rootx()
                y = self.root.winfo_rooty()
                child_root.geometry(f"620x150+{x+300}+{y+300}")  
                child_root.title("Upozornění")
                proceed_label = customtkinter.CTkLabel(master = child_root,text = "Přejete si znovu spustit aplikaci, jako administrátor?",font=("Arial",25))
                button_yes =    customtkinter.CTkButton(master = child_root,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin())
                button_no =     customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
                proceed_label   .pack(pady=(15,0),padx=10,anchor="w",expand=False,side = "top")
                button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
                button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
                child_root.update()
                child_root.update_idletasks()
                child_root.focus()
                child_root.focus_force()
                self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

            interface_index = self.connection_option_list.index(interface_name)

            def call_subprocess():
                try:

                    if ip == self.current_address_list[interface_index]:
                        Tools.add_colored_line(self.main_console,f"Pro interface {interface_name} je již tato adresa ({ip}) nastavena","orange",None,True)
                        return
                    elif ip in self.current_address_list:
                        Tools.add_colored_line(self.main_console,f"Chyba, adresa je již používána pro jiný interface","red",None,True)
                        return
                    win_change_ip_time = 7
                    for i in range(0,win_change_ip_time):
                        Tools.add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                        self.option_change("",silent=True)
                        if ip == self.current_address_list[interface_index]: # někdy dříve než 7 sekund...
                            break
                        time.sleep(1)

                    self.option_change("",silent=True)
                    if ip == self.current_address_list[interface_index]:
                        Tools.add_colored_line(self.main_console,f"IPv4 adresa u {interface_name} byla přenastavena na: {ip}","green",None,True)
                        self.refresh_ip_statuses()
                    else:
                        print("temp ip troubleshooting: ------ ",ip)
                        Tools.add_colored_line(self.main_console,f"Chyba, neplatná adresa nebo daný inteface odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)
                        open_app_as_admin_prompt()
                except Exception:
                    pass
            
            run_background = threading.Thread(target=call_subprocess,)
            run_background.start()

        def change_to_DHCP(self):
            def delay_the_refresh():
                nonlocal previous_addr
                nonlocal interface_index
                nonlocal interface
                new_addr = self.current_address_list[interface_index]
                i = 0
                while new_addr == previous_addr or new_addr == None:
                    Tools.add_colored_line(self.main_console,f"Čekám, až windows provede změny: {7-i} s...","white",None,True)
                    time.sleep(1)
                    self.option_change("",silent=True)
                    new_addr = self.current_address_list[interface_index]
                    print("current addr: ",new_addr)
                    i+=1
                    if i > 6:
                        Tools.add_colored_line(self.main_console,f"Chyba, u {interface} se nepodařilo změnit ip adresu (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)
                        return
                
                Tools.add_colored_line(self.main_console,f"IPv4 adresa interfacu: {interface} úspěšně přenastavena na DHCP (automatickou)","green",None,True)
                self.refresh_ip_statuses()
                return
            
            interface = str(self.interface_drop_options.get())
            if not main.IP_tools.check_DHCP(interface):
                if interface != None or interface != "":
                    interface_index = self.connection_option_list.index(interface)
                    previous_addr = self.current_address_list[interface_index]
                    print("previous addr: ",previous_addr)
                    try:
                        # Construct the netsh command
                        netsh_command = f"netsh interface ipv4 set address name=\"{interface}\" source=dhcp"
                        print(f"calling: {netsh_command}")
                        powershell_command = [
                            'powershell.exe',
                            '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                        ]
                        process = subprocess.Popen(powershell_command,
                                                    stdout=subprocess.PIPE,
                                                    stderr=subprocess.PIPE,
                                                    creationflags=subprocess.CREATE_NO_WINDOW)
                        
                        stdout, stderr = process.communicate()
                        stdout_str = stdout.decode('utf-8')
                        stderr_str = stderr.decode('utf-8')
                        if stderr_str:
                            print(f"Error occurred: {stderr_str}")
                        else:
                            print(f"Command executed successfully:\n{stdout_str}")                
                        
                        run_background = threading.Thread(target=delay_the_refresh,)
                        run_background.start()

                    except Exception as e:
                        print(f"Exception occurred: {str(e)}")
                else:
                    Tools.add_colored_line(self.main_console,"Nebyl zvolen žádný interface","red",None,True)
            else:
                connected_interfaces = self.refresh_interfaces()
                if interface in connected_interfaces:
                    Tools.add_colored_line(self.main_console,f"{interface} má již nastavenou DHCP","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Chyba, {interface} je odpojen od tohoto zařízení (pro nastavování odpojených interfaců spusťtě aplikaci jako administrátor)","red",None,True)

        def change_computer_ip(self,button_row,force_params = []):
            """
            button_row - index, kde se nachazi ip a maska v poli: self.all_rows
            """
            def connected_interface(interface,ip,mask):
                """
                Když jsou vyžadována admin práva, tato funkce ověří, zda není daný interface připojen nebo součástí zařízení a zkusí znovu
                """
                try:
                    # Construct the netsh command
                    netsh_command = f"netsh interface ip set address \"{interface}\" static {ip} {mask}"
                    powershell_command = [
                        'powershell.exe',
                        '-Command', f'Start-Process powershell -Verb RunAs -ArgumentList \'-Command "{netsh_command}"\' -WindowStyle Hidden -PassThru'
                    ]
                    process = subprocess.Popen(powershell_command,
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
                    
                    stdout, stderr = process.communicate()
                    stdout_str = stdout.decode('utf-8')
                    stderr_str = stderr.decode('utf-8')
                    if stderr_str:
                        print(f"Error occurred: {stderr_str}")
                        Tools.add_colored_line(self.main_console,f"Chyba, nebyla poskytnuta práva (dejte ANO :))","red",None,True)
                    else:
                        print(f"Command executed successfully:\n{stdout_str}")
                        self.make_sure_ip_changed(interface_name,ip)

                except Exception as e:
                    print(f"Exception occurred: {str(e)}")

            ip = str(self.all_rows[button_row][1])
            mask = str(self.all_rows[button_row][2])
            if len(force_params) > 0:
                ip = force_params[0]
                mask = force_params[1]

            # powershell command na zjisteni network adapter name> Get-NetAdapter | Select-Object -Property InterfaceAlias, Linkspeed, Status
            interface_name = str(self.interface_drop_options.get())
            powershell_command = f"netsh interface ip set address \"{interface_name}\" static " + ip + " " + mask
            try:
                process = subprocess.Popen(['powershell.exe', '-Command', powershell_command],
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                stdout, stderr =process.communicate()
                stdout_str = stdout.decode('utf-8')
                stderr_str = stderr.decode('utf-8')
                print(f"VÝSTUP Z IP SETTING: {str(stdout_str)}")

                if len(str(stdout_str)) > 7:
                    raise subprocess.CalledProcessError(1, powershell_command, stdout_str)
                if stderr_str:
                    raise subprocess.CalledProcessError(1, powershell_command, stderr_str)

                self.make_sure_ip_changed(interface_name,ip)

            except subprocess.CalledProcessError as e:
                if "Run as administrator" in str(stdout_str):
                    Tools.add_colored_line(self.main_console,f"Chyba, tato funkce musí být spuštěna s administrátorskými právy","red",None,True)
                    # trigger powershell potvrzení:
                    connected_interface(interface_name,ip,mask)
                elif "Invalid address" in str(stdout_str):
                    Tools.add_colored_line(self.main_console,f"Chyba, neplatná IP adresa","red",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Chyba, Nemáte tuto adresu již nastavenou pro jiný interface? (nebo daný interface na tomto zařízení neexistuje)","red",None,True)
            except Exception as e:
                # Handle any other exceptions that may occur
                Tools.add_colored_line(self.main_console, f"Nastala neočekávaná chyba: {e}", "red", None, True)

        def check_given_input(self,given_data = None):
            """
            Fills all parameters of last project
            """
            if given_data == None:
                given_data = self.search_input.get()
            if given_data == "":
                found = None
                return found
            found = False

            for i in range(0,len(self.all_rows)):
                if given_data == self.all_rows[i][0]:
                    self.last_project_name =    str(self.all_rows[i][0])
                    self.last_project_ip =      str(self.all_rows[i][1])
                    self.last_project_mask =    str(self.all_rows[i][2])
                    self.last_project_notes =   str(self.all_rows[i][3])
                    self.last_project_id = i
                    found = True
            return found    

        def clicked_on_project(self,event,widget_id,widget,textbox = "",flag = ""):
            """
            flag = notes:
            - při nakliknutí poznámky zůstanou expandnuté a při kliku na jinou je potřeba předchozí vrátit zpět
            flag = unfocus:
            - při kliku mimo se odebere focus z nakliknutých widgetů
            """
            def on_leave_entry(widget,row_of_widget):
                """
                při kliku na jiný widget:
                - upraví text pouze na první řádek
                """
                widget.configure(state = "normal")
                if "\n" in self.all_rows[row_of_widget][3]:
                    notes_rows = self.all_rows[row_of_widget][3].split("\n")
                    first_row = notes_rows[0]
                    widget.delete("1.0",tk.END)
                    widget.insert(tk.END,str(first_row))
                if self.default_note_behav == 0:
                    widget.configure(state = "disabled")

            def shrink_frame(widget_frame,widget_notes):
                widget_notes.configure(state = "normal")
                new_height = 50
                widget_frame.configure(height = new_height) #frame
                widget_notes.configure(height = new_height-10) #notes
                if self.default_note_behav == 0:
                    widget_notes.configure(state = "disabled")

            if flag == "unfocus":
                try:
                    if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                        if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                            shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                            self.last_selected_textbox = ""
                            self.last_selected_notes_widget = ""

                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        self.last_selected_widget.configure(border_color="#636363")
                        self.last_selected_widget = ""

                    for frame_and_id in self.remember_to_change_back:
                        if frame_and_id[0].winfo_exists(): 
                            frame_and_id[0].configure(border_color="#636363")
                    self.selected_list = []
                    self.remember_to_change_back = []

                except Exception as e:
                    print("chyba při odebírání focusu: ",e)
                return

            if widget_id == None:
                return
            
            print("widget_id",widget_id)
            self.search_input.delete("0","300")
            self.search_input.insert("0",str(self.all_rows[widget_id][0]))
            self.check_given_input()
            # only if it is not pressed againt the same:
            if widget != self.last_selected_widget:
                try:
                    if self.last_selected_notes_widget != "" and self.last_selected_notes_widget.winfo_exists():
                        if self.last_selected_textbox != ""  and self.last_selected_textbox.winfo_exists():
                            on_leave_entry(self.last_selected_textbox,self.last_selected_widget_id)
                            shrink_frame(self.last_selected_widget,self.last_selected_textbox)
                    if flag == "notes":
                        self.last_selected_textbox = textbox
                        self.last_selected_notes_widget = widget
                        widget.focus_set()
                    else:
                        # init the values:
                        self.last_selected_textbox = ""
                        self.last_selected_notes_widget = ""
                        widget.focus_set()

                except Exception as e:
                    print("chyba s navracenim framu do puvodniho formatu",e)
                
                try:
                    if self.last_selected_widget != "" and self.last_selected_widget.winfo_exists():
                        if len(self.selected_list) == 0 and not self.control_pressed:
                            self.last_selected_widget.configure(border_color="#636363")

                            if [self.last_selected_widget,self.last_selected_widget_id] in self.remember_to_change_back:
                                self.remember_to_change_back.pop(self.remember_to_change_back.index([self.last_selected_widget,self.last_selected_widget_id]))

                        # pokud došlo k další interakci s jiným widgeten
                        elif not self.control_pressed:
                            for frame_and_id in self.remember_to_change_back:
                                if frame_and_id[0].winfo_exists(): 
                                    frame_and_id[0].configure(border_color="#636363")
                            self.selected_list = []
                            self.remember_to_change_back = []

                    self.last_selected_widget = widget
                    widget.configure(border_color="white")

                    if not [widget,widget_id] in self.remember_to_change_back:
                        self.remember_to_change_back.append([widget,widget_id])

                    print("remember: ", self.remember_to_change_back)

                except Exception as e:
                    print("chyba pri zmene fucusu",e)
                    pass

                self.last_selected_widget_id = widget_id

        def refresh_ip_statuses(self):
            def unbind_connected_ip(widget,frame):
                widget.unbind("<Enter>")
                frame.unbind("<Enter>")
                widget.unbind("<Leave>")
                frame.unbind("<Leave>")
                frame.configure(fg_color = "black")

            def on_enter(e,interface,widget):
                widget.configure(text = interface)   

            def on_leave(e,ip,widget,frame):
                widget.configure(text = ip)
                if ip not in self.current_address_list:
                    unbind_connected_ip(widget,frame)

            for i in range(0,len(self.ip_frame_list)):
                ip_addr = self.all_rows[i][1]
                ip_frame = self.ip_frame_list[i][0]
                parameter = self.ip_frame_list[i][1]
                if ip_addr in self.current_address_list:
                    ip_frame.   configure(fg_color = "green") 
                    ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                    ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
                    parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(e,interface,widget))
                    parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(e,ip,widget,frame))
                else:
                    unbind_connected_ip(parameter,ip_frame)

        def show_context_menu(self,event,first_index,second_index,flag=""):
            """
            - first index (y) = index celeho radku
            - second index (x) = index jednoho parametru
            """
            context_menu = tk.Menu(self.root,tearoff=0,fg="white",bg="black",font=("Arial",20,"bold"))
            self.check_given_input(given_data=self.all_rows[first_index][0])
            
            if flag == "button":
                context_menu.add_command(label="Nastavit",font=("Arial",22,"bold"),command=lambda: self.change_computer_ip(first_index))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat IP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(self.all_rows[first_index][1]))
                context_menu.add_separator()
                context_menu.add_command(label="Editovat",font=("Arial",22,"bold"),command=lambda: self.add_new_project(True))
                context_menu.add_separator()
                context_menu.add_command(label="Kopírovat projekt",font=("Arial",22,"bold"),command=lambda: self.add_new_project(init_copy=True))
                context_menu.add_separator()
                context_menu.add_command(label="Přesunout na začátek",font=("Arial",22,"bold"),command=lambda: self.make_project_first(input_entry_bypass=self.all_rows[first_index][0]))
                context_menu.add_separator()
                context_menu.add_command(label="Odstranit",font=("Arial",22,"bold"),command=lambda: self.delete_project(button_trigger=True,flag="context_menu"))
            elif flag == "ip_frame":
                context_menu.add_command(label="Kopírovat IP adresu",font=("Arial",22,"bold"), command=lambda: pyperclip.copy(self.all_rows[first_index][1]))

            context_menu.tk_popup(event.x_root, event.y_root)

        def make_project_cells(self,no_read = None):
            Tools.clear_frame(self.project_tree)

            def opened_window_check():
                if self.opened_window == "":
                    return False
                try:
                    if self.opened_window.winfo_exists():
                        return True
                    else:
                        return False
                except Exception as err:
                    print(err)
                    return False
                
            def on_enter(interface,widget):
                widget.configure(text = interface)   

            def on_leave(ip,widget,frame):
                widget.configure(text = ip)
                if ip not in self.current_address_list:
                    unbind_connected_ip(widget,frame)

            def unbind_connected_ip(widget,frame):
                widget.unbind("<Enter>")
                frame.unbind("<Enter>")
                widget.unbind("<Leave>")
                frame.unbind("<Leave>")
                frame.configure(fg_color = "black")

            def filter_text_input(text):
                legit_rows = []
                legit_notes = ""
                rows = text.split("\n")
                for i in range(0,len(rows)):
                    if rows[i].replace(" ","") != "":
                        legit_rows.append(rows[i])

                for i in range(0,len(legit_rows)): 
                    if i == len(legit_rows)-1:
                        legit_notes = legit_notes + legit_rows[i]
                    else:
                        legit_notes = legit_notes + legit_rows[i]+ "\n"
                return legit_notes
            
            def save_changed_notes(notes,row):
                workbook = load_workbook(self.excel_file_path)

                def find_notes_in_whole_list(row,new_fav_status):
                    index_of_project = "no data"
                    try:
                        wanted_project = self.all_rows[row][0]
                        self.show_favourite = new_fav_status
                        self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                        for i in range(0,len(self.all_rows)):
                            if self.all_rows[i][0] == wanted_project:
                                index_of_project = i
                                break
                        return index_of_project
                    except Exception as err:
                        print(err)
                        return "no data"

                def save_to_workbook(notes,row,excel_worksheet):
                    nonlocal workbook
                    worksheet = workbook[excel_worksheet]
                    worksheet['D' + str(len(self.all_rows)-row)] = notes

                if self.show_favourite:
                    save_to_workbook(notes,row,"ip_address_fav_list")
                    index_of_project = find_notes_in_whole_list(row,new_fav_status = False)
                    print("index",index_of_project)
                    if str(index_of_project) != "no data":
                        save_to_workbook(notes,index_of_project,"ip_address_list")
                    self.show_favourite = True
                else:
                    save_to_workbook(notes,row,"ip_address_list")
                    index_of_project = find_notes_in_whole_list(row,new_fav_status = True)
                    print("index",index_of_project)
                    if str(index_of_project) != "no data":
                        save_to_workbook(notes,index_of_project,"ip_address_fav_list")
                    self.show_favourite = False

                workbook.save(filename=self.excel_file_path)
                workbook.close()
                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

            def on_enter_entry(widget,row_of_widget):
                if not opened_window_check():
                    if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        widget.configure(state = "normal")
                        widget.delete("1.0",tk.END)
                        widget.insert(tk.END,str(self.all_rows[row_of_widget][3]))
                        if self.default_note_behav == 0:
                            widget.configure(state = "disabled")

            def on_leave_entry(widget,row_of_widget):
                """
                při opuštění widgetu cursorem:
                - upraví text pouze na první řádek
                - uloží změny
                """
                if not opened_window_check():
                    notes_before = filter_text_input(str(self.all_rows[row_of_widget][3]))
                    notes_after = filter_text_input(str(widget.get("1.0",tk.END)))
                    if str(widget) != str(self.last_selected_notes_widget) + ".!ctktextbox":
                        widget.configure(state = "normal")
                        if notes_before != notes_after:
                            self.changed_notes = [self.all_rows[row_of_widget][0],notes_before]
                            self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="change_notes_back"))
                            self.all_rows[row_of_widget][3] = notes_after
                            save_changed_notes(notes_after,row_of_widget)

                        if "\n" in self.all_rows[row_of_widget][3]:
                            notes_rows = self.all_rows[row_of_widget][3].split("\n")
                            first_row = notes_rows[0]
                            widget.delete("1.0",tk.END)
                            widget.insert(tk.END,str(first_row))

                        if self.default_note_behav == 0:
                            widget.configure(state = "disabled")
                        self.root.focus_set() # unfocus widget
                    else:
                        # jinak pouze ulož změny (když je dvakrát nakliknuto to samé)
                        if notes_before != notes_after:
                            self.all_rows[row_of_widget][3] = notes_after
                            self.changed_notes = [self.all_rows[row_of_widget][0],notes_before]
                            self.undo_edit.configure(state = "normal",command = lambda: self.manage_bin(flag="change_notes_back"))
                            save_changed_notes(notes_after,row_of_widget)
                        self.root.focus_set() # unfocus widget
                        
            def shrink_frame(widget):
                tolerance = 5
                if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance:
                    return
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        widget[1].configure(state = "normal")
                        new_height = self.notes_frame_height
                        widget[0].configure(height = new_height) #frame
                        widget[1].configure(height = new_height-10) #notes
                        if self.default_note_behav == 0:
                            widget[1].configure(state = "disabled")

            def expand_frame(widget,row_of_widget):
                if not opened_window_check():
                    if str(widget[0]) != str(self.last_selected_notes_widget):
                        tolerance = 5
                        if abs(int(widget[0]._current_height)-self.notes_frame_height) <= tolerance: # if the height is not default then it means it is expanded already
                            filtered_input = filter_text_input(self.all_rows[row_of_widget][3])
                            self.all_rows[row_of_widget][3] = filtered_input
                            addition = self.notes_frame_height
                            if "\n" in self.all_rows[row_of_widget][3]:
                                notes_rows = self.all_rows[row_of_widget][3].split("\n")
                                if len(notes_rows) > 1:
                                    expanded_dim = addition + (len(notes_rows)-1) * 25
                                    widget[0].configure(height = expanded_dim)
                                    widget[1].configure(state = "normal")
                                    widget[1].configure(height = expanded_dim-10)
                                if self.default_note_behav == 0:
                                    widget[1].configure(state = "disabled")
                            # else:
                            #     if self.default_note_behav == 0:
                            #         widget[1].configure(state = "disabled")
            
            def add_row_return(widget):
                addition = widget[0]._current_height
                expanded_dim = addition + 24
                widget[0].configure(height = expanded_dim)
                widget[1].configure(height = expanded_dim-10)
           
            if no_read == None:
                self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)

            column1 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column2 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column3 =           customtkinter.CTkFrame(master = self.project_tree,corner_radius=0,border_width=0)
            column1_header =    customtkinter.CTkLabel(master = column1,text = "Projekt: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column2_header =    customtkinter.CTkLabel(master = column2,text = "IPv4 adresa: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column3_header =    customtkinter.CTkLabel(master = column3,text = "Poznámky: ",font=("Arial",20,"bold"),justify = "left",anchor = "w")
            column1_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column2_header.     pack(padx = (5,0),side = "top",anchor = "w")
            column3_header.     pack(padx = (5,0),side = "top",anchor = "w")

            self.ip_frame_list = []
            # y = widgets ve smeru y, x = widgets ve smeru x
            for y in range(0,len(self.all_rows)):
                # na pozici x = 2 je maska, kterou nevypisujeme
                for x in range(0,len(self.all_rows[y])):
                    if x == 0: # frame s názvem projektu
                        btn_frame = customtkinter.CTkFrame(master=column1,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        button =    customtkinter.CTkButton(master = btn_frame,width = 200,height=40,text = self.all_rows[y][x],font=("Arial",20,"bold"),corner_radius=0)
                        button.     pack(padx =5,pady = 5, fill= "x")
                        btn_frame.  pack(side = "top",anchor = "w",expand = False,fill= "x")
                        button.     bind("<Button-1>",lambda e,widget = btn_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        button.     bind("<Double-1>",lambda e,widget_id = y: self.change_computer_ip(widget_id))
                        button.     bind("<Button-3>",lambda e, first_index = y, second_index = x: self.show_context_menu(e,first_index,second_index,flag="button"))
                        if main.IP_tools.is_project_favourite(self.favourite_list,y):
                            button.configure(fg_color = "#1E90FF")

                    elif x == 1: # frame s ip adresou
                        ip_frame =  customtkinter.CTkFrame(master=column2,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        parameter = customtkinter.CTkLabel(master = ip_frame,text = self.all_rows[y][x],height=40,width = 250,font=("Arial",20,"bold"),justify='left',anchor = "w")
                        parameter.  pack(padx = (10,5),pady = 5)
                        ip_frame.   pack(side = "top")
                        ip_frame.   bind("<Button-1>",lambda e,widget = ip_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter.  bind("<Button-1>",lambda e,widget = ip_frame, widget_id = y: self.clicked_on_project(e, widget_id,widget))
                        parameter.  bind("<Button-3>",lambda e, first_index = y, second_index = x: self.show_context_menu(e,first_index,second_index,flag="ip_frame"))

                        ip_addr = self.all_rows[y][x]
                        self.ip_frame_list.append([ip_frame,parameter])
                        if ip_addr in self.current_address_list:
                            ip_frame.   configure(fg_color = "green")
                            ip_frame.   bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                            ip_frame.   bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))
                            parameter.  bind("<Enter>",lambda e, interface = self.connection_option_list[self.current_address_list.index(ip_addr)], widget = parameter: on_enter(interface,widget))
                            parameter.  bind("<Leave>",lambda e, ip = ip_addr, widget = parameter,frame = ip_frame: on_leave(ip,widget,frame))

                    elif x == 3: # frame s poznamkami...
                        notes_frame =   customtkinter.CTkFrame(master=column3,corner_radius=0,fg_color="black",border_color="#636363",border_width=2)
                        notes =         customtkinter.CTkTextbox(master = notes_frame,font=("Arial",20,"bold"),corner_radius=0,fg_color="black",height=40)
                        notes.          pack(padx =5,pady = 5,anchor="w",fill="x",expand = True)
                        notes_frame.    pack(pady=0,padx=0,side = "top",anchor = "w",fill="x",expand = True)
                        notes_frame.    bind("<Button-1>",lambda e,widget = notes_frame, widget_id = y, textbox_widget = notes: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))
                        notes.          bind("<Button-1>",lambda e,widget = notes_frame, widget_id = y, textbox_widget = notes: self.clicked_on_project(e, widget_id,widget,textbox_widget,flag="notes"))

                        if "\n" in self.all_rows[y][x]:
                            notes_rows = self.all_rows[y][x].split("\n")
                            first_row = notes_rows[0]
                            notes.delete("1.0",tk.END)
                            notes.insert(tk.END,str(first_row))
                        else:
                            notes.insert(tk.END,str(self.all_rows[y][x]))
                        
                        notes.bind("<Enter>",lambda e, widget = [notes_frame,notes],row=y: expand_frame(widget,row))
                        notes.bind("<Leave>",lambda e, widget = [notes_frame,notes]:       shrink_frame(widget))
                        notes.bind("<Enter>",lambda e, widget = notes,row=y:               on_enter_entry(widget,row))
                        notes.bind("<Leave>",lambda e, widget = notes,row=y:               on_leave_entry(widget,row))
                        notes.bind("<Return>",lambda e, widget = [notes_frame,notes]:      add_row_return(widget))

                        if self.default_note_behav == 0:
                            notes.configure(state = "disabled")

                if y == self.last_project_id: # případ že posouvám s projektem nahoru/ dolů/ top (zvíraznit selectnuté)
                    self.selected_list.append(y)
                    self.last_selected_widget = btn_frame
                    btn_frame.configure(border_color="white")
                    self.remember_to_change_back.append([btn_frame,y])
                    ip_frame.configure(border_color="white")
                    self.remember_to_change_back.append([ip_frame,y])
                    notes_frame.configure(border_color="white")
                    self.remember_to_change_back.append([notes_frame,y])

            column1.pack(fill="both",expand=False,side = "left")
            column2.pack(fill="both",expand=False,side = "left")
            column3.pack(fill="both",expand=True, side = "left")
            self.project_tree.update()
            self.project_tree.update_idletasks()
            self.notes_frame_height = int(notes_frame._current_height)
            try:
                self.project_tree._parent_canvas.yview_moveto(0.0)
            except Exception:
                pass
        
        def edit_project(self):
            result = self.check_given_input()
            if result == True:
                self.add_new_project(True)
            elif result == None:
                Tools.add_colored_line(self.main_console,f"Vyberte projekt pro editaci (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            else:
                Tools.add_colored_line(self.main_console,f"Projekt nenalezen","red",None,True)
            
        def option_change(self,args,only_console = False,silent = False):
            """
            Volá get_current_ip_list(), aktualizuje současně nastavené adresy (self.current_address_list)
            - only console: vypíše do konzole aktuální připojení
            - silent: nevypisuje do konzole
            """
            if not only_console:
                try:
                    self.default_connection_option = self.connection_option_list.index(self.interface_drop_options.get())
                except ValueError as e:
                    print(f"Error: {e}")
                    self.default_connection_option = 0

                #pamatovat si naposledy zvoleny zpusob pripojeni:
                Tools.save_setting_parameter(parameter="change_def_conn_option",status=int(self.default_connection_option),excel_path=self.excel_file_path)
                self.current_address_list = main.IP_tools.get_current_ip_list(self.connection_option_list)
                if self.static_label2.winfo_exists():
                    self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
            if not silent:
                # ziskat data o aktualnim pripojeni
                current_connection = main.IP_tools.get_ipv4_addresses()
                message = ""
                for items in current_connection:
                    message = message + items + " "
                if message == "":
                    message = "nenalezeno"
                Tools.add_colored_line(self.main_console,f"Současné připojení: {message}","white",None,True)

        def make_project_first(self,purpouse=None,make_cells = True,project = None, input_entry_bypass = None,upwards=False,downwards=False):
            """
            purpouse:
            - search
            - silent
            """
            def check_position():
                max_position = len(self.all_rows)
                if upwards:
                    position = self.last_project_id -1
                elif downwards:
                    position = self.last_project_id +1

                if position < 0:
                    position = max_position-1
                elif position > max_position-1:
                    position = 0
                return position

            result = self.check_given_input(input_entry_bypass)
            self.remember_to_change_back = []
            self.last_selected_widget = ""

            if result == True:
                #zmena poradi
                if project == None:
                    project = self.all_rows[self.last_project_id]
                    favourite_status = self.favourite_list[self.last_project_id]
                else:
                    favourite_status = project[4]

                if downwards or upwards:
                    position = check_position()
                else:
                    position = 0

                self.all_rows.pop(self.last_project_id)
                self.all_rows.insert(position,project)
                self.favourite_list.pop(self.last_project_id)
                self.favourite_list.insert(position,favourite_status)
                self.last_project_id = position

                for i in range(0,len(self.all_rows)):
                    row = (len(self.all_rows)-1)-i
                    main.IP_tools.save_excel_data(self.excel_file_path,
                                                        len(self.all_rows),
                                                        self.last_project_id,
                                                        self.show_favourite,
                                                        self.all_rows[i][0],
                                                        self.all_rows[i][1],
                                                        self.all_rows[i][2],
                                                        self.all_rows[i][3],
                                                        force_row_to_print=row+1,
                                                        fav_status=self.favourite_list[i])
                if make_cells:
                    self.make_project_cells()
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} nalezen","green",None,True)
                elif purpouse != "silent":
                    Tools.add_colored_line(self.main_console,f"Projekt {self.all_rows[0][0]} přesunut na začátek","green",None,True)
            elif result == None and purpouse != "silent":
                print("nevlozeno id")
                if purpouse == "search":
                    Tools.add_colored_line(self.main_console,f"Vložte hledaný projekt do vyhledávání","orange",None,True)
                else:
                    Tools.add_colored_line(self.main_console,f"Nejprve vyberte projekt (nakliknout levým na parametry daného projektu nebo pravým na tlačíko projektu)","orange",None,True)
            elif purpouse != "silent":
                Tools.add_colored_line(self.main_console,"Projekt nenalezen","red",None,True)
                print("projekt nenalezen")

        def show_favourite_toggle(self,keep_search_input = False,determine_status = None): # hlavni prepinaci tlacitko oblibene/ neoblibene
            if self.show_favourite and (determine_status == None or determine_status == "all"):
                self.show_favourite = False
                window_status = 0
                self.last_project_name = ""
                self.last_project_ip = ""
                self.last_project_mask = ""
                self.last_project_notes = ""
                self.last_project_id = ""
                self.last_selected_widget = ""
                self.last_selected_notes_widget = ""
                self.last_selected_textbox = ""
                self.last_selected_widget_id = 0           
                self.ip_frame_list = []
                if keep_search_input == False:
                    self.search_input.delete("0","300")
                    self.search_input.configure(placeholder_text="Název projektu")
                    self.make_project_cells()
                else:
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    self.check_given_input() #check ve druhem prostredi
                    self.make_project_cells(no_read=True)
                self.button_remove_main.configure(command = lambda: self.delete_project(button_trigger=True,flag="main_menu"))
                Tools.save_setting_parameter(parameter="change_def_ip_window",status=window_status,excel_path=self.excel_file_path)
                self.button_switch_favourite_ip. configure(fg_color="black")
                self.button_switch_all_ip.       configure(fg_color="#212121")
                self.button_remove_main.         configure(text="Smazat")
                # poznámky mohou být None...
                if Tools.get_none_count(self.bin_projects[0]) < 2 and len(self.bin_projects[0]) > 3:
                    self.undo_button.configure(state = "normal")
                else:
                    self.undo_button.configure(state = "disabled")

            elif self.show_favourite == False and (determine_status == None or determine_status == "fav"):
                # favourite window
                self.show_favourite = True
                window_status = 1
                self.last_project_name = ""
                self.last_project_ip = ""
                self.last_project_mask = ""
                self.last_project_notes = ""
                self.last_project_id = ""
                self.last_selected_widget = ""
                self.last_selected_notes_widget = ""
                self.last_selected_textbox = ""
                self.last_selected_widget_id = 0
                self.ip_frame_list = []
                if keep_search_input == False:
                    self.search_input.delete("0","300")
                    self.search_input.configure(placeholder_text="Název projektu")
                    self.make_project_cells()
                else:
                    self.all_rows, self.project_list, self.favourite_list = main.IP_tools.read_excel_data(self.excel_file_path,self.show_favourite)
                    self.check_given_input() #check ve druhem prostredi
                    self.make_project_cells(no_read=True)
                self.button_remove_main.configure(command = lambda: self.switch_fav_status("with_refresh"))
                Tools.save_setting_parameter(parameter="change_def_ip_window",status=window_status,excel_path=self.excel_file_path)
                self.button_switch_favourite_ip. configure(fg_color="#212121")
                self.button_switch_all_ip.       configure(fg_color="black")
                self.button_remove_main.         configure(text="Odebrat")
                self.undo_button.                configure(state = "disabled")

        def refresh_interfaces(self,all = False):
            """
            - All parametr refreshne i statusy ip adres
            """
            interfaces_data = main.IP_tools.fill_interfaces()
            self.connection_option_list = interfaces_data[0]
            self.interface_drop_options.configure(values = self.connection_option_list)
            online_list_text = ""
            if len(interfaces_data[1]) > 0:
                for data in interfaces_data[1]:
                    online_list_text = online_list_text + str(data) +", "
                online_list_text = online_list_text[:-2] # odebrat čárku s mezerou

            self.online_list.configure(text=online_list_text)
            if all:
                self.option_change("")
            self.refresh_ip_statuses()

            return interfaces_data[1]

        def setting_window(self):
            def save_new_behav_notes():
                nonlocal checkbox
                if int(checkbox.get()) == 0:
                    self.default_note_behav = 0
                    Tools.save_setting_parameter(parameter="change_def_notes_behav",status=0,excel_path=self.excel_file_path)
                    self.make_project_cells()

                elif int(checkbox.get()) == 1:
                    self.default_note_behav = 1      
                    Tools.save_setting_parameter(parameter="change_def_notes_behav",status=1,excel_path=self.excel_file_path)
                    self.make_project_cells()

            def change_make_first_behav():
                nonlocal checkbox4
                if int(checkbox4.get()) == 0:
                    self.make_edited_project_first = False
                    Tools.save_setting_parameter(parameter="change_make_first_behav",status=0,excel_path=self.excel_file_path)
                elif int(checkbox4.get()) == 1:
                    self.make_edited_project_first = True
                    Tools.save_setting_parameter(parameter="change_make_first_behav",status=1,excel_path=self.excel_file_path)
            def delete_behav():
                if int(checkbox5.get()) == 0 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 100
                elif int(checkbox5.get()) == 0 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 101
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 0:
                    self.deletion_behav = 110
                elif int(checkbox5.get()) == 1 and int(checkbox6.get()) == 1:
                    self.deletion_behav = 111
                else:
                    return
                Tools.save_setting_parameter(parameter="delete_behav",status=self.deletion_behav,excel_path=self.excel_file_path)

            child_root = customtkinter.CTkToplevel()
            self.opened_window = child_root
            # x = self.root.winfo_rootx()
            # y = self.root.winfo_rooty()
            # child_root.geometry(f"580x400+{x+350}+{y+180}")
            child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(self.app_icon)))
            child_root.title("Nastavení")
            main_frame =    customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label =         customtkinter.CTkLabel(master = main_frame, width = 100,height=40,text = "Chování poznámek (editovatelné/ needitovatelné):",font=("Arial",20,"bold"))
            checkbox =      customtkinter.CTkCheckBox(master = main_frame, text = "Přímo zapisovat a ukládat do poznámek na úvodní obrazovce",font=("Arial",16,"bold"),command=lambda: save_new_behav_notes())
            label.          pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox.       pack(pady = 10,padx=10,side="top",anchor = "w")
            
            main_frame4 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label4 =        customtkinter.CTkLabel(master = main_frame4, width = 100,height=40,text = "Nastavení chování při editaci projektů:",font=("Arial",20,"bold"))
            checkbox4 =     customtkinter.CTkCheckBox(master = main_frame4, text = "Automaticky přesouvat editovaný projekt na začátek",font=("Arial",16,"bold"),command=lambda: change_make_first_behav())
            label4.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox4.      pack(pady = 10,padx=10,side="top",anchor = "w")

            main_frame5 =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#707070",border_width=2)
            label5 =        customtkinter.CTkLabel(master = main_frame5, width = 100,height=40,text = "Odvolit dotazování při mazání:",font=("Arial",20,"bold"))
            checkbox5 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit v hlavním okně",font=("Arial",16,"bold"),command=lambda: delete_behav())
            checkbox6 =     customtkinter.CTkCheckBox(master = main_frame5, text = "Odvolit při editu",font=("Arial",16,"bold"),command=lambda: delete_behav())
            label5.         pack(pady = 10,padx=10,side="top",anchor = "w")
            checkbox5.      pack(pady = 0,padx=10,side="top",anchor = "w")
            checkbox6.      pack(pady = (5,5),padx=10,side="top",anchor = "w")

            close_frame =   customtkinter.CTkFrame(master=child_root,corner_radius=0,border_color="#303030",border_width=2)
            button_close =  customtkinter.CTkButton(master = close_frame, width = 150,height=40,text = "Zavřít",command = child_root.destroy,font=("Arial",20,"bold"),corner_radius=0)
            button_close.   pack(pady = 10,padx=10,side="bottom",anchor = "e")

            main_frame.     pack(expand=False,fill="x",side="top")
            main_frame4.    pack(expand=False,fill="x",side="top")
            main_frame5.    pack(expand=False,fill="x",side="top")
            close_frame.    pack(expand=True,fill="both",side="top")

            if self.default_note_behav == 1:
                checkbox.select()
            if self.make_edited_project_first:
                checkbox4.select()
            if self.deletion_behav == 110 or self.deletion_behav == 111:
                checkbox5.select()
            if self.deletion_behav == 101 or self.deletion_behav == 111:
                checkbox6.select()

            child_root.update()
            child_root.update_idletasks()
            child_root.focus()
            child_root.focus_force()
            self.root.bind("<Button-1>",lambda e: child_root.destroy(),"+")

        def manual_ip_setting(self):
            window = customtkinter.CTkToplevel()
            window.after(200, lambda: window.iconbitmap(Tools.resource_path(self.app_icon)))
            self.opened_window = window
            window.title("Manuální nastavení IPv4 adresy")

            def check_ip_and_mask(input_value):
                input_splitted = input_value.split(".")
                if len(input_splitted) == 4:
                    return input_value
                else:
                    return False

            def call_ip_change():
                if "DHCP" in str(select_mode.get()):
                    self.change_to_DHCP()
                    window.destroy()
                    return
                
                ip_input = ip_address_entry.get()
                mask_input = mask_entry.get()
                ip_checked = check_ip_and_mask(ip_input)
                mask_checked = check_ip_and_mask(mask_input)
                errors = 0
                if ip_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná IP adresa","red",None,True)
                    errors += 1
                if mask_checked == False and errors == 0:
                    Tools.add_colored_line(manual_console,f"Neplatná maska","red",None,True)
                    errors += 1

                if errors == 0:
                    self.change_computer_ip(0,force_params=[ip_input,mask_input])
                    window.destroy()

            def call_option_change(*args):
                nonlocal ip_address_entry
                self.interface_drop_options.set(str(*args))
                self.option_change(*args)
                ip_address_entry.delete(0,300)
                ip_address_entry.insert(0,self.current_address_list[self.default_connection_option])
                check_interface_status()

            def switch_manual_dhcp(*args):
                nonlocal ip_address_entry
                nonlocal mask_entry
                if "DHCP" in str(*args):
                    ip_address_entry.configure(state = "disabled",text_color = "gray32")
                    mask_entry.configure(state = "disabled",text_color = "gray32")
                else:
                    ip_address_entry.configure(state = "normal",text_color = "gray84")
                    mask_entry.configure(state = "normal",text_color = "gray84")

            def check_interface_status(online_list = False):
                if online_list == False:
                    online_list = self.refresh_interfaces()

                found = False
                for items in online_list:
                    if items == str(select_interface.get()):
                        found = True
                        select_interface.configure(fg_color = "green",button_color = "green")
                        interface_status.configure(text = "Online")
                        break

                if not found:
                    select_interface.configure(fg_color = "red",button_color = "red")
                    interface_status.configure(text = "Offline")

            interface_label =       customtkinter.CTkLabel(master = window,text = "Manuálně nastavit IPv4 adresu pro: ",font=("Arial",20,"bold"))
            interface_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            select_interface =      customtkinter.CTkOptionMenu(master = interface_frame,width=320,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,command= lambda args:  call_option_change(args))
            interface_status =      customtkinter.CTkLabel(master = interface_frame,text = "",font=("Arial",20,"bold"))
            select_interface.       pack(pady=(10,0),padx=10,side = "left",anchor = "w")
            interface_status.       pack(pady=(10,0),padx=10,side = "left",anchor = "w")
            mode_label =            customtkinter.CTkLabel(master = window,text = "Způsob nastavení: ",font=("Arial",20,"bold"))
            select_mode =           customtkinter.CTkOptionMenu(master = window,width=400,height=50,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,values = ["manuálně","automaticky (DHCP)"],command= lambda args: switch_manual_dhcp(args))
            ip_address =            customtkinter.CTkLabel(master = window,text = "IPv4 adresa: ",font=("Arial",20,"bold"))
            ip_address_entry =      customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            mask =                  customtkinter.CTkLabel(master = window,text = "IPv4 maska: ",font=("Arial",20,"bold"))
            mask_entry =            customtkinter.CTkEntry(master = window,width=400,height=50,font=("Arial",20),corner_radius=0)
            manual_console =        tk.Text(window, wrap="none", height=0, width=36,background="black",font=("Arial",14),state=tk.DISABLED)
            buttons_frame =         customtkinter.CTkFrame(master = window,corner_radius=0,border_width=0,fg_color="#181818")
            save_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Nastavit", command = lambda: call_ip_change(),font=("Arial",20,"bold"),corner_radius=0)
            exit_button =           customtkinter.CTkButton(master = buttons_frame, width = 190,height=40,text = "Zrušit", command = lambda: window.destroy(),font=("Arial",20,"bold"),corner_radius=0)
            interface_label.        pack(pady=(10,0),padx=10,side = "top",anchor = "w",expand = False)
            interface_frame.        pack(pady=(0),padx=0,side = "top",anchor = "w")
            mode_label.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            select_mode.            pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            ip_address.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            ip_address_entry.       pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            mask.                   pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            mask_entry.             pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            manual_console.         pack(pady=(10,0),padx=10,side = "top",anchor = "w")
            buttons_frame.          pack(pady=(10),padx=10,side = "top",anchor = "e")
            save_button.            pack(pady=0,padx=(10,0),side = "right",anchor = "w")
            exit_button.            pack(pady=0,padx=0,side = "right",anchor = "e")
        
            online_list = self.refresh_interfaces()
            select_interface.configure(values = self.connection_option_list)
            select_interface.set(self.interface_drop_options.get())
            ip_address_entry.insert(0,self.current_address_list[self.default_connection_option])
            mask_entry.insert(0,"255.255.255.0")
            check_interface_status(online_list)
            
            self.root.bind("<Button-1>",lambda e: window.destroy(),"+")
            window.update()
            window.update_idletasks()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            window.geometry(f"{window.winfo_width()}x{window.winfo_height()}+{x+150}+{y+5}")

            window.focus_force()
            window.focus()

        def sort_by_alphabet(self):
            project_names_array=[]
            for projects in self.all_rows:
                project_names_array.append(projects[0])
            project_names_sorted = sorted(project_names_array)
            whole_projects_sorted = []
            for names in project_names_sorted:
                for projects in self.all_rows:
                    if projects[0] == names:
                        whole_projects_sorted.append(projects)
                        break
            
            self.all_rows = copy.deepcopy(whole_projects_sorted)            
            for i in range(0,len(self.all_rows)):
                row = (len(self.all_rows)-1)-i
                main.IP_tools.save_excel_data(self.excel_file_path,
                                                    len(self.all_rows),
                                                    self.last_project_id,
                                                    self.show_favourite,
                                                    self.all_rows[i][0],
                                                    self.all_rows[i][1],
                                                    self.all_rows[i][2],
                                                    self.all_rows[i][3],
                                                    force_row_to_print=row+1,
                                                    fav_status=self.favourite_list[i])
            
            self.make_project_cells()
            Tools.add_colored_line(self.main_console,f"Projekty seřazeny podle abecedy","green",None,True)

        def create_widgets(self,fav_status = None,init=None,excel_load_error = False):
            if not excel_load_error:
                if init:
                    if self.window_mode == "max":
                        Tools.save_setting_parameter(parameter="change_def_window_size",status=1,excel_path=self.excel_file_path)
                    else:
                        Tools.save_setting_parameter(parameter="change_def_window_size",status=0,excel_path=self.excel_file_path)
                if fav_status:
                    self.show_favourite = True
                    Tools.save_setting_parameter(parameter="change_def_ip_window",status=1,excel_path=self.excel_file_path)
                if fav_status == False:
                    self.show_favourite = False
                    Tools.save_setting_parameter(parameter="change_def_ip_window",status=0,excel_path=self.excel_file_path)
                Tools.save_setting_parameter(parameter="change_def_main_window",status=0,excel_path=self.excel_file_path)
            
            Tools.clear_frame(self.root)
            self.control_pressed = False
            menu_cards =                        customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50,border_width=0)
            self.main_widgets =                 customtkinter.CTkFrame(master=self.root,corner_radius=0,border_width=0)
            self.project_tree =                 customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0,border_width=0)
            logo =                              customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
            image_logo =                        customtkinter.CTkLabel(master = menu_cards,text = "",image =logo,bg_color="#212121")
            main_menu_button =                  customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "MENU",command =  lambda: self.call_menu(),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            self.button_switch_all_ip =         customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - všechny",command =  lambda: self.show_favourite_toggle(determine_status="all"),font=("Arial",25,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
            self.button_switch_favourite_ip =   customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "IP - oblíbené",command =  lambda: self.show_favourite_toggle(determine_status="fav"),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            button_switch_disk =                customtkinter.CTkButton(master = menu_cards, width = 200,height=50,text = "Síťové disky",command =  lambda: main.Disk_management_gui(parent=self.parent_instance),font=("Arial",25,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
            if excel_load_error:
                self.connection_option_list = ["data nenalezena"]
                self.show_favourite = False
                self.button_switch_all_ip.configure(state = "disabled")
                self.button_switch_favourite_ip.configure(state = "disabled")
                button_switch_disk.configure(state = "disabled")

            first_row_frame =           customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,border_width=0,fg_color="#212121")
            project_label =             customtkinter.CTkLabel(master = first_row_frame, width = 100,height=40,text = "Projekt: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.search_input =         customtkinter.CTkEntry(master = first_row_frame,font=("Arial",20),width=160,height=40,placeholder_text="Název projektu",corner_radius=0)
            button_search =             customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Vyhledat",command =  lambda: self.make_project_first("search"),font=("Arial",20,"bold"),corner_radius=0)
            self.button_add_main =      customtkinter.CTkButton(master = first_row_frame, width = 150,height=40,text = "Nový projekt", command = lambda: self.add_new_project(),font=("Arial",20,"bold"),corner_radius=0)
            self.button_remove_main =   customtkinter.CTkButton(master = first_row_frame, width = 100,height=40,text = "Smazat", command =  lambda: self.delete_project(button_trigger=True,flag="main_menu"),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_button =          customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_deleted_ip"),font=(None,28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_edit_main =          customtkinter.CTkButton(master = first_row_frame, width = 110,height=40,text = "Editovat",command =  lambda: self.edit_project(),font=("Arial",20,"bold"),corner_radius=0)
            self.undo_edit =            customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↶", command =  lambda: self.manage_bin(flag="load_edited_ip"),font=(None,28,"bold"),corner_radius=0,border_width=1,text_color="red")
            button_make_first =         customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "🔝",command =  lambda: self.make_project_first(),font=(None,30),corner_radius=0)
            move_upwards =              customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↑",command =  lambda: self.make_project_first(purpouse="silent",upwards=True),font=(None,25),corner_radius=0)
            move_downwards =            customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "↓",command =  lambda: self.make_project_first(purpouse="silent",downwards=True),font=(None,25),corner_radius=0)
            sort_alphabet =             customtkinter.CTkButton(master = first_row_frame, width = 50,height=40,text = "A↑",command =  lambda: self.sort_by_alphabet(),font=(None,25),corner_radius=0)
            button_settings_behav =     customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="⚙️",command =  lambda: self.setting_window(),font=(None,22),corner_radius=0)
            manual_ip_set =             customtkinter.CTkButton(master = first_row_frame, width = 40,height=40,text="Manuálně",command =  lambda: self.manual_ip_setting(),font=("Arial",20,"bold"),corner_radius=0)
            
            if self.show_favourite:
                self.button_remove_main.            configure(text="Odebrat")
                self.button_switch_favourite_ip.    configure(fg_color="#212121")
                self.button_switch_all_ip.          configure(fg_color="black")
                self.undo_button.                   configure(state = "disabled")
            else:
                self.button_remove_main.            configure(text="Smazat")
                self.button_switch_favourite_ip.    configure(fg_color="black")
                self.button_switch_all_ip.          configure(fg_color="#212121")
                # poznámky mohou být None (delete undo)
                if Tools.get_none_count(self.bin_projects[0]) < 2 and len(self.bin_projects[0]) == 5:
                    self.undo_button.configure(state = "normal")
                else:
                    self.undo_button.configure(state = "disabled")

            # edit undo
            if Tools.get_none_count(self.bin_projects[1]) < 2 and len(self.bin_projects[1]) == 5:
                self.undo_edit.configure(state = "normal")
            else:
                self.undo_edit.configure(state = "disabled")

            second_row_frame =              customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,border_width=0,fg_color="#212121")
            connect_label =                 customtkinter.CTkLabel(master = second_row_frame, width = 100,height=40,text = "Připojení: ",font=("Arial",20,"bold"),justify="left",anchor="w")
            self.interface_drop_options =   customtkinter.CTkOptionMenu(master = second_row_frame,width=200,height=40,font=("Arial",20,"bold"),dropdown_font=("Arial",20),corner_radius=0,command=  self.option_change)
            # "⚙️", "⚒", "🔧", "🔩"
            button_settings =               customtkinter.CTkButton(master = second_row_frame, width = 40,height=40,text="⚒",command =  lambda: self.refresh_interfaces(all=True),font=("",22),corner_radius=0) #refresh interface statusů
            button_dhcp =                   customtkinter.CTkButton(master = second_row_frame, width = 100,height=40,text = "DHCP",command =  lambda: self.change_to_DHCP(),font=("Arial",20,"bold"),corner_radius=0)
            static_label =                  customtkinter.CTkLabel(master = second_row_frame, height=40,text = "Static:",font=("Arial",20,"bold"))
            self.static_label2 =            customtkinter.CTkLabel(master = second_row_frame,width=200, height=40,text = "",font=("Arial",22,"bold"),bg_color="black")
            online_label =                  customtkinter.CTkLabel(master = second_row_frame, height=40,text = "Online: ",font=("Arial",22,"bold"))
            self.online_list =              customtkinter.CTkLabel(master = second_row_frame, height=40,text = "",font=("Arial",22,"bold"))
            third_row_frame =               customtkinter.CTkFrame(master=self.main_widgets,corner_radius=0,border_width=0,fg_color="#212121")
            self.main_console =             tk.Text(third_row_frame, wrap="none", height=0, width=180,background="black",font=("Arial",22),state=tk.DISABLED)
            main_menu_button.               pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            self.button_switch_all_ip.      pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            self.button_switch_favourite_ip.pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            button_switch_disk.             pack(pady = (10,0),padx =(10,0),anchor = "s",side = "left")
            image_logo.                     pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
            menu_cards.                     pack(pady=0,padx=5,fill="x",expand=False,side = "top")
            project_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.search_input.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_search.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.button_add_main.           pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.button_remove_main.        pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_button.               pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_edit_main.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.undo_edit.                 pack(pady = (10,0),padx =(0,0),anchor="w",side="left")
            button_make_first.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_upwards.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            move_downwards.                 pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            sort_alphabet.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_settings_behav.          pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            manual_ip_set.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            connect_label.                  pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.interface_drop_options.    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_settings.                pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            button_dhcp.                    pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            static_label.                   pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.static_label2.             pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            online_label.                   pack(pady = (10,0),padx =(20,0),anchor="w",side="left")
            self.online_list.               pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            self.main_console.              pack(pady = (10,0),padx =(5,0),anchor="w",side="left")
            first_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            second_row_frame.               pack(pady=0,padx=0,fill="x",side = "top")
            third_row_frame.                pack(pady=0,padx=0,fill="x",side = "top")
            self.main_widgets.              pack(pady=0,padx=0,fill="x",side = "top")
            self.project_tree.              pack(pady=5,padx=5,fill="both",expand=True,side = "top")

            self.refresh_interfaces() # aktualizace hodnot nabídky
            if self.default_connection_option < len(self.connection_option_list):
                self.interface_drop_options.set(self.connection_option_list[self.default_connection_option])# nastavení naposledy zvoleného interfacu
            else:
                self.default_connection_option = 0             
                Tools.save_setting_parameter(parameter="change_def_conn_option",status=0,excel_path=self.excel_file_path)

                self.interface_drop_options.set(self.connection_option_list[self.default_connection_option])

            if not excel_load_error:
                self.option_change("")
                self.make_project_cells()
                self.current_address_list = main.IP_tools.get_current_ip_list(self.connection_option_list)
                self.static_label2.configure(text=self.current_address_list[self.default_connection_option])
            else:
                only_name = self.excel_file_path.split("/")
                only_name = only_name[len(only_name)-1]
                Tools.add_colored_line(self.main_console,f"Konfigurační soubor: {only_name} nebyl nalezen nebo je otevřený","red",None,True)

            def maximalize_window(e):
                self.root.update_idletasks()
                current_width = int(self.root.winfo_width())
                # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
                if Tools.focused_entry_widget(self.root): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                    return
                if int(current_width) > 1200:
                    self.root.state('normal')
                    self.root.geometry(f"260x1000+{0}+{0}")
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=2,excel_path=self.excel_file_path)
                elif int(current_width) ==260:
                    self.root.geometry("1200x900")
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=0,excel_path=self.excel_file_path)
                else:
                    self.root.state('zoomed')
                    Tools.save_setting_parameter(parameter="change_def_window_size",status=1,excel_path=self.excel_file_path)

            self.root.bind("<f>",lambda e: maximalize_window(e))

            def unfocus_widget(e):
                self.root.focus_set()
            self.root.bind("<Escape>",unfocus_widget)
            self.search_input.bind("<Return>",unfocus_widget)

            def call_search(e):
                self.make_project_first("search")
            self.search_input.bind("<Return>",call_search)

            def call_unfocus(e):
                widget = str(e.widget)
                if not ".!ctkscrollableframe" in widget and not ".!ctktoplevel" in widget and not ".!ctkbutton" in widget:
                    #odebrat focus
                    self.clicked_on_project("",None,None,None,flag="unfocus")
                    return
                
            self.root.bind("<Button-1>",call_unfocus,"+")

            def control_button(status):
                self.control_pressed = status
                if status == True:
                    if not self.last_selected_widget_id in self.selected_list:
                        self.selected_list.append(self.last_selected_widget_id)

            def multi_select():
                if not self.last_project_id in self.selected_list:
                    self.selected_list.append(self.last_project_id)
                    print(self.selected_list)

            self.root.bind("<Control_L>",lambda e: control_button(True))
            self.root.bind("<Control-Button-1>",lambda e: multi_select())
            self.root.bind("<KeyRelease-Control_L>",lambda e: control_button(False))
            self.root.bind("<Delete>",lambda e: self.delete_project(button_trigger=True,flag="main_menu"))
            # self.root.mainloop()

if testing_mode:
    # IP_assignment(root,"","max",str(os.getcwd())+"\\",100)
    print(str(os.getcwd())+"\\")
    main(root,"","max",str(os.getcwd())+"\\",100)
    root.mainloop()