import customtkinter
import os
import time
from openpyxl import load_workbook
from PIL import Image
import Deleting_option_v1 as Deleting
import trimazkon_tray_MAZ as trimazkon_tray
import string_database_MAZ
from openpyxl import Workbook

from tkinter import filedialog
import tkinter as tk
import threading
# import shutil
import sys
import ctypes
import win32pipe, win32file, pywintypes, psutil
import subprocess
from win32api import *
from win32gui import *
import win32con
import struct

testing = False

trimazkon_tray_exe_name = "trimazkon_tray_v2.exe"
global_recources_load_error = False
exe_path = sys.executable
exe_name = os.path.basename(exe_path)
print("exe name: ",exe_name)
if testing:
    exe_name = "trimazkon_test.exe"

class Tools:
    @classmethod
    def path_check(cls,path_raw,only_repair = None):
        path=path_raw
        backslash = "\\"
        if backslash[0] in path:
            newPath = path.replace(backslash[0], '/')
            path = newPath
        if path.endswith('/') == False:
            newPath = path + "/"
            path = newPath
        #oprava mezery v nazvu
        path = r"{}".format(path)
        if not os.path.exists(path) and only_repair == None:
            return False
        else:
            return path

    @classmethod
    def get_all_app_processes(cls):
        pid_list = []
        num_of_apps = 0
        for process in psutil.process_iter(['pid', 'name']):
            # if process.info['name'] == "TRIMAZKON_test.exe":
            if process.info['name'] == exe_name:
                print(process.info)
                pid_list.append(process.info['pid'])
                num_of_apps+=1
        
        return [num_of_apps,pid_list]

    @classmethod
    def check_runing_app_duplicity(cls):
        """
        Spočte procesy a názvem aplikace, pokud je jich více, jak 2 je již aplikace spuštěná
        - v top případě zajistí aby se nenačítalo gui a pouze zajistí odeslání paramterů pro image browser
        """
        found_processes = Tools.get_all_app_processes()
        print("found processes (duplicity): ",found_processes)
        if found_processes[0] > 3:
            return True
        else:
            return False

    @classmethod
    def resource_path(cls,relative_path):
        """ Get the absolute path to a resource, works for dev and for PyInstaller """
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)
    
    @classmethod
    def create_new_config(cls,config_filename,setting_list_name,default_value_list,default_labels):
        # config_filename = "config_MAZ.xlsx"
        # setting_list_name = "Settings_recources"
        def insert_new_excel_param(wb,ws,row,param,text):
            """
            Zakládá nové, chybějící parametry
            """
            ws['A' + str(row)] = str(text)
            ws['B' + str(row)] = str(param)
            print(f'inserting new parameter to excel: {text}: {param}')
            # wb.save(initial_path+config_filename)

        wb = Workbook()
        ws = wb.create_sheet(title=setting_list_name)
        for i in range(0,len(default_value_list)):
            insert_new_excel_param(wb,ws,row=i+1,param=default_value_list[i],text=default_labels[i])

        wb.save(initial_path+config_filename)
        wb.close()

    @classmethod
    def read_config_data(cls): # Funkce vraci data z config_TRIMAZKON.
        """
        Funkce vrací data z konfiguračního souboru config_TRIMAZKON.xlsx

        data jsou v pořadí:

        0 path_repaired\n
        1 supported_formats_deleting\n
        2 files_to_keep\n
        3 cutoff_date\n
        4 dir_name_deleting\n
        5 maximalized\n
        6 sorting_safe_mode\n
        7 app zoom\n
        8 app zoom checkbox\n
        9 establish tray icon in startup\n
        """
        def filter_unwanted_chars(to_filter_data, directory = False,even_space=False):
            unwanted_chars = ["\n","\"","\'","[","]"]
            if directory:
                unwanted_chars = ["\n","\"","\'","[","]","\\","/"]
            if even_space:
                unwanted_chars.append(" ")
            filtered_data = ""
            for letters in to_filter_data:
                if letters not in unwanted_chars:
                    filtered_data += letters
            return filtered_data

        def load_default_values():
            output_array = [default_setting_parameters[0],
                            default_setting_parameters[1],
                            default_setting_parameters[2],
                            default_setting_parameters[3],
                            default_setting_parameters[4],
                            default_setting_parameters[5],
                            default_setting_parameters[6],
                            default_setting_parameters[7],
                            default_setting_parameters[8],
                            default_setting_parameters[9],
                            ]
            
            print("read intern database (default values)",output_array,len(output_array))
            return output_array

        def insert_new_excel_param(wb,ws,row,param,text):
            """
            Zakládá nové, chybějící parametry
            """
            ws['A' + str(row)] = text
            ws['B' + str(row)] = param
            print(f'inserting new parameter to excel: {text}: {param}')
            wb.save(initial_path+config_filename)

        global global_recources_load_error
        config_filename = "config_MAZ.xlsx"
        setting_list_name = "Settings_recources"
        default_setting_parameters = string_database_MAZ.default_setting_database_param
        default_labels = string_database_MAZ.default_setting_database

        if os.path.exists(initial_path+config_filename):
            try:
                cutoff_date = ["","",""]
                supported_formats_deleting = []
                wb = load_workbook(initial_path+config_filename)
                ws = wb[setting_list_name]

                # checking possibly missing parameters
                # value_check = ws['B' + str(30)].value
                # if value_check is None or str(value_check) == "":
                #     insert_new_excel_param(wb,ws,row=30,param=default_setting_parameters[29],text=default_labels[29])
                
                inserted_path = str(ws['B' + str(1)].value)
                path_repaired = Tools.path_check(inserted_path)
                if path_repaired == False:
                    path_repaired = default_setting_parameters[0]

                read_formats = str(ws['B' + str(2)].value)
                read_formats = filter_unwanted_chars(read_formats,even_space=True)
                found_formats = read_formats.split(",")
                for items in found_formats:
                    supported_formats_deleting.append(str(items))
                
                files_to_keep = int(default_setting_parameters[2])
                files_to_keep_raw = str(ws['B' + str(3)].value)
                if files_to_keep_raw.isdigit():
                    files_to_keep = int(files_to_keep_raw)

                cutoff_date_raw = str(ws['B' + str(4)].value)
                cutoff_date_filtered = filter_unwanted_chars(cutoff_date_raw,even_space=True)
                if "." in cutoff_date_filtered:
                    cutoff_date = cutoff_date_filtered.split(".")
                elif "," in cutoff_date_filtered:
                    cutoff_date = cutoff_date_filtered.split(",")

                dir_name_raw = str(ws['B' + str(5)].value)
                dir_name_filtered = filter_unwanted_chars(dir_name_raw,directory=True)
                if dir_name_filtered == "":
                    dir_name_filtered = default_setting_parameters[4]
                dir_name_deleting = dir_name_filtered

                maximalized = str(ws['B' + str(6)].value)
                if maximalized != "ano":
                    maximalized = "ne"

                safe_mode = str(ws['B' + str(7)].value)
                if safe_mode != "ne":
                    safe_mode = "ano"

                app_zoom = int(default_setting_parameters[7])
                app_zoom_raw = str(ws['B' + str(8)].value)
                if app_zoom_raw.isdigit():
                    app_zoom = int(app_zoom_raw)

                app_zoom_checkbox = str(ws['B' + str(9)].value)
                if app_zoom_checkbox != "ano":
                    app_zoom_checkbox = "ne"

                tray_startup_status = str(ws['B' + str(10)].value)
                if tray_startup_status != "ano":
                    tray_startup_status = "ne"

                global_recources_load_error = False
                output_array = [path_repaired,
                                supported_formats_deleting,
                                files_to_keep,
                                cutoff_date,
                                dir_name_deleting,
                                maximalized,
                                safe_mode,
                                app_zoom,
                                app_zoom_checkbox,
                                tray_startup_status,
                                ]
                
                print("read config",output_array,len(output_array))
                wb.close()
                return output_array

            except Exception as e:
                print(f"Nejdřív zavřete soubor {config_filename} Chyba: {e}")   
                print("Budou načteny defaultní hodnoty")
                global_recources_load_error = True
                output_array = load_default_values()
                return output_array
        else:
            print(f"Chybí konfigurační soubor {config_filename}")
            print("Bude vytvořen")
            # print("Budou načteny defaultní hodnoty")
            # global_recources_load_error = True
            Tools.create_new_config(config_filename,setting_list_name,default_setting_parameters,default_labels)
            output_array = load_default_values()
            return output_array
            
    @classmethod
    def save_to_config(cls,input_data,which_parameter): # Funkce zapisuje data do souboru config_TRIMAZKON.xlsx
        """
        Funkce zapisuje data do konfiguračního souboru

        vraci vystupni zpravu: report

        which_parameter je bud: 
        
        1 default_path\n
        2 add_supported_deleting_formats\n
        3 pop_supported_deleting_formats\n
        4 default_files_to_keep\n
        5 default_cutoff_date\n
        6 new_default_static_dir_name\n
        7 maximalized\n
        8 sorting_safe_mode\n
        9 app_zoom\n
        10 app_zoom_checkbox\n
        11 tray_icon_startup\n
        """

        def filter_unwanted_chars(to_filter_data, directory = False,formats = False):
            unwanted_chars = ["\n","\"","\'","[","]"]
            if directory:
                unwanted_chars = ["\n","\"","\'","[","]","\\","/"]
            if formats:
                unwanted_chars = ["\n","\"","\'","[","]"," ","."]

            filtered_data = ""
            for letters in to_filter_data:
                if letters not in unwanted_chars:
                    filtered_data += letters
            return filtered_data
        
        config_filename = "config_MAZ.xlsx"
        setting_list_name = "Settings_recources"
        parameter_row_mapping = {
            "default_path": 1,
            "add_supported_deleting_formats": 2,
            "pop_supported_deleting_formats": 2,
            "default_files_to_keep": 3,
            "default_cutoff_date": 4,
            "new_default_dir_name":5,
            "maximalized": 6,
            "sorting_safe_mode": 7,
            "app_zoom": 8,
            "app_zoom_checkbox": 9,
            "tray_icon_startup": 10,
            }
        
        if os.path.exists(initial_path + config_filename):
            wb = load_workbook(initial_path+config_filename)
            ws = wb[setting_list_name]
            formats_changes = False
            report = ""
            supported_formats_deleting = []
            found_formats = str(ws['B' + str(2)].value)
            found_formats = filter_unwanted_chars(found_formats,formats=True)
            found_formats = found_formats.split(",")
            supported_formats_deleting = found_formats
            
            if which_parameter == "add_supported_deleting_formats":
                corrected_input = filter_unwanted_chars(str(input_data),formats=True)
                if str(corrected_input) not in supported_formats_deleting:
                    supported_formats_deleting.append(str(corrected_input))
                    report =  (f"Byl přidán formát: \"{corrected_input}\" do podporovaných formátů pro možnosti mazání")
                    formats_changes = True
                else:
                    report =  (f"Formát: \"{corrected_input}\" je již součástí podporovaných formátů možností mazání")
                
            elif which_parameter == "pop_supported_deleting_formats":
                poped = 0
                found = False
                range_to = len(supported_formats_deleting)
                for i in range(0,range_to):
                    if i < range_to:
                        if str(input_data) == supported_formats_deleting[i] and len(str(input_data)) == len(supported_formats_deleting[i]):
                            supported_formats_deleting.pop(i)
                            poped+=1
                            range_to = range_to - poped
                            report =  (f"Z podporovaných formátů možností mazání byl odstraněn formát: \".{input_data}\"")
                            formats_changes = True
                            found = True
                if found == False:
                    report =  (f"Formát: \"{input_data}\" nebyl nalezen v podporovaných formátech možností mazání, nemůže tedy být odstraněn")
            
            elif which_parameter == "default_path":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
                report = (f"Základní cesta přenastavena na: {str(input_data)}")
            
            elif which_parameter == "default_files_to_keep":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
            
            elif which_parameter == "default_cutoff_date":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data[0])+"."+str(input_data[1])+"."+str(input_data[2])

            elif which_parameter == "maximalized":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "new_default_static_dir_name":
                input_data_splitted = str(input_data).split(" | ")
                input_data = input_data_splitted[0]
                increment = int(input_data_splitted[1])
                cells_with_names = parameter_row_mapping.get(which_parameter)
                excel_row = cells_with_names[increment]
                input_filtered = filter_unwanted_chars(str(input_data),directory=True)
                ws['B' + str(excel_row)] = str(input_filtered)
            
            elif which_parameter == "sorting_safe_mode":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "app_zoom":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)
            
            elif which_parameter == "app_zoom_checkbox":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            elif which_parameter == "tray_icon_startup":
                ws['B' + str(parameter_row_mapping.get(which_parameter))] = str(input_data)

            if formats_changes:
                #navraceni poli zpet do stringu radku:
                deleting_formats_string = ""  
                for items in supported_formats_deleting:
                    if deleting_formats_string == "":
                        deleting_formats_string += str(items)
                    else:
                        deleting_formats_string += ("," + str(items))

                ws['B' + str(2)] = str(deleting_formats_string)

            wb.save(initial_path+config_filename)
            wb.close()
            return report
        
        else:
            print("Chybí konfigurační soubor config_TRIMAZKON.xlsx (nelze ukládat změny)")
            return "Chybí konfigurační soubor config_TRIMAZKON.xlsx (nelze ukládat změny)"
   
    @classmethod
    def browseDirectories(cls,visible_files,start_path=None): # Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat
        """
        Funkce spouští průzkumníka systému windows pro definování cesty, kde má program pracovat

        Vstupní data:

        0: visible_files = "all" / "only_dirs"\n
        1: start_path = None -optimalni, docasne se ulozi posledni nastavena cesta v exploreru

        Výstupní data:

        0: výstupní chybová hlášení
        1: opravená cesta
        2: nazev vybraneho souboru (option: all)
        """
        corrected_path = ""
        output= ""
        name_of_selected_file = ""
        text_file_data = Tools.read_config_data()
        if start_path == None:
            start_path = str(text_file_data[2]) #defaultni cesta
        else: # byla zadana docasna cesta pro explorer
            checked_path = Tools.path_check(start_path)
            if checked_path == False:
                output = "Změněná dočasná základní cesta pro explorer již neexistuje"
                start_path = str(text_file_data[2]) #defaultni cesta
            else:
                start_path = checked_path

        if start_path != False:
            if not os.path.exists(start_path):
                start_path = ""
                output="Konfigurační soubor obsahuje neplatnou cestu"

        else:
            output="Chybí konfigurační soubor config_TRIMAZKON.xlsx s počáteční cestou...\n"
            start_path=""

        # pripad vyberu files, aby byly viditelne
        if visible_files == "all":
            if(start_path != ""):
                foldername_path = filedialog.askopenfile(initialdir = start_path,title = "Klikněte na soubor v požadované cestě")
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:           
                foldername_path = filedialog.askopenfile(initialdir = "/",title = "Klikněte na soubor v požadované cestě")
                path_to_directory= ""
                if foldername_path != None:
                    path_to_file = str(foldername_path.name)
                    path_to_file_split = path_to_file.split("/")
                    i=0
                    for parts in path_to_file_split:
                        i+=1
                        if i<len(path_to_file_split):
                            if i == 1:
                                path_to_directory = path_to_directory + parts
                            else:
                                path_to_directory = path_to_directory +"/"+ parts
                        else:
                            name_of_selected_file = parts
                else:
                    output = "Přes explorer nebyla vložena žádná cesta"

        # pripad vyberu slozek
        if visible_files == "only_dirs":
            if(start_path != ""):
                path_to_directory = filedialog.askdirectory(initialdir = start_path, title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"
            else:
                path_to_directory = filedialog.askdirectory(initialdir = "/", title = "Vyberte adresář")
                if path_to_directory == None or path_to_directory == "":
                    output = "Přes explorer nebyla vložena žádná cesta"

        check = Tools.path_check(path_to_directory)
        corrected_path = check
        return [output,corrected_path,name_of_selected_file]

    @classmethod
    def add_colored_line(cls,text_widget, text, color,font=None,delete_line = None,no_indent=None):
        """
        Vloží řádek do console
        """
        try:
            text_widget.configure(state=tk.NORMAL)
            if font == None:
                font = ("Arial",16)
            if delete_line != None:
                text_widget.delete("current linestart","current lineend")
                text_widget.tag_configure(color, foreground=color,font=font)
                text_widget.insert("current lineend",text, color)
            else:
                text_widget.tag_configure(color, foreground=color,font=font)
                if no_indent:
                    text_widget.insert(tk.END,text+"\n", color)
                else:
                    text_widget.insert(tk.END,"    > "+ text+"\n", color)

            text_widget.configure(state=tk.DISABLED)
        except Exception as e:
            print(f"Error při psaní do konzole: {e}")

    @classmethod
    def save_path(cls,console,path_entered):
        path_given = path_entered
        path_checked = Tools.path_check(path_given)
        if path_checked != False and path_checked != "/":
            console_input = Tools.save_to_config(path_checked,"default_path")
            Tools.add_colored_line(console,console_input,"green",None,True)
        elif path_checked != "/":
            Tools.add_colored_line(console,f"Zadaná cesta: {path_given} nebyla nalezena, nebude tedy uložena","red",None,True)
        elif path_checked == "/":
            Tools.add_colored_line(console,"Nebyla vložena žádná cesta k souborům","red",None,True)

    @classmethod
    def clear_console(cls,text_widget,from_where=None):
        """
        Vymaže celou consoli
        """
        if from_where == None:
            from_where = 1.0
        text_widget.configure(state=tk.NORMAL)
        text_widget.delete(from_where, tk.END)
        text_widget.configure(state=tk.DISABLED)

    @classmethod
    def check_task_existence_in_TS(cls,taskname):
        process = subprocess.Popen(f'schtasks /query /tn \"{taskname}\" /v /fo LIST',
                                                stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                creationflags=subprocess.CREATE_NO_WINDOW)
        stdout, stderr = process.communicate()
        try:
            stdout_str = stdout.decode('utf-8')
            stderr_str = stderr.decode('utf-8')
            data = str(stdout_str)
            error_data = str(stderr_str)
        except UnicodeDecodeError:
            try:
                stdout_str = stdout.decode('cp1250')
                stderr_str = stderr.decode('cp1250')
                data = str(stdout_str)
                error_data = str(stderr_str)
            except UnicodeDecodeError:
                data = str(stdout)
                error_data = str(stderr)
        if "ERROR" in error_data:
            return False
        else:
            return True
    
    @classmethod
    def is_thread_running(cls,name):
        for thread in threading.enumerate():
            if thread.name == name:
                return True
        return False

    @classmethod
    def tray_startup_cmd(cls):
        """
        Sepnutí aplikace v system tray nabídce

        """
        Tray_thread_name = "Main_app_tray_thread"
        if Tools.is_thread_running(Tray_thread_name): # Pokud tray aplikace už běží nezapínej novou
            print("tray app is already running")
            return

        def call_tray_class():
            tray_app_instance = trimazkon_tray.tray_app_service(initial_path,Tools.resource_path('images/logo_TRIMAZKON.ico'),exe_name,"config_MAZ.xlsx")
            tray_app_instance.main()

        blocking_task = threading.Thread(target=call_tray_class,name=Tray_thread_name)
        blocking_task.start()
    
    @classmethod
    def establish_startup_tray(cls):
        """
        Sets the startup task of switching on the tray application icon
        - if it doesnt exist already
        """
        task_name = "jhv_MAZ_startup_tray_setup"
        task_presence = Tools.check_task_existence_in_TS(task_name)
        print("task presence: ",task_presence)

        if not task_presence:
            path_app_location = str(initial_path + exe_name)
            task_command = "\"" + path_app_location + " run_tray" + "\" /sc onlogon"
            process = subprocess.Popen(f"schtasks /Create /TN {task_name} /TR {task_command}",
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        creationflags=subprocess.CREATE_NO_WINDOW)
            
            stdout, stderr = process.communicate()
            output_message = "out"+str(stdout) +"err"+str(stderr)
            print(output_message)
            if "Access is denied" in output_message:
                return "need_access"
            
        Tools.tray_startup_cmd() # init sepnutí po prvním zavedení tasku
    
    # @classmethod
    # def stop_tray_thread(cls):
    #     Tray_thread_name = "Main_app_tray_thread"
    #     if Tools.is_thread_running(Tray_thread_name):
    #         trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path,Tools.resource_path('images/logo_TRIMAZKON.ico'),exe_name,"config_MAZ.xlsx")
    #         trimazkon_tray_instance.quit_application()
    #         print("tray app terminated")

    @classmethod
    def is_admin(cls):
        try:
            return ctypes.windll.shell32.IsUserAnAdmin()
        except:
            return False

    @classmethod
    def call_again_as_admin(cls,input_flag:str,window_title,main_title):
        def run_as_admin():# Vyžádání admin práv: nefunkční ve vscode
            if not Tools.is_admin():
                pid = "None"
                try:
                    pid = os.getpid()
                except Exception:
                    pass
                ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join([input_flag,str(pid)]), None, 1)
                sys.exit()

        def close_prompt(child_root):
            child_root.grab_release()
            child_root.destroy()

        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(app_icon)))
        # child_root.geometry(f"620x150+{300}+{300}")  
        child_root.title(window_title)
        label_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        proceed_label = customtkinter.CTkLabel(master = label_frame,text = main_title,font=("Arial",25),anchor="w",justify="left")
        proceed_label.pack(pady=5,padx=10,anchor="w",side = "left")
        button_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_yes =    customtkinter.CTkButton(master = button_frame,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: run_as_admin())
        button_no =     customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
        button_no       .pack(pady = 5, padx = 10,anchor="e",side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="e",side="right")
        label_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        button_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)

        child_root.update()
        child_root.update_idletasks()
        # child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{300}+{300}")
        child_root.focus()
        child_root.focus_force()
        child_root.grab_set()

    @classmethod
    def remove_task_from_TS(cls,name_of_task):
        cmd_command = f"schtasks /Delete /TN {name_of_task} /F"
        # subprocess.call(cmd_command,shell=True,text=True)

        process = subprocess.Popen(cmd_command,
                                   stdout=subprocess.PIPE,
                                   stderr=subprocess.PIPE,
                                   creationflags=subprocess.CREATE_NO_WINDOW)
            
        stdout, stderr = process.communicate()
        output_message = "out"+str(stdout) +"err"+str(stderr)
        print(output_message)
        if "Access is denied" in output_message:
            return "need_access"

    @classmethod
    def terminate_pid(cls,pid:int):
        print("pid to terminate: ",pid)

        try:
            process = psutil.Process(pid)
            process.terminate()
            process.wait(timeout=5)
            print(f"Process with PID {pid} terminated.")
        except psutil.NoSuchProcess:
            print(f"No process with PID {pid} found.")
        except psutil.AccessDenied:
            print(f"Permission denied to terminate PID {pid}.")
        except psutil.TimeoutExpired:
            print(f"Process with PID {pid} did not terminate in time.")
        
def get_init_path():
    initial_path = Tools.path_check(os.getcwd())
    if len(sys.argv) > 1: #spousteni pres cmd (kliknuti na obrazek) nebo task scheduler - mazání
        raw_path = str(sys.argv[0])
        initial_path = Tools.path_check(raw_path,True)
        initial_path_splitted = initial_path.split("/")
        initial_path = ""
        for i in range(0,len(initial_path_splitted)-2):
            initial_path += str(initial_path_splitted[i])+"/"
        print("SYSTEM: ",sys.argv)

    return initial_path

initial_path = get_init_path()
print("init path: ",initial_path)

class WindowsBalloonTip:
    """
    Systémová notifikace
    """
    def __init__(self, title, msg,app_icon):
        message_map = {
                win32con.WM_DESTROY: self.OnDestroy,
        }
        # Register the Window class.
        wc = WNDCLASS()
        hinst = wc.hInstance = GetModuleHandle(None)
        wc.lpszClassName = "PythonTaskbar"
        wc.lpfnWndProc = message_map # could also specify a wndproc.
        classAtom = RegisterClass(wc)
        # Create the Window.
        style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
        self.hwnd = CreateWindow( classAtom, "Taskbar", style, \
                0, 0, win32con.CW_USEDEFAULT, win32con.CW_USEDEFAULT, \
                0, 0, hinst, None)
        UpdateWindow(self.hwnd)
        # iconPathName = os.path.abspath(os.path.join( sys.path[0], "balloontip.ico" ))
        iconPathName = app_icon
        icon_flags = win32con.LR_LOADFROMFILE | win32con.LR_DEFAULTSIZE
        try:
            hicon = LoadImage(hinst, iconPathName, \
                    win32con.IMAGE_ICON, 0, 0, icon_flags)
        except:
            hicon = LoadIcon(0, win32con.IDI_APPLICATION)
        flags = NIF_ICON | NIF_MESSAGE | NIF_TIP
        nid = (self.hwnd, 0, flags, win32con.WM_USER+20, hicon, "tooltip")
        Shell_NotifyIcon(NIM_ADD, nid)
        Shell_NotifyIcon(NIM_MODIFY, \
                         (self.hwnd, 0, NIF_INFO, win32con.WM_USER+20,\
                          hicon, "Balloon  tooltip",msg,200,title))
        # self.show_balloon(title, msg)
        time.sleep(10)
        DestroyWindow(self.hwnd)
    def OnDestroy(self, hwnd, msg, wparam, lparam):
        nid = (self.hwnd, 0)
        Shell_NotifyIcon(NIM_DELETE, nid)
        PostQuitMessage(0) # Terminate the app.

def deleting_via_cmd():
    print("deleting system entry: ",sys.argv)
    task_name = str(sys.argv[2])
    deleting_path = str(sys.argv[3])
    max_days = int(sys.argv[4])
    files_to_keep = int(sys.argv[5])
    cutoff_date = Deleting.get_cutoff_date(days=max_days)
    text_file_data = Tools.read_config_data()
    supported_formats_deleting = text_file_data[1]
    to_delete_folder_name = text_file_data[4]

    del_instance = Deleting.whole_deleting_function(
        deleting_path,
        more_dirs=False,
        del_option=1,
        files_to_keep=files_to_keep,
        cutoff_date_given=cutoff_date,
        supported_formats=supported_formats_deleting,
        testing_mode=False,
        to_delete_folder_name=to_delete_folder_name
        )
    output_data = del_instance.main()
    output_message = f"|||Datum provedení: {output_data[3]}||Zkontrolováno: {output_data[0]} souborů||Starších: {output_data[1]} souborů||Smazáno: {output_data[2]} souborů"
    output_message_clear = f"Provedeno: {output_data[3]}\nZkontrolováno: {output_data[0]} souborů\nStarších: {output_data[1]} souborů\nSmazáno: {output_data[2]} souborů"
    print(output_message)

    icon_path = Tools.resource_path('images/logo_TRIMAZKON.ico')
    trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path,icon_path,exe_name,"config_MAZ.xlsx")
    trimazkon_tray_instance.save_new_log(task_name,output_message)

    WindowsBalloonTip("Bylo provedeno automatické mazání",
                        str(output_message_clear),
                        icon_path)
   
    return output_message_clear

class system_pipeline_communication: # vytvoření pipeline serveru s pipe názvem TRIMAZKON_pipe_ + pid (id systémového procesu)
    """
    aby bylo možné posílat běžící aplikaci parametry:
    - mám otevřené okno ip setting - kliknu na obrázek - jen pošlu parametry
    """
    def __init__(self,exe_name,no_server = False):
        self.root = None #define later (to prevend gui loading when 2 apps opened)
        self.current_pid = None
        self.exe_name = exe_name
        self.current_pid = os.getpid()
        if not no_server:
            self.start_server()
            # run_server_background = threading.Thread(target=self.start_server,)
            # run_server_background.start()

    def check_root_existence(self):
        try:
            if self.root.winfo_exists():
                return True
        except Exception as e:
            # if "main thread is not in main loop" in str(e):
            # new_root = start_new_root()
            return False

    def server(self,pipe_input):
        """
        Endless loop listening for commands
        """
        pipe_name = fr'\\.\pipe\{pipe_input}'
        while True:
            print(f"Waiting for a jhv_MAZ to connect on {pipe_name}...") 
            pipe = win32pipe.CreateNamedPipe(
                pipe_name,
                win32pipe.PIPE_ACCESS_DUPLEX,
                win32pipe.PIPE_TYPE_MESSAGE | win32pipe.PIPE_READMODE_MESSAGE | win32pipe.PIPE_WAIT,
                1,
                512,
                512,
                0,
                None
            )

            win32pipe.ConnectNamedPipe(pipe, None)
            print("jhv_MAZ connected.")

            try:
                while True:
                    hr, data = win32file.ReadFile(pipe, 64 * 1024)
                    received_data = data.decode()
                    print(f"Received: {received_data}")
                    if "Establish main menu gui" in received_data:
                        # self.root.after(0,menu.command_landed,received_data)
                        root_existance = self.check_root_existence()
                        print("root_status: ",root_existance)

                        if root_existance == True:
                            self.root.after(0,menu.menu(clear_root=True))
                        else:
                            start_new_root()
                            # self.root.after(0,menu.menu(clear_root=True))
                    else:
                        self.root.after(0,menu.command_landed,received_data)

            except pywintypes.error as e:
                if e.args[0] == 109:  # ERROR_BROKEN_PIPE
                    print("jhv_MAZ disconnected.")
            finally:
                # Close the pipe after disconnection
                win32file.CloseHandle(pipe)
            # Loop back to wait for new client connections

    def client(self,pipe_name_given,command,parameters):
        """
        odesílá zprávu
        """
        pipe_name = fr'\\.\pipe\{pipe_name_given}'
        handle = win32file.CreateFile(
            pipe_name,
            win32file.GENERIC_READ | win32file.GENERIC_WRITE,
            0,
            None,
            win32file.OPEN_EXISTING,
            0,
            None
        )
        if "Open image browser" in command:
            message = str(parameters[0]) + ",," + str(parameters[1])
            print("Message sent.")
            win32file.WriteFile(handle, message.encode())
        elif "Establish main menu gui" in command:
            message = "Establish main menu gui"
            print("Message sent.")
            win32file.WriteFile(handle, message.encode())

    def start_server(self):
        self.pipe_name = f"jhv_MAZ_pipe_{self.current_pid}"
        running_server = threading.Thread(target=self.server, args=(self.pipe_name,), daemon=True)
        running_server.start()
        time.sleep(0.5)  # Wait for the server to start

    def call_checking(self,command,parameters):
        """
        for every found process with name of an application: send given command
        """
        checking = Tools.get_all_app_processes()
        print("SYSTEM application processes: ",checking)
        # if it is running more then one application, execute (root + self.root)
        if checking[0]>2:
            pid_list = checking[1]
            # try to send command to every process which has application name
            for pids in pid_list:
                if pids != self.current_pid:
                    try:
                        pipe_name = f"jhv_MAZ_pipe_{pids}"
                        self.client(pipe_name,command,parameters)
                    except Exception:
                        pass
            return True
        else:
            return False

load_gui=True
print("SYSTEM: ",sys.argv)
if len(sys.argv) > 1:
    if sys.argv[1] == "deleting":
        deleting_output_message = deleting_via_cmd()
        load_gui = False
        sys.exit(f"0: {deleting_output_message}")
    
    elif sys.argv[1] == "run_tray":
        pipeline_duplex = system_pipeline_communication(exe_name)# potřeba spustit server, protože neběží nic
        Tools.tray_startup_cmd()
        load_gui = False

    # elif sys.argv[1] == "run_tray_admin":
    #     pipeline_duplex = system_pipeline_communication(exe_name)# potřeba spustit server, protože neběží nic
    #     Tools.tray_startup_cmd()
    #     load_gui = False
        # sys.exit(0)

    # elif len(sys.argv) > 2:
    #     if sys.argv[2] == "admin_menu":
    #         # Tools.tray_startup_cmd()
    #         print("spoustim pres admina, pid: ",os.getpid())
    #         print(Tools.get_all_app_processes())
    #         # load_gui = False
    #         sys.exit(0)

    elif sys.argv[1] == "settings_tray" or sys.argv[1] == "settings_tray_del" or sys.argv[1] == "admin_menu":
        pid = int(sys.argv[2])
        Tools.terminate_pid(pid) #vypnout thread s tray aplikací

    # elif sys.argv[1] == "tray_startup_call":
    #     tray_startup_cmd()
    #     load_gui = False
    #     sys.exit(0)


if load_gui:
    app_icon = Tools.resource_path('images/logo_TRIMAZKON.ico')

    # pipeline_duplex = system_pipeline_communication(exe_name)# Establishment of pipeline server for duplex communication between running applications
    app_running_status = Tools.check_runing_app_duplicity()
    print("already opened app status: ",app_running_status)
    # if len(sys.argv) > 1: # VÝJIMKA: pukud nové spuštění s admin právy načti i gui...
    #     if sys.argv[0] == sys.argv[1]:
    #         app_running_status = False

    if not app_running_status: # aplikace ještě neběží -> spustit server
        pipeline_duplex = system_pipeline_communication(exe_name)# Establishment of pipeline server for duplex communication between running applications
        customtkinter.set_appearance_mode("dark")
        customtkinter.set_default_color_theme("dark-blue")
        root=customtkinter.CTk()
        root.geometry("1200x900")
        root.title("jhv_MAZ v_1.0.0")
        root.wm_iconbitmap(Tools.resource_path(app_icon))

    else:# předání parametrů pipeline komunikací
        # pass
        # pipeline_duplex_instance = system_pipeline_communication(exe_name,no_server=True) # pokud už je aplikace spuštěná nezapínej server, trvá to...
        # if len(sys.argv) > 1:
        #     raw_path = str(sys.argv[1]) #klik na spusteni trimazkonu s admin právy
        #     if sys.argv[0] != sys.argv[1]: # pokud se nerovnají jedná se nejspíše o volání základního prohlížeče obrázků (spuštění kliknutím na obrázek...)
        #         IB_as_def_browser_path=Tools.path_check(raw_path,True)
        #         IB_as_def_browser_path_splitted = IB_as_def_browser_path.split("/")
        #         IB_as_def_browser_path = ""
        #         for i in range(0,len(IB_as_def_browser_path_splitted)-2):
        #             IB_as_def_browser_path += IB_as_def_browser_path_splitted[i]+"/"
        #         selected_image = IB_as_def_browser_path_splitted[len(IB_as_def_browser_path_splitted)-2]
        #         pipeline_duplex_instance.call_checking(f"Open image browser starting with image: {IB_as_def_browser_path}, {selected_image}",[IB_as_def_browser_path,selected_image])
        
        pipeline_duplex_instance = system_pipeline_communication(exe_name,no_server=True) # pokud už je aplikace spuštěná nezapínej server, trvá to...
        pipeline_duplex_instance.call_checking(f"Establish main menu gui",[])

def set_zoom(zoom_factor):
    try:
        root.after(0, lambda: customtkinter.set_widget_scaling(zoom_factor / 100))
        # customtkinter.set_widget_scaling(zoom_factor / 100)
    except Exception as e:
        print(f"error with zoom scaling: {e}")
    
    root.tk.call('tk', 'scaling', zoom_factor / 100)

class main_menu:
    def __init__(self,root,new_loop=False):
        self.root = root
        if not new_loop:
            pipeline_duplex.root = self.root # předání rootu do pipeline_duplex až ve chvílí, kdy je jasné, že aplikace není vícekrát spuštěná:
        self.config_filename = "config_MAZ.xlsx"
        setting_list_name = "Settings_recources"
        self.data_read_in_txt = Tools.read_config_data()
        self.database_downloaded = False
        self.ib_running = False
        self.run_as_admin = False
        #init spínání tray podle nastavení
        if self.data_read_in_txt[9] == "ano":
            task_success = Tools.establish_startup_tray()
            if str(task_success) == "need_access":
                self.run_as_admin = True
        else: # když nezaškrtnuto aut. spouštění ujisti se, že není nastavené - potřeba taky admin
            if Tools.check_task_existence_in_TS("jhv_MAZ_startup_tray_setup"):
                Tools.remove_task_from_TS("jhv_MAZ_startup_tray_setup")
        
    def clear_frames(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        
    def call_deleting_option(self):
        self.clear_frames()
        self.root.unbind("<f>")
        Deleting_option(self.root)

    def call_advanced_option(self,success_message = None):
        self.clear_frames()
        self.root.unbind("<f>")
        Advanced_option(self.root,tray_setting_status_message=success_message)

    def fill_changelog(self,change_log):
        # Iterate through each <string> element and print its text
        for string_element in string_database_MAZ.change_log_list:
            change_log.insert("current lineend",string_element + "\n")
        change_log.see(tk.END)

    def command_landed(self,command):
        """
        tato funkce přijímá příkazy z pipeline serveru
        """
        print("received in menu: ",command)
        params = command.split(",,")
        print("Image browser running status: ",self.ib_running)
        if self.ib_running == False:
            for widget in self.root.winfo_children():
                widget.destroy()
            self.root.unbind("<Button-1>")
            self.call_view_option(params[0],params[1])
        else:
            print("previous path: ",self.IB_class.image_browser_path)
            print("previous path: ",self.IB_class.IB_as_def_browser_path)
            print("previous image: ",self.IB_class.selected_image)
            print("new path: ",params[0])
            print("new image: ",params[1])

            for widget in self.root.winfo_children():
                widget.destroy()
            self.root.unbind("<Button-1>")
            self.call_view_option(params[0],params[1])

    def on_closing(self):
        # def run_command_outside_admin_context(task_command):
        #     command = f'cmd /min /C "set __COMPAT_LAYER=RUNASINVOKER && start \"\" \"{app_path}\""'
        #     subprocess.Popen(task_command, shell=True, creationflags=subprocess.CREATE_NO_WINDOW)

        if Tools.is_admin(): # pokud se vypíná admin app - vypnout i admin tray a zapnout bez práv
            data_read_in_config = Tools.read_config_data()
            if data_read_in_config[9] == "ano":
                task_name = "jhv_MAZ_startup_tray_setup"
                try:
                    run_task_command = f'schtasks /Run /TN "{task_name}"'
                    print("Running task with command:", run_task_command)
                    subprocess.run(run_task_command, shell=True)
                except:
                    pass
            Tools.terminate_pid(os.getpid()) #vypnout thread s tray aplikací
        else:
            self.root.destroy()

    def menu(self,initial=False,catalogue_downloaded = False,zoom_disable = False,clear_root=False): # Funkce spouští základní menu při spuštění aplikace (MAIN)
        """
        Funkce spouští základní menu při spuštění aplikace (MAIN)

        -obsahuje 3 rámce:

        list_of_menu_frames = [frame_with_buttons,frame_with_logo,frame_with_buttons_right]
        """
        # if root_given != None:
        # #     self.root = root_given
        #     try:
        #         print("existuje root given?",root_given.winfo_exists())
        #     except Exception as e:
        #         print("chyba  ",e)
        # try:
        #     print("existuje root?",self.root.winfo_exists())
        # except Exception as e:
        #     print("chyba  ",e)

        if clear_root:
            self.root.after(0, lambda:self.clear_frames())
            # for widget in root_given.winfo_children():
            #     widget.destroy()

        if self.data_read_in_txt[5] == "ano" and initial:
            self.root.after(0, lambda:self.root.state('zoomed')) # max zoom, porad v okne
            
        if self.data_read_in_txt[8] == "ne" and initial: # pokud není využito nastavení windows
            try:
                root.after(0, lambda: set_zoom(int(self.data_read_in_txt[7])))
            except Exception as e:
                print("error with menu scaling")

        frame_with_logo = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        logo = customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo = customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        
        frame_with_buttons_right = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        frame_with_buttons = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        frame_with_logo.pack(pady=0,padx=0,fill="x",side = "top")
        image_logo.pack()
        frame_with_buttons_right.pack(pady=0,padx=0,fill="both",expand=True,side = "right")
        frame_with_buttons.pack(pady=0,padx=0,fill="both",expand=True,side = "left")
        trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path,Tools.resource_path('images/logo_TRIMAZKON.ico'),exe_name,"config_MAZ.xlsx")
        new_deleting =         customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Nastavit nové mazání", command = lambda: self.call_deleting_option(),font=("Arial",25,"bold"))
        task_manager =         customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Zobrazit nastavené mazání", command = lambda: trimazkon_tray_instance.show_all_tasks(toplevel=True),font=("Arial",25,"bold"))
        deleting_history =      customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Zobrazit záznamy o mazání", command = lambda: trimazkon_tray_instance.show_task_log(),font=("Arial",25,"bold"))
        advanced_button =       customtkinter.CTkButton(master = frame_with_buttons, width = 400,height=100, text = "Nastavení", command = lambda: self.call_advanced_option(),font=("Arial",25,"bold"))
        change_log_label =      customtkinter.CTkLabel(master=frame_with_buttons_right, width= 600,height=50,font=("Arial",24,"bold"),text="Seznam posledně provedených změn: ")
        change_log =            customtkinter.CTkTextbox(master=frame_with_buttons_right, width= 600,height=450,fg_color="#212121",font=("Arial",20),border_color="#636363",border_width=3,corner_radius=0)
        resources_load_error =  customtkinter.CTkLabel(master=frame_with_buttons_right, width= 600,height=50,font=("Arial",24,"bold"),text=f"Nepodařilo se načíst konfigurační soubor ({self.config_filename})",text_color="red")
        new_deleting.           pack(pady =(105,0), padx=20,side="top",anchor="e")
        task_manager.           pack(pady =(10,0), padx=20,side="top",anchor="e")
        deleting_history.       pack(pady =(10,0), padx=20,side="top",anchor="e")
        advanced_button.        pack(pady = (10,0), padx=20,side="top",anchor="e")
        change_log_label.       pack(pady = (50,5), padx=20,side="top",anchor="w")
        change_log.             pack(pady =0,       padx=20,side="top",anchor="w")
        if global_recources_load_error:
            resources_load_error.pack(pady = (5,5), padx=20,side="top",anchor="w")

        self.fill_changelog(change_log)
        
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
            self.root.update()
        
        self.root.bind("<f>",maximalize_window)
        # initial promenna aby se to nespoustelo porad do kola pri navratu do menu (system argumenty jsou stále uložené v aplikaci)
        if len(sys.argv) > 1 and initial == True:
            raw_path = str(sys.argv[1])
            #klik na spusteni trimazkonu s admin právy
            # if sys.argv[0] == sys.argv[1]:
            #     self.call_ip_manager()
            if sys.argv[1] == "settings_tray":
                self.call_advanced_option(success_message="Automatické spouštění úspěšně nastaveno")
            elif sys.argv[1] == "settings_tray_del":
                self.call_advanced_option(success_message="Automatické spouštění úspěšně odstraněno")
            
        
        if self.run_as_admin:
            self.root.after(1000, lambda: Tools.call_again_as_admin("admin_menu","Upozornění","Aplikace vyžaduje práva pro nastavení aut. spouštění na pozadí\n     - možné změnit v nastavení\n\nPřejete si znovu spustit aplikaci, jako administrátor?"))
        if initial and len(sys.argv) == 1:
            self.root.after(100, lambda: self.call_deleting_option())
        try:
            self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
            self.root.mainloop()
        except Exception as e:
            print("already looped? ",e)

class Advanced_option: # Umožňuje nastavit základní parametry, které ukládá do config souboru
    """
    Umožňuje nastavit základní parametry, které ukládá do config souboru
    """
    def __init__(self,root,windowed=None,spec_location=None,path_to_remember = None,last_params = None,tray_setting_status_message = None):
        self.spec_location = spec_location
        self.path_to_remember = path_to_remember
        self.ib_last_params = last_params
        self.windowed = windowed
        self.root = root
        self.tray_setting_status_message = tray_setting_status_message
        self.unbind_list = []
        self.drop_down_prefix_dir_names_list = []
        self.drop_down_static_dir_names_list = []
        self.default_displayed_prefix_dir = "cam"
        self.default_displayed_static_dir = 0
        self.submenu_option = "default_path"
        self.text_file_data = Tools.read_config_data()
        self.default_dir_name = self.text_file_data[4]
        
        self.creating_advanced_option_widgets()
    
    def set_zoom(self,zoom_factor):
        try:
            root.after(0,customtkinter.set_widget_scaling(zoom_factor / 100))
        except Exception as e:
            print(f"error with zoom scaling: {e}")
        
        root.tk.call('tk', 'scaling', zoom_factor / 100)

    def call_menu(self): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu
        """
        self.list_of_frames = [self.top_frame,
                               self.bottom_frame_default_path,
                               self.menu_buttons_frame]
        for frames in self.list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()
        
        for binds in self.unbind_list:
            self.root.unbind(binds)
        menu.menu(zoom_disable = True)

    def clear_frame(self,frame): # Smaže widgets na daném framu
        """
        Smaže widgets na daném framu
        """
        try:
            children = frame.winfo_children()
        except Exception:
            return
        for widget in children:
            widget.destroy()

    def maximalized(self): # Nastavení základního spouštění (v okně/ maximalizované)
        option = self.checkbox_maximalized.get()
        if option == 1:
            Tools.save_to_config("ano","maximalized")
        else:
            Tools.save_to_config("ne","maximalized")
    
    def tray_startup_setup(self,main_console): # Nastavení základního spouštění (v okně/ maximalizované)
        option = self.tray_checkbox.get()
        if option == 1:
            Tools.save_to_config("ano","tray_icon_startup")
            new_task_success = Tools.establish_startup_tray()
            if str(new_task_success) == "need_access":
                menu.run_as_admin = True
                Tools.call_again_as_admin("settings_tray","Upozornění","Aplikace vyžaduje práva pro nastavení aut. spouštění na pozadí\n\n- přejete si znovu spustit aplikaci, jako administrátor?")
                main_console.configure(text = "Jsou vyžadována admin práva",text_color="red")
            else:
                # Tools.establish_startup_tray()
                menu.run_as_admin = False
                main_console.configure(text = "Automatické spouštění úspěšně nastaveno",text_color="green")

        else:
            Tools.save_to_config("ne","tray_icon_startup")
            remove_task_success = Tools.remove_task_from_TS("jhv_MAZ_startup_tray_setup")
            if str(remove_task_success) == "need_access":
                menu.run_as_admin = True
                Tools.call_again_as_admin("settings_tray_del","Upozornění","Aplikace vyžaduje práva pro odstranění aut. spouštění na pozadí\n\n- přejete si znovu spustit aplikaci, jako administrátor?")
                main_console.configure(text = "Jsou vyžadována admin práva",text_color="red")
            else:
                menu.run_as_admin = False
                main_console.configure(text = "Automatické spouštění úspěšně odstraněno",text_color="green")

    def refresh_main_window(self):
        self.clear_frame(self.root)
        self.clear_frame(self.current_root)
        self.current_root.destroy()
        # if self.spec_location == "image_browser":
        #     Image_browser(root=self.root,path_given=self.path_to_remember,params_given=self.ib_last_params)
        # elif self.spec_location == "converting_option":
        #     Converting_option(self.root)
        if self.spec_location == "deleting_option":
            Deleting_option(self.root)
        # elif self.spec_location == "sorting_option":
        #     Sorting_option(self.root)

    def setting_widgets(self,exception=False,main_console_text = "",main_console_text_color = "white",submenu_option = None): # samotné možnosti úprav parametrů uložených v config souboru
        """
        Nabídka možností úprav

        0 = default_path
        1 = set_folder_names
        2 = set_default_parametres
        3 = set_supported_formats

        """

        if self.tray_setting_status_message != None:
            main_console_text = self.tray_setting_status_message
            main_console_text_color = "green"

        self.clear_frame(self.bottom_frame_default_path)
        text_file_data = Tools.read_config_data()
        if exception == False:
            cutoff_date = text_file_data[3]
        else:
            cutoff_date = exception
        
        files_to_keep = text_file_data[2]
        deleting_dir_name = text_file_data[4]
        row_index = 0
        for buttons in self.option_buttons:
            buttons.configure(fg_color = "black")

        def call_browseDirectories(): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if select_by_dir.get() == 1:
                output = Tools.browseDirectories("only_dirs")
            else:
                output = Tools.browseDirectories("all")
            if str(output[1]) != "/":
                self.path_set.delete("0","200")
                self.path_set.insert("0", output[1])
                console_input = Tools.save_to_config(output[1],"default_path") # hlaska o nove vlozene ceste
                default_path_insert_console.configure(text="")
                default_path_insert_console.configure(text = "Aktuálně nastavená základní cesta k souborům: " + str(output[1]),text_color="white")
                main_console.configure(text="")
                main_console.configure(text=console_input,text_color="green")
            else:
                main_console.configure(text = str(output[0]),text_color="red")

        def save_path():
            path_given = str(self.path_set.get())
            path_checked = Tools.path_check(path_given)
            if path_checked != False and path_checked != "/":
                console_input = Tools.save_to_config(path_checked,"default_path")
                default_path_insert_console.configure(text="")
                default_path_insert_console.configure(text = "Aktuálně nastavená základní cesta k souborům: " + str(path_checked),text_color="white")
                main_console.configure(text="")
                main_console.configure(text=console_input,text_color="green")
            elif path_checked != "/":
                main_console.configure(text="")
                main_console.configure(text=f"Zadaná cesta: {path_given} nebyla nalezena, nebude tedy uložena",text_color="red")
            elif path_checked == "/":
                main_console.configure(text="")
                main_console.configure(text="Nebyla vložena žádná cesta k souborům",text_color="red")
        
        def select_path_by_file():
            select_by_file.select()
            select_by_dir.deselect()

        def select_path_by_dir():
            select_by_dir.select()
            select_by_file.deselect()

        def set_default_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(cutoff_date[1]))
                        if int(cutoff_date[0]) > max_days_in_month:
                            cutoff_date[0] = str(max_days_in_month)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Měsíc: " + str(input_month) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení měsíce jste nezadali číslo",text_color="red")

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        cutoff_date[0] = int(input_day)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Den: " + str(input_day) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení dne jste nezadali číslo",text_color="red")

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        cutoff_date[2] = int(input_year) + 2000
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    elif len(str(input_year)) == 4:
                        cutoff_date[2] = int(input_year)
                        main_console.configure(text="")
                        main_console.configure(text="Datum přenastaveno na: "+ str(cutoff_date[0])+ "."+str(cutoff_date[1])+"."+ str(cutoff_date[2]),text_color="green")
                    else:
                        main_console.configure(text="")
                        main_console.configure(text="Rok: " + str(input_year) + " je mimo rozsah",text_color="red")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="U nastavení roku jste nezadali číslo",text_color="red")

            Tools.save_to_config(cutoff_date,"default_cutoff_date")
            self.setting_widgets(False, main_console._text,main_console._text_color,submenu_option="set_default_parametres")

        def set_files_to_keep():
            nonlocal main_console
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    files_to_keep = int(input_files_to_keep)
                    Tools.save_to_config(files_to_keep,"default_files_to_keep")
                    main_console.configure(text="")
                    main_console.configure(text="Základní počet ponechaných starších souborů nastaven na: " + str(files_to_keep),text_color="green")
                    console_files_to_keep.configure(text = "Aktuálně nastavené minimum: "+str(files_to_keep),text_color="white")
                else:
                    main_console.configure(text="")
                    main_console.configure(text="Mimo rozsah",text_color="red")
            else:
                main_console.configure(text="")
                main_console.configure(text="Nazadali jste číslo",text_color="red")

            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_default_parametres")

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                cutoff_date[i-1]=items
            main_console.configure(text="")
            main_console.configure(text="Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)",text_color="orange")
            self.setting_widgets(cutoff_date,main_console._text,main_console._text_color,submenu_option="set_default_parametres")


        def add_format(which_operation):
            if which_operation == 1:
                new_format = str(formats_deleting_input.get())
                if new_format !="":
                    main_console_text_add = Tools.save_to_config(new_format,"add_supported_deleting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_add,text_color="white")
            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_supported_formats")

        def pop_format(which_operation):
            if which_operation == 1:
                format_to_delete = str(formats_deleting_input.get())
                if format_to_delete !="":
                    main_console_text_pop = Tools.save_to_config(format_to_delete,"pop_supported_deleting_formats")
                    main_console.configure(text="")
                    main_console.configure(text=main_console_text_pop,text_color="white")

            self.setting_widgets(False,main_console._text,main_console._text_color,submenu_option="set_supported_formats")
        
        def manage_app_zoom(*args):
            app_zoom_percent.configure(text = str(int(*args)) + " %")

        def windows_zoom_setting():
            def get_screen_dpi():
                user32 = ctypes.windll.user32
                user32.SetProcessDPIAware()  # Make sure the process is DPI aware
                hdc = user32.GetDC(0)
                dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # 88 is the index for LOGPIXELSX
                return dpi

            if checkbox_app_zoom.get() == 1:
                Tools.save_to_config("ano","app_zoom_checkbox")
                current_dpi = get_screen_dpi()
                if current_dpi == 96:
                    set_zoom(100)
                elif current_dpi == 120:
                    set_zoom(125)
                elif current_dpi == 144:
                    set_zoom(150)
                app_zoom_slider.configure(state = "disabled",button_color = "gray50",button_hover_color = "gray50")
            else:
                app_zoom_slider.configure(state = "normal",button_color = "#3a7ebf",button_hover_color = "#3a7ebf")
                Tools.save_to_config("ne","app_zoom_checkbox")
                set_zoom(int(app_zoom_slider.get()))

        if submenu_option == "default_path":
            self.option_buttons[0].configure(fg_color="#212121")
            row_index = 1
            first_option_frame =        customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            self.checkbox_maximalized = customtkinter.CTkCheckBox(master = first_option_frame,height=40,text = "Spouštět v maximalizovaném okně",command = lambda: self.maximalized(),font=("Arial",22,"bold"))
            first_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")

            tray_option_frame =         customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            self.tray_checkbox =        customtkinter.CTkCheckBox(master = tray_option_frame,height=40,text = "Spouštět TRIMAZKON na pozadí (v systémové nabídce \"tray_icons\") při zapnutí systému Windows?",command = lambda: self.tray_startup_setup(main_console),font=("Arial",22,"bold"))
            tray_option_frame.          pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")

            current_zoom = self.text_file_data[7]
            new_option_frame =          customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            new_option_frame.           pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            zomm_app_label =            customtkinter.CTkLabel(master = new_option_frame,height=20,text = "Nastavte celkové přiblížení aplikace:",justify = "left",font=("Arial",22,"bold"))
            checkbox_app_zoom =         customtkinter.CTkCheckBox(master = new_option_frame,height=40,text = "Použít nastavení Windows",command = lambda: windows_zoom_setting(),font=("Arial",22,"bold"))
            app_zoom_slider =           customtkinter.CTkSlider(master = new_option_frame,width=300,height=15,from_=60,to=200,number_of_steps= 14,command = lambda e: manage_app_zoom(e))
            app_zoom_percent =          customtkinter.CTkLabel(master= new_option_frame,height=20,text = str(current_zoom) + " %",justify = "left",font=("Arial",20))
            zomm_app_label.             grid(column =0,row=0,sticky = tk.W,pady =(10,10),padx=10)
            app_zoom_slider.            grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=10)
            app_zoom_percent.           grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=320)
            checkbox_app_zoom.          grid(column =0,row=1,sticky = tk.W,pady =(10,20),padx=400)

            second_option_frame =        customtkinter.CTkFrame(    master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            second_option_frame.         pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label5 =                    customtkinter.CTkLabel(     master = second_option_frame,height=40,text = "Nastavte základní cestu k souborům při spuštění:",justify = "left",font=("Arial",22,"bold"))
            explorer_settings_label =   customtkinter.CTkLabel(     master = second_option_frame,height=40,text = "Nastavení EXPLORERU: ",justify = "left",font=("Arial",20,"bold"))
            select_by_dir =             customtkinter.CTkCheckBox(  master = second_option_frame,height=40,text = "Vybrat cestu zvolením složky",font=("Arial",20),command = lambda: select_path_by_dir())
            select_by_file =            customtkinter.CTkCheckBox(  master = second_option_frame,height=40,text = "Vybrat cestu zvolením souboru (jsou viditelné při vyhledávání)",font=("Arial",20),command = lambda: select_path_by_file())
            self.path_set =             customtkinter.CTkEntry(     master = second_option_frame,width=800,height=40,font=("Arial",20),placeholder_text="")
            button_save5 =              customtkinter.CTkButton(    master = second_option_frame,width=100,height=40, text = "Uložit", command = lambda: save_path(),font=("Arial",22,"bold"))
            button_explorer =           customtkinter.CTkButton(    master = second_option_frame,width=100,height=40, text = "EXPLORER", command = lambda: call_browseDirectories(),font=("Arial",22,"bold"))
            default_path_insert_console=customtkinter.CTkLabel(     master = second_option_frame,height=40,text ="",justify = "left",font=("Arial",22),text_color="white")
            console_frame =             customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.              pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =              customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =            customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.             pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =   customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            self.checkbox_maximalized.  grid(column =0,row=row_index-1,sticky = tk.W,pady =20,padx=10)
            self.tray_checkbox.         grid(column =0,row=row_index-1,sticky = tk.W,pady =20,padx=10)
            label5.                     grid(column =0,row=row_index,sticky = tk.W,pady =(5,0),padx=10)
            explorer_settings_label.    grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            select_by_dir .             grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=250)
            select_by_file.             grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=550)
            self.path_set.              grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
            button_save5.               grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=815)
            button_explorer.            grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=920)
            default_path_insert_console.grid(column =0,row=row_index+3,sticky = tk.W,pady =10,padx=10)
            main_console.               grid(column =0,row=row_index+4,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.    pack(pady =5,padx=10,anchor = "e")
            select_by_dir.select()

            def save_path_enter_btn(e):
                save_path()
                self.current_root.focus_set()
            self.path_set.bind("<Return>",save_path_enter_btn)

            app_zoom_slider.set(self.text_file_data[7])
            if self.text_file_data[8] == "ano":
                checkbox_app_zoom.select()
                windows_zoom_setting()

            def slider_released(e):
                """
                save after the slider is released - it still opening and closing excel otherwise
                """
                if not checkbox_app_zoom.get() == 1:
                    current_zoom = int(app_zoom_slider.get())
                    Tools.save_to_config(current_zoom,"app_zoom")
                    self.set_zoom(current_zoom)

            app_zoom_slider.bind("<ButtonRelease-1>",lambda e: slider_released(e))

            if text_file_data[0] != False and text_file_data[0] != "/":
                default_path_insert_console.configure(text="Aktuálně nastavená základní cesta k souborům: " + str(text_file_data[0]),text_color="white")
                self.path_set.configure(placeholder_text=str(text_file_data[0]))
                self.path_set.delete("0","200")
                self.path_set.insert("0", str(text_file_data[0]))
            else:
                default_path_insert_console.configure(text="Aktuálně nastavená základní cesta k souborům v konfiguračním souboru je neplatná",text_color="red")
                self.path_set.configure(placeholder_text="Není nastavena žádná základní cesta")
            
            if text_file_data[5] == "ano":
                self.checkbox_maximalized.select()
            else:
                self.checkbox_maximalized.deselect()

            if text_file_data[9] == "ano":
                self.tray_checkbox.select()
            else:
                self.tray_checkbox.deselect()

        if submenu_option == "set_default_parametres":
            self.option_buttons[1].configure(fg_color="#212121")
            #widgets na nastaveni zakladniho poctu files_to_keep
            files_to_keep_console_text ="Aktuálně nastavené minimum: "+str(files_to_keep)
            second_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            second_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_files_to_keep =               customtkinter.CTkLabel(master = second_option_frame,height=50,text = "2. Nastavte základní počet ponechaných souborů, vyhodnocených jako starších:",justify = "left",font=("Arial",22,"bold"))
            files_to_keep_set =                 customtkinter.CTkEntry(master = second_option_frame,width=100,height=50,font=("Arial",20), placeholder_text= files_to_keep)
            button_save2 =                      customtkinter.CTkButton(master = second_option_frame,width=100,height=50, text = "Uložit", command = lambda: set_files_to_keep(),font=("Arial",22,"bold"))
            console_files_to_keep=              customtkinter.CTkLabel(master = second_option_frame,height=50,text =files_to_keep_console_text,justify = "left",font=("Arial",22))
            label_files_to_keep.                grid(column =0,row=row_index+3,sticky = tk.W,pady =10,padx=10)
            files_to_keep_set.                  grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
            button_save2.                       grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=115)
            console_files_to_keep.              grid(column =0,row=row_index+5,sticky = tk.W,pady =(0,10),padx=10)
            def files_to_keep_enter_btn(e):
                set_files_to_keep()
                self.current_root.focus_set()
            files_to_keep_set.bind("<Return>",files_to_keep_enter_btn)
            
            #widgets na nastaveni zakladniho dne
            third_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
            third_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_set_default_date =            customtkinter.CTkLabel(master = third_option_frame,height=50,text = "3. Nastavte základní datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",22,"bold"))
            set_day =                           customtkinter.CTkEntry(master = third_option_frame,width=40,height=50,font=("Arial",20), placeholder_text= cutoff_date[0])
            sep1 =                              customtkinter.CTkLabel(master = third_option_frame,height=50,width=10,text = ".",font=("Arial",22))
            set_month =                         customtkinter.CTkEntry(master = third_option_frame,width=40,height=50,font=("Arial",20), placeholder_text= cutoff_date[1])
            sep2 =                              customtkinter.CTkLabel(master = third_option_frame,height=50,width=10,text = ".",font=("Arial",22))
            set_year =                          customtkinter.CTkEntry(master = third_option_frame,width=60,height=50,font=("Arial",20), placeholder_text= cutoff_date[2])
            button_save_date =                  customtkinter.CTkButton(master = third_option_frame,width=100,height=50, text = "Uložit", command = lambda: set_default_cutoff_date(),font=("Arial",22,"bold"))
            insert_button =                     customtkinter.CTkButton(master = third_option_frame,width=285,height=50, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",22,"bold"))
            console_frame =                     customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.                      pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =                      customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =                    customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.                     pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =               customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))
            label_set_default_date.             grid(column =0,row=row_index+6,sticky = tk.W,pady =10,padx=10)
            set_day.                            grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=10)
            sep1.                               grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=55)
            set_month.                          grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=70)
            sep2.                               grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=115)
            set_year.                           grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=130)
            button_save_date.                   grid(column =0,row=row_index+7,sticky = tk.W,pady =0,padx=195)
            insert_button.                      grid(column =0,row=row_index+8,sticky = tk.W,pady =5,padx=10)
            main_console.                       grid(column =0,row=row_index+9,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.            pack(pady =5,padx=10,anchor = "e")

            def new_date_enter_btn(e):
                set_default_cutoff_date()
                self.current_root.focus_set()
            set_day.bind("<Return>",new_date_enter_btn)
            set_month.bind("<Return>",new_date_enter_btn)
            set_year.bind("<Return>",new_date_enter_btn)

        if submenu_option == "set_supported_formats":
            self.option_buttons[2].configure(fg_color="#212121")
            #widgets pro nastavovani podporovanych formatu
            supported_formats_deleting = "Aktuálně nastavené podporované formáty pro možnosti mazání: " + str(text_file_data[1])
            first_option_frame =                customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=20,corner_radius=0,border_width=1)
            first_option_frame.                 pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            label_supported_formats_deleting =  customtkinter.CTkLabel(master = first_option_frame,height=50,text = "1. Nastavte podporované formáty pro možnosti: MAZÁNÍ:",justify = "left",font=("Arial",22,"bold"))
            formats_deleting_input =            customtkinter.CTkEntry(master = first_option_frame,height=50,font=("Arial",20),width=200)
            button_save4 =                      customtkinter.CTkButton(master = first_option_frame,width=50,height=50, text = "Uložit", command = lambda: add_format(1),font=("Arial",22,"bold"))
            button_pop2 =                       customtkinter.CTkButton(master = first_option_frame,width=70,height=50, text = "Odebrat", command = lambda: pop_format(1),font=("Arial",22,"bold"))
            console_bottom_frame_4=             customtkinter.CTkLabel(master = first_option_frame,height=50,text =supported_formats_deleting,justify = "left",font=("Arial",22))
            label_supported_formats_deleting.   grid(column =0,row=row_index+1,sticky = tk.W,pady =10,padx=10)
            formats_deleting_input.             grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
            button_save4.                       grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=215)
            button_pop2.                        grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=290)
            console_bottom_frame_4.             grid(column =0,row=row_index+3,sticky = tk.W,pady =0,padx=10)

            console_frame =                     customtkinter.CTkFrame(master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1,fg_color="black")
            console_frame.                      pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top")
            main_console =                      customtkinter.CTkLabel(master = console_frame,height=20,text = str(main_console_text),text_color=str(main_console_text_color),justify = "left",font=("Arial",22))
            if self.windowed:
                save_frame =                    customtkinter.CTkFrame(     master = self.bottom_frame_default_path,height=50,corner_radius=0,border_width=1)
                save_frame.                     pack(pady=(10,0),padx=5,fill="x",expand=False,side = "top",anchor = "e")
                save_changes_button =               customtkinter.CTkButton(master = save_frame,width=150,height=40, text = "Aplikovat/ načíst změny", command = lambda: self.refresh_main_window(),font=("Arial",22,"bold"))

            main_console.                       grid(column =0,row=row_index+7,sticky = tk.W,pady =10,padx=10)
            if self.windowed:
                save_changes_button.            pack(pady =5,padx=10,anchor = "e")

            def add_or_rem_formats(e):
                self.current_root.focus_set()
            formats_deleting_input.bind("<Return>",add_or_rem_formats)

    def creating_advanced_option_widgets(self): # Vytváří veškeré widgets (advance option MAIN)
        if self.windowed:
            self.current_root=customtkinter.CTkToplevel()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            self.current_root.geometry(f"1250x900+{x+200}+{y+200}")
            self.current_root.title("Pokročilá nastavení")
            self.current_root.after(200, lambda: self.current_root.iconbitmap(Tools.resource_path(app_icon)))
        else:
            self.current_root = self.root
        self.bottom_frame_default_path   = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,border_width = 0)
        self.top_frame                   = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,border_width = 0)
        self.menu_buttons_frame          = customtkinter.CTkFrame(master=self.current_root,corner_radius=0,fg_color="#636363",height=50,border_width = 0)
        self.top_frame.                 pack(pady=(2.5,0),padx=5,fill="x",expand=False,side = "top")
        self.menu_buttons_frame.        pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.bottom_frame_default_path. pack(pady=(0,2.5),padx=5,fill="both",expand=True,side = "bottom")
        
        label0          = customtkinter.CTkLabel(master = self.top_frame,height=20,text = "Nastavte požadované parametry (nastavení bude uloženo i po vypnutí aplikace): ",justify = "left",font=("Arial",22,"bold"))
        main_menu_button =  customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "MENU",                  command =  lambda: self.call_menu(),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options0 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Základní nastavení",    command =  lambda: self.setting_widgets(submenu_option="default_path"),font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        # options1 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Názvy složek",          command =  lambda: self.setting_widgets(submenu_option="set_folder_names"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options2 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Počáteční parametry",   command =  lambda: self.setting_widgets(submenu_option="set_default_parametres"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        options3 =          customtkinter.CTkButton(master = self.menu_buttons_frame, width = 200,height=50,text = "Podporované formáty",   command =  lambda: self.setting_widgets(submenu_option="set_supported_formats"),font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        label0.             grid(column = 0,row=0,sticky = tk.W,pady =10,padx=10)
        shift_const = 210
        if not self.windowed:
            main_menu_button.grid(column = 0,row=0,pady = (10,0),padx =10,sticky = tk.W)
            shift_const = 0
        options0.           grid(column = 0,row=0,pady = (10,0),padx =220-shift_const,sticky = tk.W)
        # options1.           grid(column = 0,row=0,pady = (10,0),padx =430-shift_const,sticky = tk.W)
        options2.           grid(column = 0,row=0,pady = (10,0),padx =430-shift_const,sticky = tk.W)
        options3.           grid(column = 0,row=0,pady = (10,0),padx =640-shift_const,sticky = tk.W)
        self.option_buttons = [options0,options2,options3]

        if self.windowed and not global_recources_load_error:
            if self.spec_location == "image_browser":
                self.setting_widgets(submenu_option="set_image_browser_setting")
            else:
                self.setting_widgets(submenu_option="default_path")
        elif not global_recources_load_error:
            self.setting_widgets(submenu_option="default_path")
        elif global_recources_load_error:
            error_label = customtkinter.CTkLabel(master = self.bottom_frame_default_path,height=20,text = "Nepodařilo se načíst konfigurační soubor config_TRIMAZKON.xlsx (nastavení se nemá kam uložit)",justify = "left",font=("Arial",22,"bold"),text_color="red")
            error_label.grid(column = 0,row=0,pady = (10,0),padx =20,sticky = tk.W)
            options0.configure(state = "disabled")
            options2.configure(state = "disabled")
            options3.configure(state = "disabled")

        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.current_root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.current_root._current_width) > 1200:
                self.current_root.after(0, lambda:self.current_root.state('normal'))
                self.current_root.geometry("1250x900")
            else:
                self.current_root.after(0, lambda:self.current_root.state('zoomed'))
        self.current_root.bind("<f>",maximalize_window)
        self.unbind_list.append("<f>")

        def unfocus_widget(e):
            self.current_root.focus_set()
        self.current_root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")

        if self.windowed:
            self.current_root.update()
            self.current_root.update_idletasks()
            self.current_root.focus_force()
            self.current_root.focus()
            # click outside the window - kill it
            self.root.bind("<Button-1>",lambda e: self.current_root.destroy())

class Deleting_option: # Umožňuje mazat soubory podle nastavených specifikací
    """
    Umožňuje mazat soubory podle nastavených specifikací

    -obsahuje i režim testování, kde soubory pouze přesune do složky ke smazání
    -umožňuje procházet více subsložek
    
    """
    @classmethod
    def confirm_window(cls,prompt_message):
        selected_option = False
        def selected_yes(child_root):# Vyžádání admin práv: nefunkční ve vscode
            child_root.grab_release()
            child_root.destroy()
            nonlocal selected_option
            selected_option = True

        def close_prompt(child_root):
            child_root.grab_release()
            child_root.destroy()
            nonlocal selected_option
            selected_option = False
            
        child_root = customtkinter.CTkToplevel()
        child_root.after(200, lambda: child_root.iconbitmap(Tools.resource_path(app_icon)))
        # child_root.geometry(f"620x150+{300}+{300}")  
        child_root.title("Upozornění (první spuštění aplikace)")
        label_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        proceed_label = customtkinter.CTkLabel(master = label_frame,text = prompt_message,font=("Arial",25),anchor="w",justify="left")
        proceed_label.pack(pady=5,padx=10,anchor="w",side = "left")
        button_frame = customtkinter.CTkFrame(master = child_root,corner_radius=0)
        button_yes =   customtkinter.CTkButton(master = button_frame,text = "ANO",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: selected_yes(child_root))
        button_no =    customtkinter.CTkButton(master = button_frame,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  close_prompt(child_root))
        button_no      .pack(pady = 5, padx = 10,anchor="e",side="right")
        button_yes     .pack(pady = 5, padx = 10,anchor="e",side="right")
        label_frame    .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        button_frame   .pack(pady=0,padx=0,anchor="w",side = "top",fill="x",expand=True)
        child_root.update()
        child_root.update_idletasks()
        # child_root.geometry(f"{child_root.winfo_width()}x{child_root.winfo_height()}+{300}+{300}")
        child_root.focus()
        child_root.focus_force()
        child_root.grab_set()
        child_root.wait_window()
        return selected_option

    def __init__(self,root):
        self.root = root
        self.unbind_list = []
        self.text_file_data = Tools.read_config_data()
        self.supported_formats_deleting = self.text_file_data[1]
        self.files_to_keep = self.text_file_data[2]
        self.cutoff_date = self.text_file_data[3]
        self.to_delete_folder_name = self.text_file_data[4]
        self.console_frame_right_1_text = "","white"
        self.console_frame_right_2_text = "","white"
        self.config_filename = "config_MAZ.xlsx"
        self.temp_path_for_explorer = None
        self.create_deleting_option_widgets()
 
    def call_extern_function(self,list_of_frames,function:str): # Tlačítko menu (konec, návrat do menu)
        """
        Funkce čistí všechny zaplněné rámečky a funguje, jako tlačítko zpět do menu\n
        function:
        - menu
        - sorting
        - (deleting)
        - converting
        """
        
        for frames in list_of_frames:
            frames.pack_forget()
            frames.grid_forget()
            frames.destroy()
        
        for binds in self.unbind_list:
            self.root.unbind(binds)

        if function == "menu":
            menu.menu()
        # elif function == "sorting":
        #     Sorting_option(self.root)
        # elif function == "converting":
        #     Converting_option(self.root)

    def start(self):# Ověřování cesty, init, spuštění
        """
        Ověřování cesty, init, spuštění
        """
        if self.checkbox.get()+self.checkbox2.get()+self.checkbox3.get() == 0:
            Tools.add_colored_line(self.console,"Nevybrali jste žádný způsob mazání","red")
            self.info.configure(text = "")

        else:
            path = self.path_set.get() 
            if path != "":
                check = Tools.path_check(path)
                if check == False:
                    Tools.add_colored_line(self.console,"Zadaná cesta: "+str(path)+" nebyla nalezena","red")
                else:
                    path = check
                    if self.checkbox_testing.get() != 1:
                        if self.checkbox6.get() == 1 and self.checkbox3.get() != 1: # sublozky u adresaru
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání souborů v cestě:\n{path}\na procházet přitom i SUBSLOŽKY?"
                        elif self.checkbox3.get() == 1:
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání ADRESÁŘŮ v cestě:\n{path}"
                        else:
                            confirm_prompt_msg = f"Opravdu si přejete spustit navolené mazání souborů v cestě:\n{path}"
                        # confirm = tk.messagebox.askokcancel("Potvrzení", confirm_prompt_msg)
                        confirm = Deleting_option.confirm_window(confirm_prompt_msg)
                    else: # pokud je zapnut rezim testovani
                        confirm = True

                    if confirm == True:
                        Tools.add_colored_line(self.console,"- Provádím navolené možnosti mazání v cestě: " + str(path) + "\n","orange")
                        self.console.update_idletasks()
                        self.root.update_idletasks()
                        self.del_files(path)
                    else:
                        Tools.add_colored_line(self.console,"Zrušeno uživatelem","red")
            else:
                Tools.add_colored_line(self.console,"Nebyla vložena cesta k souborům","red")

    def del_files(self,path): # zde se volá externí script: Deleting
        testing_mode = True
        del_option = 0
        if self.checkbox.get() == 1:
            del_option = 1
        if self.checkbox2.get() == 1:
            del_option = 2
        if self.checkbox3.get() == 1:
            del_option = 3
        if self.checkbox6.get() == 1:
            self.more_dirs = True
        else:
            self.more_dirs = False
        if self.checkbox_testing.get() == 1:
            testing_mode = True
        else:
            testing_mode = False

        def call_deleting_main(whole_instance):
            whole_instance.main()

        running_deleting = Deleting.whole_deleting_function(
            path,
            self.more_dirs,
            del_option,
            self.files_to_keep,
            self.cutoff_date,
            self.supported_formats_deleting,
            testing_mode,
            self.to_delete_folder_name
            )

        run_del_background = threading.Thread(target=call_deleting_main, args=(running_deleting,))
        run_del_background.start()

        completed = False
        previous_len = 0

        while not running_deleting.finish or completed == False:
            time.sleep(0.05)
            if int(len(running_deleting.output)) > previous_len:
                new_row = str(running_deleting.output[previous_len])
                if "Mazání dokončeno" in new_row or "Zkontrolováno" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"green",("Arial",15,"bold"))
                elif "Chyba" in new_row or "Nebyly nalezeny" in new_row or "- zrušeno" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"red",("Arial",15,"bold"))
                elif "Smazalo by se" in new_row or "Smazáno souborů" in new_row:
                    Tools.add_colored_line(self.console,str(new_row),"orange",("Arial",15,"bold"))
                else:
                    Tools.add_colored_line(self.console,str(new_row),"white")
                self.console.update_idletasks()
                self.root.update_idletasks()
                previous_len +=1

            if running_deleting.finish and (int(len(running_deleting.output)) == previous_len):
                completed = True
        
        run_del_background.join()

    def call_browseDirectories(self): # Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
        """
        if self.checkbox6.get() == 1: # pokud je zvoleno more_dirs v exploreru pouze slozky...
            output = Tools.browseDirectories("only_dirs",self.temp_path_for_explorer)
        else:
            output = Tools.browseDirectories("all",self.temp_path_for_explorer)
        if str(output[1]) != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", output[1])
            Tools.add_colored_line(self.console,f"Byla vložena cesta: {output[1]}","green")
            self.temp_path_for_explorer = output[1]
        else:
            Tools.add_colored_line(self.console,str(output[0]),"red")

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def selected(self,clear:bool): # První možnost mazání, od nejstarších
        """
        Nastavení widgets pro první možnost mazání

        -vstup: console text a barva textu

        -Budou smazány soubory starší než nastavené datum, přičemž bude ponechán nastavený počet souborů, vyhodnocených, jako starších\n
        -Podporované formáty jsou uživatelem nastavené a uložené v textovém souboru
        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        #self.console.configure(text = "")
        Tools.clear_console(self.console)
        self.checkbox.select()
        self.checkbox2.deselect()
        self.checkbox3.deselect()
        self.info.configure(text = f"- Budou smazány soubory starší než nastavené datum, přičemž bude ponechán nastavený počet souborů, vyhodnocených, jako starších\n(Ponechány budou všechny novější než nastavené datum a k tomu bude ponecháno: {self.files_to_keep} starších souborů)\nPodporované formáty: {self.supported_formats_deleting}\n\n",
                            font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            # if set_month.get() == self.cutoff_date[1] and set_day.get() == self.cutoff_date[0] and set_day.get() == self.cutoff_date[2]
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení měsíce jste nezadali číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení dne jste nezadali číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "U nastavení roku jste nezadali číslo","red"

            self.selected(False)

        def set_files_to_keep():
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)
                    self.console_frame_right_2_text = "Počet ponechaných starších souborů nastaven na: " + str(self.files_to_keep),"green"
                else:
                    self.console_frame_right_2_text = "Mimo rozsah","red"
            else:
                self.console_frame_right_2_text = "Nazadali jste číslo","red"

            self.selected(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)","orange"

            self.selected(False)

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)

        def set_max_days():
            new_cutoff = Deleting.get_cutoff_date(int(max_days_entry.get()))
            set_day.insert(0,new_cutoff[0])
            set_month.insert(0,new_cutoff[1])
            set_year.insert(0,new_cutoff[2])
            set_cutoff_date()

        console_1_text, console_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        label1      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",16))
        set_day     = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month   = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year    = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1 = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        max_days_entry = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        max_days_label = customtkinter.CTkLabel(master = self.frame_right,text = "dní",font=("Arial",16))
        max_days_save = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_max_days(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1 = customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_1_text,justify = "left",font=("Arial",18),text_color=console_1_color)
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=150)
        max_days_entry.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=260)
        max_days_label.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=320)
        max_days_save.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=350)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =5,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        max_days_entry.insert(0,Deleting.get_max_days(self.cutoff_date))

        console_2_text, console_2_color = self.console_frame_right_2_text
        label2          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte počet ponechaných souborů, vyhodnocených jako starších:",justify = "left",font=("Arial",16))
        files_to_keep_set = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.files_to_keep)
        button_save2    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_files_to_keep(),font=("Arial",18,"bold"))
        console_frame_right_2 = customtkinter.CTkLabel(master = self.frame_right,height=30,text =console_2_text,justify = "left",font=("Arial",18),text_color=console_2_color)
        label2.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)
        files_to_keep_set.grid(column =0,row=6,sticky = tk.W,pady =0,padx=10)
        button_save2.grid(column =0,row=6,sticky = tk.W,pady =0,padx=60)
        console_frame_right_2.grid(column =0,row=7,sticky = tk.W,pady =0,padx=10)
        def new_FTK_enter_btn(e):
            set_files_to_keep()
        files_to_keep_set.bind("<Return>",new_FTK_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů
        
    def selected2(self,clear:bool): # Druhá možnost mazání, mazání všech starých, redukce nových
        """
        Nastavení widgets pro druhou možnost mazání

        -Budou smazány VŠECHNY soubory starší než nastavené datum, přičemž budou redukovány i soubory novější\n
        -Souborů, vyhodnocených, jako novější, bude ponechán nastavený počet\n
        -(vhodné při situacích rychlého pořizování velkého množství fotografií, kde je potřebné ponechat nějaké soubory pro referenci)\n
        -Podporované formáty jsou uživatelem nastavené a uložené v textovém souboru
        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        Tools.clear_console(self.console)
        self.checkbox.deselect()
        self.checkbox2.select()
        self.checkbox3.deselect()
        self.info.configure(text = f"- Budou smazány VŠECHNY soubory starší než nastavené datum, přičemž budou redukovány i soubory novější\n(Ošetřeno: pokud se v dané cestě nacházejí pouze starší soubory, než nastavené datum, zruší se mazání)\n- Souborů, vyhodnocených, jako novější, než nastavené datum, bude ponecháno: {self.files_to_keep}\n(vhodné při situacích rychlého pořizování velkého množství fotografií, kde je potřebné ponechat nějaké soubory pro referenci)\nPodporované formáty: {self.supported_formats_deleting}",font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"            
            self.selected2(False)

        def set_files_to_keep():
            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)
                    self.console_frame_right_2_text = "Počet ponechaných starších souborů nastaven na: " + str(self.files_to_keep),"green"
                else:
                    self.console_frame_right_2_text = "Mimo rozsah","red"
            else:
                self.console_frame_right_2_text = "Nazadali jste číslo","red"

            self.selected2(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny soubory vyhodnoceny, jako starší!)","orange"
            self.selected2(False)

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

            input_files_to_keep = files_to_keep_set.get()
            if input_files_to_keep.isdigit():
                if int(input_files_to_keep) >= 0:
                    self.files_to_keep = int(input_files_to_keep)

        def set_max_days():
            new_cutoff = Deleting.get_cutoff_date(int(max_days_entry.get()))
            set_day.insert(0,new_cutoff[0])
            set_month.insert(0,new_cutoff[1])
            set_year.insert(0,new_cutoff[2])
            set_cutoff_date()

        console_frame_right_1_text, console_frame_right_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        label1      = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení souborů, jako starších:",justify = "left",font=("Arial",16))
        set_day     = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month   = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2        = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year    = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1 = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        max_days_entry = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        max_days_label = customtkinter.CTkLabel(master = self.frame_right,text = "dní",font=("Arial",16))
        max_days_save = customtkinter.CTkButton(master = self.frame_right,width=100,height=30, text = "Nastavit", command = lambda: set_max_days(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1=customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_frame_right_1_text,justify = "left",font=("Arial",18),text_color=console_frame_right_1_color)
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=150)
        max_days_entry.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=260)
        max_days_label.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=320)
        max_days_save.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=350)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =5,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        max_days_entry.insert(0,Deleting.get_max_days(self.cutoff_date))
        
        console_frame_right_2_text, console_frame_right_2_color = self.console_frame_right_2_text
        label2          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte počet ponechaných novějších souborů:",justify = "left",font=("Arial",16))
        files_to_keep_set = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.files_to_keep)
        button_save2    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_files_to_keep(),font=("Arial",18,"bold"))
        console_frame_right_2=customtkinter.CTkLabel(master = self.frame_right,height=30,text =console_frame_right_2_text,justify = "left",font=("Arial",18),text_color=console_frame_right_2_color)
        label2.grid(column =0,row=5,sticky = tk.W,pady =0,padx=10)
        files_to_keep_set.grid(column =0,row=6,sticky = tk.W,pady =0,padx=10)
        button_save2.grid(column =0,row=6,sticky = tk.W,pady =0,padx=60)
        console_frame_right_2.grid(column =0,row=7,sticky = tk.W,pady =0,padx=10)
        def new_FTK_enter_btn(e):
            set_files_to_keep()
        files_to_keep_set.bind("<Return>",new_FTK_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů
   
    def selected3(self,clear:bool): # Třetí možnost mazání, mazání datumových adresářů
        """
        Nastavení widgets pro třetí možnost mazání

        Budou smazány VŠECHNY adresáře (včetně všech subadresářů), které obsahují v názvu podporovaný formát datumu a jsou vyhodnoceny,jako starší než nastavené datum\n
        -Podporované datumové formáty jsou ["YYYYMMDD","DDMMYYYY","YYMMDD"] a podporované datumové separátory: [".","/","_"]

        """
        self.clear_frame(self.frame_right)
        self.bottom_frame2.unbind("<Enter>")
        Tools.clear_console(self.console)
        self.checkbox2.deselect()
        self.checkbox3.select()
        self.checkbox.deselect()
        self.info.configure(text = f"- Budou smazány VŠECHNY adresáře (včetně všech subadresářů), které obsahují v názvu podporovaný formát datumu a jsou vyhodnoceny,\njako starší než nastavené datum\nPodporované datumové formáty: {Deleting.supported_date_formats}\nPodporované separátory datumu: {Deleting.supported_separators}",font = ("Arial",16,"bold"),justify="left")
        self.selected6() #update

        if clear == True:
            self.console_frame_right_1_text = "","white"
            self.console_frame_right_2_text = "","white"

        def set_cutoff_date():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Měsíc: " + str(input_month) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))

            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Den: " + str(input_day) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)
                        self.console_frame_right_1_text = "Datum přenastaveno na: "+ str(self.cutoff_date[0])+ "."+str(self.cutoff_date[1])+"."+ str(self.cutoff_date[2]),"green"
                    else:
                        self.console_frame_right_1_text = "Rok: " + str(input_year) + " je mimo rozsah","red"
                else:
                    self.console_frame_right_1_text = "Nezadali jste číslo","red"

                        
            self.selected3(False)

        def insert_current_date():
            today = Deleting.get_current_date()
            today_split = today[1].split(".")
            i=0
            for items in today_split:
                i+=1
                self.cutoff_date[i-1]=items

            self.console_frame_right_1_text = "Bylo vloženo dnešní datum (Momentálně všechny adresáře vyhodnoceny, jako starší!)","orange"

            self.selected3(False) #refresh

        def save_before_execution():
            input_month = set_month.get()
            if input_month != "":
                if input_month.isdigit():
                    if int(input_month) < 13 and int(input_month) > 0:
                        self.cutoff_date[1] = int(input_month)
                        max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
                        if int(self.cutoff_date[0]) > max_days_in_month:
                            self.cutoff_date[0] = str(max_days_in_month)

            input_day = set_day.get()
            max_days_in_month = Deleting.calc_days_in_month(int(self.cutoff_date[1]))
            if input_day != "":
                if input_day.isdigit():
                    if int(input_day) <= int(max_days_in_month) and int(input_day) > 0:
                        self.cutoff_date[0] = int(input_day)

            input_year = set_year.get()
            if input_year != "":
                if input_year.isdigit():
                    if len(str(input_year)) == 2:
                        self.cutoff_date[2] = int(input_year) + 2000
                    elif len(str(input_year)) == 4:
                        self.cutoff_date[2] = int(input_year)

        console_frame_right_1_text, console_frame_right_1_color = self.console_frame_right_1_text
        today = Deleting.get_current_date()
        row_index = 0
        label0          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Dnešní datum: "+today[1],justify = "left",font=("Arial",16,"bold"))
        images2         = customtkinter.CTkLabel(master = self.frame_right,text = "")
        label1          = customtkinter.CTkLabel(master = self.frame_right,height=20,text = "Nastavte datum pro vyhodnocení datumu v názvu adresářů, jako staršího:",justify = "left",font=("Arial",16))
        set_day         = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[0])
        sep1            = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_month       = customtkinter.CTkEntry(master = self.frame_right,width=30,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[1])
        sep2            = customtkinter.CTkLabel(master = self.frame_right,height=20,width=10,text = ".",font=("Arial",20))
        set_year        = customtkinter.CTkEntry(master = self.frame_right,width=50,height=30,font=("Arial",16), placeholder_text= self.cutoff_date[2])
        button_save1    = customtkinter.CTkButton(master = self.frame_right,width=50,height=30, text = "Nastavit", command = lambda: set_cutoff_date(),font=("Arial",18,"bold"))
        insert_button = customtkinter.CTkButton(master = self.frame_right,width=190,height=30, text = "Vložit dnešní datum", command = lambda: insert_current_date(),font=("Arial",18,"bold"))
        console_frame_right_1 = customtkinter.CTkLabel(master = self.frame_right,height=30,text = console_frame_right_1_text,justify = "left",font=("Arial",18),text_color=console_frame_right_1_color)
        directories     = customtkinter.CTkImage(Image.open(Tools.resource_path("images/directories.png")),size=(240, 190))
        label0.grid(column =0,row=row_index,sticky = tk.W,pady =0,padx=10)
        images2.grid(column =0,row=row_index,sticky = tk.W,pady =15,padx=600,rowspan=5)
        label1.grid(column =0,row=row_index+1,sticky = tk.W,pady =0,padx=10)
        set_day.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=10)
        sep1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=40)
        set_month.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=50)
        sep2.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=80)
        set_year.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=90)
        button_save1.grid(column =0,row=row_index+2,sticky = tk.W,pady =0,padx=140)
        insert_button.grid(column =0,row=row_index+3,sticky = tk.W,pady =0,padx=10)
        console_frame_right_1.grid(column =0,row=row_index+4,sticky = tk.W,pady =0,padx=10)
        images2.configure(image = directories)
        def new_date_enter_btn(e):
            set_cutoff_date()
        set_day.bind("<Return>",new_date_enter_btn)
        set_month.bind("<Return>",new_date_enter_btn)
        set_year.bind("<Return>",new_date_enter_btn)
        self.bottom_frame2.bind("<Enter>",lambda e: save_before_execution()) # případ, že se nestiskne uložit - aby nedošlo ke ztrátě souborů

    def selected6(self): # checkbox s dotazem procházet subsložky
        """
        checkbox s dotazem procházet subsložky
        """
        if self.checkbox6.get() == 1:
            if self.checkbox3.get() == 1:
                self.info2.configure(text = "- Pro tuto možnost třídění není tato funkce podporována",font=("Arial",16,"bold"),text_color="white")
            else:
                self.info2.configure(text = "- VAROVÁNÍ: Máte spuštěné možnosti mazání obrázkových souborů i ve všech subsložkách vložené cesty (max:6 subsložek)",font=("Arial",16,"bold"),text_color="yellow")
        else:
            self.info2.configure(text = "")

    def save_new_task(self):
        def call_browse_directories():
            """
            Volání průzkumníka souborů (kliknutí na tlačítko EXPLORER)
            """
            if os.path.exists(str(operating_path.get())):
                output = Tools.browseDirectories("only_dirs",start_path=str(operating_path.get()))
            else:
                output = Tools.browseDirectories("only_dirs")
            if str(output[1]) != "/":
                operating_path.delete(0,300)
                operating_path.insert(0, str(output[1]))
                Tools.add_colored_line(console,"Byla vložena cesta pro vykonávání úkolu","green",None,True)
            print(output[0])
            window.focus()
            window.focus_force()

        def save_task_to_config():
            if check_entry("",hour_format=True,input_char=str(frequency_entry.get())) == False:
                return
            def get_task_name(current_tasks):
                names_taken = []
                new_task_name = "jhv_MAZ_task_xx"
                for tasks in current_tasks:
                    names_taken.append(tasks[0])
                for i in range(1,100):
                    name_suggestion = "jhv_MAZ_task_" + str(i)
                    if not name_suggestion in names_taken:
                        new_task_name = name_suggestion
                        break
                return new_task_name

            def set_up_task_in_ts():
                def check_freq_format(freq_input):
                    input_splitted = freq_input.split(":")
                    if len(str(input_splitted[0])) == 1:
                        corrected = "0"+str(input_splitted[0]) +":"+ str(input_splitted[1])
                        return corrected
                    else:
                        return freq_input
                        
                name_of_task = new_task[0]
                repaired_freq_param = check_freq_format(str(new_task[4]))
                path_app_location = str(initial_path+"/"+exe_name) 
                # task_command = "/c start \""+ path_app_location+ " deleting " + name_of_task + " " + str(new_task[1]) + " " + str(new_task[2]) + " " + str(new_task[3]) + "\" /SC DAILY /ST " + repaired_freq_param
                task_command = "\""+ path_app_location+ " deleting " + name_of_task + " " + str(new_task[1]) + " " + str(new_task[2]) + " " + str(new_task[3]) + "\" /SC DAILY /ST " + repaired_freq_param
                process = subprocess.Popen(f"schtasks /Create /TN {name_of_task} /TR {task_command}",
                                            stdout=subprocess.PIPE,
                                            stderr=subprocess.PIPE,
                                            creationflags=subprocess.CREATE_NO_WINDOW)
                stdout, stderr = process.communicate()
                try:
                    stdout_str = stdout.decode('utf-8')
                    data = str(stdout_str)
                except UnicodeDecodeError:
                    try:
                        stdout_str = stdout.decode('cp1250')
                        data = str(stdout_str)
                    except UnicodeDecodeError:
                        data = str(stdout)
                output_message = "out"+str(stdout) +"err"+str(stderr)
                print(output_message)
                if "SUCCESS" in stdout_str:
                    os.startfile("taskschd.msc")
                    return True
                else:
                    return False

            current_tasks = trimazkon_tray_instance.read_config()
            wb = load_workbook(initial_path + self.config_filename)
            ws = wb["Task_settings"]
            # subexe_path = Tools.resource_path(trimazkon_tray_exe_name)
            # process_output = subprocess.run(subexe_path + " " + initial_path + " read_config",
            #                 creationflags=subprocess.CREATE_NO_WINDOW,
            #                 stdout=subprocess.PIPE,
            #                 text = True)
            # current_tasks = list(process_output.stdout)
            if len(current_tasks) == 0:
                current_tasks = []
            print("current tasks: ",current_tasks)
            new_task_name = get_task_name(current_tasks)
            new_task = [new_task_name,operating_path.get(),self.older_then_entry.get(),minimum_file_entry.get(),frequency_entry.get(),str(Deleting.get_current_date()[2]),""]
            current_tasks.insert(0,new_task)

            row_to_print = 1
            for tasks in current_tasks:
                ws['A' + str(row_to_print)] = tasks[0] # název tasku
                ws['B' + str(row_to_print)] = tasks[1] # cesta vykonavani
                ws['C' + str(row_to_print)] = tasks[2] # max days
                ws['D' + str(row_to_print)] = tasks[3] # min left
                ws['E' + str(row_to_print)] = tasks[4] # frequency
                ws['F' + str(row_to_print)] = tasks[5] # datum přidání tasku
                ws['G' + str(row_to_print)] = tasks[6] # log mazání (pocet smazanych,datum,seznam smazanych)
                row_to_print +=1
            try:
                success_status = set_up_task_in_ts()
                if success_status:
                    Tools.add_colored_line(console,"Nový úkol byl uložen a zaveden do task scheduleru","green",None,True)
                    wb.save(self.config_filename)
                else:
                    Tools.add_colored_line(console,"Neočekávaná chyba, nepovedlo se nastavit nový úkol","red",None,True)
                wb.close()

            except Exception as e:
                Tools.add_colored_line(console,f"Prosím zavřete konfigurační soubor ({e})","red",None,True)
                wb.close()

        def refresh_cutoff_date():
            self.older_then_entry.update()
            self.older_then_entry.update_idletasks()
            try:
                cutoffdate_list = Deleting.get_cutoff_date(int(self.older_then_entry.get()))
                new_date = "(starší než: "+str(cutoffdate_list[0])+"."+str(cutoffdate_list[1])+"."+str(cutoffdate_list[2])+")"
                if older_then_label3.cget("text") != new_date:
                    older_then_label3.configure(text = new_date)
            except Exception:
                pass

        def check_entry(event,number=False,hour_format=False,input_char=None,flag=""):
            if flag == "cutoff":
                window.after(100, lambda: refresh_cutoff_date())
            if event != "":
                if event.keysym == "BackSpace" or event.keysym == "Return":
                    return

            if number:
                if not event.char.isdigit():
                    Tools.add_colored_line(console,"Vkládejte pouze čísla","red",None,True)
                    event.widget.insert(tk.INSERT,"")
                    return "break"  # Stop the event from inserting the original character
                
            elif hour_format:
                if not ":" in input_char:
                    Tools.add_colored_line(console,"Neplatný formát času, chybí separátor (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif len(input_char.split(":")) != 2:
                    Tools.add_colored_line(console,"Neplatný formát času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif len(str(input_char.split(":")[1])) != 2:
                    Tools.add_colored_line(console,"Neplatný formát času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif not input_char.split(":")[0].isdigit() or not input_char.split(":")[1].isdigit():
                    Tools.add_colored_line(console,"Neplatné znaky u času (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                elif int(input_char.split(":")[0]) > 23 or int(input_char.split(":")[0]) < 0 or int(input_char.split(":")[1]) > 59 or int(input_char.split(":")[1]) < 0:
                    Tools.add_colored_line(console,"Neplatný formát času, mimo rozsah (vkládejte ve formátu: 00:00)","red",None,True)
                    return False
                
        window = customtkinter.CTkToplevel()
        window.after(200, lambda: window.iconbitmap(app_icon))
        window.title("Nastavení nového úkolu")
        trimazkon_tray_instance = trimazkon_tray.tray_app_service(initial_path,Tools.resource_path('images/logo_TRIMAZKON.ico'),exe_name,"config_MAZ.xlsx")

        parameter_frame = customtkinter.CTkFrame(master = window,corner_radius=0)
        path_label = customtkinter.CTkLabel(master = parameter_frame,text = "Zadejte cestu, kde bude úkol spouštěn:",font=("Arial",22,"bold"))
        path_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        operating_path = customtkinter.CTkEntry(master = path_frame,font=("Arial",20),height=50,corner_radius=0)
        explorer_btn = customtkinter.CTkButton(master = path_frame,text = "...",font=("Arial",22,"bold"),width = 40,height=50,corner_radius=0,command=lambda: call_browse_directories())
        path_label.pack(pady = (10,0),padx = (10,0),side="top",anchor="w")
        operating_path.pack(pady = (10,0),padx = (10,0),side="left",anchor="w",expand = True,fill="x")
        explorer_btn.pack(pady = (10,0),padx = (0,10),side="left",anchor="w")
        path_frame.pack(side="top",anchor="w",fill="x",expand = True)

        older_then_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        older_then_label = customtkinter.CTkLabel(master = older_then_frame,text = "Odstanit soubory starší než:",font=("Arial",22,"bold"))
        self.older_then_entry = customtkinter.CTkEntry(master = older_then_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        older_then_label2 = customtkinter.CTkLabel(master = older_then_frame,text = "dní",font=("Arial",22,"bold"))
        older_then_label3 = customtkinter.CTkLabel(master = older_then_frame,text = "",font=("Arial",22,"bold"))
        older_then_label.pack(pady = (10,0),padx = (10,10),side="left")
        self.older_then_entry.pack(pady = (10,0),padx = (0,0),side="left")
        older_then_label2.pack(pady = (10,0),padx = (10,0),side="left")
        older_then_label3.pack(pady = (10,0),padx = (10,10),side="left")
        older_then_frame.pack(side="top",fill="x",anchor="w")
        self.older_then_entry.bind("<Key>",lambda e: check_entry(e,number=True,flag="cutoff"))

        minimum_file_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        minimum_file_label = customtkinter.CTkLabel(master = minimum_file_frame,text = "Ponechat souborů:",font=("Arial",22,"bold"))
        minimum_file_entry = customtkinter.CTkEntry(master = minimum_file_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        minimum_file_label.pack(pady = (10,0),padx = (10,10),side="left")
        minimum_file_entry.pack(pady = (10,0),padx = (0,10),side="left")
        minimum_file_frame.pack(side="top",fill="x",anchor="w")
        minimum_file_entry.bind("<Key>",lambda e: check_entry(e,number=True))

        frequency_frame = customtkinter.CTkFrame(master = parameter_frame,corner_radius=0)
        frequency_label = customtkinter.CTkLabel(master = frequency_frame,text = "Frekvence: denně, ",font=("Arial",22,"bold"))
        frequency_entry = customtkinter.CTkEntry(master = frequency_frame,font=("Arial",20),width=100,height=40,corner_radius=0)
        frequency_label2 = customtkinter.CTkLabel(master = frequency_frame,text = "hodin (př.: 0:00, 6:00, 14:30)",font=("Arial",22,"bold"))
        frequency_label.pack(pady = (10,0),padx = (10,10),side="left",anchor="w")
        frequency_entry.pack(pady = (10,0),padx = (0,0),side="left",anchor="w")
        frequency_label2.pack(pady = (10,0),padx = (10,10),side="left",anchor="w")
        frequency_frame.pack(side="top",fill="x",anchor="w")
        console = tk.Text(parameter_frame, wrap="none", height=0, width=30,background="black",font=("Arial",22),state=tk.DISABLED)
        console.pack(pady = 10,padx =10,side="top",anchor="w",fill="x")

        button_frame =   customtkinter.CTkFrame(master = window,corner_radius=0)
        show_tasks_btn = customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zobrazit nastavené úkoly", command =  lambda: trimazkon_tray_instance.show_all_tasks(toplevel=True),font=("Arial",20,"bold"),corner_radius=0)
        # show_tasks_btn = customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zobrazit nastavené úkoly", command =  lambda: call_show_all_tasks(),font=("Arial",20,"bold"),corner_radius=0)
        save_task_btn =  customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Uložit nový úkol", command =  lambda: save_task_to_config(),font=("Arial",20,"bold"),corner_radius=0)
        cancel_btn =  customtkinter.CTkButton(master = button_frame, width = 300,height=50,text = "Zavřít", command =  lambda: window.destroy(),font=("Arial",20,"bold"),corner_radius=0)
        cancel_btn.   pack(pady=10,padx=(10,10),side="right",anchor="e")
        save_task_btn.   pack(pady=10,padx=(10,0),side="right",anchor="e")
        show_tasks_btn.  pack(pady=10,padx=(10,0),side="right",anchor="e")
        parameter_frame.pack(side="top",fill="both")
        button_frame.pack(side="top",fill="x")
        operating_path.insert("0",self.path_set.get())
        max_days = Deleting.get_max_days(self.cutoff_date)
        self.older_then_entry.insert("0",max_days)
        minimum_file_entry.insert("0",self.files_to_keep)
        frequency_entry.insert("0","12:00")
        refresh_cutoff_date()
        window.update()
        window.update_idletasks()
        window_width = window.winfo_width()
        if window_width < 1200:
            window_width = 1200
        window.geometry(f"{window_width}x{window.winfo_height()}")
        window.after(100,window.focus_force())
        window.focus()

    def create_deleting_option_widgets(self):  # Vytváří veškeré widgets (delete option MAIN)
        #definice ramcu

        top_frame =             customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#212121")
        frame_with_logo =       customtkinter.CTkFrame(master=top_frame,corner_radius=0)
        logo =                  customtkinter.CTkImage(Image.open(Tools.resource_path("images/jhv_logo.png")),size=(300, 100))
        image_logo =            customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        frame_with_logo.        pack(pady=0,padx=0,expand=False,side = "right")
        image_logo.pack(pady = 0,padx =(15,0),anchor = "e",side = "right",ipadx = 20,ipady = 10,expand=False)
        
        frame_with_cards =      customtkinter.CTkFrame(master=top_frame,corner_radius=0,fg_color="#636363",height=100)
        self.frame_path_input = customtkinter.CTkFrame(master=self.root,corner_radius=0)
        self.bottom_frame2 =    customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        self.bottom_frame1 =    customtkinter.CTkFrame(master=self.root,height = 80,corner_radius=0)
        checkbox_frame =        customtkinter.CTkFrame(master=self.root,width=400,height = 150,corner_radius=0)
        self.frame_right =      customtkinter.CTkFrame(master=self.root,corner_radius=0,height = 150)
        frame_with_cards.       pack(pady=0,padx=0,fill="both",expand=True,side = "right",anchor="w")
        top_frame.              pack(pady=0,padx=0,fill="both",side = "top")
        self.frame_path_input.  pack(pady=0,padx=0,fill="both",expand=False,side = "top")
        self.bottom_frame2.     pack(pady=0,padx=5,fill="both",expand=True,side = "bottom")
        self.bottom_frame1.     pack(pady=5,padx=5,fill="x",expand=False,side = "bottom")
        checkbox_frame.         pack(pady=0,padx=5,fill="y",expand=False,side="left")
        self.frame_right.       pack(pady=0,padx=0,fill="both",expand=True,side="right")
        self.frame_with_checkboxes = checkbox_frame
        list_of_frames = [self.frame_path_input,self.bottom_frame1,self.bottom_frame2,self.frame_right,self.frame_with_checkboxes,top_frame]

        shift_const = 250
        menu_button =       customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "MENU",                  command =  lambda: self.call_extern_function(list_of_frames,function="menu"),
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="black",hover_color="#212121")
        deleting_button =   customtkinter.CTkButton(master = frame_with_cards, width = 250,height=50,text = "Mazání souborů",
                                                    font=("Arial",20,"bold"),corner_radius=0,fg_color="#212121",hover_color="#212121")
        menu_button.        grid(column = 0,row=0,pady = (70,0),padx =260-shift_const,sticky = tk.W)
        deleting_button.    grid(column = 0,row=0,pady = (70,0),padx =520-shift_const,sticky = tk.W)
        
        # menu_button =           customtkinter.CTkButton(master = self.frame_path_input, width = 180, text = "MENU", command = lambda: self.call_menu(),font=("Arial",20,"bold"))
        self.path_set    =      customtkinter.CTkEntry(master = self.frame_path_input,font=("Arial",18),placeholder_text="Zadejte cestu k souborům z kamery (kde se přímo nacházejí soubory nebo datumové složky)")
        tree        =           customtkinter.CTkButton(master = self.frame_path_input, width = 180,text = "EXPLORER", command = self.call_browseDirectories,font=("Arial",20,"bold"))
        button_save_path =      customtkinter.CTkButton(master = self.frame_path_input,width=50,text = "Uložit cestu", command = lambda: Tools.save_path(self.console,self.path_set.get()),font=("Arial",20,"bold"))
        button_open_setting =   customtkinter.CTkButton(master = self.frame_path_input,width=30,height=30, text = "⚙️", command = lambda: Advanced_option(self.root,windowed=True,spec_location="deleting_option"),font=("Arial",16))
        # menu_button.            pack(pady =12,padx=10,anchor ="w",side = "left")
        self.path_set.          pack(pady = 12,padx =(10,0),anchor ="w",side = "left",fill="both",expand=True)
        tree.                   pack(pady = 12,padx =10,anchor ="w",side = "left")
        button_save_path.       pack(pady = 12,padx =0,anchor ="w",side = "left")
        button_open_setting.    pack(pady = 12,padx =10,anchor = "w",side = "left")

        self.checkbox  =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Mazání souborů starších než: určité datum",command = lambda: self.selected(True))
        self.checkbox2 =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Redukce novějších, mazání souborů starších než: určité datum",command = lambda: self.selected2(True))
        self.checkbox3 =        customtkinter.CTkCheckBox(master = self.frame_with_checkboxes,font=("Arial",16), text = "Mazání adresářů s názvem ve formátu určitého datumu",command = lambda: self.selected3(True))
        self.checkbox.          pack(pady =10,padx=10,anchor ="w")
        self.checkbox2.         pack(pady =10,padx=10,anchor ="w")
        self.checkbox3.         pack(pady =10,padx=10,anchor ="w")

        self.checkbox6 =        customtkinter.CTkCheckBox(master = self.bottom_frame1, text = "Procházet subsložky? (max:6)",command = self.selected6,font=("Arial",16,"bold"))
        self.info2 =            customtkinter.CTkLabel(   master = self.bottom_frame1,text = "",font=("Arial",16,"bold"))
        self.checkbox_testing = customtkinter.CTkCheckBox(master = self.bottom_frame1, text = f"Režim TESTOVÁNÍ (Soubory vyhodnocené ke smazání se pouze přesunou do složky s názvem: \"{self.to_delete_folder_name}\")",font=("Arial",16,"bold"))
        self.checkbox6.         grid(column =0,row=0,sticky = tk.W,pady =5,padx=10)
        self.info2.             grid(column =0,row=0,sticky = tk.W,pady =5,padx=280)
        self.checkbox_testing.  grid(column =0,row=1,sticky = tk.W,pady =5,padx=10)
        self.info =             customtkinter.CTkLabel(master = self.bottom_frame2,text = "",font=("Arial",16,"bold"))
        execution_btn_frame =   customtkinter.CTkFrame(master=self.bottom_frame2,corner_radius=0)
        button =                customtkinter.CTkButton(master = execution_btn_frame,width = 300,height = 60,text = "SPUSTIT", command = self.start,font=("Arial",20,"bold"))
        create_task_btn =       customtkinter.CTkButton(master = execution_btn_frame,width = 300,height = 60,text = "Nastavit aut. spouštění",command = lambda: self.save_new_task(),font=("Arial",20,"bold"))
        button.                 pack(pady=10,padx=(10,0),side="left",anchor="w")
        create_task_btn.        pack(pady=10,padx=(10,0),side="left",anchor="w")
        self.console =          tk.Text(self.bottom_frame2, wrap="word", height=20, width=1200,background="black",font=("Arial",16),state=tk.DISABLED)
        self.info.              pack(pady = 12,padx =10,anchor="w",side = "top")
        execution_btn_frame.    pack(pady =20,padx=10,side = "top",anchor="n")
        self.console.           pack(pady =10,padx=10,side = "top")
        #default:
        self.checkbox.select()
        self.checkbox_testing.select()
        self.selected(False)

        if global_recources_load_error:
            create_task_btn.configure(state = "disabled")

        recources_path = self.text_file_data[0]
        if recources_path != False and recources_path != "/":
            self.path_set.delete("0","200")
            self.path_set.insert("0", str(recources_path))
            Tools.add_colored_line(self.console,"Byla vložena cesta z konfiguračního souboru","white")
        else:
            Tools.add_colored_line(self.console,"Konfigurační soubor obsahuje neplatnou cestu k souborům (můžete vložit v pokročilém nastavení)","orange")
        def maximalize_window(e):
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            currently_focused = str(self.root.focus_get())
            if ".!ctkentry" in currently_focused:
                return
            if int(self.root._current_width) > 1200:
                self.root.after(0, lambda:self.root.state('normal'))
                self.root.geometry("1200x900")
            else:
                self.root.after(0, lambda:self.root.state('zoomed'))
        self.root.bind("<f>",maximalize_window)
        self.unbind_list.append("<f>")

        def unfocus_widget(e):
            self.root.focus_set()
        self.root.bind("<Escape>",unfocus_widget)
        self.unbind_list.append("<Escape>")
        self.path_set.bind("<Return>",unfocus_widget)

if load_gui:
    if not app_running_status:
        menu = main_menu(root)
        menu.menu(initial=True)

def start_new_root():
    global menu
    global root
    global app_icon
    global initial_path
    initial_path = get_init_path()

    app_icon = Tools.resource_path('images/logo_TRIMAZKON.ico')
    customtkinter.set_appearance_mode("dark")
    customtkinter.set_default_color_theme("dark-blue")
    root=customtkinter.CTk()
    root.geometry("1200x900")
    root.title("jhv_MAZ v_1.0.0")
    root.wm_iconbitmap(Tools.resource_path(app_icon))

    menu = main_menu(root)
    menu.menu(initial=True)