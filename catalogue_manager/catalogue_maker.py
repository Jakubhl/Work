import customtkinter
import tkinter as tk
import openpyxl
from openpyxl import load_workbook
import xlwings as xw
from openpyxl.worksheet.copier import WorksheetCopy
from PIL import Image

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")
root=customtkinter.CTk()
root.geometry("1200x900")
root.title("Catalogue maker v1.0")

class Catalogue_gui:
    def __init__(self,root):
        self.root = root
        self.root.after(0, lambda:self.root.state('zoomed'))
        self.station_list = []
        self.camera_type_database = ["kamera1","kamera2","kamera3"]
        self.controller_database = ["kontoler1","kontoler2","kontoler3"]
        self.optics_database = ["optika1","optika2","optika3"]
        self.accessory_database = ["svetlo","kabel","drzak"]
        self.favourite_colors = [""]

        self.create_main_widgets()

    def clear_frame(self,frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def switch_widget_info(self,args,widget_tier,widget):
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            if widget._text == str(self.station_list[station_index]["inspection_description"]):
                widget.configure(text=str(self.station_list[station_index]["name"]))
            else:
                widget.configure(text=str(self.station_list[station_index]["inspection_description"]))
        
        elif len(widget_tier) == 4: # 0101-9999 kamery
            camera_index = int(widget_tier[2:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["type"]):
                details = "Kontroler: " + str(self.station_list[station_index]["camera_list"][camera_index]["controller"]) + "\n"
                details = details + str(self.station_list[station_index]["camera_list"][camera_index]["description"])

                widget.configure(text=details)
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["type"]))

        elif len(widget_tier) == 6: # 010101-999999 optika
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"]):
                details = "Alternativa: " +  str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["alternative"]) + "\n"
                details = details + str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["description"])
                widget.configure(text=details)
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["type"]))

        elif len(widget_tier) == 8: # 01010101-99999999 prislusenstvi
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:6])
            accessory_index = int(widget_tier[6:])
            if widget._text == str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index]["type"]):
                details = str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index]["dimension"]) + "\n"
                details = details + str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index]["description"])
                widget.configure(text=details)
            else:
                widget.configure(text=str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index]["type"]))
            
    def make_block(self,master_widget,height,width,fg_color,text,side,dummy_block = False,tier = ""):
        if dummy_block:
            dummy_block_widget =    customtkinter.CTkFrame(master=master_widget,corner_radius=0,height=height,width =width,fg_color="#212121")
            dummy_block_widget.     pack(pady = (0,0),padx =0,expand = False,side = side,anchor="w")
            return dummy_block
        else:
            block_widget =    customtkinter.CTkFrame(master=master_widget,corner_radius=0,fg_color=fg_color,height=height,width =width,border_width= 2,border_color="#636363")
            block_widget.     pack(pady = (0,0),padx =0,expand = False,side = side,anchor="w")
            block_name =      customtkinter.CTkLabel(master = block_widget,text = text,font=("Arial",25,"bold"),height=height-15,width =width-15)
            block_name.       pack(pady = 5,padx =5,anchor="n",expand=False)

            block_widget.bind("<Button-1>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
            block_name.bind("<Button-1>",lambda e, widget_tier=tier,widget = block_name: self.switch_widget_info(e, widget_tier,widget))
            return block_widget
    
    def make_new_object(self,which_one,object_to_edit = None,cam_index = None,optic_index = None):
        """
        which_one:
        - station
        - camera
        - optic
        - accessory
        """
        if which_one == "station":
            # accessory = {
            #     "type": "typ prislusenstvi",
            #     "dimension":"rozmery/ velikost",
            #     "description":"pozn",
            # }
            optic = {
                "type": "Objektiv",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            
            camera = {
                "type": "Typ kamery",
                "controller": "",
                "optics_list": [optic],
                "description": "",
            }
            station = {
                "name": "Název stanice",
                "inspection_description": "",
                "camera_list": [camera],
            }

            return station
        
        elif which_one == "camera":
            # accessory = {
            #     "type": "typ prislusenstvi",
            #     "dimension":"rozmery/ velikost",
            #     "description":"pozn",
            # }
            optic = {
                "type": "Objektiv",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }
            camera = {
                "type": "Typ kamery",
                "controller": "",
                "optics_list": [optic],
                "description": "",
            }

            object_to_edit["camera_list"].append(camera)
            return object_to_edit
        
        elif which_one == "optic":
            # accessory = {
            #     "type": "typ prislusenstvi",
            #     "dimension":"rozmery/ velikost",
            #     "description":"pozn",
            # }
            optic = {
                "type": "Objektiv",
                "alternative":"",
                "accessory_list": [],
                "description":"",
            }

            object_to_edit["camera_list"][cam_index]["optics_list"].append(optic)
            return object_to_edit
        
        elif which_one == "accessory":
            accessory = {
                "type": "Příslušenství",
                "dimension":"",
                "description":"",
            }

            object_to_edit["camera_list"][cam_index]["optics_list"][optic_index]["accessory_list"].append(accessory)
            return object_to_edit

    def manage_widgets(self,args,widget_tier,btn):
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            if btn == "add_line": # nova stanice
                new_station = self.make_new_object("station")
                self.station_list.append(new_station)
                self.make_project_widgets()

            elif btn == "add_object": # nova kamera ke stanici 0101-9999 kamery
                station_with_new_camera = self.make_new_object("camera",object_to_edit = self.station_list[station_index])
                self.station_list[station_index] = station_with_new_camera
                self.make_project_widgets()
        
        elif len(widget_tier) == 4: # 0101-9999 kamery, nove bude pridano: 010101-999999 optika
            if btn == "add_object": # nova optika kamery
                camera_index = int(widget_tier[2:])
                camera_with_new_optics = self.make_new_object("optic",object_to_edit = self.station_list[station_index],cam_index = camera_index)
                self.station_list[station_index] = camera_with_new_optics
                self.make_project_widgets()

        elif len(widget_tier) == 6: # 010101-999999 optika, nove bude pridano: 01010101-99999999 prislusenstvi
            if btn == "add_object": # nove prislusenstvi ka kamere
                camera_index = int(widget_tier[2:4])
                optic_index = int(widget_tier[4:])
                camera_with_new_accessoryes = self.make_new_object("accessory",object_to_edit = self.station_list[station_index],cam_index = camera_index,optic_index = optic_index)
                self.station_list[station_index] = camera_with_new_accessoryes
                self.make_project_widgets()

        # print("STATION_LIST: ",self.station_list)
    
    def confirm_delete(self,to_del_object):
        def make_decision(decision):
            if decision == True:
                self.station_list.pop(to_del_object)
            self.make_project_widgets()
            child_root.destroy()
        
        child_root=customtkinter.CTk()
        x = self.root.winfo_rootx()
        y = self.root.winfo_rooty()
        child_root.geometry(f"650x130+{x+80}+{y+80}")  
        child_root.title("Upozornění")

        proceed_label = customtkinter.CTkLabel(master = child_root,text = "Opravdu si přejete odstranit celou stanici a všechna zařízení k ní připojená?",font=("Arial",18))
        button_yes =    customtkinter.CTkButton(master = child_root,text = "Pokračovat",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda: make_decision(True))
        button_no =     customtkinter.CTkButton(master = child_root,text = "Zrušit",font=("Arial",20,"bold"),width = 200,height=50,corner_radius=0,command=lambda:  make_decision(False))

        proceed_label   .pack(pady=(15,0),padx=10,expand=False,side = "top")
        button_no       .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        button_yes      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="right")
        child_root.mainloop()
        
    def delete_block(self,args,widget_tier):
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            print("deleting",self.station_list[station_index])
            self.confirm_delete(station_index)
            return
        
        elif len(widget_tier) == 4: # 0101-9999 kamery
            camera_index = int(widget_tier[2:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index])
            self.station_list[station_index]["camera_list"].pop(camera_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"].pop(optic_index)
            
        elif len(widget_tier) == 8: # 01010101-99999999 prislusenstvi
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:6])
            accessory_index = int(widget_tier[6:])
            print("deleting",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"].pop(accessory_index)
        
        #refresh
        self.make_project_widgets()

    def edit_object_gui(self,object:str,station_index,camera_index = None,optics_index = None,accessory_index = None):
        """
        Object:
        - station
        - camera
        - optics
        - accessory
        """
        def save_station_changes(new_name,new_description,child_root):
            self.station_list[station_index]["name"] = new_name
            self.station_list[station_index]["inspection_description"] = new_description
            self.make_project_widgets()
            child_root.destroy()
        def edit_station():
            # lze editovat nazev a popis inspekce
            child_root=customtkinter.CTk()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"350x450+{x+80}+{y+80}")  
            child_root.title("Editování stanice: " + str(self.station_list[station_index]["name"]))
            station_name_label =        customtkinter.CTkLabel(master = child_root,text = "Název stanice:",font=("Arial",22,"bold"))
            new_name =                  customtkinter.CTkEntry(master = child_root,font=("Arial",22),width=300,height=50,corner_radius=0)
            inspection_description =    customtkinter.CTkLabel(master = child_root,text = "Popis inspekce:",font=("Arial",22,"bold"))
            # new_description =           customtkinter.CTkEntry(master = child_root,font=("Arial",22),width=300,height=50,corner_radius=0)
            new_description =           customtkinter.CTkTextbox(master = child_root,font=("Arial",22),width=300,height=200)
            button_save =               customtkinter.CTkButton(master = child_root,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,
                                                                command=lambda: save_station_changes(new_name.get(),new_description.get("0.0", "end"),child_root))
            station_name_label          .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
            new_name                    .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            inspection_description      .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            # new_description             .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            new_description             .pack(pady = 5, padx = 10,expand=True,side="top")
            button_save                 .pack(pady = 5, padx = 10,expand=True,side="bottom")
            

            # initial prefill:
            new_name.insert(0,str(self.station_list[station_index]["name"]))
            new_description.insert("0.0",str(self.station_list[station_index]["inspection_description"]))
            child_root.mainloop()

        def save_cam_changes(new_camera_type,new_controller,new_notes,child_root):
            self.station_list[station_index]["camera_list"][camera_index]["type"] = new_camera_type
            self.station_list[station_index]["camera_list"][camera_index]["controller"] = new_controller
            self.station_list[station_index]["camera_list"][camera_index]["description"] = new_notes
            self.make_project_widgets() #refresh
            child_root.destroy()
        def edit_cam():
            # lze editovat nazev a popis inspekce
            child_root=customtkinter.CTk()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"350x520+{x+80}+{y+80}")  
            child_root.title("Editování kamery: " + str(self.station_list[station_index]["camera_list"][camera_index]["type"]))
            camera_type =               customtkinter.CTkLabel(master = child_root,text = "Typ kamery:",font=("Arial",22,"bold"))
            camera_type_entry =         customtkinter.CTkOptionMenu(master = child_root,font=("Arial",22),width=300,height=50,values=self.camera_type_database,corner_radius=0)
            controller =                customtkinter.CTkLabel(master = child_root,text = "Kontroler:",font=("Arial",22,"bold"))
            controller_entry =          customtkinter.CTkOptionMenu(master = child_root,font=("Arial",22),width=300,height=50,values=self.controller_database,corner_radius=0)
            note_label =                customtkinter.CTkLabel(master = child_root,text = "Poznámky:",font=("Arial",22,"bold"))
            notes_input =               customtkinter.CTkTextbox(master = child_root,font=("Arial",22),width=300,height=200)
            button_save =               customtkinter.CTkButton(master = child_root,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,
                                                                command=lambda: save_cam_changes(camera_type_entry.get(),controller_entry.get(),notes_input.get("0.0", "end"),child_root))
            camera_type                 .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
            camera_type_entry           .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            controller                  .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            controller_entry            .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            note_label                  .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            notes_input                 .pack(pady = 5, padx = 10,expand=True,side="top")
            button_save                 .pack(pady = 5, padx = 10,expand=True,side="bottom")

            # initial prefill:
            if str(self.station_list[station_index]["camera_list"][camera_index]["type"]) in self.camera_type_database:
                camera_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["type"]))
            if str(self.station_list[station_index]["camera_list"][camera_index]["controller"]) in self.controller_database:
                controller_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["controller"]))
            notes_input.insert("0.0",str(self.station_list[station_index]["camera_list"][camera_index]["description"]))
            child_root.mainloop()

        def save_optics_changes(new_optics_type,new_optics_alternative,new_notes,child_root):
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"] = new_optics_type
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"] = new_optics_alternative
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"] = new_notes
            self.make_project_widgets() #refresh
            child_root.destroy()
        def edit_optics():
            # lze editovat nazev a popis inspekce
            child_root=customtkinter.CTk()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"350x520+{x+80}+{y+80}")  
            child_root.title("Editování optiky: " + str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
            optic_type =                customtkinter.CTkLabel(master = child_root,text = "Typ objektivu:",font=("Arial",22,"bold"))
            optic_type_entry =          customtkinter.CTkOptionMenu(master = child_root,font=("Arial",22),width=300,height=50,values=self.optics_database,corner_radius=0)
            alternative_type =          customtkinter.CTkLabel(master = child_root,text = "Alternativa:",font=("Arial",22,"bold"))
            alternative_entry =         customtkinter.CTkOptionMenu(master = child_root,font=("Arial",22),width=300,height=50,values=self.optics_database,corner_radius=0)
            note_label =                customtkinter.CTkLabel(master = child_root,text = "Poznámky:",font=("Arial",22,"bold"))
            notes_input =               customtkinter.CTkTextbox(master = child_root,font=("Arial",22),width=300,height=200)
            button_save =               customtkinter.CTkButton(master = child_root,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,
                                                                command=lambda: save_optics_changes(optic_type_entry.get(),alternative_entry.get(),notes_input.get("0.0", "end"),child_root))
            optic_type                  .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
            optic_type_entry            .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            alternative_type            .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            alternative_entry           .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            note_label                  .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            notes_input                 .pack(pady = 5, padx = 10,expand=True,side="top")
            button_save                 .pack(pady = 5, padx = 10,expand=True,side="bottom")

            # initial prefill:
            if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]) in self.optics_database:
                optic_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["type"]))
            if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]) in self.optics_database:
                alternative_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["alternative"]))
            notes_input.insert("0.0",str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["description"]))
            child_root.mainloop()

        def save_accessory_changes(new_accessory_type,new_notes,child_root):
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["type"] = new_accessory_type
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["description"] = new_notes
            self.make_project_widgets() #refresh
            child_root.destroy()
        def edit_accessory():
            # lze editovat nazev a popis inspekce
            child_root=customtkinter.CTk()
            x = self.root.winfo_rootx()
            y = self.root.winfo_rooty()
            child_root.geometry(f"350x470+{x+80}+{y+80}")  
            child_root.title("Editování příslušenství: " + str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["type"]))
            hw_type =                   customtkinter.CTkLabel(master = child_root,text = "Zařízení:",font=("Arial",22,"bold"))
            hw_type_entry =             customtkinter.CTkOptionMenu(master = child_root,font=("Arial",22),width=300,height=50,values=self.accessory_database,corner_radius=0)
            note_label =                customtkinter.CTkLabel(master = child_root,text = "Poznámky:",font=("Arial",22,"bold"))
            notes_input =               customtkinter.CTkTextbox(master = child_root,font=("Arial",22),width=300,height=200)
            button_save =               customtkinter.CTkButton(master = child_root,text = "Uložit",font=("Arial",22,"bold"),width = 200,height=50,corner_radius=0,
                                                                command=lambda: save_accessory_changes(hw_type_entry.get(),notes_input.get("0.0", "end"),child_root))
            hw_type                     .pack(pady=(15,5),padx=10,anchor="w",expand=False,side = "top")
            hw_type_entry               .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            note_label                  .pack(pady = 5, padx = 10,anchor="w",expand=False,side="top")
            notes_input                 .pack(pady = 5, padx = 10,expand=True,side="top")
            button_save                 .pack(pady = 5, padx = 10,expand=True,side="bottom")

            # initial prefill:
            if str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["type"]) in self.accessory_database:
                hw_type_entry.set(str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["type"]))
           
            notes_input.insert("0.0",str(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"][accessory_index]["description"]))
            child_root.mainloop()

        if object == "station":
            edit_station()
        elif object == "camera":
            edit_cam()
        elif object == "optics":
            edit_optics()
        elif object == "accessory":
            edit_accessory()

    def edit_object(self,args,widget_tier):
        station_index = int(widget_tier[:2])
        if len(widget_tier) == 2: #01-99 stanice
            print("editing",self.station_list[station_index])
            self.edit_object_gui("station",station_index)
        
        elif len(widget_tier) == 4: # 0101-9999 kamery
            camera_index = int(widget_tier[2:])
            print("editing",self.station_list[station_index]["camera_list"][camera_index])
            self.edit_object_gui("camera",station_index,camera_index)

        elif len(widget_tier) == 6: # 010101-999999 optika
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:])
            print("editing",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index])
            self.edit_object_gui("optics",station_index,camera_index,optic_index)
            
        elif len(widget_tier) == 8: # 01010101-99999999 prislusenstvi
            camera_index = int(widget_tier[2:4])
            optic_index = int(widget_tier[4:6])
            accessory_index = int(widget_tier[6:])
            print("editing",self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optic_index]["accessory_list"][accessory_index])
            self.edit_object_gui("accessory",station_index,camera_index,optic_index,accessory_index)
        
        #refresh
        self.make_project_widgets()

    def make_block_buttons(self,master_widget,tier:str,station:bool,accessory=False):#,btn_add_line:str,btn_add_object:str
        button_add_line = customtkinter.CTkButton(master = master_widget, width = 25,height=25,text = "+",font=("",15),corner_radius=0,fg_color="#009933",hover_color="green")
        if station:
            button_add_line.pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_edit_object = customtkinter.CTkButton(master = master_widget,text = "🖌",font=("",15),width = 25,height=25,corner_radius=0)
        button_edit_object.pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_edit_color = customtkinter.CTkButton(master = master_widget,text = "🎨",font=("",15),width = 25,height=25,corner_radius=0)
        button_edit_color.pack(pady = 5, padx = (5,0),anchor="w",expand=False,side="left")
        button_del_object = customtkinter.CTkButton(master = master_widget, width = 25,height=25,text = "×",font=("",15),corner_radius=0,fg_color="#cc0000",hover_color="red")
        button_del_object.pack(pady = 5, padx = (5,0),anchor="w",expand=True,side="left")
        button_add_object = customtkinter.CTkButton(master = master_widget, width = 25,height=25,text = "+",font=("",15),corner_radius=0,fg_color="#009933",hover_color="green")
        if not accessory:
            button_add_object.pack(pady = 5, padx = 5,anchor="e",expand=True,side="left")

        if station:
            button_add_line.bind("<Button-1>",lambda e, widget_tier=tier, btn = "add_line": self.manage_widgets(e, widget_tier,btn))
        if not accessory:
            button_add_object.bind("<Button-1>",lambda e, widget_tier=tier, btn = "add_object": self.manage_widgets(e, widget_tier,btn))
        
        button_del_object.bind("<Button-1>",lambda e, widget_tier=tier: self.delete_block(e, widget_tier))
        button_edit_object.bind("<Button-1>",lambda e, widget_tier=tier: self.edit_object(e, widget_tier))

        master_widget.update()
        # print(master_widget._current_height)
    
    def create_main_widgets(self):
        self.clear_frame(self.root)
        main_header =               customtkinter.CTkFrame(master=self.root,corner_radius=0,height=100)
        main_header                 .pack(pady=0,padx=5,fill="x",expand=False,side = "top")

        # logo =                      customtkinter.CTkImage(Image.open(initial_path+"images/logo.png"),size=(1200, 100))
        # image_logo =                customtkinter.CTkLabel(master = frame_with_logo,text = "",image =logo)
        export_button =             customtkinter.CTkButton(master = main_header,text = "Exportovat .xlsm",font=("Arial",25,"bold"),width=250,height=50,corner_radius=0,
                                                            command=lambda:Save_excel(station_list = self.station_list))
        export_button               .pack(pady = 10, padx = 10,anchor="w",expand=False,side="left")

        column_labels =             customtkinter.CTkFrame(master=self.root,corner_radius=0,fg_color="#636363",height=50)
        self.project_tree =         customtkinter.CTkScrollableFrame(master=self.root,corner_radius=0)
        column_labels               .pack(pady=0,padx=5,fill="x",expand=False,side = "top")
        self.project_tree           .pack(pady=5,padx=5,fill="both",expand=True,side = "top")
        stations_column_header =    customtkinter.CTkLabel(master = column_labels,text = "Stanice",font=("Arial",25,"bold"),bg_color="#212121",width=275,height=50)
        stations_column_header      .pack(pady=(15,0),padx=10,expand=False,side = "left")
        camera_column_header =      customtkinter.CTkLabel(master = column_labels,text = "Kamera",font=("Arial",25,"bold"),bg_color="#212121",width=275,height=50)
        camera_column_header        .pack(pady=(15,0),padx=10,expand=False,side = "left")
        optics_column_header =      customtkinter.CTkLabel(master = column_labels,text = "Objektiv",font=("Arial",25,"bold"),bg_color="#212121",width=275,height=50)
        optics_column_header        .pack(pady=(15,0),padx=10,expand=False,side = "left")
        accessory_column_header =   customtkinter.CTkLabel(master = column_labels,text = "Příslušenství",font=("Arial",25,"bold"),bg_color="#212121",width=275,height=50)
        accessory_column_header     .pack(pady=(15,0),padx=10,expand=False,side = "left")
        
        self.project_column =   customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)
        self.project_column     .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.camera_column =    customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)    
        self.camera_column      .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.optic_column =     customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)    
        self.optic_column       .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.accessory_column = customtkinter.CTkFrame(master=self.project_tree,corner_radius=0,border_width=0)    
        self.accessory_column   .pack(pady=0,padx=0,fill="y",expand=False,side = "left")
        self.station_list.append(self.make_new_object("station"))
        self.make_project_widgets()
        
        def maximalize_window(e):
            self.root.update_idletasks()
            current_width = int(self.root.winfo_width())
            # netrigguj fullscreen zatimco pisu do vstupniho textovyho pole
            if self.focused_entry_widget(): # pokud nabindovane pismeno neni vepisovano do entry widgetu
                return
            if int(current_width) > 1200:
                #self.root.after(0, lambda:self.root.state('normal'))
                self.root.state('normal')
                self.root.geometry(f"260x500+{0}+{0}")
                # self.root.geometry("210x500")
                self.save_setting_parameter(parameter="change_def_window_size",status=2)
            elif int(current_width) ==260:
                self.root.geometry("1200x900")
                self.save_setting_parameter(parameter="change_def_window_size",status=0)
            else:
                #self.root.after(0, lambda:self.root.state('zoomed'))
                self.root.state('zoomed')
                self.save_setting_parameter(parameter="change_def_window_size",status=1)

        self.root.bind("<f>",lambda e: maximalize_window(e))
    
    def check_widget_growth(self,widget:str,station_index,camera_index=None,optics_index=None):
        """
        widget:
        - station
        - camera
        - optics
        """
        # station_optic_count = 0
        station_accessory_count = 0 # dummy block...
        station_widget_growth_accessory = 0
        # station_widget_growth_optics = 0

        if widget == "station":
            for camera in self.station_list[station_index]["camera_list"]:
                # station_optic_count += len(camera["optics_list"])
                for optics in camera["optics_list"]:
                    station_accessory_count += len(optics["accessory_list"])
                    if len(optics["accessory_list"]) == 0:
                        station_accessory_count +=1
            # station_widget_growth_optics = ((station_optic_count*100)-100)
            if station_accessory_count>0:
                station_widget_growth_accessory = ((station_accessory_count*100)-100)
            
            self.station_list[station_index]["row_count"] = station_accessory_count
            return station_widget_growth_accessory

        elif widget == "camera":
            for optics in self.station_list[station_index]["camera_list"][camera_index]["optics_list"]:
                station_accessory_count += len(optics["accessory_list"])
                if len(optics["accessory_list"]) == 0:
                    station_accessory_count +=1

            if station_accessory_count>0:
                station_widget_growth_accessory = ((station_accessory_count*100)-100)
            
            self.station_list[station_index]["camera_list"][camera_index]["row_count"] = station_accessory_count
            return station_widget_growth_accessory
        
        elif widget == "optics":
            station_accessory_count = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"])
            if station_accessory_count>0:
                station_widget_growth_accessory = ((station_accessory_count*100)-100)
            
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["row_count"] = station_accessory_count
            return station_widget_growth_accessory

    def make_project_widgets(self):
        self.clear_frame(self.project_column)
        self.clear_frame(self.camera_column)
        self.clear_frame(self.optic_column)
        self.clear_frame(self.accessory_column)

        # creating stations ------------------------------------------------------------------------------------------------------------------------------
        for i in range(0,len(self.station_list)):
            station_name = self.station_list[i]["name"]
            if i < 10:
                station_tier =  "0" + str(i) #01-99 
            else:
                station_tier =  str(i) #01-99

            station_camera_list = self.station_list[i]["camera_list"]
            camera_count = len(station_camera_list)

            station_widget_growth = self.check_widget_growth("station",station_index=i)
            station_widget = self.make_block(master_widget=self.project_column,height=70+station_widget_growth,width=300,fg_color="#181818",side = "top",text=station_name,tier=station_tier)
            self.make_block_buttons(master_widget=station_widget,tier=station_tier,station=True)
            # creating cameras ------------------------------------------------------------------------------------------------------------------------------
            for x in range(0,camera_count):
                camera_type = station_camera_list[x]["type"]
                station_camera_optic_list = station_camera_list[x]["optics_list"]
                optic_count = len(station_camera_optic_list)
                # camera_widget_growth = ((optic_count*100)-100)
                if x < 10:
                    camera_tier =  station_tier + "0" + str(x) #0101-9999
                else:    
                    camera_tier =  station_tier + str(x) #0101-9999

                camera_widget_growth = self.check_widget_growth("camera",station_index=i,camera_index=x)
                camera_widget = self.make_block(master_widget=self.camera_column,height=70+camera_widget_growth,width=300,fg_color="#181818",side = "top",text=camera_type,tier = camera_tier)
                self.make_block_buttons(master_widget=camera_widget,tier=camera_tier,station=False)

                # creating optics ------------------------------------------------------------------------------------------------------------------------------
                for y in range(0,optic_count):
                    optic_type = station_camera_optic_list[y]["type"]
                    accessory_list = station_camera_optic_list[y]["accessory_list"]
                    accessory_count = len(accessory_list)
                    if y < 10:
                        optic_tier =  camera_tier + "0" + str(y) #010101-999999
                    else:
                        optic_tier =  camera_tier + str(y) #010101-999999

                    optic_widget_growth = self.check_widget_growth("optics",station_index=i,camera_index=x,optics_index=y)
                    optic_widget = self.make_block(master_widget=self.optic_column,height=70+optic_widget_growth,width=300,fg_color="#181818",side = "top",text=optic_type,tier=optic_tier)
                    self.make_block_buttons(master_widget=optic_widget,tier=optic_tier,station=False)

                    # creating accessories ------------------------------------------------------------------------------------------------------------------------------
                    for z in range(0,accessory_count):
                        accessory_type = accessory_list[z]["type"]
                        if z < 10:
                            accessory_tier =  optic_tier + "0" + str(z) #01010101-99999999
                        else:
                            accessory_tier =  optic_tier + str(z) #01010101-99999999

                        accessory_widget = self.make_block(master_widget=self.accessory_column,height=70,width=300,fg_color="#181818",side = "top",text=accessory_type,tier = accessory_tier)
                        self.make_block_buttons(master_widget=accessory_widget,tier=accessory_tier,station=False,accessory=True)
                    if accessory_count == 0:
                        accessory_widget = self.make_block(master_widget=self.accessory_column,height=100,width=300,fg_color="#181818",side = "top",text="",dummy_block=True)
        
class Save_excel:
    def __init__(self,station_list):
        # self.change_vba_script()
        self.station_list = station_list
        self.excel_file_path = "example.xlsx"
        self.main() 

    def merge_cells(self,sheet_name:str,cell_range:str):
        """
        cell range format: A1:A2
        """
        # wb = load_workbook(filename=self.excel_file_path, read_only=False, keep_vba=True)
        wb = load_workbook(filename=self.excel_file_path)
        ws = wb[sheet_name]
        ws.merge_cells(cell_range)
        wb.save(filename=self.excel_file_path)
        wb.close()

    def update_sheet_vba_code(self,file_path, sheet_name, new_code):
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        vb_project = wb.api.VBProject
        sheet = wb.sheets[sheet_name]
        code_module = vb_project.VBComponents(sheet.name).CodeModule
        code_module.DeleteLines(1, code_module.CountOfLines)
        code_module.AddFromString(new_code)
        wb.save()
        wb.close()
        app.quit()

    def check_row_count(self,widget,station_index,camera_index=None,optics_index = None):
        """
        pridavame novy parametr, informace o poctu radku u kazde stanice, kazde kamery a kazde optiky\n
        nemohu to číst a zapisovat dříve, kvůli zpětnému přidávání bloků...\n
        widget:
        - station
        - camera
        - optics
        """
        station_accessory_count = 0 # dummy block...
        if widget == "station":
            for camera in self.station_list[station_index]["camera_list"]:
                for optics in camera["optics_list"]:
                    station_accessory_count += len(optics["accessory_list"])
                    if len(optics["accessory_list"]) == 0:
                        station_accessory_count +=1
            self.station_list[station_index]["row_count"] = station_accessory_count

        elif widget == "camera":
            for optics in self.station_list[station_index]["camera_list"][camera_index]["optics_list"]:
                station_accessory_count += len(optics["accessory_list"])
                if len(optics["accessory_list"]) == 0:
                    station_accessory_count +=1
            self.station_list[station_index]["camera_list"][camera_index]["row_count"] = station_accessory_count

        elif widget == "optics":
            station_accessory_count = len(self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["accessory_list"])
            self.station_list[station_index]["camera_list"][camera_index]["optics_list"][optics_index]["row_count"] = station_accessory_count


    def get_cells_to_merge(self):
        last_row = 1
        rows_to_merge = []
        for stations in self.station_list:
            station_index = self.station_list.index(stations)
            if stations["row_count"] > 1:
                self.station_list[station_index]["excel_position"] = "A"+str(last_row)
                rows_to_merge.append("A" + str(last_row) + ":A" + str(last_row + int(stations["row_count"]) - 1))
                last_row = last_row + (stations["row_count"])
            else:
                self.station_list[station_index]["excel_position"] = "A"+str(last_row)
                last_row = last_row + 1
        
        #     print(self.station_list[station_index]["excel_position"])
        # print(rows_to_merge)
        return rows_to_merge


    def change_vba_script(self):
        # Create a new Excel workbook and select the active sheet

        # Define the VBA code to be added
        sheet_name="Sheet1"
        file_path="formular2.xlsm"
        vba_code = """
        Private Sub Worksheet_BeforeRightClick(ByVal Target As Range, Cancel As Boolean)
            ' Define the cell you chraplavý kašel want to toggle
            Dim targetCell As Range
            Set targetCell = Me.Range("A1") ' Change "A1" to the cell reference you want to toggle

            ' Read text values from hidden worksheet
            Dim text1 As String
            Dim text2 As String
            text1 = Worksheets("HiddenSheet").Range("AAA1").Value
            text2 = Worksheets("HiddenSheet").Range("AAA2").Value

            ' Read toggle status from hidden worksheet
            Dim toggle_status As Integer
            toggle_status = Worksheets("HiddenSheet").Range("AAA3").Value

            ' Check if the right-clicked cell is the target cell
            If Not Intersect(Target, targetCell) Is Nothing Then
                ' Toggle the cell value
                If toggle_status = 1 Then
                    Worksheets("HiddenSheet").Range("AAA1").Value = targetCell.Value
                    targetCell.Value = text2
                    toggle_status = 0
                Else
                    Worksheets("HiddenSheet").Range("AAA2").Value = targetCell.Value
                    targetCell.Value = text1
                    toggle_status = 1
                End If

                ' Update toggle status on hidden worksheet
                Worksheets("HiddenSheet").Range("AAA3").Value = toggle_status
                ' Cancel the default right-click menu
                Cancel = True
            End If
        End Sub
        """

        self.update_sheet_vba_code(file_path, sheet_name, new_code=vba_code)

    def fill_values(self,sheet_name):
        # wb = load_workbook(filename=self.excel_file_path, read_only=False, keep_vba=True)
        wb = load_workbook(filename=self.excel_file_path)
        ws = wb[sheet_name]
        
        
        wb.save(filename=self.excel_file_path)
        wb.close()

    def main(self):
        rows_to_merge = self.get_cells_to_merge()
        for rows in rows_to_merge:
            self.merge_cells(sheet_name="List1",cell_range=rows)
Catalogue_gui(root)
# Manage_excel()

root.mainloop()